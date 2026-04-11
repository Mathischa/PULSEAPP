# -*- coding: utf-8 -*-
"""
loader.py — Fonctions de chargement robuste des fichiers Excel (OneDrive/SharePoint).
"""
from __future__ import annotations

import os
import subprocess
import time
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")


def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)


def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(
            ["attrib", path], 
            capture_output=True, 
            text=True, 
            shell=True
        )
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None


def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")


def pin_and_hydrate(path: str, timeout: int = 30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l'hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False


def robust_load_workbook(path: str, retries: int = 3, delay: float = 1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(
                f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} "
                f"({type(e).__name__}: {e})"
            )
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e


def diag_path(path: str) -> None:
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
