# -*- coding: utf-8 -*-
"""
api/import_profils.py — Endpoints REST pour l'import des profils de prévision.

POST /api/import_profils/scan      — Scan le dossier et détecte les fichiers profil.
POST /api/import_profils/launch    — Lance l'import en arrière-plan.
GET  /api/import_profils/progress/<job_id> — Retourne la progression d'un job.
"""
from __future__ import annotations

import os
import re
import uuid
import traceback
import threading

from flask import Blueprint, request, jsonify

bp = Blueprint("import_profils", __name__, url_prefix="/api/import_profils")

# ─── In-memory job store ────────────────────────────────────────────────────
_JOBS: dict[str, dict] = {}


# ─── Helpers ────────────────────────────────────────────────────────────────

def _build_pattern(year: int) -> re.Pattern:
    """Construit le regex de détection des fichiers profil pour une année donnée."""
    y = year
    y_prev = y - 1

    def y_or_date(yy: int) -> str:
        return rf"{yy}(?:-\d{{2}}-\d{{2}})?"

    return re.compile(
        rf"^Profil\s+Tr[ée]so\s+SNCF\s+"
        rf"(?:"
            rf"{y_prev}\s*[-–—]\s*{y_or_date(y)}"
            rf"|"
            rf"{y}\s*[-–—]\s*{y_or_date(y)}"
        rf")\b.*\.xlsx$",
        re.IGNORECASE,
    )


def _scan_folder(folder: str, pattern: re.Pattern) -> list[str]:
    """Parcourt récursivement le dossier et retourne les chemins qui matchent le pattern."""
    resultats = []
    for dossier, _, fichiers in os.walk(folder):
        for fichier in fichiers:
            if pattern.match(fichier):
                resultats.append(os.path.join(dossier, fichier))
    return resultats


def _find_reel(year: int, base_dir: str) -> str | None:
    """
    Cherche le fichier réel pour l'année donnée.
    1) Essai exact : «Réel {year}.xlsx»
    2) Fallback regex insensible à la casse.
    """
    # Tentative exacte
    attendu = os.path.join(base_dir, f"Réel {year}.xlsx")
    if os.path.isfile(attendu):
        return attendu

    # Tentative flexible
    patt = re.compile(rf"^R[ée]el\s+{year}\.xlsx$", re.IGNORECASE)
    try:
        for f in os.listdir(base_dir):
            if patt.match(f):
                return os.path.join(base_dir, f)
    except OSError:
        pass

    return None


# ─── Background import ──────────────────────────────────────────────────────

def _run_import(job_id: str, year: int, fichiers_trouves: list[str]) -> None:
    """Fonction d'import exécutée dans un thread dédié."""
    import pandas as pd
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    from collections import defaultdict
    from datetime import datetime, date as date_type
    from pulse_v2.config import (
        FICHIER_EXCEL_DIR,
        FICHIER_CONFIG_SECTIONS,
        BASE_DONNEES_DIR,
        FEUILLE_CONFIG_SECTIONS,
        COL_DEST,
        COL_SOURCE,
        COL_PREV,
    )

    def _set(pct: int, msg: str) -> None:
        _JOBS[job_id]["progress"] = pct
        _JOBS[job_id]["message"] = msg

    try:
        _set(2, "Lecture des dates des fichiers profil…")

        # ── Chargement des sections ─────────────────────────────────────────
        def load_sections() -> list[dict]:
            if not os.path.isfile(FICHIER_CONFIG_SECTIONS):
                return []
            try:
                wb = load_workbook(FICHIER_CONFIG_SECTIONS, read_only=True, data_only=True)
            except Exception as e:
                print(f"[ERROR] load_sections: {e}")
                return []
            if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[FEUILLE_CONFIG_SECTIONS]
            mapping = []
            row = 2
            while True:
                dest_val   = ws.cell(row=row, column=COL_DEST).value
                source_val = ws.cell(row=row, column=COL_SOURCE).value
                prev_val   = ws.cell(row=row, column=COL_PREV).value
                if dest_val is None or str(dest_val).strip() == "":
                    break
                mapping.append({
                    "dest":   str(dest_val).strip(),
                    "source": str(source_val).strip() if source_val else str(dest_val).strip(),
                    "prev":   str(prev_val).strip()   if prev_val   else str(dest_val).strip(),
                })
                row += 1
            wb.close()
            return mapping

        # ── Lecture date C descendant ───────────────────────────────────────
        def lire_date_c(fichier: str, feuille: str = "SA_SNCF",
                        start_row: int = 6, max_lookahead: int = 50):
            try:
                wb = load_workbook(fichier, read_only=True, data_only=True)
                if feuille not in wb.sheetnames:
                    wb.close()
                    return None, None
                ws = wb[feuille]

                def parse(val):
                    if isinstance(val, datetime):  return val.date()
                    if isinstance(val, date_type): return val
                    if isinstance(val, str):
                        s = val.strip()
                        try: return datetime.strptime(s, "%d/%m/%Y").date()
                        except: pass
                        try: return date_type.fromisoformat(s)
                        except: pass
                        m = re.search(r"\b(19\d{2}|20\d{2})\b", s)
                        if m: return date_type(int(m.group(1)), 1, 1)
                    return None

                for r in range(start_row, start_row + max_lookahead):
                    val = ws[f"C{r}"].value
                    if val is None:
                        continue
                    d = parse(val)
                    if not d:
                        continue
                    if d.year != year:
                        if isinstance(val, str) and re.search(rf"\b{year}\b", val):
                            wb.close()
                            return d, r
                        continue
                    wb.close()
                    return d, r

                wb.close()
                return None, None
            except Exception as e:
                print(f"[ERROR] lire_date_c {os.path.basename(fichier)}: {e}")
                return None, None

        def choisir_meilleure_version(fichiers: list[str]) -> str | None:
            fichiers = sorted(fichiers)
            for tag in ("V3", "V2"):
                tagged = [f for f in fichiers if tag in os.path.basename(f)]
                if tagged:
                    return tagged[0]
            return fichiers[0] if fichiers else None

        # ── 1) Dates par fichier ────────────────────────────────────────────
        dates_par_fichier = []
        for f in fichiers_trouves:
            d, row_found = lire_date_c(f)
            if d:
                dates_par_fichier.append((f, d, row_found))

        if not dates_par_fichier:
            _set(0, "Aucune date trouvée dans les fichiers profil.")
            _JOBS[job_id]["done"] = True
            return

        # ── 2) Groupement + meilleure version ──────────────────────────────
        groupes: dict = defaultdict(list)
        for fichier, d, row_found in dates_par_fichier:
            groupes[d].append((fichier, row_found))

        dates_uniques = []
        for d, lst in groupes.items():
            fichiers_g = [f for f, _ in lst]
            best = choisir_meilleure_version(fichiers_g)
            if best:
                row_found = next(r for f, r in lst if f == best)
                dates_uniques.append((best, d, row_found))
        dates_uniques.sort(key=lambda x: x[1])

        fichier_prev = [f for f, _, _ in dates_uniques]
        rows_prev    = [r for _, _, r in dates_uniques]

        # ── 3) Deltas cumulés ───────────────────────────────────────────────
        deltas = []
        for i in range(1, len(dates_uniques)):
            _, d1, _ = dates_uniques[i - 1]
            _, d2, _ = dates_uniques[i]
            deltas.append((d2 - d1).days)
        cumul = [0]
        s = 0
        for d in deltas:
            s += d
            cumul.append(s)

        # ── 4) Fichier réel ─────────────────────────────────────────────────
        fichier_source = _find_reel(year, BASE_DONNEES_DIR)
        if not fichier_source or not os.path.isfile(fichier_source):
            _set(0, f"Fichier 'Réel {year}.xlsx' introuvable dans {BASE_DONNEES_DIR}")
            _JOBS[job_id]["done"] = True
            return

        # ── 5) Sections ─────────────────────────────────────────────────────
        sections = load_sections()
        if not sections:
            _set(0, "Aucune section dans le fichier de configuration.")
            _JOBS[job_id]["done"] = True
            return

        # ── 6) Labels de prévisions ─────────────────────────────────────────
        dates_prev = []
        for path in fichier_prev:
            m = re.search(r"\d{4}-\d{2}-\d{2}", path)
            dates_prev.append(
                datetime.strptime(m.group(), "%Y-%m-%d").strftime("%d/%m") if m else "N/A"
            )

        # ── 7) Styles openpyxl ──────────────────────────────────────────────
        fill_jaune   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_bleu    = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        border       = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        align_center = Alignment(horizontal="center", vertical="center")
        font_bold    = Font(bold=True)

        # ── 8) Préchargement matrices de prévisions ─────────────────────────
        _set(5, "Préchargement des données de prévision…")
        valeurs_prev_by_section: dict[str, list] = {}
        for section in sections:
            feuille_prev_name = section["prev"]
            frames = []
            for f, row_found in zip(fichier_prev, rows_prev):
                try:
                    df_prev = pd.read_excel(
                        f, sheet_name=feuille_prev_name, header=None, skiprows=row_found - 1
                    )
                    raw_noms  = df_prev.iloc[0, 5:]
                    mask_valid = raw_noms.notna() & (raw_noms.astype(str).str.strip() != "")
                    noms_prev  = raw_noms[mask_valid].astype(str).str.strip().reset_index(drop=True)
                    valeurs    = df_prev.iloc[:, 5:].loc[:, mask_valid.values].reset_index(drop=True)
                    valeurs.columns = noms_prev
                    frames.append(valeurs.reset_index(drop=True))
                except Exception as e:
                    print(f"[ERROR] prev {os.path.basename(f)}: {e}")
                    frames.append(pd.DataFrame())
            valeurs_prev_by_section[feuille_prev_name] = frames

        # ── 9) Boucle principale ─────────────────────────────────────────────
        os.makedirs(FICHIER_EXCEL_DIR, exist_ok=True)
        total    = len(sections)
        OVERWRITE = False

        for idx_section, section in enumerate(sections, start=1):
            feuille_source    = section["source"]
            feuille_prev_name = section["prev"]
            feuille_dest      = section["dest"]

            _set(
                int(10 + (idx_section - 1) / total * 85),
                f"Section {idx_section}/{total} : {feuille_dest}…",
            )

            try:
                df = pd.read_excel(
                    fichier_source, sheet_name=feuille_source, header=None, skiprows=4
                )
                dates           = df.iloc[0, 4:].dropna().reset_index(drop=True)
                dates_ts        = pd.to_datetime(dates, dayfirst=True)
                lignes_valides  = df.iloc[4:, [2] + list(range(4, df.shape[1]))].dropna(subset=[2])
                noms            = lignes_valides.iloc[:, 0].astype(str).str.strip().reset_index(drop=True)
                valeurs_reelles = lignes_valides.iloc[:, 1:].reset_index(drop=True)
            except Exception as e:
                print(f"[ERROR] Lecture réel section {feuille_source}: {e}")
                continue

            groupes_mois: dict = defaultdict(list)
            for i, ts in enumerate(dates_ts):
                groupes_mois[(ts.year, ts.month)].append(i)

            valeurs_prev_all = valeurs_prev_by_section[feuille_prev_name]

            for (yr, month), idxs in sorted(groupes_mois.items()):
                year_dir = os.path.join(FICHIER_EXCEL_DIR, str(yr))
                os.makedirs(year_dir, exist_ok=True)
                out_path = os.path.join(
                    year_dir,
                    f"Historique_prev_reel_filiales_{yr}_{month:02d}.xlsx",
                )

                if os.path.exists(out_path):
                    try:
                        wb_out = load_workbook(out_path)
                    except Exception:
                        wb_out = Workbook()
                        if "Sheet" in wb_out.sheetnames:
                            wb_out.remove(wb_out["Sheet"])
                        for sec in sections:
                            wb_out.create_sheet(sec["dest"])
                else:
                    wb_out = Workbook()
                    if "Sheet" in wb_out.sheetnames:
                        wb_out.remove(wb_out["Sheet"])
                    for sec in sections:
                        wb_out.create_sheet(sec["dest"])

                if feuille_dest not in wb_out.sheetnames:
                    wb_out.create_sheet(feuille_dest)
                ws_out = wb_out[feuille_dest]

                def _has_payload(_ws):
                    if _ws.max_row <= 3:
                        return False
                    for row in _ws.iter_rows(
                        min_row=4, max_row=_ws.max_row, values_only=True
                    ):
                        if any(v is not None and v != "" for v in row):
                            return True
                    return False

                if not OVERWRITE and _has_payload(ws_out):
                    pass
                else:
                    ws_out.delete_rows(1, ws_out.max_row)
                    sous_titres = ["Dates", "Réel (K€)"]
                    for date_str in dates_prev:
                        sous_titres.append(f"Prévision {date_str} (K€)")
                        sous_titres.append(f"Écart {date_str} (K€)")

                    start_col = 3
                    for idx_flux, nom in enumerate(noms, start=1):
                        col1 = get_column_letter(start_col)
                        colN = get_column_letter(start_col - 1 + len(sous_titres))
                        ws_out.merge_cells(f"{col1}2:{colN}2")
                        tc = ws_out[f"{col1}2"]
                        tc.value     = nom
                        tc.fill      = fill_jaune
                        tc.font      = font_bold
                        tc.alignment = align_center
                        tc.border    = border

                        for i, titre in enumerate(sous_titres):
                            cell           = ws_out.cell(row=3, column=start_col - 1 + i)
                            cell.value     = titre
                            cell.fill      = fill_bleu
                            cell.font      = font_bold
                            cell.alignment = align_center
                            cell.border    = border

                        for r, i_global in enumerate(idxs):
                            rn = 4 + r
                            ws_out.cell(
                                row=rn, column=start_col - 1,
                                value=dates_ts.iloc[i_global].date()
                            ).alignment = align_center
                            ws_out.cell(row=rn, column=start_col - 1).border = border

                            valeur = valeurs_reelles.iloc[idx_flux - 1, i_global]
                            val_k  = round(valeur / 1000) if pd.notna(valeur) else None
                            ws_out.cell(row=rn, column=start_col, value=val_k).alignment = align_center
                            ws_out.cell(row=rn, column=start_col).border = border

                            for j, decal in enumerate(cumul):
                                prev_col  = start_col + 1 + j * 2
                                ecart_col = start_col + 2 + j * 2
                                if j >= len(valeurs_prev_all):
                                    ws_out.cell(row=rn, column=prev_col,  value=None)
                                    ws_out.cell(row=rn, column=ecart_col, value=None)
                                else:
                                    prev_data = valeurs_prev_all[j]
                                    idx_prev  = i_global - decal
                                    try:
                                        val = (
                                            prev_data.iloc[idx_prev, idx_flux - 1]
                                            if 0 <= idx_prev < len(prev_data)
                                            else None
                                        )
                                    except Exception:
                                        val = None
                                    val_prev_k = (
                                        round(val / 1000)
                                        if val is not None and pd.notna(val)
                                        else None
                                    )
                                    ecart = (
                                        val_prev_k - val_k
                                        if (val_prev_k is not None and val_k is not None)
                                        else None
                                    )
                                    ws_out.cell(
                                        row=rn, column=prev_col, value=val_prev_k
                                    ).alignment = align_center
                                    ws_out.cell(row=rn, column=prev_col).border = border
                                    ws_out.cell(
                                        row=rn, column=ecart_col, value=ecart
                                    ).alignment = align_center
                                    ws_out.cell(row=rn, column=ecart_col).border = border

                                    if r == 0:
                                        cl    = get_column_letter(ecart_col)
                                        plage = f"{cl}4:{cl}{3 + len(idxs)}"
                                        ws_out.conditional_formatting.add(
                                            plage,
                                            FormulaRule(
                                                formula=[f"${cl}4<0"],
                                                font=Font(color="FF0000", bold=True),
                                            ),
                                        )
                                        ws_out.conditional_formatting.add(
                                            plage,
                                            FormulaRule(
                                                formula=[f"${cl}4>0"],
                                                font=Font(color="00B050", bold=True),
                                            ),
                                        )

                        start_col += 2 + 2 * len(dates_prev) + 1

                tmp_path = out_path + ".tmp"
                try:
                    wb_out.save(tmp_path)
                    os.replace(tmp_path, out_path)
                except PermissionError as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"[ERROR] PermissionError {out_path}: {e}")
                except Exception as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"[ERROR] Save {out_path}: {e}")
                    traceback.print_exc()

        _set(100, "Import terminé avec succès !")
        _JOBS[job_id]["done"] = True

    except Exception as e:
        traceback.print_exc()
        _JOBS[job_id]["error"]   = str(e)
        _JOBS[job_id]["message"] = f"Erreur inattendue : {e}"
        _JOBS[job_id]["done"]    = True


# ─── Routes ─────────────────────────────────────────────────────────────────

@bp.route("/scan", methods=["POST"])
def scan():
    """
    POST /api/import_profils/scan
    Body: {"year": int, "folder": str}
    Returns: {"files": [{"name": str, "path": str}], "fichier_reel": str|null, "base_donnees_dir": str}
    """
    from pulse_v2.config import BASE_DONNEES_DIR

    data   = request.get_json(force=True) or {}
    year   = int(data.get("year", 0))
    folder = str(data.get("folder", "")).strip()

    if not year or not folder:
        return jsonify({"error": "Paramètres 'year' et 'folder' requis."}), 400

    if not os.path.isdir(folder):
        return jsonify({"error": f"Dossier introuvable : {folder}"}), 400

    pattern = _build_pattern(year)
    paths   = _scan_folder(folder, pattern)

    files = [{"name": os.path.basename(p), "path": p} for p in sorted(paths)]

    fichier_reel = _find_reel(year, BASE_DONNEES_DIR)
    fichier_reel_path = fichier_reel if fichier_reel and os.path.isfile(fichier_reel) else None

    return jsonify({
        "files":           files,
        "fichier_reel":    fichier_reel_path,
        "base_donnees_dir": BASE_DONNEES_DIR,
    })


@bp.route("/launch", methods=["POST"])
def launch():
    """
    POST /api/import_profils/launch
    Body: {"year": int, "files": [str]}
    Returns: {"job_id": str}
    """
    data  = request.get_json(force=True) or {}
    year  = int(data.get("year", 0))
    files = data.get("files", [])

    if not year or not files:
        return jsonify({"error": "Paramètres 'year' et 'files' requis."}), 400

    job_id = str(uuid.uuid4())
    _JOBS[job_id] = {
        "progress": 0,
        "message":  "Initialisation…",
        "done":     False,
        "error":    None,
    }

    t = threading.Thread(
        target=_run_import,
        args=(job_id, year, list(files)),
        daemon=True,
    )
    t.start()

    return jsonify({"job_id": job_id})


@bp.route("/progress/<job_id>", methods=["GET"])
def progress(job_id: str):
    """
    GET /api/import_profils/progress/<job_id>
    Returns: {"progress": 0-100, "message": str, "done": bool, "error": str|null}
    """
    job = _JOBS.get(job_id)
    if job is None:
        return jsonify({"error": "Job introuvable."}), 404

    return jsonify({
        "progress": job.get("progress", 0),
        "message":  job.get("message", ""),
        "done":     job.get("done", False),
        "error":    job.get("error", None),
    })
