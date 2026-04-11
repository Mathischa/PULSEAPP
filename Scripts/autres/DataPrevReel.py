import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import re
from datetime import datetime
import os
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

def chercher_excel_commence_par(root_folder, prefix):
    resultats = []
    for dossier, _, fichiers in os.walk(root_folder):
        for fichier in fichiers:
            if fichier.endswith(".xlsx") and fichier.startswith(prefix):
                resultats.append(os.path.join(dossier, fichier))
    return resultats

def lire_date_c6(fichier, feuille="SA_SNCF"):
    try:
        wb = load_workbook(fichier, read_only=True, data_only=True)
        if feuille not in wb.sheetnames:
            return None
        ws = wb[feuille]
        valeur = ws["C6"].value
        wb.close()
        if isinstance(valeur, datetime):
            return valeur.date()
        elif isinstance(valeur, str):
            try:
                return datetime.strptime(valeur, "%d/%m/%Y").date()
            except:
                return None
        return None
    except:
        return None

def choisir_meilleure_version(fichiers):
    fichiers = sorted(fichiers)
    v3 = [f for f in fichiers if "V3" in os.path.basename(f)]
    if v3:
        return v3[0]
    v2 = [f for f in fichiers if "V2" in os.path.basename(f)]
    if v2:
        return v2[0]
    return fichiers[0] if fichiers else None

# --- Utilisation ---
dossier_depart = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Unité Reporting\Prévisions de trésorerie\Production\1 PROFIL SNCF\Profil 2025"

fichiers_trouves = chercher_excel_commence_par(dossier_depart, "Profil Tréso SNCF 2025")

# Récupérer dates par fichier
dates_par_fichier = []
for f in fichiers_trouves:
    date_c6 = lire_date_c6(f)
    if date_c6:
        dates_par_fichier.append((f, date_c6))

# Grouper par date
groupes = defaultdict(list)
for fichier, date in dates_par_fichier:
    groupes[date].append(fichier)

# Choisir une seule version par date
dates_uniques = []
for date, fichiers in groupes.items():
    meilleur_fichier = choisir_meilleure_version(fichiers)
    if meilleur_fichier:
        dates_uniques.append((meilleur_fichier, date))

# Trier les résultats par date
dates_uniques.sort(key=lambda x: x[1])

# Calcul des différences
deltas = []
for i in range(1, len(dates_uniques)):
    fichier1, date1 = dates_uniques[i-1]
    fichier2, date2 = dates_uniques[i]
    delta = (date2 - date1).days
    deltas.append(delta)

# Calcul cumulatif en incluant 0
cumul = [0]
jours = [0]
somme = 0
for d in deltas:
    somme += d
    cumul.append(somme)
    jours.append(d)

# --- Affichage ---
print("📂 Fichiers retenus avec leurs dates et cumuls :")
for (fichier, date), c in zip(dates_uniques, cumul):
    print(f"- {fichier} → {date.strftime('%d/%m/%Y')} | Cumul = {c} jours")

print("\n⏱️ Différences entre fichiers :")
for j, c in zip(jours, cumul):
    print(f"{j} jours | Cumul = {c} jours")

# === Chemins des fichiers ===
fichier_source = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Données\Données Cashsolve 0101 310525.xlsx"

fichier_prev = [fichier for fichier, _ in dates_uniques]

fichier_destination = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Résultats\Historique_prev_reel_filiales2.xlsx"

# === Feuilles à traiter ===
sections = [
    {"source": "1_-_SA_SNCF", "prev": "SA_SNCF", "dest": "SNCF_SA"},
    {"source": "3_-_SA_VOYAGEURS", "prev": "SA_VOYAGEURS", "dest": "SA_VOYAGEURS"},
    {"source": "2_-_RESEAU", "prev": "RESEAU", "dest": "RESEAU"},
    {"source": "4_-_G&amp;amp;C", "prev": "G&C", "dest": "G&C"},
    {"source": "5_-_GEODIS", "prev": "GEODIS", "dest": "GEODIS"},
]

# === Styles ===
fill_jaune = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
fill_bleu = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))
align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
font_bold = Font(bold=True)

# (tout ce qui précède est inchangé...)


# Extraire les dates des chemins
dates_prev = []
for path in fichier_prev:
    match = re.search(r"\d{4}-\d{2}-\d{2}", path)
    if match:
        date_obj = datetime.strptime(match.group(), "%Y-%m-%d")
        dates_prev.append(date_obj.strftime("%d/%m"))  # Format JJ/MM
    else:
        dates_prev.append("N/A")
        
nb_prev = len(fichier_prev)
taille_bloc = 2 + 2 * nb_prev
# === Boucle principale ===
for section in sections:
    feuille_source = section["source"]
    feuille_prev = section["prev"]
    feuille_dest = section["dest"]

    df = pd.read_excel(fichier_source, sheet_name=feuille_source, header=None, skiprows=4)
    dates = df.iloc[0, 4:].dropna().reset_index(drop=True)
    lignes_valides = df.iloc[4:, [2] + list(range(4, df.shape[1]))].dropna(subset=[2])
    noms = lignes_valides.iloc[:, 0].astype(str).str.strip().reset_index(drop=True)
    valeurs_reelles = lignes_valides.iloc[:, 1:].reset_index(drop=True)

    wb = load_workbook(fichier_destination)
    ws = wb[feuille_dest]

    start_col = 3
    for idx, nom in enumerate(noms):
        col1 = get_column_letter(start_col)
        col10 = get_column_letter(start_col + 9)
        ws.merge_cells(f"{col1}2:{col10}2")
        titre_cell = ws[f"{col1}2"]
        titre_cell.value = nom
        titre_cell.fill = fill_jaune
        titre_cell.font = font_bold
        titre_cell.alignment = align_center
        titre_cell.border = border

        sous_titres = ["Dates", "Réel (K€)"]
        for i, date_str in enumerate(dates_prev):
            sous_titres.append(f"Prévision {date_str} (K€)")
            sous_titres.append(f"Écart {date_str} (K€)")

        for i, titre in enumerate(sous_titres):
            cell = ws.cell(row=3, column=start_col - 1 + i)
            cell.value = titre
            cell.fill = fill_bleu
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = border

        for i, date in enumerate(dates):
            row = 4 + i
            ws.cell(row=row, column=start_col - 1, value=pd.to_datetime(date, dayfirst=True).date()).alignment = align_center
            ws.cell(row=row, column=start_col - 1).border = border

            valeur = valeurs_reelles.iloc[idx, i]
            val_k = round(valeur / 1000) if pd.notna(valeur) else None
            ws.cell(row=row, column=start_col, value=val_k).alignment = align_center
            ws.cell(row=row, column=start_col).border = border

        start_col += taille_bloc + 1

    valeurs_prev_all = []
    for f in fichier_prev:
        df_prev = pd.read_excel(f, sheet_name=feuille_prev, header=None, skiprows=5)
        raw_noms = df_prev.iloc[0, 5:]
        mask_valid = raw_noms.notna() & (raw_noms.astype(str).str.strip() != "")
        noms_prev = raw_noms[mask_valid].astype(str).str.strip().reset_index(drop=True)
        valeurs = df_prev.iloc[:, 5:].loc[:, mask_valid.values].reset_index(drop=True)
        valeurs.columns = noms_prev
        valeurs_prev_all.append(valeurs.reset_index(drop=True))

    nb_lignes = len(dates)
    start_col = 3
    for idx in range(len(noms)):
        for i in range(nb_lignes):
            row = 4 + i
            val_reel = ws.cell(row=row, column=start_col).value

            for j, decal in enumerate(cumul):
                prev_col = start_col + 1 + j * 2
                ecart_col = start_col + 2 + j * 2
                if i < decal or j >= len(valeurs_prev_all):
                    ws.cell(row=row, column=prev_col, value=None)
                    ws.cell(row=row, column=ecart_col, value=None)
                else:
                    prev_data = valeurs_prev_all[j]
                    val = prev_data.iloc[i - decal, idx] if (i - decal < len(prev_data)) else None
                    val_k = round(val / 1000) if pd.notna(val) else None
                    ecart = val_k - val_reel if val_k is not None and val_reel is not None else None

                    ws.cell(row=row, column=prev_col, value=val_k).alignment = align_center
                    ws.cell(row=row, column=prev_col).border = border

                    ws.cell(row=row, column=ecart_col, value=ecart).alignment = align_center
                    ws.cell(row=row, column=ecart_col).border = border

                    col_letter = get_column_letter(ecart_col)
                    plage = f"{col_letter}4:{col_letter}{3 + nb_lignes}"
                    rule_rouge = FormulaRule(formula=[f"${col_letter}4<0"], font=Font(color="FF0000", bold=True))
                    rule_vert = FormulaRule(formula=[f"${col_letter}4>0"], font=Font(color="00B050", bold=True))
                    ws.conditional_formatting.add(plage, rule_rouge)
                    ws.conditional_formatting.add(plage, rule_vert)

        start_col += taille_bloc + 1

    wb.save(fichier_destination)

    print(f"✅ Feuille {feuille_dest} traitée avec succès.")
