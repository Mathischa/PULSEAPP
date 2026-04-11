# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict
import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter  # (si nécessaire ailleurs)

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# === Répertoire contenant les fichiers mensuels ===
from pathlib import Path
import os
import os, time, subprocess, unicodedata
from pathlib import Path
import winreg



from pathlib import Path
import os
import unicodedata
# === Modèle ML (scikit-learn) ===
try:
    from sklearn.ensemble import RandomForestRegressor
except ImportError:
    RandomForestRegressor = None
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor

USER_ID = Path.home().name
BASE_DIR = Path(fr"C:\Users\{USER_ID}\SNCF")
import tkinter as tk

# --- PATCH pour empêcher les RuntimeError "main thread is not in main loop" ---

# Sauvegarde des __del__ originaux
_orig_image_del = getattr(tk.Image, "__del__", None)
_orig_var_del = getattr(tk.Variable, "__del__", None)

def _safe_image_del(self):
    if _orig_image_del is None:
        return
    try:
        _orig_image_del(self)
    except RuntimeError:
        # On ignore seulement le cas où Tk n'est plus dans la boucle principale
        pass

def _safe_var_del(self):
    if _orig_var_del is None:
        return
    try:
        _orig_var_del(self)
    except RuntimeError:
        # Pareil ici : on ne fait rien si Tk est déjà "mort"
        pass

if _orig_image_del is not None:
    tk.Image.__del__ = _safe_image_del

if _orig_var_del is not None:
    tk.Variable.__del__ = _safe_var_del

# On n’impose QUE la "queue" du chemin (fin de chemin), flexible sur tout le reste.
REQUIRED_TAIL = ["Partage - Invités", "Projet PULSE", "4. Données historiques", "Développement","Données"]

def _norm(s: str) -> str:
    """Normalise pour comparaisons: minuscules, accents retirés, espaces/underscores compressés."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("_", " ").strip()
    s = " ".join(s.split())
    return s

def _match_tail(path: Path, tail: list[str]) -> bool:
    """Vérifie que 'path' se termine par les éléments de 'tail' (nom pour nom), comparés normalisés."""
    parts = list(path.parts)
    if len(parts) < len(tail):
        return False
    # Compare en partant de la fin
    for i in range(1, len(tail)+1):
        if _norm(parts[-i]) != _norm(tail[-i]):
            return False
    return True

def find_sharepoint_base():
    """
    1) Si ton chemin 'par défaut' existe, on l’utilise.
    2) Sinon, on parcourt TOUT 'C:\\Users\\<USER_ID>\\SNCF' pour trouver un dossier
       qui SE TERMINE par 'Partage - Invités/Projet PULSE/4. Données historiques/Développement'.
    """
    # 1) Essai direct sur ton chemin "canonique"
    default_path = BASE_DIR / "DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions" / Path(*REQUIRED_TAIL)
    if default_path.exists():
        print(f"[OK] Bibliothèque détectée (DCF par défaut) : {default_path}")
        return default_path

    # 2) Scan générique : on ne regarde QUE la fin de chemin (tail), robustes aux variations amont
    print("[INFO] Recherche générique de    la queue de chemin '.../Partage - Invités/Projet PULSE/4. Données historiques/Développement'...")
    candidates = []
    for root, dirs, _ in os.walk(BASE_DIR):
        # accélère: on ne teste que si "Développement" est présent dans le dossier courant
        if any(_norm(d) == _norm(REQUIRED_TAIL[-1]) for d in dirs):
            for d in dirs:
                full = Path(root) / d
                if _match_tail(full, REQUIRED_TAIL):
                    candidates.append(full)

    if candidates:
        # Choix simple : on prend le plus long chemin (le plus profond) ou simplement le premier trouvé
        best = sorted(candidates, key=lambda p: len(p.parts), reverse=True)[0]
        print(f"[OK] Bibliothèque trouvée automatiquement : {best}")
        return best

    print("❌ Aucune bibliothèque SharePoint locale trouvée correspondant à la queue requise.")
    return None

# --- Résolution finale (inchangé) ---
DEV_PATH = find_sharepoint_base()
if not DEV_PATH:
    raise FileNotFoundError("Impossible de localiser la bibliothèque SharePoint (vérifie la synchro OneDrive/SharePoint).")

FICHIER_EXCEL_DIR = str(DEV_PATH / "Historique Prévisions Réel Filiales")
FICHIER_CONFIG_SECTIONS = str(DEV_PATH / "Filiales Analysées.xlsx")
image_path = str(DEV_PATH / "Images" / "logo_Pulse.png")
BASE_DONNEES_DIR = str(DEV_PATH / "Données Réelles")


print("\n[INIT] Dossier Résultats :", FICHIER_EXCEL_DIR)
print("[INIT] Fichier Config    :", FICHIER_CONFIG_SECTIONS)
print("[INIT] Image            :", image_path)
print("[INIT] Dossier Données   :\n", BASE_DONNEES_DIR)

if not Path(FICHIER_EXCEL_DIR).exists():
    print(f"⚠️  Dossier 'Résultats' introuvable : {FICHIER_EXCEL_DIR}")
if not Path(FICHIER_CONFIG_SECTIONS).exists():
    print(f"⚠️  Fichier 'Filiales Analysées.xlsx' introuvable : {FICHIER_CONFIG_SECTIONS}")
if not Path(image_path).exists():
    print(f"⚠️  Image 'logo_Pulse.png' introuvable : {image_path}")

# ======================
#   FONCTIONS DE SÉCURISATION
# ======================

def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")

def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)

def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(["attrib", path], capture_output=True, text=True, shell=True)
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None

def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")

def pin_and_hydrate(path: str, timeout=30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l’hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False

def robust_load_workbook(path: str, retries=3, delay=1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} ({type(e).__name__}: {e})")
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e

def diag_path(path: str):
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
    
# === Fichier Excel contenant la configuration des sections ===
FEUILLE_CONFIG_SECTIONS = "CONFIG_SECTIONS"  # ou le vrai nom de la feuille dans ton fichier
COL_DEST = 1   # A
COL_SOURCE = 2 # B
COL_PREV = 3   # C

def charger_sections_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit le fichier Excel de configuration des sections et renvoie une liste
    de dictionnaires {"source": ..., "prev": ..., "dest": ...}.
    """
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return []

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return []

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return []

    ws = wb[FEUILLE_CONFIG_SECTIONS]
    mapping = []

    row = 2  # ligne 1 = en-têtes
    while True:
        dest_val = ws.cell(row=row, column=COL_DEST).value
        source_val = ws.cell(row=row, column=COL_SOURCE).value
        prev_val = ws.cell(row=row, column=COL_PREV).value

        if dest_val is None or str(dest_val).strip() == "":
            break  # fin du tableau

        mapping.append({
            "dest": str(dest_val).strip(),
            "source": str(source_val).strip() if source_val else str(dest_val).strip(),
            "prev": str(prev_val).strip() if prev_val else str(dest_val).strip(),
        })
        row += 1

    wb.close()

    print("[INFO] Sections chargées depuis 'Filiales Analysées.xlsx' :")
    for m in mapping:
        print(f"   - dest='{m['dest']}' | source='{m['source']}' | prev='{m['prev']}'")

    return mapping

def charger_noms_feuilles_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit uniquement la colonne A (dest) de la feuille de config et
    retourne la liste des noms de feuilles à utiliser (strings non vides, uniques, ordonnées).
    """
    noms = []
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return noms

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return noms

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return noms

    ws = wb[FEUILLE_CONFIG_SECTIONS]

    row = 2  # ligne 1 = en-têtes
    seen = set()
    while True:
        val = ws.cell(row=row, column=COL_DEST).value
        if val is None or str(val).strip() == "":
            break
        name = str(val).strip()
        if name not in seen:
            noms.append(name)
            seen.add(name)
        row += 1

    wb.close()

    print("[INFO] Feuilles (dest) chargées depuis 'Filiales Analysées.xlsx' :")
    for n in noms:
        print(f"   - {n}")
    return noms

# === Construire dynamiquement `sections` à partir de la colonne A ===
_dest_names = charger_noms_feuilles_depuis_cells()

# Création directe du dictionnaire sections
sections = {name: name for name in _dest_names}  # clé logique -> nom de feuille Excel


# Optionnel mais pratique : caler la feuille de référence sur la première feuille disponible
if _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]   # ex: "SNCF_SA"

SECTIONS_CONFIG = charger_sections_depuis_cells()
print(SECTIONS_CONFIG)

print(FICHIER_EXCEL_DIR)

# === Regex utilitaires ===
_MENSUEL_RE = re.compile(r"^Historique_prev_reel_filiales_(\d{4})_(\d{2})\.xlsx$")
_RE_PREV = re.compile(r"^Prévision\s+\d{2}/\d{2}", re.I)

def _is_prev(h):  return isinstance(h, str) and _RE_PREV.search(h.strip())
def _is_dates(h): return isinstance(h, str) and h.strip().lower().startswith("date")
def _is_reel(h):  return isinstance(h, str) and "réel" in h.lower()

def _parse_excel_date(v):
    """Convertit proprement tout format de date Excel (float, datetime, str...)"""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        # borne basse/haute pour éviter les overflow absurdes
        if 10000 <= v <= 60000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")).normalize()
            except Exception:
                return None
        else:
            return None
    if isinstance(v, str):
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        return ts.normalize() if pd.notna(ts) else None
    ts = pd.to_datetime(v, errors="coerce")
    return ts.normalize() if pd.notna(ts) else None

# -------------------------------------------------------------------------
# 1️⃣ Structures globales
# -------------------------------------------------------------------------
STRUCT = {}
TOKENS = {}
PREV_UNION = set()
FEUILLE_REFERENCE = "SNCF_SA"
CACHE = {}

# Index annuel: (section, flux) -> { "years": { year: { "row_idx": [...], "prof_idx": [...], "headers": [...] } } }
YEAR_INDEX = {}

def _clean_profil_label(raw, i):
    """Nettoyage identique à l'historique: (K€) retiré, 'Prévision' -> 'Profil', fallback."""
    if raw is None:
        return f"Profil {i+1}"
    s = str(raw).replace("(K€)", "").replace("Prévision", "Profil").strip()
    return s if s else f"Profil {i+1}"

def _flux_name_from_token(section, token_col):
    """Retrouve le nom de flux à partir de la colonne 'token' (col_start) dans TOKENS[section]."""
    for name, tok in TOKENS.get(section, []):
        if tok == token_col:
            return name
    return None


# -------------------------------------------------------------------------
# 2️⃣ Lister les fichiers mensuels
# -------------------------------------------------------------------------

def _lister_fichiers_mensuels():
    """
    Parcourt récursivement FICHIER_EXCEL_DIR pour trouver tous les fichiers
    'Historique_prev_reel_filiales_YYYY_MM.xlsx', y compris dans des sous-dossiers d'années.
    En cas de doublons pour un même (YYYY, MM), conserve le fichier le plus récent (mtime).
    """
    print("🔍 Recherche des fichiers mensuels (scan récursif des sous-dossiers d'années)...")

    if not os.path.isdir(FICHIER_EXCEL_DIR):
        print(f"⚠️ Dossier introuvable : {FICHIER_EXCEL_DIR}")
        return []

    best = {}  # (year, month) -> {"path": str, "mtime": float}
    try:
        for root, _, files in os.walk(FICHIER_EXCEL_DIR):
            for fname in files:
                m = _MENSUEL_RE.match(fname)
                if not m:
                    continue
                y, mth = int(m.group(1)), int(m.group(2))
                fullpath = os.path.join(root, fname)
                try:
                    mtime = os.path.getmtime(fullpath)
                except OSError:
                    mtime = 0.0

                key = (y, mth)
                if key not in best or mtime > best[key]["mtime"]:
                    best[key] = {"path": fullpath, "mtime": mtime}
    except Exception as e:
        print(f"⚠️ Erreur pendant le scan récursif : {e}")
        return []

    if not best:
        print("⚠️ Aucun fichier Historique_prev_reel_filiales_YYYY_MM.xlsx trouvé (sous-dossiers inclus).")
        return []

    files = sorted([(y, mth, info["path"]) for (y, mth), info in best.items()])
    print(f"✅ {len(files)} fichiers retenus : {[os.path.basename(f[2]) for f in files]}")
    return files

def _parse_prev_header_sort_key(h):
    # Essaie d'en déduire un tri MM/YY si possible, sinon garde l’ordre d’apparition
    import re
    s = str(h)
    m = re.search(r'(\d{2})/(\d{2,4})', s)
    if not m:
        return (9999, 99, s)
    mm = int(m.group(1))
    yy = int(m.group(2))
    if yy < 100:
        yy += 2000
    return (yy, mm, s)

def _reconcile_headers(B, new_headers):
    """
    Fusionne les headers déjà présents dans B avec ceux du fichier courant :
    - construit l’union ordonnée
    - re-map B['prev_vals'] dans ce nouvel ordre
    - rétro-remplit de None pour les séries nouvellement apparues
    """
    if "prev_headers" not in B or "prev_vals" not in B:
        B["prev_headers"] = []
        B["prev_vals"] = []

    old_headers = list(B["prev_headers"])
    # union en conservant l’ordre d’apparition, puis tente un tri chronologique
    seen = {}
    for h in old_headers + list(new_headers):
        if h not in seen:
            seen[h] = None
    union = list(seen.keys())
    union.sort(key=_parse_prev_header_sort_key)

    # rien à faire ?
    if union == old_headers:
        return

    # re-map des anciennes séries dans le nouveau canevas
    old_idx = {h: i for i, h in enumerate(old_headers)}
    n_rows = len(B.get("dates", []))
    new_prev_vals = []
    for h in union:
        if h in old_idx:
            new_prev_vals.append(B["prev_vals"][old_idx[h]])
        else:
            # nouvelle série : rétro-remplir pour les lignes déjà connues
            new_prev_vals.append([None] * n_rows)

    B["prev_headers"] = union
    B["prev_vals"] = new_prev_vals

# -------------------------------------------------------------------------
# 3️⃣ Lecture de la structure sur un fichier de référence
# -------------------------------------------------------------------------
def _lire_structure_reference(path_ref):
    print(f"\n📘 Lecture de la structure de référence ({FEUILLE_REFERENCE}) → {os.path.basename(path_ref)}")
    df_head = pd.read_excel(path_ref, sheet_name=FEUILLE_REFERENCE, header=None, nrows=3)
    row2 = df_head.iloc[1].tolist()
    row3 = df_head.iloc[2].tolist()

    ref_od = OrderedDict()
    ref_tokens = []
    col = 2  # colonne C (0-based)
    token_col = 3

    while col < len(row2):
        flux_name = row2[col]
        if pd.isna(flux_name):
            break
        flux_name = str(flux_name).strip()

        prev_headers, prev_cols = [], []
        c = col + 1
        while c < len(row3):
            h = row3[c]
            if isinstance(h, str) and "Prévision" in h:
                prev_headers.append(h.strip())
                prev_cols.append(c)
                c += 2
                continue
            break

        if prev_headers:
            ref_od[flux_name] = {
                "col_start": col + 1,
                "prev_headers": prev_headers,
                "prev_cols": [x + 1 for x in prev_cols],
                "dates_col": col,
                "reel_col": col + 1,
            }
            ref_tokens.append((flux_name, token_col))
            PREV_UNION.update(prev_headers)

        col += (2 + 2 * len(prev_headers) + 1)
        token_col += (2 + 2 * len(prev_headers) + 1)

    print(f"   → {len(ref_od)} flux détectés ({len(PREV_UNION)} prévisions).")
    for sec in sections:
        STRUCT[sec] = ref_od.copy()
        TOKENS[sec] = ref_tokens.copy()
    print("✅ Structure copiée sur toutes les feuilles.\n")

# -------------------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# -------------------------------------------------------------------------
def _ensure_flux_bucket(section, flux_name, headers=None):
    key = (section, flux_name)
    if key not in CACHE:
        CACHE[key] = {
            "dates": [],
            "reel": [],
            "prev_headers": [],   # ← vide
            "prev_vals": [],      # ← vide
        }
    return CACHE[key]

def _accumuler_valeurs_tous_mois(files_list):
    """
    Lit chaque fichier mensuel et alimente CACHE[(section, flux)] en alignant
    les séries de prévision sur l'union ordonnée des en-têtes (headers).
    - Réconciliation des headers à chaque fichier (ajout rétro-rempli de None).
    - Insertion directe ligne par ligne dans B["dates"], B["reel"], B["prev_vals"][k].
    """
    print(f"📊 Lecture de {len(files_list)} fichiers mensuels (toutes feuilles)...")
    for idx, (_, mois, path) in enumerate(files_list, 1):
        print(f"   [{idx}/{len(files_list)}] → {os.path.basename(path)}")

        try:
            all_sheets = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as e:
            print(f"⚠️ Erreur lecture {path}: {e}")
            continue

        for sec, feuille in sections.items():
            if feuille not in all_sheets:
                continue

            df = all_sheets[feuille].copy()
            if df.shape[0] < 5:
                continue

            row2 = df.iloc[1].tolist()
            row3 = df.iloc[2].tolist()
            col = 2  # colonne C (0-based -> ici index pandas)

            while col < len(row2):
                flux_name = row2[col]
                if pd.isna(flux_name):
                    break
                flux_name = str(flux_name).strip()

                # Détection des colonnes "Prévision"
                prev_headers, prev_cols = [], []
                c = col + 1
                while c < len(row3):
                    h = row3[c]
                    if isinstance(h, str) and "Prévision" in h:
                        prev_headers.append(h.strip())
                        prev_cols.append(c)
                        c += 2  # on saute la colonne (k€) qui suit chaque prévision
                        continue
                    break

                if not prev_headers:
                    # Bloc sans prévisions (ex: espaces, colonnes diverses) → on saute "gros"
                    col += 10
                    continue

                # 1) Bucket pour ce flux
                B = _ensure_flux_bucket(sec, flux_name)

                # 2) Réconcilier les en-têtes avec ceux du fichier courant
                _reconcile_headers(B, prev_headers)

                # 3) Map header -> colonne pour CE fichier
                header_to_col = {h: prev_cols[j] for j, h in enumerate(prev_headers)}

                # 4) Pousser les lignes : date, réel, et séries alignées
                rows_appended = 0
                for r in range(3, df.shape[0]):
                    date_val = df.iat[r, col - 1]
                    if pd.isna(date_val):
                        continue
                    d = _parse_excel_date(date_val)
                    if d is None:
                        continue

                    B["dates"].append(d)

                    r_val = df.iat[r, col]
                    B["reel"].append(r_val if pd.notna(r_val) else 0)

                    # Pour chaque header du bucket, si la colonne existe ce mois-ci → valeur ; sinon → None
                    for k, h in enumerate(B["prev_headers"]):
                        if h in header_to_col:
                            cpr = header_to_col[h]
                            v = df.iat[r, cpr]
                            B["prev_vals"][k].append(v if pd.notna(v) else None)
                        else:
                            B["prev_vals"][k].append(None)

                    rows_appended += 1

                # Avance au bloc suivant : (dates + réel) + 2 colonnes par prévision + 1 de séparation
                col += (2 + 2 * len(prev_headers) + 1)

    print(f"✅ Cache complété : {len(CACHE)} flux au total (valeurs réelles et prévisions incluses).\n")

    # Récapitulatif par section (contrôle)
    print("📊 Récapitulatif des flux chargés par section :")
    counts = {}
    for (section, flux), data in CACHE.items():
        counts.setdefault(section, 0)
        counts[section] += 1
    for section, n in counts.items():
        total_lignes = sum(len(CACHE[(section, f)]['dates']) for f in [x for x, _ in CACHE if _ == section])
        print(f"   - {section} : {n} flux ({total_lignes} lignes cumulées)")
    print()

# -------------------------------------------------------------------------
# 5️⃣ Index annuel
# -------------------------------------------------------------------------
def _build_year_index():
    """
    Pour chaque (section, flux) dans CACHE :
      - regroupe les indices de lignes par année
      - détermine quels profils (colonnes) sont 'actifs' sur l'année (au moins une valeur != 0 / non None)
      - prépare la liste de headers nettoyés pour l'année
    Remplit YEAR_INDEX[(section, flux)].
    """
    YEAR_INDEX.clear()
    for (section, flux_name), B in CACHE.items():
        dates = B.get("dates", [])
        prev_vals = B.get("prev_vals", [])
        headers = B.get("prev_headers", [])

        # indices de lignes par année
        rows_by_year = {}
        for i, d in enumerate(dates):
            y = d.year
            rows_by_year.setdefault(y, []).append(i)

        years_map = {}
        for y, row_idx in rows_by_year.items():
            prof_idx = []
            for k, serie in enumerate(prev_vals):
                # activité = au moins une valeur non None et != 0 sur ces lignes
                active = False
                for i in row_idx:
                    if i < len(serie):
                        v = serie[i]
                        if v is None:
                            continue
                        try:
                            if float(v) != 0.0:
                                active = True
                                break
                        except Exception:
                            active = True
                            break
                if active:
                    prof_idx.append(k)

            headers_year = [_clean_profil_label(headers[k] if k < len(headers) else None, k) for k in prof_idx]

            years_map[y] = {
                "row_idx": row_idx,       # indices de lignes pour cette année
                "prof_idx": prof_idx,     # indices de colonnes actives pour cette année
                "headers": headers_year,  # labels propres des profils actifs
            }

        YEAR_INDEX[(section, flux_name)] = {"years": years_map}

# -------------------------------------------------------------------------
# 6️⃣ Initialisation complète
# -------------------------------------------------------------------------
def _init_full_load():
    print("🚀 Initialisation complète du cache de données...")
    files = _lister_fichiers_mensuels()
    if not files:
        print("ℹ️ Aucun fichier trouvé — initialisation du cache ignorée.\n")
        return

    ref_path = files[-1][2]  # dernier mois = référence
    _lire_structure_reference(ref_path)
    _accumuler_valeurs_tous_mois(files)
    _build_year_index()  # ✅ construire l’index après remplissage du CACHE
    print("✅ Chargement complet terminé.\n")


# Lance le chargement maintenant que tout est défini
_init_full_load()

# -------------------------------------------------------------------------
# 7️⃣ Variables dérivées compatibles avec l'existant
# -------------------------------------------------------------------------
previsions_triees = sorted(PREV_UNION)
nb_prev = len(previsions_triees)
taille_bloc = 2 + 2 * nb_prev + 1
print(f"📈 Nombre total de prévisions : {nb_prev}")
print(f"📦 Taille d’un bloc de flux : {taille_bloc}\n")


# -------------------------------------------------------------------------
# 8️⃣ Fonctions publiques
# -------------------------------------------------------------------------
def charger_donnees(feuille, taille_bloc_param):
    """Retourne le nom de la feuille (section) et la liste des flux détectés, en excluant certains flux inutiles."""
    print(f"➡️ charger_donnees() appelé pour '{feuille}'")

    # Liste des flux à ignorer
    noms_a_exclure = [
        "Trésorerie de fin",
        "Cashpool",
        "Emprunts",
        "Tirages Lignes CT",
        "Variation de collatéral",
        "Créances CDP",
        "Placements",
        "CC financiers",
        "Emprunts / Prêts - Groupe",
        "Encours de financement",
        "Endettement Net"
    ]

    # Récupère tous les flux
    all_noms = list(TOKENS.get(feuille, []))

    # Filtrage : on garde uniquement ceux qui ne sont pas exclus
    noms = [(name, col) for name, col in all_noms if name not in noms_a_exclure]
    for j in noms :
        print(f"noms : {j}")
    print(len(noms))
    
    if not noms:
        print(f"⚠️ Aucun flux valide trouvé pour la feuille '{feuille}'")
    else:
        print(f"   ✅ {len(noms)} flux conservés (sur {len(all_noms)} totaux)")

    return feuille, noms
# Profils à exclure pour 2023 (d’après tes listes / cases à cocher)
EXCLUDED_DATES_2023 = {
    "02/01",
    "09/01",
    "16/01",
    "30/05",
    "05/06",
    "19/06",
    "26/06",
    "03/07",
    "10/07",
    "17/07",
}

def extraire_valeurs(ws, col_start, nb_prev_param, annee=None, annee_min=None, annee_max=None):
    """
    Retourne (dates, reel, previsions, noms_profils)

    - Si 'annee' est fourni : utilise YEAR_INDEX pour ne garder que
      *les lignes* de l'année et *les profils actifs* de l'année,
      puis on filtre éventuellement certains profils (2023) → le RÉEL reste inchangé.
    - Sinon : filtre sur plage ou aucun filtre (comportement historique).
    """

    DEBUG_EXTRAIRE = True
    def _dbg(*args):
        if DEBUG_EXTRAIRE:
            print("[extraire_valeurs]", *args)

    section = ws
    flux_name = _flux_name_from_token(section, col_start)
    _dbg(f"CALL: section={section!r} col_start={col_start} nb_prev_param={nb_prev_param} "
         f"annee={annee} annee_min={annee_min} annee_max={annee_max}")

    if not flux_name:
        _dbg("ABORT: flux_name introuvable -> retour listes vides/pad prév.")
        return [], [], [[] for _ in range(nb_prev_param)], []

    B = CACHE.get((section, flux_name))
    if not B or not B.get("dates"):
        _dbg(f"ABORT: pas de cache ou dates vides pour {(section, flux_name)}")
        return [], [], [[] for _ in range(nb_prev_param)], []

    dates_all = B["dates"]
    reel_all  = B["reel"]
    prev_all  = B.get("prev_vals", [])
    headers   = B.get("prev_headers", [])   # ex: "Prévision 16/01 (K€)"

    _dbg(f"CACHE: len(dates_all)={len(dates_all)}, len(reel_all)={len(reel_all)}, "
         f"nb_profils_total={len(prev_all)}, nb_headers={len(headers)}")

    # ==============================
    # CAS 1 : année précise
    # ==============================
    if annee is not None:
        years = YEAR_INDEX.get((section, flux_name), {}).get("years", {})
        info = years.get(annee)
        if not info:
            _dbg(f"ANNEE={annee}: YEAR_INDEX absent/vide -> retour listes vides.")
            return [], [], [], []

        # Ces indices NE CHANGENT PAS pour les lignes → le RÉEL reste intact
        row_idx  = info["row_idx"]      # indices des dates pour cette année
        prof_idx = info["prof_idx"]     # indices des colonnes de prévision actives

        _dbg(f"ANNEE={annee}: rows={len(row_idx)} profils_actifs={len(prof_idx)}")

        # ---- Filtre spécial 2023 : on enlève certains profils, mais on garde le réel ----
        if annee == 2023:
            def _is_excluded(k: int) -> bool:
                if k < 0 or k >= len(headers):
                    return False
                lab = str(headers[k])  # "Prévision 16/01 (K€)"
                for d in EXCLUDED_DATES_2023:
                    if d in lab:
                        return True
                return False

            prof_idx_before = list(prof_idx)
            prof_idx = [k for k in prof_idx if not _is_excluded(k)]
            _dbg(f"ANNEE=2023: prof_idx avant filtre = {prof_idx_before}")
            _dbg(f"ANNEE=2023: prof_idx après  filtre = {prof_idx}")

        # ---- Extraction des lignes (réel inchangé) ----
        dates = [dates_all[i] for i in row_idx]
        reel  = [reel_all[i]  for i in row_idx]

        # ---- Extraction des prévisions uniquement pour les profils retenus ----
        previsions = []
        for k in prof_idx:
            serie = prev_all[k] if k < len(prev_all) else []
            previsions.append([serie[i] if i < len(serie) else None for i in row_idx])

        # Noms de profils alignés sur prof_idx
        noms_profils = [
            _clean_profil_label(headers[k] if k < len(headers) else None, k)
            for k in prof_idx
        ]

        _dbg(f"RETURN(annee): len(dates)={len(dates)}, len(reel)={len(reel)}, "
             f"nb_profils={len(previsions)}, noms_profils={noms_profils}")
        return dates, reel, previsions, noms_profils

    # ==============================
    # CAS 2 : plage / historique
    # ==============================
    def _keep(d):
        y = d.year
        if annee_min is not None and annee_max is not None:
            return annee_min <= y <= annee_max
        if annee_min is not None:
            return y >= annee_min
        if annee_max is not None:
            return y <= annee_max
        return True

    idxs = [i for i, d in enumerate(dates_all) if _keep(d)]
    dates = [dates_all[i] for i in idxs]
    reel  = [reel_all[i]  for i in idxs]
    previsions = [[serie[i] if i < len(serie) else None for i in idxs] for serie in prev_all]

    noms_profils = [_clean_profil_label(h, i) for i, h in enumerate(headers)]

    _dbg(f"PLAGE/HISTO: indices_retenus={len(idxs)} "
         f"(min_year={min((d.year for d in dates), default=None)}, "
         f"max_year={max((d.year for d in dates), default=None)})")
    _dbg(f"RETURN(plage): len(dates)={len(dates)}, len(reel)={len(reel)}, "
         f"nb_profils={len(previsions)}, len(noms_profils)={len(noms_profils)}; "
         f"exemple_date={dates[0] if dates else None}")

    return dates, reel, previsions, noms_profils


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        self.attributes("-fullscreen", True)
        self.geometry("1300x900")
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

#===================Page Accueil + fonctions de navigation===================
    def creer_accueil(self):
        # -- RESET & imports locaux
        import customtkinter as ctk
        from PIL import Image
        import tkinter as tk
        from customtkinter import CTkImage
        from datetime import datetime
        import os

        # === THEME / BASE ===
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color="#0b1220")
        except Exception:
            pass

        # Détruit proprement les enfants
        self.vider_fenetre()

        # =========================
        #  Helpers data pour KPI
        # =========================
        def _fmt_dt(ts):
            try:
                return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return "—"

        # Fichiers mensuels
        try:
            fichiers = _lister_fichiers_mensuels()  # supposé renvoyer (annee, mois, path)
            nb_fichiers = len(fichiers)

            if fichiers:
                last_mtime = max(os.path.getmtime(p) for _, _, p in fichiers if os.path.exists(p))
                derniere_maj = _fmt_dt(last_mtime)
                # années disponibles (1er champ du tuple)
                annees_dispos = sorted({int(a) for a, _, _ in fichiers if a is not None})
            else:
                derniere_maj = "—"
                annees_dispos = []
        except Exception:
            nb_fichiers, derniere_maj = 0, "—"
            annees_dispos = []

        # Filiales
        try:
            if SECTIONS_CONFIG:
                filiales_set = {s.get("dest") for s in SECTIONS_CONFIG if s.get("dest")}
            else:
                filiales_set = set(_dest_names) if '_dest_names' in globals() else set()
            filiales_list = sorted(f for f in filiales_set if f)
            nb_filiales = len(filiales_list)
        except Exception:
            filiales_list = []
            nb_filiales = 0

        # =========================
        #  GRID RACINE (window)
        # =========================
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        # =========================
        #  HEADER
        # =========================
        header = ctk.CTkFrame(self, fg_color="#0f1b31", corner_radius=0, height=72)
        header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        header.grid_columnconfigure(0, weight=0)
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        # Barre blanche sous le header
        barre_header = ctk.CTkFrame(self, fg_color="white", height=2)
        barre_header.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Logo à gauche
        logo_wrap = ctk.CTkFrame(header, fg_color="transparent")
        logo_wrap.grid(row=0, column=0, sticky="w", padx=24, pady=12)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h, new_w = 44, int(44 * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(logo_wrap, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            ctk.CTkLabel(logo_wrap, text="PULSE", font=("Segoe UI Semibold", 20),
                        text_color="white").pack(side="left")

        titre = ctk.CTkLabel(
            header,
            text="PROJET PULSE — ANALYSE DE L’EXISTANT",
            font=("Segoe UI Semibold", 18, "bold"),
            text_color="white",
        )
        titre.grid(row=0, column=1, sticky="w", padx=8)

        # Actions header à droite
        header_actions = ctk.CTkFrame(header, fg_color="transparent")
        header_actions.grid(row=0, column=2, sticky="e", padx=20)

        ctk.CTkButton(
            header_actions, text="❌ Quitter",
            fg_color="#C21515", hover_color="#A00000",
            command=self.demander_confirmation_quit
        ).grid(row=0, column=1)

        # =========================
        #  SIDEBAR
        # =========================
        sidebar = ctk.CTkFrame(self, fg_color="#0e162a", corner_radius=0, width=260)
        sidebar.grid(row=1, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self.grid_rowconfigure(1, weight=1)
        sidebar.grid_rowconfigure(11, weight=1)

        inner = ctk.CTkFrame(sidebar, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=14, pady=14)

        def _nav_btn(parent, text, cmd, emoji=""):
            btn = ctk.CTkButton(
                parent, text=f"{emoji}  {text}", anchor="w",
                width=220, height=42, corner_radius=12,
                fg_color="#15305b", hover_color="#1d3e77",
                font=("Segoe UI", 13, "bold"),
                command=cmd
            )

            def _enter(_):
                btn.configure(fg_color="#1d3e77")

            def _leave(_):
                btn.configure(fg_color="#15305b")

            btn.bind("<Enter>", _enter)
            btn.bind("<Leave>", _leave)
            return btn

        ctk.CTkLabel(inner, text="Navigation", text_color="#9fb7dd",
                    font=("Segoe UI Semibold", 14)).pack(anchor="w", pady=(4, 12))

        _nav_btn(inner, "IA — Prédiction", self.creer_page_ia_prediction, "🤖").pack(fill="x", pady=6)

        separator = ctk.CTkFrame(sidebar, fg_color="#1a2745", height=2)
        separator.pack(fill="x", padx=14, pady=(4, 2))

        foot = ctk.CTkLabel(
            sidebar,
            text=f"v1.0 • Dernière MAJ : {derniere_maj}",
            text_color="#6e86ad",
            font=("Segoe UI", 11)
        )
        foot.pack(anchor="w", padx=20, pady=(8, 10))

        # =========================
        #  MAIN CONTENT
        # =========================
        main = ctk.CTkScrollableFrame(self, fg_color="#0b1220", corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        self.grid_columnconfigure(1, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # ====== HERO ======
        hero = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 14))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(
            hero,
            text="Bienvenue 👋",
            font=("Segoe UI Semibold", 24, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 0))

        subtitle_text = (
            "Console d'analyse des flux de trésorerie : réels, prévisions, écarts et détection d'anomalies."
        )
        subtitle_label = ctk.CTkLabel(
            hero,
            text="",
            font=("Segoe UI", 13),
            text_color="#c9defa",
            wraplength=520,
            justify="left"
        )
        subtitle_label.grid(row=1, column=0, sticky="w", padx=18, pady=(6, 10))

        def _typewriter(i=0):
            if i <= len(subtitle_text):
                subtitle_label.configure(text=subtitle_text[:i])
                self.after(12, _typewriter, i + 1)

        _typewriter()

        # Pills
        pills = ctk.CTkFrame(hero, fg_color="transparent")
        pills.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 16))

        def _pill(parent, text, emoji=""):
            f = ctk.CTkFrame(parent, fg_color="#152544", corner_radius=20)
            f.pack(side="left", padx=6)
            ctk.CTkLabel(
                f,
                text=f"{emoji}  {text}",
                font=("Segoe UI", 11),
                text_color="#c9defa"
            ).pack(padx=12, pady=5)

        _pill(pills, "Réel vs prévisions multi-horizon", "📈")
        _pill(pills, "Analyse d'écarts & anomalies", "🧠")
        _pill(pills, "Vue filiales / flux / profils", "📊")

        # Illustration graphique (3 mini graphes)
        viz = ctk.CTkFrame(hero, fg_color="#111b33", corner_radius=18)
        viz.grid(row=0, column=1, rowspan=3, sticky="e", padx=18, pady=12)
        viz.grid_columnconfigure((0, 1, 2), weight=1)
        viz.grid_rowconfigure(0, weight=1)

        # --- Graph 1 : barres animées ---
        graph1 = ctk.CTkFrame(viz, fg_color="transparent")
        graph1.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        bars_h = [52, 34, 78, 60]
        bar_widgets = []
        for h_val in bars_h:
            col = ctk.CTkFrame(graph1, fg_color="transparent")
            col.pack(side="left", padx=4, fill="y", expand=True)

            base = ctk.CTkFrame(col, fg_color="#1b2945", corner_radius=8, width=18, height=85)
            base.pack(side="bottom", pady=(0, 4))

            bar = ctk.CTkFrame(col, fg_color="#2563eb", corner_radius=8, width=18, height=1)
            bar.place(relx=0.5, rely=1.0, anchor="s")
            bar_widgets.append((bar, h_val))

        def _animate_bars(step=0, steps=18):
            factor = min(1.0, step / float(steps))
            for bar, target in bar_widgets:
                bar.configure(height=int(target * factor))
            if step < steps:
                self.after(25, _animate_bars, step + 1, steps)

        _animate_bars()

        # --- Graph 2 : courbe ---
        graph2 = ctk.CTkFrame(viz, fg_color="transparent")
        graph2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        pts = [10, 40, 22, 70, 50, 90, 60, 80, 55, 100]
        w, h = 120, 90

        canvas2 = ctk.CTkCanvas(graph2, width=w, height=h, bg="#111b33", bd=0, highlightthickness=0)
        canvas2.pack(fill="both", expand=True)

        max_v = max(pts)
        coords = [(i * (w // (len(pts) - 1)), h - int((v / max_v) * (h - 10))) for i, v in enumerate(pts)]
        for i in range(len(coords) - 1):
            canvas2.create_line(
                coords[i][0], coords[i][1],
                coords[i + 1][0], coords[i + 1][1],
                fill="#2563eb", width=3, smooth=True
            )
        for x, y in coords:
            canvas2.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#3c82ff", outline="")

        # --- Graph 3 : donut ---
        graph3 = ctk.CTkFrame(viz, fg_color="transparent")
        graph3.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)

        canvas3 = tk.Canvas(graph3, width=110, height=110,
                            bg="#111b33", bd=0, highlightthickness=0)
        canvas3.pack(fill="both", expand=True, padx=6, pady=6)

        values = [30, 15, 22, 18, 15]
        total = sum(values)
        start_angle = 0
        colors = ["#1d4ed8", "#2563eb", "#3b82f6", "#60a5fa", "#93c5fd"]
        bbox_outer = (8, 8, 102, 102)

        for i, v in enumerate(values):
            extent = 360 * v / total
            canvas3.create_arc(
                *bbox_outer,
                start=start_angle,
                extent=extent,
                fill=colors[i % len(colors)],
                outline="#111b33",
                style="pieslice"
            )
            start_angle += extent

        bbox_inner = (28, 28, 82, 82)
        canvas3.create_oval(*bbox_inner, fill="#111b33", outline="#111b33")

        # ====== KPI CARDS ======
        kpi = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        kpi.grid(row=1, column=0, sticky="ew", padx=24, pady=10)
        for i in range(3):
            kpi.grid_columnconfigure(i, weight=1)

        # ----- Zone "Détails" dynamique juste sous les KPI -----
        details = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        details.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 10))
        details.grid_columnconfigure(0, weight=1)

        details_title = ctk.CTkLabel(
            details,
            text="Détails",
            font=("Segoe UI Semibold", 15, "bold"),
            text_color="white"
        )
        details_title.grid(row=0, column=0, sticky="w", padx=16, pady=(14, 4))

        details_box = ctk.CTkTextbox(
            details,
            height=110,
            wrap="word",
            fg_color="#142544",
            text_color="white",
            font=("Segoe UI", 12),
            border_width=1,
            border_color="#223658",
            corner_radius=12
        )
        details_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 14))
        details_box.insert("1.0", "Clique sur un KPI pour afficher les détails (années ou filiales)…")
        details_box.configure(state="disabled")

        def _set_details(title: str, lines: list[str]):
            details_title.configure(text=title)
            details_box.configure(state="normal")
            details_box.delete("1.0", "end")
            if not lines:
                details_box.insert("1.0", "Aucune donnée disponible.")
            else:
                details_box.insert("1.0", "\n".join(lines))
            details_box.configure(state="disabled")

        def _show_years():
            if annees_dispos:
                lignes = ["Années disponibles :"]
                lignes += [f"• {a}" for a in annees_dispos]
            else:
                lignes = ["Aucune année détectée (pas de fichiers mensuels trouvés)."]
            _set_details("Détails — Fichiers mensuels détectés", lignes)

        def _show_filiales():
            if filiales_list:
                lignes = ["Filiales configurées :"]
                lignes += [f"• {f}" for f in filiales_list]
            else:
                lignes = ["Aucune filiale détectée dans la configuration."]
            _set_details("Détails — Filiales", lignes)

        def _kpi_card(parent, title, value, subtitle, col, detail_cb=None):
            card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16,
                                border_width=1, border_color="#223658")
            card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
            ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
            ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 28, "bold"),
                        text_color="white") \
                .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 4))
            ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                        text_color="#7ea2d8") \
                .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 6))

            if detail_cb is not None:
                btn = ctk.CTkButton(
                    card,
                    text="Voir détails",
                    height=28,
                    corner_radius=10,
                    fg_color="#2563eb",
                    hover_color="#1d4ed8",
                    command=detail_cb
                )
                btn.grid(row=3, column=0, sticky="w", padx=14, pady=(0, 10))

        _kpi_card(
            kpi,
            "Fichiers mensuels détectés",
            f"{nb_fichiers:,}".replace(",", " "),
            "Historique *_YYYY_MM.xlsx*",
            0,
            detail_cb=_show_years
        )
        _kpi_card(
            kpi,
            "Nombre de filiales",
            f"{nb_filiales:,}".replace(",", " "),
            "Distinctes dans la config",
            1,
            detail_cb=_show_filiales
        )
        _kpi_card(
            kpi,
            "Dernière actualisation",
            f"{derniere_maj}",
            "Basée sur les mtime",
            2,
            detail_cb=None
        )



        # ====== ACTIVITÉ RÉCENTE ======
        recent = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        recent.grid(row=4, column=0, sticky="ew", padx=24, pady=(10, 20))
        recent.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            recent, text="Activité récente",
            font=("Segoe UI Semibold", 16, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        table = ctk.CTkFrame(recent, fg_color="#142544", corner_radius=12,
                            border_width=1, border_color="#223658")
        table.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 16))
        for i in range(3):
            table.grid_columnconfigure(i, weight=[3, 1, 1][i])

        def _header(r, text):
            ctk.CTkLabel(table, text=text, text_color="#9fb7dd",
                        font=("Segoe UI Semibold", 12)) \
                .grid(row=r, column=["Fichier", "Taille", "Modifié le"].index(text),
                    sticky="w", padx=12, pady=(10, 6))

        _header(0, "Fichier")
        _header(0, "Taille")
        _header(0, "Modifié le")

        recent_files = []
        try:
            if fichiers:
                rec = sorted(
                    [(p, os.path.getmtime(p)) for _, _, p in fichiers if os.path.exists(p)],
                    key=lambda x: x[1], reverse=True
                )[:5]
                for p, ts in rec:
                    size = f"{(os.path.getsize(p) / 1024 / 1024):.1f} Mo"
                    recent_files.append((os.path.basename(p), size, _fmt_dt(ts)))
        except Exception:
            pass

        if not recent_files:
            recent_files = [("—", "—", "—")]

        for i, (name, size, dt_str) in enumerate(recent_files, start=1):
            ctk.CTkLabel(table, text=name, text_color="white", font=("Segoe UI", 12)) \
                .grid(row=i, column=0, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=size, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=1, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=dt_str, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=2, sticky="w", padx=12, pady=6)

    def vider_fenetre(self):
        for widget in self.winfo_children():
            widget.destroy()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.destroy()

#===================Page Prédiction IA===================
    def _rebuild_profils_ui2(self, section, flux_name, annee):
        """Reconstruit la select-box de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()

        # nouveau : on n'utilise plus self.vars_prev ici
        self.vars_prev = []
        self.profils_names_order = []   # ordre des profils pour mapping
        self.profils_listbox = None     # widget de sélection multiple

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils

        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # Label d'info
        lbl = tk.Label(
            self.profils_frame,
            text="Profils (Ctrl+clic pour multi-sélection) :",
            bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
        )
        lbl.pack(anchor="w", padx=4, pady=(0, 4))

        # Frame contenant la listbox + scrollbar
        container = tk.Frame(self.profils_frame, bg="#00122e")
        container.pack(fill="x", padx=4, pady=(0, 6))

        scrollbar = tk.Scrollbar(container, orient="vertical")
        self.profils_listbox = tk.Listbox(
            container,
            selectmode="multiple",
            exportselection=False,
            bg="#00122e",
            fg="white",
            font=('Segoe UI', 10),
            height=min(8, len(noms_profils)),
            highlightthickness=0,
            activestyle="dotbox"
        )
        self.profils_listbox.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.profils_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.profils_listbox.yview)

        for name in noms_profils:
            self.profils_listbox.insert("end", name)

        # petite sélection par défaut : tout sélectionner
        self.profils_listbox.select_set(0, "end")

    def creer_page_ia_prediction(self):
        """
        Page IA de prévision Enc. Autres Produits (N -> N+1).
        Cette méthode ne fait plus que :
        1) Vérifier les dépendances
        2) Réinitialiser la fenêtre
        3) Construire le dataset global
        4) Construire toute la page IA (UI + callbacks + entraînement)
        """
        # 1) Imports (idéalement à mettre en haut du fichier Python)
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # 2) sécurité sklearn
        if RandomForestRegressor is None:
            messagebox.showerror(
                "Dépendance manquante",
                "Le module 'scikit-learn' n'est pas installé.\n"
                "Installe-le avec : pip install scikit-learn"
            )
            return

        # 3) reset page
        self.vider_fenetre()

        # 4) construire le dataset global (toutes filiales, tous jours)
        df = self._ia_build_dataset_enc_autres_produits()
        if df is None:
            # _ia_build_dataset_enc_autres_produits affiche déjà le message et fait retour_menu()
            return

        # 5) construire la page IA complète (UI + callbacks + entraînement)
        self._ia_build_prediction_page(df)
        
    def _ia_build_dataset_enc_autres_produits(self):
        """
        Construit le DataFrame global df sur le flux 'Enc. Autres Produits'
        pour toutes les filiales, avec colonnes :
        section, date, y, year, month, dayofyear, section_id, roll_mean_7, roll_mean_30
        Retourne df ou None si aucune donnée.
        """
        import pandas as pd
        import numpy as np
        from tkinter import messagebox

        flux_cible = "Trafic Voyageurs"
        lignes = []

        # ---- helpers numériques ----
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        # ---- construction dataset (reprends ton code) ----
        for section_name, feuille in sections.items():
            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees pour {feuille} : {e}")
                continue

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                continue

            _, col_start = cible[0]
            try:
                dates, reel, _previsions, _noms_profils = extraire_valeurs(
                    ws, col_start, nb_prev, annee=None
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs pour {section_name}/{flux_cible} : {e}")
                continue

            for d, r in zip(dates, reel):
                y_val = _to_number(r)
                if y_val is None:
                    continue
                try:
                    d_ts = pd.to_datetime(d)
                except Exception:
                    continue
                lignes.append({
                    "section": section_name,
                    "date": d_ts,
                    "y": float(y_val)
                })

        if not lignes:
            messagebox.showinfo(
                "Information",
                "Aucune donnée trouvée pour le flux 'Trafic Voyageurs' sur les filiales."
            )
            self.retour_menu()
            return None

        df = pd.DataFrame(lignes).sort_values(["section", "date"]).reset_index(drop=True)
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["dayofyear"] = df["date"].dt.dayofyear

        # Encodage des filiales
        cat = df["section"].astype("category")
        df["section_id"] = cat.cat.codes

        # Lissages
        df["roll_mean_7"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(7, min_periods=1).mean()
        )
        df["roll_mean_30"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(30, min_periods=1).mean()
        )

        return df

    def _ia_build_prediction_page(self, df):
        """
        Construit toute l'interface IA + callbacks d'entraînement
        à partir du DataFrame global df déjà préparé.
        """
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # ---------- helpers numériques locaux ----------
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_float_or_nan(x):
            val = _to_number(x)
            return float(val) if val is not None else np.nan

        flux_cible = "Trafic Voyageurs"
        default_n_estimators = 200

        # ---------- variables profils & graph 2 ----------
        ia_profils_vars = []        # liste de tk.BooleanVar
        ia_profils_names = []       # noms de profils actifs N+1
        ia_profils_dates = []       # dates N+1
        ia_profils_series = []      # séries (une par profil actif)

        current_pred_df = None          # df_future_all pour N+1
        current_real_target_df = None   # réel sur N+1 (si dispo)
        current_target_year = None
        current_filiale_name = None

        ia_graph2_widget = None
        graph2_container_packed = False

        exported_pred_df = None         # pour export Excel prédictions
        analysis_table_frame = None
        analysis_tree = None
        analysis_export_button = None
        export_button = None            # bouton d’export des prédictions

        # ============ HEADER ============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - IA PRÉDICTION  (N → N+1)",
            font=titre_font,
            text_color="white"
        ).pack(side="left", anchor="w")

        # Logo
        try:
            logo_img = Image.open(image_path)
            test = tk.Label(self, text="Test", font=titre_font)
            test.update_idletasks()
            text_h = test.winfo_reqheight()
            test.destroy()
            ratio = logo_img.width / max(logo_img.height, 1)
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"[IA] Erreur chargement logo IA : {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ============ CONTAINER SCROLLABLE ============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # ============ TITRE / TEXTE ============
        ctk.CTkLabel(
            scrollable_frame,
            text="IA - Prédiction sur le flux 'Trafic Voyageurs' (prévision année N+1)",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        ).pack(pady=15)

        ctk.CTkLabel(
            scrollable_frame,
            text=(
                "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                "Il apprend comment la valeur de 'Trafic Voyageurs' évolue d'une année à l'autre "
                "(N → N+1) pour chaque jour de l'année.\n"
                "Il prédit ensuite l’ensemble de l’année N+1 pour cette filiale.\n"
                "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
            ),
            font=("Segoe UI", 12),
            text_color="#c9defa",
            justify="left"
        ).pack(pady=(0, 10))

        # ============ ZONE PARAMÈTRES (CARD) ============
        params_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        params_frame.pack(fill="x", padx=10, pady=(5, 10))

        # Filiale
        ctk.CTkLabel(
            params_frame,
            text="Filiale :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))

        filiales = sorted(df["section"].unique().tolist())
        selected_filiale = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(
            params_frame,
            textvariable=selected_filiale,
            values=filiales,
            state="readonly",
            width=28
        )
        filiale_box.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        # Année N
        ctk.CTkLabel(
            params_frame,
            text="Année N (historique → N+1) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=1, sticky="w", padx=12, pady=(10, 2))

        years_sorted = sorted(df["year"].unique().tolist())
        annees_var = tk.StringVar(value=str(years_sorted[-1]))
        annees_box = ttk.Combobox(
            params_frame,
            textvariable=annees_var,
            values=[str(y) for y in years_sorted],
            state="readonly",
            width=18
        )
        annees_box.grid(row=1, column=1, sticky="w", padx=12, pady=(0, 10))

        # Hyperparamètres RF
        ctk.CTkLabel(
            params_frame,
            text="Hyperparamètres Random Forest (point de départ) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=2, sticky="w", padx=12, pady=(10, 2))

        # n_estimators
        ctk.CTkLabel(
            params_frame,
            text="n_estimators",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=2, sticky="w", padx=12, pady=(0, 2))

        n_estimators_var = tk.IntVar(value=default_n_estimators)
        slider_n = ctk.CTkSlider(
            params_frame,
            from_=50,
            to=600,
            number_of_steps=11,
            command=lambda v: n_estimators_var.set(int(v))
        )
        slider_n.set(default_n_estimators)
        slider_n.grid(row=2, column=2, sticky="we", padx=12, pady=(0, 8))

        lbl_n = ctk.CTkLabel(
            params_frame,
            text=f"{default_n_estimators}",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_n.grid(row=3, column=2, sticky="w", padx=12, pady=(0, 6))

        def _update_lbl_n(_=None):
            lbl_n.configure(text=str(n_estimators_var.get()))
        slider_n.configure(command=lambda v: (n_estimators_var.set(int(v)), _update_lbl_n()))

        # max_depth
        ctk.CTkLabel(
            params_frame,
            text="max_depth",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=3, sticky="w", padx=12, pady=(0, 2))

        max_depth_var = tk.IntVar(value=15)
        slider_d = ctk.CTkSlider(
            params_frame,
            from_=3,
            to=25,
            number_of_steps=11,
            command=lambda v: max_depth_var.set(int(v))
        )
        slider_d.set(15)
        slider_d.grid(row=2, column=3, sticky="we", padx=12, pady=(0, 8))

        lbl_d = ctk.CTkLabel(
            params_frame,
            text="15",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_d.grid(row=3, column=3, sticky="w", padx=12, pady=(0, 2))

        def _update_lbl_d(_=None):
            lbl_d.configure(text=str(max_depth_var.get()))
        slider_d.configure(command=lambda v: (max_depth_var.set(int(v)), _update_lbl_d()))

        use_depth_var = tk.BooleanVar(value=True)
        chk_depth = ctk.CTkCheckBox(
            params_frame,
            text="Limiter la profondeur",
            variable=use_depth_var,
            text_color="#c9defa",
            font=("Segoe UI", 11)
        )
        chk_depth.grid(row=4, column=3, sticky="w", padx=12, pady=(0, 10))

        # Bouton entraîner
        bouton_train = ctk.CTkButton(
            params_frame,
            text="🚀 Entraîner / ré-entraîner le modèle",
            width=240,
            height=40,
            corner_radius=16,
            fg_color="#008C4B",
            hover_color="#006C39",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_train.grid(row=1, column=4, rowspan=3, padx=14, pady=(0, 10), sticky="e")

        for c in range(5):
            params_frame.grid_columnconfigure(c, weight=1 if c in (2, 3) else 0)

        # ============ CARD PROFILS N+1 (CASES À COCHER) ============
        profils_card = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        profils_card.pack(fill="x", padx=10, pady=(0, 15))

        ctk.CTkLabel(
            profils_card,
            text="Profils de prévision disponibles sur l'année N+1 :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=12, pady=(10, 4))

        ia_profils_frame = tk.Frame(profils_card, bg="#00122e")
        ia_profils_frame.pack(fill="x", padx=12, pady=(0, 10))

        # ============ ZONE RÉSULTATS ============
        graph_widgets = []   # kpi, graphe 1, tableaux
        graph2_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe 2
        monthly_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe mensuel
        monthly_graph_widget = None
        monthly_container_packed = False

        # ---------- helpers graphiques / profils ----------

        def _clear_graph_widgets():
            """Supprime les widgets graphiques (sauf le conteneur du graphe 2)."""
            nonlocal graph_widgets
            for w in graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            graph_widgets = []

        def _redraw_graph2():
            """Redessine le graphe détaillé N+1 avec :
            - Réel N+1 (si dispo)
            - Prévision IA N+1
            - Profils cochés N+1
            """
            nonlocal ia_graph2_widget, graph2_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Si rien à tracer et aucun profil, on ne fait rien
            if current_pred_df is None and current_real_target_df is None and not ia_profils_names:
                return

            # S'assurer que le conteneur est packé une fois pour toutes
            if not graph2_container_packed:
                graph2_container.pack(pady=10, fill="both", expand=True)
                graph2_container_packed = True

            # Nettoyer le conteneur
            for child in graph2_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            ia_graph2_widget = None

            # Créer la figure
            fig2, ax2 = plt.subplots(
                figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True
            )
            ax2.set_facecolor("#00122e")

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            # Réel N+1
            if current_real_target_df is not None and not current_real_target_df.empty:
                ax2.plot(
                    current_real_target_df["date"], current_real_target_df["y"],
                    label=f"Réel {target_year}", linewidth=2, color="#5DADE2"
                )

            # Prévision IA N+1
            if current_pred_df is not None and not current_pred_df.empty:
                ax2.plot(
                    current_pred_df["date"], current_pred_df["pred_value"],
                    label=f"Prévision IA {target_year}",
                    linewidth=2, linestyle="--", color="#F4D03F"
                )

            # Profils cochés N+1
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = [pd.to_datetime(d) for d in ia_profils_dates]
                palette = plt.cm.tab10.colors
                prof_idx = 0
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    y_prof = [_to_float_or_nan(v) for v in serie]
                    ax2.plot(
                        dates_prof, y_prof,
                        label=f"Profil '{name}' {target_year}",
                        linewidth=1.8,
                        linestyle="-.",
                        color=palette[prof_idx % len(palette)]
                    )
                    prof_idx += 1

            filiale_for_title = current_filiale_name or selected_filiale.get()
            titre2 = f"Trafic Voyageurs – année {target_year} – {filiale_for_title}"
            ax2.set_title(titre2, color="white", fontsize=14)
            ax2.set_xlabel("Date", color="white", fontsize=12)
            ax2.set_ylabel("Montant (K€)", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white", rotation=30)
            ax2.tick_params(axis='y', colors="white")
            ax2.legend(facecolor="#00122e", edgecolor="white", labelcolor="white")

            canvas_fig2 = FigureCanvasTkAgg(fig2, master=graph2_container)
            canvas_fig2.draw()
            ia_graph2_widget = canvas_fig2.get_tk_widget()
            ia_graph2_widget.pack(fill="both", expand=True)
            plt.close(fig2)
        
        def _redraw_monthly_graph():
            """
            Graphe mensuel N+1 :
            - Barres : Réel N+1 (si dispo) vs Prévision IA N+1
            - Barres supplémentaires : Profils cochés (somme mensuelle)
            """
            nonlocal monthly_graph_widget, monthly_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Il faut au moins la prédiction N+1 pour tracer quelque chose
            if current_pred_df is None or current_pred_df.empty:
                return

            # Pack du conteneur une seule fois
            if not monthly_container_packed:
                monthly_container.pack(pady=10, fill="both", expand=True)
                monthly_container_packed = True

            # Nettoyage du conteneur
            for child in monthly_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            monthly_graph_widget = None

            import numpy as np
            from matplotlib.ticker import FuncFormatter

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            color_real = "#1f77b4"   # bleu
            color_pred = "#F4D03F"   # orange
            # Palette dédiée aux profils (change-les comme tu veux)
            profile_colors = [
                "#e74c3c",  # rouge vif
                "#9b59b6",  # violet
                "#2ecc71",  # vert clair
                "#1abc9c",  # turquoise
                "#f1c40f",  # jaune
                "#d35400",  # orange foncé (≠ #ff7f0e)
                "#8e44ad",  # violet profond
                "#27ae60",  # vert riche
                "#16a085",  # vert/turquoise sombre
                "#c0392b",  # rouge foncé
                "#7f8c8d",  # gris bleuté
                "#95a5a6",  # gris clair
                "#34495e",  # bleu/gris métallique
                "#bdc3c7",  # argent clair
                "#f39c12",  # jaune/orangé
                "#c27ba0",  # rose/mauve mat
                "#76d7c4",  # aqua pastel
                "#7dcea0",  # vert pastel
                "#af7ac5",  # mauve pastel
                "#5dade2",  # bleu pastel clair (≠ réel, beaucoup plus clair)
            ]

            # ---------- Prévision IA N+1 : somme mensuelle ----------
            df_pred = current_pred_df.copy()
            df_pred["month"] = df_pred["date"].dt.month
            pred_monthly = (
                df_pred.groupby("month", as_index=False)["pred_value"]
                       .sum()
            ).rename(columns={"pred_value": "pred_value"})

            # ---------- Réel N+1 : somme mensuelle (si dispo) ----------
            if current_real_target_df is not None and not current_real_target_df.empty:
                df_real = current_real_target_df.copy()
                df_real["month"] = df_real["date"].dt.month
                real_monthly = (
                    df_real.groupby("month", as_index=False)["y"]
                           .sum()
                ).rename(columns={"y": "real_value"})

                monthly_cmp = pd.merge(
                    real_monthly,
                    pred_monthly,
                    on="month",
                    how="outer"
                ).fillna(0.0)
            else:
                monthly_cmp = pred_monthly.copy()
                monthly_cmp["real_value"] = 0.0

            monthly_cmp = monthly_cmp.sort_values("month")

            # ---------- Profils cochés : agrégés par mois (pour barres) ----------
            # ---------- Profils cochés : calcul "Prévision sinon Réel" + masquage ----------
            active_profiles = []

            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = pd.to_datetime(ia_profils_dates)

                # Base réelle jour par jour
                df_day = pd.DataFrame({
                    "date": dates_prof,
                    "month": dates_prof.month,
                })

                # Ajouter le réel si dispo
                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = dict(zip(current_real_target_df["date"], current_real_target_df["y"]))
                    df_day["real"] = df_day["date"].map(real_map).fillna(0.0)
                else:
                    df_day["real"] = 0.0

                # Calcul par profil
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    vals = [_to_float_or_nan(v) for v in serie]

                    df_day["prev"] = vals

                    # Prévision sinon réel
                    df_day["comb"] = df_day["prev"]
                    df_day.loc[df_day["comb"].isna(), "comb"] = df_day["real"]

                    # Mois avec prévision
                    mois_avec_prev = set(df_day.loc[df_day["prev"].notna(), "month"].unique())
                    mois_tous = set(monthly_cmp["month"].unique())
                    mois_masques = sorted(mois_tous - mois_avec_prev)

                    # Masquage complet des mois sans prévision
                    df_day["comb_masked"] = df_day["comb"]
                    df_day.loc[df_day["month"].isin(mois_masques), "comb_masked"] = np.nan

                    # Agrégation mensuelle
                    prof_month = df_day.groupby("month")["comb_masked"].sum(min_count=1)

                    # Alignement
                    yvals = [prof_month.get(m, np.nan) for m in monthly_cmp["month"]]

                    active_profiles.append((name, yvals))


            # ---------- Placement des barres ----------
            x = np.arange(len(monthly_cmp))

            # nb séries : Réel + Prévision + profils actifs
            nb_series = 2 + len(active_profiles)
            # largeur de chaque barre pour que tout tienne dans le mois
            width = 0.8 / max(nb_series, 1)
            offsets = (np.arange(nb_series) - (nb_series - 1) / 2.0) * width

            fig_m, ax_m = plt.subplots(figsize=(11, 4.5),
                                       facecolor="#00122e",
                                       constrained_layout=True)
            ax_m.set_facecolor("#00122e")

            # Barres Réel N+1
            ax_m.bar(
                x + offsets[0],
                monthly_cmp["real_value"],
                width,
                label=f"Réel {target_year}",
                color=color_real
            )

            # Barres Prévision IA N+1
            ax_m.bar(
                x + offsets[1],
                monthly_cmp["pred_value"],
                width,
                label=f"Prévision IA {target_year}",
                color=color_pred
            )

            # Barres Profils cochés
            if active_profiles:
                for i, (name, y_vals) in enumerate(active_profiles):
                    serie_idx = 2 + i
                    profile_color = profile_colors[i % len(profile_colors)]
                    ax_m.bar(
                        x + offsets[serie_idx],
                        y_vals,
                        width,
                        label=f"Profil '{name}' {target_year}",
                        color=profile_color
                    )


            # ---------- X / Y / style ----------
            month_nums = monthly_cmp["month"].astype(int)
            month_dates = pd.to_datetime(
                {"year": [target_year] * len(month_nums),
                 "month": month_nums,
                 "day": 1}
            )
            mois_labels = month_dates.dt.strftime("%Y-%m")

            ax_m.set_xticks(x)
            ax_m.set_xticklabels(mois_labels, rotation=45, ha="right",
                                 fontsize=9, color="white")

            ax_m.set_xlabel("Mois", fontsize=11, color="white")
            ax_m.set_ylabel("Valeur cumulée (K€)", fontsize=11, color="white")
            ax_m.tick_params(axis="y", colors="white")

            def _fmt_milliers(val, pos):
                try:
                    return f"{int(val):,}".replace(",", " ")
                except Exception:
                    return ""
            ax_m.yaxis.set_major_formatter(FuncFormatter(_fmt_milliers))

            filiale_for_title = current_filiale_name or selected_filiale.get()
            ax_m.set_title(
                f"{filiale_for_title} - Trafic Voyageurs - {target_year}",
                fontsize=13,
                fontweight="bold",
                color="white",
                pad=12
            )

            # pas de bordures
            fig_m.patch.set_edgecolor("#00122e")
            fig_m.patch.set_linewidth(0)
            for spine in ax_m.spines.values():
                spine.set_visible(False)

            # légende sans cadre
            leg_m = ax_m.legend(frameon=False, facecolor="#00122e")
            for text in leg_m.get_texts():
                text.set_color("white")

            canvas_m = FigureCanvasTkAgg(fig_m, master=monthly_container)
            canvas_m.draw()
            monthly_graph_widget = canvas_m.get_tk_widget()
            monthly_graph_widget.pack(fill="both", expand=True)
            plt.close(fig_m)

        # ---------- analyse des écarts (IA vs Réel, Profil vs Réel) ----------

        def _compute_metrics(label, ref_dates, ref_values, cmp_dates, cmp_values,
                     min_abs_value=0.0, ecart_seuil=40.0):
            """
            ref = série réelle (Réel)
            cmp = série comparée (IA ou Profil)

            - aligne les dates
            - convertit proprement en float (gère '', None, etc.)
            - filtre (optionnel) les petits montants sur le RÉEL via min_abs_value
            - calcule l'écart relatif |ref - cmp| / |cmp|
            - renvoie toujours un dict SI au moins une date est commune,
            même si après filtrage il n'y a plus de points 'significatifs'
            """

            if ref_dates is None or cmp_dates is None:
                print(f"[METRICS] {label} -> ref_dates ou cmp_dates None")
                return None

            # Conversion en float en utilisant ton helper
            ref_vals = [_to_float_or_nan(v) for v in ref_values]
            cmp_vals = [_to_float_or_nan(v) for v in cmp_values]

            ref_idx = pd.to_datetime(ref_dates)
            cmp_idx = pd.to_datetime(cmp_dates)

            ref = pd.Series(ref_vals, index=ref_idx, dtype="float64")
            cmp_ = pd.Series(cmp_vals, index=cmp_idx, dtype="float64")

            # Intersection des dates
            common_idx = ref.index.intersection(cmp_.index)
            print(f"[METRICS] {label} -> nb dates communes = {len(common_idx)}")

            if len(common_idx) == 0:
                # dans ce cas on ne sait vraiment rien comparer
                return None

            df_m = pd.DataFrame({
                "ref": ref.loc[common_idx],
                "cmp": cmp_.loc[common_idx],
            })

            # Drop NaN
            df_m = df_m.dropna()
            print(f"[METRICS] {label} -> après dropna, nb points = {len(df_m)}")

            if df_m.empty:
                # aucune valeur exploitable → on renvoie quand même une ligne "0"
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # 🔎 Filtre optionnel : SEULEMENT sur le RÉEL
            # ex : min_abs_value=10  -> on garde uniquement les jours où |Réel| > 10
            if min_abs_value is not None and min_abs_value > 0:
                before = len(df_m)
                df_m = df_m[df_m["ref"].abs() >= min_abs_value]
                print(f"[METRICS] {label} -> filtre |ref| >= {min_abs_value} : {before} → {len(df_m)} points")

            if df_m.empty:
                # il y a des dates communes, mais tous les réels sont trop petits
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # Éviter la division par zéro (sur la prévision)
            df_m = df_m[df_m["cmp"] != 0]
            if df_m.empty:
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # Écart relatif basé sur la prévision cmp
            rel = (df_m["ref"] - df_m["cmp"]).abs() / df_m["cmp"].abs() * 100.0
            rel = rel.replace([np.inf, -np.inf], np.nan).dropna()

            if rel.empty:
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            mask = rel > ecart_seuil

            nb_points = int(df_m.shape[0])
            nb_ecarts = int(mask.sum())
            taux = 0.0 if nb_points == 0 else 100.0 * nb_ecarts / nb_points
            valo = float((df_m["ref"] - df_m["cmp"]).abs()[mask].sum())

            print(f"[METRICS] {label} -> nb_points={nb_points}, nb_ecarts={nb_ecarts}, taux={taux:.1f}, valo={valo:.1f}")

            return {
                "Comparaison": label,
                "Nb écarts (> 40 %)": nb_ecarts,
                "Taux d'écarts (%)": round(taux, 1),
                "Valeur totale des écarts (K€)": round(valo, 1),
            }


        def _rebuild_analysis_table():
            nonlocal analysis_table_frame, analysis_tree, analysis_export_button
            nonlocal current_pred_df, current_real_target_df, current_target_year

            if analysis_table_frame is None:
                return
            for child in analysis_table_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            analysis_tree = None
            analysis_export_button = None

            # s'il n'y a pas de réel, on ne peut rien comparer
            if current_real_target_df is None or current_real_target_df.empty:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune donnée réelle pour l'année N+1 : analyse des écarts impossible.",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                print("[ANALYSE] Pas de réel N+1.")
                return

            print(f"[ANALYSE] réel N+1 : {len(current_real_target_df)} lignes")
            if current_pred_df is not None:
                print(f"[ANALYSE] préd N+1 : {len(current_pred_df)} lignes")
            else:
                print("[ANALYSE] current_pred_df est None")

            rows = []

            # 1) IA vs Réel
            if current_pred_df is not None and not current_pred_df.empty:
                m = _compute_metrics(
                    label="Prévision IA vs Réel",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    # tu peux remonter le seuil ici plus tard si tu veux
                    min_abs_value=10.0,
                    ecart_seuil=40.0,
                )
                print("[ANALYSE] IA vs Réel ->", m)
                if m is not None:
                    rows.append(m)

            # 2) Profils cochés vs Réel
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    m = _compute_metrics(
                        label=f"Profil '{name}' vs Réel",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=ia_profils_dates,
                        cmp_values=serie,
                        min_abs_value=10.0,
                        ecart_seuil=40.0,
                    )
                    print(f"[ANALYSE] Profil {name} vs Réel ->", m)
                    if m is not None:
                        rows.append(m)

            ctk.CTkLabel(
                analysis_table_frame,
                text=f"Analyse des écarts (seuil 40 %, ref/prev ≥ 0) année {current_target_year or (int(annees_var.get())+1)}",
                font=("Segoe UI", 13, "bold"),
                text_color="white"
            ).pack(anchor="w", pady=(0, 6))

            if not rows:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune comparaison disponible (vérifie que le modèle est entraîné et que des profils sont cochés).",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                print("[ANALYSE] Aucune ligne de comparaison construite.")
                return

            cols = [
                "Comparaison",
                "Nb écarts (> 40 %)",
                "Taux d'écarts (%)",
                "Valeur totale des écarts (K€)"
            ]
            tree = ttk.Treeview(
                analysis_table_frame, columns=cols, show="headings",
                height=len(rows)
            )
            for col in cols:
                tree.heading(col, text=col)

            tree.column("Comparaison", anchor="w", width=260)
            tree.column("Nb écarts (> 40 %)", anchor="center", width=150)
            tree.column("Taux d'écarts (%)", anchor="center", width=150)
            tree.column("Valeur totale des écarts (K€)", anchor="center", width=180)

            for r in rows:
                tree.insert(
                    "", "end",
                    values=(
                        r["Comparaison"],
                        r["Nb écarts (> 40 %)"],
                        r["Taux d'écarts (%)"],
                        r["Valeur totale des écarts (K€)"],
                    )
                )

            tree.pack(fill="x", expand=True, pady=(0, 8))
            analysis_tree = tree

            # bouton export Excel du tableau d'analyse
            def _export_analysis_excel():
                if not rows:
                    messagebox.showinfo("Export", "Aucune donnée à exporter.")
                    return
                df_export = pd.DataFrame(rows)
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Fichiers Excel", "*.xlsx")],
                        title="Exporter le tableau d'analyse en Excel"
                    )
                    if not file_path:
                        return
                    df_export.to_excel(file_path, index=False)
                    messagebox.showinfo("Export", f"Tableau d'analyse exporté vers :\n{file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur export", str(e))

            analysis_export_button = ctk.CTkButton(
                analysis_table_frame,
                text="📤 Exporter le tableau d'analyse (Excel)",
                width=260, height=32,
                corner_radius=10,
                fg_color="#2563eb", hover_color="#1d4ed8",
                text_color="white",
                command=_export_analysis_excel
            )
            analysis_export_button.pack(anchor="w", pady=(6, 4))

        # ---------- export Excel des prédictions N+1 ----------

        def _export_predictions_to_excel():
            nonlocal exported_pred_df, current_target_year, current_filiale_name

            if exported_pred_df is None or exported_pred_df.empty:
                messagebox.showinfo(
                    "Export",
                    "Aucune prédiction à exporter. Lance d'abord l'entraînement / la prévision."
                )
                return

            try:
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Fichiers Excel", "*.xlsx")],
                    title="Exporter les prédictions N+1 en Excel"
                )
                if not file_path:
                    return

                df_export = exported_pred_df.copy()
                df_export.insert(0, "Filiale", current_filiale_name or "")
                df_export.insert(1, "Année", current_target_year or "")

                df_export.to_excel(file_path, index=False)
                messagebox.showinfo(
                    "Export",
                    f"Prédictions N+1 exportées vers :\n{file_path}"
                )
            except Exception as e:
                messagebox.showerror("Erreur export", str(e))

        # ---------- PROFILS UI ----------

        def _build_ia_profils_ui(filiale, base_year):
            """
            Construit les cases à cocher de profils pour l'année N+1 de la filiale sélectionnée.
            Dès qu'on coche/décoche, _redraw_graph2() et _rebuild_analysis_table() sont appelés.
            """
            nonlocal ia_profils_vars, ia_profils_names, ia_profils_dates, ia_profils_series

            # Nettoyage du frame
            for w in ia_profils_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            ia_profils_vars = []
            ia_profils_names = []
            ia_profils_dates = []
            ia_profils_series = []

            if not filiale:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Aucune filiale sélectionnée.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                base_year_int = int(base_year)
            except Exception:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Année N invalide.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            target_year = base_year_int + 1
            feuille = sections.get(filiale)
            if not feuille:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune feuille trouvée pour {filiale}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees (profils N+1) pour {feuille} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Erreur de chargement des données N+1.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Flux 'Trafic Voyageurs' introuvable pour cette filiale.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            _, col_start = cible[0]

            try:
                dates_p, reel_p, previsions_p, noms_profils_p = extraire_valeurs(
                    ws, col_start, nb_prev, annee=target_year
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs (profils N+1) {filiale}/{flux_cible}/{target_year} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée de profils pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            if not dates_p:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            # filtrer les profils actifs (au moins une valeur non nulle)
            actifs = []
            for serie in previsions_p:
                exist = any(v not in (None, 0, 0.0, "") for v in serie)
                actifs.append(exist)

            noms_actifs = [n for n, ok in zip(noms_profils_p, actifs) if ok]
            series_actives = [s for s, ok in zip(previsions_p, actifs) if ok]

            if not noms_actifs:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucun profil actif pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            ia_profils_dates = dates_p
            ia_profils_names = noms_actifs
            ia_profils_series = series_actives

            # Label d'info
            lbl = tk.Label(
                ia_profils_frame,
                text=f"Profils N+1 ({target_year}) : coche pour afficher dans le graphe détaillé et l'analyse",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
            )
            lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 6))

            nb_lignes = 4
            for i, name in enumerate(noms_actifs):
                var = tk.BooleanVar(value=False)

                def _on_toggle(v=var):
                    _redraw_graph2()
                    _redraw_monthly_graph()
                    _rebuild_analysis_table()


                cb = tk.Checkbutton(
                    ia_profils_frame, text=name, variable=var,
                    bg="#00122e", fg="white", font=('Segoe UI', 10),
                    selectcolor="#00aced", activebackground="#003366",
                    activeforeground="white",
                    command=_on_toggle
                )

                row = 1 + (i % nb_lignes)   # +1 car première ligne = label
                col = i // nb_lignes

                cb.grid(row=row, column=col, sticky="w", padx=12, pady=4)
                cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
                cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

                ia_profils_vars.append(var)

        def _on_filiale_or_year_change(_event=None):
            """Quand filiale ou année N change, on recharge les profils N+1."""
            filiale = selected_filiale.get()
            base_year = annees_var.get()
            if not filiale or not base_year:
                return
            _build_ia_profils_ui(filiale, base_year)

        filiale_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        annees_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        _on_filiale_or_year_change()

        # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------
        def _train_model():
            """
            Entraîne le modèle uniquement sur :
            - la filiale SA_VOYAGEURS
            - le flux 'Trafic Voyageurs' (colonne y)
            Puis prédit N+1 pour SA_VOYAGEURS et met à jour les graphes + analyse.
            """
            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame, export_button

            _clear_graph_widgets()
            print("======================== ENTRAINEMENT DU MODELE (SA_VOYAGEURS – RF + XGB résidus + WEEKEND=0) ========================")

            # ----------- Filiale fixée : SA_VOYAGEURS -----------
            filiale = "SA_VOYAGEURS"
            current_filiale_name = filiale
            print(f"Filiale utilisée (fixe) : {filiale}")

            # On ne prend que SA_VOYAGEURS dans le df global
            df_filiale = df[df["section"] == filiale].copy()
            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée trouvée pour la filiale {filiale}.")
                return

            # ----------------------- ANNÉE N -----------------------
            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["year"].max())
            print(f"Base year for training {filiale} : {base_year}")

            try:
                # ----------------------- HISTORIQUE -----------------------
                if df_filiale["year"].nunique() < 2:
                    messagebox.showinfo(
                        "Information",
                        f"Pas assez d'historique pour {filiale} (au moins 2 années)."
                    )
                    return

                # ----------------------- FEATURES TEMPORELLES -----------------------
                df_filiale["dow"] = df_filiale["date"].dt.weekday
                df_filiale["month"] = df_filiale["date"].dt.month
                df_filiale["dom"] = df_filiale["date"].dt.day
                df_filiale["is_eom"] = df_filiale["date"].dt.is_month_end.astype(int)
                df_filiale["lag_1"] = df_filiale["y"].shift(1)
                df_filiale["lag_7"] = df_filiale["y"].shift(7)

                # ========== PAIRS YEAR-OVER-YEAR (N -> N+1) ==========
                s = df_filiale.copy()
                s_prev = s.copy()
                s_prev["year_target"] = s_prev["year"] + 1

                s_prev = s_prev.rename(columns={
                    "y": "y_prev_year",
                    "roll_mean_7": "roll_prev_7",
                    "roll_mean_30": "roll_prev_30",
                    "dow": "dow_prev",
                    "lag_1": "lag_1_prev",
                    "lag_7": "lag_7_prev",
                    "month": "month_prev",
                    "dom": "dom_prev",
                    "is_eom": "is_eom_prev"
                })

                merged = pd.merge(
                    s_prev,
                    s[["section", "year", "dayofyear", "y"]],
                    left_on=["section", "year_target", "dayofyear"],
                    right_on=["section", "year", "dayofyear"],
                    how="inner",
                )

                if merged.empty:
                    messagebox.showinfo("Information", f"Aucune paire année→année pour {filiale}.")
                    return

                merged = merged.rename(columns={"y": "y_target"})

                df_pairs = merged[[  
                    "section", "year_target", "dayofyear",
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev",
                    "y_target"
                ]].copy()

                # Une seule filiale ⇒ section_id fixe (peu utile mais on le laisse)
                df_pairs["section_id"] = 0

                # ----------------------- TRAIN SET = années <= N -----------------------
                df_train_pairs = df_pairs[df_pairs["year_target"] <= base_year]
                if df_train_pairs.shape[0] < 30:
                    messagebox.showinfo(
                        "Information",
                        f"Pas assez de données pour entraîner le modèle jusqu'à {base_year}."
                    )
                    return

                # ----------------------- FEATURES UTILISÉES -----------------------
                features = [
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dayofyear", "section_id",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev"
                ]

                df_train_pairs = df_train_pairs.dropna(subset=features + ["y_target"])
                # Conserver DataFrame/Series pour préserver les noms de colonnes (évite warnings LGBM)
                X_df = df_train_pairs[features].reset_index(drop=True)
                y_series = df_train_pairs["y_target"].reset_index(drop=True)

                # ----------------------- SPLIT TRAIN / VALID (chronologique) -----------------------
                # Utilise un split chronologique (pas de shuffle) : les dernières observations servent de validation
                n_total = len(X_df)
                # Recherche du meilleur split chronologique parmi plusieurs candidats
                print(f"[SPLIT SEARCH] total rows available={n_total}")
                from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
                from sklearn.ensemble import RandomForestRegressor

                # paramètres légers pour l'évaluation des splits (rapide)
                temp_rf = RandomForestRegressor(n_estimators=100, max_depth=None, random_state=42, n_jobs=-1)

                candidate_fracs = [0.5, 0.6, 0.65, 0.7, 0.75, 0.8, 0.85]
                split_results = []
                for frac in candidate_fracs:
                    cut_i = int(n_total * frac)
                    if cut_i < 10 or n_total - cut_i < 5:
                        continue
                    X_tr = X_df.iloc[:cut_i]
                    X_va = X_df.iloc[cut_i:]
                    y_tr = y_series.iloc[:cut_i]
                    y_va = y_series.iloc[cut_i:]
                    try:
                        temp_rf.fit(X_tr, y_tr)
                        pred_va = temp_rf.predict(X_va)
                        r2_va = r2_score(y_va, pred_va)
                    except Exception as e:
                        r2_va = float('-inf')
                    split_results.append({'frac': frac, 'cut': cut_i, 'r2': r2_va, 'train': len(X_tr), 'valid': len(X_va)})
                    print(f"[SPLIT] frac={frac:.2f} cut={cut_i} train={len(X_tr)} valid={len(X_va)} -> R2={r2_va:.4f}")

                if not split_results:
                    # fallback simple split
                    cut = int(n_total * 0.8)
                    best_cut = cut
                else:
                    best = max(split_results, key=lambda x: x['r2'])
                    best_cut = best['cut']
                    print(f"[SPLIT BEST] frac={best['frac']:.2f} cut={best_cut} train={best['train']} valid={best['valid']} R2={best['r2']:.4f}")

                # Appliquer le meilleur split trouvé
                cut = best_cut
                X_train = X_df.iloc[:cut].reset_index(drop=True)
                X_valid = X_df.iloc[cut:].reset_index(drop=True)
                y_train = y_series.iloc[:cut].reset_index(drop=True)
                y_valid = y_series.iloc[cut:].reset_index(drop=True)

                # Affichage final des tailles de split choisies
                print(f"[DATA SPLIT chrono chosen] total={n_total}, train={len(X_train)}, valid={len(X_valid)}")

                base_n = int(n_estimators_var.get())
                base_d = int(max_depth_var.get()) if use_depth_var.get() else None

                # ========================================================================================
                #                      MODELE AMELIORE : GBM + Stacking
                # - Essaye LightGBM si disponible, sinon XGBoost, sinon HistGradientBoosting
                # - Recherche d'hyperparamètres légère (RandomizedSearchCV)
                # - Stacking entre RandomForest et GBM avec meta-regressor RidgeCV
                # - Cross-validation pour robustesse et metrics finales
                # ========================================================================================
                print("\n===== MODELE AMELIORE : GBM + Stacking =====")

                # imports locaux (léger) — évite de casser l'espace de noms global
                from sklearn.ensemble import StackingRegressor, HistGradientBoostingRegressor
                from sklearn.linear_model import RidgeCV
                from sklearn.model_selection import RandomizedSearchCV, KFold, cross_val_score
                from sklearn.pipeline import Pipeline
                from sklearn.preprocessing import StandardScaler
                from scipy.stats import randint, uniform

                # tentative import LightGBM
                try:
                    from lightgbm import LGBMRegressor
                except Exception:
                    LGBMRegressor = None

                # fonction utilitaire : tune léger pour GBM
                def _tune_gbm(X_t, y_t, base_n, base_d):
                    """Retourne un estimateur GBM prêt à l'emploi (déjà entraîné sur X_t/y_t).
                    Priorité : LGBM -> HistGradientBoosting. Utilise RandomizedSearchCV pour 20 itérations max.
                    """
                    # small search space
                    param_dist = {
                        'n_estimators': randint(max(50, int(base_n/2)), max(100, base_n*2)),
                        'max_depth': [None, 4, 6, 8, int(base_d) if base_d else None],
                        'learning_rate': uniform(0.02, 0.2)
                    }

                    if LGBMRegressor is not None:
                        model = LGBMRegressor(n_estimators=base_n, max_depth=base_d if base_d is not None else -1, random_state=42)
                        rs = RandomizedSearchCV(model, param_distributions=param_dist, n_iter=20, cv=3,
                                                scoring='neg_root_mean_squared_error', n_jobs=-1, random_state=42, verbose=0)
                        rs.fit(X_t, y_t)
                        print(f"[TUNING] LGBM best params: {rs.best_params_} | score={rs.best_score_:.4f}")
                        best = rs.best_estimator_
                        return best

                    # fallback : HistGradientBoosting (scikit-learn) — tuning léger
                    print("[INFO] LightGBM non disponible — utilisation de HistGradientBoosting (scikit-learn) avec tuning léger.")
                    from sklearn.ensemble import HistGradientBoostingRegressor as _HGB
                    h_param_dist = {
                        'max_iter': randint(50, max(100, base_n)),
                        'max_depth': [None, 4, 6, 8],
                        'learning_rate': uniform(0.02, 0.2)
                    }
                    hmodel = _HGB(random_state=42)
                    hrs = RandomizedSearchCV(hmodel, param_distributions=h_param_dist, n_iter=12, cv=3,
                                             scoring='neg_root_mean_squared_error', n_jobs=-1, random_state=42, verbose=0)
                    hrs.fit(X_t, y_t)
                    print(f"[TUNING] HistGB best params: {hrs.best_params_} | score={hrs.best_score_:.4f}")
                    return hrs.best_estimator_

                # Construire et tuner les estimateurs
                gbm = _tune_gbm(X_train, y_train, base_n, base_d)

                rf = RandomForestRegressor(
                    n_estimators=max(50, int(base_n/2)),
                    max_depth=base_d,
                    max_features='sqrt',
                    min_samples_leaf=2,
                    random_state=42,
                    n_jobs=-1
                )

                # stacking
                estimators = [('rf', rf), ('gbm', gbm)]
                stack = StackingRegressor(estimators=estimators, final_estimator=RidgeCV(alphas=(0.1, 1.0, 10.0)), n_jobs=-1, passthrough=False)

                # pipeline simple : scaler -> stacking (scaler helps Ridge meta-estimator)
                pipeline = Pipeline([
                    ('scaler', StandardScaler()),
                    ('stack', stack)
                ])

                # entraînement
                pipeline.fit(X_train, y_train)

                # prédictions
                pred_train = pipeline.predict(X_train)
                pred_valid = pipeline.predict(X_valid)

                # metrics CV rapides
                # CV : utiliser KFold (shuffle) pour estimation générale — on utilise les DataFrame/Series
                kf = KFold(n_splits=5, shuffle=True, random_state=42)
                cv_r2 = cross_val_score(pipeline, X_df, y_series, scoring='r2', cv=kf, n_jobs=-1)
                cv_rmse = -cross_val_score(pipeline, X_df, y_series, scoring='neg_root_mean_squared_error', cv=kf, n_jobs=-1)

                mae = mean_absolute_error(y_valid, pred_valid)
                rmse = mean_squared_error(y_valid, pred_valid)**0.5
                r2 = r2_score(y_valid, pred_valid)

                print(f"[MODEL FINAL] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")
                print(f"[CV 5-fold] R² mean={cv_r2.mean():.3f} (std={cv_r2.std():.3f}), RMSE mean={cv_rmse.mean():.3f})")

                # ========================================================================================
                #                                    KPI
                # ========================================================================================
                kpi_frame = ctk.CTkFrame(scrollable_frame, fg_color="#0f1b31", corner_radius=18)
                kpi_frame.pack(fill="x", padx=10, pady=(0, 10))
                graph_widgets.append(kpi_frame)

                for i in range(3):
                    kpi_frame.grid_columnconfigure(i, weight=1)

                def _kpi(parent, title, value, subtitle, col):
                    card = ctk.CTkFrame(
                        parent, fg_color="#142544", corner_radius=16,
                        border_width=1, border_color="#223658"
                    )
                    card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
                    ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                        .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
                    ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 22, "bold"),
                                text_color="white") \
                        .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
                    ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                                text_color="#7ea2d8") \
                        .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

                _kpi(kpi_frame, "MAE", f"{mae:.1f}", "Erreur moyenne", 0)
                _kpi(kpi_frame, "RMSE", f"{rmse:.1f}", "Racine MSE", 1)
                _kpi(kpi_frame, "R²", f"{r2:.3f}", "Pouvoir explicatif", 2)

                # ========================================================================================
                #                     PRÉVISION N+1 (WEEK-END = 0 + rattrapage lundi)
                # ========================================================================================

                target_year = base_year + 1
                current_target_year = target_year

                df_prev_year = df_filiale[df_filiale["year"] == base_year].copy()
                df_prev_year = df_prev_year.sort_values("date")  # important pour gérer le carryover

                future_rows = []
                carryover = 0.0  # somme des prédictions des jours de week-end à reporter

                for _, row in df_prev_year.iterrows():
                    d_prev = row["date"]
                    d_next = d_prev + pd.DateOffset(years=1)
                    dow_next = d_next.weekday()  # jour de la semaine dans l'année N+1

                    feat_vec = [[
                        float(row["y"]),
                        float(row["roll_mean_7"]),
                        float(row["roll_mean_30"]),
                        int(row["dayofyear"]),
                        0,  # section_id
                        int(row["dow"]),
                        float(row["lag_1"]),
                        float(row["lag_7"]),
                        int(row["month"]),
                        int(row["dom"]),
                        int(row["is_eom"])
                    ]]

                    # créer un DataFrame avec les mêmes colonnes que l'entraînement — évite warning LGBM
                    try:
                        feat_df = pd.DataFrame(feat_vec, columns=features)
                    except Exception:
                        # en dernier recours, prédire depuis le numpy array
                        feat_df = None

                    # Prédiction brute (sans logique week-end) — on utilise le pipeline entraîné
                    try:
                        if feat_df is not None:
                            pred_raw = float(pipeline.predict(feat_df)[0])
                        else:
                            pred_raw = float(pipeline.predict(feat_vec)[0])
                    except Exception:
                        # fallback si pipeline non disponible : prédire avec rf seul (si fitted)
                        try:
                            if feat_df is not None:
                                pred_raw = float(rf.predict(feat_df)[0])
                            else:
                                pred_raw = float(rf.predict(feat_vec)[0])
                        except Exception:
                            pred_raw = 0.0
                    raw_pred = pred_raw

                    # WEEK-END = 0 mais on accumule pour le lundi suivant
                    if dow_next >= 5:  # samedi (5) ou dimanche (6)
                        carryover += raw_pred
                        pred_final = 0.0
                    else:
                        pred_final = raw_pred + carryover
                        carryover = 0.0

                    future_rows.append({
                        "section": filiale,
                        "date": d_next,
                        "year": d_next.year,
                        "month": d_next.month,
                        "dayofyear": d_next.dayofyear,
                        "pred_value": pred_final
                    })

                df_future_all = pd.DataFrame(future_rows)

                # ========================================================================================
                #     AJUSTEMENT HEBDO : sur chaque semaine, le plus haut pic est mis le lundi
                # ========================================================================================
                if not df_future_all.empty:
                    iso = df_future_all["date"].dt.isocalendar()
                    df_future_all["iso_year"] = iso.year.astype(int)
                    df_future_all["iso_week"] = iso.week.astype(int)
                    df_future_all["dow"] = df_future_all["date"].dt.weekday

                    for (y_iso, w_iso), g in df_future_all.groupby(["iso_year", "iso_week"]):
                        monday_idx = g[g["dow"] == 0].index
                        if monday_idx.empty:
                            continue
                        monday_idx = monday_idx[0]

                        max_idx = g["pred_value"].idxmax()
                        if max_idx == monday_idx:
                            continue

                        val_monday = df_future_all.at[monday_idx, "pred_value"]
                        val_max = df_future_all.at[max_idx, "pred_value"]
                        df_future_all.at[monday_idx, "pred_value"] = val_max
                        df_future_all.at[max_idx, "pred_value"] = val_monday

                    df_future_all = df_future_all.drop(columns=["iso_year", "iso_week", "dow"])

                # ========================================================================================
                #                                    GRAPHE 1 (jours)
                # ========================================================================================
                fig1, ax1 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")

                color_real = "#1f77b4"   # bleu
                color_pred = "#ff7f0e"   # orange

                df_hist_plot = df_filiale[df_filiale["year"] <= base_year]

                ax1.plot(
                    df_hist_plot["date"],
                    df_hist_plot["y"],
                    label=f"Réel (≤ {base_year})",
                    linewidth=2,
                    color=color_real
                )

                ax1.plot(
                    df_future_all["date"],
                    df_future_all["pred_value"],
                    label=f"Prévision Modèle 2 niveaux {target_year}",
                    linewidth=2,
                    linestyle="--",
                    color=color_pred
                )

                ax1.set_title(f"SA_VOYAGEURS – Prévision N+1 – Modèle RF + XGB résidus (Weekend=0 + pic hebdo lundi)",
                            color="white")
                ax1.tick_params(axis='x', colors="white", rotation=30)
                ax1.tick_params(axis='y', colors="white")

                for spine in ax1.spines.values():
                    spine.set_color("#00122e")

                leg1 = ax1.legend(facecolor="#00122e", edgecolor="white")
                for text in leg1.get_texts():
                    text.set_color("white")

                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                w_fig1 = canvas_fig1.get_tk_widget()
                w_fig1.pack(pady=10, fill="both", expand=True)
                graph_widgets.append(w_fig1)
                plt.close(fig1)

                # ========================================================================================
                #                   MISE À JOUR GLOBAL + ANALYSE
                # ========================================================================================
                current_pred_df = df_future_all
                current_real_target_df = df_filiale[df_filiale["year"] == target_year]
                exported_pred_df = df_future_all.copy()

                _redraw_monthly_graph()
                _redraw_graph2()

                if analysis_table_frame is not None:
                    analysis_table_frame.destroy()

                analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
                analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
                graph_widgets.append(analysis_table_frame)

                _rebuild_analysis_table()

            except Exception:
                messagebox.showerror("Erreur", traceback.format_exc())
                print("[IA] ERREUR:\n", traceback.format_exc())

        bouton_train.configure(command=_train_model)

        # ============ SCROLL MOLETTE ============

        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # ============ BOUTON RETOUR ============

        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)

if __name__ == "__main__":
    app = Application()
    app.mainloop()
    