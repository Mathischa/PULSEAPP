# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict
import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter  # (si nécessaire ailleurs)

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# === Répertoire contenant les fichiers mensuels ===
from pathlib import Path
import os
import os, time, subprocess, unicodedata
from pathlib import Path
import winreg



from pathlib import Path
import os
import unicodedata
# === Modèle ML (scikit-learn) ===
try:
    from sklearn.ensemble import RandomForestRegressor
except ImportError:
    RandomForestRegressor = None
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor

USER_ID = Path.home().name
BASE_DIR = Path(fr"C:\Users\{USER_ID}\SNCF")
import tkinter as tk

# --- PATCH pour empêcher les RuntimeError "main thread is not in main loop" ---

# Sauvegarde des __del__ originaux
_orig_image_del = getattr(tk.Image, "__del__", None)
_orig_var_del = getattr(tk.Variable, "__del__", None)

def _safe_image_del(self):
    if _orig_image_del is None:
        return
    try:
        _orig_image_del(self)
    except RuntimeError:
        # On ignore seulement le cas où Tk n'est plus dans la boucle principale
        pass

def _safe_var_del(self):
    if _orig_var_del is None:
        return
    try:
        _orig_var_del(self)
    except RuntimeError:
        # Pareil ici : on ne fait rien si Tk est déjà "mort"
        pass

if _orig_image_del is not None:
    tk.Image.__del__ = _safe_image_del

if _orig_var_del is not None:
    tk.Variable.__del__ = _safe_var_del

# On n’impose QUE la "queue" du chemin (fin de chemin), flexible sur tout le reste.
REQUIRED_TAIL = ["Partage - Invités", "Projet PULSE", "4. Données historiques", "Développement","Données"]

def _norm(s: str) -> str:
    """Normalise pour comparaisons: minuscules, accents retirés, espaces/underscores compressés."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("_", " ").strip()
    s = " ".join(s.split())
    return s

def _match_tail(path: Path, tail: list[str]) -> bool:
    """Vérifie que 'path' se termine par les éléments de 'tail' (nom pour nom), comparés normalisés."""
    parts = list(path.parts)
    if len(parts) < len(tail):
        return False
    # Compare en partant de la fin
    for i in range(1, len(tail)+1):
        if _norm(parts[-i]) != _norm(tail[-i]):
            return False
    return True

def find_sharepoint_base():
    """
    1) Si ton chemin 'par défaut' existe, on l’utilise.
    2) Sinon, on parcourt TOUT 'C:\\Users\\<USER_ID>\\SNCF' pour trouver un dossier
       qui SE TERMINE par 'Partage - Invités/Projet PULSE/4. Données historiques/Développement'.
    """
    # 1) Essai direct sur ton chemin "canonique"
    default_path = BASE_DIR / "DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions" / Path(*REQUIRED_TAIL)
    if default_path.exists():
        print(f"[OK] Bibliothèque détectée (DCF par défaut) : {default_path}")
        return default_path

    # 2) Scan générique : on ne regarde QUE la fin de chemin (tail), robustes aux variations amont
    print("[INFO] Recherche générique de    la queue de chemin '.../Partage - Invités/Projet PULSE/4. Données historiques/Développement'...")
    candidates = []
    for root, dirs, _ in os.walk(BASE_DIR):
        # accélère: on ne teste que si "Développement" est présent dans le dossier courant
        if any(_norm(d) == _norm(REQUIRED_TAIL[-1]) for d in dirs):
            for d in dirs:
                full = Path(root) / d
                if _match_tail(full, REQUIRED_TAIL):
                    candidates.append(full)

    if candidates:
        # Choix simple : on prend le plus long chemin (le plus profond) ou simplement le premier trouvé
        best = sorted(candidates, key=lambda p: len(p.parts), reverse=True)[0]
        print(f"[OK] Bibliothèque trouvée automatiquement : {best}")
        return best

    print("❌ Aucune bibliothèque SharePoint locale trouvée correspondant à la queue requise.")
    return None

# --- Résolution finale (inchangé) ---
DEV_PATH = find_sharepoint_base()
if not DEV_PATH:
    raise FileNotFoundError("Impossible de localiser la bibliothèque SharePoint (vérifie la synchro OneDrive/SharePoint).")

FICHIER_EXCEL_DIR = str(DEV_PATH / "Historique Prévisions Réel Filiales")
FICHIER_CONFIG_SECTIONS = str(DEV_PATH / "Filiales Analysées.xlsx")
image_path = str(DEV_PATH / "Images" / "logo_Pulse.png")
BASE_DONNEES_DIR = str(DEV_PATH / "Données Réelles")


print("\n[INIT] Dossier Résultats :", FICHIER_EXCEL_DIR)
print("[INIT] Fichier Config    :", FICHIER_CONFIG_SECTIONS)
print("[INIT] Image            :", image_path)
print("[INIT] Dossier Données   :\n", BASE_DONNEES_DIR)

if not Path(FICHIER_EXCEL_DIR).exists():
    print(f"⚠️  Dossier 'Résultats' introuvable : {FICHIER_EXCEL_DIR}")
if not Path(FICHIER_CONFIG_SECTIONS).exists():
    print(f"⚠️  Fichier 'Filiales Analysées.xlsx' introuvable : {FICHIER_CONFIG_SECTIONS}")
if not Path(image_path).exists():
    print(f"⚠️  Image 'logo_Pulse.png' introuvable : {image_path}")

# ======================
#   FONCTIONS DE SÉCURISATION
# ======================

def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")

def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)

def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(["attrib", path], capture_output=True, text=True, shell=True)
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None

def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")

def pin_and_hydrate(path: str, timeout=30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l’hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False

def robust_load_workbook(path: str, retries=3, delay=1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} ({type(e).__name__}: {e})")
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e

def diag_path(path: str):
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
    
# === Fichier Excel contenant la configuration des sections ===
FEUILLE_CONFIG_SECTIONS = "CONFIG_SECTIONS"  # ou le vrai nom de la feuille dans ton fichier
COL_DEST = 1   # A
COL_SOURCE = 2 # B
COL_PREV = 3   # C

def charger_sections_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit le fichier Excel de configuration des sections et renvoie une liste
    de dictionnaires {"source": ..., "prev": ..., "dest": ...}.
    """
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return []

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return []

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return []

    ws = wb[FEUILLE_CONFIG_SECTIONS]
    mapping = []

    row = 2  # ligne 1 = en-têtes
    while True:
        dest_val = ws.cell(row=row, column=COL_DEST).value
        source_val = ws.cell(row=row, column=COL_SOURCE).value
        prev_val = ws.cell(row=row, column=COL_PREV).value

        if dest_val is None or str(dest_val).strip() == "":
            break  # fin du tableau

        mapping.append({
            "dest": str(dest_val).strip(),
            "source": str(source_val).strip() if source_val else str(dest_val).strip(),
            "prev": str(prev_val).strip() if prev_val else str(dest_val).strip(),
        })
        row += 1

    wb.close()

    print("[INFO] Sections chargées depuis 'Filiales Analysées.xlsx' :")
    for m in mapping:
        print(f"   - dest='{m['dest']}' | source='{m['source']}' | prev='{m['prev']}'")

    return mapping

def charger_noms_feuilles_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit uniquement la colonne A (dest) de la feuille de config et
    retourne la liste des noms de feuilles à utiliser (strings non vides, uniques, ordonnées).
    """
    noms = []
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return noms

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return noms

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return noms

    ws = wb[FEUILLE_CONFIG_SECTIONS]

    row = 2  # ligne 1 = en-têtes
    seen = set()
    while True:
        val = ws.cell(row=row, column=COL_DEST).value
        if val is None or str(val).strip() == "":
            break
        name = str(val).strip()
        if name not in seen:
            noms.append(name)
            seen.add(name)
        row += 1

    wb.close()

    print("[INFO] Feuilles (dest) chargées depuis 'Filiales Analysées.xlsx' :")
    for n in noms:
        print(f"   - {n}")
    return noms

# === Construire dynamiquement `sections` à partir de la colonne A ===
_dest_names = charger_noms_feuilles_depuis_cells()

# Création directe du dictionnaire sections
sections = {name: name for name in _dest_names}  # clé logique -> nom de feuille Excel


# Optionnel mais pratique : caler la feuille de référence sur la première feuille disponible
if _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]   # ex: "SNCF_SA"

SECTIONS_CONFIG = charger_sections_depuis_cells()
print(SECTIONS_CONFIG)

print(FICHIER_EXCEL_DIR)

# === Regex utilitaires ===
_MENSUEL_RE = re.compile(r"^Historique_prev_reel_filiales_(\d{4})_(\d{2})\.xlsx$")
_RE_PREV = re.compile(r"^Prévision\s+\d{2}/\d{2}", re.I)

def _is_prev(h):  return isinstance(h, str) and _RE_PREV.search(h.strip())
def _is_dates(h): return isinstance(h, str) and h.strip().lower().startswith("date")
def _is_reel(h):  return isinstance(h, str) and "réel" in h.lower()

def _parse_excel_date(v):
    """Convertit proprement tout format de date Excel (float, datetime, str...)"""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        # borne basse/haute pour éviter les overflow absurdes
        if 10000 <= v <= 60000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")).normalize()
            except Exception:
                return None
        else:
            return None
    if isinstance(v, str):
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        return ts.normalize() if pd.notna(ts) else None
    ts = pd.to_datetime(v, errors="coerce")
    return ts.normalize() if pd.notna(ts) else None

# -------------------------------------------------------------------------
# 1️⃣ Structures globales
# -------------------------------------------------------------------------
STRUCT = {}
TOKENS = {}
PREV_UNION = set()
FEUILLE_REFERENCE = "SNCF_SA"
CACHE = {}

# Index annuel: (section, flux) -> { "years": { year: { "row_idx": [...], "prof_idx": [...], "headers": [...] } } }
YEAR_INDEX = {}

def _clean_profil_label(raw, i):
    """Nettoyage identique à l'historique: (K€) retiré, 'Prévision' -> 'Profil', fallback."""
    if raw is None:
        return f"Profil {i+1}"
    s = str(raw).replace("(K€)", "").replace("Prévision", "Profil").strip()
    return s if s else f"Profil {i+1}"

def _flux_name_from_token(section, token_col):
    """Retrouve le nom de flux à partir de la colonne 'token' (col_start) dans TOKENS[section]."""
    for name, tok in TOKENS.get(section, []):
        if tok == token_col:
            return name
    return None


# -------------------------------------------------------------------------
# 2️⃣ Lister les fichiers mensuels
# -------------------------------------------------------------------------

def _lister_fichiers_mensuels():
    """
    Parcourt récursivement FICHIER_EXCEL_DIR pour trouver tous les fichiers
    'Historique_prev_reel_filiales_YYYY_MM.xlsx', y compris dans des sous-dossiers d'années.
    En cas de doublons pour un même (YYYY, MM), conserve le fichier le plus récent (mtime).
    """
    print("🔍 Recherche des fichiers mensuels (scan récursif des sous-dossiers d'années)...")

    if not os.path.isdir(FICHIER_EXCEL_DIR):
        print(f"⚠️ Dossier introuvable : {FICHIER_EXCEL_DIR}")
        return []

    best = {}  # (year, month) -> {"path": str, "mtime": float}
    try:
        for root, _, files in os.walk(FICHIER_EXCEL_DIR):
            for fname in files:
                m = _MENSUEL_RE.match(fname)
                if not m:
                    continue
                y, mth = int(m.group(1)), int(m.group(2))
                fullpath = os.path.join(root, fname)
                try:
                    mtime = os.path.getmtime(fullpath)
                except OSError:
                    mtime = 0.0

                key = (y, mth)
                if key not in best or mtime > best[key]["mtime"]:
                    best[key] = {"path": fullpath, "mtime": mtime}
    except Exception as e:
        print(f"⚠️ Erreur pendant le scan récursif : {e}")
        return []

    if not best:
        print("⚠️ Aucun fichier Historique_prev_reel_filiales_YYYY_MM.xlsx trouvé (sous-dossiers inclus).")
        return []

    files = sorted([(y, mth, info["path"]) for (y, mth), info in best.items()])
    print(f"✅ {len(files)} fichiers retenus : {[os.path.basename(f[2]) for f in files]}")
    return files

def _parse_prev_header_sort_key(h):
    # Essaie d'en déduire un tri MM/YY si possible, sinon garde l’ordre d’apparition
    import re
    s = str(h)
    m = re.search(r'(\d{2})/(\d{2,4})', s)
    if not m:
        return (9999, 99, s)
    mm = int(m.group(1))
    yy = int(m.group(2))
    if yy < 100:
        yy += 2000
    return (yy, mm, s)

def _reconcile_headers(B, new_headers):
    """
    Fusionne les headers déjà présents dans B avec ceux du fichier courant :
    - construit l’union ordonnée
    - re-map B['prev_vals'] dans ce nouvel ordre
    - rétro-remplit de None pour les séries nouvellement apparues
    """
    if "prev_headers" not in B or "prev_vals" not in B:
        B["prev_headers"] = []
        B["prev_vals"] = []

    old_headers = list(B["prev_headers"])
    # union en conservant l’ordre d’apparition, puis tente un tri chronologique
    seen = {}
    for h in old_headers + list(new_headers):
        if h not in seen:
            seen[h] = None
    union = list(seen.keys())
    union.sort(key=_parse_prev_header_sort_key)

    # rien à faire ?
    if union == old_headers:
        return

    # re-map des anciennes séries dans le nouveau canevas
    old_idx = {h: i for i, h in enumerate(old_headers)}
    n_rows = len(B.get("dates", []))
    new_prev_vals = []
    for h in union:
        if h in old_idx:
            new_prev_vals.append(B["prev_vals"][old_idx[h]])
        else:
            # nouvelle série : rétro-remplir pour les lignes déjà connues
            new_prev_vals.append([None] * n_rows)

    B["prev_headers"] = union
    B["prev_vals"] = new_prev_vals

# -------------------------------------------------------------------------
# 3️⃣ Lecture de la structure sur un fichier de référence
# -------------------------------------------------------------------------
def _lire_structure_reference(path_ref):
    print(f"\n📘 Lecture de la structure de référence ({FEUILLE_REFERENCE}) → {os.path.basename(path_ref)}")
    df_head = pd.read_excel(path_ref, sheet_name=FEUILLE_REFERENCE, header=None, nrows=3)
    row2 = df_head.iloc[1].tolist()
    row3 = df_head.iloc[2].tolist()

    ref_od = OrderedDict()
    ref_tokens = []
    col = 2  # colonne C (0-based)
    token_col = 3

    while col < len(row2):
        flux_name = row2[col]
        if pd.isna(flux_name):
            break
        flux_name = str(flux_name).strip()

        prev_headers, prev_cols = [], []
        c = col + 1
        while c < len(row3):
            h = row3[c]
            if isinstance(h, str) and "Prévision" in h:
                prev_headers.append(h.strip())
                prev_cols.append(c)
                c += 2
                continue
            break

        if prev_headers:
            ref_od[flux_name] = {
                "col_start": col + 1,
                "prev_headers": prev_headers,
                "prev_cols": [x + 1 for x in prev_cols],
                "dates_col": col,
                "reel_col": col + 1,
            }
            ref_tokens.append((flux_name, token_col))
            PREV_UNION.update(prev_headers)

        col += (2 + 2 * len(prev_headers) + 1)
        token_col += (2 + 2 * len(prev_headers) + 1)

    print(f"   → {len(ref_od)} flux détectés ({len(PREV_UNION)} prévisions).")
    for sec in sections:
        STRUCT[sec] = ref_od.copy()
        TOKENS[sec] = ref_tokens.copy()
    print("✅ Structure copiée sur toutes les feuilles.\n")

# -------------------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# -------------------------------------------------------------------------
def _ensure_flux_bucket(section, flux_name, headers=None):
    key = (section, flux_name)
    if key not in CACHE:
        CACHE[key] = {
            "dates": [],
            "reel": [],
            "prev_headers": [],   # ← vide
            "prev_vals": [],      # ← vide
        }
    return CACHE[key]

def _accumuler_valeurs_tous_mois(files_list):
    """
    Lit chaque fichier mensuel et alimente CACHE[(section, flux)] en alignant
    les séries de prévision sur l'union ordonnée des en-têtes (headers).
    - Réconciliation des headers à chaque fichier (ajout rétro-rempli de None).
    - Insertion directe ligne par ligne dans B["dates"], B["reel"], B["prev_vals"][k].
    """
    print(f"📊 Lecture de {len(files_list)} fichiers mensuels (toutes feuilles)...")
    for idx, (_, mois, path) in enumerate(files_list, 1):
        print(f"   [{idx}/{len(files_list)}] → {os.path.basename(path)}")

        try:
            all_sheets = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as e:
            print(f"⚠️ Erreur lecture {path}: {e}")
            continue

        for sec, feuille in sections.items():
            if feuille not in all_sheets:
                continue

            df = all_sheets[feuille].copy()
            if df.shape[0] < 5:
                continue

            row2 = df.iloc[1].tolist()
            row3 = df.iloc[2].tolist()
            col = 2  # colonne C (0-based -> ici index pandas)

            while col < len(row2):
                flux_name = row2[col]
                if pd.isna(flux_name):
                    break
                flux_name = str(flux_name).strip()

                # Détection des colonnes "Prévision"
                prev_headers, prev_cols = [], []
                c = col + 1
                while c < len(row3):
                    h = row3[c]
                    if isinstance(h, str) and "Prévision" in h:
                        prev_headers.append(h.strip())
                        prev_cols.append(c)
                        c += 2  # on saute la colonne (k€) qui suit chaque prévision
                        continue
                    break

                if not prev_headers:
                    # Bloc sans prévisions (ex: espaces, colonnes diverses) → on saute "gros"
                    col += 10
                    continue

                # 1) Bucket pour ce flux
                B = _ensure_flux_bucket(sec, flux_name)

                # 2) Réconcilier les en-têtes avec ceux du fichier courant
                _reconcile_headers(B, prev_headers)

                # 3) Map header -> colonne pour CE fichier
                header_to_col = {h: prev_cols[j] for j, h in enumerate(prev_headers)}

                # 4) Pousser les lignes : date, réel, et séries alignées
                rows_appended = 0
                for r in range(3, df.shape[0]):
                    date_val = df.iat[r, col - 1]
                    if pd.isna(date_val):
                        continue
                    d = _parse_excel_date(date_val)
                    if d is None:
                        continue

                    B["dates"].append(d)

                    r_val = df.iat[r, col]
                    B["reel"].append(r_val if pd.notna(r_val) else 0)

                    # Pour chaque header du bucket, si la colonne existe ce mois-ci → valeur ; sinon → None
                    for k, h in enumerate(B["prev_headers"]):
                        if h in header_to_col:
                            cpr = header_to_col[h]
                            v = df.iat[r, cpr]
                            B["prev_vals"][k].append(v if pd.notna(v) else None)
                        else:
                            B["prev_vals"][k].append(None)

                    rows_appended += 1

                # Avance au bloc suivant : (dates + réel) + 2 colonnes par prévision + 1 de séparation
                col += (2 + 2 * len(prev_headers) + 1)

    print(f"✅ Cache complété : {len(CACHE)} flux au total (valeurs réelles et prévisions incluses).\n")

    # Récapitulatif par section (contrôle)
    print("📊 Récapitulatif des flux chargés par section :")
    counts = {}
    for (section, flux), data in CACHE.items():
        counts.setdefault(section, 0)
        counts[section] += 1
    for section, n in counts.items():
        total_lignes = sum(len(CACHE[(section, f)]['dates']) for f in [x for x, _ in CACHE if _ == section])
        print(f"   - {section} : {n} flux ({total_lignes} lignes cumulées)")
    print()

# -------------------------------------------------------------------------
# 5️⃣ Index annuel
# -------------------------------------------------------------------------
def _build_year_index():
    """
    Pour chaque (section, flux) dans CACHE :
      - regroupe les indices de lignes par année
      - détermine quels profils (colonnes) sont 'actifs' sur l'année (au moins une valeur != 0 / non None)
      - prépare la liste de headers nettoyés pour l'année
    Remplit YEAR_INDEX[(section, flux)].
    """
    YEAR_INDEX.clear()
    for (section, flux_name), B in CACHE.items():
        dates = B.get("dates", [])
        prev_vals = B.get("prev_vals", [])
        headers = B.get("prev_headers", [])

        # indices de lignes par année
        rows_by_year = {}
        for i, d in enumerate(dates):
            y = d.year
            rows_by_year.setdefault(y, []).append(i)

        years_map = {}
        for y, row_idx in rows_by_year.items():
            prof_idx = []
            for k, serie in enumerate(prev_vals):
                # activité = au moins une valeur non None et != 0 sur ces lignes
                active = False
                for i in row_idx:
                    if i < len(serie):
                        v = serie[i]
                        if v is None:
                            continue
                        try:
                            if float(v) != 0.0:
                                active = True
                                break
                        except Exception:
                            active = True
                            break
                if active:
                    prof_idx.append(k)

            headers_year = [_clean_profil_label(headers[k] if k < len(headers) else None, k) for k in prof_idx]

            years_map[y] = {
                "row_idx": row_idx,       # indices de lignes pour cette année
                "prof_idx": prof_idx,     # indices de colonnes actives pour cette année
                "headers": headers_year,  # labels propres des profils actifs
            }

        YEAR_INDEX[(section, flux_name)] = {"years": years_map}

# -------------------------------------------------------------------------
# 6️⃣ Initialisation complète
# -------------------------------------------------------------------------
def _init_full_load():
    print("🚀 Initialisation complète du cache de données...")
    files = _lister_fichiers_mensuels()
    if not files:
        print("ℹ️ Aucun fichier trouvé — initialisation du cache ignorée.\n")
        return

    ref_path = files[-1][2]  # dernier mois = référence
    _lire_structure_reference(ref_path)
    _accumuler_valeurs_tous_mois(files)
    _build_year_index()  # ✅ construire l’index après remplissage du CACHE
    print("✅ Chargement complet terminé.\n")


# Lance le chargement maintenant que tout est défini
_init_full_load()

# -------------------------------------------------------------------------
# 7️⃣ Variables dérivées compatibles avec l'existant
# -------------------------------------------------------------------------
previsions_triees = sorted(PREV_UNION)
nb_prev = len(previsions_triees)
taille_bloc = 2 + 2 * nb_prev + 1
print(f"📈 Nombre total de prévisions : {nb_prev}")
print(f"📦 Taille d’un bloc de flux : {taille_bloc}\n")


# -------------------------------------------------------------------------
# 8️⃣ Fonctions publiques
# -------------------------------------------------------------------------
def charger_donnees(feuille, taille_bloc_param):
    """Retourne le nom de la feuille (section) et la liste des flux détectés, en excluant certains flux inutiles."""
    print(f"➡️ charger_donnees() appelé pour '{feuille}'")

    # Liste des flux à ignorer
    noms_a_exclure = [
        "Trésorerie de fin",
        "Cashpool",
        "Emprunts",
        "Tirages Lignes CT",
        "Variation de collatéral",
        "Créances CDP",
        "Placements",
        "CC financiers",
        "Emprunts / Prêts - Groupe",
        "Encours de financement",
        "Endettement Net"
    ]

    # Récupère tous les flux
    all_noms = list(TOKENS.get(feuille, []))

    # Filtrage : on garde uniquement ceux qui ne sont pas exclus
    noms = [(name, col) for name, col in all_noms if name not in noms_a_exclure]
    for j in noms :
        print(f"noms : {j}")
    print(len(noms))
    
    if not noms:
        print(f"⚠️ Aucun flux valide trouvé pour la feuille '{feuille}'")
    else:
        print(f"   ✅ {len(noms)} flux conservés (sur {len(all_noms)} totaux)")

    return feuille, noms
# Profils à exclure pour 2023 (d’après tes listes / cases à cocher)
EXCLUDED_DATES_2023 = {
    "02/01",
    "09/01",
    "16/01",
    "30/05",
    "05/06",
    "19/06",
    "26/06",
    "03/07",
    "10/07",
    "17/07",
}

def extraire_valeurs(ws, col_start, nb_prev_param, annee=None, annee_min=None, annee_max=None):
    """
    Retourne (dates, reel, previsions, noms_profils)

    - Si 'annee' est fourni : utilise YEAR_INDEX pour ne garder que
      *les lignes* de l'année et *les profils actifs* de l'année,
      puis on filtre éventuellement certains profils (2023) → le RÉEL reste inchangé.
    - Sinon : filtre sur plage ou aucun filtre (comportement historique).
    """

    DEBUG_EXTRAIRE = True
    def _dbg(*args):
        if DEBUG_EXTRAIRE:
            print("[extraire_valeurs]", *args)

    section = ws
    flux_name = _flux_name_from_token(section, col_start)
    _dbg(f"CALL: section={section!r} col_start={col_start} nb_prev_param={nb_prev_param} "
         f"annee={annee} annee_min={annee_min} annee_max={annee_max}")

    if not flux_name:
        _dbg("ABORT: flux_name introuvable -> retour listes vides/pad prév.")
        return [], [], [[] for _ in range(nb_prev_param)], []

    B = CACHE.get((section, flux_name))
    if not B or not B.get("dates"):
        _dbg(f"ABORT: pas de cache ou dates vides pour {(section, flux_name)}")
        return [], [], [[] for _ in range(nb_prev_param)], []

    dates_all = B["dates"]
    reel_all  = B["reel"]
    prev_all  = B.get("prev_vals", [])
    headers   = B.get("prev_headers", [])   # ex: "Prévision 16/01 (K€)"

    _dbg(f"CACHE: len(dates_all)={len(dates_all)}, len(reel_all)={len(reel_all)}, "
         f"nb_profils_total={len(prev_all)}, nb_headers={len(headers)}")

    # ==============================
    # CAS 1 : année précise
    # ==============================
    if annee is not None:
        years = YEAR_INDEX.get((section, flux_name), {}).get("years", {})
        info = years.get(annee)
        if not info:
            _dbg(f"ANNEE={annee}: YEAR_INDEX absent/vide -> retour listes vides.")
            return [], [], [], []

        # Ces indices NE CHANGENT PAS pour les lignes → le RÉEL reste intact
        row_idx  = info["row_idx"]      # indices des dates pour cette année
        prof_idx = info["prof_idx"]     # indices des colonnes de prévision actives

        _dbg(f"ANNEE={annee}: rows={len(row_idx)} profils_actifs={len(prof_idx)}")

        # ---- Filtre spécial 2023 : on enlève certains profils, mais on garde le réel ----
        if annee == 2023:
            def _is_excluded(k: int) -> bool:
                if k < 0 or k >= len(headers):
                    return False
                lab = str(headers[k])  # "Prévision 16/01 (K€)"
                for d in EXCLUDED_DATES_2023:
                    if d in lab:
                        return True
                return False

            prof_idx_before = list(prof_idx)
            prof_idx = [k for k in prof_idx if not _is_excluded(k)]
            _dbg(f"ANNEE=2023: prof_idx avant filtre = {prof_idx_before}")
            _dbg(f"ANNEE=2023: prof_idx après  filtre = {prof_idx}")

        # ---- Extraction des lignes (réel inchangé) ----
        dates = [dates_all[i] for i in row_idx]
        reel  = [reel_all[i]  for i in row_idx]

        # ---- Extraction des prévisions uniquement pour les profils retenus ----
        previsions = []
        for k in prof_idx:
            serie = prev_all[k] if k < len(prev_all) else []
            previsions.append([serie[i] if i < len(serie) else None for i in row_idx])

        # Noms de profils alignés sur prof_idx
        noms_profils = [
            _clean_profil_label(headers[k] if k < len(headers) else None, k)
            for k in prof_idx
        ]

        _dbg(f"RETURN(annee): len(dates)={len(dates)}, len(reel)={len(reel)}, "
             f"nb_profils={len(previsions)}, noms_profils={noms_profils}")
        return dates, reel, previsions, noms_profils

    # ==============================
    # CAS 2 : plage / historique
    # ==============================
    def _keep(d):
        y = d.year
        if annee_min is not None and annee_max is not None:
            return annee_min <= y <= annee_max
        if annee_min is not None:
            return y >= annee_min
        if annee_max is not None:
            return y <= annee_max
        return True

    idxs = [i for i, d in enumerate(dates_all) if _keep(d)]
    dates = [dates_all[i] for i in idxs]
    reel  = [reel_all[i]  for i in idxs]
    previsions = [[serie[i] if i < len(serie) else None for i in idxs] for serie in prev_all]

    noms_profils = [_clean_profil_label(h, i) for i, h in enumerate(headers)]

    _dbg(f"PLAGE/HISTO: indices_retenus={len(idxs)} "
         f"(min_year={min((d.year for d in dates), default=None)}, "
         f"max_year={max((d.year for d in dates), default=None)})")
    _dbg(f"RETURN(plage): len(dates)={len(dates)}, len(reel)={len(reel)}, "
         f"nb_profils={len(previsions)}, len(noms_profils)={len(noms_profils)}; "
         f"exemple_date={dates[0] if dates else None}")

    return dates, reel, previsions, noms_profils


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        self.attributes("-fullscreen", True)
        self.geometry("1300x900")
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

#===================Page Accueil + fonctions de navigation===================
    def creer_accueil(self):
        # -- RESET & imports locaux
        import customtkinter as ctk
        from PIL import Image
        import tkinter as tk
        from customtkinter import CTkImage
        from datetime import datetime
        import os

        # === THEME / BASE ===
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color="#0b1220")
        except Exception:
            pass

        # Détruit proprement les enfants
        self.vider_fenetre()

        # =========================
        #  Helpers data pour KPI
        # =========================
        def _fmt_dt(ts):
            try:
                return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return "—"

        # Fichiers mensuels
        try:
            fichiers = _lister_fichiers_mensuels()  # supposé renvoyer (annee, mois, path)
            nb_fichiers = len(fichiers)

            if fichiers:
                last_mtime = max(os.path.getmtime(p) for _, _, p in fichiers if os.path.exists(p))
                derniere_maj = _fmt_dt(last_mtime)
                # années disponibles (1er champ du tuple)
                annees_dispos = sorted({int(a) for a, _, _ in fichiers if a is not None})
            else:
                derniere_maj = "—"
                annees_dispos = []
        except Exception:
            nb_fichiers, derniere_maj = 0, "—"
            annees_dispos = []

        # Filiales
        try:
            if SECTIONS_CONFIG:
                filiales_set = {s.get("dest") for s in SECTIONS_CONFIG if s.get("dest")}
            else:
                filiales_set = set(_dest_names) if '_dest_names' in globals() else set()
            filiales_list = sorted(f for f in filiales_set if f)
            nb_filiales = len(filiales_list)
        except Exception:
            filiales_list = []
            nb_filiales = 0

        # =========================
        #  GRID RACINE (window)
        # =========================
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        # =========================
        #  HEADER
        # =========================
        header = ctk.CTkFrame(self, fg_color="#0f1b31", corner_radius=0, height=72)
        header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        header.grid_columnconfigure(0, weight=0)
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        # Barre blanche sous le header
        barre_header = ctk.CTkFrame(self, fg_color="white", height=2)
        barre_header.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Logo à gauche
        logo_wrap = ctk.CTkFrame(header, fg_color="transparent")
        logo_wrap.grid(row=0, column=0, sticky="w", padx=24, pady=12)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h, new_w = 44, int(44 * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(logo_wrap, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            ctk.CTkLabel(logo_wrap, text="PULSE", font=("Segoe UI Semibold", 20),
                        text_color="white").pack(side="left")

        titre = ctk.CTkLabel(
            header,
            text="PROJET PULSE — ANALYSE DE L’EXISTANT",
            font=("Segoe UI Semibold", 18, "bold"),
            text_color="white",
        )
        titre.grid(row=0, column=1, sticky="w", padx=8)

        # Actions header à droite
        header_actions = ctk.CTkFrame(header, fg_color="transparent")
        header_actions.grid(row=0, column=2, sticky="e", padx=20)

        ctk.CTkButton(
            header_actions, text="❌ Quitter",
            fg_color="#C21515", hover_color="#A00000",
            command=self.demander_confirmation_quit
        ).grid(row=0, column=1)

        # =========================
        #  SIDEBAR
        # =========================
        sidebar = ctk.CTkFrame(self, fg_color="#0e162a", corner_radius=0, width=260)
        sidebar.grid(row=1, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self.grid_rowconfigure(1, weight=1)
        sidebar.grid_rowconfigure(11, weight=1)

        inner = ctk.CTkFrame(sidebar, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=14, pady=14)

        def _nav_btn(parent, text, cmd, emoji=""):
            btn = ctk.CTkButton(
                parent, text=f"{emoji}  {text}", anchor="w",
                width=220, height=42, corner_radius=12,
                fg_color="#15305b", hover_color="#1d3e77",
                font=("Segoe UI", 13, "bold"),
                command=cmd
            )

            def _enter(_):
                btn.configure(fg_color="#1d3e77")

            def _leave(_):
                btn.configure(fg_color="#15305b")

            btn.bind("<Enter>", _enter)
            btn.bind("<Leave>", _leave)
            return btn

        ctk.CTkLabel(inner, text="Navigation", text_color="#9fb7dd",
                    font=("Segoe UI Semibold", 14)).pack(anchor="w", pady=(4, 12))

        _nav_btn(inner, "IA — Prédiction", self.creer_page_ia_prediction, "🤖").pack(fill="x", pady=6)

        separator = ctk.CTkFrame(sidebar, fg_color="#1a2745", height=2)
        separator.pack(fill="x", padx=14, pady=(4, 2))

        foot = ctk.CTkLabel(
            sidebar,
            text=f"v1.0 • Dernière MAJ : {derniere_maj}",
            text_color="#6e86ad",
            font=("Segoe UI", 11)
        )
        foot.pack(anchor="w", padx=20, pady=(8, 10))

        # =========================
        #  MAIN CONTENT
        # =========================
        main = ctk.CTkScrollableFrame(self, fg_color="#0b1220", corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        self.grid_columnconfigure(1, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # ====== HERO ======
        hero = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 14))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(
            hero,
            text="Bienvenue 👋",
            font=("Segoe UI Semibold", 24, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 0))

        subtitle_text = (
            "Console d'analyse des flux de trésorerie : réels, prévisions, écarts et détection d'anomalies."
        )
        subtitle_label = ctk.CTkLabel(
            hero,
            text="",
            font=("Segoe UI", 13),
            text_color="#c9defa",
            wraplength=520,
            justify="left"
        )
        subtitle_label.grid(row=1, column=0, sticky="w", padx=18, pady=(6, 10))

        def _typewriter(i=0):
            if i <= len(subtitle_text):
                subtitle_label.configure(text=subtitle_text[:i])
                self.after(12, _typewriter, i + 1)

        _typewriter()

        # Pills
        pills = ctk.CTkFrame(hero, fg_color="transparent")
        pills.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 16))

        def _pill(parent, text, emoji=""):
            f = ctk.CTkFrame(parent, fg_color="#152544", corner_radius=20)
            f.pack(side="left", padx=6)
            ctk.CTkLabel(
                f,
                text=f"{emoji}  {text}",
                font=("Segoe UI", 11),
                text_color="#c9defa"
            ).pack(padx=12, pady=5)

        _pill(pills, "Réel vs prévisions multi-horizon", "📈")
        _pill(pills, "Analyse d'écarts & anomalies", "🧠")
        _pill(pills, "Vue filiales / flux / profils", "📊")

        # Illustration graphique (3 mini graphes)
        viz = ctk.CTkFrame(hero, fg_color="#111b33", corner_radius=18)
        viz.grid(row=0, column=1, rowspan=3, sticky="e", padx=18, pady=12)
        viz.grid_columnconfigure((0, 1, 2), weight=1)
        viz.grid_rowconfigure(0, weight=1)

        # --- Graph 1 : barres animées ---
        graph1 = ctk.CTkFrame(viz, fg_color="transparent")
        graph1.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        bars_h = [52, 34, 78, 60]
        bar_widgets = []
        for h_val in bars_h:
            col = ctk.CTkFrame(graph1, fg_color="transparent")
            col.pack(side="left", padx=4, fill="y", expand=True)

            base = ctk.CTkFrame(col, fg_color="#1b2945", corner_radius=8, width=18, height=85)
            base.pack(side="bottom", pady=(0, 4))

            bar = ctk.CTkFrame(col, fg_color="#2563eb", corner_radius=8, width=18, height=1)
            bar.place(relx=0.5, rely=1.0, anchor="s")
            bar_widgets.append((bar, h_val))

        def _animate_bars(step=0, steps=18):
            factor = min(1.0, step / float(steps))
            for bar, target in bar_widgets:
                bar.configure(height=int(target * factor))
            if step < steps:
                self.after(25, _animate_bars, step + 1, steps)

        _animate_bars()

        # --- Graph 2 : courbe ---
        graph2 = ctk.CTkFrame(viz, fg_color="transparent")
        graph2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        pts = [10, 40, 22, 70, 50, 90, 60, 80, 55, 100]
        w, h = 120, 90

        canvas2 = ctk.CTkCanvas(graph2, width=w, height=h, bg="#111b33", bd=0, highlightthickness=0)
        canvas2.pack(fill="both", expand=True)

        max_v = max(pts)
        coords = [(i * (w // (len(pts) - 1)), h - int((v / max_v) * (h - 10))) for i, v in enumerate(pts)]
        for i in range(len(coords) - 1):
            canvas2.create_line(
                coords[i][0], coords[i][1],
                coords[i + 1][0], coords[i + 1][1],
                fill="#2563eb", width=3, smooth=True
            )
        for x, y in coords:
            canvas2.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#3c82ff", outline="")

        # --- Graph 3 : donut ---
        graph3 = ctk.CTkFrame(viz, fg_color="transparent")
        graph3.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)

        canvas3 = tk.Canvas(graph3, width=110, height=110,
                            bg="#111b33", bd=0, highlightthickness=0)
        canvas3.pack(fill="both", expand=True, padx=6, pady=6)

        values = [30, 15, 22, 18, 15]
        total = sum(values)
        start_angle = 0
        colors = ["#1d4ed8", "#2563eb", "#3b82f6", "#60a5fa", "#93c5fd"]
        bbox_outer = (8, 8, 102, 102)

        for i, v in enumerate(values):
            extent = 360 * v / total
            canvas3.create_arc(
                *bbox_outer,
                start=start_angle,
                extent=extent,
                fill=colors[i % len(colors)],
                outline="#111b33",
                style="pieslice"
            )
            start_angle += extent

        bbox_inner = (28, 28, 82, 82)
        canvas3.create_oval(*bbox_inner, fill="#111b33", outline="#111b33")

        # ====== KPI CARDS ======
        kpi = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        kpi.grid(row=1, column=0, sticky="ew", padx=24, pady=10)
        for i in range(3):
            kpi.grid_columnconfigure(i, weight=1)

        # ----- Zone "Détails" dynamique juste sous les KPI -----
        details = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        details.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 10))
        details.grid_columnconfigure(0, weight=1)

        details_title = ctk.CTkLabel(
            details,
            text="Détails",
            font=("Segoe UI Semibold", 15, "bold"),
            text_color="white"
        )
        details_title.grid(row=0, column=0, sticky="w", padx=16, pady=(14, 4))

        details_box = ctk.CTkTextbox(
            details,
            height=110,
            wrap="word",
            fg_color="#142544",
            text_color="white",
            font=("Segoe UI", 12),
            border_width=1,
            border_color="#223658",
            corner_radius=12
        )
        details_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 14))
        details_box.insert("1.0", "Clique sur un KPI pour afficher les détails (années ou filiales)…")
        details_box.configure(state="disabled")

        def _set_details(title: str, lines: list[str]):
            details_title.configure(text=title)
            details_box.configure(state="normal")
            details_box.delete("1.0", "end")
            if not lines:
                details_box.insert("1.0", "Aucune donnée disponible.")
            else:
                details_box.insert("1.0", "\n".join(lines))
            details_box.configure(state="disabled")

        def _show_years():
            if annees_dispos:
                lignes = ["Années disponibles :"]
                lignes += [f"• {a}" for a in annees_dispos]
            else:
                lignes = ["Aucune année détectée (pas de fichiers mensuels trouvés)."]
            _set_details("Détails — Fichiers mensuels détectés", lignes)

        def _show_filiales():
            if filiales_list:
                lignes = ["Filiales configurées :"]
                lignes += [f"• {f}" for f in filiales_list]
            else:
                lignes = ["Aucune filiale détectée dans la configuration."]
            _set_details("Détails — Filiales", lignes)

        def _kpi_card(parent, title, value, subtitle, col, detail_cb=None):
            card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16,
                                border_width=1, border_color="#223658")
            card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
            ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
            ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 28, "bold"),
                        text_color="white") \
                .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 4))
            ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                        text_color="#7ea2d8") \
                .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 6))

            if detail_cb is not None:
                btn = ctk.CTkButton(
                    card,
                    text="Voir détails",
                    height=28,
                    corner_radius=10,
                    fg_color="#2563eb",
                    hover_color="#1d4ed8",
                    command=detail_cb
                )
                btn.grid(row=3, column=0, sticky="w", padx=14, pady=(0, 10))

        _kpi_card(
            kpi,
            "Fichiers mensuels détectés",
            f"{nb_fichiers:,}".replace(",", " "),
            "Historique *_YYYY_MM.xlsx*",
            0,
            detail_cb=_show_years
        )
        _kpi_card(
            kpi,
            "Nombre de filiales",
            f"{nb_filiales:,}".replace(",", " "),
            "Distinctes dans la config",
            1,
            detail_cb=_show_filiales
        )
        _kpi_card(
            kpi,
            "Dernière actualisation",
            f"{derniere_maj}",
            "Basée sur les mtime",
            2,
            detail_cb=None
        )



        # ====== ACTIVITÉ RÉCENTE ======
        recent = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        recent.grid(row=4, column=0, sticky="ew", padx=24, pady=(10, 20))
        recent.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            recent, text="Activité récente",
            font=("Segoe UI Semibold", 16, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        table = ctk.CTkFrame(recent, fg_color="#142544", corner_radius=12,
                            border_width=1, border_color="#223658")
        table.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 16))
        for i in range(3):
            table.grid_columnconfigure(i, weight=[3, 1, 1][i])

        def _header(r, text):
            ctk.CTkLabel(table, text=text, text_color="#9fb7dd",
                        font=("Segoe UI Semibold", 12)) \
                .grid(row=r, column=["Fichier", "Taille", "Modifié le"].index(text),
                    sticky="w", padx=12, pady=(10, 6))

        _header(0, "Fichier")
        _header(0, "Taille")
        _header(0, "Modifié le")

        recent_files = []
        try:
            if fichiers:
                rec = sorted(
                    [(p, os.path.getmtime(p)) for _, _, p in fichiers if os.path.exists(p)],
                    key=lambda x: x[1], reverse=True
                )[:5]
                for p, ts in rec:
                    size = f"{(os.path.getsize(p) / 1024 / 1024):.1f} Mo"
                    recent_files.append((os.path.basename(p), size, _fmt_dt(ts)))
        except Exception:
            pass

        if not recent_files:
            recent_files = [("—", "—", "—")]

        for i, (name, size, dt_str) in enumerate(recent_files, start=1):
            ctk.CTkLabel(table, text=name, text_color="white", font=("Segoe UI", 12)) \
                .grid(row=i, column=0, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=size, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=1, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=dt_str, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=2, sticky="w", padx=12, pady=6)

    def vider_fenetre(self):
        for widget in self.winfo_children():
            widget.destroy()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.destroy()

#===================Page Prédiction IA===================
    def _rebuild_profils_ui2(self, section, flux_name, annee):
        """Reconstruit la select-box de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()

        # nouveau : on n'utilise plus self.vars_prev ici
        self.vars_prev = []
        self.profils_names_order = []   # ordre des profils pour mapping
        self.profils_listbox = None     # widget de sélection multiple

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils

        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # Label d'info
        lbl = tk.Label(
            self.profils_frame,
            text="Profils (Ctrl+clic pour multi-sélection) :",
            bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
        )
        lbl.pack(anchor="w", padx=4, pady=(0, 4))

        # Frame contenant la listbox + scrollbar
        container = tk.Frame(self.profils_frame, bg="#00122e")
        container.pack(fill="x", padx=4, pady=(0, 6))

        scrollbar = tk.Scrollbar(container, orient="vertical")
        self.profils_listbox = tk.Listbox(
            container,
            selectmode="multiple",
            exportselection=False,
            bg="#00122e",
            fg="white",
            font=('Segoe UI', 10),
            height=min(8, len(noms_profils)),
            highlightthickness=0,
            activestyle="dotbox"
        )
        self.profils_listbox.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.profils_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.profils_listbox.yview)

        for name in noms_profils:
            self.profils_listbox.insert("end", name)

        # petite sélection par défaut : tout sélectionner
        self.profils_listbox.select_set(0, "end")

    def creer_page_ia_prediction(self):
        """
        Page IA de prévision Enc. Autres Produits (N -> N+1).
        Cette méthode ne fait plus que :
        1) Vérifier les dépendances
        2) Réinitialiser la fenêtre
        3) Construire le dataset global
        4) Construire toute la page IA (UI + callbacks + entraînement)
        """
        # 1) Imports (idéalement à mettre en haut du fichier Python)
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # 2) sécurité sklearn
        if RandomForestRegressor is None:
            messagebox.showerror(
                "Dépendance manquante",
                "Le module 'scikit-learn' n'est pas installé.\n"
                "Installe-le avec : pip install scikit-learn"
            )
            return

        # 3) reset page
        self.vider_fenetre()

        # 4) construire le dataset global (toutes filiales, tous jours)
        df = self._ia_build_dataset_enc_autres_produits()
        if df is None:
            # _ia_build_dataset_enc_autres_produits affiche déjà le message et fait retour_menu()
            return

        # 5) construire la page IA complète (UI + callbacks + entraînement)
        self._ia_build_prediction_page(df)
        
    def _ia_build_dataset_enc_autres_produits(self):
        """
        Construit le DataFrame global df sur le flux 'Enc. Autres Produits'
        pour toutes les filiales, avec colonnes :
        section, date, y, year, month, dayofyear, section_id, roll_mean_7, roll_mean_30
        Retourne df ou None si aucune donnée.
        """
        import pandas as pd
        import numpy as np
        from tkinter import messagebox

        flux_cible = "Trafic Voyageurs"
        lignes = []

        # ---- helpers numériques ----
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        # ---- construction dataset (reprends ton code) ----
        for section_name, feuille in sections.items():
            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees pour {feuille} : {e}")
                continue

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                continue

            _, col_start = cible[0]
            try:
                dates, reel, _previsions, _noms_profils = extraire_valeurs(
                    ws, col_start, nb_prev, annee=None
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs pour {section_name}/{flux_cible} : {e}")
                continue

            for d, r in zip(dates, reel):
                y_val = _to_number(r)
                if y_val is None:
                    continue
                try:
                    d_ts = pd.to_datetime(d)
                except Exception:
                    continue
                lignes.append({
                    "section": section_name,
                    "date": d_ts,
                    "y": float(y_val)
                })

        if not lignes:
            messagebox.showinfo(
                "Information",
                "Aucune donnée trouvée pour le flux 'Trafic Voyageurs' sur les filiales."
            )
            self.retour_menu()
            return None

        df = pd.DataFrame(lignes).sort_values(["section", "date"]).reset_index(drop=True)
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["dayofyear"] = df["date"].dt.dayofyear

        # Encodage des filiales
        cat = df["section"].astype("category")
        df["section_id"] = cat.cat.codes

        # Lissages
        df["roll_mean_7"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(7, min_periods=1).mean()
        )
        df["roll_mean_30"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(30, min_periods=1).mean()
        )

        return df

    def _ia_build_prediction_page(self, df):
        """
        Construit toute l'interface IA + callbacks d'entraînement
        à partir du DataFrame global df déjà préparé.
        """
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # ---------- helpers numériques locaux ----------
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_float_or_nan(x):
            val = _to_number(x)
            return float(val) if val is not None else np.nan

        flux_cible = "Trafic Voyageurs"
        default_n_estimators = 200

        # ---------- variables profils & graph 2 ----------
        ia_profils_vars = []        # liste de tk.BooleanVar
        ia_profils_names = []       # noms de profils actifs N+1
        ia_profils_dates = []       # dates N+1
        ia_profils_series = []      # séries (une par profil actif)

        current_pred_df = None          # df_future_all pour N+1
        current_real_target_df = None   # réel sur N+1 (si dispo)
        current_target_year = None
        current_filiale_name = None

        ia_graph2_widget = None
        graph2_container_packed = False

        exported_pred_df = None         # pour export Excel prédictions
        analysis_table_frame = None
        analysis_tree = None
        analysis_export_button = None
        export_button = None            # bouton d’export des prédictions
        export_graph2_button = None

        model_artifacts = {
            "X_train": None,
            "y_train": None,
            "features": None,
            "cls_model": None,
            "reg_model": None,
        }


        # ============ HEADER ============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - IA PRÉDICTION  (N → N+1)",
            font=titre_font,
            text_color="white"
        ).pack(side="left", anchor="w")

        # Logo
        try:
            logo_img = Image.open(image_path)
            test = tk.Label(self, text="Test", font=titre_font)
            test.update_idletasks()
            text_h = test.winfo_reqheight()
            test.destroy()
            ratio = logo_img.width / max(logo_img.height, 1)
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"[IA] Erreur chargement logo IA : {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ============ CONTAINER SCROLLABLE ============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # ============ TITRE / TEXTE ============
        ctk.CTkLabel(
            scrollable_frame,
            text="IA - Prédiction sur le flux 'Trafic Voyageurs' (prévision année N+1)",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        ).pack(pady=15)

        ctk.CTkLabel(
            scrollable_frame,
            text=(
                "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                "Il apprend comment la valeur de 'Trafic Voyageurs' évolue d'une année à l'autre "
                "(N → N+1) pour chaque jour de l'année.\n"
                "Il prédit ensuite l’ensemble de l’année N+1 pour cette filiale.\n"
                "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
            ),
            font=("Segoe UI", 12),
            text_color="#c9defa",
            justify="left"
        ).pack(pady=(0, 10))

        # ============ ZONE PARAMÈTRES (CARD) ============
        params_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        params_frame.pack(fill="x", padx=10, pady=(5, 10))

        # Filiale
        ctk.CTkLabel(
            params_frame,
            text="Filiale :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))

        filiales = sorted(df["section"].unique().tolist())
        selected_filiale = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(
            params_frame,
            textvariable=selected_filiale,
            values=filiales,
            state="readonly",
            width=28
        )
        filiale_box.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        # Année N
        ctk.CTkLabel(
            params_frame,
            text="Année N (historique → N+1) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=1, sticky="w", padx=12, pady=(10, 2))

        years_sorted = sorted(df["year"].unique().tolist())
        annees_var = tk.StringVar(value=str(years_sorted[-1]))
        annees_box = ttk.Combobox(
            params_frame,
            textvariable=annees_var,
            values=[str(y) for y in years_sorted],
            state="readonly",
            width=18
        )
        annees_box.grid(row=1, column=1, sticky="w", padx=12, pady=(0, 10))

        # Hyperparamètres RF
        ctk.CTkLabel(
            params_frame,
            text="Hyperparamètres Random Forest (point de départ) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=2, sticky="w", padx=12, pady=(10, 2))

        # n_estimators
        ctk.CTkLabel(
            params_frame,
            text="n_estimators",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=2, sticky="w", padx=12, pady=(0, 2))

        n_estimators_var = tk.IntVar(value=default_n_estimators)
        slider_n = ctk.CTkSlider(
            params_frame,
            from_=50,
            to=600,
            number_of_steps=11,
            command=lambda v: n_estimators_var.set(int(v))
        )
        slider_n.set(default_n_estimators)
        slider_n.grid(row=2, column=2, sticky="we", padx=12, pady=(0, 8))

        lbl_n = ctk.CTkLabel(
            params_frame,
            text=f"{default_n_estimators}",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_n.grid(row=3, column=2, sticky="w", padx=12, pady=(0, 6))

        def _update_lbl_n(_=None):
            lbl_n.configure(text=str(n_estimators_var.get()))
        slider_n.configure(command=lambda v: (n_estimators_var.set(int(v)), _update_lbl_n()))

        # max_depth
        ctk.CTkLabel(
            params_frame,
            text="max_depth",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=3, sticky="w", padx=12, pady=(0, 2))

        max_depth_var = tk.IntVar(value=15)
        slider_d = ctk.CTkSlider(
            params_frame,
            from_=3,
            to=25,
            number_of_steps=11,
            command=lambda v: max_depth_var.set(int(v))
        )
        slider_d.set(15)
        slider_d.grid(row=2, column=3, sticky="we", padx=12, pady=(0, 8))

        lbl_d = ctk.CTkLabel(
            params_frame,
            text="15",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_d.grid(row=3, column=3, sticky="w", padx=12, pady=(0, 2))

        def _update_lbl_d(_=None):
            lbl_d.configure(text=str(max_depth_var.get()))
        slider_d.configure(command=lambda v: (max_depth_var.set(int(v)), _update_lbl_d()))

        use_depth_var = tk.BooleanVar(value=True)
        chk_depth = ctk.CTkCheckBox(
            params_frame,
            text="Limiter la profondeur",
            variable=use_depth_var,
            text_color="#c9defa",
            font=("Segoe UI", 11)
        )
        chk_depth.grid(row=4, column=3, sticky="w", padx=12, pady=(0, 10))

        # Bouton entraîner
        bouton_train = ctk.CTkButton(
            params_frame,
            text="🚀 Entraîner / ré-entraîner le modèle",
            width=240,
            height=40,
            corner_radius=16,
            fg_color="#008C4B",
            hover_color="#006C39",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_train.grid(row=1, column=4, rowspan=3, padx=14, pady=(0, 10), sticky="e")

        for c in range(5):
            params_frame.grid_columnconfigure(c, weight=1 if c in (2, 3) else 0)

        # ============ CARD PROFILS N+1 (CASES À COCHER) ============
        profils_card = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        profils_card.pack(fill="x", padx=10, pady=(0, 15))

        ctk.CTkLabel(
            profils_card,
            text="Profils de prévision disponibles sur l'année N+1 :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=12, pady=(10, 4))

        ia_profils_frame = tk.Frame(profils_card, bg="#00122e")
        ia_profils_frame.pack(fill="x", padx=12, pady=(0, 10))

        def _export_graph2_to_excel():
            """
            Export Excel des valeurs utilisées pour le graphe 2 :
            - date
            - réel (si dispo)
            - prédiction IA (si dispo)
            - profils cochés (si dispo)
            """
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal ia_profils_names, ia_profils_vars, ia_profils_dates, ia_profils_series

            if (current_pred_df is None or current_pred_df.empty) and (current_real_target_df is None or current_real_target_df.empty):
                messagebox.showinfo("Export", "Aucune donnée à exporter (ni réel, ni prévision).")
                return

            try:
                # ---------- base dates = union de toutes les dates visibles ----------
                all_dates = set()

                if current_real_target_df is not None and not current_real_target_df.empty:
                    all_dates |= set(pd.to_datetime(current_real_target_df["date"]).dt.normalize())

                if current_pred_df is not None and not current_pred_df.empty:
                    all_dates |= set(pd.to_datetime(current_pred_df["date"]).dt.normalize())

                if ia_profils_dates:
                    all_dates |= set(pd.to_datetime(ia_profils_dates).normalize())

                if not all_dates:
                    messagebox.showinfo("Export", "Aucune date exploitable à exporter.")
                    return

                dates_sorted = sorted(all_dates)
                df_out = pd.DataFrame({"date": pd.to_datetime(dates_sorted)})

                # ---------- colonne réel ----------
                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = (
                        current_real_target_df.assign(date_norm=pd.to_datetime(current_real_target_df["date"]).dt.normalize())
                        .set_index("date_norm")["y"]
                        .to_dict()
                    )
                    df_out["reel"] = df_out["date"].dt.normalize().map(real_map)
                else:
                    df_out["reel"] = np.nan

                # ---------- colonne prévision IA ----------
                if current_pred_df is not None and not current_pred_df.empty:
                    pred_map = (
                        current_pred_df.assign(date_norm=pd.to_datetime(current_pred_df["date"]).dt.normalize())
                        .set_index("date_norm")["pred_value"]
                        .to_dict()
                    )
                    df_out["prev_ia"] = df_out["date"].dt.normalize().map(pred_map)
                else:
                    df_out["prev_ia"] = np.nan

                # ---------- colonnes profils cochés ----------
                if ia_profils_names and ia_profils_series and ia_profils_dates and ia_profils_vars:
                    prof_dates = pd.to_datetime(ia_profils_dates).normalize()

                    for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                        if not var.get():
                            continue

                        # conversion float safe
                        vals = [_to_float_or_nan(v) for v in serie]
                        prof_map = pd.Series(vals, index=prof_dates, dtype="float64").to_dict()

                        col_name = f"profil_{name}".replace("\n", " ").strip()
                        df_out[col_name] = df_out["date"].dt.normalize().map(prof_map)

                # ---------- métadonnées utiles ----------
                df_out.insert(0, "filiale", current_filiale_name or selected_filiale.get())
                df_out.insert(1, "annee", current_target_year or (int(annees_var.get()) + 1))

                # ---------- save ----------
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Fichiers Excel", "*.xlsx")],
                    title="Exporter les données du graphe détaillé (Graph2) en Excel"
                )
                if not file_path:
                    return

                df_out.to_excel(file_path, index=False)
                messagebox.showinfo("Export", f"Export Graph2 OK :\n{file_path}")

            except Exception as e:
                messagebox.showerror("Erreur export Graph2", str(e))

        def _show_pyvista_3d():
            import numpy as np
            import pandas as pd

            try:
                import pyvista as pv
            except Exception as e:
                messagebox.showerror(
                    "PyVista",
                    "PyVista n'est pas dispo dans cet environnement.\n"
                    "Installe-le dans le même Python/venv que ton script.\n\n"
                    f"Détail: {e}"
                )
                return

            feats = model_artifacts.get("features")
            cls_model = model_artifacts.get("cls_model")
            reg_model = model_artifacts.get("reg_model")
            X_train = model_artifacts.get("X_train")

            if feats is None or cls_model is None or reg_model is None or X_train is None:
                messagebox.showinfo("3D", "Lance d'abord l'entraînement du modèle.")
                return

            features = list(feats)

            # -----------------------
            # Helpers
            # -----------------------
            def _norm01(x):
                x = np.asarray(x, dtype=float)
                if x.size == 0:
                    return x.astype(np.float32)
                mn = np.nanmin(x)
                mx = np.nanmax(x)
                if not np.isfinite(mn) or not np.isfinite(mx) or abs(mx - mn) < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return ((x - mn) / (mx - mn)).astype(np.float32)

            def _norm_signed(x):
                """Normalize to [-1,1] using max abs."""
                x = np.asarray(x, dtype=float)
                m = np.nanmax(np.abs(x))
                if not np.isfinite(m) or m < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return (x / m).astype(np.float32)

            def build_lines(src_pts, dst_pts, w):
                e = len(src_pts)
                if e == 0:
                    m = pv.PolyData(np.zeros((0, 3), dtype=np.float32))
                    m.lines = np.array([], dtype=np.int64)
                    m.point_data["w"] = np.array([], dtype=np.float32)
                    return m

                pts = np.empty((2 * e, 3), dtype=np.float32)
                pts[0::2] = src_pts
                pts[1::2] = dst_pts

                lines = np.empty((3 * e,), dtype=np.int64)
                lines[0::3] = 2
                lines[1::3] = np.arange(0, 2 * e, 2, dtype=np.int64)
                lines[2::3] = np.arange(1, 2 * e, 2, dtype=np.int64)

                m = pv.PolyData(pts)
                m.lines = lines
                m.point_data["w"] = np.repeat(np.asarray(w, dtype=np.float32), 2)
                return m

            def get_gain_importance(model, features):
                imp = np.zeros(len(features), dtype=np.float32)
                try:
                    booster = getattr(model, "booster_", None)
                    if booster is not None:
                        names = booster.feature_name()
                        gains = booster.feature_importance(importance_type="gain")
                        d = dict(zip(names, gains))
                        for i, f in enumerate(features):
                            imp[i] = float(d.get(f, 0.0))
                    else:
                        gains = getattr(model, "feature_importances_", None)
                        if gains is not None:
                            imp[:] = np.asarray(gains[: len(features)], dtype=np.float32)
                except Exception:
                    pass
                return imp

            gain_cls = get_gain_importance(cls_model, features)
            gain_reg = get_gain_importance(reg_model, features)

            # -----------------------
            # SHAP (best effort)
            # -----------------------
            shap_ok = False
            shap_cls = np.zeros(len(features), dtype=np.float32)
            shap_reg = np.zeros(len(features), dtype=np.float32)

            def try_compute_shap():
                nonlocal shap_ok, shap_cls, shap_reg
                try:
                    import shap
                except Exception:
                    shap_ok = False
                    return

                try:
                    Xs = X_train[features].copy()
                except Exception:
                    Xs = pd.DataFrame(X_train, columns=features).copy()

                n_sample = min(800, len(Xs))
                if n_sample < 80:
                    shap_ok = False
                    return
                Xs = Xs.sample(n=n_sample, random_state=42)

                ok_any = False
                # CLS
                try:
                    expl = shap.TreeExplainer(cls_model)
                    sv = expl.shap_values(Xs)
                    if isinstance(sv, list) and len(sv) >= 2:
                        sv = sv[1]
                    shap_cls[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_cls)) > 0)
                except Exception:
                    shap_cls[:] = 0.0

                # REG
                try:
                    expl = shap.TreeExplainer(reg_model)
                    sv = expl.shap_values(Xs)
                    shap_reg[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_reg)) > 0)
                except Exception:
                    shap_reg[:] = 0.0

                shap_ok = bool(ok_any)

            try_compute_shap()

            # -----------------------
            # Scene config
            # -----------------------
            TOPK = 18          # nb features affichées par arc (CLS + REG)
            LABEL_TOP = 10     # nb labels par arc
            NODE_R = 0.20

            pv.set_plot_theme("dark")
            plotter = pv.Plotter(window_size=(1600, 900))
            plotter.set_background("#0b1220")

            # pipeline positions
            P_OUT   = np.array([-9.0, 0.0, -2.6], dtype=np.float32)
            P_RULES = np.array([-6.0, 0.0, -1.0], dtype=np.float32)
            P_COMB  = np.array([-3.0, 0.0,  0.4], dtype=np.float32)
            P_CLS   = np.array([ 0.0, -1.2,  1.8], dtype=np.float32)
            P_REG   = np.array([ 0.0, +1.2,  1.8], dtype=np.float32)

            def add_block(center, size=(2.6, 1.0, 0.9), label=""):
                box = pv.Cube(center=center, x_length=size[0], y_length=size[1], z_length=size[2])
                box.point_data["v"] = np.full(box.n_points, 1.0, dtype=np.float32)
                plotter.add_mesh(box, scalars="v", cmap="turbo", opacity=0.90, show_edges=True)
                plotter.add_point_labels([center], [label], font_size=14, text_color="white",
                                        point_color="white", shape=False)

            def add_arrow(p0, p1):
                arr = pv.Arrow(start=p0, direction=(p1 - p0),
                            tip_length=0.25, tip_radius=0.07, shaft_radius=0.03, scale="auto")
                plotter.add_mesh(arr, color="white", opacity=0.9)

            add_block(P_CLS, label="LGBM CLS\nP(y>0)")
            add_block(P_REG, label="LGBM REG\nE[y|y>0)")
            add_block(P_COMB, label="Combine\np*amt")
            add_block(P_RULES, label="Rules\nWE/Férié=0\ncarry")
            add_block(P_OUT, label="Output\npred/p10/p90")

            add_arrow(P_CLS, P_COMB)
            add_arrow(P_REG, P_COMB)
            add_arrow(P_COMB, P_RULES)
            add_arrow(P_RULES, P_OUT)

            # -----------------------
            # Arcs placement (2 arcs)
            # -----------------------
            def arc_points(n, z_shift=0.0, y_shift=0.0):
                ang = np.linspace(-1.05, 1.05, n)
                r = 6.2
                x = 6.6 + 0.28 * np.cos(ang)
                y = (r * np.cos(ang) * 0.18) + y_shift
                z = (r * np.sin(ang) * 0.55) + 2.2 + z_shift
                return np.stack([x, y, z], axis=1).astype(np.float32)

            # -----------------------
            # State
            # -----------------------
            state = {
                "edge_mode": 3,          # 1=CLS 2=REG 3=BOTH
                "score_mode": ("shap" if shap_ok else "gain"),  # gain/shap
                "diff_mode": False,      # False: turbo on v ; True: diverging on diff
            }

            a_cls_nodes = None
            a_reg_nodes = None
            a_cls_edges = None
            a_reg_edges = None
            label_actors = []

            def clear_labels():
                nonlocal label_actors
                for act in label_actors:
                    try:
                        plotter.remove_actor(act)
                    except Exception:
                        pass
                label_actors = []

            def scores():
                if state["score_mode"] == "shap" and shap_ok:
                    s_cls = shap_cls.copy()
                    s_reg = shap_reg.copy()
                else:
                    s_cls = gain_cls.copy()
                    s_reg = gain_reg.copy()
                return s_cls, s_reg

            def hud_text():
                s = f"Edges 1/2/3=CLS/REG/BOTH | Score G/S=GAIN/SHAP | D=DIFF"
                s += f"  ||  score={state['score_mode'].upper()}  diff={'ON' if state['diff_mode'] else 'OFF'}"
                if state["score_mode"] == "shap" and not shap_ok:
                    s += " (SHAP indispo -> GAIN)"
                return s

            plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            # -----------------------
            # Rebuild
            # -----------------------
            def rebuild():
                nonlocal a_cls_nodes, a_reg_nodes, a_cls_edges, a_reg_edges

                s_cls, s_reg = scores()

                # TOPK CLS et TOPK REG (union)
                idx_cls = np.argsort(-s_cls)[: min(TOPK, len(features))]
                idx_reg = np.argsort(-s_reg)[: min(TOPK, len(features))]

                f_cls = [features[i] for i in idx_cls]
                f_reg = [features[i] for i in idx_reg]

                v_cls = s_cls[idx_cls]
                v_reg = s_reg[idx_reg]

                # node colors:
                # - normal : importance normalisée
                # - diff : (reg - cls) normalisé signé
                if state["diff_mode"]:
                    # pour diff, on calcule diff sur l'ensemble des features de chaque arc
                    d_cls = (s_reg[idx_cls] - s_cls[idx_cls])  # sur les features TOP CLS
                    d_reg = (s_reg[idx_reg] - s_cls[idx_reg])  # sur les features TOP REG
                    c_cls = _norm_signed(d_cls)  # [-1,1]
                    c_reg = _norm_signed(d_reg)  # [-1,1]
                else:
                    c_cls = _norm01(v_cls)       # [0,1]
                    c_reg = _norm01(v_reg)       # [0,1]

                # edges weights (normalisés séparément)
                w_cls = _norm01(v_cls)
                w_reg = _norm01(v_reg)

                pts_cls = arc_points(len(f_cls), z_shift=-0.9, y_shift=-0.25)
                pts_reg = arc_points(len(f_reg), z_shift=+0.9, y_shift=+0.25)

                # nodes glyphs
                sphere = pv.Sphere(radius=NODE_R, theta_resolution=16, phi_resolution=16)

                cloud_cls = pv.PolyData(pts_cls)
                cloud_reg = pv.PolyData(pts_reg)

                # scalar name
                scalar_name = "d" if state["diff_mode"] else "v"
                cloud_cls[scalar_name] = c_cls.astype(np.float32)
                cloud_reg[scalar_name] = c_reg.astype(np.float32)

                glyph_cls = cloud_cls.glyph(scale=False, geom=sphere, orient=False)
                glyph_reg = cloud_reg.glyph(scale=False, geom=sphere, orient=False)

                # edges meshes (feature -> model)
                PCLS = np.repeat(P_CLS[None, :], len(f_cls), axis=0)
                PREG = np.repeat(P_REG[None, :], len(f_reg), axis=0)

                e_cls = build_lines(pts_cls, PCLS, w_cls)
                e_reg = build_lines(pts_reg, PREG, w_reg)

                # create/update actors
                if a_cls_nodes is None:
                    if state["diff_mode"]:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_cls_nodes.mapper.dataset.shallow_copy(glyph_cls)

                if a_reg_nodes is None:
                    if state["diff_mode"]:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_reg_nodes.mapper.dataset.shallow_copy(glyph_reg)

                if a_cls_edges is None:
                    a_cls_edges = plotter.add_mesh(e_cls, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_cls_edges.mapper.dataset.shallow_copy(e_cls)

                if a_reg_edges is None:
                    a_reg_edges = plotter.add_mesh(e_reg, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_reg_edges.mapper.dataset.shallow_copy(e_reg)

                # edge visibility
                if state["edge_mode"] == 1:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(False)
                elif state["edge_mode"] == 2:
                    a_cls_edges.SetVisibility(False)
                    a_reg_edges.SetVisibility(True)
                else:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(True)

                # make edge thickness depend on mean importance (simple)
                # (PyVista can't vary per-line width easily, but you can tune global)
                a_cls_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_cls)))
                a_reg_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_reg)))

                # labels
                clear_labels()
                k1 = min(LABEL_TOP, len(f_cls))
                k2 = min(LABEL_TOP, len(f_reg))
                if k1 > 0:
                    act = plotter.add_point_labels(pts_cls[:k1], [f"CLS: {x}" for x in f_cls[:k1]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)
                if k2 > 0:
                    act = plotter.add_point_labels(pts_reg[:k2], [f"REG: {x}" for x in f_reg[:k2]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)

                plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            rebuild()

            # -----------------------
            # Key events
            # -----------------------
            def set_edge_mode(m):
                state["edge_mode"] = m
                rebuild()

            def set_score_mode(mode):
                if mode == "shap" and not shap_ok:
                    state["score_mode"] = "gain"
                else:
                    state["score_mode"] = mode
                rebuild()

            def toggle_diff():
                state["diff_mode"] = not state["diff_mode"]
                rebuild()

            plotter.add_key_event("1", lambda: set_edge_mode(1))
            plotter.add_key_event("2", lambda: set_edge_mode(2))
            plotter.add_key_event("3", lambda: set_edge_mode(3))
            plotter.add_key_event("g", lambda: set_score_mode("gain"))
            plotter.add_key_event("G", lambda: set_score_mode("gain"))
            plotter.add_key_event("s", lambda: set_score_mode("shap"))
            plotter.add_key_event("S", lambda: set_score_mode("shap"))
            plotter.add_key_event("d", toggle_diff)
            plotter.add_key_event("D", toggle_diff)

            plotter.add_axes()
            plotter.camera_position = "iso"
            plotter.camera.zoom(1.35)
            plotter.show()

        # ============ ZONE RÉSULTATS ============
        graph_widgets = []   # kpi, graphe 1, tableaux
        graph2_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe 2
        monthly_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe mensuel
        monthly_graph_widget = None
        monthly_container_packed = False

        # ---------- bouton export Graph2 ----------

        export_tools_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=0)
        export_tools_frame.pack(fill="x", padx=10, pady=(0, 10))

        export_graph2_button = ctk.CTkButton(
            export_tools_frame,
            text="📤 Export Graph2 (Réel / IA / Profils) - Excel",
            width=320, height=34,
            corner_radius=10,
            fg_color="#2563eb", hover_color="#1d4ed8",
            text_color="white",
            state="disabled",  # 👈 désactivé tant que pas de modèle
            command=_export_graph2_to_excel
        )
        export_graph2_button.pack(anchor="w")

        btn_3d = ctk.CTkButton(
            export_tools_frame,
            text="🧊 Voir 3D (PyVista)",
            width=200, height=34,
            corner_radius=10,
            fg_color="#7c3aed", hover_color="#6d28d9",
            text_color="white",
            state="disabled",
            command=_show_pyvista_3d
        )
        btn_3d.pack(anchor="w", pady=(6, 0))

        # ---------- helpers graphiques / profils ----------

        def _clear_graph_widgets():
            """Supprime les widgets graphiques (sauf le conteneur du graphe 2)."""
            nonlocal graph_widgets
            for w in graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            graph_widgets = []

        def _redraw_graph2():
            """Redessine le graphe détaillé N+1 avec :
            - Réel N+1 (si dispo)
            - Prévision IA N+1
            - Profils cochés N+1
            """
            nonlocal ia_graph2_widget, graph2_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Si rien à tracer et aucun profil, on ne fait rien
            if current_pred_df is None and current_real_target_df is None and not ia_profils_names:
                return

            # S'assurer que le conteneur est packé une fois pour toutes
            if not graph2_container_packed:
                graph2_container.pack(pady=10, fill="both", expand=True)
                graph2_container_packed = True

            # Nettoyer le conteneur
            for child in graph2_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            ia_graph2_widget = None

            # Créer la figure
            fig2, ax2 = plt.subplots(
                figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True
            )
            ax2.set_facecolor("#00122e")

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            # Réel N+1
            if current_real_target_df is not None and not current_real_target_df.empty:
                ax2.plot(
                    current_real_target_df["date"], current_real_target_df["y"],
                    label=f"Réel {target_year}", linewidth=2, color="#5DADE2"
                )

            # Prévision IA N+1
            if current_pred_df is not None and not current_pred_df.empty:
                ax2.plot(
                    current_pred_df["date"], current_pred_df["pred_value"],
                    label=f"Prévision IA {target_year}",
                    linewidth=2, linestyle="--", color="#F4D03F"
                )

            # Profils cochés N+1
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = [pd.to_datetime(d) for d in ia_profils_dates]
                palette = plt.cm.tab10.colors
                prof_idx = 0
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    y_prof = [_to_float_or_nan(v) for v in serie]
                    ax2.plot(
                        dates_prof, y_prof,
                        label=f"Profil '{name}' {target_year}",
                        linewidth=1.8,
                        linestyle="-.",
                        color=palette[prof_idx % len(palette)]
                    )
                    prof_idx += 1

            filiale_for_title = current_filiale_name or selected_filiale.get()
            titre2 = f"Trafic Voyageurs – année {target_year} – {filiale_for_title}"
            ax2.set_title(titre2, color="white", fontsize=14)
            ax2.set_xlabel("Date", color="white", fontsize=12)
            ax2.set_ylabel("Montant (K€)", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white", rotation=30)
            ax2.tick_params(axis='y', colors="white")
            ax2.legend(facecolor="#00122e", edgecolor="white", labelcolor="white")

            canvas_fig2 = FigureCanvasTkAgg(fig2, master=graph2_container)
            canvas_fig2.draw()
            ia_graph2_widget = canvas_fig2.get_tk_widget()
            ia_graph2_widget.pack(fill="both", expand=True)
            plt.close(fig2)
        
        def _redraw_monthly_graph():
            """
            Graphe mensuel N+1 :
            - Barres : Réel N+1 (si dispo) vs Prévision IA N+1
            - Barres supplémentaires : Profils cochés (somme mensuelle)
            """
            nonlocal monthly_graph_widget, monthly_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Il faut au moins la prédiction N+1 pour tracer quelque chose
            if current_pred_df is None or current_pred_df.empty:
                return

            # Pack du conteneur une seule fois
            if not monthly_container_packed:
                monthly_container.pack(pady=10, fill="both", expand=True)
                monthly_container_packed = True

            # Nettoyage du conteneur
            for child in monthly_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            monthly_graph_widget = None

            import numpy as np
            from matplotlib.ticker import FuncFormatter

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            color_real = "#1f77b4"   # bleu
            color_pred = "#F4D03F"   # orange
            # Palette dédiée aux profils (change-les comme tu veux)
            profile_colors = [
                "#e74c3c",  # rouge vif
                "#9b59b6",  # violet
                "#2ecc71",  # vert clair
                "#1abc9c",  # turquoise
                "#f1c40f",  # jaune
                "#d35400",  # orange foncé (≠ #ff7f0e)
                "#8e44ad",  # violet profond
                "#27ae60",  # vert riche
                "#16a085",  # vert/turquoise sombre
                "#c0392b",  # rouge foncé
                "#7f8c8d",  # gris bleuté
                "#95a5a6",  # gris clair
                "#34495e",  # bleu/gris métallique
                "#bdc3c7",  # argent clair
                "#f39c12",  # jaune/orangé
                "#c27ba0",  # rose/mauve mat
                "#76d7c4",  # aqua pastel
                "#7dcea0",  # vert pastel
                "#af7ac5",  # mauve pastel
                "#5dade2",  # bleu pastel clair (≠ réel, beaucoup plus clair)
            ]

            # ---------- Prévision IA N+1 : somme mensuelle ----------
            df_pred = current_pred_df.copy()
            df_pred["month"] = df_pred["date"].dt.month
            pred_monthly = (
                df_pred.groupby("month", as_index=False)["pred_value"]
                       .sum()
            ).rename(columns={"pred_value": "pred_value"})

            # ---------- Réel N+1 : somme mensuelle (si dispo) ----------
            if current_real_target_df is not None and not current_real_target_df.empty:
                df_real = current_real_target_df.copy()
                df_real["month"] = df_real["date"].dt.month
                real_monthly = (
                    df_real.groupby("month", as_index=False)["y"]
                           .sum()
                ).rename(columns={"y": "real_value"})

                monthly_cmp = pd.merge(
                    real_monthly,
                    pred_monthly,
                    on="month",
                    how="outer"
                ).fillna(0.0)
            else:
                monthly_cmp = pred_monthly.copy()
                monthly_cmp["real_value"] = 0.0

            monthly_cmp = monthly_cmp.sort_values("month")

            # ---------- Profils cochés : agrégés par mois (pour barres) ----------
            # ---------- Profils cochés : calcul "Prévision sinon Réel" + masquage ----------
            active_profiles = []

            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = pd.to_datetime(ia_profils_dates)

                # Base réelle jour par jour
                df_day = pd.DataFrame({
                    "date": dates_prof,
                    "month": dates_prof.month,
                })

                # Ajouter le réel si dispo
                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = dict(zip(current_real_target_df["date"], current_real_target_df["y"]))
                    df_day["real"] = df_day["date"].map(real_map).fillna(0.0)
                else:
                    df_day["real"] = 0.0

                # Calcul par profil
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    vals = [_to_float_or_nan(v) for v in serie]

                    df_day["prev"] = vals

                    # Prévision sinon réel
                    df_day["comb"] = df_day["prev"]
                    df_day.loc[df_day["comb"].isna(), "comb"] = df_day["real"]

                    # Mois avec prévision
                    mois_avec_prev = set(df_day.loc[df_day["prev"].notna(), "month"].unique())
                    mois_tous = set(monthly_cmp["month"].unique())
                    mois_masques = sorted(mois_tous - mois_avec_prev)

                    # Masquage complet des mois sans prévision
                    df_day["comb_masked"] = df_day["comb"]
                    df_day.loc[df_day["month"].isin(mois_masques), "comb_masked"] = np.nan

                    # Agrégation mensuelle
                    prof_month = df_day.groupby("month")["comb_masked"].sum(min_count=1)

                    # Alignement
                    yvals = [prof_month.get(m, np.nan) for m in monthly_cmp["month"]]

                    active_profiles.append((name, yvals))


            # ---------- Placement des barres ----------
            x = np.arange(len(monthly_cmp))

            # nb séries : Réel + Prévision + profils actifs
            nb_series = 2 + len(active_profiles)
            # largeur de chaque barre pour que tout tienne dans le mois
            width = 0.8 / max(nb_series, 1)
            offsets = (np.arange(nb_series) - (nb_series - 1) / 2.0) * width

            fig_m, ax_m = plt.subplots(figsize=(11, 4.5),
                                       facecolor="#00122e",
                                       constrained_layout=True)
            ax_m.set_facecolor("#00122e")

            # Barres Réel N+1
            ax_m.bar(
                x + offsets[0],
                monthly_cmp["real_value"],
                width,
                label=f"Réel {target_year}",
                color=color_real
            )

            # Barres Prévision IA N+1
            ax_m.bar(
                x + offsets[1],
                monthly_cmp["pred_value"],
                width,
                label=f"Prévision IA {target_year}",
                color=color_pred
            )

            # Barres Profils cochés
            if active_profiles:
                for i, (name, y_vals) in enumerate(active_profiles):
                    serie_idx = 2 + i
                    profile_color = profile_colors[i % len(profile_colors)]
                    ax_m.bar(
                        x + offsets[serie_idx],
                        y_vals,
                        width,
                        label=f"Profil '{name}' {target_year}",
                        color=profile_color
                    )


            # ---------- X / Y / style ----------
            month_nums = monthly_cmp["month"].astype(int)
            month_dates = pd.to_datetime(
                {"year": [target_year] * len(month_nums),
                 "month": month_nums,
                 "day": 1}
            )
            mois_labels = month_dates.dt.strftime("%Y-%m")

            ax_m.set_xticks(x)
            ax_m.set_xticklabels(mois_labels, rotation=45, ha="right",
                                 fontsize=9, color="white")

            ax_m.set_xlabel("Mois", fontsize=11, color="white")
            ax_m.set_ylabel("Valeur cumulée (K€)", fontsize=11, color="white")
            ax_m.tick_params(axis="y", colors="white")

            def _fmt_milliers(val, pos):
                try:
                    return f"{int(val):,}".replace(",", " ")
                except Exception:
                    return ""
            ax_m.yaxis.set_major_formatter(FuncFormatter(_fmt_milliers))

            filiale_for_title = current_filiale_name or selected_filiale.get()
            ax_m.set_title(
                f"{filiale_for_title} - Trafic Voyageurs - {target_year}",
                fontsize=13,
                fontweight="bold",
                color="white",
                pad=12
            )

            # pas de bordures
            fig_m.patch.set_edgecolor("#00122e")
            fig_m.patch.set_linewidth(0)
            for spine in ax_m.spines.values():
                spine.set_visible(False)

            # légende sans cadre
            leg_m = ax_m.legend(frameon=False, facecolor="#00122e")
            for text in leg_m.get_texts():
                text.set_color("white")

            canvas_m = FigureCanvasTkAgg(fig_m, master=monthly_container)
            canvas_m.draw()
            monthly_graph_widget = canvas_m.get_tk_widget()
            monthly_graph_widget.pack(fill="both", expand=True)
            plt.close(fig_m)

        # ---------- analyse des écarts (IA vs Réel, Profil vs Réel) ----------

        def _compute_metrics(
            label,
            ref_dates, ref_values,
            cmp_dates, cmp_values,
            # --- métier / robustesse ---
            min_abs_ref=10.0,          # filtre sur le RÉEL : focus jours "significatifs" (ex: >10 K€)
            rel_seuil=40.0,            # seuil d'alerte (%)
            rel_mode="smape",          # "smape" (reco) | "mape_ref" | "mape_cmp"
            eps=1e-9,                  # anti-division par zéro
        ):
            """
            Analyse quant + compréhensible métier:
            - aligne les dates (normalize)
            - convertit proprement en float
            - filtre optionnel sur |Réel| >= min_abs_ref (focus business)
            - calcule plusieurs métriques:
                * MAE, RMSE (K€)
                * Biais (moyenne(cmp-ref)) (K€)
                * sMAPE (%) (robuste)
                * WMAPE (%) (pondéré par volume réel)
                * taux d’alertes > rel_seuil (%) + valorisation des écarts (K€)
            - renvoie toujours un dict si au moins une date commune existe.
            """

            if ref_dates is None or cmp_dates is None:
                print(f"[METRICS] {label} -> ref_dates ou cmp_dates None")
                return None

            # --- conversions robustes ---
            ref_vals = [_to_float_or_nan(v) for v in ref_values]
            cmp_vals = [_to_float_or_nan(v) for v in cmp_values]

            ref_idx = pd.to_datetime(ref_dates).dt.normalize()
            cmp_idx = pd.to_datetime(cmp_dates).dt.normalize()


            ref = pd.Series(ref_vals, index=ref_idx, dtype="float64")
            cmp_ = pd.Series(cmp_vals, index=cmp_idx, dtype="float64")

            # --- intersection dates ---
            common_idx = ref.index.intersection(cmp_.index)
            print(f"[METRICS] {label} -> nb dates communes = {len(common_idx)}")
            if len(common_idx) == 0:
                return None

            df_m = pd.DataFrame({
                "ref": ref.loc[common_idx],
                "cmp": cmp_.loc[common_idx],
            }).dropna()

            # Si rien exploitable
            if df_m.empty:
                return {
                    "Comparaison": label,
                    "Nb points (communs)": int(len(common_idx)),
                    "Nb points (utilisés)": 0,
                    "MAE (K€)": 0.0,
                    "RMSE (K€)": 0.0,
                    "Biais (K€)": 0.0,
                    "sMAPE (%)": 0.0,
                    "WMAPE (%)": 0.0,
                    f"Nb alertes (> {rel_seuil}%)": 0,
                    f"Taux alertes (> {rel_seuil}%)": 0.0,
                    "Valeur totale écarts alertes (K€)": 0.0,
                }

            # --- filtre métier sur le RÉEL ---
            if min_abs_ref is not None and min_abs_ref > 0:
                before = len(df_m)
                df_m = df_m[df_m["ref"].abs() >= float(min_abs_ref)]
                print(f"[METRICS] {label} -> filtre |ref| >= {min_abs_ref} : {before} → {len(df_m)} points")

            if df_m.empty:
                return {
                    "Comparaison": label,
                    "Nb points (communs)": int(len(common_idx)),
                    "Nb points (utilisés)": 0,
                    "MAE (K€)": 0.0,
                    "RMSE (K€)": 0.0,
                    "Biais (K€)": 0.0,
                    "sMAPE (%)": 0.0,
                    "WMAPE (%)": 0.0,
                    f"Nb alertes (> {rel_seuil}%)": 0,
                    f"Taux alertes (> {rel_seuil}%)": 0.0,
                    "Valeur totale écarts alertes (K€)": 0.0,
                }

            # --- erreurs de base ---
            err = (df_m["cmp"] - df_m["ref"])
            abs_err = err.abs()

            mae = float(abs_err.mean())
            rmse = float(np.sqrt(np.mean(err.values ** 2)))
            bias = float(err.mean())  # >0 => sur-prédit, <0 => sous-prédit

            # --- métrique relative (robuste) ---
            if rel_mode == "mape_ref":
                denom = df_m["ref"].abs().clip(lower=eps)
                rel = (abs_err / denom) * 100.0
            elif rel_mode == "mape_cmp":
                denom = df_m["cmp"].abs().clip(lower=eps)
                rel = (abs_err / denom) * 100.0
            else:
                # sMAPE : 2|err|/(|ref|+|cmp|)
                denom = (df_m["ref"].abs() + df_m["cmp"].abs()).clip(lower=eps)
                rel = (2.0 * abs_err / denom) * 100.0

            # cohérence: rel est alignée sur df_m (pas de drop séparé)
            rel = rel.replace([np.inf, -np.inf], np.nan)
            df_m = df_m.assign(rel=rel).dropna(subset=["rel"])

            if df_m.empty:
                return {
                    "Comparaison": label,
                    "Nb points (communs)": int(len(common_idx)),
                    "Nb points (utilisés)": 0,
                    "MAE (K€)": 0.0,
                    "RMSE (K€)": 0.0,
                    "Biais (K€)": 0.0,
                    "sMAPE (%)": 0.0,
                    "WMAPE (%)": 0.0,
                    f"Nb alertes (> {rel_seuil}%)": 0,
                    f"Taux alertes (> {rel_seuil}%)": 0.0,
                    "Valeur totale écarts alertes (K€)": 0.0,
                }

            # --- alertes "métier" ---
            mask = df_m["rel"] > float(rel_seuil)
            nb_points = int(len(df_m))
            nb_alert = int(mask.sum())
            taux_alert = 0.0 if nb_points == 0 else 100.0 * nb_alert / nb_points

            # valo des alertes = somme des |écarts| sur les jours alertes
            valo_alert = float((df_m.loc[mask, "cmp"] - df_m.loc[mask, "ref"]).abs().sum())

            # --- WMAPE (pondéré par volume réel) ---
            # = sum(|err|) / sum(|ref|) ; très “métier” car pondère les gros jours
            sum_ref = float(df_m["ref"].abs().sum())
            wmape = float((df_m["cmp"] - df_m["ref"]).abs().sum() / (sum_ref + eps) * 100.0)

            # sMAPE moyen (plus stable que MAPE)
            smape = float(df_m["rel"].mean()) if rel_mode == "smape" else float((2.0 * (df_m["cmp"]-df_m["ref"]).abs() / (df_m["ref"].abs()+df_m["cmp"].abs()).clip(lower=eps) * 100.0).mean())

            print(f"[METRICS] {label} -> nb_points={nb_points}, nb_alert={nb_alert}, taux={taux_alert:.1f}%, MAE={mae:.1f}, WMAPE={wmape:.1f}%")

            return {
                "Comparaison": label,
                "Nb points (communs)": int(len(common_idx)),
                "Nb points (utilisés)": nb_points,
                "MAE (K€)": round(mae, 1),
                "RMSE (K€)": round(rmse, 1),
                "Biais (K€)": round(bias, 1),
                "sMAPE (%)": round(smape, 1),
                "WMAPE (%)": round(wmape, 1),
                f"Nb alertes (> {rel_seuil}%)": nb_alert,
                f"Taux alertes (> {rel_seuil}%)": round(taux_alert, 1),
                "Valeur totale écarts alertes (K€)": round(valo_alert, 1),
            }


        def _rebuild_analysis_table():
            nonlocal analysis_table_frame, analysis_tree, analysis_export_button
            nonlocal current_pred_df, current_real_target_df, current_target_year

            if analysis_table_frame is None:
                return
            for child in analysis_table_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            analysis_tree = None
            analysis_export_button = None

            if current_real_target_df is None or current_real_target_df.empty:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune donnée réelle pour l'année N+1 : analyse des écarts impossible.",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                return

            rows = []

            # Paramètres d'analyse (tu peux les exposer UI plus tard)
            rel_seuil = 40.0
            min_abs_ref = 10.0  # focus business
            rel_mode = "smape"  # recommandé

            # IA vs Réel
            if current_pred_df is not None and not current_pred_df.empty:
                m = _compute_metrics(
                    label="Prévision IA vs Réel",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    min_abs_ref=min_abs_ref,
                    rel_seuil=rel_seuil,
                    rel_mode=rel_mode,
                )
                if m is not None:
                    rows.append(m)

            # Profils cochés vs Réel
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    m = _compute_metrics(
                        label=f"Profil '{name}' vs Réel",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=ia_profils_dates,
                        cmp_values=serie,
                        min_abs_ref=min_abs_ref,
                        rel_seuil=rel_seuil,
                        rel_mode=rel_mode,
                    )
                    if m is not None:
                        rows.append(m)

            ctk.CTkLabel(
                analysis_table_frame,
                text=(
                    f"Analyse quant (année {current_target_year or (int(annees_var.get())+1)}) "
                    f"— focus jours |Réel|≥{min_abs_ref} K€ — alertes si erreur relative > {rel_seuil}% (sMAPE)"
                ),
                font=("Segoe UI", 13, "bold"),
                text_color="white"
            ).pack(anchor="w", pady=(0, 6))

            if not rows:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune comparaison disponible.",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                return

            cols = list(rows[0].keys())

            tree = ttk.Treeview(
                analysis_table_frame, columns=cols, show="headings",
                height=len(rows)
            )
            for col in cols:
                tree.heading(col, text=col)

            # widths (ajuste à ton goût)
            widths = {
                "Comparaison": 260,
                "Nb points (communs)": 140,
                "Nb points (utilisés)": 140,
                "MAE (K€)": 110,
                "RMSE (K€)": 110,
                "Biais (K€)": 110,
                "sMAPE (%)": 110,
                "WMAPE (%)": 110,
                f"Nb alertes (> {rel_seuil}%)": 160,
                f"Taux alertes (> {rel_seuil}%)": 170,
                "Valeur totale écarts alertes (K€)": 220,
            }
            for col in cols:
                tree.column(col, anchor="center", width=widths.get(col, 140))
            tree.column("Comparaison", anchor="w", width=widths.get("Comparaison", 260))

            for r in rows:
                tree.insert("", "end", values=[r.get(c, "") for c in cols])

            tree.pack(fill="x", expand=True, pady=(0, 8))
            analysis_tree = tree

            # export Excel
            def _export_analysis_excel():
                if not rows:
                    messagebox.showinfo("Export", "Aucune donnée à exporter.")
                    return
                df_export = pd.DataFrame(rows)
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Fichiers Excel", "*.xlsx")],
                        title="Exporter le tableau d'analyse quant en Excel"
                    )
                    if not file_path:
                        return
                    df_export.to_excel(file_path, index=False)
                    messagebox.showinfo("Export", f"Tableau d'analyse exporté vers :\n{file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur export", str(e))

            analysis_export_button = ctk.CTkButton(
                analysis_table_frame,
                text="📤 Exporter analyse quant (Excel)",
                width=260, height=32,
                corner_radius=10,
                fg_color="#2563eb", hover_color="#1d4ed8",
                text_color="white",
                command=_export_analysis_excel
            )
            analysis_export_button.pack(anchor="w", pady=(6, 4))


        # ---------- PROFILS UI ----------

        def _build_ia_profils_ui(filiale, base_year):
            """
            Construit les cases à cocher de profils pour l'année N+1 de la filiale sélectionnée.
            Dès qu'on coche/décoche, _redraw_graph2() et _rebuild_analysis_table() sont appelés.
            """
            nonlocal ia_profils_vars, ia_profils_names, ia_profils_dates, ia_profils_series

            # Nettoyage du frame
            for w in ia_profils_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            ia_profils_vars = []
            ia_profils_names = []
            ia_profils_dates = []
            ia_profils_series = []

            if not filiale:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Aucune filiale sélectionnée.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                base_year_int = int(base_year)
            except Exception:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Année N invalide.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            target_year = base_year_int + 1
            feuille = sections.get(filiale)
            if not feuille:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune feuille trouvée pour {filiale}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees (profils N+1) pour {feuille} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Erreur de chargement des données N+1.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Flux 'Trafic Voyageurs' introuvable pour cette filiale.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            _, col_start = cible[0]

            try:
                dates_p, reel_p, previsions_p, noms_profils_p = extraire_valeurs(
                    ws, col_start, nb_prev, annee=target_year
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs (profils N+1) {filiale}/{flux_cible}/{target_year} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée de profils pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            if not dates_p:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            # filtrer les profils actifs (au moins une valeur non nulle)
            actifs = []
            for serie in previsions_p:
                exist = any(v not in (None, 0, 0.0, "") for v in serie)
                actifs.append(exist)

            noms_actifs = [n for n, ok in zip(noms_profils_p, actifs) if ok]
            series_actives = [s for s, ok in zip(previsions_p, actifs) if ok]

            if not noms_actifs:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucun profil actif pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            ia_profils_dates = dates_p
            ia_profils_names = noms_actifs
            ia_profils_series = series_actives

            # Label d'info
            lbl = tk.Label(
                ia_profils_frame,
                text=f"Profils N+1 ({target_year}) : coche pour afficher dans le graphe détaillé et l'analyse",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
            )
            lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 6))

            nb_lignes = 4
            for i, name in enumerate(noms_actifs):
                var = tk.BooleanVar(value=False)

                def _on_toggle(v=var):
                    _redraw_graph2()
                    _redraw_monthly_graph()
                    _rebuild_analysis_table()


                cb = tk.Checkbutton(
                    ia_profils_frame, text=name, variable=var,
                    bg="#00122e", fg="white", font=('Segoe UI', 10),
                    selectcolor="#00aced", activebackground="#003366",
                    activeforeground="white",
                    command=_on_toggle
                )

                row = 1 + (i % nb_lignes)   # +1 car première ligne = label
                col = i // nb_lignes

                cb.grid(row=row, column=col, sticky="w", padx=12, pady=4)
                cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
                cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

                ia_profils_vars.append(var)

        def _on_filiale_or_year_change(_event=None):
            """Quand filiale ou année N change, on recharge les profils N+1."""
            filiale = selected_filiale.get()
            base_year = annees_var.get()
            if not filiale or not base_year:
                return
            _build_ia_profils_ui(filiale, base_year)

        filiale_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        annees_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        _on_filiale_or_year_change()

        # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------
        def _train_model():
            """
            Train SA_VOYAGEURS
            - Hurdle: LGBMClassifier (y>0) + LGBMRegressor Tweedie (y)
            - Features YoY + calendrier + événement "22" + "22 effectif ouvré"
            - Règles métier forecast: WE/Fériés=0 + report
            - Calibration: Conformal P10/P90 par bins de niveau (plus robuste que résidu global)
            """

            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame, export_button

            _clear_graph_widgets()
            print("======================== ENTRAINEMENT DU MODELE (SA_VOYAGEURS – HURDLE + TWEEDIE + Event 22) ========================")

            import numpy as np
            import pandas as pd
            from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
            from sklearn.model_selection import RandomizedSearchCV
            from scipy.stats import randint, uniform

            # ---------------- Filiale fixée ----------------
            filiale = "SA_VOYAGEURS"
            current_filiale_name = filiale
            df_filiale = df[df["section"] == filiale].copy()
            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée trouvée pour la filiale {filiale}.")
                return

            # base year
            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["year"].max())
            print(f"Base year for training {filiale} : {base_year}")

            try:
                # ---------------- Historique ----------------
                if df_filiale["year"].nunique() < 2:
                    messagebox.showinfo("Information", f"Pas assez d'historique pour {filiale} (au moins 2 années).")
                    return

                df_filiale = df_filiale.sort_values("date").copy()

                # ---------------- Jours fériés ----------------
                holiday_sets = {}  # {year: set(datetime.date)}
                if "is_holiday" not in df_filiale.columns:
                    df_filiale["is_holiday"] = 0
                    try:
                        import holidays
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            fr_h = holidays.country_holidays("FR", years=[int(yy)])
                            holiday_sets[int(yy)] = set(fr_h.keys())
                        df_filiale["is_holiday"] = df_filiale["date"].dt.date.apply(
                            lambda d: 1 if d in holiday_sets.get(int(d.year), set()) else 0
                        )
                        print(f"[HOLIDAYS] Calcul jours fériés FR OK (years={years})")
                    except Exception:
                        print("[HOLIDAYS] Librairie 'holidays' non dispo ou erreur -> is_holiday=0 (fallback)")
                        holiday_sets = {}
                else:
                    try:
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            sub = df_filiale[df_filiale["year"] == int(yy)]
                            holiday_sets[int(yy)] = set(sub.loc[sub["is_holiday"] == 1, "date"].dt.date.tolist())
                    except Exception:
                        holiday_sets = {}

                # ---------------- Features calendrier de base ----------------
                df_filiale["dow"] = df_filiale["date"].dt.weekday
                df_filiale["month"] = df_filiale["date"].dt.month
                df_filiale["dom"] = df_filiale["date"].dt.day
                df_filiale["is_eom"] = df_filiale["date"].dt.is_month_end.astype(int)

                # ============================================================
                # OPTION (fortement recommandée si c’est la règle “réelle”):
                # Forcer les jours fermés à 0 AVANT de calculer lags/rolls,
                # pour aligner apprentissage et forecast (report).
                # ============================================================
                # Si tu ne veux pas toucher à l’historique, commente ces 2 lignes.
                is_closed_hist = (df_filiale["dow"] >= 5) | (df_filiale["is_holiday"] == 1)
                df_filiale.loc[is_closed_hist, "y"] = 0.0

                # ---------------- Lags / rolls ----------------
                if "lag_1" not in df_filiale.columns:
                    df_filiale["lag_1"] = df_filiale["y"].shift(1)
                if "lag_7" not in df_filiale.columns:
                    df_filiale["lag_7"] = df_filiale["y"].shift(7)

                if "roll_mean_7" not in df_filiale.columns:
                    df_filiale["roll_mean_7"] = df_filiale["y"].rolling(7, min_periods=1).mean()
                if "roll_mean_30" not in df_filiale.columns:
                    df_filiale["roll_mean_30"] = df_filiale["y"].rolling(30, min_periods=1).mean()

                # ======================================================================
                #   FEATURES “PIC DU 22” + “22 EFFECTIF JOUR OUVRÉ” (vectorisé)
                # ======================================================================
                def _is_open_day(ts: pd.Timestamp, holiday_set: set) -> bool:
                    return (ts.weekday() < 5) and (ts.date() not in holiday_set)

                def _effective_22_date(year: int, month: int, holiday_set: set) -> pd.Timestamp:
                    d = pd.Timestamp(year=year, month=month, day=22)
                    while not _is_open_day(d, holiday_set):
                        d += pd.Timedelta(days=1)
                    return d

                # Table eff22: une ligne par (year, month)
                years_present = sorted(df_filiale["year"].unique().tolist())
                eff22_rows = []
                for yy in years_present:
                    hset = holiday_sets.get(int(yy), set())
                    for mm in range(1, 13):
                        try:
                            d_eff = _effective_22_date(int(yy), int(mm), hset)
                            eff22_rows.append({"year": int(yy), "month": int(mm), "eff22_date": d_eff.normalize()})
                        except Exception:
                            pass
                eff22_df = pd.DataFrame(eff22_rows)

                df_filiale["is_dom_22"] = (df_filiale["dom"] == 22).astype(int)
                df_filiale["days_to_22"] = (df_filiale["dom"] - 22).astype(int)

                df_filiale = df_filiale.merge(eff22_df, on=["year", "month"], how="left")
                df_filiale["date_norm"] = df_filiale["date"].dt.normalize()
                df_filiale["is_22_effective"] = (df_filiale["date_norm"] == df_filiale["eff22_date"]).astype(int)
                df_filiale["days_to_22_effective"] = (df_filiale["date_norm"] - df_filiale["eff22_date"]).dt.days.fillna(0).astype(int)

                # ---------------- Pairs YoY ----------------
                s = df_filiale.copy()
                s_prev = s.copy()
                s_prev["year_target"] = s_prev["year"] + 1

                # Prev-side rename
                s_prev = s_prev.rename(columns={
                    "y": "y_prev_year",
                    "roll_mean_7": "roll_prev_7",
                    "roll_mean_30": "roll_prev_30",
                    "dow": "dow_prev",
                    "lag_1": "lag_1_prev",
                    "lag_7": "lag_7_prev",
                    "month": "month_prev",
                    "dom": "dom_prev",
                    "is_eom": "is_eom_prev",
                    "is_holiday": "is_holiday_prev",
                    "is_dom_22": "is_dom_22_prev",
                    "days_to_22": "days_to_22_prev",
                    "is_22_effective": "is_22_effective_prev",
                    "days_to_22_effective": "days_to_22_effective_prev",
                })

                # Target-side features
                merged = pd.merge(
                    s_prev,
                    s[[
                        "section", "year", "dayofyear",
                        "y",
                        "dow", "month", "dom", "is_eom", "is_holiday",
                        "is_dom_22", "days_to_22", "is_22_effective", "days_to_22_effective"
                    ]],
                    left_on=["section", "year_target", "dayofyear"],
                    right_on=["section", "year", "dayofyear"],
                    how="inner",
                )
                if merged.empty:
                    messagebox.showinfo("Information", f"Aucune paire année→année pour {filiale}.")
                    return

                merged = merged.rename(columns={"y": "y_target"})
                merged = merged.rename(columns={
                    "dow": "dow_tgt",
                    "month": "month_tgt",
                    "dom": "dom_tgt",
                    "is_eom": "is_eom_tgt",
                    "is_holiday": "is_holiday_tgt",
                    "is_dom_22": "is_dom_22_tgt",
                    "days_to_22": "days_to_22_tgt",
                    "is_22_effective": "is_22_effective_tgt",
                    "days_to_22_effective": "days_to_22_effective_tgt",
                })

                df_pairs = merged[[
                    "section", "year_target", "dayofyear",
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev", "is_holiday_prev",
                    "is_dom_22_prev", "days_to_22_prev", "is_22_effective_prev", "days_to_22_effective_prev",
                    "dow_tgt", "month_tgt", "dom_tgt", "is_eom_tgt", "is_holiday_tgt",
                    "is_dom_22_tgt", "days_to_22_tgt", "is_22_effective_tgt", "days_to_22_effective_tgt",
                    "y_target"
                ]].copy()

                df_pairs["section_id"] = 0

                # ---------------- FEATURES ----------------
                features = [
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dayofyear", "section_id",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev", "is_holiday_prev",
                    "is_dom_22_prev", "days_to_22_prev", "is_22_effective_prev", "days_to_22_effective_prev",
                    "dow_tgt", "month_tgt", "dom_tgt", "is_eom_tgt", "is_holiday_tgt",
                    "is_dom_22_tgt", "days_to_22_tgt", "is_22_effective_tgt", "days_to_22_effective_tgt",
                ]

                df_pairs = df_pairs.dropna(subset=features + ["y_target"]).reset_index(drop=True)

                # ---------------- SPLIT plus réaliste: val = dernière année_target dispo (base_year) ----------------
                df_train_pairs = df_pairs[df_pairs["year_target"] <= (base_year - 1)].copy()
                df_valid_pairs = df_pairs[df_pairs["year_target"] == base_year].copy()

                if df_train_pairs.shape[0] < 90 or df_valid_pairs.shape[0] < 60:
                    # fallback si pas assez d'années : split chrono simple
                    print("[WARN] Pas assez d'années pour split 'année'. Fallback split chrono 80/20.")
                    X_all = df_pairs[features].copy()
                    y_all = df_pairs["y_target"].astype(float).copy()
                    y_cls_all = (y_all > 0).astype(int)
                    n_total = len(X_all)
                    cut = int(n_total * 0.8)
                    cut = max(30, min(cut, n_total - 10))
                    X_train = X_all.iloc[:cut].reset_index(drop=True)
                    X_valid = X_all.iloc[cut:].reset_index(drop=True)
                    y_train = y_all.iloc[:cut].reset_index(drop=True)
                    y_valid = y_all.iloc[cut:].reset_index(drop=True)
                    ycls_train = y_cls_all.iloc[:cut].reset_index(drop=True)
                else:
                    X_train = df_train_pairs[features].copy()
                    y_train = df_train_pairs["y_target"].astype(float).copy()
                    ycls_train = (y_train > 0).astype(int)

                    X_valid = df_valid_pairs[features].copy()
                    y_valid = df_valid_pairs["y_target"].astype(float).copy()

                print(f"[SPLIT] train={len(X_train)} | valid={len(X_valid)} | base_year={base_year}")

                # ---------------- Models ----------------
                try:
                    from lightgbm import LGBMClassifier, LGBMRegressor
                except Exception:
                    messagebox.showerror("Erreur", "LightGBM n'est pas disponible. Installe lightgbm.")
                    return

                def _tune(model, param_dist, X_t, y_t, scoring, n_iter=20):
                    # CV simple car on a déjà une vraie validation (année complète)
                    rs = RandomizedSearchCV(
                        model,
                        param_distributions=param_dist,
                        n_iter=n_iter,
                        cv=3,
                        scoring=scoring,
                        random_state=42,
                        n_jobs=-1,
                        verbose=0
                    )
                    rs.fit(X_t, y_t)
                    return rs.best_estimator_, rs.best_params_, rs.best_score_

                # ---- Stage 1: classifier ----
                cls_base = LGBMClassifier(
                    random_state=42,
                    n_estimators=800,
                    learning_rate=0.03,
                    subsample=0.8,
                    colsample_bytree=0.8
                )
                cls_params = {
                    "n_estimators": randint(300, 1400),
                    "num_leaves": randint(15, 127),
                    "max_depth": randint(2, 10),
                    "min_child_samples": randint(10, 120),
                    "learning_rate": uniform(0.01, 0.10),
                    "subsample": uniform(0.6, 0.4),
                    "colsample_bytree": uniform(0.6, 0.4),
                }
                cls_model, cls_bp, cls_bs = _tune(cls_base, cls_params, X_train, ycls_train, scoring="roc_auc", n_iter=20)
                print(f"[TUNING] LGBMClassifier best params: {cls_bp} | CV AUC={cls_bs:.4f}")

                # ---- Stage 2: Tweedie reg (train on y>0) ----
                mask_pos = (y_train > 0).values
                X_train_pos = X_train.loc[mask_pos].reset_index(drop=True)
                y_train_pos = y_train.loc[mask_pos].reset_index(drop=True)
                if len(X_train_pos) < 40:
                    print("[WARN] Peu de jours >0 : régression entraînée sur tout le train.")
                    X_train_pos, y_train_pos = X_train, y_train

                reg_base = LGBMRegressor(
                    random_state=42,
                    objective="tweedie",
                    n_estimators=1200,
                    learning_rate=0.02,
                    subsample=0.8,
                    colsample_bytree=0.8,
                )
                reg_params = {
                    "n_estimators": randint(500, 2000),
                    "num_leaves": randint(15, 255),
                    "max_depth": randint(2, 12),
                    "min_child_samples": randint(10, 140),
                    "learning_rate": uniform(0.01, 0.10),
                    "subsample": uniform(0.6, 0.4),
                    "colsample_bytree": uniform(0.6, 0.4),
                    "reg_lambda": uniform(0.0, 10.0),
                    "tweedie_variance_power": uniform(1.1, 0.8),  # [1.1, 1.9]
                }
                reg_model, reg_bp, reg_bs = _tune(
                    reg_base, reg_params, X_train_pos, y_train_pos,
                    scoring="neg_root_mean_squared_error", n_iter=25
                )
                print(f"[TUNING] LGBM Tweedie best params: {reg_bp} | CV negRMSE={reg_bs:.4f}")

                # ===================== VALIDATION =====================
                p_valid = cls_model.predict_proba(X_valid)[:, 1]
                amt_valid = np.clip(reg_model.predict(X_valid), 0.0, None)
                pred_valid = p_valid * amt_valid

                mae = mean_absolute_error(y_valid, pred_valid)
                rmse = mean_squared_error(y_valid, pred_valid) ** 0.5
                r2 = r2_score(y_valid, pred_valid)

                print(f"[MODEL – HURDLE+TWEEDIE+22] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")

                # ===================== CALIBRATION améliorée: conformal par bins =====================
                # On calcule des quantiles de résidus par tranche de niveau prédit
                res = (y_valid.values - pred_valid)
                # 5 bins de niveau (quantiles)
                qs = np.quantile(pred_valid, [0.0, 0.2, 0.4, 0.6, 0.8, 1.0])
                bins = np.digitize(pred_valid, qs[1:-1], right=True)  # 0..4

                q_low_by_bin = np.zeros(5, dtype=float)
                q_high_by_bin = np.zeros(5, dtype=float)
                for b in range(5):
                    rr = res[bins == b]
                    if len(rr) < 20:
                        # fallback global si bin trop vide
                        q_low_by_bin[b] = float(np.quantile(res, 0.10))
                        q_high_by_bin[b] = float(np.quantile(res, 0.90))
                    else:
                        q_low_by_bin[b] = float(np.quantile(rr, 0.10))
                        q_high_by_bin[b] = float(np.quantile(rr, 0.90))

                pred_p10_valid = np.clip(pred_valid + q_low_by_bin[bins], 0.0, None)
                pred_p90_valid = np.clip(pred_valid + q_high_by_bin[bins], 0.0, None)
                coverage = float(np.mean((y_valid.values >= pred_p10_valid) & (y_valid.values <= pred_p90_valid)))
                width = float(np.mean(pred_p90_valid - pred_p10_valid))

                print(f"[CALIBRATION bins] Coverage(P10-P90)={coverage:.3f}, Width={width:.1f}")
                print(f"[CALIBRATION bins] q_low={q_low_by_bin.round(1).tolist()} | q_high={q_high_by_bin.round(1).tolist()}")

                # ===================== KPI UI =====================
                kpi_frame = ctk.CTkFrame(scrollable_frame, fg_color="#0f1b31", corner_radius=18)
                kpi_frame.pack(fill="x", padx=10, pady=(0, 10))
                graph_widgets.append(kpi_frame)
                for i in range(7):
                    kpi_frame.grid_columnconfigure(i, weight=1)

                def _kpi(parent, title, value, subtitle, col):
                    card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16, border_width=1, border_color="#223658")
                    card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
                    ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                        .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
                    ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 22, "bold"), text_color="white") \
                        .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
                    ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11), text_color="#7ea2d8") \
                        .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

                _kpi(kpi_frame, "MAE", f"{mae:.1f}", "Erreur moyenne", 0)
                _kpi(kpi_frame, "RMSE", f"{rmse:.1f}", "Risque pics", 1)
                _kpi(kpi_frame, "R²", f"{r2:.3f}", "Pouvoir explicatif", 2)
                _kpi(kpi_frame, "Coverage", f"{coverage:.1%}", "Réel ∈ [P10,P90]", 3)
                _kpi(kpi_frame, "Width", f"{width:.0f}", "Largeur bande", 4)
                _kpi(kpi_frame, "Bins", "5", "Calibration niveau", 5)
                _kpi(kpi_frame, "Event", "22 du mois", "Encodé", 6)

                # ========================================================================================
                #                             PRÉVISION N+1 (WE & FERIES = 0 + report)
                # ========================================================================================
                target_year = base_year + 1
                current_target_year = target_year

                df_real_target = df_filiale[df_filiale["year"] == target_year].copy().sort_values("date")
                current_real_target_df = df_real_target

                df_prev_year = df_filiale[df_filiale["year"] == base_year].copy().sort_values("date")

                target_holidays = holiday_sets.get(int(target_year), set())

                # Pré-calc 22 effectifs année cible
                eff22_target = {}
                for mm in range(1, 13):
                    try:
                        eff22_target[int(mm)] = _effective_22_date(int(target_year), int(mm), target_holidays).normalize()
                    except Exception:
                        pass

                def _calib_offsets(pred_value: float):
                    # retrouve le bin selon qs (issus de la validation)
                    b = int(np.digitize([pred_value], qs[1:-1], right=True)[0])
                    return q_low_by_bin[b], q_high_by_bin[b]

                future_rows = []
                carryover = 0.0

                for _, row in df_prev_year.iterrows():
                    d_prev = row["date"]
                    d_next = d_prev + pd.DateOffset(years=1)
                    if int(d_next.year) != int(target_year):
                        continue

                    dow_next = int(d_next.weekday())
                    is_h_next = 1 if (d_next.date() in target_holidays) else 0

                    dom_tgt = int(d_next.day)
                    month_tgt = int(d_next.month)

                    is_dom_22_tgt = 1 if dom_tgt == 22 else 0
                    days_to_22_tgt = int(dom_tgt - 22)

                    d_eff_tgt = eff22_target.get(month_tgt, None)
                    is_22_eff_tgt = 1 if (d_eff_tgt is not None and d_next.normalize() == d_eff_tgt) else 0
                    days_to_22_eff_tgt = int((d_next.normalize() - d_eff_tgt).days) if d_eff_tgt is not None else 0

                    feat_df = pd.DataFrame([[
                        float(row["y"]),
                        float(row["roll_mean_7"]),
                        float(row["roll_mean_30"]),
                        int(row["dayofyear"]),
                        0,
                        int(row["dow"]),
                        float(row["lag_1"]) if pd.notna(row["lag_1"]) else 0.0,
                        float(row["lag_7"]) if pd.notna(row["lag_7"]) else 0.0,
                        int(row["month"]),
                        int(row["dom"]),
                        int(row["is_eom"]),
                        int(row["is_holiday"]),
                        int(row.get("is_dom_22", 1 if int(row["dom"]) == 22 else 0)),
                        int(row.get("days_to_22", int(row["dom"]) - 22)),
                        int(row.get("is_22_effective", 0)),
                        int(row.get("days_to_22_effective", 0)),
                        dow_next,
                        month_tgt,
                        dom_tgt,
                        int(d_next.is_month_end),
                        is_h_next,
                        is_dom_22_tgt,
                        days_to_22_tgt,
                        is_22_eff_tgt,
                        days_to_22_eff_tgt
                    ]], columns=features)

                    p = float(cls_model.predict_proba(feat_df)[:, 1][0])
                    amt = float(reg_model.predict(feat_df)[0])
                    amt = max(0.0, amt)
                    pred_raw = p * amt

                    is_closed = (dow_next >= 5) or (is_h_next == 1)
                    if is_closed:
                        carryover += pred_raw
                        pred_final = 0.0
                    else:
                        pred_final = pred_raw + carryover
                        carryover = 0.0

                    ql, qh = _calib_offsets(pred_final)
                    pred_p10 = max(0.0, pred_final + float(ql))
                    pred_p90 = max(0.0, pred_final + float(qh))

                    future_rows.append({
                        "section": filiale,
                        "date": d_next,
                        "year": int(d_next.year),
                        "month": int(d_next.month),
                        "dayofyear": int(d_next.dayofyear),
                        "pred_p10": float(pred_p10),
                        "pred_value": float(pred_final),
                        "pred_p90": float(pred_p90)
                    })

                df_future_all = pd.DataFrame(future_rows).sort_values("date").reset_index(drop=True)

                # Align dates réelles N+1
                if not df_real_target.empty:
                    df_future_all = pd.merge(
                        df_real_target[["date"]].copy(),
                        df_future_all,
                        on="date",
                        how="left"
                    )
                    for c in ["pred_p10", "pred_value", "pred_p90"]:
                        df_future_all[c] = df_future_all[c].fillna(0.0)

                    df_future_all["section"] = filiale
                    df_future_all["year"] = df_future_all["date"].dt.year
                    df_future_all["month"] = df_future_all["date"].dt.month
                    df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

                # ========================================================================================
                #                                    GRAPHE 1 + bande
                # ========================================================================================
                fig1, ax1 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")

                df_hist_plot = df_filiale[df_filiale["year"] <= base_year]
                ax1.plot(df_hist_plot["date"], df_hist_plot["y"], label=f"Réel (≤ {base_year})", linewidth=2)

                ax1.plot(df_future_all["date"], df_future_all["pred_value"], label=f"Prévision {target_year}", linewidth=2, linestyle="--")
                ax1.fill_between(df_future_all["date"], df_future_all["pred_p10"], df_future_all["pred_p90"], alpha=0.25, label="Bande calibrée (bins) P10–P90")

                ax1.set_title("SA_VOYAGEURS – Prévision N+1 – Hurdle + Tweedie + Event 22 (WE/Fériés=0)", color="white")
                ax1.tick_params(axis='x', colors="white", rotation=30)
                ax1.tick_params(axis='y', colors="white")
                for spine in ax1.spines.values():
                    spine.set_color("#00122e")

                leg1 = ax1.legend(facecolor="#00122e", edgecolor="white")
                for text in leg1.get_texts():
                    text.set_color("white")

                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                w_fig1 = canvas_fig1.get_tk_widget()
                w_fig1.pack(pady=10, fill="both", expand=True)
                graph_widgets.append(w_fig1)
                plt.close(fig1)

                # ========================================================================================
                #                   MISE À JOUR GLOBAL + ANALYSE
                # ========================================================================================
                current_pred_df = df_future_all
                exported_pred_df = df_future_all.copy()
                model_artifacts["X_train"] = X_train.copy()
                model_artifacts["y_train"] = y_train.copy()
                model_artifacts["features"] = features[:]
                model_artifacts["cls_model"] = cls_model
                model_artifacts["reg_model"] = reg_model
                model_artifacts["calib_qs"] = qs
                model_artifacts["calib_q_low_by_bin"] = q_low_by_bin
                model_artifacts["calib_q_high_by_bin"] = q_high_by_bin

                try:
                    btn_3d.configure(state="normal")
                except Exception:
                    pass

                try:
                    export_graph2_button.configure(state="normal")
                except Exception:
                    pass

                _redraw_monthly_graph()
                _redraw_graph2()

                if analysis_table_frame is not None:
                    analysis_table_frame.destroy()

                analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
                analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
                graph_widgets.append(analysis_table_frame)

                _rebuild_analysis_table()

            except Exception:
                messagebox.showerror("Erreur", traceback.format_exc())
                print("[IA] ERREUR:\n", traceback.format_exc())

        bouton_train.configure(command=_train_model)

        # ============ SCROLL MOLETTE ============

        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"
                                                        
        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # ============ BOUTON RETOUR ============

        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)

if __name__ == "__main__":
    app = Application()
    app.mainloop()
    