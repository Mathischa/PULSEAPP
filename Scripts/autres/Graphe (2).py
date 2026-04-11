# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict
import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter  # (si nécessaire ailleurs)

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# === Répertoire contenant les fichiers mensuels ===
from pathlib import Path
import os
import os, time, subprocess, unicodedata
from pathlib import Path
import winreg



from pathlib import Path
import os
import unicodedata
# === Modèle ML (scikit-learn) ===
try:
    from sklearn.ensemble import RandomForestRegressor
except ImportError:
    RandomForestRegressor = None
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor


USER_ID = Path.home().name
BASE_DIR = Path(fr"C:\Users\{USER_ID}\SNCF")
import tkinter as tk

# --- PATCH pour empêcher les RuntimeError "main thread is not in main loop" ---

# Sauvegarde des __del__ originaux
_orig_image_del = getattr(tk.Image, "__del__", None)
_orig_var_del = getattr(tk.Variable, "__del__", None)

def _safe_image_del(self):
    if _orig_image_del is None:
        return
    try:
        _orig_image_del(self)
    except RuntimeError:
        # On ignore seulement le cas où Tk n'est plus dans la boucle principale
        pass

def _safe_var_del(self):
    if _orig_var_del is None:
        return
    try:
        _orig_var_del(self)
    except RuntimeError:
        # Pareil ici : on ne fait rien si Tk est déjà "mort"
        pass

if _orig_image_del is not None:
    tk.Image.__del__ = _safe_image_del

if _orig_var_del is not None:
    tk.Variable.__del__ = _safe_var_del

# On n’impose QUE la "queue" du chemin (fin de chemin), flexible sur tout le reste.
REQUIRED_TAIL = ["Partage - Invités", "Projet PULSE", "4. Données historiques", "Développement"]

def _norm(s: str) -> str:
    """Normalise pour comparaisons: minuscules, accents retirés, espaces/underscores compressés."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("_", " ").strip()
    s = " ".join(s.split())
    return s

def _match_tail(path: Path, tail: list[str]) -> bool:
    """Vérifie que 'path' se termine par les éléments de 'tail' (nom pour nom), comparés normalisés."""
    parts = list(path.parts)
    if len(parts) < len(tail):
        return False
    # Compare en partant de la fin
    for i in range(1, len(tail)+1):
        if _norm(parts[-i]) != _norm(tail[-i]):
            return False
    return True

def find_sharepoint_base():
    """
    1) Si ton chemin 'par défaut' existe, on l’utilise.
    2) Sinon, on parcourt TOUT 'C:\\Users\\<USER_ID>\\SNCF' pour trouver un dossier
       qui SE TERMINE par 'Partage - Invités/Projet PULSE/4. Données historiques/Développement'.
    """
    # 1) Essai direct sur ton chemin "canonique"
    default_path = BASE_DIR / "DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions" / Path(*REQUIRED_TAIL)
    if default_path.exists():
        print(f"[OK] Bibliothèque détectée (DCF par défaut) : {default_path}")
        return default_path

    # 2) Scan générique : on ne regarde QUE la fin de chemin (tail), robustes aux variations amont
    print("[INFO] Recherche générique de    la queue de chemin '.../Partage - Invités/Projet PULSE/4. Données historiques/Développement'...")
    candidates = []
    for root, dirs, _ in os.walk(BASE_DIR):
        # accélère: on ne teste que si "Développement" est présent dans le dossier courant
        if any(_norm(d) == _norm(REQUIRED_TAIL[-1]) for d in dirs):
            for d in dirs:
                full = Path(root) / d
                if _match_tail(full, REQUIRED_TAIL):
                    candidates.append(full)

    if candidates:
        # Choix simple : on prend le plus long chemin (le plus profond) ou simplement le premier trouvé
        best = sorted(candidates, key=lambda p: len(p.parts), reverse=True)[0]
        print(f"[OK] Bibliothèque trouvée automatiquement : {best}")
        return best

    print("❌ Aucune bibliothèque SharePoint locale trouvée correspondant à la queue requise.")
    return None

# --- Résolution finale (inchangé) ---
DEV_PATH = find_sharepoint_base()
if not DEV_PATH:
    raise FileNotFoundError("Impossible de localiser la bibliothèque SharePoint (vérifie la synchro OneDrive/SharePoint).")

FICHIER_EXCEL_DIR = str(DEV_PATH / "Résultats")
FICHIER_CONFIG_SECTIONS = str(DEV_PATH / "Filiales Analysées.xlsx")
image_path = str(DEV_PATH / "Images" / "logo_Pulse.png")
BASE_DONNEES_DIR = str((DEV_PATH / "Données"))


print("\n[INIT] Dossier Résultats :", FICHIER_EXCEL_DIR)
print("[INIT] Fichier Config    :", FICHIER_CONFIG_SECTIONS)
print("[INIT] Image            :", image_path)
print("[INIT] Dossier Données   :\n", BASE_DONNEES_DIR)

if not Path(FICHIER_EXCEL_DIR).exists():
    print(f"⚠️  Dossier 'Résultats' introuvable : {FICHIER_EXCEL_DIR}")
if not Path(FICHIER_CONFIG_SECTIONS).exists():
    print(f"⚠️  Fichier 'Filiales Analysées.xlsx' introuvable : {FICHIER_CONFIG_SECTIONS}")
if not Path(image_path).exists():
    print(f"⚠️  Image 'logo_Pulse.png' introuvable : {image_path}")

# ======================
#   FONCTIONS DE SÉCURISATION
# ======================

def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")

def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)

def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(["attrib", path], capture_output=True, text=True, shell=True)
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None

def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")

def pin_and_hydrate(path: str, timeout=30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l’hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False

def robust_load_workbook(path: str, retries=3, delay=1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} ({type(e).__name__}: {e})")
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e

def diag_path(path: str):
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
    
# === Fichier Excel contenant la configuration des sections ===
FEUILLE_CONFIG_SECTIONS = "CONFIG_SECTIONS"  # ou le vrai nom de la feuille dans ton fichier
COL_DEST = 1   # A
COL_SOURCE = 2 # B
COL_PREV = 3   # C

def charger_sections_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit le fichier Excel de configuration des sections et renvoie une liste
    de dictionnaires {"source": ..., "prev": ..., "dest": ...}.
    """
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return []

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return []

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return []

    ws = wb[FEUILLE_CONFIG_SECTIONS]
    mapping = []

    row = 2  # ligne 1 = en-têtes
    while True:
        dest_val = ws.cell(row=row, column=COL_DEST).value
        source_val = ws.cell(row=row, column=COL_SOURCE).value
        prev_val = ws.cell(row=row, column=COL_PREV).value

        if dest_val is None or str(dest_val).strip() == "":
            break  # fin du tableau

        mapping.append({
            "dest": str(dest_val).strip(),
            "source": str(source_val).strip() if source_val else str(dest_val).strip(),
            "prev": str(prev_val).strip() if prev_val else str(dest_val).strip(),
        })
        row += 1

    wb.close()

    print("[INFO] Sections chargées depuis 'Filiales Analysées.xlsx' :")
    for m in mapping:
        print(f"   - dest='{m['dest']}' | source='{m['source']}' | prev='{m['prev']}'")

    return mapping

def charger_noms_feuilles_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit uniquement la colonne A (dest) de la feuille de config et
    retourne la liste des noms de feuilles à utiliser (strings non vides, uniques, ordonnées).
    """
    noms = []
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return noms

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return noms

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return noms

    ws = wb[FEUILLE_CONFIG_SECTIONS]

    row = 2  # ligne 1 = en-têtes
    seen = set()
    while True:
        val = ws.cell(row=row, column=COL_DEST).value
        if val is None or str(val).strip() == "":
            break
        name = str(val).strip()
        if name not in seen:
            noms.append(name)
            seen.add(name)
        row += 1

    wb.close()

    print("[INFO] Feuilles (dest) chargées depuis 'Filiales Analysées.xlsx' :")
    for n in noms:
        print(f"   - {n}")
    return noms

# === Construire dynamiquement `sections` à partir de la colonne A ===
_dest_names = charger_noms_feuilles_depuis_cells()

# Création directe du dictionnaire sections
sections = {name: name for name in _dest_names}  # clé logique -> nom de feuille Excel


# Optionnel mais pratique : caler la feuille de référence sur la première feuille disponible
if _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]   # ex: "SNCF_SA"

SECTIONS_CONFIG = charger_sections_depuis_cells()
print(SECTIONS_CONFIG)

print(FICHIER_EXCEL_DIR)

# === Regex utilitaires ===
_MENSUEL_RE = re.compile(r"^Historique_prev_reel_filiales_(\d{4})_(\d{2})\.xlsx$")
_RE_PREV = re.compile(r"^Prévision\s+\d{2}/\d{2}", re.I)

def _is_prev(h):  return isinstance(h, str) and _RE_PREV.search(h.strip())
def _is_dates(h): return isinstance(h, str) and h.strip().lower().startswith("date")
def _is_reel(h):  return isinstance(h, str) and "réel" in h.lower()

def _parse_excel_date(v):
    """Convertit proprement tout format de date Excel (float, datetime, str...)"""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        # borne basse/haute pour éviter les overflow absurdes
        if 10000 <= v <= 60000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")).normalize()
            except Exception:
                return None
        else:
            return None
    if isinstance(v, str):
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        return ts.normalize() if pd.notna(ts) else None
    ts = pd.to_datetime(v, errors="coerce")
    return ts.normalize() if pd.notna(ts) else None

# -------------------------------------------------------------------------
# 1️⃣ Structures globales
# -------------------------------------------------------------------------
STRUCT = {}
TOKENS = {}
PREV_UNION = set()
FEUILLE_REFERENCE = "SNCF_SA"
CACHE = {}

# Index annuel: (section, flux) -> { "years": { year: { "row_idx": [...], "prof_idx": [...], "headers": [...] } } }
YEAR_INDEX = {}

def _clean_profil_label(raw, i):
    """Nettoyage identique à l'historique: (K€) retiré, 'Prévision' -> 'Profil', fallback."""
    if raw is None:
        return f"Profil {i+1}"
    s = str(raw).replace("(K€)", "").replace("Prévision", "Profil").strip()
    return s if s else f"Profil {i+1}"

def _flux_name_from_token(section, token_col):
    """Retrouve le nom de flux à partir de la colonne 'token' (col_start) dans TOKENS[section]."""
    for name, tok in TOKENS.get(section, []):
        if tok == token_col:
            return name
    return None


# -------------------------------------------------------------------------
# 2️⃣ Lister les fichiers mensuels
# -------------------------------------------------------------------------

def _lister_fichiers_mensuels():
    """
    Parcourt récursivement FICHIER_EXCEL_DIR pour trouver tous les fichiers
    'Historique_prev_reel_filiales_YYYY_MM.xlsx', y compris dans des sous-dossiers d'années.
    En cas de doublons pour un même (YYYY, MM), conserve le fichier le plus récent (mtime).
    """
    print("🔍 Recherche des fichiers mensuels (scan récursif des sous-dossiers d'années)...")

    if not os.path.isdir(FICHIER_EXCEL_DIR):
        print(f"⚠️ Dossier introuvable : {FICHIER_EXCEL_DIR}")
        return []

    best = {}  # (year, month) -> {"path": str, "mtime": float}
    try:
        for root, _, files in os.walk(FICHIER_EXCEL_DIR):
            for fname in files:
                m = _MENSUEL_RE.match(fname)
                if not m:
                    continue
                y, mth = int(m.group(1)), int(m.group(2))
                fullpath = os.path.join(root, fname)
                try:
                    mtime = os.path.getmtime(fullpath)
                except OSError:
                    mtime = 0.0

                key = (y, mth)
                if key not in best or mtime > best[key]["mtime"]:
                    best[key] = {"path": fullpath, "mtime": mtime}
    except Exception as e:
        print(f"⚠️ Erreur pendant le scan récursif : {e}")
        return []

    if not best:
        print("⚠️ Aucun fichier Historique_prev_reel_filiales_YYYY_MM.xlsx trouvé (sous-dossiers inclus).")
        return []

    files = sorted([(y, mth, info["path"]) for (y, mth), info in best.items()])
    print(f"✅ {len(files)} fichiers retenus : {[os.path.basename(f[2]) for f in files]}")
    return files

def _parse_prev_header_sort_key(h):
    # Essaie d'en déduire un tri MM/YY si possible, sinon garde l’ordre d’apparition
    import re
    s = str(h)
    m = re.search(r'(\d{2})/(\d{2,4})', s)
    if not m:
        return (9999, 99, s)
    mm = int(m.group(1))
    yy = int(m.group(2))
    if yy < 100:
        yy += 2000
    return (yy, mm, s)

def _reconcile_headers(B, new_headers):
    """
    Fusionne les headers déjà présents dans B avec ceux du fichier courant :
    - construit l’union ordonnée
    - re-map B['prev_vals'] dans ce nouvel ordre
    - rétro-remplit de None pour les séries nouvellement apparues
    """
    if "prev_headers" not in B or "prev_vals" not in B:
        B["prev_headers"] = []
        B["prev_vals"] = []

    old_headers = list(B["prev_headers"])
    # union en conservant l’ordre d’apparition, puis tente un tri chronologique
    seen = {}
    for h in old_headers + list(new_headers):
        if h not in seen:
            seen[h] = None
    union = list(seen.keys())
    union.sort(key=_parse_prev_header_sort_key)

    # rien à faire ?
    if union == old_headers:
        return

    # re-map des anciennes séries dans le nouveau canevas
    old_idx = {h: i for i, h in enumerate(old_headers)}
    n_rows = len(B.get("dates", []))
    new_prev_vals = []
    for h in union:
        if h in old_idx:
            new_prev_vals.append(B["prev_vals"][old_idx[h]])
        else:
            # nouvelle série : rétro-remplir pour les lignes déjà connues
            new_prev_vals.append([None] * n_rows)

    B["prev_headers"] = union
    B["prev_vals"] = new_prev_vals

# -------------------------------------------------------------------------
# 3️⃣ Lecture de la structure sur un fichier de référence
# -------------------------------------------------------------------------
def _lire_structure_reference(path_ref):
    print(f"\n📘 Lecture de la structure de référence ({FEUILLE_REFERENCE}) → {os.path.basename(path_ref)}")
    df_head = pd.read_excel(path_ref, sheet_name=FEUILLE_REFERENCE, header=None, nrows=3)
    row2 = df_head.iloc[1].tolist()
    row3 = df_head.iloc[2].tolist()

    ref_od = OrderedDict()
    ref_tokens = []
    col = 2  # colonne C (0-based)
    token_col = 3

    while col < len(row2):
        flux_name = row2[col]
        if pd.isna(flux_name):
            break
        flux_name = str(flux_name).strip()

        prev_headers, prev_cols = [], []
        c = col + 1
        while c < len(row3):
            h = row3[c]
            if isinstance(h, str) and "Prévision" in h:
                prev_headers.append(h.strip())
                prev_cols.append(c)
                c += 2
                continue
            break

        if prev_headers:
            ref_od[flux_name] = {
                "col_start": col + 1,
                "prev_headers": prev_headers,
                "prev_cols": [x + 1 for x in prev_cols],
                "dates_col": col,
                "reel_col": col + 1,
            }
            ref_tokens.append((flux_name, token_col))
            PREV_UNION.update(prev_headers)

        col += (2 + 2 * len(prev_headers) + 1)
        token_col += (2 + 2 * len(prev_headers) + 1)

    print(f"   → {len(ref_od)} flux détectés ({len(PREV_UNION)} prévisions).")
    for sec in sections:
        STRUCT[sec] = ref_od.copy()
        TOKENS[sec] = ref_tokens.copy()
    print("✅ Structure copiée sur toutes les feuilles.\n")


# -------------------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# -------------------------------------------------------------------------
def _ensure_flux_bucket(section, flux_name, headers=None):
    key = (section, flux_name)
    if key not in CACHE:
        CACHE[key] = {
            "dates": [],
            "reel": [],
            "prev_headers": [],   # ← vide
            "prev_vals": [],      # ← vide
        }
    return CACHE[key]


def _accumuler_valeurs_tous_mois(files_list):
    """
    Lit chaque fichier mensuel et alimente CACHE[(section, flux)] en alignant
    les séries de prévision sur l'union ordonnée des en-têtes (headers).
    - Réconciliation des headers à chaque fichier (ajout rétro-rempli de None).
    - Insertion directe ligne par ligne dans B["dates"], B["reel"], B["prev_vals"][k].
    """
    print(f"📊 Lecture de {len(files_list)} fichiers mensuels (toutes feuilles)...")
    for idx, (_, mois, path) in enumerate(files_list, 1):
        print(f"   [{idx}/{len(files_list)}] → {os.path.basename(path)}")

        try:
            all_sheets = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as e:
            print(f"⚠️ Erreur lecture {path}: {e}")
            continue

        for sec, feuille in sections.items():
            if feuille not in all_sheets:
                continue

            df = all_sheets[feuille].copy()
            if df.shape[0] < 5:
                continue

            row2 = df.iloc[1].tolist()
            row3 = df.iloc[2].tolist()
            col = 2  # colonne C (0-based -> ici index pandas)

            while col < len(row2):
                flux_name = row2[col]
                if pd.isna(flux_name):
                    break
                flux_name = str(flux_name).strip()

                # Détection des colonnes "Prévision"
                prev_headers, prev_cols = [], []
                c = col + 1
                while c < len(row3):
                    h = row3[c]
                    if isinstance(h, str) and "Prévision" in h:
                        prev_headers.append(h.strip())
                        prev_cols.append(c)
                        c += 2  # on saute la colonne (k€) qui suit chaque prévision
                        continue
                    break

                if not prev_headers:
                    # Bloc sans prévisions (ex: espaces, colonnes diverses) → on saute "gros"
                    col += 10
                    continue

                # 1) Bucket pour ce flux
                B = _ensure_flux_bucket(sec, flux_name)

                # 2) Réconcilier les en-têtes avec ceux du fichier courant
                _reconcile_headers(B, prev_headers)

                # 3) Map header -> colonne pour CE fichier
                header_to_col = {h: prev_cols[j] for j, h in enumerate(prev_headers)}

                # 4) Pousser les lignes : date, réel, et séries alignées
                rows_appended = 0
                for r in range(3, df.shape[0]):
                    date_val = df.iat[r, col - 1]
                    if pd.isna(date_val):
                        continue
                    d = _parse_excel_date(date_val)
                    if d is None:
                        continue

                    B["dates"].append(d)

                    r_val = df.iat[r, col]
                    B["reel"].append(r_val if pd.notna(r_val) else 0)

                    # Pour chaque header du bucket, si la colonne existe ce mois-ci → valeur ; sinon → None
                    for k, h in enumerate(B["prev_headers"]):
                        if h in header_to_col:
                            cpr = header_to_col[h]
                            v = df.iat[r, cpr]
                            B["prev_vals"][k].append(v if pd.notna(v) else None)
                        else:
                            B["prev_vals"][k].append(None)

                    rows_appended += 1

                # Avance au bloc suivant : (dates + réel) + 2 colonnes par prévision + 1 de séparation
                col += (2 + 2 * len(prev_headers) + 1)

    print(f"✅ Cache complété : {len(CACHE)} flux au total (valeurs réelles et prévisions incluses).\n")

    # Récapitulatif par section (contrôle)
    print("📊 Récapitulatif des flux chargés par section :")
    counts = {}
    for (section, flux), data in CACHE.items():
        counts.setdefault(section, 0)
        counts[section] += 1
    for section, n in counts.items():
        total_lignes = sum(len(CACHE[(section, f)]['dates']) for f in [x for x, _ in CACHE if _ == section])
        print(f"   - {section} : {n} flux ({total_lignes} lignes cumulées)")
    print()


# -------------------------------------------------------------------------
# 5️⃣ Index annuel
# -------------------------------------------------------------------------
def _build_year_index():
    """
    Pour chaque (section, flux) dans CACHE :
      - regroupe les indices de lignes par année
      - détermine quels profils (colonnes) sont 'actifs' sur l'année (au moins une valeur != 0 / non None)
      - prépare la liste de headers nettoyés pour l'année
    Remplit YEAR_INDEX[(section, flux)].
    """
    YEAR_INDEX.clear()
    for (section, flux_name), B in CACHE.items():
        dates = B.get("dates", [])
        prev_vals = B.get("prev_vals", [])
        headers = B.get("prev_headers", [])

        # indices de lignes par année
        rows_by_year = {}
        for i, d in enumerate(dates):
            y = d.year
            rows_by_year.setdefault(y, []).append(i)

        years_map = {}
        for y, row_idx in rows_by_year.items():
            prof_idx = []
            for k, serie in enumerate(prev_vals):
                # activité = au moins une valeur non None et != 0 sur ces lignes
                active = False
                for i in row_idx:
                    if i < len(serie):
                        v = serie[i]
                        if v is None:
                            continue
                        try:
                            if float(v) != 0.0:
                                active = True
                                break
                        except Exception:
                            active = True
                            break
                if active:
                    prof_idx.append(k)

            headers_year = [_clean_profil_label(headers[k] if k < len(headers) else None, k) for k in prof_idx]

            years_map[y] = {
                "row_idx": row_idx,       # indices de lignes pour cette année
                "prof_idx": prof_idx,     # indices de colonnes actives pour cette année
                "headers": headers_year,  # labels propres des profils actifs
            }

        YEAR_INDEX[(section, flux_name)] = {"years": years_map}

# -------------------------------------------------------------------------
# 6️⃣ Initialisation complète
# -------------------------------------------------------------------------
def _init_full_load():
    print("🚀 Initialisation complète du cache de données...")
    files = _lister_fichiers_mensuels()
    if not files:
        print("ℹ️ Aucun fichier trouvé — initialisation du cache ignorée.\n")
        return

    ref_path = files[-1][2]  # dernier mois = référence
    _lire_structure_reference(ref_path)
    _accumuler_valeurs_tous_mois(files)
    _build_year_index()  # ✅ construire l’index après remplissage du CACHE
    print("✅ Chargement complet terminé.\n")


# Lance le chargement maintenant que tout est défini
_init_full_load()

# -------------------------------------------------------------------------
# 7️⃣ Variables dérivées compatibles avec l'existant
# -------------------------------------------------------------------------
previsions_triees = sorted(PREV_UNION)
nb_prev = len(previsions_triees)
taille_bloc = 2 + 2 * nb_prev + 1
print(f"📈 Nombre total de prévisions : {nb_prev}")
print(f"📦 Taille d’un bloc de flux : {taille_bloc}\n")


# -------------------------------------------------------------------------
# 8️⃣ Fonctions publiques
# -------------------------------------------------------------------------
def charger_donnees(feuille, taille_bloc_param):
    """Retourne le nom de la feuille (section) et la liste des flux détectés, en excluant certains flux inutiles."""
    print(f"➡️ charger_donnees() appelé pour '{feuille}'")

    # Liste des flux à ignorer
    noms_a_exclure = [
        "Trésorerie de fin",
        "Cashpool",
        "Emprunts",
        "Tirages Lignes CT",
        "Variation de collatéral",
        "Créances CDP",
        "Placements",
        "CC financiers",
        "Emprunts / Prêts - Groupe",
        "Encours de financement",
        "Endettement Net"
    ]

    # Récupère tous les flux
    all_noms = list(TOKENS.get(feuille, []))

    # Filtrage : on garde uniquement ceux qui ne sont pas exclus
    noms = [(name, col) for name, col in all_noms if name not in noms_a_exclure]
    for j in noms :
        print(f"noms : {j}")
    print(len(noms))
    
    if not noms:
        print(f"⚠️ Aucun flux valide trouvé pour la feuille '{feuille}'")
    else:
        print(f"   ✅ {len(noms)} flux conservés (sur {len(all_noms)} totaux)")

    return feuille, noms
# Profils à exclure pour 2023 (d’après tes listes / cases à cocher)
EXCLUDED_DATES_2023 = {
    "02/01",
    "09/01",
    "16/01",
    "30/05",
    "05/06",
    "19/06",
    "26/06",
    "03/07",
    "10/07",
    "17/07",
}

def extraire_valeurs(ws, col_start, nb_prev_param, annee=None, annee_min=None, annee_max=None):
    """
    Retourne (dates, reel, previsions, noms_profils)

    - Si 'annee' est fourni : utilise YEAR_INDEX pour ne garder que
      *les lignes* de l'année et *les profils actifs* de l'année,
      puis on filtre éventuellement certains profils (2023) → le RÉEL reste inchangé.
    - Sinon : filtre sur plage ou aucun filtre (comportement historique).
    """

    DEBUG_EXTRAIRE = True
    def _dbg(*args):
        if DEBUG_EXTRAIRE:
            print("[extraire_valeurs]", *args)

    section = ws
    flux_name = _flux_name_from_token(section, col_start)
    _dbg(f"CALL: section={section!r} col_start={col_start} nb_prev_param={nb_prev_param} "
         f"annee={annee} annee_min={annee_min} annee_max={annee_max}")

    if not flux_name:
        _dbg("ABORT: flux_name introuvable -> retour listes vides/pad prév.")
        return [], [], [[] for _ in range(nb_prev_param)], []

    B = CACHE.get((section, flux_name))
    if not B or not B.get("dates"):
        _dbg(f"ABORT: pas de cache ou dates vides pour {(section, flux_name)}")
        return [], [], [[] for _ in range(nb_prev_param)], []

    dates_all = B["dates"]
    reel_all  = B["reel"]
    prev_all  = B.get("prev_vals", [])
    headers   = B.get("prev_headers", [])   # ex: "Prévision 16/01 (K€)"

    _dbg(f"CACHE: len(dates_all)={len(dates_all)}, len(reel_all)={len(reel_all)}, "
         f"nb_profils_total={len(prev_all)}, nb_headers={len(headers)}")

    # ==============================
    # CAS 1 : année précise
    # ==============================
    if annee is not None:
        years = YEAR_INDEX.get((section, flux_name), {}).get("years", {})
        info = years.get(annee)
        if not info:
            _dbg(f"ANNEE={annee}: YEAR_INDEX absent/vide -> retour listes vides.")
            return [], [], [], []

        # Ces indices NE CHANGENT PAS pour les lignes → le RÉEL reste intact
        row_idx  = info["row_idx"]      # indices des dates pour cette année
        prof_idx = info["prof_idx"]     # indices des colonnes de prévision actives

        _dbg(f"ANNEE={annee}: rows={len(row_idx)} profils_actifs={len(prof_idx)}")

        # ---- Filtre spécial 2023 : on enlève certains profils, mais on garde le réel ----
        if annee == 2023:
            def _is_excluded(k: int) -> bool:
                if k < 0 or k >= len(headers):
                    return False
                lab = str(headers[k])  # "Prévision 16/01 (K€)"
                for d in EXCLUDED_DATES_2023:
                    if d in lab:
                        return True
                return False

            prof_idx_before = list(prof_idx)
            prof_idx = [k for k in prof_idx if not _is_excluded(k)]
            _dbg(f"ANNEE=2023: prof_idx avant filtre = {prof_idx_before}")
            _dbg(f"ANNEE=2023: prof_idx après  filtre = {prof_idx}")

        # ---- Extraction des lignes (réel inchangé) ----
        dates = [dates_all[i] for i in row_idx]
        reel  = [reel_all[i]  for i in row_idx]

        # ---- Extraction des prévisions uniquement pour les profils retenus ----
        previsions = []
        for k in prof_idx:
            serie = prev_all[k] if k < len(prev_all) else []
            previsions.append([serie[i] if i < len(serie) else None for i in row_idx])

        # Noms de profils alignés sur prof_idx
        noms_profils = [
            _clean_profil_label(headers[k] if k < len(headers) else None, k)
            for k in prof_idx
        ]

        _dbg(f"RETURN(annee): len(dates)={len(dates)}, len(reel)={len(reel)}, "
             f"nb_profils={len(previsions)}, noms_profils={noms_profils}")
        return dates, reel, previsions, noms_profils

    # ==============================
    # CAS 2 : plage / historique
    # ==============================
    def _keep(d):
        y = d.year
        if annee_min is not None and annee_max is not None:
            return annee_min <= y <= annee_max
        if annee_min is not None:
            return y >= annee_min
        if annee_max is not None:
            return y <= annee_max
        return True

    idxs = [i for i, d in enumerate(dates_all) if _keep(d)]
    dates = [dates_all[i] for i in idxs]
    reel  = [reel_all[i]  for i in idxs]
    previsions = [[serie[i] if i < len(serie) else None for i in idxs] for serie in prev_all]

    noms_profils = [_clean_profil_label(h, i) for i, h in enumerate(headers)]

    _dbg(f"PLAGE/HISTO: indices_retenus={len(idxs)} "
         f"(min_year={min((d.year for d in dates), default=None)}, "
         f"max_year={max((d.year for d in dates), default=None)})")
    _dbg(f"RETURN(plage): len(dates)={len(dates)}, len(reel)={len(reel)}, "
         f"nb_profils={len(previsions)}, len(noms_profils)={len(noms_profils)}; "
         f"exemple_date={dates[0] if dates else None}")

    return dates, reel, previsions, noms_profils


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        self.attributes("-fullscreen", True)
        self.geometry("1300x900")
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

#===================Page Accueil + fonctions de navigation===================
    def creer_accueil(self):
        # -- RESET & imports locaux
        import customtkinter as ctk
        from PIL import Image
        import tkinter as tk
        from customtkinter import CTkImage
        from datetime import datetime
        import os

        # === THEME / BASE ===
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color="#0b1220")
        except Exception:
            pass

        # Détruit proprement les enfants
        self.vider_fenetre()

        # =========================
        #  Helpers data pour KPI
        # =========================
        def _fmt_dt(ts):
            try:
                return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return "—"

        try:
            fichiers = _lister_fichiers_mensuels()
            nb_fichiers = len(fichiers)
            if fichiers:
                last_mtime = max(os.path.getmtime(p) for _, _, p in fichiers if os.path.exists(p))
                derniere_maj = _fmt_dt(last_mtime)
            else:
                derniere_maj = "—"
        except Exception:
            nb_fichiers, derniere_maj = 0, "—"

        try:
            if SECTIONS_CONFIG:
                nb_filiales = len({s.get("dest") for s in SECTIONS_CONFIG if s.get("dest")})
            else:
                nb_filiales = len(_dest_names) if '_dest_names' in globals() else 0
        except Exception:
            nb_filiales = 0

        # =========================
        #  GRID RACINE (window)
        # =========================
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        # =========================
        #  HEADER
        # =========================
        header = ctk.CTkFrame(self, fg_color="#0f1b31", corner_radius=0, height=72)
        header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        header.grid_columnconfigure(0, weight=0)
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        # Barre blanche sous le header
        barre_header = ctk.CTkFrame(self, fg_color="white", height=2)
        barre_header.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Logo à gauche
        logo_wrap = ctk.CTkFrame(header, fg_color="transparent")
        logo_wrap.grid(row=0, column=0, sticky="w", padx=24, pady=12)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h, new_w = 44, int(44 * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(logo_wrap, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            ctk.CTkLabel(logo_wrap, text="PULSE", font=("Segoe UI Semibold", 20),
                        text_color="white").pack(side="left")

        titre = ctk.CTkLabel(
            header,
            text="PROJET PULSE — ANALYSE DE L’EXISTANT",
            font=("Segoe UI Semibold", 18, "bold"),
            text_color="white",
        )
        titre.grid(row=0, column=1, sticky="w", padx=8)

        # Actions header à droite (recherche + quit)
        header_actions = ctk.CTkFrame(header, fg_color="transparent")
        header_actions.grid(row=0, column=2, sticky="e", padx=20)

        search_entry = ctk.CTkEntry(header_actions, placeholder_text="Rechercher une filiale, un flux...", width=320)
        search_entry.grid(row=0, column=0, padx=(0, 8))
        ctk.CTkButton(header_actions, text="❌ Quitter", fg_color="#C21515", hover_color="#A00000",
                    command=self.demander_confirmation_quit).grid(row=0, column=1)

        # =========================
        #  SIDEBAR
        # =========================
        sidebar = ctk.CTkFrame(self, fg_color="#0e162a", corner_radius=0, width=260)
        sidebar.grid(row=1, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self.grid_rowconfigure(1, weight=1)
        sidebar.grid_rowconfigure(11, weight=1)

        inner = ctk.CTkFrame(sidebar, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=14, pady=14)

        def _nav_btn(parent, text, cmd, emoji=""):
            btn = ctk.CTkButton(
                parent, text=f"{emoji}  {text}", anchor="w",
                width=220, height=42, corner_radius=12,
                fg_color="#15305b", hover_color="#1d3e77",
                font=("Segoe UI", 13, "bold"),
                command=cmd
            )

            # petit hover animé (couleur qui “boost” légèrement)
            def _enter(_):
                btn.configure(fg_color="#1d3e77")
            def _leave(_):
                btn.configure(fg_color="#15305b")
            btn.bind("<Enter>", _enter)
            btn.bind("<Leave>", _leave)

            return btn

        ctk.CTkLabel(inner, text="Navigation", text_color="#9fb7dd",
                    font=("Segoe UI Semibold", 14)).pack(anchor="w", pady=(4, 12))
        
        _nav_btn(inner, "Importer les profils", self.importer_les_profils, "📁").pack(fill="x", pady=6)
        _nav_btn(inner, "Voir Graphique (Filiale + flux)", self.creer_page_graphique, "📊").pack(fill="x", pady=6)
        _nav_btn(inner, "Voir Écarts Importants", self.afficher_ecarts, "⚠️").pack(fill="x", pady=6)
        _nav_btn(inner, "Écarts par Filiale", self.afficher_repartition, "📈").pack(fill="x", pady=6)
        _nav_btn(inner, "Écarts par Profil", self.afficher_repartition_par_prevision, "🧩").pack(fill="x", pady=6)
        _nav_btn(inner, "Écarts par Flux", self.afficher_repartition_flux, "🔀").pack(fill="x", pady=6)
        _nav_btn(inner, "Heatmap des anomalies", self.afficher_heatmap_anomalies, "🔥").pack(fill="x", pady=6)
        _nav_btn(inner, "Heatmap des écarts", self.afficher_heatmap_ecarts, "🗺️").pack(fill="x", pady=6)
        _nav_btn(inner, "Clustering des écarts", self.analyser_ecarts_ml, "📈").pack(fill="x", pady=6)
        _nav_btn(inner, "Backtesting multi-horizon", self.afficher_backtesting_multi_horizon, "📉").pack(fill="x", pady=6)
        _nav_btn(inner, "IA — Prédiction Enc. Autre Prod.", self.creer_page_ia_prediction, "🤖").pack(fill="x", pady=6)

        separator = ctk.CTkFrame(sidebar, fg_color="#1a2745", height=2)
        separator.pack(fill="x", padx=14, pady=(4, 2))

        foot = ctk.CTkLabel(sidebar,
                            text=f"v1.0 • Dernière MAJ : {derniere_maj}",
                            text_color="#6e86ad",
                            font=("Segoe UI", 11))
        foot.pack(anchor="w", padx=20, pady=(8, 10))

        # =========================
        #  MAIN CONTENT
        # =========================
        main = ctk.CTkScrollableFrame(self, fg_color="#0b1220", corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        self.grid_columnconfigure(1, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # ====== HERO ======
        hero = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 14))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_columnconfigure(1, weight=0)

        # Titre
        ctk.CTkLabel(
            hero,
            text="Bienvenue 👋",
            font=("Segoe UI Semibold", 24, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 0))

        # Sous-titre avec effet “machine à écrire”
        subtitle_text = (
            "Console de pilotage des flux de trésorerie : réels, prévisions, écarts et détection d'anomalies."
        )
        subtitle_label = ctk.CTkLabel(
            hero,
            text="",
            font=("Segoe UI", 13),
            text_color="#c9defa",
            wraplength=520,
            justify="left"
        )
        subtitle_label.grid(row=1, column=0, sticky="w", padx=18, pady=(6, 10))

        def _typewriter(i=0):
            # Animation très légère, safe
            if i <= len(subtitle_text):
                subtitle_label.configure(text=subtitle_text[:i])
                self.after(12, _typewriter, i + 1)

        _typewriter()

        # Pills
        pills = ctk.CTkFrame(hero, fg_color="transparent")
        pills.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 16))

        def _pill(parent, text, emoji=""):
            f = ctk.CTkFrame(parent, fg_color="#152544", corner_radius=20)
            f.pack(side="left", padx=6)
            ctk.CTkLabel(
                f,
                text=f"{emoji}  {text}",
                font=("Segoe UI", 11),
                text_color="#c9defa"
            ).pack(padx=12, pady=5)

        _pill(pills, "Réel vs prévisions multi-horizon", "📈")
        _pill(pills, "Analyse d'écarts & anomalies", "🧠")
        _pill(pills, "Vue filiales / flux / profils", "📊")

        # =========================
        #  Illustration graphique (3 graphes)
        # =========================
        viz = ctk.CTkFrame(hero, fg_color="#111b33", corner_radius=18)
        viz.grid(row=0, column=1, rowspan=3, sticky="e", padx=18, pady=12)
        viz.grid_columnconfigure((0, 1, 2), weight=1)
        viz.grid_rowconfigure(0, weight=1)

        # --- Graph 1 : barres animées ---
        graph1 = ctk.CTkFrame(viz, fg_color="transparent")
        graph1.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        bars_h = [52, 34, 78, 60]
        bar_widgets = []

        for h in bars_h:
            col = ctk.CTkFrame(graph1, fg_color="transparent")
            col.pack(side="left", padx=4, fill="y", expand=True)

            base = ctk.CTkFrame(col, fg_color="#1b2945", corner_radius=8, width=18, height=85)
            base.pack(side="bottom", pady=(0, 4))

            bar = ctk.CTkFrame(col, fg_color="#2563eb", corner_radius=8, width=18, height=1)
            bar.place(relx=0.5, rely=1.0, anchor="s")
            bar_widgets.append((bar, h))

        def _animate_bars(step=0, steps=18):
            # animation simple : on augmente la hauteur progressivement
            factor = min(1.0, step / float(steps))
            for bar, target in bar_widgets:
                bar.configure(height=int(target * factor))
            if step < steps:
                self.after(25, _animate_bars, step + 1, steps)

        _animate_bars()

        # --- Graph 2 : courbe ---
        graph2 = ctk.CTkFrame(viz, fg_color="transparent")
        graph2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        pts = [10, 40, 22, 70, 50, 90, 60, 80, 55, 100]
        w, h = 120, 90

        canvas2 = ctk.CTkCanvas(graph2, width=w, height=h, bg="#111b33", bd=0, highlightthickness=0)
        canvas2.pack(fill="both", expand=True)

        max_v = max(pts)
        coords = [(i * (w // (len(pts) - 1)), h - int((v / max_v) * (h - 10))) for i, v in enumerate(pts)]

        for i in range(len(coords) - 1):
            canvas2.create_line(
                coords[i][0], coords[i][1],
                coords[i + 1][0], coords[i + 1][1],
                fill="#2563eb", width=3, smooth=True
            )
        for x, y in coords:
            canvas2.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#3c82ff", outline="")

        # --- Graph 3 : camembert / donut ---
        graph3 = ctk.CTkFrame(viz, fg_color="transparent")
        graph3.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)

        canvas3 = tk.Canvas(graph3, width=110, height=110,
                            bg="#111b33", bd=0, highlightthickness=0)
        canvas3.pack(fill="both", expand=True, padx=6, pady=6)

        values = [30, 15, 22, 18, 15]
        total = sum(values)
        start_angle = 0
        colors = ["#1d4ed8", "#2563eb", "#3b82f6", "#60a5fa", "#93c5fd"]
        bbox_outer = (8, 8, 102, 102)

        for i, v in enumerate(values):
            extent = 360 * v / total
            canvas3.create_arc(
                *bbox_outer,
                start=start_angle,
                extent=extent,
                fill=colors[i % len(colors)],
                outline="#111b33",
                style="pieslice"
            )
            start_angle += extent

        bbox_inner = (28, 28, 82, 82)
        canvas3.create_oval(*bbox_inner, fill="#111b33", outline="#111b33")

        # ====== KPI CARDS ======
        kpi = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        kpi.grid(row=1, column=0, sticky="ew", padx=24, pady=10)
        for i in range(3):
            kpi.grid_columnconfigure(i, weight=1)

        def _kpi_card(parent, title, value, subtitle, col):
            card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16,
                                border_width=1, border_color="#223658")
            card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
            ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd")\
                .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
            ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 28, "bold"),
                        text_color="white")\
                .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
            ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                        text_color="#7ea2d8")\
                .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

        _kpi_card(kpi, "Fichiers mensuels détectés",
                f"{nb_fichiers:,}".replace(",", " "), "Historique *_YYYY_MM.xlsx*", 0)
        _kpi_card(kpi, "Nombre de filiales",
                f"{nb_filiales:,}".replace(",", " "), "Distincts dans la config", 1)
        _kpi_card(kpi, "Dernière actualisation",
                f"{derniere_maj}", "Basée sur les mtime", 2)

        # ====== GRILLE ACTIONS ======
        actions = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        actions.grid(row=2, column=0, sticky="ew", padx=24, pady=10)
        for i in range(3):
            actions.grid_columnconfigure(i, weight=1)

        def _action_card(parent, title, desc, btn_text, cmd, col, emoji=""):
            card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16,
                                border_width=1, border_color="#223658")
            card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
            ctk.CTkLabel(card, text=title, font=("Segoe UI Semibold", 16, "bold"),
                        text_color="white")\
                .grid(row=0, column=0, sticky="w", padx=16, pady=(14, 4))
            ctk.CTkLabel(card, text=desc, font=("Segoe UI", 12),
                        text_color="#c9defa", wraplength=360, justify="left")\
                .grid(row=1, column=0, sticky="w", padx=16, pady=(0, 12))
            ctk.CTkButton(card, text=f"{emoji}  {btn_text}", height=36,
                        corner_radius=10, fg_color="#2563eb",
                        hover_color="#1d4ed8", command=cmd)\
                .grid(row=2, column=0, sticky="w", padx=16, pady=(0, 16))

        _action_card(actions, "Visualisation Graphique Réel vs Prévisions",
                    "Visualise les données réelles et prévisionnelles pour une filiale, une année et un flux donnés.",
                    "Visualiser graphiquement", self.creer_page_graphique, 0, "🧭")
        _action_card(actions, "Heatmaps des écarts critiques",
                    "Repère visuellement les zones critiques mois/profil/flux avec une heatmaps.",
                    "Voir les heatmaps", self.afficher_heatmap_ecarts, 1, "🌡️")
        _action_card(actions, "Importer des profils",
                    "Ingestion guidée des fichiers profil avec contrôle d’intégrité.",
                    "Importer maintenant", self.importer_les_profils, 2, "📥")

        # ====== ACTIVITÉ RÉCENTE ======
        recent = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        recent.grid(row=3, column=0, sticky="ew", padx=24, pady=(10, 20))
        recent.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(recent, text="Activité récente", font=("Segoe UI Semibold", 16, "bold"),
                    text_color="white").grid(row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        table = ctk.CTkFrame(recent, fg_color="#142544", corner_radius=12,
                            border_width=1, border_color="#223658")
        table.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 16))
        for i in range(3):
            table.grid_columnconfigure(i, weight=[3, 1, 1][i])

        def _header(r, text):
            ctk.CTkLabel(table, text=text, text_color="#9fb7dd",
                        font=("Segoe UI Semibold", 12))\
                .grid(row=r, column=["Fichier", "Taille", "Modifié le"].index(text),
                    sticky="w", padx=12, pady=(10, 6))

        _header(0, "Fichier")
        _header(0, "Taille")
        _header(0, "Modifié le")

        recent_files = []
        try:
            if fichiers:
                rec = sorted([(p, os.path.getmtime(p)) for _, _, p in fichiers if os.path.exists(p)],
                            key=lambda x: x[1], reverse=True)[:5]
                for p, ts in rec:
                    size = f"{(os.path.getsize(p)/1024/1024):.1f} Mo"
                    recent_files.append((os.path.basename(p), size, _fmt_dt(ts)))
        except Exception:
            pass

        if not recent_files:
            recent_files = [("—", "—", "—")]

        for i, (name, size, dt) in enumerate(recent_files, start=1):
            ctk.CTkLabel(table, text=name, text_color="white", font=("Segoe UI", 12))\
                .grid(row=i, column=0, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=size, text_color="#c9defa", font=("Segoe UI", 12))\
                .grid(row=i, column=1, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=dt, text_color="#c9defa", font=("Segoe UI", 12))\
                .grid(row=i, column=2, sticky="w", padx=12, pady=6)

                # ====== NOTES / JOURNAL D'ANALYSE ======
        notes = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        notes.grid(row=5, column=0, sticky="ew", padx=24, pady=(0, 24))
        notes.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            notes,
            text="Notes / journal d’analyse",
            font=("Segoe UI Semibold", 16, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 4))

        ctk.CTkLabel(
            notes,
            text="Utilise cet espace pour garder une trace des points d’attention, hypothèses, et conclusions au fil de tes analyses.",
            font=("Segoe UI", 11),
            text_color="#9fb7dd",
            wraplength=700,
            justify="left"
        ).grid(row=1, column=0, sticky="w", padx=16, pady=(0, 8))

        notes_box = ctk.CTkTextbox(
            notes,
            height=130,
            wrap="word",
            fg_color="#142544",
            text_color="white",
            font=("Segoe UI", 12),
            border_width=1,
            border_color="#223658",
            corner_radius=12
        )
        notes_box.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 14))

        # Texte de départ léger
        notes_placeholder = (
            "Exemples de notes :\n"
            "- Filiale X : écarts récurrents sur les flux intra-groupes.\n"
            "- Vérifier la qualité des prévisions pour les profils Y et Z.\n"
            "- À creuser : saisonnalité forte sur les mois de mars/avril.\n"
        )
        try:
            notes_box.insert("1.0", notes_placeholder)
        except Exception:
            pass


        # Raccourci clavier pour focus sur la recherche
        def _focus_search(event=None):
            try:
                search_entry.focus_set()
            except Exception:
                pass

        self.bind("<Control-f>", _focus_search)

    def vider_fenetre(self):
        for widget in self.winfo_children():
            widget.destroy()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.destroy()

#===================Page Importation des profils===================
    def importer_les_profils(self):
        import customtkinter as ctk
        from tkinter import filedialog, ttk
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import os
        import re

        # ===================== PARAMÈTRES (centralisés) =====================
        # Par défaut : aucune année sélectionnée tant que l'utilisateur n'a pas choisi
        if not hasattr(self, "baseline_year"):
            self.baseline_year = None  # None = pas encore choisi

        # Liste interne des chemins complets des fichiers profil sélectionnés
        # (utilisée plus tard par lancer_import_profils)
        self.profils_import = []

        def build_pattern(year: int | None):
            if year is None:
                return None
            y = year
            y_prev = y - 1

            # autorise que la 2e "année" soit soit juste YYYY, soit une date complète YYYY-MM-DD
            def y_or_date(yy: int) -> str:
                return rf"{yy}(?:-\d{{2}}-\d{{2}})?"

            return re.compile(
                rf'^Profil\s+Tr[ée]so\s+SNCF\s+'
                rf'(?:'
                    rf'{y_prev}\s*[-–—]\s*{y_or_date(y)}'  # année précédente → année courante
                    rf'|'
                    rf'{y}\s*[-–—]\s*{y_or_date(y)}'      # même année → même année
                rf')\b.*\.xlsx$',
                re.IGNORECASE
            )

        # helpers accessibles depuis d'autres méthodes
        self.build_pattern = build_pattern
        self.pattern_profil = build_pattern(self.baseline_year)

        def chercher_profils_regex(root_folder: str, pattern):
            """Retourne tous les fichiers qui matchent la regex. Si pattern=None → []."""
            if pattern is None:
                return []
            resultats = []
            for dossier, _, fichiers in os.walk(root_folder):
                for fichier in fichiers:
                    if pattern.match(fichier):
                        resultats.append(os.path.join(dossier, fichier))
            return resultats

        self.chercher_profils_regex = chercher_profils_regex

        # === PARAMÈTRES DOSSIER DONNÉES + HELPERS Réel AAAA.xlsx ===
        self.BASE_DONNEES_DIR = BASE_DONNEES_DIR
        
        def find_fichier_reel_exact(year: int, base_dir: str):
            """Cherche strictement 'Réel {year}.xlsx' dans base_dir (non récursif)."""
            attendu = f"Réel {year}.xlsx"
            plein = os.path.join(base_dir, attendu)
            return plein if os.path.isfile(plein) else None

        def find_fichier_reel_flexible(year: int, root_dir: str):
            """
            Cherche 'Réel {year}.xlsx' de manière tolérante (R/É/é), d'abord non récursif puis récursif.
            """
            patt = re.compile(rf"^R[ée]el\s+{year}\.xlsx$", re.IGNORECASE)

            # 1) Non récursif
            try:
                for f in os.listdir(root_dir):
                    if patt.match(f):
                        return os.path.join(root_dir, f)
            except FileNotFoundError:
                pass

            # 2) Récursif (fallback)
            for d, _, files in os.walk(root_dir):
                for f in files:
                    if patt.match(f):
                        return os.path.join(d, f)
            return None

        self.find_fichier_reel_exact = find_fichier_reel_exact
        self.find_fichier_reel_flexible = find_fichier_reel_flexible
        self.fichier_source = None  # mémorisé à la sélection d'année

        # =====================================================

        self.vider_fenetre()

        # --- HEADER ---
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 28, "bold")
        titre_label = ctk.CTkLabel(
            header_frame,
            text="🚀 PROJET PULSE - IMPORTER LES PROFILS",
            font=titre_font,
            text_color="white"
        )
        titre_label.pack(side="left", anchor="w", padx=(10, 0))

        # Logo SNCF
        try:
            image_path = image_path  
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(20, 10))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # --- CONTAINER PRINCIPAL (arrondi) ---
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=20)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        # configure grid pour que le canvas prenne tout l'espace
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # frame qui va contenir le canvas (facilite sticky/expand)
        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")

        # canvas principal
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")

        # faire en sorte que le canvas prenne tout l'espace du canvas_frame
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        # scrollbars dans le container
        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")

        # frame scrollable (contenu) placé à l'intérieur du canvas
        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # mise à jour de la scrollregion quand le contenu change
        def _on_configure(e):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_configure)

        # lier les scrollbars au canvas
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # ------------------------------------------------------------------
        # === CONTENU (identique, placé dans scrollable_frame)
        # ------------------------------------------------------------------

        # Deux colonnes : gauche = tableau, droite = boutons
        left_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        left_frame.pack(side="left", fill="both", expand=True, padx=(30, 15), pady=30)

        right_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        right_frame.pack(side="right", fill="y", padx=(15, 30), pady=30)

        # --- Tableau fichiers trouvés ---
        colonnes = ("Nom des fichiers profil",)
        style = ttk.Style()
        style.configure("Treeview", rowheight=30, font=("Segoe UI", 12))
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 13))

        self.tableau = ttk.Treeview(left_frame, columns=colonnes, show="headings", height=20)
        for col in colonnes:
            self.tableau.heading(col, text=col)
            self.tableau.column(col, width=500, anchor="w")
        self.tableau.pack(fill="both", expand=True)

        # --- Label chemin (sous tableau à gauche) ---
        label_chemin = ctk.CTkLabel(
            left_frame, text="",
            font=("Segoe UI", 13), text_color="white",
            wraplength=500, justify="center"
        )
        label_chemin.pack(pady=(15, 0))

        # --- Liste déroulante Année (pilote la regex) ---
        lbl_annee = ctk.CTkLabel(
            right_frame,
            text="Choisir l’année des fichiers profil à lister :",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_annee.pack(anchor="w", pady=(0, 4))

        years = ["Choisir..."] + [str(y) for y in range(2018, 2036)]

        # Label dédié pour le statut du fichier "Réel AAAA.xlsx"
        lbl_reel = ctk.CTkLabel(
            right_frame,
            text="",
            anchor="w",
            justify="left",
            text_color="white",
            font=("Segoe UI", 12)
        )
        lbl_reel.pack(anchor="w", pady=(8, 10))
        self.lbl_reel = lbl_reel  # pour y accéder ailleurs

        # Helper pour (dés)activer le bouton d'import si disponible
        def set_import_button_state(enabled: bool):
            try:
                self.bouton_valider.configure(state=("normal" if enabled else "disabled"))
            except Exception:
                pass
        self.set_import_button_state = set_import_button_state

        def on_year_change(value: str):
            # Choix "placeholder" -> reset propre
            if value == "Choisir...":
                self.baseline_year = None
                self.pattern_profil = None
                self.fichier_source = None
                self.profils_import = []

                # vider la table si un dossier était choisi
                if hasattr(self, "chemin_dossier") and self.chemin_dossier:
                    vider_tableau(silent=True, hard=False)

                # statut du Réel
                try:
                    self.lbl_reel.configure(
                        text="ℹ️ Sélectionnez une année pour détecter le fichier 'Réel AAAA.xlsx'.",
                        text_color="white"
                    )
                    self.update_idletasks()
                except Exception:
                    pass

                # désactiver le bouton import (si existant)
                self.set_import_button_state(False)
                return

            # maj de l'année + regex
            self.baseline_year = int(value)
            self.pattern_profil = self.build_pattern(self.baseline_year)

            # === Résoudre le fichier Réel AAAA.xlsx pour l'année choisie ===
            year = self.baseline_year
            chemin_reel = self.find_fichier_reel_exact(year, self.BASE_DONNEES_DIR)
            if not chemin_reel:
                chemin_reel = self.find_fichier_reel_flexible(year, self.BASE_DONNEES_DIR)

            self.fichier_source = chemin_reel  # peut rester None

            # Feedback utilisateur
            try:
                if self.fichier_source:
                    self.lbl_reel.configure(
                        text=f"✅ Fichier réel détecté pour {year} :\n{self.fichier_source}",
                        text_color="white"
                    )
                    self.set_import_button_state(True)
                else:
                    self.lbl_reel.configure(
                        text=(
                            f"⚠️ Aucun fichier 'Réel {year}.xlsx' trouvé dans :\n"
                            f"{self.BASE_DONNEES_DIR}\n"
                            "L’import ne sera pas lancé tant que ce fichier n’existe pas."
                        ),
                        text_color="#FFCC00"
                    )
                    self.set_import_button_state(False)
                self.update_idletasks()
            except Exception:
                pass

            # Rafraîchir la table si un dossier est déjà choisi
            if hasattr(self, "chemin_dossier") and self.chemin_dossier:
                vider_tableau(silent=True, hard=False)
                fichiers = self.chercher_profils_regex(self.chemin_dossier, self.pattern_profil)

                # MAJ liste interne de chemins complets
                self.profils_import = list(fichiers)

                for f in fichiers:
                    self.tableau.insert("", "end", values=(os.path.basename(f),))

        annee_menu = ctk.CTkOptionMenu(
            right_frame,
            values=years,
            command=on_year_change,
            width=200,
            height=45,
        )

        # Valeur initiale : "Choisir..." si baseline_year None, sinon baseline_year
        if self.baseline_year is None:
            annee_menu.set("Choisir...")
        else:
            annee_menu.set(str(self.baseline_year))
        annee_menu.pack(pady=(0, 12))

        # --- Boutons (droite) + logique ---
        buttons_frame = ctk.CTkFrame(right_frame, fg_color="#00122e")
        buttons_frame.pack(fill="x", padx=0, pady=0)

        buttons_frame.grid_columnconfigure(0, weight=1)   # libellés
        buttons_frame.grid_columnconfigure(1, weight=0)   # boutons

        def vider_tableau(silent: bool = False, hard: bool = False):
            """
            Vide la table et, si hard=True, réinitialise aussi les variables et éléments d'UI.
            - silent=True : pas de message utilisateur.
            - hard=True : supprime toutes les lignes, efface sélection, remet 'Choisir...' pour l'année,
                        remet baseline_year=None, pattern_profil=None, chemin_dossier=None,
                        vide label_chemin, et réinitialise la progress bar.
            """
            # 1) Vider les lignes du Treeview
            rows = self.tableau.get_children()
            nb = len(rows)
            for row in rows:
                self.tableau.delete(row)

            # 2) Effacer la sélection résiduelle
            try:
                self.tableau.selection_remove(self.tableau.selection())
            except Exception:
                pass

            # 3) Si hard reset : réinitialiser les variables et l’UI liée
            if hard:
                self.chemin_dossier = None
                self.baseline_year = None
                self.pattern_profil = None
                self.profils_import = []   # <-- on vide aussi la liste interne

                # Remettre la liste déroulante sur "Choisir..."
                try:
                    annee_menu.set("Choisir...")
                except Exception:
                    pass

                # Réinitialiser le label de chemin
                label_chemin.configure(text="")

                # Réinitialiser progress bar/label si existants
                try:
                    if hasattr(self, "progress_label") and self.progress_label is not None:
                        self.progress_label.configure(text="")
                        self.progress_label.pack_forget()
                    if hasattr(self, "progress_bar") and self.progress_bar is not None:
                        self.progress_bar.set(0)
                        self.progress_bar.pack_forget()
                except Exception:
                    pass

            # 4) Message utilisateur si non-silent
            if not silent:
                if nb == 0 and not hard:
                    label_chemin.configure(text="ℹ️ Le tableau était déjà vide.")
                elif hard:
                    label_chemin.configure(text="🧹 Réinitialisation complète effectuée (tableau + paramètres).")
                else:
                    label_chemin.configure(text=f"🧹 Tableau vidé ({nb} ligne(s) supprimée(s)).")

        def choisir_dossier():
            chemin = filedialog.askdirectory(title="Sélectionner un dossier où les profils sont stockés")
            if chemin:
                self.chemin_dossier = chemin
                label_chemin.configure(text=f"✅ Dossier choisi :\n{chemin}")

                # vider tableau (soft reset)
                vider_tableau(silent=True, hard=False)

                # On ne peut chercher que si une année est choisie
                if self.pattern_profil is None:
                    label_chemin.configure(text="ℹ️ Choisissez d’abord une année dans la liste à droite.")
                    return

                fichiers = self.chercher_profils_regex(chemin, self.pattern_profil)

                # stocke la liste interne de chemins complets
                self.profils_import = list(fichiers)

                for f in fichiers:
                    self.tableau.insert("", "end", values=(os.path.basename(f),))

        # === Ligne 1 : Parcourir ===
        lbl_parcourir = ctk.CTkLabel(
            buttons_frame,
            text="Sélectionner le dossier racine où se trouvent les profils (fichiers Excel).",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_parcourir.grid(row=0, column=0, sticky="w", padx=(0, 10), pady=(0, 6))

        bouton_parcourir = ctk.CTkButton(
            buttons_frame, text="📁 Parcourir...",
            command=choisir_dossier,
            width=200, height=45, corner_radius=20,
            fg_color="#005288", hover_color="#00396B",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_parcourir.grid(row=0, column=1, sticky="e", pady=(0, 6))

        # === Ligne 2 : Vider tableau (hard reset) ===
        lbl_nettoyer = ctk.CTkLabel(
            buttons_frame,
            text="Vider la liste et réinitialiser les paramètres (dossier + année + progression).",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_nettoyer.grid(row=1, column=0, sticky="w", padx=(0, 10), pady=(6, 6))

        bouton_nettoyer = ctk.CTkButton(
            buttons_frame, text="🧹 Vider tableau",
            command=lambda: vider_tableau(silent=False, hard=True),
            width=200, height=45, corner_radius=20,
            fg_color="#888888", hover_color="#555555",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_nettoyer.grid(row=1, column=1, sticky="e", pady=(6, 6))

        # === Ligne 3 : Supprimer élément(s) sélectionné(s) ===
        def supprimer_selection():
            """Supprime les lignes sélectionnées dans le Treeview ET les chemins associés dans self.profils_import."""
            selection = self.tableau.selection()
            if not selection:
                label_chemin.configure(text="ℹ️ Sélectionnez d’abord au moins un élément dans la liste à gauche.")
                return

            nb_suppr = 0

            for item_id in selection:
                # Récupérer le nom de fichier affiché dans la ligne
                values = self.tableau.item(item_id, "values")
                if not values:
                    continue
                nom_fichier = values[0]

                # Supprimer du Treeview
                try:
                    self.tableau.delete(item_id)
                    nb_suppr += 1
                except Exception:
                    pass

                # Supprimer de la liste interne des chemins complets
                chemins_a_supprimer = [
                    p for p in list(self.profils_import)
                    if os.path.basename(p) == nom_fichier
                ]
                for p in chemins_a_supprimer:
                    try:
                        self.profils_import.remove(p)
                    except ValueError:
                        pass

            # Nettoyer la sélection résiduelle
            try:
                self.tableau.selection_remove(self.tableau.selection())
            except Exception:
                pass

            label_chemin.configure(text=f"🗑️ {nb_suppr} élément(s) supprimé(s) de la liste (interface + chemins).")

        lbl_supprimer = ctk.CTkLabel(
            buttons_frame,
            text="Supprimer l’élément (ou les éléments) actuellement sélectionné(s) dans la liste.",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_supprimer.grid(row=2, column=0, sticky="w", padx=(0, 10), pady=(6, 6))

        bouton_supprimer = ctk.CTkButton(
            buttons_frame, text="🗑️ Supprimer sélection",
            command=supprimer_selection,
            width=200, height=45, corner_radius=20,
            fg_color="#B00020", hover_color="#8A001A",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_supprimer.grid(row=2, column=1, sticky="e", pady=(6, 6))

        # === Ligne 4 : Lancer import ===
        lbl_valider = ctk.CTkLabel(
            buttons_frame,
            text="Lancer l’import des profils trouvés (selon l’année choisie) vers les fichiers mensuels.",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_valider.grid(row=3, column=0, sticky="w", padx=(0, 10), pady=(6, 6))

        bouton_valider = ctk.CTkButton(
            buttons_frame, text="✅ Lancer import",
            command=self.lancer_import_profils,
            width=200, height=45, corner_radius=20,
            fg_color="#008C4B", hover_color="#006C39",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_valider.grid(row=3, column=1, sticky="e", pady=(6, 6))
        self.bouton_valider = bouton_valider  # pour set_import_button_state

        # === Ligne 5 : Retour menu ===
        lbl_retour = ctk.CTkLabel(
            buttons_frame,
            text="",
            anchor="w", justify="left", text_color="white", font=("Segoe UI", 12)
        )
        lbl_retour.grid(row=4, column=0, sticky="w", padx=(0, 10), pady=(6, 0))

        bouton_retour = ctk.CTkButton(
            buttons_frame, text="⬅️ Retour menu",
            command=self.creer_accueil,
            width=200, height=45, corner_radius=20,
            fg_color="#444",
            hover_color="#666",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_retour.grid(row=4, column=1, sticky="e", pady=(6, 0))

        # --- Barre de progression (masquée par défaut) ---
        self.progress_label = ctk.CTkLabel(self, text="", font=("Segoe UI", 13), text_color="white")
        self.progress_label.pack_forget()

        self.progress_bar = ctk.CTkProgressBar(self, width=400, height=20, progress_color="#00B050")
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()

    def lancer_import_profils(self):
        import pandas as pd
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
        import re
        from datetime import datetime, date
        import os
        from collections import defaultdict
        import traceback

        print("=== [START] lancer_import_profils ===")

        # ===================== PARAMÈTRES / RECHERCHE (réutilisés) =====================
        if not hasattr(self, "baseline_year"):
            self.baseline_year = 2025  # fallback si l'UI n'a pas été ouverte
            print(f"[INFO] baseline_year absent → fallback {self.baseline_year}")
        else:
            print(f"[INFO] baseline_year récupérée depuis l'UI : {self.baseline_year}")

        # Sécurité : build_pattern au cas où importer_les_profils n'aurait pas été appelé
        if not hasattr(self, "build_pattern"):
            print("[INIT] Création de la fonction build_pattern (sécurité)")
            def build_pattern(year: int | None):
                if year is None:
                    return None
                y = year
                y_prev = y - 1

                def y_or_date(yy: int) -> str:
                    return rf"{yy}(?:-\d{{2}}-\d{{2}})?"

                return re.compile(
                    rf'^Profil\s+Tr[ée]so\s+SNCF\s+'
                    rf'(?:'
                        rf'{y_prev}\s*[-–—]\s*{y_or_date(y)}'
                        rf'|'
                        rf'{y}\s*[-–—]\s*{y_or_date(y)}'
                    rf')\b.*\.xlsx$',
                    re.IGNORECASE
                )
            self.build_pattern = build_pattern

        if not hasattr(self, "pattern_profil"):
            self.pattern_profil = self.build_pattern(self.baseline_year)
            print("[INIT] pattern_profil (regex) compilé")

        if not hasattr(self, "chercher_profils_regex"):
            print("[INIT] Création de la fonction chercher_profils_regex (sécurité)")
            def _chercher_profils_regex(root_folder: str, pattern):
                resultats = []
                for dossier, _, fichiers in os.walk(root_folder):
                    for fichier in fichiers:
                        if pattern.match(fichier):
                            resultats.append(os.path.join(dossier, fichier))
                return resultats
            self.chercher_profils_regex = _chercher_profils_regex

        # ===================== LECTURE DATE COLONNE C =====================
        def lire_date_c_descendant(fichier, feuille="SA_SNCF", start_row=6, year=None, max_lookahead=50):
            """
            Parcourt C{start_row}..C{start_row+max_lookahead-1} pour trouver une date cohérente.
            Filtre sur 'year' si fourni.
            """
            try:
                wb = load_workbook(fichier, read_only=True, data_only=True)
                if feuille not in wb.sheetnames:
                    wb.close()
                    print(f"[WARN] Feuille '{feuille}' absente dans {os.path.basename(fichier)}")
                    return None, None
                ws = wb[feuille]

                def parse_cell_to_date(val):
                    if isinstance(val, datetime):
                        return val.date()
                    if isinstance(val, date):
                        return val
                    if isinstance(val, str):
                        s = val.strip()
                        try:
                            return datetime.strptime(s, "%d/%m/%Y").date()
                        except Exception:
                            pass
                        try:
                            return date.fromisoformat(s)
                        except Exception:
                            pass
                        m = re.search(r"\b(19\d{2}|20\d{2})\b", s)
                        if m:
                            return date(int(m.group(1)), 1, 1)
                    return None

                for r in range(start_row, start_row + max_lookahead):
                    val = ws[f"C{r}"].value
                    if val is None:
                        continue

                    d = parse_cell_to_date(val)
                    if not d:
                        print(f"[INFO] {os.path.basename(fichier)} C{r}: valeur non parsable → {val!r}")
                        continue

                    if year is not None and d.year != year:
                        if isinstance(val, str) and re.search(rf"\b{year}\b", val):
                            wb.close()
                            print(f"[OK] {os.path.basename(fichier)} C{r}: année {year} repérée dans le texte → {d.isoformat()}")
                            return d, r
                        print(f"[SKIP] {os.path.basename(fichier)} C{r}: année {d.year} ≠ {year}")
                        continue

                    wb.close()
                    print(f"[OK] {os.path.basename(fichier)} C{r} → {d.isoformat()}")
                    return d, r

                wb.close()
                print(f"[WARN] Aucune date (année={year}) trouvée en C{start_row}..C{start_row+max_lookahead-1} "
                    f"dans {os.path.basename(fichier)}")
                return None, None

            except Exception as e:
                print(f"[ERROR] Lecture colonne C échouée pour {os.path.basename(fichier)} : {e}")
                traceback.print_exc()
                return None, None

        def choisir_meilleure_version(fichiers):
            try:
                fichiers = sorted(fichiers)
                v3 = [f for f in fichiers if "V3" in os.path.basename(f)]
                if v3:
                    return v3[0]
                v2 = [f for f in fichiers if "V2" in os.path.basename(f)]
                if v2:
                    return v2[0]
                return fichiers[0] if fichiers else None
            except Exception as e:
                print(f"[ERROR] choisir_meilleure_version: {e}")
                traceback.print_exc()
                return fichiers[0] if fichiers else None

        # === Progress bar ===
        try:
            self.progress_label.configure(text="⏳ Export en cours (division par mois)...")
            self.progress_label.pack(pady=(10, 0))
            self.progress_bar.pack(pady=(5, 20))
            self.progress_bar.set(0)
            self.update_idletasks()
        except Exception:
            print("[WARN] UI progress bar indisponible dans ce contexte.")

        # === LISTE DES FICHIERS PROFILS À UTILISER ===
        # Priorité : self.profils_import (mise à jour par importer_les_profils + supprimer_selection)
        profils_import = getattr(self, "profils_import", [])

        if profils_import:
            fichiers_trouves = list(profils_import)
            print(f"[INFO] Utilisation de self.profils_import : {len(fichiers_trouves)} fichier(s).")
        else:
            # Fallback : on re-scanne le dossier avec la regex (cas où importer_les_profils n'a pas été utilisé)
            dossier_depart = getattr(self, "chemin_dossier", None)
            if not dossier_depart:
                print("[ERROR] Aucun dossier de profils sélectionné et self.profils_import est vide.")
                try:
                    self.progress_label.configure(text="⚠️ Aucun fichier profil sélectionné.")
                except Exception:
                    pass
                print("=== [STOP] lancer_import_profils (erreur aucun fichier) ===")
                return

            print(f"[INFO] Dossier de départ : {dossier_depart}")
            print("[STEP] Recherche des fichiers de profils par regex (fallback)...")
            fichiers_trouves = self.chercher_profils_regex(dossier_depart, self.pattern_profil)
            print(f"[OK] {len(fichiers_trouves)} fichier(s) trouvé(s) (fallback).")

        if not fichiers_trouves:
            print("[ERROR] Liste de fichiers profils vide après filtrage.")
            try:
                self.progress_label.configure(text="⚠️ Aucun fichier profil à traiter (liste vide).")
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (liste vide) ===")
            return

        # --- Récupération des dates par fichier ---
        print("[STEP] Lecture des dates (colonne C) en descendant depuis C6...")
        dates_par_fichier = []
        for f in fichiers_trouves:
            date_trouvee, row_found = lire_date_c_descendant(
                f, feuille="SA_SNCF", start_row=6, year=self.baseline_year, max_lookahead=50
            )
            if date_trouvee:
                dates_par_fichier.append((f, date_trouvee, row_found))
                print(f"   - {os.path.basename(f)} → date={date_trouvee.isoformat()} (C{row_found})")
            else:
                print(f"   - {os.path.basename(f)} → aucune date conforme trouvée")

        print("[STEP] Groupement des fichiers par date Cx...")
        groupes = defaultdict(list)
        for fichier, d, row_found in dates_par_fichier:
            groupes[d].append((fichier, row_found))

        print("[STEP] Sélection de la meilleure version (V3>V2>baseline) pour chaque date...")
        dates_uniques = []
        for d, lst in groupes.items():
            fichiers = [f for f, _r in lst]
            meilleur_fichier = choisir_meilleure_version(fichiers)
            if meilleur_fichier:
                row_found = next(r for (f, r) in lst if f == meilleur_fichier)
                dates_uniques.append((meilleur_fichier, d, row_found))
                print(f"   - {d.isoformat()} → {os.path.basename(meilleur_fichier)} (C{row_found})")

        dates_uniques.sort(key=lambda x: x[1])
        print("[OK] Dates uniques triées :", [(os.path.basename(f), d.isoformat()) for f, d, _r in dates_uniques])

        fichier_prev = [f for f, _d, _r in dates_uniques]
        rows_prev    = [r for _f, _d, r in dates_uniques]

        print("[STEP] Calcul des décalages (jours) entre prévisions successives...")
        deltas = []
        if len(dates_uniques) < 2:
            print("[WARN] Pas assez de dates pour calculer des décalages (0 ou 1).")
        else:
            for i in range(1, len(dates_uniques)):
                _, date1, _ = dates_uniques[i-1]
                _, date2, _ = dates_uniques[i]
                delta = (date2 - date1).days
                deltas.append(delta)
                print(f"   - Δ({date1.isoformat()} → {date2.isoformat()}) = {delta} jour(s)")

        cumul = [0]
        somme = 0
        for d in deltas:
            somme += d
            cumul.append(somme)
        print(f"[OK] Cumul des décalages : {cumul}")

        # === Résolution du fichier source (réel) depuis la sélection d'année ===
        fichier_source = getattr(self, "fichier_source", None)

        if not hasattr(self, "find_fichier_reel_exact") or not hasattr(self, "find_fichier_reel_flexible"):
            print("[ERROR] Helpers find_fichier_reel_* manquants (importer_les_profils non appelé ?)")
            # Tu peux éventuellement recréer ici des versions simplifiées si besoin.
            # Pour l’instant on stoppe proprement :
            try:
                self.progress_label.configure(text="⚠️ Helpers pour trouver 'Réel AAAA.xlsx' manquants.")
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (helpers manquants) ===")
            return

        if not fichier_source:
            base_dir = getattr(self, "BASE_DONNEES_DIR", None)
            if base_dir and self.baseline_year:
                fichier_source = self.find_fichier_reel_exact(self.baseline_year, base_dir) \
                                or self.find_fichier_reel_flexible(self.baseline_year, base_dir)

        if not fichier_source or not os.path.isfile(fichier_source):
            print("[ERROR] Fichier 'Réel AAAA.xlsx' introuvable : import annulé.")
            try:
                self.progress_label.configure(
                    text=(
                        f"⚠️ Fichier réel manquant pour {getattr(self, 'baseline_year', '???')}.\n"
                        f"Attendu : 'Réel {getattr(self, 'baseline_year', 'AAAA')}.xlsx' dans "
                        f"{getattr(self, 'BASE_DONNEES_DIR', '(dossier inconnu)')}\n"
                        f"→ Aucun import exécuté."
                    )
                )
                self.progress_label.pack(pady=(10, 0))
                self.progress_bar.pack_forget()
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (pas de fichier réel) ===")
            return

        print(f"[INFO] Fichier source (réel) : {fichier_source}")
        print(f"[INFO] {len(fichier_prev)} fichier(s) de prévision retenu(s).")

        # === Feuilles à traiter ===
        # === Feuilles à traiter ===
        sections = charger_sections_depuis_cells()

        if not sections:
            print("[WARN] Aucune section trouvée → traitement annulé.")
            return

        print(f"[INFO] Sections à traiter : {[s['dest'] for s in sections]}")

        # === Styles ===
        fill_jaune = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        fill_bleu = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
        align_center = Alignment(horizontal='center', vertical='center')
        font_bold = Font(bold=True)

        # === Libellés de prév JJ/MM ===
        print("[STEP] Extraction des libellés de dates pour les entêtes de prévisions...")
        dates_prev = []
        for path in fichier_prev:
            match = re.search(r"\d{4}-\d{2}-\d{2}", path)
            if match:
                date_obj = datetime.strptime(match.group(), "%Y-%m-%d")
                label = date_obj.strftime("%d/%m")
            else:
                label = "N/A"
            dates_prev.append(label)
            print(f"   - {os.path.basename(path)} → label '{label}'")
        nb_prev = len(fichier_prev)
        print(f"[OK] {nb_prev} label(s) de prévision.")

        # === Précharger les prévisions ===
        print("[STEP] Préchargement des matrices de prévisions par section...")
        valeurs_prev_all_by_section = {}
        for section in sections:
            feuille_prev = section["prev"]
            frames = []
            print(f"   > Chargement prévisions feuille '{feuille_prev}' sur {len(fichier_prev)} fichier(s)...")
            for idx_f, (f, row_found) in enumerate(zip(fichier_prev, rows_prev), start=1):
                try:
                    df_prev = pd.read_excel(f, sheet_name=feuille_prev, header=None, skiprows=row_found-1)
                    raw_noms = df_prev.iloc[0, 5:]
                    mask_valid = raw_noms.notna() & (raw_noms.astype(str).str.strip() != "")
                    noms_prev = raw_noms[mask_valid].astype(str).str.strip().reset_index(drop=True)
                    valeurs = df_prev.iloc[:, 5:].loc[:, mask_valid.values].reset_index(drop=True)
                    valeurs.columns = noms_prev
                    frames.append(valeurs.reset_index(drop=True))
                    print(f"      - [{idx_f}/{len(fichier_prev)}] {os.path.basename(f)}: "
                        f"skiprows={row_found-1}, shape={valeurs.shape}")
                except Exception as e:
                    print(f"[ERROR] Lecture prévisions '{feuille_prev}' échouée pour {os.path.basename(f)} : {e}")
                    traceback.print_exc()
                    frames.append(pd.DataFrame())
            valeurs_prev_all_by_section[feuille_prev] = frames

        # === Répertoire de sortie ===
        base_out_dir = FICHIER_EXCEL_DIR
        os.makedirs(base_out_dir, exist_ok=True)
        print(f"[INFO] Répertoire de sortie : {base_out_dir}")

        # === Boucle principale : split par mois ===
        total = len(sections)
        OVERWRITE = False
        print(f"[INFO] OVERWRITE = {OVERWRITE}")

        for idx_section, section in enumerate(sections, start=1):
            feuille_source = section["source"]
            feuille_prev = section["prev"]
            feuille_dest = section["dest"]
            print(f"\n[SECTION {idx_section}/{total}] '{feuille_dest}' — source='{feuille_source}', prev='{feuille_prev}'")

            # Lecture du réel
            print("   [STEP] Lecture du fichier source (réel)...")
            try:
                df = pd.read_excel(fichier_source, sheet_name=feuille_source, header=None, skiprows=4)
                dates = df.iloc[0, 4:].dropna().reset_index(drop=True)
                dates_ts = pd.to_datetime(dates, dayfirst=True)
                lignes_valides = df.iloc[4:, [2] + list(range(4, df.shape[1]))].dropna(subset=[2])
                noms = lignes_valides.iloc[:, 0].astype(str).str.strip().reset_index(drop=True)
                valeurs_reelles = lignes_valides.iloc[:, 1:].reset_index(drop=True)
                print(f"   [OK] Réel: {valeurs_reelles.shape[0]} lignes (flux) x {valeurs_reelles.shape[1]} dates")
            except Exception as e:
                print(f"   [ERROR] Lecture du réel échouée pour feuille '{feuille_source}' : {e}")
                traceback.print_exc()
                continue

            print("   [STEP] Constitution des groupes par (année, mois) ...")
            groupes_mois = defaultdict(list)
            for i, ts in enumerate(dates_ts):
                groupes_mois[(ts.year, ts.month)].append(i)
            print(f"   [OK] {len(groupes_mois)} mois détecté(s).")

            valeurs_prev_all = valeurs_prev_all_by_section[feuille_prev]

            for (year, month), idxs in sorted(groupes_mois.items()):
                print(f"   [MOIS] {year}-{month:02d} → {len(idxs)} jour(s) dans le mois")
                year_dir = os.path.join(base_out_dir, str(year))
                os.makedirs(year_dir, exist_ok=True)
                out_path = os.path.join(year_dir, f"Historique_prev_reel_filiales_{year}_{month:02d}.xlsx")
                print(f"      - Fichier cible: {out_path}")

                new_wb = False
                if os.path.exists(out_path):
                    try:
                        wb_out = load_workbook(out_path)
                        print("      - Workbook existant ouvert.")
                    except Exception as e:
                        print(f"      [WARN] Fichier existant illisible → recréation. Détail: {e}")
                        wb_out = Workbook()
                        if "Sheet" in wb_out.sheetnames:
                            wb_out.remove(wb_out["Sheet"])
                        for sec in sections:
                            if sec["dest"] not in wb_out.sheetnames:
                                wb_out.create_sheet(sec["dest"])
                        new_wb = True
                else:
                    wb_out = Workbook()
                    if "Sheet" in wb_out.sheetnames:
                        wb_out.remove(wb_out["Sheet"])
                    for sec in sections:
                        if sec["dest"] not in wb_out.sheetnames:
                            wb_out.create_sheet(sec["dest"])
                    new_wb = True
                    print("      - Nouveau workbook créé (toutes feuilles destination préparées).")

                if feuille_dest not in wb_out.sheetnames:
                    wb_out.create_sheet(feuille_dest)
                    print(f"      - Feuille '{feuille_dest}' créée.")
                ws = wb_out[feuille_dest]

                def _sheet_has_payload(_ws):
                    if _ws.max_row <= 3:
                        return False
                    for row in _ws.iter_rows(min_row=4, max_row=_ws.max_row, values_only=True):
                        if any(v is not None and v != "" for v in row):
                            return True
                    return False

                if not OVERWRITE and _sheet_has_payload(ws):
                    print("      - Données déjà présentes → SKIP (OVERWRITE=False).")
                else:
                    print("      - Écriture/rafraîchissement de la feuille...")
                    ws.delete_rows(1, ws.max_row)

                    sous_titres = ["Dates", "Réel (K€)"]
                    for date_str in dates_prev:
                        sous_titres.append(f"Prévision {date_str} (K€)")
                        sous_titres.append(f"Écart {date_str} (K€)")

                    start_col = 3
                    print(f"      - {len(noms)} flux à écrire. Largeur bloc par flux = 2 + 2*{len(dates_prev)} + 1 (séparation)")

                    for idx_flux, nom in enumerate(noms, start=1):
                        if idx_flux % 10 == 0 or idx_flux == 1 or idx_flux == len(noms):
                            print(f"         • Flux {idx_flux}/{len(noms)} : '{nom}'")

                        col1 = get_column_letter(start_col)
                        colN = get_column_letter(start_col - 1 + len(sous_titres))
                        ws.merge_cells(f"{col1}2:{colN}2")
                        titre_cell = ws[f"{col1}2"]
                        titre_cell.value = nom
                        titre_cell.fill = fill_jaune
                        titre_cell.font = font_bold
                        titre_cell.alignment = align_center
                        titre_cell.border = border

                        for i, titre in enumerate(sous_titres):
                            cell = ws.cell(row=3, column=start_col - 1 + i)
                            cell.value = titre
                            cell.fill = fill_bleu
                            cell.font = font_bold
                            cell.alignment = align_center
                            cell.border = border

                        for r, i_global in enumerate(idxs):
                            row = 4 + r

                            ws.cell(row=row, column=start_col - 1, value=dates_ts.iloc[i_global].date()).alignment = align_center
                            ws.cell(row=row, column=start_col - 1).border = border

                            valeur = valeurs_reelles.iloc[idx_flux-1, i_global]
                            val_k = round(valeur / 1000) if pd.notna(valeur) else None
                            ws.cell(row=row, column=start_col, value=val_k).alignment = align_center
                            ws.cell(row=row, column=start_col).border = border

                            for j, decal in enumerate(cumul):
                                prev_col = start_col + 1 + j * 2
                                ecart_col = start_col + 2 + j * 2

                                if j >= len(valeurs_prev_all):
                                    ws.cell(row=row, column=prev_col, value=None)
                                    ws.cell(row=row, column=ecart_col, value=None)
                                else:
                                    prev_data = valeurs_prev_all[j]
                                    idx_prev = i_global - decal
                                    if 0 <= idx_prev < len(prev_data):
                                        try:
                                            val = prev_data.iloc[idx_prev, idx_flux-1]
                                        except Exception:
                                            val = None
                                    else:
                                        val = None
                                    val_prev_k = round(val / 1000) if pd.notna(val) else None
                                    ecart = val_prev_k - val_k if (val_prev_k is not None and val_k is not None) else None

                                    ws.cell(row=row, column=prev_col, value=val_prev_k).alignment = align_center
                                    ws.cell(row=row, column=prev_col).border = border

                                    ws.cell(row=row, column=ecart_col, value=ecart).alignment = align_center
                                    ws.cell(row=row, column=ecart_col).border = border

                                    if r == 0:
                                        col_letter = get_column_letter(ecart_col)
                                        plage = f"{col_letter}4:{col_letter}{3 + len(idxs)}"
                                        rule_rouge = FormulaRule(formula=[f"${col_letter}4<0"],
                                                                font=Font(color="FF0000", bold=True))
                                        rule_vert = FormulaRule(formula=[f"${col_letter}4>0"],
                                                                font=Font(color="00B050", bold=True))
                                        ws.conditional_formatting.add(plage, rule_rouge)
                                        ws.conditional_formatting.add(plage, rule_vert)

                        start_col += 2 + 2 * len(dates_prev)
                        start_col += 1

                    print("      - Écriture terminée pour ce mois.")

                tmp_path = out_path + ".tmp"
                try:
                    wb_out.save(tmp_path)
                    os.replace(tmp_path, out_path)
                    print(f"      [SAVE] OK → {out_path}")
                except PermissionError as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"      [ERROR] PermissionError sur {out_path} (fichier ouvert ?) : {e}")
                except Exception as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"      [ERROR] Erreur pendant la sauvegarde de {out_path}: {e}")
                    traceback.print_exc()

            try:
                progress = idx_section / total
                self.progress_bar.set(progress)
                self.progress_label.configure(text=f"⏳ Export en cours... {int(progress*100)}%")
                self.update_idletasks()
            except Exception:
                pass

        try:
            self.progress_bar.set(1)
            self.progress_label.configure(text="✅ Export terminé (1 fichier par mois, rangé par année) !")
        except Exception:
            pass

        print("=== [END] lancer_import_profils — Export terminé ===")

#===================Page écarts détails + k-means===================
    def creer_fichier_excel(self):
        # === Demande de confirmation ===
        confirmer = messagebox.askyesno(
            "Confirmation",
            "Voulez-vous vraiment créer le fichier Excel ?\n"
            "Le fichier existant sera remplacé si présent."
        )
        if not confirmer:
            return  # Sort de la fonction si l'utilisateur clique sur Non

        # === Chemin du fichier à créer ===
        fichier_destination = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Résultats\Historique_prev_reel_filiales3.xlsx"

        # === Sections avec noms de feuilles ===
        sections = [
            {"source": "1_-_SA_SNCF", "prev": "SA_SNCF", "dest": "SNCF_SA"},
            {"source": "3_-_SA_VOYAGEURS", "prev": "SA_VOYAGEURS", "dest": "SA_VOYAGEURS"},
            {"source": "2_-_RESEAU", "prev": "RESEAU", "dest": "RESEAU"},
            {"source": "4_-_G&C", "prev": "G&C", "dest": "G&C"},
            {"source": "5_-_GEODIS", "prev": "GEODIS", "dest": "GEODIS"},
        ]

        # === Crée le dossier s'il n'existe pas ===
        dossier_resultats = os.path.dirname(fichier_destination)
        os.makedirs(dossier_resultats, exist_ok=True)

        # === Supprime le fichier s'il existe déjà ===
        if os.path.exists(fichier_destination):
            os.remove(fichier_destination)
            print(f"🗑️ Fichier existant supprimé : {fichier_destination}")

        # === Crée un nouveau classeur ===
        wb = Workbook()

        # Supprime la feuille par défaut
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crée toutes les feuilles définies dans sections
        for s in sections:
            wb.create_sheet(title=s["dest"])

        # Sauvegarde le fichier
        wb.save(fichier_destination)

        print(f"✅ Nouveau fichier Excel créé avec les feuilles : {[s['dest'] for s in sections]}")
         # === Message visuel à l'utilisateur ===
        messagebox.showinfo(
            "Fichier créé",
            f"Le fichier Excel a été créé avec succès.\n\nChemin :\n{fichier_destination}"
        )
   
    def afficher_ecarts(self):
        import customtkinter as ctk
        from tkinter import ttk
        import tkinter as tk
        from PIL import Image
        from customtkinter import CTkImage

        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - ÉCARTS IMPORTANTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === Colonnes du tableau ===
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (k€)", "Écart (%)"]

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral", "Créances CDP",
            "Placements", "CC financiers", "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]

        decaissements = [
            "Péages", "Charges de personnel", "ACE & Investissements"
        ]

        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital",
            "Sous total financier", "Free cash Flow", "Emprunts",
            "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cash flow de financement",
            "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
            "Sous total dépenses"
        ]

        # === Fonction pour convertir en flux ===
        def en_flux(values):
            values = [float(v) if v is not None else None for v in values]
            if not values or all(v is None for v in values):
                return values
            flux = [0 if values[0] is not None else None]
            for i in range(1, len(values)):
                v, v_prev = values[i], values[i - 1]
                flux.append(v - v_prev if v is not None and v_prev is not None else None)
            return flux

        # === Récupération et calcul des écarts ===
        ecarts_data = []
        repartition = {feuille: 0 for feuille in sections.values()}

        for feuille in sections.values():
            ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
            for nom, col_start in noms_colonnes:
                dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                for i, date in enumerate(dates):
                    if i >= len(reel) or reel[i] is None:
                        continue
                    for idx, prev_list in enumerate(previsions):
                        if i >= len(prev_list) or prev_list[i] is None:
                            continue

                        r = reel[i]
                        prev_val = prev_list[i]

                        if r == 0 and prev_val == 0:
                            continue
                        elif prev_val == 0:
                            prev_val = 1

                        ecart = (r - prev_val) / prev_val

                        if abs(ecart) >= 0.4:
                            profil_label = noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}"
                            repartition[feuille] += 1

                            ecarts_data.append((
                                date,                    # Date
                                profil_label,            # Profil
                                feuille,                 # Filiale / Feuille
                                nom,                     # Flux
                                round(reel[i], 2),       # Réel
                                round(prev_val, 2),      # Prévision
                                round(r - prev_val, 2),  # Écart k€
                                round(ecart * 100, 1)    # Écart %
                            ))

        # === PIE CHART : % d'écarts par rapport au total (console) ===
        feuilles = list(repartition.keys())
        total_ecarts = sum(repartition.values())

        valeurs = []
        for f in feuilles:
            if total_ecarts > 0:
                pourcentage = (repartition[f] / total_ecarts) * 100
                print(f"{pourcentage:.1f}%, nb écarts: {repartition[f]}, total écarts : {total_ecarts}")
            else:
                pourcentage = 0
            valeurs.append(pourcentage)

        ecarts_data.sort(key=lambda x: abs(x[7]), reverse=True)

        # === Frame filtres et boutons ===
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(padx=30, pady=(10, 0), fill="x")

        filtre_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        filtre_frame.pack(side="left", fill="x", expand=True)

        # === ✅ AJOUT : filtre Année
        annees_disponibles = sorted({d.year for (d, *_ ) in ecarts_data})
        annee_label = ctk.CTkLabel(filtre_frame, text="Année :", font=("Segoe UI", 11, "bold"))
        annee_label.pack(side="left", padx=(0, 2), pady=5)

        annee_combo_frame = ctk.CTkFrame(filtre_frame, fg_color="#0084ff", corner_radius=8)
        annee_combo_frame.pack(side="left", padx=(0, 8), pady=5)

        annee_combo = ttk.Combobox(
            annee_combo_frame,
            values=(["Toutes"] + [str(y) for y in annees_disponibles]),
            state="readonly", width=8, font=("Segoe UI", 11, "bold")
        )
        annee_combo.set(str(annees_disponibles[-1]) if annees_disponibles else "Toutes")
        annee_combo.pack(padx=5, pady=2, fill="x")

        # === Filtres existants ===
        colonnes_filtrables = ["Date", "Profil", "Filiales", "Flux"]
        filtres = {}
        valeurs_uniques = {col: set() for col in colonnes_filtrables}
        for row in ecarts_data:
            for i, col in enumerate(colonnes):
                if col in colonnes_filtrables:
                    valeurs_uniques[col].add(str(row[i]))

        for col in colonnes_filtrables:
            label = ctk.CTkLabel(filtre_frame, text=f"{col} :", font=("Segoe UI", 11, "bold"))
            label.pack(side="left", padx=(0, 2), pady=5)
            combo_frame = ctk.CTkFrame(filtre_frame, fg_color="#0084ff", corner_radius=8)
            combo_frame.pack(side="left", padx=(0, 8), pady=5)
            valeurs = ["Tous"] + sorted(valeurs_uniques[col])
            combo = ttk.Combobox(combo_frame, values=valeurs, state="readonly", width=15, font=("Segoe UI", 11, "bold"))
            combo.set("Tous")
            combo.pack(padx=5, pady=2, fill="x")
            filtres[col] = combo

        # === Boutons ===
        btn_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        btn_frame.pack(side="right")

        btn_retour = ctk.CTkButton(
            btn_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_retour.pack(side="top", pady=5)

        btn_export = ctk.CTkButton(
            btn_frame,
            text="📊 Exporter en Excel",
            command=lambda: self.exporter_ecarts_excel(ecarts_data),
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#0078D7",
            hover_color="#005A9E",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_export.pack(side="top", pady=5)

        btn_graphique = ctk.CTkButton(
            btn_frame,
            text="📈 Visualiser graphiquement",
            command=lambda: self.analyser_ecarts_ml(),
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#FC7100",
            hover_color="#6C4100",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_graphique.pack(side="top", pady=5)

        # === Treeview ===
        tree = ttk.Treeview(self, columns=colonnes, show="headings", height=25)
        for col in colonnes:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=130)
        tree.pack(pady=10, padx=30, fill="both", expand=True)

        tree.tag_configure("neg", foreground="red")
        tree.tag_configure("pos", foreground="green")

        # === Fonctions de formatage ===
        def format_milliers(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ")
                return str(val)
            except Exception:
                return str(val)

        def format_pourcentage(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ") + "%"
                return str(val)
            except Exception:
                return str(val)

        # === Favorabilité ===
        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                if prev_val >= 0:
                    return reel_val >= prev_val
                else:
                    return abs(reel_val) <= abs(prev_val)
            else:
                return reel_val >= prev_val

        # === Affichage dans le tree ===
        def afficher_donnees(donnees):
            for row in tree.get_children():
                tree.delete(row)
            for data in donnees:
                date_str = data[0].strftime("%Y-%m-%d")
                reel_str = format_milliers(data[4])
                prev_str = format_milliers(data[5])
                ecart_k_str = format_milliers(data[6])
                ecart_pct_str = format_pourcentage(data[7])

                flux_nom = data[3]
                reel_val, prev_val = data[4], data[5]
                favorable = est_favorable(flux_nom, reel_val, prev_val)
                tags = ("pos",) if favorable else ("neg",)

                tree.insert(
                    "",
                    "end",
                    values=(date_str, data[1], data[2], flux_nom, reel_str, prev_str, ecart_k_str, ecart_pct_str),
                    tags=tags
                )

        # Premier affichage (non filtré)
        afficher_donnees(ecarts_data)

        # === Filtrage (incluant Année) ===
        def appliquer_filtre(event=None):
            filtred = ecarts_data

            # 1) Filtre Année
            sel_annee = annee_combo.get()
            if sel_annee != "Toutes":
                try:
                    y = int(sel_annee)
                    filtred = [row for row in filtred if row[0].year == y]  # row[0] = Date
                except Exception:
                    pass

            # 2) Filtres existants
            for i, col in enumerate(colonnes):
                if col in filtres:
                    val = filtres[col].get()
                    if val != "Tous":
                        filtred = [row for row in filtred if str(row[i]) == val]

            afficher_donnees(filtred)

        # Bind des filtres
        for combo in filtres.values():
            combo.bind("<<ComboboxSelected>>", appliquer_filtre)
        annee_combo.bind("<<ComboboxSelected>>", appliquer_filtre)

        # Appliquer une première fois (pour prendre en compte l'année par défaut)
        appliquer_filtre()

    def analyser_ecarts_ml(self):
        import tkinter as tk
        import customtkinter as ctk
        import pandas as pd
        import numpy as np
        from sklearn.cluster import KMeans
        from sklearn.preprocessing import StandardScaler
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox, ttk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors
        import matplotlib.colors as mcolors
        import matplotlib.cm as cm
        import re
        import datetime as _dt

        # ---------------- Helpers ----------------
        def _to_number(x):
            if x is None: return None
            if isinstance(x, (int, float)): return float(x)
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
                if s in {"", "-", "—", "NA", "N/A"}: return None
                try: return float(s)
                except Exception: return None
            try: return float(x)
            except Exception: return None

        def _year_of(d):
            if d is None: return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        def _annees_disponibles_filiale(nom_filiale_ui: str):
            """Retourne l'union d'années disponibles sur la filiale (ou toutes)."""
            annees = set()
            feuilles = list(sections.values()) if nom_filiale_ui == "Toutes les filiales" else [nom_filiale_ui]
            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue
                for _nom_flux, col_start in noms_colonnes:
                    try:
                        dates, reel, prevs, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue
                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            annees.add(y)
            return sorted(annees)

        # ---------------- UI ----------------
        try:
            self.vider_fenetre()

            # Header
            header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
            header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

            titre_label = ctk.CTkLabel(
                header_frame,
                text="PROJET PULSE - ANALYSE ML DES ÉCARTS (2D: % vs Valorisation signée)",
                font=("Segoe UI Semibold", 26, "bold")
            )
            titre_label.pack(side="left", anchor="w")

            # Logo (optionnel)
            try:
                image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
                logo_img = Image.open(image_path)
                resized_logo = logo_img.resize((90, 40), Image.Resampling.LANCZOS)
                ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo)
                logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
                logo_label.image = ctk_logo
                logo_label.pack(side="right", anchor="e", padx=(10, 0))
            except Exception as e:
                print(f"Erreur chargement logo: {e}")

            ctk.CTkFrame(self, height=2, fg_color="#00aced").pack(side="top", fill="x", pady=(0, 15))

            # Container scrollable
            container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
            container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

            main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
            scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
            scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
            scrollable_frame.bind("<Configure>", lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")))
            main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            main_canvas.configure(yscrollcommand=scrollbar.set)
            main_canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # Contrôles
            control_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=15)
            control_frame.pack(padx=20, pady=10, fill="x")

            ttk.Label(control_frame, text="Filiale :", background="#00122e",
                    foreground="white", font=('Segoe UI', 12)).pack(side="left", padx=(0, 8))

            valeurs_combo = ["Toutes les filiales"] + list(sections.values())
            filiale_combo = ttk.Combobox(control_frame, values=valeurs_combo, state="readonly", width=35)
            filiale_combo.current(0)
            filiale_combo.pack(side="left", padx=(0, 15))

            ttk.Label(control_frame, text="Année :", background="#00122e",
                    foreground="white", font=('Segoe UI', 12)).pack(side="left", padx=(0, 8))

            annee_combo = ttk.Combobox(control_frame, values=["Toutes années"], state="readonly", width=18)
            annee_combo.current(0)
            annee_combo.pack(side="left", padx=(0, 15))

            btn_afficher = ctk.CTkButton(control_frame, text="Afficher l'analyse 2D (ML)",
                                        width=220, fg_color="#00aced", hover_color="#0099e6")
            btn_afficher.pack(side="left", padx=(0, 10))

            btn_retour_tableau = ctk.CTkButton(control_frame, text="Retour au tableau",
                                            command=self.afficher_ecarts, width=180,
                                            fg_color="#FC7100", hover_color="#6C4100")
            btn_retour_tableau.pack(side="left", padx=(0, 10))

            btn_retour_menu = ctk.CTkButton(control_frame, text="Retour au menu",
                                            command=self.retour_menu, width=160,
                                            fg_color="#444", hover_color="#666")
            btn_retour_menu.pack(side="left", padx=(0, 10))

            # Cadre graphique
            graph_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=15)
            graph_frame.pack(padx=20, pady=20, fill="both", expand=True)

            # ---------------- cœur : calcul + tracé ----------------
            def afficher_graphique():
                feuille_sel = filiale_combo.get() or "Toutes les filiales"
                val_annee = annee_combo.get()
                annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)

                feuilles_a_traiter = list(sections.values()) if feuille_sel == "Toutes les filiales" else [feuille_sel]

                # Collecte points 2D : (ecart %, valorisation k€)
                # et infos tooltip
                points = []  # dicts: {x_pct, y_k, filiale, date}
                for ws_feuille in feuilles_a_traiter:
                    try:
                        ws, noms_colonnes = charger_donnees(ws_feuille, taille_bloc)
                    except Exception:
                        continue

                    for nom, col_start in noms_colonnes:
                        try:
                            dates, reel, previsions, profils = extraire_valeurs(ws, col_start, nb_prev, annee=annee)
                        except Exception:
                            # si YEAR_INDEX pas présent pour cette année, retester sans filtre puis filtrer
                            try:
                                dates, reel, previsions, profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                            except Exception:
                                continue

                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue

                            r = _to_number(reel[i] if i < len(reel) else None)
                            if r is None:
                                continue

                            for prev_list in previsions:
                                pv = prev_list[i] if i < len(prev_list) else None
                                prev_val = _to_number(pv)
                                if prev_val is None:
                                    continue

                                # éviter divisions par zéro
                                denom = r if r != 0 else (prev_val if prev_val != 0 else None)
                                if denom is None:
                                    continue

                                # X = écart en %
                                try:
                                    ecart_pct = (prev_val - r) / denom * 100.0
                                except ZeroDivisionError:
                                    continue

                                # Y = valorisation signée (k€)
                                # On suppose valeurs en €; si déjà en k€, enlève la division par 1000.
                                y_k = (r - prev_val) / 1000.0

                                points.append({
                                    "x_pct": ecart_pct,
                                    "y_k": y_k,
                                    "filiale": ws_feuille,
                                    "date": d
                                })

                if not points:
                    messagebox.showinfo("Analyse ML", "Aucune donnée exploitable pour ces filtres.")
                    return

                df = pd.DataFrame(points)

                # Nettoyage frame graphique
                for widget in graph_frame.winfo_children():
                    widget.destroy()

                # Standardisation + KMeans 2D
                X = df[['x_pct', 'y_k']].values
                scaler = StandardScaler()
                Xs = scaler.fit_transform(X)

                # Choix n_clusters = 4 (assez lisible)
                kmeans = KMeans(n_clusters=4, random_state=42, n_init=20)
                labels = kmeans.fit_predict(Xs)
                df['cluster'] = labels
                centers_s = kmeans.cluster_centers_
                centers = scaler.inverse_transform(centers_s)

                # Palette
                palette = ["#FC7100", "#005BBB", "#00CC66", "#FF3B30"]
                colors = [palette[l % len(palette)] for l in labels]

                # Figure
                plt.style.use("seaborn-v0_8-darkgrid")
                fig, ax = plt.subplots(figsize=(11, 6))
                sc = ax.scatter(df['x_pct'], df['y_k'], c=colors, s=45, alpha=0.9, edgecolors='none')

                # Centroides
                ax.scatter(centers[:, 0], centers[:, 1], c="black", marker="X", s=160, label="Centroids")

                # Axes & titres
                titre = "Dispersion des écarts — 2D (% vs valorisation signée)"
                suffix = f" — {feuille_sel}" if feuille_sel != "Toutes les filiales" else " — Ensemble des filiales"
                if annee is not None:
                    suffix += f" — {annee}"
                ax.set_title(titre + suffix, fontsize=15, fontweight="bold")
                ax.set_xlabel("Écart (%)  —  (prev - réel) / (réf) × 100")
                ax.set_ylabel("Valorisation signée (r − prev) en k€")
                ax.grid(True, linestyle="--", alpha=0.4)

                # Légende clusters (couleurs)
                for i in range(4):
                    ax.scatter([], [], color=palette[i], label=f"Cluster {i+1}", s=70)
                ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

                # Tooltips
                cursor = mplcursors.cursor(sc, hover=True)
                @cursor.connect("add")
                def on_hover(sel):
                    idx = sel.index
                    fil = df.iloc[idx]["filiale"]
                    x = df.iloc[idx]["x_pct"]
                    y = df.iloc[idx]["y_k"]
                    d = df.iloc[idx]["date"]
                    dtxt = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)
                    sel.annotation.set_text(f"Filiale : {fil}\nDate : {dtxt}\nÉcart : {x:.2f} %\nValo : {y:.0f} k€")
                    sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

                # Intégration Tkinter
                canvas_fig = FigureCanvasTkAgg(fig, master=graph_frame)
                canvas_fig.draw()
                canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)

                toolbar_frame = tk.Frame(graph_frame)
                toolbar_frame.pack()
                toolbar = NavigationToolbar2Tk(canvas_fig, toolbar_frame)
                toolbar.update()
                plt.close(fig)

                # Summary cluster (count, mean %, mean k€)
                summary = (df.groupby('cluster')
                            .agg(count=('x_pct', 'count'),
                                mean_pct=('x_pct', 'mean'),
                                mean_k=('y_k', 'mean'),
                                sum_k=('y_k', 'sum'))
                            .round({'mean_pct': 2, 'mean_k': 1, 'sum_k': 0})
                        )

                table_frame = ctk.CTkFrame(graph_frame, fg_color="#00122e")
                table_frame.pack(pady=10, fill="x")

                ttk.Label(table_frame, text="Résumé des clusters :",
                        background="#00122e", foreground="white",
                        font=('Segoe UI Semibold', 13)
                        ).pack(anchor="w", padx=5, pady=(0,5))

                tree = ttk.Treeview(table_frame,
                                    columns=["Cluster", "Nb points", "Moy. écart (%)", "Moy. valo (k€)", "Somme valo (k€)"],
                                    show='headings', height=min(6, len(summary)))
                for col in ["Cluster", "Nb points", "Moy. écart (%)", "Moy. valo (k€)", "Somme valo (k€)"]:
                    tree.heading(col, text=col)
                    tree.column(col, width=150, anchor="center")
                for cl in sorted(summary.index):
                    tree.insert(
                        "", "end",
                        values=[
                            cl + 1,
                            int(summary.loc[cl, 'count']),
                            f"{summary.loc[cl, 'mean_pct']:.2f}",
                            f"{summary.loc[cl, 'mean_k']:.1f}",
                            f"{int(summary.loc[cl, 'sum_k']):,}".replace(",", " ")
                        ]
                    )
                # Totaux
                tree.insert(
                    "", "end",
                    values=[
                        "Total",
                        int(summary['count'].sum()),
                        f"{(df['x_pct'].mean() if len(df)>0 else 0):.2f}",
                        f"{(df['y_k'].mean() if len(df)>0 else 0):.1f}",
                        f"{int(df['y_k'].sum()):,}".replace(",", " ")
                    ],
                    tags=("total",)
                )
                tree.tag_configure("total", background="#444", foreground="white", font=('Segoe UI', 12, 'bold'))
                tree.pack(pady=5, padx=5, fill="x")

            # Binding : quand la filiale change, on recharge la liste d'années
            def _on_filiale_change(_evt=None):
                fil = filiale_combo.get() or "Toutes les filiales"
                annees = _annees_disponibles_filiale(fil)
                if annees:
                    annee_combo.config(values=(["Toutes années"] + [str(a) for a in annees]))
                    annee_combo.set(str(annees[-1]))  # dernière année dispo par défaut
                else:
                    annee_combo.config(values=["Toutes années"])
                    annee_combo.set("Toutes années")

            filiale_combo.bind("<<ComboboxSelected>>", _on_filiale_change)
            _on_filiale_change()  # init

            btn_afficher.configure(command=afficher_graphique)

        except Exception as e:
            messagebox.showerror("Erreur ML", f"Une erreur est survenue : {e}")

    def exporter_ecarts_excel(self, ecarts_data):
        # Création du fichier Excel
        wb_out = Workbook()
        ws = wb_out.active
        ws.title = "Écarts Importants"

        # Titres
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]

        # Style pour titres
        titre_font = Font(bold=True, color="000000")
        titre_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Ajout des titres
        for col_num, titre in enumerate(colonnes, 1):
            cell = ws.cell(row=1, column=col_num, value=titre)
            cell.font = titre_font
            cell.fill = titre_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajout des données
        for row_num, data in enumerate(ecarts_data, start=2):
            date_str = data[0].strftime("%Y-%m-%d")
            row_values = [date_str, data[1], data[2], data[3], data[4], data[5], f"{data[6]}%"]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Sauvegarde
        from tkinter import filedialog
        fichier_sortie = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Enregistrer le fichier Excel"
        )
        if fichier_sortie:
            wb_out.save(fichier_sortie)
            messagebox.showinfo("Export réussi", f"Fichier exporté : {fichier_sortie}")

#===================Page repartition (flux/ filiales/ Profils)===================
    def afficher_repartition(self):
        import customtkinter as ctk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from collections import defaultdict
        import mplcursors
        import numpy as np
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors
        import datetime as _dt  # NEW (année)

        # --- helpers: compatibilité des signatures (lecture) ---
        def _charger(feuille):
            """Wrapper tolérant: essaye charger_donnees(feuille) puis charger_donnees(feuille, taille_bloc)."""
            try:
                return charger_donnees(feuille)
            except TypeError:
                return charger_donnees(feuille, taille_bloc)

        def _extraire(ws, col_start, annee=None):  # NEW (année) -> même signature partout
            """Wrapper tolérant: essaye extraire_valeurs(ws, col_start[, nb_prev][, annee]).
            Si 'annee' n'est pas supportée en paramètre par extraire_valeurs, on filtre après coup."""
            # 1) Essais avec annee
            try:
                return extraire_valeurs(ws, col_start, nb_prev, annee=annee)
            except TypeError:
                try:
                    return extraire_valeurs(ws, col_start, annee=annee)
                except TypeError:
                    # 2) Sans annee -> on filtrera manuellement
                    try:
                        return extraire_valeurs(ws, col_start, nb_prev)
                    except TypeError:
                        return extraire_valeurs(ws, col_start)

        # --- utilitaires année ---  # NEW (année)
        def _year_of(d):
            """Renvoie l'année (int) si possible, sinon None."""
            if d is None:
                return None
            # datetime/date
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    pass
            # chaîne 'YYYY-MM-DD' ou 'DD/MM/YYYY' ou 'MM/YYYY' etc.
            if isinstance(d, str):
                s = d.strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(s, fmt).year
                    except Exception:
                        continue
                # extraction brute de 4 chiffres contigus
                import re
                m = re.search(r"(20\d{2}|19\d{2})", s)
                if m:
                    return int(m.group(1))
            return None

        def _collect_years(feuilles):
            """Parcourt toutes les feuilles/colonnes et collecte les années présentes."""
            years = set()
            for feuille in feuilles:
                ws, noms_colonnes = _charger(feuille)
                for _nom, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = _extraire(ws, col_start, annee=None)
                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            years.add(y)
            return sorted(years)

        # ================== UI ==================
        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - RÉPARTITION DES ÉCARTS PAR FILIALE",
            font=titre_font
        )
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE ...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="#00aced")
        barre.pack(side="top", fill="x", pady=(0, 15))

        # === FRAME PRINCIPALE SCROLLABLE AVEC MARGES ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg="#00122e")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=v_scrollbar.set)

        main_canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

        # =============== Sélecteur d'année (NEW) ===============
        feuilles_all = list(sections.values())  # on travaille sur toutes les filiales
        annees_dispo = _collect_years(feuilles_all)  # NEW (année)

        ctk.CTkLabel(
            scrollable_frame,
            text="Sélectionnez une année :",
            font=("Arial", 12, "bold"),
            text_color="white",
            fg_color="#00122e"
        ).pack(pady=(6, 2), anchor="center")

        annees_box = ttk.Combobox(
            scrollable_frame,
            state="readonly",
            width=20,
            values=(["Toutes années"] + [str(a) for a in annees_dispo])  # NEW (année)
        )
        # Valeur par défaut = dernière année dispo si existante
        if annees_dispo:
            annees_box.set(str(annees_dispo[-1]))
        else:
            annees_box.set("Toutes années")
        annees_box.pack(pady=(0, 12), anchor="center")

        # =============== Types de flux, logique "favorable" (inchangé) ===============
        encaissements = ["Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
                        "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"]
        decaissements = ["Péages", "Charges de personnel", "ACE & Investissements"]
        mixtes = ["Sous total Investissements nets et ACE", "Charges et produits financiers",
                "Dividendes reçus et versés", "Augmentations de capital",
                "Sous total financier", "Free cash Flow", "Emprunts",
                "Tirages Lignes CT", "Change", "Variation de collatéral",
                "Créances CDP", "Placements", "CC financiers",
                "Emprunts / Prêts - Groupe", "Cash flow de financement",
                "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
                "Sous total dépenses"]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                return (reel_val >= prev_val) if prev_val >= 0 else (abs(reel_val) <= abs(prev_val))
            else:
                return reel_val >= prev_val

        # Widgets pour graphes, afin de les détruire entre recalculs
        _graph_widgets = []

        def _clear_graphs():
            for w in _graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            _graph_widgets.clear()

        # ================== Recalcul complet selon l'année ==================  # NEW (année)
        def recalcul(annee: int | None):
            _clear_graphs()
        
            # === CALCUL DES ÉCARTS ET VALORISATION (filtrés par année) ===
            repartition = {feuille: 0 for feuille in feuilles_all}
            valorisation_ecarts = {feuille: 0 for feuille in feuilles_all}
            details_ecarts = {feuille: [] for feuille in feuilles_all}

            for feuille in feuilles_all:
                ws, noms_colonnes = _charger(feuille)
                for nom, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = _extraire(ws, col_start, annee=None)  # on lit tout...
                    # ...puis filtre manuel si extraire_valeurs ne gère pas annee
                    # (si elle la gère déjà, les dates hors année seront probablement absentes mais ce filtre ne gêne pas)
                    for i, date in enumerate(dates):
                        y = _year_of(date)
                        if annee is not None and y is not None and y != annee:
                            continue

                        if i >= len(reel) or reel[i] is None:
                            continue
                        r = reel[i]
                        for idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list) or prev_list[i] is None:
                                continue
                            prev_val = prev_list[i]

                            # éviter division par zéro
                            if r == 0 and prev_val == 0:
                                continue
                            elif prev_val == 0:
                                prev_val = 1

                            ecart = (r - prev_val) / prev_val  # signe: positif si prev > reel
                            if abs(ecart) >= 0.4:
                                repartition[feuille] += 1
                                valorisation_ecarts[feuille] += (r - prev_val)  # valorisation signée en k€
                                details_ecarts[feuille].append({
                                    "date": date,
                                    "profil": noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}",
                                    "filiale": feuille,
                                    "flux": nom,
                                    "reel": round(reel[i], 2),
                                    "prevision": round(prev_val, 2),
                                    "ecart_pct": round(ecart * 100, 1)
                                })

            # === PIE CHART : % d'écarts par rapport au total ===
            feuilles = feuilles_all
            total_ecarts = sum(repartition.values())
            valeurs = [(repartition[f] / total_ecarts) * 100 if total_ecarts > 0 else 0 for f in feuilles]

            feuilles_filtrees = [f for i, f in enumerate(feuilles) if valeurs[i] > 0]
            valeurs_filtrees = [v for v in valeurs if v > 0]

            fig, ax = plt.subplots(figsize=(8, 6), facecolor="#00122e")
            ax.set_facecolor("#00122e")
            palette_pie = ["#5D5F83", "#34495E", "#5D6D7E", "#85929E", "#AAB7B8", "#D5DBDB"]

            if not valeurs_filtrees:
                ax.text(0.5, 0.5, "Aucun écart important détecté", ha='center', va='center', fontsize=12, color="white")
                ax.axis('off')
                wedges = []
            else:
                wedges, texts, autotexts = ax.pie(
                    valeurs_filtrees,
                    labels=feuilles_filtrees,
                    autopct='%1.1f%%',
                    startangle=140,
                    colors=palette_pie[:len(valeurs_filtrees)],
                    textprops={'color': 'white', 'fontsize': 10}
                )
                for t in texts + autotexts:
                    t.set_color("white")

            titre_suffix = "" if annee is None else f" – Année {annee}"
            ax.set_title(
                "Répartition des écarts significatifs par filiale – en nombre d’occurrences" + titre_suffix,
                fontsize=14,
                color="white"
            )

            canvas_fig = FigureCanvasTkAgg(fig, master=scrollable_frame)
            canvas_fig.draw()
            w1 = canvas_fig.get_tk_widget()
            w1.pack(pady=20, padx=30, fill="both", expand=True)
            _graph_widgets.append(w1)

            # === BAR CHART : % d'écarts par filiale (nb écarts / nb prévisions non-None filtrées par année) ===
            pourcentage_ecarts_filiales = {}
            nombre_previsions_par_feuille = {}

            for feuille in feuilles_all:
                ecarts = details_ecarts.get(feuille, [])
                nombre_ecarts = len(ecarts)

                ws, noms_colonnes_local = _charger(feuille)
                nombre_previsions = 0
                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = _extraire(ws, col_start, annee=None)
                    for i, d in enumerate(dates_local):
                        y = _year_of(d)
                        if annee is not None and y is not None and y != annee:
                            continue
                        for prev_list in previsions_local:
                            if i < len(prev_list) and prev_list[i] is not None:
                                nombre_previsions += 1

                nombre_previsions_par_feuille[feuille] = nombre_previsions
                pourcentage = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0
                pourcentage_ecarts_filiales[feuille] = pourcentage

            feuilles_bar = list(pourcentage_ecarts_filiales.keys())
            valeurs_bar = list(pourcentage_ecarts_filiales.values())

            fig_bar, ax_bar = plt.subplots(figsize=(10, 5), facecolor="#00122e")
            ax_bar.set_facecolor("#00122e")

            if valeurs_bar:
                vmin = min(valeurs_bar)
                vmax = max(valeurs_bar) if max(valeurs_bar) != vmin else vmin + 1
                norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
                cmap = cm.Blues
                colors_bar = [cmap(norm(v)) for v in valeurs_bar]
            else:
                colors_bar = []

            bars = ax_bar.bar(feuilles_bar, valeurs_bar, color=colors_bar, alpha=0.9)

            ax_bar.set_title("Fréquence des écarts significatifs par filiale" + titre_suffix,
                            fontsize=14, color="white")
            ax_bar.set_ylabel("% Écarts", color="white")
            ax_bar.tick_params(axis='x', rotation=45, colors="white")
            ax_bar.tick_params(axis='y', colors="white")
            ax_bar.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig_bar.tight_layout(pad=2.0)

            cursor_bar = mplcursors.cursor(bars, hover=True)
            @cursor_bar.connect("add")
            def on_hover_bar(sel):
                idx = sel.index
                feuille = feuilles_bar[idx]
                sel.annotation.set_text(
                    f"{feuille}\n"
                    f"Nombre d'écarts: {len(details_ecarts.get(feuille, []))}\n"
                    f"Nombre de Prévisions: {nombre_previsions_par_feuille.get(feuille, 0)}\n"
                    f"Pourcentage: {valeurs_bar[idx]:.1f}%"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

            canvas_bar = FigureCanvasTkAgg(fig_bar, master=scrollable_frame)
            canvas_bar.draw()
            w2 = canvas_bar.get_tk_widget()
            w2.pack(pady=20, padx=30, fill="both", expand=True)
            _graph_widgets.append(w2)

            # === BAR CHART : valorisation signée ===
            graph_valorisation_frame = tk.Frame(scrollable_frame, bg="#00122e")
            graph_valorisation_frame.pack(fill="both", expand=True, padx=10, pady=(10, 20))
            _graph_widgets.append(graph_valorisation_frame)

            feuilles_val = list(valorisation_ecarts.keys())
            valeurs_val = [0 if valorisation_ecarts[f] is None else valorisation_ecarts[f] for f in feuilles_val]

            fig_val, ax_val = plt.subplots(figsize=(10, 5), facecolor="#00122e")
            ax_val.set_facecolor("#00122e")

            if any(v != 0 for v in valeurs_val):
                max_abs = max(abs(v) for v in valeurs_val) or 1
                norm_val = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
                cmap_val = cm.RdBu_r
                colors_val = [cmap_val(norm_val(v)) for v in valeurs_val]
            else:
                colors_val = ["#8395a7"] * len(valeurs_val)

            bars_val = ax_val.bar(feuilles_val, valeurs_val, color=colors_val, alpha=0.95)

            ax_val.axhline(0, color="white", linewidth=1, alpha=0.7)
            ax_val.set_title("Valorisation totale des écarts pour chaque filiale (k€)" + titre_suffix,
                            fontsize=14, color="white")
            ax_val.set_ylabel("Écarts cumulés (k€)", color="white")
            ax_val.tick_params(axis='x', rotation=45, colors="white")
            ax_val.tick_params(axis='y', colors="white")
            ax_val.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig_val.tight_layout(pad=2.0)

            cursor_val = mplcursors.cursor(bars_val, hover=True)
            @cursor_val.connect("add")
            def on_add(sel):
                idx = sel.index
                feuille = feuilles_val[idx]
                val = valeurs_val[idx]
                nb_ecarts = repartition.get(feuille, 0)
                # recompute nb_prevs déjà fait plus haut
                nb_prevs = nombre_previsions_par_feuille.get(feuille, 0)
                val_str = f"{int(val):,}".replace(",", " ")
                sel.annotation.set_text(
                    f"{feuille}\nÉcarts: {nb_ecarts}\nPrévisions: {nb_prevs}\nValorisation: {val_str} k€"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            canvas_val = FigureCanvasTkAgg(fig_val, master=graph_valorisation_frame)
            canvas_val.draw()
            w3 = canvas_val.get_tk_widget()
            w3.pack(pady=10, fill="both", expand=True)
            _graph_widgets.append(w3)

            # === HOVER & CLIC SUR PIE (détails, filtrés par année) ===
            if wedges:
                original_colors = [w.get_facecolor() for w in wedges]

                def reset_colors():
                    for i, w in enumerate(wedges):
                        w.set_facecolor(original_colors[i])
                    canvas_fig.draw_idle()

                def on_hover(event):
                    if event.inaxes == ax and wedges:
                        found = False
                        for i, w in enumerate(wedges):
                            if w.contains_point((event.x, event.y)):
                                w.set_facecolor("blue")
                                for j, w2 in enumerate(wedges):
                                    if j != i:
                                        w2.set_facecolor(original_colors[j])
                                canvas_fig.draw_idle()
                                found = True
                                break
                        if not found:
                            reset_colors()

                def afficher_details_feuille(feuille):
                    if hasattr(self, 'frame_details') and self.frame_details is not None:
                        self.frame_details.destroy()

                    self.frame_details = tk.Frame(scrollable_frame, bg="#00122e")
                    self.frame_details.pack(fill='x', padx=30, pady=(10, 5))
                    _graph_widgets.append(self.frame_details)

                    ecarts = details_ecarts.get(feuille, [])
                    nombre_ecarts = len(ecarts)

                    # recompute nombre_previsions pour cette feuille (filtré par année)
                    ws, noms_colonnes_local = _charger(feuille)
                    nombre_previsions = 0
                    for nom_colonne, col_start in noms_colonnes_local:
                        dates_local, reel_local, previsions_local, noms_profils_local = _extraire(ws, col_start, annee=None)
                        for i, d in enumerate(dates_local):
                            y = _year_of(d)
                            if annee is not None and y is not None and y != annee:
                                continue
                            for prev_list in previsions_local:
                                if i < len(prev_list) and prev_list[i] is not None:
                                    nombre_previsions += 1

                    pourcentage_ecarts = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0

                    ctk.CTkLabel(
                        self.frame_details,
                        text=(f"Détails pour {feuille} : "
                            f"{nombre_ecarts} écarts sur {nombre_previsions} prévisions "
                            f"({pourcentage_ecarts:.1f}%)" + titre_suffix),
                        font=("Arial", 12, "bold")
                    ).pack(anchor='w', pady=(0, 10))

                    colonnes = ["Date", "Profil", "Filiale", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]
                    tree = ttk.Treeview(self.frame_details, columns=colonnes, show="headings", height=5)
                    for col in colonnes:
                        tree.heading(col, text=col)
                        tree.column(col, anchor="center", width=120)
                    tree.pack(pady=10, fill="x")

                    cinq_plus_gros = sorted(ecarts, key=lambda x: abs(x["ecart_pct"]), reverse=True)[:5]
                    for e in cinq_plus_gros:
                        date_str = e["date"].strftime("%Y-%m-%d") if hasattr(e["date"], 'strftime') else str(e["date"])
                        ecart_str = f"{e['ecart_pct']}%"
                        tags = ("pos",) if est_favorable(e["flux"], e["reel"], e["prevision"]) else ("neg",)
                        tree.insert("", "end", values=(date_str, e["profil"], e["filiale"], e["flux"],
                                                    e["reel"], e["prevision"], ecart_str), tags=tags)
                    tree.tag_configure("neg", foreground="red")
                    tree.tag_configure("pos", foreground="green")

                    self.update_idletasks()
                    try:
                        main_canvas.yview_moveto(self.frame_details.winfo_y() / max(1, scrollable_frame.winfo_height()))
                    except Exception:
                        pass

                def on_click(event):
                    if event.inaxes == ax and wedges:
                        for i, w in enumerate(wedges):
                            if w.contains_point((event.x, event.y)):
                                afficher_details_feuille(feuilles_filtrees[i])
                                break

                canvas_fig.mpl_connect("motion_notify_event", on_hover)
                canvas_fig.mpl_connect("button_press_event", on_click)

        # Lancement initial
        def _parse_annee_box():
            val = annees_box.get()
            return None if (not val or val == "Toutes années") else int(val)

        recalcul(_parse_annee_box())

        # Binding de la combobox année
        def _on_annee_change(_evt=None):
            recalcul(_parse_annee_box())

        annees_box.bind("<<ComboboxSelected>>", _on_annee_change)

        # === Bouton retour ===
        bouton_retour = ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=220,
            height=40,
            corner_radius=12,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_retour.pack(pady=20)
        
        # ===== Scroll molette (Canvas) =====
        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units"); return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units"); return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

    def afficher_repartition_par_prevision(self):
        from collections import defaultdict
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors
        import re

        # ===== RESET PAGE =====
        self.vider_fenetre()

        # ===== HEADER (titre + logo) =====
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - RÉPARTITION PAR PROFIL", font=titre_font)\
            .pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) ...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_height = 40
            new_width = int(new_height * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ===== CONTAINER (canvas + scrollbars) =====
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        canvas_container = tk.Frame(container, bg="#00122e")
        canvas_container.pack(fill="both", expand=True)

        main_canvas = tk.Canvas(canvas_container, bg="#00122e", highlightthickness=0)
        v_scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=main_canvas.yview)
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)

        scrollable_frame = tk.Frame(main_canvas, bg="#00122e")
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(event):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        main_canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        # ===== Titre =====
        ctk.CTkLabel(
            scrollable_frame,
            text="Répartition des Écarts Importants par Profil (en % et k€)",
            font=("Arial", 18, "bold"),
            text_color="white",
            fg_color="#00122e"
        ).pack(pady=15)

        # ===== Sélecteurs (filiale + année) =====
        ctk.CTkLabel(scrollable_frame, text="Sélectionnez une filiale :", font=("Arial", 12, "bold"),
                    text_color="white", fg_color="#00122e").pack(pady=(10, 2), anchor="center")
        filiales = ["Toutes filiales"] + list(sections.values())
        selected_filiale = tk.StringVar(value=filiales[0])
        select_box = ttk.Combobox(scrollable_frame, textvariable=selected_filiale, values=filiales,
                                state="readonly", width=25)
        select_box.pack(pady=(0, 8), anchor="center")

        ctk.CTkLabel(scrollable_frame, text="Sélectionnez une année :", font=("Arial", 12, "bold"),
                    text_color="white", fg_color="#00122e").pack(pady=(6, 2), anchor="center")
        annees_var = tk.StringVar(value="Toutes années")
        annees_box = ttk.Combobox(scrollable_frame, textvariable=annees_var,
                                values=["Toutes années"], state="readonly", width=20)
        annees_box.pack(pady=(0, 12), anchor="center")

        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)
        # ===== Zones graphique et tableau =====
        graph_frame = tk.Frame(scrollable_frame, bg="#00122e")
        graph_frame.pack(fill="both", expand=True, padx=10, pady=(5, 20))

        table_frame = tk.Frame(scrollable_frame, bg="#00122e")
        table_frame.pack(fill="x", padx=10, pady=(10, 30))

        colonnes = ("Profil", "Nb prévisions", "Nb écarts >=40%", "Taux (%)", "Valorisation (k€)")
        table = ttk.Treeview(table_frame, columns=colonnes, show="headings", height=6)
        for col in colonnes:
            table.heading(col, text=col)
            table.column(col, anchor="center", width=160)
        table.pack(pady=5, fill="x", expand=True)

        # ===== Helpers =====
        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]
            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _tok in noms_colonnes:
                    try:
                        years = self._annees_disponibles(feuille, nom_flux)
                        annees.update(years)
                    except Exception:
                        pass
            return sorted(annees)


        # ===== Fonction principale =====
        # ===== Fonction principale =====
        # ===== Fonction principale : REP PAR PREV + TRI JJ/MM =====
        def maj_graphique(filiale, annee):
            from collections import defaultdict
            import re
            import mplcursors
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            # Nettoyer les anciens graphes
            for w in graph_frame.winfo_children():
                w.destroy()

            # ---------- Helpers locaux ----------
            def _year_of(d):
                if d is None:
                    return None
                if hasattr(d, "year"):
                    try:
                        return int(d.year)
                    except Exception:
                        return None
                if isinstance(d, str):
                    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                        try:
                            import datetime as _dt
                            return _dt.datetime.strptime(d, fmt).year
                        except Exception:
                            pass
                    m = re.search(r"(20\d{2}|19\d{2})", d)
                    if m:
                        return int(m.group(1))
                return None

            def _to_number(x):
                if x is None: return None
                if isinstance(x, str):
                    s = x.strip().replace("\xa0", " ").replace(" ", "")
                    if s in {"", "-", "—", "NA", "N/A"}: return None
                    s = s.replace(",", ".")
                    try: return float(s)
                    except Exception: return None
                try: return float(x)
                except Exception: return None

            def _is_filled(x):
                return _to_number(x) is not None

            def _parse_jj_mm(nom: str):
                m = re.search(r'(\d{1,2})[/-](\d{1,2})', str(nom))
                if not m:
                    return 99, 99
                jj = int(m.group(1)); mm = int(m.group(2))
                if 1 <= jj <= 31 and 1 <= mm <= 12:
                    return mm, jj
                return 99, 99

            # ---------- Filiales à parcourir ----------
            filiales_calc = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            # ================= PASS 1 =================
            # Numérateur : nombre d'écarts >= 40% + valorisation, par profil, filtrés par année
            compteur_ecarts = defaultdict(int)
            valorisation_ecarts = defaultdict(float)

            for f in filiales_calc:
                ws, noms_colonnes = charger_donnees(f, taille_bloc)
                for nom_colonne, col_start in noms_colonnes:
                    # on lit TOUT puis on filtre nous-mêmes par année (robuste si extraire_valeurs ignore annee)
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    for p_idx, nom_profil in enumerate(noms_profils):
                        if p_idx >= len(previsions):
                            continue
                        prev_list = previsions[p_idx]

                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue

                            pv_raw = prev_list[i] if i < len(prev_list) else None
                            prev_val = _to_number(pv_raw)
                            r_val   = _to_number(reel[i] if i < len(reel) else None)

                            if prev_val is None:
                                continue
                            if r_val is None:
                                r_val = 0.0

                            if prev_val == 0:
                                if r_val == 0:
                                    continue
                                prev_val = 1.0  # convention anti-division par 0 côté prev

                            ecart = (r_val - prev_val) / prev_val
                            if abs(ecart) >= 0.4:
                                compteur_ecarts[nom_profil] += 1
                                valorisation_ecarts[nom_profil] += abs((prev_val or 0.0) - (r_val or 0.0))

            # ================= PASS 2 =================
            # Dénominateur : nb de prévisions non vides alignées sur dates et filtrées par année, par profil
            compteur_total = defaultdict(int)

            for f in filiales_calc:
                ws, noms_colonnes_local = charger_donnees(f, taille_bloc)
                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    for p_idx, nom_profil in enumerate(noms_profils_local):
                        if p_idx >= len(previsions_local):
                            continue
                        prev_list = previsions_local[p_idx]

                        nb_prev_non_vides = 0
                        for i, d in enumerate(dates_local):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue
                            if i < len(prev_list) and _is_filled(prev_list[i]):
                                nb_prev_non_vides += 1
                        compteur_total[nom_profil] += nb_prev_non_vides

            # ---------- Préparation + TRI + FILTRAGE PROFILS SANS DONNÉES ----------
            rows = []
            all_profils = set(compteur_total.keys()) | set(compteur_ecarts.keys())

            for nom in all_profils:
                total = compteur_total.get(nom, 0)
                ecarts = compteur_ecarts.get(nom, 0)

                # ⚠️ Masquer les profils sans aucune donnée sur l'année choisie
                if annee is not None and total == 0 and ecarts == 0:
                    continue

                taux = (ecarts / total * 100) if total > 0 else 0.0
                valo = valorisation_ecarts.get(nom, 0.0)
                mm, jj = _parse_jj_mm(nom)
                rows.append((mm, jj, nom, taux, valo, total, ecarts))

            rows.sort(key=lambda r: (r[0], r[1], str(r[2]).casefold()))

            # Séquences ordonnées pour graphes/table
            noms_final        = [r[2] for r in rows]
            pourcentages      = [r[3] for r in rows]
            valorisations     = [r[4] for r in rows]
            totaux_prev       = [r[5] for r in rows]
            totaux_ecarts     = [r[6] for r in rows]

            suffix = "" if annee is None else f" - {annee}"

            # ---------- FIGURE 1 : Taux d'écarts ----------
            fig1, ax1 = plt.subplots(figsize=(14, 5), facecolor="#00122e")
            ax1.set_facecolor("#00122e")
            if noms_final:
                bars1 = ax1.bar(noms_final, pourcentages, color="#1f77b4", alpha=0.9)
                ax1.plot(noms_final, pourcentages, color="white", marker="o", linewidth=2)
                cursor = mplcursors.cursor(bars1, hover=True)
                @cursor.connect("add")
                def on_add(sel):
                    idx = sel.index
                    sel.annotation.set_text(f"Profil : {noms_final[idx]}\nTaux : {pourcentages[idx]:.2f}%")
                    sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

            titre1 = ("Toutes filiales" if filiale == "Toutes filiales" else filiale) + f" - Taux d'Écarts (rep/prev){suffix}"
            ax1.set_title(titre1, fontsize=14, color="white")
            ax1.set_ylabel("Taux d'écarts (%)", color="white")
            ax1.tick_params(axis='y', colors="white")
            ax1.tick_params(axis='x', rotation=60, labelcolor="white")
            ax1.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig1.tight_layout(pad=2.0)

            canvas1 = FigureCanvasTkAgg(fig1, master=graph_frame)
            canvas1.draw()
            canvas1.get_tk_widget().pack(pady=(10, 20), fill="both", expand=True)

            # ---------- FIGURE 2 : Valorisation des écarts ----------
            fig2, ax2 = plt.subplots(figsize=(14, 5), facecolor="#00122e")
            ax2.set_facecolor("#00122e")
            if noms_final:
                ax2.bar(noms_final, valorisations, color="#28B463", alpha=0.8)

            titre2 = ("Toutes filiales" if filiale == "Toutes filiales" else filiale) + f" - Valorisation des écarts (rep/prev){suffix}"
            ax2.set_title(titre2, fontsize=14, color="white")
            ax2.set_ylabel("Valorisation (k€)", color="white")
            ax2.tick_params(axis='y', colors="white")
            ax2.tick_params(axis='x', rotation=60, labelcolor="white")
            ax2.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig2.tight_layout(pad=2.0)

            canvas2 = FigureCanvasTkAgg(fig2, master=graph_frame)
            canvas2.draw()
            canvas2.get_tk_widget().pack(pady=(10, 40), fill="both", expand=True)

            # ---------- Tableau ----------
            for row in table.get_children():
                table.delete(row)

            total_valorisation = 0.0
            for nom, taux, valo, tot, ec in zip(noms_final, pourcentages, valorisations, totaux_prev, totaux_ecarts):
                total_valorisation += valo
                table.insert(
                    "",
                    "end",
                    values=(
                        nom,
                        tot,
                        ec,
                        f"{taux:.2f}%",
                        f"{valo:,.0f}".replace(",", " ")
                    )
                )

            total_previsions = sum(totaux_prev)
            total_ecarts = sum(totaux_ecarts)
            taux_total = (total_ecarts / total_previsions * 100) if total_previsions > 0 else 0.0
            table.insert(
                "",
                "end",
                values=(
                    "TOTAL",
                    total_previsions,
                    total_ecarts,
                    f"{taux_total:.2f}%",
                    f"{total_valorisation:,.0f}".replace(",", " ")
                )
            )


        # ===== Callbacks sélection =====
        def _on_filiale_change(*_):
            filiale = selected_filiale.get()
            annees = _annees_pour_filiale(filiale)
            annees_box.config(values=(["Toutes années"] + [str(a) for a in annees]))
            annees_var.set(str(annees[-1]) if annees else "Toutes années")
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            maj_graphique(filiale, annee)

        def _on_annee_change(*_):
            filiale = selected_filiale.get()
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            maj_graphique(filiale, annee)

        select_box.bind("<<ComboboxSelected>>", lambda e: _on_filiale_change())
        annees_box.bind("<<ComboboxSelected>>", lambda e: _on_annee_change())

        # ===== Affichage initial =====
        _on_filiale_change()

        # ===== Scroll molette (Canvas) =====
        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units"); return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units"); return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")
                # ===== Bouton retour =====

    def afficher_repartition_flux(self):
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.colors as mcolors
        import matplotlib.cm as cm
        import re
        import datetime as _dt

        # =============== RESET PAGE ===============
        self.vider_fenetre()

        # =============== HEADER ===============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - RÉPARTITION DES FLUX", font=titre_font)\
            .pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            test = tk.Label(self, text="Test", font=titre_font); test.update_idletasks()
            text_h = test.winfo_reqheight(); test.destroy()
            ratio = logo_img.width / logo_img.height
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # =============== CONTAINER SCROLLABLE ===============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # Frame scrollable dans le Canvas
        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # =============== SÉLECTEURS ===============
        ctk.CTkLabel(
            scrollable_frame,
            text="Top flux avec le plus d'écarts importants",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        ).pack(pady=15)

        # Filiale
        ctk.CTkLabel(
            scrollable_frame, text="Sélectionnez une filiale :", font=("Arial", 12, "bold"),
            text_color="white", fg_color="#00122e"
        ).pack(pady=(10, 2), anchor="center")

        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = tk.StringVar(value=filiales[0])
        select_box = ttk.Combobox(scrollable_frame, textvariable=selected_filiale, values=filiales,
                                state="readonly", width=25)
        select_box.pack(pady=(0, 5), anchor="center")

        # Année
        ctk.CTkLabel(
            scrollable_frame, text="Sélectionnez une année :", font=("Arial", 12, "bold"),
            text_color="white", fg_color="#00122e"
        ).pack(pady=(10, 2), anchor="center")

        annees_var = tk.StringVar(value="Toutes années")
        annees_box = ttk.Combobox(scrollable_frame, textvariable=annees_var,
                                values=["Toutes années"], state="readonly", width=20)
        annees_box.pack(pady=(0, 10), anchor="center")

        # Profil (uniquement pour le graphe 2 — valorisation)
        ctk.CTkLabel(
            scrollable_frame, text="Filtrer par profil (valorisation seulement) :", font=("Arial", 12, "bold"),
            text_color="white", fg_color="#00122e"
        ).pack(pady=(5, 2), anchor="center")

        profils_var = tk.StringVar(value="Tous profils")
        profils_box = ttk.Combobox(scrollable_frame, textvariable=profils_var,
                                values=["Tous profils"], state="readonly", width=40)
        profils_box.pack(pady=(0, 10), anchor="center")

        # Mémorisation des widgets graphiques
        self.graph_widgets = []

        # =============== HELPERS ===============
        def _annees_pour_filiale(filiale):
            """Union des années disponibles via self._annees_disponibles(section, flux)."""
            annees = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]
            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _tok in noms_colonnes:
                    try:
                        years = self._annees_disponibles(feuille, nom_flux)
                        annees.update(years)
                    except Exception:
                        pass
            return sorted(annees)

        def _parse_profil_day_month(nom: str):
            m = re.search(r'(\d{1,2})[/-](\d{1,2})', str(nom))
            if not m:
                return None, None
            jj = int(m.group(1)); mm = int(m.group(2))
            if 1 <= jj <= 31 and 1 <= mm <= 12:
                return jj, mm
            return None, None

        def _profils_uniques_ordonnes_par_mois(filiale: str, annee: int | None):
            if annee is None:
                return []
            profils = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]
            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _tok in noms_colonnes:
                    try:
                        noms_ok, _flags = self._profils_for_year(feuille, nom_flux, annee)
                        for p in noms_ok:
                            if p and str(p).strip():
                                profils.add(str(p).strip())
                    except Exception:
                        pass
            def keyer(name: str):
                jj, mm = _parse_profil_day_month(name)
                if mm is None:
                    return (99, 99, name.lower())
                return (mm, jj if jj is not None else 31, name.lower())
            return sorted(profils, key=keyer)

        def _to_number(x):
            if x is None: return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}: return None
                s = s.replace(",", ".")
                try: return float(s)
                except Exception: return None
            try: return float(x)
            except Exception: return None

        def _year_of(d):
            if d is None: return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        # =============== COEUR : CALCUL & GRAPHIQUES ===============
        def afficher_graphes():
            # Nettoyage
            for w in self.graph_widgets:
                try: w.destroy()
                except Exception: pass
            self.graph_widgets.clear()

            # Filtres
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            suffix_annee = "" if annee is None else f" - {annee}"

            filiale = selected_filiale.get()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]

            # flux à exclure (agrégats)
            flux_a_exclure = {
                "Cash flow de financement", "Cash flow net", "Sous total financier",
                "Sous total Investissements nets et ACE", "Free cash Flow",
                "Sous total recettes", "Sous total dépenses", "C/C - Groupe"
            }

            # --- liste complète des flux (pour afficher aussi ceux à 0) ---
            tous_flux = []
            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _ in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    if nom_flux not in tous_flux:
                        tous_flux.append(nom_flux)

            # ==== PASS A : nombre d'écarts ≥ 40% par FLUX ====
            ecarts_par_flux = {f: 0 for f in tous_flux}

            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    try:
                        dates, reel, previsions, _noms_prof = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for i, d in enumerate(dates):
                        # filtre année
                        if annee is not None:
                            y = _year_of(d)
                            if y is not None and y != annee:
                                continue

                        # réel
                        r_val = _to_number(reel[i] if i < len(reel) else None)
                        if r_val is None:
                            r_val = 0.0

                        # chaque profil = une prévision
                        for prev_list in previsions:
                            pv = prev_list[i] if i < len(prev_list) else None
                            prev_val = _to_number(pv)
                            if prev_val is None:
                                continue
                            if prev_val == 0:
                                if r_val == 0:
                                    continue  # 0 vs 0 → pas d'écart
                                prev_val = 1.0  # convention anti-division par zéro

                            ecart = (r_val - prev_val) / prev_val
                            if abs(ecart) >= 0.4:
                                ecarts_par_flux[nom_flux] += 1

            # ==== PASS B : nombre de prévisions non vides par FLUX (dénominateur PAR PROFIL) ====
            prevs_par_flux = {f: 0 for f in tous_flux}

            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    try:
                        dates, _reel, previsions, _noms_prof = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    # compter par profil puis sommer
                    for prev_list in previsions:
                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue
                            if i < len(prev_list) and _to_number(prev_list[i]) is not None:
                                prevs_par_flux[nom_flux] += 1

            # --- ordre d’affichage : tous les flux, triés par nb d’écarts décroissant (0 inclus) ---
            noms_flux = sorted(tous_flux, key=lambda f: ecarts_par_flux.get(f, 0), reverse=True)
            valeurs   = [ecarts_par_flux.get(f, 0) for f in noms_flux]

            # Tri (par nb d'écarts décroissant)
            flux_tries = sorted(ecarts_par_flux.items(), key=lambda x: x[1], reverse=True)
            noms_flux = [f for f, _ in flux_tries]
            valeurs   = [v for _, v in flux_tries]

            # >>>>>> Filtrage DISPLAY-ONLY pour retirer les 0
            noms_flux_aff = [f for f, v in zip(noms_flux, valeurs) if v != 0]
            valeurs_aff   = [v for v in valeurs if v != 0]

            # Graphe 1 — Nombre d'écarts
            if noms_flux_aff:
                fig1, ax1 = plt.subplots(figsize=(12, max(6, len(noms_flux_aff) * 0.5)),
                                        facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")
                bars1 = ax1.barh(noms_flux_aff, valeurs_aff, color="#5DADE2")
                ax1.set_xlabel("Nombre d'écarts importants (≥ 40%)", color="white", fontsize=12)
                ax1.set_ylabel("Flux", color="white", fontsize=12)
                ax1.tick_params(axis='x', colors="white")
                ax1.tick_params(axis='y', colors="white")
                ax1.set_title(f"Flux – Nombre d'écarts ({filiale}){suffix_annee}", color="white", fontsize=14)
                for bar in bars1:
                    ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                            str(int(bar.get_width())), va='center', color="white")
                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                canvas_fig1.get_tk_widget().pack(pady=10, fill="both", expand=True)
                self.graph_widgets.append(canvas_fig1.get_tk_widget())
                plt.close(fig1)
            else:
                fig, ax = plt.subplots(figsize=(8, 4), facecolor="#00122e", constrained_layout=True)
                ax.set_facecolor("#00122e")
                ax.text(0.5, 0.5, "Aucun écart important détecté", ha='center', va='center', fontsize=12, color="white")
                ax.axis("off")
                canvas_fig = FigureCanvasTkAgg(fig, master=scrollable_frame)
                canvas_fig.draw()
                canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)
                self.graph_widgets.append(canvas_fig.get_tk_widget())
                plt.close(fig)
                return  # rien d'autre si vide

            # ==== 1bis) Graphe % d'écarts par flux ====
            pourcentage_par_flux = []
            for f in noms_flux_aff:
                nb_e = ecarts_par_flux.get(f, 0)
                nb_p = prevs_par_flux.get(f, 0)
                pct  = (nb_e / nb_p * 100) if nb_p > 0 else 0.0
                pourcentage_par_flux.append(pct)

            fig_pct, ax_pct = plt.subplots(figsize=(12, max(5, len(noms_flux_aff) * 0.45)),
                                        facecolor="#00122e", constrained_layout=True)
            ax_pct.set_facecolor("#00122e")

            if pourcentage_par_flux:
                vmin = min(pourcentage_par_flux)
                vmax = max(pourcentage_par_flux) if max(pourcentage_par_flux) != vmin else vmin + 1
                norm_pct = mcolors.Normalize(vmin=vmin, vmax=vmax)
                cmap_pct = cm.Blues
                colors_pct = [cmap_pct(norm_pct(v)) for v in pourcentage_par_flux]
            else:
                colors_pct = []

            # >>>>>> utiliser la liste filtrée
            bars_pct = ax_pct.barh(noms_flux_aff, pourcentage_par_flux, color=colors_pct, alpha=0.95)
            ax_pct.set_xlabel("% d'écarts (≥ 40%) / nb de prévisions", color="white", fontsize=12)
            ax_pct.set_ylabel("Flux", color="white", fontsize=12)
            ax_pct.tick_params(axis='x', colors="white")
            ax_pct.tick_params(axis='y', colors="white")
            ax_pct.set_title(f"Flux – % d'écarts ({filiale}){suffix_annee}", color="white", fontsize=14)
            for bar, v in zip(bars_pct, pourcentage_par_flux):
                ax_pct.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                            f"{v:.1f}%", va='center', color="white")
            canvas_pct = FigureCanvasTkAgg(fig_pct, master=scrollable_frame)
            canvas_pct.draw()
            canvas_pct.get_tk_widget().pack(pady=10, fill="both", expand=True)
            self.graph_widgets.append(canvas_pct.get_tk_widget())
            plt.close(fig_pct)

            # ==== Tableau sous le graphe % ====
            table_frame = tk.Frame(scrollable_frame, bg="#00122e")
            table_frame.pack(fill="x", padx=10, pady=(0, 15))
            self.graph_widgets.append(table_frame)

            ctk.CTkLabel(
                table_frame,
                text="Données utilisées pour le % d'écarts (par flux)",
                font=("Segoe UI", 13, "bold"),
                text_color="white"
            ).pack(anchor="w", pady=(0, 6))

            colonnes = ("Flux", "Prévisions", "Écarts ≥40%", "% Écarts")
            tree = ttk.Treeview(table_frame, columns=colonnes, show="headings",
                                height=min(12, len(noms_flux_aff)))
            for col in colonnes:
                tree.heading(col, text=col)

            tree.column("Flux", anchor="w", width=280)
            tree.column("Prévisions", anchor="center", width=120)
            tree.column("Écarts ≥40%", anchor="center", width=120)
            tree.column("% Écarts", anchor="center", width=100)

            for f in noms_flux_aff:
                nb_e = ecarts_par_flux.get(f, 0)
                nb_p = prevs_par_flux.get(f, 0)
                pct  = (nb_e / nb_p * 100) if nb_p > 0 else 0.0
                tree.insert("", "end", values=(f, nb_p, nb_e, f"{pct:.1f}%"))

            tree.pack(fill="x", expand=True)

            # >>>>>> totaux calculés sur la sélection affichée uniquement
            tot_prev = sum(prevs_par_flux.get(f, 0) for f in noms_flux_aff)
            tot_ecarts = sum(ecarts_par_flux.get(f, 0) for f in noms_flux_aff)
            pct_global = (tot_ecarts / tot_prev * 100) if tot_prev > 0 else 0.0

            

            # ==== 2) Valorisation signée par flux (filtrable par PROFIL) ====
            prof_sel = profils_var.get()
            prof_sel = None if (not prof_sel or prof_sel == "Tous profils") else prof_sel

            valeur_ecarts = []
            for flux in noms_flux_aff:
                total_ecart = 0.0
                for feuille in feuilles:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for nom, col_start in noms_colonnes:
                        if nom != flux or nom in flux_a_exclure:
                            continue
                        try:
                            dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        except Exception:
                            continue

                        # Périmètre commun d'indices pour la valorisation aussi
                        idxs = []
                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue
                            idxs.append(i)
                        if not idxs:
                            continue

                        # Profils à inclure
                        if prof_sel is None:
                            idx_profils = range(len(previsions))
                        else:
                            idx_profils = [i for i, p in enumerate(noms_profils) if str(p).strip() == prof_sel]

                        for i in idxs:
                            r_val = _to_number(reel[i] if i < len(reel) else None)
                            if r_val is None:
                                r_val = 0.0
                            for idxp in idx_profils:
                                if idxp >= len(previsions):
                                    continue
                                prev_list = previsions[idxp]
                                pv = prev_list[i] if i < len(prev_list) else None
                                prev_val = _to_number(pv)
                                if prev_val is None:
                                    continue
                                if prev_val == 0:
                                    if r_val == 0:
                                        continue
                                    prev_val = 1.0
                                ecart = (r_val - prev_val) / prev_val
                                if abs(ecart) >= 0.4:
                                    total_ecart += (r_val - prev_val)  # signé (k€)

                valeur_ecarts.append(total_ecart)

            # Graphe 3 — Valorisation signée
            fig2, ax2 = plt.subplots(figsize=(12, max(6, len(noms_flux_aff) * 0.5)), facecolor="#00122e", constrained_layout=True)
            ax2.set_facecolor("#00122e")

            max_abs = max([abs(v) for v in valeur_ecarts] or [1])
            if max_abs == 0:
                max_abs = 1

            norm = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
            cmap = cm.RdBu_r
            colors2 = [cmap(norm(v)) for v in valeur_ecarts]

            # >>>>>> utiliser la liste filtrée pour cohérence
            bars2 = ax2.barh(noms_flux_aff, valeur_ecarts, color=colors2, alpha=0.9)
            ax2.axvline(0, color="white", linewidth=1, alpha=0.8)
            ax2.set_xlabel("Valorisation cash des écarts (k€)", color="white", fontsize=12)
            ax2.set_ylabel("Flux", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white")
            ax2.tick_params(axis='y', colors="white")

            titre2 = f"Flux – Valorisation des écarts ({filiale}){suffix_annee}"
            if prof_sel:
                jj, mm = _parse_profil_day_month(prof_sel)
                if mm is not None:
                    titre2 += f" | Profil : {prof_sel} (M{mm:02d})"
                else:
                    titre2 += f" | Profil : {prof_sel}"
            else:
                titre2 += " | Profils : tous"
            ax2.set_title(titre2, color="white", fontsize=14)

            # Labels intelligents
            inside_threshold = 0.20 * max_abs
            pad_data = 0.03 * max_abs
            longest_out_right_chars = 0
            longest_out_left_chars = 0
            labels_info = []
            for bar, v in zip(bars2, valeur_ecarts):
                place_inside = abs(v) >= inside_threshold
                sign = 1 if v >= 0 else -1
                labels_info.append((bar, v, place_inside, sign))
                txt = f"{int(v):,} k€".replace(",", " ")
                if not place_inside:
                    if sign > 0:
                        longest_out_right_chars = max(longest_out_right_chars, len(txt))
                    else:
                        longest_out_left_chars = max(longest_out_left_chars, len(txt))
            char_unit = 0.012 * max_abs
            extra_right = longest_out_right_chars * char_unit + (pad_data if longest_out_right_chars > 0 else 0)
            extra_left  = longest_out_left_chars  * char_unit + (pad_data if longest_out_left_chars  > 0 else 0)
            ax2.set_xlim(-max_abs - extra_left, max_abs + extra_right)

            for bar, v, inside, sign in labels_info:
                txt = f"{int(v):,} k€".replace(",", " ")
                y = bar.get_y() + bar.get_height()/2
                if inside:
                    x = v - pad_data if sign > 0 else v + pad_data
                    ha = 'right' if sign > 0 else 'left'
                    ax2.text(x, y, txt, va='center', ha=ha, color="white", clip_on=True)
                else:
                    x = v + pad_data if sign > 0 else v - pad_data
                    ha = 'left' if sign > 0 else 'right'
                    ax2.text(x, y, txt, va='center', ha=ha, color="white", clip_on=True)

            fig2.tight_layout(pad=2.0)
            canvas_fig2 = FigureCanvasTkAgg(fig2, master=scrollable_frame)
            canvas_fig2.draw()
            canvas_fig2.get_tk_widget().pack(pady=10, fill="both", expand=True)
            self.graph_widgets.append(canvas_fig2.get_tk_widget())
            plt.close(fig2)

        # =============== LIAISONS (filiale / année / profil) ===============
        def _on_filiale_change(*_):
            filiale = selected_filiale.get()
            # années
            annees = _annees_pour_filiale(filiale)
            annees_box.config(values=(["Toutes années"] + [str(a) for a in annees]))
            annees_var.set(str(annees[-1]) if annees else "Toutes années")

            # profils : union sur l’année, triés par mois/jour
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            profils_list = _profils_uniques_ordonnes_par_mois(filiale, annee) if annee is not None else []
            profils_box.config(values=(["Tous profils"] + profils_list))
            profils_var.set("Tous profils")

            afficher_graphes()

        def _on_annee_change(*_):
            filiale = selected_filiale.get()
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)

            profils_list = _profils_uniques_ordonnes_par_mois(filiale, annee) if annee is not None else []
            valeurs = (["Tous profils"] + profils_list)
            profils_box.config(values=valeurs)
            if profils_var.get() not in valeurs:
                profils_var.set("Tous profils")

            afficher_graphes()

        selected_filiale.trace_add("write", _on_filiale_change)
        annees_var.trace_add("write", _on_annee_change)
        profils_var.trace_add("write", lambda *_: afficher_graphes())

        # =============== AFFICHAGE INITIAL ===============
        _on_filiale_change()

        # =============== SCROLL MOLETTE (Canvas) ===============
        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # =============== BOUTON RETOUR ===============
        ctk.CTkButton(
            scrollable_frame, text="⬅️ Retour au menu", command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)

#===================Page heatmaps (ecart/anomalies)===================
    def afficher_heatmap_anomalies(self):
        import customtkinter as ctk
        from tkinter import ttk, messagebox
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import numpy as np
        import pandas as pd
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.text import Annotation
        from sklearn.ensemble import IsolationForest
        import seaborn as sns
        import matplotlib.patches as patches
        import re
        import datetime as _dt

        self.vider_fenetre()

        # === HEADER ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - DÉTECTION D’ANOMALIES", font=titre_font)\
            .pack(side="left", anchor="w")

        # Logo
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365)\Projet PULSE\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_height = 40
            new_width = int(new_height * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Logo ignoré: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === CONTROLES HORIZONTAUX (retour + filtres) ===
        control_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        control_frame.pack(side="top", fill="x", pady=10, padx=30)

        # Bouton retour (en haut)
        ctk.CTkButton(
            control_frame, text="⬅️ Retour au menu", command=self.retour_menu,
            width=180, height=40, corner_radius=15, fg_color="#444", hover_color="#666",
            text_color="white", font=("Segoe UI", 13, "bold")
        ).pack(side="left", padx=(0, 16))

        def _label(parent, text):
            return ttk.Label(parent, text=text, background="#001f3f", foreground="white",
                            font=('Segoe UI', 12, 'bold'))

        # --- Filiale ---
        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = tk.StringVar(value=filiales[0])
        _label(control_frame, "Filiale :").pack(side="left", padx=(0, 8))
        select_box = ttk.Combobox(control_frame, textvariable=selected_filiale,
                                values=filiales, state="readonly", width=25)
        select_box.pack(side="left", padx=(0, 20))

        # --- Année (sans 'Toutes années') ---
        _label(control_frame, "Année :").pack(side="left", padx=(0, 8))
        annees_var = tk.StringVar(value="")
        annees_box = ttk.Combobox(control_frame, textvariable=annees_var,
                                values=[], state="readonly", width=12)
        annees_box.pack(side="left")

        # === CONTENEUR SCROLLABLE ===
        container = tk.Frame(self, bg="#001f3f")
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg="#001f3f", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=canvas.xview)

        scrollable_frame = tk.Frame(canvas, bg="#001f3f")
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        # === Tableau dynamique ===
        cols = ["Date", "Flux", "Profil", "Réel (k€)", "Prévision (k€)", "Écart (k€)"]
        tree = ttk.Treeview(scrollable_frame, columns=cols, show="headings", height=10)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=150, anchor="center")
        tree.pack(pady=10, fill="x")

        tree.tag_configure("neg", foreground="red")
        tree.tag_configure("pos", foreground="green")

        self.canvas_fig = None
        self.fig = None

        # === Catégorisation favorable/défavorable ===
        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]
        decaissements = ["Péages", "Charges de personnel", "ACE & Investissements"]
        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital", "Sous total financier",
            "Free cash Flow", "Emprunts", "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers", "Emprunts / Prêts - Groupe",
            "Cash flow de financement", "Cash flow net", "Cessions d'immobilisations",
            "Impôts et Taxes", "Sous total dépenses"
        ]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                return (reel_val >= prev_val) if prev_val >= 0 else (abs(reel_val) <= abs(prev_val))
            else:
                return reel_val >= prev_val

        # ===== Helpers année =====
        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, (int, float)):
                y = int(d); return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]
            for feuille in feuilles:
                try:
                    ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)
                    for nom_flux, col_start in noms_colonnes_local:
                        dates, _, _, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        for d in dates:
                            y = _year_of(d)
                            if y is not None:
                                annees.add(y)
                except Exception:
                    pass
            return sorted(annees)

        # === Rendu principal ===
        def afficher_heatmap(annee:int):
            if self.canvas_fig is not None:
                self.canvas_fig.get_tk_widget().destroy()
                self.canvas_fig = None
                self.fig = None

            filiale_actuelle = selected_filiale.get()
            df_all = []

            feuilles = sections.values() if filiale_actuelle == "Toute filiale" else [sections[filiale_actuelle]]

            # Flux exclus
            flux_a_exclure = [
                "Cash flow de financement", "Cash flow net", "Sous total financier",
                "Sous total Investissements nets et ACE", "Free cash Flow",
                "Sous total recettes", "Sous total dépenses"
            ]

            for feuille in feuilles:
                ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes_local:
                    if nom_flux in flux_a_exclure:
                        continue
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    for i, date in enumerate(dates):
                        # Filtre par année SEULEMENT (pas d'option "Toutes années")
                        y = _year_of(date)
                        if y is None or y != annee:
                            continue
                        if i >= len(reel) or reel[i] is None:
                            continue
                        for idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list) or prev_list[i] is None:
                                continue
                            r = reel[i]
                            prev_val = prev_list[i]
                            if r == 0:
                                r = 1
                            ecart = r - prev_val
                            df_all.append({
                                "Date": date,
                                "Flux": nom_flux,
                                "Profil": noms_profils[idx],
                                "Réel": reel[i],
                                "Prévision": prev_val,
                                "Écart": ecart
                            })

            if not df_all:
                messagebox.showinfo("Info", "Aucune donnée exploitable trouvée pour cette année.")
                return

            df_all = pd.DataFrame(df_all)
            df_all["Écart_abs"] = df_all["Écart"].abs()

            # Isolation Forest avec contamination dynamique
            seuil = 2 * df_all["Écart"].std()
            contamination_dyn = min(0.05, max(0.01, (df_all["Écart"].abs() > seuil).mean()))
            iso = IsolationForest(contamination=contamination_dyn, random_state=42)
            df_all["Anomalie"] = iso.fit_predict(df_all[["Écart"]])
            df_all["Anomalie"] = df_all["Anomalie"].map({1: 0, -1: 1})

            # Matrices
            heatmap_data = df_all.pivot_table(index="Profil", columns="Flux", values="Anomalie",
                                            aggfunc="sum", fill_value=0).astype(int)
            mean_data = df_all.pivot_table(index="Profil", columns="Flux", values="Écart_abs",
                                        aggfunc="mean", fill_value=0)
            heatmap_data = heatmap_data.reindex(df_all['Profil'].drop_duplicates()[::-1])
            mean_data = mean_data.reindex(heatmap_data.index)

            # === Graphique ===
            plt.close("all")
            self.fig, ax = plt.subplots(figsize=(18, max(6, heatmap_data.shape[0]*0.5)), facecolor="#001f3f")
            ax.set_facecolor("#001f3f")

            # Annotations entières
            annot_data = heatmap_data.values  # ndarray d'entiers
            sns.heatmap(
                heatmap_data,
                cmap="Reds",
                annot=annot_data,   # annotations = entiers
                fmt="d",            # format entier
                linewidths=0.5,
                linecolor="#444",
                ax=ax,
                cbar=True
            )

            # Axes et titre
            ax.set_xlabel("Flux", color="white", fontsize=12, fontweight="bold")
            ax.set_ylabel("Profil", color="white", fontsize=12, fontweight="bold")
            ax.set_title(f"Heatmap des anomalies ({filiale_actuelle}) - {annee}",
                        color="white", fontsize=16, fontweight="bold")
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=10, color="white")
            plt.setp(ax.get_yticklabels(), rotation=0, fontsize=10, color="white")

            # Colorbar
            cbar = ax.collections[0].colorbar
            cbar.set_label("Nombre d'anomalies", color="white", fontsize=12, fontweight="bold")
            cbar.ax.yaxis.set_tick_params(color='white')
            plt.setp(cbar.ax.yaxis.get_ticklabels(), color='white')
            cbar.outline.set_edgecolor('white')

            self.fig.tight_layout(pad=2.0)
            self.fig.subplots_adjust(bottom=0.25)

            # === Tooltip & survol ===
            tooltip = Annotation("", xy=(0, 0), xytext=(15, 15), textcoords="offset points",
                                ha="left", va="bottom",
                                bbox=dict(boxstyle="round", fc="black", ec="white", lw=1, alpha=0.8),
                                color="white", fontsize=9)
            tooltip.set_visible(False)
            ax.add_artist(tooltip)

            hover_rect = patches.Rectangle((0, 0), 1, 1, fill=True, edgecolor="black",
                                        linewidth=2, facecolor="blue", alpha=0.3)
            hover_rect.set_visible(False)
            ax.add_patch(hover_rect)

            def get_cell(event):
                if event.inaxes != ax or event.xdata is None or event.ydata is None:
                    return None, None
                x, y = int(event.xdata), int(event.ydata)
                if x < 0 or y < 0 or x >= heatmap_data.shape[1] or y >= heatmap_data.shape[0]:
                    return None, None
                return x, y

            def on_hover(event):
                x, y = get_cell(event)
                if x is None:
                    hover_rect.set_visible(False)
                    tooltip.set_visible(False)
                    self.fig.canvas.draw_idle()
                    return
                hover_rect.set_xy((x, y))
                hover_rect.set_visible(True)

                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]
                n_anomalies = int(heatmap_data.iloc[y, x])

                tooltip.xy = (event.xdata, event.ydata)
                tooltip.set_text(f"{profil} / {flux}\nAnomalies : {n_anomalies}")
                tooltip.set_visible(True)
                self.fig.canvas.draw_idle()

            def on_click(event):
                x, y = get_cell(event)
                if x is None:
                    return
                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]
                tree.delete(*tree.get_children())

                filtered = df_all[(df_all["Flux"] == flux) & (df_all["Profil"] == profil) & (df_all["Anomalie"] == 1)].copy()
                filtered["Écart_abs"] = filtered["Écart"].abs()
                filtered = filtered.sort_values(by="Écart_abs", ascending=False)

                for _, row in filtered.iterrows():
                    date_str = row["Date"].strftime("%d/%m/%Y") if hasattr(row["Date"], "strftime") else str(row["Date"])
                    reel_val, prev_val = row["Réel"], row["Prévision"]
                    favorable = est_favorable(row["Flux"], reel_val, prev_val)
                    tag = "pos" if favorable else "neg"

                    tree.insert("", "end", values=[
                        date_str,
                        row.get("Flux", ""),
                        row.get("Profil", ""),
                        f"{row.get('Réel', 0):,.0f}".replace(",", " "),
                        f"{row.get('Prévision', 0):,.0f}".replace(",", " "),
                        f"{row.get('Écart', 0):,.0f}".replace(",", " ")
                    ], tags=(tag,))

            self.fig.canvas.mpl_connect("motion_notify_event", on_hover)
            self.fig.canvas.mpl_connect("button_press_event", on_click)

            self.canvas_fig = FigureCanvasTkAgg(self.fig, master=scrollable_frame)
            self.canvas_fig.draw()
            self.canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)

        # === Callbacks filtres ===
        def _on_filiale_change(*_):
            filiale = selected_filiale.get()
            annees = _annees_pour_filiale(filiale)
            values = [str(a) for a in annees] if annees else []
            annees_box.config(values=values)
            if values:
                annees_var.set(values[-1])      # dernière année dispo
                afficher_heatmap(int(values[-1]))
            else:
                annees_var.set("")
                messagebox.showinfo("Info", "Aucune année disponible pour cette filiale.")

        def _on_annee_change(*_):
            val = annees_var.get()
            if not val:
                return
            afficher_heatmap(int(val))

        select_box.bind("<<ComboboxSelected>>", lambda e: _on_filiale_change())
        annees_box.bind("<<ComboboxSelected>>", lambda e: _on_annee_change())

        # === Affichage initial ===
        _on_filiale_change()

    def afficher_heatmap_ecarts(self):
        import customtkinter as ctk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        from collections import defaultdict
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import tkinter as tk
        import re
        import datetime as _dt

        # ===== RESET PAGE =====
        self.vider_fenetre()

        # ===== HEADER =====
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - VISUALISATION GRAPHIQUE DES ÉCARTS", font=titre_font)\
            .pack(side="left", anchor="w")

        # --- Logo ---
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365)\Projet PULSE\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_height = 40
            new_width = int(new_height * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        # Barre blanche
        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ===== CONTENEUR + SCROLLBARS =====
        container = tk.Frame(self, bg="#001f3f")
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg="#001f3f", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=canvas.xview)

        scrollable_frame = tk.Frame(canvas, bg="#001f3f")
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        # ===== Barre d’outils (Retour + sélecteurs alignés) =====
        toolbar = tk.Frame(scrollable_frame, bg="#001f3f")
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        # Bouton retour en haut
        ctk.CTkButton(
            toolbar,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=36,
            corner_radius=12,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(side="left", padx=(0, 16))

        # Sélecteurs horizontaux
        def _label(parent, txt):
            return ctk.CTkLabel(parent, text=txt, font=("Arial", 12, "bold"),
                                text_color="white", fg_color="#001f3f")

        _label(toolbar, "Filiale :").pack(side="left", padx=(0, 6))
        filiales = ["Toutes filiales"] + list(sections.values())
        selected_filiale = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(toolbar, textvariable=selected_filiale, values=filiales,
                                state="readonly", width=28)
        filiale_box.pack(side="left", padx=(0, 16))

        _label(toolbar, "Année :").pack(side="left", padx=(0, 6))
        annees_var = tk.StringVar(value="")
        annees_box = ttk.Combobox(toolbar, textvariable=annees_var, values=[],
                                state="readonly", width=12)
        annees_box.pack(side="left")

        # ===== Titre secondaire =====
        ttk.Label(
            scrollable_frame,
            text="Carte thermique des écarts par mois et profil",
            font=("Segoe UI Semibold", 20, "bold"),
            foreground="white",
            background="#001f3f"
        ).pack(pady=(8, 10))

        # ===== Zone graphique =====
        graph_holder = tk.Frame(scrollable_frame, bg="#001f3f")
        graph_holder.pack(fill="both", expand=True, padx=10, pady=(5, 15))

        # ===== Helpers =====
        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, (int, float)):
                y = int(d)
                return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]
            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for _nom, col_start in noms_colonnes:
                        dates, _reel, _prevs, _profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        for d in dates:
                            y = _year_of(d)
                            if y is not None:
                                annees.add(y)
                except Exception:
                    pass
            return sorted(annees)

        # ===== Rendu heatmap =====
        def maj_heatmap(filiale, annee:int):
            for w in graph_holder.winfo_children():
                w.destroy()

            noms_a_convertir_flux = [
                "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
                "Créances CDP", "Placements", "CC financiers",
                "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement",
                "Endettement Net"
            ]

            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]
            data_dict = defaultdict(lambda: defaultdict(int))  # <-- entiers

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for nom, col_start in noms_colonnes:
                    if nom in noms_a_convertir_flux:
                        continue

                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)

                    for i, date in enumerate(dates):
                        y = _year_of(date)
                        if y is not None and y != annee:
                            continue
                        if i >= len(reel) or reel[i] in (None, 0):
                            continue

                        try:
                            mois = pd.to_datetime(date).strftime("%Y-%m")
                        except Exception:
                            continue

                        for p_idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list):
                                continue
                            prev_val = prev_list[i]
                            if prev_val is None:
                                continue
                            try:
                                ecart = (prev_val - reel[i]) / reel[i]
                            except Exception:
                                continue
                            if abs(ecart) >= 0.4:
                                profil_name = noms_profils[p_idx]
                                data_dict[profil_name][mois] += 1  # <-- compte entier

            if not data_dict:
                ctk.CTkLabel(graph_holder, text="Aucune donnée à afficher pour ces filtres.",
                            font=("Segoe UI", 14, "italic"),
                            text_color="white", fg_color="#001f3f").pack(pady=20)
                return

            heatmap_df = pd.DataFrame(data_dict).T.fillna(0).astype(int)  # <-- entiers
            heatmap_df = heatmap_df.reindex(sorted(heatmap_df.columns), axis=1)

            # Figure
            fig, ax = plt.subplots(figsize=(16, 10), facecolor="#001f3f")
            ax.set_facecolor("#001f3f")

            # Annotations entières (plus de floats)
            annot_data = heatmap_df.values  # ndarray d'entiers
            sns.heatmap(
                heatmap_df,
                cmap="coolwarm",
                annot=annot_data,   # <-- on passe l'array d’annotations
                fmt="d",            # <-- format entier
                linewidths=0.5,
                linecolor="#444",
                ax=ax,
                cbar_kws={'label': "Nombre d'écarts significatifs"}
            )

            cbar = ax.collections[0].colorbar
            cbar.ax.yaxis.label.set_color("white")
            cbar.ax.tick_params(colors="white")

            ax.set_title(("Toutes filiales" if filiale == "Toutes filiales" else filiale) + f" - {annee}",
                        fontsize=16, color="white")
            ax.set_xlabel("Mois", color="white")
            ax.set_ylabel("Profil", color="white")
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", color="white")
            plt.setp(ax.get_yticklabels(), color="white")

            canvas_fig = FigureCanvasTkAgg(fig, master=graph_holder)
            canvas_fig.draw()
            canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)

        # ===== Callbacks =====
        def _on_filiale_change(*_):
            filiale = selected_filiale.get()
            annees = _annees_pour_filiale(filiale)
            # plus d’option "Toutes années" -> uniquement années disponibles
            values = [str(a) for a in annees] if annees else []
            annees_box.config(values=values)

            # défaut = dernière année disponible
            if values:
                annees_var.set(values[-1])
                annee = int(values[-1])
                maj_heatmap(filiale, annee)
            else:
                # aucune année dispo
                annees_var.set("")
                for w in graph_holder.winfo_children():
                    w.destroy()
                ctk.CTkLabel(graph_holder, text="Aucune année disponible pour cette filiale.",
                            font=("Segoe UI", 14, "italic"),
                            text_color="white", fg_color="#001f3f").pack(pady=20)

        def _on_annee_change(*_):
            filiale = selected_filiale.get()
            val_annee = annees_var.get()
            if not val_annee:
                return
            maj_heatmap(filiale, int(val_annee))

        filiale_box.bind("<<ComboboxSelected>>", lambda e: _on_filiale_change())
        annees_box.bind("<<ComboboxSelected>>", lambda e: _on_annee_change())

        # ===== Affichage initial =====
        _on_filiale_change()

#===================Page backtesting multi-horizon===================
    def afficher_backtesting_multi_horizon(self):
        import customtkinter as ctk
        from tkinter import ttk, filedialog, messagebox
        import tkinter as tk
        import matplotlib.pyplot as plt
        import seaborn as sns
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import pandas as pd
        from PIL import Image
        from customtkinter import CTkImage
        import numpy as np
        import re
        import datetime as _dt
        from functools import lru_cache

        # =============== RESET PAGE ===============
        self.vider_fenetre()

        # =============== HEADER ===============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - BACKTESTING MULTI-HORIZON", font=titre_font)\
            .pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_h = 40
            resized_logo = logo_img.resize((int(new_h * ratio), new_h), Image.Resampling.LANCZOS)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(int(new_h * ratio), new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement logo: {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # =============== CONTENEUR SCROLLABLE ===============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ctk.CTkFrame(canvas, fg_color="#00122e", corner_radius=0)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # =============== HELPERS ===============
        def _to_number(x):
            if x is None: return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}: return None
                s = s.replace(",", ".")
                try: return float(s)
                except Exception: return None
            try: return float(x)
            except Exception: return None

        def _year_of(d):
            if d is None: return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, (int, float)):
                y = int(d); return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        def _parse_mois_from_profil(name: str):
            """
            Extrait des indices d’horizon depuis le libellé de profil.
            Retourne (mm, jj, rel):
            - JJ/MM -> (mm, jj, None)
            - 'Mxx' -> (mm, None, None)
            - 'M-2' -> (None, None, 2)
            """
            s = str(name)
            m = re.search(r'(\d{1,2})[/-](\d{1,2})', s)
            if m:
                jj = int(m.group(1)); mm = int(m.group(2))
                if 1 <= mm <= 12:
                    return mm, jj, None
            m2 = re.search(r'[Mm]\s*0?(\d{1,2})', s)
            if m2:
                mm = int(m2.group(1))
                if 1 <= mm <= 12:
                    return mm, None, None
            m3 = re.search(r'[Mm]\s*-\s*(\d+)', s)
            if m3:
                return None, None, int(m3.group(1))
            return None, None, None

        def _horizon_from_name(p_name, previsions_len, max_m):
            mm, jj, rel = _parse_mois_from_profil(p_name)
            if rel is not None:
                return max(1, int(rel))
            if mm is not None and max_m is not None:
                return max(1, int(max_m - mm + 1))
            # fallback : plus l’index est grand, plus l’horizon est court
            return max(1, previsions_len)

        def _profil_max_m(noms_profils):
            mms = [m for m, _, _ in map(_parse_mois_from_profil, noms_profils) if m is not None]
            return max(mms) if mms else None

        @lru_cache(maxsize=64)
        def _annees_pour_filiale_cached(filiale_key: str):
            """Union des années disponibles selon la filiale sélectionnée."""
            if filiale_key == "Toutes filiales":
                feuilles = list(sections.values())
            else:
                feuilles = [sections[filiale_key]]
            years = set()
            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue
                for nom_flux, col_start in noms_colonnes:
                    try:
                        dates, _, _, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        for d in dates:
                            y = _year_of(d)
                            if y is not None:
                                years.add(y)
                    except Exception:
                        pass
            return sorted(years)

        def _compute_metric(prev, reel, metric: str):
            if metric == "MAPE (%)":
                denom = np.where(np.abs(reel) < 1e-9, 1.0, np.abs(reel))
                return np.abs(prev - reel) / denom * 100.0
            elif metric == "MAE (k€)":
                return np.abs(prev - reel) / 1000.0
            elif metric == "RMSE (k€)":
                return ( (prev - reel) ** 2 ) / (1000.0**2)
            else:
                return np.abs(prev - reel) / np.where(np.abs(reel) < 1e-9, 1.0, np.abs(reel)) * 100.0

        # =============== BARRE D’OUTILS (EN HAUT) ===============
        toolbar = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        toolbar.pack(fill="x", padx=10, pady=(10, 0))

        # Retour
        ctk.CTkButton(
            toolbar, text="⬅️ Retour au menu", command=self.retour_menu,
            width=180, height=40, corner_radius=15, fg_color="#444", hover_color="#666",
            text_color="white", font=("Segoe UI", 13, "bold")
        ).pack(side="left", padx=(0, 16))

        # Filiale
        def _label(parent, txt):
            return ctk.CTkLabel(parent, text=txt, text_color="white", font=("Segoe UI", 12, "bold"), fg_color="#00122e")

        _label(toolbar, "Filiale :").pack(side="left", padx=(0, 6))
        filiales = ["Toutes filiales"] + list(sections.keys())
        filiale_var = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(toolbar, values=filiales, textvariable=filiale_var, state="readonly", width=28)
        filiale_box.pack(side="left", padx=(0, 16))

        # Année (OBLIGATOIRE, pas de "Toutes années")
        _label(toolbar, "Année :").pack(side="left", padx=(0, 6))
        annee_var = tk.StringVar(value="")
        annee_box = ttk.Combobox(toolbar, values=[], textvariable=annee_var, state="readonly", width=12)
        annee_box.pack(side="left", padx=(0, 16))

        # Choix du métrique
        _label(toolbar, "Métrique :").pack(side="left", padx=(0, 6))
        metric_var = tk.StringVar(value="MAPE (%)")
        metric_box = ttk.Combobox(toolbar, values=["MAPE (%)", "MAE (k€)", "RMSE (k€)"], textvariable=metric_var,
                                state="readonly", width=12)
        metric_box.pack(side="left")

        # =============== FRAME GRAPHIQUE ===============
        graph_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=10)
        graph_frame.pack(pady=16, padx=16, fill="both", expand=True)

        # =============== TABLEAU RÉCAP ===============
        table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        table_frame.pack(pady=(4, 16), fill="x")

        colonnes = ["Filiale", "Année", "Horizon (mois)", "Métrique", "Valeur", "Nb points"]
        tree = ttk.Treeview(table_frame, columns=colonnes, show="headings", height=10)
        for col in colonnes:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=160)
        tree.pack(pady=5, padx=5, fill="x")

        # =============== BOUTONS ACTION ===============
        actions_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        actions_frame.pack(pady=(0, 12), fill="x")

        export_csv_btn = ctk.CTkButton(actions_frame, text="Exporter CSV", fg_color="#28B463", hover_color="#1E8449")
        export_png_btn = ctk.CTkButton(actions_frame, text="Exporter PNG", fg_color="#3498DB", hover_color="#2E86C1")
        export_csv_btn.pack(side="left", padx=(16, 8))
        export_png_btn.pack(side="left", padx=(8, 8))

        # =============== CALCUL & AFFICHAGE ===============
        last_df_shown = {"precision": None, "fig": None}

        def _load_backtest_rows(filiale_sel: str, annee_sel: int) -> pd.DataFrame:
            """
            Retourne un DF lignes élémentaires pour backtest:
            cols: filiale (code feuille), annee (str), horizon (int), prev, reel
            """
            records = []
            feuilles = list(sections.values()) if filiale_sel == "Toutes filiales" else [sections[filiale_sel]]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                # pour chaque flux, lit toutes les prévisions; on filtre l'année ensuite
                for nom_flux, col_start in noms_colonnes:
                    try:
                        dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    max_m = _profil_max_m(noms_profils)
                    previsions_len = len(previsions)

                    for i, d in enumerate(dates):
                        y = _year_of(d)
                        if y is None or y != annee_sel:
                            continue

                        r = _to_number(reel[i] if i < len(reel) else None)
                        if r is None:
                            continue
                        r = 1.0 if abs(r) < 1e-9 else r  # évite /0

                        for p_idx, p_list in enumerate(previsions):
                            pv = _to_number(p_list[i] if i < len(p_list) else None)
                            if pv is None:
                                continue

                            # horizon: basé sur nom + structure
                            p_name = noms_profils[p_idx] if p_idx < len(noms_profils) else f"P{p_idx+1}"
                            # le fallback utilise la "profondeur restante"
                            horizon = _horizon_from_name(p_name, previsions_len - p_idx, max_m)

                            records.append({
                                "filiale": feuille,
                                "annee": str(annee_sel),
                                "horizon": int(horizon),
                                "prev": float(pv),
                                "reel": float(r),
                                "flux": nom_flux,           # utile si besoin d’affiner plus tard
                                "profil": str(p_name)       # debug / audit
                            })
            return pd.DataFrame(records)

        def _aggregate_precision(df_raw: pd.DataFrame, metric_name: str) -> pd.DataFrame:
            if df_raw.empty:
                return df_raw

            df = df_raw.copy()
            # métrique élémentaire
            base = _compute_metric(df["prev"].values, df["reel"].values, metric_name)
            if metric_name == "RMSE (k€)":
                # on agrège la MSE et on racine après groupby
                df["metric_base"] = base  # (k€)^2
                grouped = df.groupby(["filiale", "annee", "horizon"]).agg(
                    metric=("metric_base", "mean"),
                    N=("metric_base", "size")
                ).reset_index()
                grouped["metric"] = np.sqrt(grouped["metric"])  # RMSE
            else:
                df["metric_base"] = base  # MAPE% ou MAE (k€)
                grouped = df.groupby(["filiale", "annee", "horizon"]).agg(
                    metric=("metric_base", "mean"),
                    N=("metric_base", "size")
                ).reset_index()
                grouped["metric"] = grouped["metric"]

            grouped = grouped.sort_values(["annee", "filiale", "horizon"], ascending=[True, True, False])
            return grouped

        def _refresh_graph_and_table():
            # Nettoyer graph + table
            for w in graph_frame.winfo_children():
                w.destroy()
            for r in tree.get_children():
                tree.delete(r)

            filiale = filiale_var.get()
            metric_name = metric_var.get()
            try:
                annee = int(annee_var.get())
            except Exception:
                tk.Label(graph_frame, text="Sélectionnez une année valide.", bg="#00122e", fg="white").pack()
                return

            df_raw = _load_backtest_rows(filiale, annee)
            if df_raw.empty:
                tk.Label(graph_frame, text="Aucune donnée disponible pour ces filtres.", bg="#00122e", fg="white").pack()
                last_df_shown["precision"] = None
                last_df_shown["fig"] = None
                return

            precision = _aggregate_precision(df_raw, metric_name)
            if precision.empty:
                tk.Label(graph_frame, text="Pas de points agrégés.", bg="#00122e", fg="white").pack()
                last_df_shown["precision"] = None
                last_df_shown["fig"] = None
                return

            # -------- GRAPHIQUE : lignes par filiale pour l'année sélectionnée --------
            plot_df = precision[precision["annee"] == str(annee)].copy()
            fig, ax = plt.subplots(figsize=(10, 6), facecolor="#00122e")
            ax.set_facecolor("#00122e")

            sns.lineplot(data=plot_df, x="horizon", y="metric", hue="filiale", marker="o", linewidth=2, ax=ax)

            ax.invert_xaxis()
            ax.set_xlabel("Horizon (mois avant la réalisation)", color="white")
            ylabel = "MAPE moyen (%)" if metric_name == "MAPE (%)" else metric_name
            ax.set_ylabel(ylabel, color="white")
            ax.grid(alpha=0.3)
            ax.tick_params(colors="white")

            # Légende : noms humains si possible
            handles, labels = ax.get_legend_handles_labels()
            new_labels = []
            for lab in labels:
                # lab est code feuille -> on le remappe au libellé si dispo
                for k, v in sections.items():
                    if v == lab:
                        new_labels.append(k)
                        break
                else:
                    new_labels.append(lab)
            ax.legend(handles, new_labels, title="Filiale", facecolor="#00122e", edgecolor="white", labelcolor="white", title_fontsize=10)

            title_suffix = f" – {annee}"
            ax.set_title(f"Précision par filiale ({metric_name}){title_suffix}", color="white", fontsize=14)

            canvas_fig = FigureCanvasTkAgg(fig, master=graph_frame)
            canvas_fig.draw()
            canvas_widget = canvas_fig.get_tk_widget()
            canvas_widget.pack(pady=10, fill="both", expand=True)
            plt.close(fig)

            # -------- TABLEAU RÉCAP --------
            for _, row in plot_df.sort_values(["filiale", "horizon"], ascending=[True, False]).iterrows():
                # Remap pour affichage
                fil = row["filiale"]
                for k, v in sections.items():
                    if v == fil:
                        fil_aff = k
                        break
                else:
                    fil_aff = fil

                val_disp = f"{row['metric']:.1f}" if metric_name == "MAPE (%)" else f"{row['metric']:.2f}"
                tree.insert("", "end", values=(
                    fil_aff,
                    row["annee"],
                    int(row["horizon"]),
                    metric_name,
                    val_disp,
                    int(row["N"])
                ))

            # TOTAL pondéré
            tot = plot_df["N"].sum()
            if tot > 0:
                if metric_name == "RMSE (k€)":
                    # moyenne pondérée déjà faite, mais on peut afficher la moyenne simple pondérée :
                    weighted = np.average(plot_df["metric"], weights=plot_df["N"])
                    total_val = f"{weighted:.2f}"
                else:
                    weighted = np.average(plot_df["metric"], weights=plot_df["N"])
                    total_val = f"{weighted:.1f}" if metric_name == "MAPE (%)" else f"{weighted:.2f}"
            else:
                total_val = "-"

            tree.insert("", "end",
                        values=("TOTAL", str(annee), "-", metric_name, total_val, int(tot)),
                        tags=("total",))
            tree.tag_configure("total", background="#444", foreground="white", font=('Segoe UI', 12, 'bold'))

            # Mémorise pour export
            last_df_shown["precision"] = plot_df.copy()
            last_df_shown["fig"] = canvas_fig

        # =============== CALLBACKS ===============
        def _on_filiale_change(*_):
            filiale = filiale_var.get()
            years = _annees_pour_filiale_cached(filiale)
            if not years:
                annee_box.config(values=[])
                annee_var.set("")
                messagebox.showinfo("Info", "Aucune année disponible pour cette filiale.")
                # Nettoyage affichages
                for w in graph_frame.winfo_children():
                    w.destroy()
                for r in tree.get_children():
                    tree.delete(r)
                return
            vals = [str(y) for y in years]
            annee_box.config(values=vals)
            annee_var.set(vals[-1])  # dernière année disponible
            _refresh_graph_and_table()

        def _on_annee_change(*_):
            _refresh_graph_and_table()

        def _on_metric_change(*_):
            _refresh_graph_and_table()

        def _export_csv():
            df = last_df_shown["precision"]
            if df is None or df.empty:
                messagebox.showinfo("Export CSV", "Aucune donnée à exporter pour ces filtres.")
                return
            # Remap filiale -> libellé humain pour export
            df = df.copy()
            df["Filiale"] = df["filiale"].apply(lambda code: next((k for k, v in sections.items() if v == code), code))
            out = df[["Filiale", "annee", "horizon", "metric", "N"]].rename(columns={
                "annee": "Année", "horizon": "Horizon (mois)", "metric": metric_var.get(), "N": "Nb points"
            })
            default_name = f"backtest_multi_horizon_{filiale_var.get().replace(' ','_')}_{annee_var.get()}_{metric_var.get().split()[0]}.csv"
            path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=default_name,
                                                filetypes=[("CSV","*.csv")])
            if not path:
                return
            try:
                out.to_csv(path, index=False, sep=";")
                messagebox.showinfo("Export CSV", f"Fichier exporté :\n{path}")
            except Exception as e:
                messagebox.showerror("Export CSV", f"Échec de l'export : {e}")

        def _export_png():
            fig_canvas = last_df_shown["fig"]
            if fig_canvas is None:
                messagebox.showinfo("Export PNG", "Rien à exporter pour le moment.")
                return
            default_name = f"backtest_multi_horizon_{filiale_var.get().replace(' ','_')}_{annee_var.get()}_{metric_var.get().split()[0]}.png"
            path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=default_name,
                                                filetypes=[("Image PNG","*.png")])
            if not path:
                return
            try:
                fig_canvas.figure.savefig(path, dpi=160, bbox_inches="tight")
                messagebox.showinfo("Export PNG", f"Image exportée :\n{path}")
            except Exception as e:
                messagebox.showerror("Export PNG", f"Échec de l'export : {e}")

        filiale_box.bind("<<ComboboxSelected>>", lambda e: _on_filiale_change())
        annee_box.bind("<<ComboboxSelected>>", lambda e: _on_annee_change())
        metric_box.bind("<<ComboboxSelected>>", lambda e: _on_metric_change())
        export_csv_btn.configure(command=_export_csv)
        export_png_btn.configure(command=_export_png)

        # =============== AFFICHAGE INITIAL ===============
        _on_filiale_change()

        # =============== SCROLL (souris) ===============
        def _on_mousewheel(event):
            if event.delta == 0: return "break"
            step = -1 if event.delta > 0 else 1
            canvas.yview_scroll(step, "units"); return "break"
        def _on_linux_scroll_up(event): canvas.yview_scroll(-1, "units"); return "break"
        def _on_linux_scroll_down(event): canvas.yview_scroll(1, "units"); return "break"
        def _on_mousewheel_shift(event):
            if event.delta == 0: return "break"
            step = -1 if event.delta > 0 else 1
            canvas.xview_scroll(step, "units"); return "break"
        def _bind_mousewheel(_event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")
        def _unbind_mousewheel(_event=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Shift-MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")
        canvas.bind("<Enter>", _bind_mousewheel, add="+")
        canvas.bind("<Leave>", _unbind_mousewheel, add="+")

#===================Page Graphique des profils===================
    def _annees_disponibles(self, section, flux_name):
        """Retourne la liste triée des années disponibles pour (section, flux)."""
        B = CACHE.get((section, flux_name))
        if not B or not B.get("dates"):
            return []
        return sorted({d.year for d in B["dates"]})

    def _token_from_flux(self, section, flux_name):
        """Retrouve col_start (token) à partir du nom de flux."""
        for name, tok in TOKENS.get(section, []):
            if name == flux_name:
                return tok
        return None

    def _profils_for_year(self, section, flux_name, annee):
        """
        Calcule les profils à afficher pour une année donnée : on ne garde que
        ceux qui ont au moins une valeur non-nulle/non-None sur l'année.
        Renvoie (noms_profils_filtrés, previsions_filtrées_existence_bool).
        """
        col_start = self._token_from_flux(section, flux_name)
        if col_start is None:
            return [], []

        dates, reel, previsions, noms_profils = extraire_valeurs(section, col_start, nb_prev, annee=annee)

        actifs = []
        for serie in previsions:  # une série par profil
            exist = any(v not in (None, 0, 0.0, "") for v in serie)
            actifs.append(exist)

        noms_ok = [np for np, ok in zip(noms_profils, actifs) if ok]
        return noms_ok, actifs

    def _on_flux_change(self, event=None):
        """Quand l'utilisateur choisit un flux : calcule les années disponibles, active la combo et met à jour les profils."""
        section = self.feuille_combo.get()
        flux_name = self.nom_combo.get()
        if not section or not flux_name:
            return

        annees = self._annees_disponibles(section, flux_name)
        if not annees:
            # Pas d'années disponibles → on vide l'UI profils
            self.annee_combo.config(values=[], state="disabled")
            self._rebuild_profils_ui(section, flux_name, None)
            return

        # Active la combobox et pré-sélectionne la plus récente
        self.annee_combo.config(values=[str(a) for a in annees], state="readonly")
        self.annee_combo.set(str(annees[-1]))

        # Recharge les profils pour l’année par défaut
        self._rebuild_profils_ui(section, flux_name, annees[-1])

    def _on_annee_change(self, event=None):
        """Quand l'année change : profils dynamiques mis à jour."""
        section = self.feuille_combo.get()
        flux_name = self.nom_combo.get()
        if not section or not flux_name:
            return
        try:
            annee = int(self.annee_combo.get())
        except Exception:
            annee = None
        self._rebuild_profils_ui(section, flux_name, annee)

    def _rebuild_profils_ui(self, section, flux_name, annee):
        """Reconstruit les checkboxes de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()
        self.vars_prev = []
        self.profils_names_order = []  # ✅ reset au début

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils  # ✅ mémorise l’ordre affiché

        # Si aucun profil
        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # ✅ Disposition sur 4 lignes max
        nb_lignes = 4
        nb_cols = max(1, len(noms_profils) // nb_lignes + (1 if len(noms_profils) % nb_lignes else 0))

        for i, nom_profil in enumerate(noms_profils):
            var = tk.BooleanVar(value=False)
            cb = tk.Checkbutton(
                self.profils_frame, text=nom_profil, variable=var,
                bg="#00122e", fg="white", font=('Segoe UI', 10),
                selectcolor="#00aced", activebackground="#003366",
                activeforeground="white"
            )
            # ✅ Répartition sur 4 lignes
            row = i % nb_lignes
            col = i // nb_lignes

            cb.grid(row=row, column=col, sticky="w", padx=12, pady=6)
            cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
            cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))
            self.vars_prev.append(var)

    def creer_page_graphique(self):
        import tkinter as tk
        from tkinter import ttk
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage

        self.vider_fenetre()

        # === HEADER + logo (inchangé) ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - VISUALISATION GRAPHIQUE DES ÉCARTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\...\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="#00aced")
        barre.pack(side="top", fill="x", pady=(0, 15))

        # ⚠️ Tu avais ici une boucle d'extraction fixe sur 2024, on la retire : ce sera fait après sélection.

        # === FRAME PRINCIPALE SCROLLABLE ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        self.main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
        self.scrollbar_y = tk.Scrollbar(container, orient="vertical", command=self.main_canvas.yview)
        self.scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=self.main_canvas.xview)
        self.scrollable_frame = ctk.CTkFrame(self.main_canvas, fg_color="#00122e", corner_radius=0)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.main_canvas.configure(yscrollcommand=self.scrollbar_y.set, xscrollcommand=self.scrollbar_x.set)

        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar_y.pack(side="right", fill="y")
        self.scrollbar_x.pack(side="bottom", fill="x")

        # === CONTENU ===
        self.center_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="#00122e", corner_radius=15)
        self.center_frame.pack(pady=20, padx=30, fill="both", expand=True)

        # --- Sélection filiale
        ttk.Label(self.center_frame, text="Sélectionnez une filiale :", background="#00122e", foreground="white",
                font=('Segoe UI', 12)).pack(pady=(10, 5))
        self.feuille_combo = ttk.Combobox(self.center_frame, values=list(sections.values()), state="readonly", width=35)
        self.feuille_combo.pack(pady=(0, 10))
        self.feuille_combo.bind("<<ComboboxSelected>>", self.charger_noms)  # suppose que ta méthode existe

        # --- Sélection flux
        ttk.Label(self.center_frame, text="Sélectionnez un flux :", background="#00122e", foreground="white",
                font=('Segoe UI', 12)).pack(pady=(5, 5))
        self.nom_combo = ttk.Combobox(self.center_frame, state="readonly", width=35)
        self.nom_combo.pack(pady=(0, 10))
        self.nom_combo.bind("<<ComboboxSelected>>", self._on_flux_change)  # ★ nouveau

        # --- Sélection année (nouveau)
        ttk.Label(self.center_frame, text="Sélectionnez une année :", background="#00122e", foreground="white",
                font=('Segoe UI', 12)).pack(pady=(5, 5))
        self.annee_combo = ttk.Combobox(self.center_frame, state="disabled", width=20)  # ★ désactivée tant que pas de flux
        self.annee_combo.pack(pady=(0, 15))
        self.annee_combo.bind("<<ComboboxSelected>>", self._on_annee_change)  # ★

        # --- Profils (dynamiques)
        self.profils_title = ttk.Label(self.center_frame, text="Profils disponibles :", background="#00122e",
                                    foreground="white", font=('Segoe UI', 12))
        self.profils_title.pack(pady=(10, 5))
        self.profils_frame = tk.Frame(self.center_frame, bg="#00122e")  # ★ on le garde en attribut
        self.profils_frame.pack(pady=10, padx=15, fill="x")
        self.vars_prev = []  # sera rempli dynamiquement

        # --- Checkbox Réel
        import tkinter as tk  # (au cas où)
        self.var_reel = tk.BooleanVar(value=True)
        chk_reel = tk.Checkbutton(self.center_frame, text="Réel", variable=self.var_reel,
                                bg="#00122e", fg="white", font=('Segoe UI', 11),
                                selectcolor="#00aced", activebackground="#003366",
                                activeforeground="white")
        chk_reel.pack(anchor="w", padx=20, pady=5)
        chk_reel.bind("<Enter>", lambda e, c=chk_reel: c.config(bg="#003366"))
        chk_reel.bind("<Leave>", lambda e, c=chk_reel: c.config(bg="#00122e"))

        # --- Boutons
        self.btn_afficher = ctk.CTkButton(self.center_frame, text="Afficher Graphique",
                                        command=self.afficher_graphique, width=240,
                                        fg_color="#00aced", hover_color="#0099e6",
                                        corner_radius=12)
        self.btn_afficher.pack(pady=(25, 10))

        self.btn_afficher_cumule = ctk.CTkButton(
            self.center_frame, text="Afficher Graphique Cumulé",
            command=self.afficher_graphique_cumule, width=240,
            fg_color="#0078D7", hover_color="#005A9E",
            corner_radius=12
        )
        self.btn_afficher_cumule.pack(pady=(10, 10))

        self.btn_retour = ctk.CTkButton(self.center_frame, text="Retour au menu",
                                        command=self.retour_menu, width=240,
                                        fg_color="#444", hover_color="#666",
                                        corner_radius=12)
        self.btn_retour.pack(pady=(5, 20))

        self.canvas = None

    def charger_noms(self, event=None):
        """Remplit la liste des flux pour la filiale sélectionnée et réinitialise année/profils."""
        section = self.feuille_combo.get()
        if not section:
            return

        ws, noms_colonnes = charger_donnees(section, taille_bloc)
        flux_disponibles = [name for (name, _tok) in noms_colonnes]

        self.nom_combo.config(values=flux_disponibles, state="readonly")
        self.nom_combo.set("")  # reset du flux

        # Réinitialisation des autres éléments
        self.annee_combo.config(values=[], state="disabled")
        self.annee_combo.set("")
        for w in self.profils_frame.winfo_children():
            w.destroy()
        self.vars_prev = []
        self.profils_names_order = []  # ✅ reset ordre profils
                                                                                                                                                                                       
    def afficher_graphique(self):
        import matplotlib.pyplot as plt
        import mplcursors
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import matplotlib.ticker as mticker
        import tkinter as tk

        # --- Préparation
        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        section = self.feuille_combo.get()
        if not section:
            messagebox.showwarning("Attention", "Veuillez sélectionner une filiale.")
            return

        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un flux.")
            return

        # Récupère le token/col_start depuis le nom de flux (évite dépendance à self.ws/self.noms_colonnes)
        if not hasattr(self, "_token_from_flux"):
            messagebox.showerror("Erreur", "Méthode _token_from_flux absente (voir patch précédent).")
            return
        col_start = self._token_from_flux(section, nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Flux sélectionné invalide.")
            return

        # Année sélectionnée (optionnelle)
        try:
            annee = int(self.annee_combo.get()) if self.annee_combo.get() else None
        except Exception:
            annee = None

        # --- Extraction des séries pour l'année
        dates, reel, previsions, noms_profils_complets = extraire_valeurs(section, col_start, nb_prev, annee=annee)

        # Vérif basique
        if not dates:
            messagebox.showinfo("Info", "Aucune donnée pour l'année sélectionnée.")
            return

        # Conversion en flux si nécessaire
        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i-1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux

            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        # --- Nettoyage ancien graphique
        if hasattr(self, "canvas") and self.canvas:
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
        if hasattr(self, "toolbar_frame") and self.toolbar_frame:
            self.toolbar_frame.destroy()
            self.toolbar_frame = None

        # === Figure ===
        plt.style.use("seaborn-v0_8-darkgrid")
        fig, ax = plt.subplots(figsize=(16, 8))
        palette = plt.cm.tab10.colors

        # --- Réel
        if getattr(self, "var_reel", None) is not None and self.var_reel.get():
            ax.plot(dates, reel, label="Réel", color="black", linewidth=2, marker="o")

        # --- Prévisions
        # On mappe les checkboxes (dynamiques) aux vraies séries par leur nom.
        # self.profils_names_order = liste affichée (filtres/année) dans _rebuild_profils_ui
        profils_affiches = getattr(self, "profils_names_order", None)
        if profils_affiches is None:
            # fallback : on suppose que les checkboxes correspondent 1:1 aux entêtes complets
            profils_affiches = noms_profils_complets[:len(self.vars_prev)]

        # index par nom dans la table complète
        index_by_name = {name: i for i, name in enumerate(noms_profils_complets)}

        plotted = 0
        for i, var in enumerate(self.vars_prev):
            if not var.get():
                continue
            if i >= len(profils_affiches):
                continue
            nom_profil = profils_affiches[i]
            idx = index_by_name.get(nom_profil, None)
            if idx is None or idx >= len(previsions):
                continue

            y = previsions[idx]
            label = nom_profil
            ax.plot(
                dates, y,
                label=label,
                alpha=0.9,
                linewidth=1.8,
                marker=".",
                color=palette[plotted % len(palette)]
            )
            plotted += 1

        # --- Axes & style
        titre_suffix = f" - {annee}" if annee is not None else ""
        ax.set_title(f"{section} - {nom_selectionne}{titre_suffix}", fontsize=18, fontweight="bold")
        ax.set_xlabel("Date", fontsize=12)
        ax.set_ylabel("Valeur (k€)" if nom_selectionne not in noms_a_convertir_flux else "Flux", fontsize=12)
        ax.legend(loc="upper left", bbox_to_anchor=(1, 1))
        ax.grid(True, linestyle="--", alpha=0.6)

        # Format des nombres (espaces pour milliers)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " ")))

        # Interaction
        mplcursors.cursor(ax, hover=True)

        # === Intégration Tkinter
        self.canvas = FigureCanvasTkAgg(fig, master=self.center_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(pady=20)

        self.toolbar_frame = tk.Frame(self.center_frame, bg="#00122e")
        self.toolbar_frame.pack(pady=(5, 10))
        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()

    def afficher_graphique_cumule(self):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import tkinter as tk
        import matplotlib.ticker as mticker
        import pandas as pd
        import numpy as np
        import mplcursors

        # --- Sélection / validation ---
        section = self.feuille_combo.get()
        if not section:
            messagebox.showwarning("Attention", "Veuillez sélectionner une filiale.")
            return

        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un flux.")
            return

        # Résoudre col_start via le helper (évite self.ws / self.noms_colonnes)
        if not hasattr(self, "_token_from_flux"):
            messagebox.showerror("Erreur", "Méthode _token_from_flux absente (voir patch précédent).")
            return
        col_start = self._token_from_flux(section, nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Flux sélectionné invalide.")
            return

        # Année sélectionnée (optionnelle)
        try:
            annee = int(self.annee_combo.get()) if self.annee_combo.get() else None
        except Exception:
            annee = None

        # --- Lecture des séries (filtrées par année) ---
        dates, reel, previsions, noms_profils_complets = extraire_valeurs(section, col_start, nb_prev, annee=annee)
        if not dates:
            messagebox.showinfo("Info", "Aucune donnée pour l'année sélectionnée.")
            return

        # --- Conversion en flux si nécessaire ---
        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement", "Endettement Net"
        ]
        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i-1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux
            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        # --- Nettoyage ancien rendu ---
        if hasattr(self, "canvas") and self.canvas:
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
        if hasattr(self, "toolbar_frame") and self.toolbar_frame:
            self.toolbar_frame.destroy()
            self.toolbar_frame = None

        # --- DataFrame de base (quotidien) ---
        df = pd.DataFrame({"Date": dates, "Réel": reel})
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Mois"] = df["Date"].dt.strftime("%Y-%m")

        # --- Profils cochés (mapping dynamique par nom) ---
        profils_affiches = getattr(self, "profils_names_order", None)
        if profils_affiches is None:
            profils_affiches = noms_profils_complets[:len(self.vars_prev)]  # fallback

        index_by_name = {name: i for i, name in enumerate(noms_profils_complets)}

        profils_selectionnes = []
        for i, var in enumerate(self.vars_prev):
            if not var.get():
                continue
            if i >= len(profils_affiches):
                continue
            nom_prof = profils_affiches[i]
            idx = index_by_name.get(nom_prof, None)
            if idx is None or idx >= len(previsions):
                continue
            df[nom_prof] = previsions[idx]
            profils_selectionnes.append(nom_prof)

        # --- Debug console (optionnel) ---
        print("\n=== CALCUL CUMULÉ (par profil & par mois) ===")
        titre_flux = f"{section} - {nom_selectionne}" + (f" - {annee}" if annee is not None else "")
        print(f"Flux : {titre_flux}")
        print(f"Profils cochés : {profils_selectionnes}\n")

        # --- Réel de base (impression mensuelle) ---
        print("--- Réel de base ---")
        for mois, g in df.groupby("Mois", dropna=True):
            nb_j_reel = int(g["Réel"].notna().sum())
            val_reel = float(g["Réel"].sum()) if nb_j_reel > 0 else 0.0
            print(f"Mois {mois} | Réel total = {val_reel} (jours={nb_j_reel})")
        print()

        # --- Combinaison par profil : Prévision sinon Réel + masquage des mois sans aucune prévision ---
        series_combinees = {}
        for nom_prof in profils_selectionnes:
            comb = pd.Series(df[nom_prof] if nom_prof in df.columns else np.nan)
            comb = comb.where(~comb.isna(), df["Réel"])  # prev sinon réel

            # Mois ayant au moins 1 prévision
            tmp_prev = pd.DataFrame({"Mois": df["Mois"], "Prev": df.get(nom_prof)})
            mois_avec_prev = set(tmp_prev.loc[tmp_prev["Prev"].notna(), "Mois"].dropna().unique())
            mois_tous = set(df["Mois"].dropna().unique())
            mois_masques = sorted(mois_tous - mois_avec_prev)

            # Masquer les mois sans prévision (entièrement)
            mask_keep = df["Mois"].isin(mois_avec_prev)
            comb = comb.where(mask_keep, np.nan)

            # Debug détaillé
            if mois_masques:
                print(f"[{nom_prof}] Mois masqués (0 prévision) : {', '.join(mois_masques)}")
            print(f"--- Profil : {nom_prof} ---")
            tmp = pd.DataFrame({
                "Mois": df["Mois"],
                "Reel": df["Réel"],
                "Prev": df.get(nom_prof),
            })
            tmp = tmp[tmp["Mois"].isin(mois_avec_prev)]
            if tmp.empty:
                print("(Aucun mois avec prévision — ce profil n'apparaîtra pas sur le graphe.)\n")
            else:
                tmp["has_prev"] = tmp["Prev"].notna()
                tmp["use_reel_equil"] = (~tmp["has_prev"]) & (tmp["Reel"].notna())
                for mois, g in tmp.groupby("Mois", dropna=True):
                    nb_j_prev = int(g["has_prev"].sum())
                    val_prev = float(g.loc[g["has_prev"], "Prev"].sum()) if nb_j_prev > 0 else 0.0
                    nb_j_reel_eq = int(g["use_reel_equil"].sum())
                    val_reel_eq = float(g.loc[g["use_reel_equil"], "Reel"].sum()) if nb_j_reel_eq > 0 else 0.0
                    nb_j_reel = int(g["Reel"].notna().sum())
                    val_reel = float(g["Reel"].sum()) if nb_j_reel > 0 else 0.0
                    total = val_prev + val_reel_eq
                    print(
                        f"Mois {mois} | Réel pur = {val_reel} (jours={nb_j_reel}) | "
                        f"Prévision = {val_prev} (jours={nb_j_prev}) | "
                        f"Réel ajouté (équilibrage) = {val_reel_eq} (jours={nb_j_reel_eq}) | "
                        f"Total combiné = {total}"
                    )
                print()

            series_combinees[f"{nom_prof} (Prévision sinon Réel)"] = comb

        # --- Passage au mensuel ---
        df_comb = pd.DataFrame({"Mois": df["Mois"], "Réel": df["Réel"]})
        for nom_serie, s in series_combinees.items():
            df_comb[nom_serie] = s
        df_cumule = df_comb.groupby("Mois", as_index=True).sum(min_count=1)

        # --- Tracé ---
        plt.style.use("seaborn-v0_8-darkgrid")
        fig, ax = plt.subplots(figsize=(14, 7))

        n_mois = len(df_cumule)
        n_series = 1 + len(series_combinees)  # +1 pour Réel
        positions = np.arange(n_mois)
        largeur_barre = 0.8 / max(1, n_series)
        palette = plt.cm.tab10.colors
        bar_containers = []

        # Réel en noir
        bars_reel = ax.bar(
            positions,
            df_cumule["Réel"].fillna(0).values,
            width=largeur_barre,
            label="Réel",
            color="black",
        )
        bar_containers.append(bars_reel)

        # Profils
        for i, (nom_serie, _) in enumerate(series_combinees.items()):
            offset = (i + 1) * largeur_barre
            bars = ax.bar(
                positions + offset,
                df_cumule[nom_serie].fillna(0).values,  # mois sans prev → NaN → somme=0
                width=largeur_barre,
                label=nom_serie,
                color=palette[i % len(palette)],
            )
            bar_containers.append(bars)

        # Axes & style
        ax.set_xticks(positions + largeur_barre * (n_series - 1) / 2)
        ax.set_xticklabels(df_cumule.index, rotation=45)

        titre_suffix = f" - {annee}" if annee is not None else ""
        ax.set_title(f"{section} - {nom_selectionne}{titre_suffix}", fontsize=16, fontweight="bold")
        ax.set_xlabel("Mois", fontsize=12)
        ax.set_ylabel("Valeur cumulée (k€)", fontsize=12)
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " ")))
        ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))

        # Hover interactif
        cursor = mplcursors.cursor(bar_containers, hover=True)
        @cursor.connect("add")
        def on_add(sel):
            bars = sel.artist
            idx = sel.index
            serie = bars.get_label()
            val = bars.datavalues[idx]
            mois = df_cumule.index[idx]
            sel.annotation.set_text(
                f"Série : {serie}\nMois : {mois}\nValeur : {int(val):,} K€".replace(",", " ")
            )
            sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

        # Intégration Tkinter
        self.canvas = FigureCanvasTkAgg(fig, master=self.center_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(pady=10, fill="x", expand=True)

        self.toolbar_frame = tk.Frame(self.center_frame, bg="#00122e")
        self.toolbar_frame.pack()
        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()

#===================Page Prédiction IA===================
    def _rebuild_profils_ui2(self, section, flux_name, annee):
        """Reconstruit la select-box de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()

        # nouveau : on n'utilise plus self.vars_prev ici
        self.vars_prev = []
        self.profils_names_order = []   # ordre des profils pour mapping
        self.profils_listbox = None     # widget de sélection multiple

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils

        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # Label d'info
        lbl = tk.Label(
            self.profils_frame,
            text="Profils (Ctrl+clic pour multi-sélection) :",
            bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
        )
        lbl.pack(anchor="w", padx=4, pady=(0, 4))

        # Frame contenant la listbox + scrollbar
        container = tk.Frame(self.profils_frame, bg="#00122e")
        container.pack(fill="x", padx=4, pady=(0, 6))

        scrollbar = tk.Scrollbar(container, orient="vertical")
        self.profils_listbox = tk.Listbox(
            container,
            selectmode="multiple",
            exportselection=False,
            bg="#00122e",
            fg="white",
            font=('Segoe UI', 10),
            height=min(8, len(noms_profils)),
            highlightthickness=0,
            activestyle="dotbox"
        )
        self.profils_listbox.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.profils_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.profils_listbox.yview)

        for name in noms_profils:
            self.profils_listbox.insert("end", name)

        # petite sélection par défaut : tout sélectionner
        self.profils_listbox.select_set(0, "end")

    def creer_page_ia_prediction(self):
        """
        Page IA de prévision Enc. Autres Produits (N -> N+1).
        Cette méthode ne fait plus que :
        1) Vérifier les dépendances
        2) Réinitialiser la fenêtre
        3) Construire le dataset global
        4) Construire toute la page IA (UI + callbacks + entraînement)
        """
        # 1) Imports (idéalement à mettre en haut du fichier Python)
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # 2) sécurité sklearn
        if RandomForestRegressor is None:
            messagebox.showerror(
                "Dépendance manquante",
                "Le module 'scikit-learn' n'est pas installé.\n"
                "Installe-le avec : pip install scikit-learn"
            )
            return

        # 3) reset page
        self.vider_fenetre()

        # 4) construire le dataset global (toutes filiales, tous jours)
        df = self._ia_build_dataset_enc_autres_produits()
        if df is None:
            # _ia_build_dataset_enc_autres_produits affiche déjà le message et fait retour_menu()
            return

        # 5) construire la page IA complète (UI + callbacks + entraînement)
        self._ia_build_prediction_page(df)
        
    def _ia_build_dataset_enc_autres_produits(self):
        """
        Construit le DataFrame global df sur le flux 'Enc. Autres Produits'
        pour toutes les filiales, avec colonnes :
        section, date, y, year, month, dayofyear, section_id, roll_mean_7, roll_mean_30
        Retourne df ou None si aucune donnée.
        """
        import pandas as pd
        import numpy as np
        from tkinter import messagebox

        flux_cible = "Trafic Voyageurs"
        lignes = []

        # ---- helpers numériques ----
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        # ---- construction dataset (reprends ton code) ----
        for section_name, feuille in sections.items():
            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees pour {feuille} : {e}")
                continue

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                continue

            _, col_start = cible[0]
            try:
                dates, reel, _previsions, _noms_profils = extraire_valeurs(
                    ws, col_start, nb_prev, annee=None
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs pour {section_name}/{flux_cible} : {e}")
                continue

            for d, r in zip(dates, reel):
                y_val = _to_number(r)
                if y_val is None:
                    continue
                try:
                    d_ts = pd.to_datetime(d)
                except Exception:
                    continue
                lignes.append({
                    "section": section_name,
                    "date": d_ts,
                    "y": float(y_val)
                })

        if not lignes:
            messagebox.showinfo(
                "Information",
                "Aucune donnée trouvée pour le flux 'Enc. Autres Produits' sur les filiales."
            )
            self.retour_menu()
            return None

        df = pd.DataFrame(lignes).sort_values(["section", "date"]).reset_index(drop=True)
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["dayofyear"] = df["date"].dt.dayofyear

        # Encodage des filiales
        cat = df["section"].astype("category")
        df["section_id"] = cat.cat.codes

        # Lissages
        df["roll_mean_7"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(7, min_periods=1).mean()
        )
        df["roll_mean_30"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(30, min_periods=1).mean()
        )

        return df

    def _ia_build_prediction_page(self, df):
        """
        Construit toute l'interface IA + callbacks d'entraînement
        à partir du DataFrame global df déjà préparé.
        """
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # ---------- helpers numériques locaux ----------
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_float_or_nan(x):
            val = _to_number(x)
            return float(val) if val is not None else np.nan

        flux_cible = "Trafic Voyageurs"
        default_n_estimators = 200

        # ---------- variables profils & graph 2 ----------
        ia_profils_vars = []        # liste de tk.BooleanVar
        ia_profils_names = []       # noms de profils actifs N+1
        ia_profils_dates = []       # dates N+1
        ia_profils_series = []      # séries (une par profil actif)

        current_pred_df = None          # df_future_all pour N+1
        current_real_target_df = None   # réel sur N+1 (si dispo)
        current_target_year = None
        current_filiale_name = None

        ia_graph2_widget = None
        graph2_container_packed = False

        exported_pred_df = None         # pour export Excel prédictions
        analysis_table_frame = None
        analysis_tree = None
        analysis_export_button = None
        export_button = None            # bouton d’export des prédictions

        # ============ HEADER ============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - IA PRÉDICTION ENC. AUTRES PRODUITS (N → N+1)",
            font=titre_font,
            text_color="white"
        ).pack(side="left", anchor="w")

        # Logo
        try:
            logo_img = Image.open(image_path)
            test = tk.Label(self, text="Test", font=titre_font)
            test.update_idletasks()
            text_h = test.winfo_reqheight()
            test.destroy()
            ratio = logo_img.width / max(logo_img.height, 1)
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"[IA] Erreur chargement logo IA : {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ============ CONTAINER SCROLLABLE ============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # ============ TITRE / TEXTE ============
        ctk.CTkLabel(
            scrollable_frame,
            text="IA - Random Forest sur le flux 'Enc. Autres Produits' (prévision année N+1)",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        ).pack(pady=15)

        ctk.CTkLabel(
            scrollable_frame,
            text=(
                "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                "Il apprend comment la valeur de 'Enc. Autres Produits' évolue d'une année à l'autre "
                "(N → N+1) pour chaque jour de l'année.\n"
                "Il prédit ensuite l’ensemble de l’année N+1 pour cette filiale.\n"
                "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
            ),
            font=("Segoe UI", 12),
            text_color="#c9defa",
            justify="left"
        ).pack(pady=(0, 10))

        # ============ ZONE PARAMÈTRES (CARD) ============
        params_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        params_frame.pack(fill="x", padx=10, pady=(5, 10))

        # Filiale
        ctk.CTkLabel(
            params_frame,
            text="Filiale :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))

        filiales = sorted(df["section"].unique().tolist())
        selected_filiale = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(
            params_frame,
            textvariable=selected_filiale,
            values=filiales,
            state="readonly",
            width=28
        )
        filiale_box.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        # Année N
        ctk.CTkLabel(
            params_frame,
            text="Année N (historique → N+1) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=1, sticky="w", padx=12, pady=(10, 2))

        years_sorted = sorted(df["year"].unique().tolist())
        annees_var = tk.StringVar(value=str(years_sorted[-1]))
        annees_box = ttk.Combobox(
            params_frame,
            textvariable=annees_var,
            values=[str(y) for y in years_sorted],
            state="readonly",
            width=18
        )
        annees_box.grid(row=1, column=1, sticky="w", padx=12, pady=(0, 10))

        # Hyperparamètres RF
        ctk.CTkLabel(
            params_frame,
            text="Hyperparamètres Random Forest (point de départ) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=2, sticky="w", padx=12, pady=(10, 2))

        # n_estimators
        ctk.CTkLabel(
            params_frame,
            text="n_estimators",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=2, sticky="w", padx=12, pady=(0, 2))

        n_estimators_var = tk.IntVar(value=default_n_estimators)
        slider_n = ctk.CTkSlider(
            params_frame,
            from_=50,
            to=600,
            number_of_steps=11,
            command=lambda v: n_estimators_var.set(int(v))
        )
        slider_n.set(default_n_estimators)
        slider_n.grid(row=2, column=2, sticky="we", padx=12, pady=(0, 8))

        lbl_n = ctk.CTkLabel(
            params_frame,
            text=f"{default_n_estimators}",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_n.grid(row=3, column=2, sticky="w", padx=12, pady=(0, 6))

        def _update_lbl_n(_=None):
            lbl_n.configure(text=str(n_estimators_var.get()))
        slider_n.configure(command=lambda v: (n_estimators_var.set(int(v)), _update_lbl_n()))

        # max_depth
        ctk.CTkLabel(
            params_frame,
            text="max_depth",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=3, sticky="w", padx=12, pady=(0, 2))

        max_depth_var = tk.IntVar(value=15)
        slider_d = ctk.CTkSlider(
            params_frame,
            from_=3,
            to=25,
            number_of_steps=11,
            command=lambda v: max_depth_var.set(int(v))
        )
        slider_d.set(15)
        slider_d.grid(row=2, column=3, sticky="we", padx=12, pady=(0, 8))

        lbl_d = ctk.CTkLabel(
            params_frame,
            text="15",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_d.grid(row=3, column=3, sticky="w", padx=12, pady=(0, 2))

        def _update_lbl_d(_=None):
            lbl_d.configure(text=str(max_depth_var.get()))
        slider_d.configure(command=lambda v: (max_depth_var.set(int(v)), _update_lbl_d()))

        use_depth_var = tk.BooleanVar(value=True)
        chk_depth = ctk.CTkCheckBox(
            params_frame,
            text="Limiter la profondeur",
            variable=use_depth_var,
            text_color="#c9defa",
            font=("Segoe UI", 11)
        )
        chk_depth.grid(row=4, column=3, sticky="w", padx=12, pady=(0, 10))

        # Bouton entraîner
        bouton_train = ctk.CTkButton(
            params_frame,
            text="🚀 Entraîner / ré-entraîner le modèle",
            width=240,
            height=40,
            corner_radius=16,
            fg_color="#008C4B",
            hover_color="#006C39",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_train.grid(row=1, column=4, rowspan=3, padx=14, pady=(0, 10), sticky="e")

        for c in range(5):
            params_frame.grid_columnconfigure(c, weight=1 if c in (2, 3) else 0)

        # ============ CARD PROFILS N+1 (CASES À COCHER) ============
        profils_card = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        profils_card.pack(fill="x", padx=10, pady=(0, 15))

        ctk.CTkLabel(
            profils_card,
            text="Profils de prévision disponibles sur l'année N+1 :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=12, pady=(10, 4))

        ia_profils_frame = tk.Frame(profils_card, bg="#00122e")
        ia_profils_frame.pack(fill="x", padx=12, pady=(0, 10))

        # ============ ZONE RÉSULTATS ============
        graph_widgets = []   # kpi, graphe 1, tableaux
        graph2_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe 2
        monthly_container = tk.Frame(scrollable_frame, bg="#00122e")  # conteneur stable pour le graphe mensuel
        monthly_graph_widget = None
        monthly_container_packed = False

        # ---------- helpers graphiques / profils ----------

        def _clear_graph_widgets():
            """Supprime les widgets graphiques (sauf le conteneur du graphe 2)."""
            nonlocal graph_widgets
            for w in graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            graph_widgets = []

        def _redraw_graph2():
            """Redessine le graphe détaillé N+1 avec :
            - Réel N+1 (si dispo)
            - Prévision IA N+1
            - Profils cochés N+1
            """
            nonlocal ia_graph2_widget, graph2_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Si rien à tracer et aucun profil, on ne fait rien
            if current_pred_df is None and current_real_target_df is None and not ia_profils_names:
                return

            # S'assurer que le conteneur est packé une fois pour toutes
            if not graph2_container_packed:
                graph2_container.pack(pady=10, fill="both", expand=True)
                graph2_container_packed = True

            # Nettoyer le conteneur
            for child in graph2_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            ia_graph2_widget = None

            # Créer la figure
            fig2, ax2 = plt.subplots(
                figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True
            )
            ax2.set_facecolor("#00122e")

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            # Réel N+1
            if current_real_target_df is not None and not current_real_target_df.empty:
                ax2.plot(
                    current_real_target_df["date"], current_real_target_df["y"],
                    label=f"Réel {target_year}", linewidth=2, color="#5DADE2"
                )

            # Prévision IA N+1
            if current_pred_df is not None and not current_pred_df.empty:
                ax2.plot(
                    current_pred_df["date"], current_pred_df["pred_value"],
                    label=f"Prévision IA {target_year}",
                    linewidth=2, linestyle="--", color="#F4D03F"
                )

            # Profils cochés N+1
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = [pd.to_datetime(d) for d in ia_profils_dates]
                palette = plt.cm.tab10.colors
                prof_idx = 0
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    y_prof = [_to_float_or_nan(v) for v in serie]
                    ax2.plot(
                        dates_prof, y_prof,
                        label=f"Profil '{name}' {target_year}",
                        linewidth=1.8,
                        linestyle="-.",
                        color=palette[prof_idx % len(palette)]
                    )
                    prof_idx += 1

            filiale_for_title = current_filiale_name or selected_filiale.get()
            titre2 = f"Enc. Autres Produits – année {target_year} – {filiale_for_title}"
            ax2.set_title(titre2, color="white", fontsize=14)
            ax2.set_xlabel("Date", color="white", fontsize=12)
            ax2.set_ylabel("Montant (K€)", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white", rotation=30)
            ax2.tick_params(axis='y', colors="white")
            ax2.legend(facecolor="#00122e", edgecolor="white", labelcolor="white")

            canvas_fig2 = FigureCanvasTkAgg(fig2, master=graph2_container)
            canvas_fig2.draw()
            ia_graph2_widget = canvas_fig2.get_tk_widget()
            ia_graph2_widget.pack(fill="both", expand=True)
            plt.close(fig2)
        
        def _redraw_monthly_graph():
            """
            Graphe mensuel N+1 :
            - Barres : Réel N+1 (si dispo) vs Prévision IA N+1
            - Barres supplémentaires : Profils cochés (somme mensuelle)
            """
            nonlocal monthly_graph_widget, monthly_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            # Il faut au moins la prédiction N+1 pour tracer quelque chose
            if current_pred_df is None or current_pred_df.empty:
                return

            # Pack du conteneur une seule fois
            if not monthly_container_packed:
                monthly_container.pack(pady=10, fill="both", expand=True)
                monthly_container_packed = True

            # Nettoyage du conteneur
            for child in monthly_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            monthly_graph_widget = None

            import numpy as np
            from matplotlib.ticker import FuncFormatter

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            color_real = "#1f77b4"   # bleu
            color_pred = "#F4D03F"   # orange
            # Palette dédiée aux profils (change-les comme tu veux)
            profile_colors = [
                "#e74c3c",  # rouge vif
                "#9b59b6",  # violet
                "#2ecc71",  # vert clair
                "#1abc9c",  # turquoise
                "#f1c40f",  # jaune
                "#d35400",  # orange foncé (≠ #ff7f0e)
                "#8e44ad",  # violet profond
                "#27ae60",  # vert riche
                "#16a085",  # vert/turquoise sombre
                "#c0392b",  # rouge foncé
                "#7f8c8d",  # gris bleuté
                "#95a5a6",  # gris clair
                "#34495e",  # bleu/gris métallique
                "#bdc3c7",  # argent clair
                "#f39c12",  # jaune/orangé
                "#c27ba0",  # rose/mauve mat
                "#76d7c4",  # aqua pastel
                "#7dcea0",  # vert pastel
                "#af7ac5",  # mauve pastel
                "#5dade2",  # bleu pastel clair (≠ réel, beaucoup plus clair)
            ]

            # ---------- Prévision IA N+1 : somme mensuelle ----------
            df_pred = current_pred_df.copy()
            df_pred["month"] = df_pred["date"].dt.month
            pred_monthly = (
                df_pred.groupby("month", as_index=False)["pred_value"]
                       .sum()
            ).rename(columns={"pred_value": "pred_value"})

            # ---------- Réel N+1 : somme mensuelle (si dispo) ----------
            if current_real_target_df is not None and not current_real_target_df.empty:
                df_real = current_real_target_df.copy()
                df_real["month"] = df_real["date"].dt.month
                real_monthly = (
                    df_real.groupby("month", as_index=False)["y"]
                           .sum()
                ).rename(columns={"y": "real_value"})

                monthly_cmp = pd.merge(
                    real_monthly,
                    pred_monthly,
                    on="month",
                    how="outer"
                ).fillna(0.0)
            else:
                monthly_cmp = pred_monthly.copy()
                monthly_cmp["real_value"] = 0.0

            monthly_cmp = monthly_cmp.sort_values("month")

            # ---------- Profils cochés : agrégés par mois (pour barres) ----------
            # ---------- Profils cochés : calcul "Prévision sinon Réel" + masquage ----------
            active_profiles = []

            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = pd.to_datetime(ia_profils_dates)

                # Base réelle jour par jour
                df_day = pd.DataFrame({
                    "date": dates_prof,
                    "month": dates_prof.month,
                })

                # Ajouter le réel si dispo
                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = dict(zip(current_real_target_df["date"], current_real_target_df["y"]))
                    df_day["real"] = df_day["date"].map(real_map).fillna(0.0)
                else:
                    df_day["real"] = 0.0

                # Calcul par profil
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    vals = [_to_float_or_nan(v) for v in serie]

                    df_day["prev"] = vals

                    # Prévision sinon réel
                    df_day["comb"] = df_day["prev"]
                    df_day.loc[df_day["comb"].isna(), "comb"] = df_day["real"]

                    # Mois avec prévision
                    mois_avec_prev = set(df_day.loc[df_day["prev"].notna(), "month"].unique())
                    mois_tous = set(monthly_cmp["month"].unique())
                    mois_masques = sorted(mois_tous - mois_avec_prev)

                    # Masquage complet des mois sans prévision
                    df_day["comb_masked"] = df_day["comb"]
                    df_day.loc[df_day["month"].isin(mois_masques), "comb_masked"] = np.nan

                    # Agrégation mensuelle
                    prof_month = df_day.groupby("month")["comb_masked"].sum(min_count=1)

                    # Alignement
                    yvals = [prof_month.get(m, np.nan) for m in monthly_cmp["month"]]

                    active_profiles.append((name, yvals))


            # ---------- Placement des barres ----------
            x = np.arange(len(monthly_cmp))

            # nb séries : Réel + Prévision + profils actifs
            nb_series = 2 + len(active_profiles)
            # largeur de chaque barre pour que tout tienne dans le mois
            width = 0.8 / max(nb_series, 1)
            offsets = (np.arange(nb_series) - (nb_series - 1) / 2.0) * width

            fig_m, ax_m = plt.subplots(figsize=(11, 4.5),
                                       facecolor="#00122e",
                                       constrained_layout=True)
            ax_m.set_facecolor("#00122e")

            # Barres Réel N+1
            ax_m.bar(
                x + offsets[0],
                monthly_cmp["real_value"],
                width,
                label=f"Réel {target_year}",
                color=color_real
            )

            # Barres Prévision IA N+1
            ax_m.bar(
                x + offsets[1],
                monthly_cmp["pred_value"],
                width,
                label=f"Prévision IA {target_year}",
                color=color_pred
            )

            # Barres Profils cochés
            if active_profiles:
                for i, (name, y_vals) in enumerate(active_profiles):
                    serie_idx = 2 + i
                    profile_color = profile_colors[i % len(profile_colors)]
                    ax_m.bar(
                        x + offsets[serie_idx],
                        y_vals,
                        width,
                        label=f"Profil '{name}' {target_year}",
                        color=profile_color
                    )


            # ---------- X / Y / style ----------
            month_nums = monthly_cmp["month"].astype(int)
            month_dates = pd.to_datetime(
                {"year": [target_year] * len(month_nums),
                 "month": month_nums,
                 "day": 1}
            )
            mois_labels = month_dates.dt.strftime("%Y-%m")

            ax_m.set_xticks(x)
            ax_m.set_xticklabels(mois_labels, rotation=45, ha="right",
                                 fontsize=9, color="white")

            ax_m.set_xlabel("Mois", fontsize=11, color="white")
            ax_m.set_ylabel("Valeur cumulée (K€)", fontsize=11, color="white")
            ax_m.tick_params(axis="y", colors="white")

            def _fmt_milliers(val, pos):
                try:
                    return f"{int(val):,}".replace(",", " ")
                except Exception:
                    return ""
            ax_m.yaxis.set_major_formatter(FuncFormatter(_fmt_milliers))

            filiale_for_title = current_filiale_name or selected_filiale.get()
            ax_m.set_title(
                f"{filiale_for_title} - Enc. Autres Produits - {target_year}",
                fontsize=13,
                fontweight="bold",
                color="white",
                pad=12
            )

            # pas de bordures
            fig_m.patch.set_edgecolor("#00122e")
            fig_m.patch.set_linewidth(0)
            for spine in ax_m.spines.values():
                spine.set_visible(False)

            # légende sans cadre
            leg_m = ax_m.legend(frameon=False, facecolor="#00122e")
            for text in leg_m.get_texts():
                text.set_color("white")

            canvas_m = FigureCanvasTkAgg(fig_m, master=monthly_container)
            canvas_m.draw()
            monthly_graph_widget = canvas_m.get_tk_widget()
            monthly_graph_widget.pack(fill="both", expand=True)
            plt.close(fig_m)

        # ---------- analyse des écarts (IA vs Réel, Profil vs Réel) ----------

        def _compute_metrics(label, ref_dates, ref_values, cmp_dates, cmp_values,
                     min_abs_value=0.0, ecart_seuil=40.0):
            """
            ref = série réelle (Réel)
            cmp = série comparée (IA ou Profil)

            - aligne les dates
            - convertit proprement en float (gère '', None, etc.)
            - filtre (optionnel) les petits montants sur le RÉEL via min_abs_value
            - calcule l'écart relatif |ref - cmp| / |cmp|
            - renvoie toujours un dict SI au moins une date est commune,
            même si après filtrage il n'y a plus de points 'significatifs'
            """

            if ref_dates is None or cmp_dates is None:
                print(f"[METRICS] {label} -> ref_dates ou cmp_dates None")
                return None

            # Conversion en float en utilisant ton helper
            ref_vals = [_to_float_or_nan(v) for v in ref_values]
            cmp_vals = [_to_float_or_nan(v) for v in cmp_values]

            ref_idx = pd.to_datetime(ref_dates)
            cmp_idx = pd.to_datetime(cmp_dates)

            ref = pd.Series(ref_vals, index=ref_idx, dtype="float64")
            cmp_ = pd.Series(cmp_vals, index=cmp_idx, dtype="float64")

            # Intersection des dates
            common_idx = ref.index.intersection(cmp_.index)
            print(f"[METRICS] {label} -> nb dates communes = {len(common_idx)}")

            if len(common_idx) == 0:
                # dans ce cas on ne sait vraiment rien comparer
                return None

            df_m = pd.DataFrame({
                "ref": ref.loc[common_idx],
                "cmp": cmp_.loc[common_idx],
            })

            # Drop NaN
            df_m = df_m.dropna()
            print(f"[METRICS] {label} -> après dropna, nb points = {len(df_m)}")

            if df_m.empty:
                # aucune valeur exploitable → on renvoie quand même une ligne "0"
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # 🔎 Filtre optionnel : SEULEMENT sur le RÉEL
            # ex : min_abs_value=10  -> on garde uniquement les jours où |Réel| > 10
            if min_abs_value is not None and min_abs_value > 0:
                before = len(df_m)
                df_m = df_m[df_m["ref"].abs() >= min_abs_value]
                print(f"[METRICS] {label} -> filtre |ref| >= {min_abs_value} : {before} → {len(df_m)} points")

            if df_m.empty:
                # il y a des dates communes, mais tous les réels sont trop petits
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # Éviter la division par zéro (sur la prévision)
            df_m = df_m[df_m["cmp"] != 0]
            if df_m.empty:
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            # Écart relatif basé sur la prévision cmp
            rel = (df_m["ref"] - df_m["cmp"]).abs() / df_m["cmp"].abs() * 100.0
            rel = rel.replace([np.inf, -np.inf], np.nan).dropna()

            if rel.empty:
                return {
                    "Comparaison": label,
                    "Nb écarts (> 40 %)": 0,
                    "Taux d'écarts (%)": 0.0,
                    "Valeur totale des écarts (K€)": 0.0,
                }

            mask = rel > ecart_seuil

            nb_points = int(df_m.shape[0])
            nb_ecarts = int(mask.sum())
            taux = 0.0 if nb_points == 0 else 100.0 * nb_ecarts / nb_points
            valo = float((df_m["ref"] - df_m["cmp"]).abs()[mask].sum())

            print(f"[METRICS] {label} -> nb_points={nb_points}, nb_ecarts={nb_ecarts}, taux={taux:.1f}, valo={valo:.1f}")

            return {
                "Comparaison": label,
                "Nb écarts (> 40 %)": nb_ecarts,
                "Taux d'écarts (%)": round(taux, 1),
                "Valeur totale des écarts (K€)": round(valo, 1),
            }


        def _rebuild_analysis_table():
            nonlocal analysis_table_frame, analysis_tree, analysis_export_button
            nonlocal current_pred_df, current_real_target_df, current_target_year

            if analysis_table_frame is None:
                return
            for child in analysis_table_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            analysis_tree = None
            analysis_export_button = None

            # s'il n'y a pas de réel, on ne peut rien comparer
            if current_real_target_df is None or current_real_target_df.empty:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune donnée réelle pour l'année N+1 : analyse des écarts impossible.",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                print("[ANALYSE] Pas de réel N+1.")
                return

            print(f"[ANALYSE] réel N+1 : {len(current_real_target_df)} lignes")
            if current_pred_df is not None:
                print(f"[ANALYSE] préd N+1 : {len(current_pred_df)} lignes")
            else:
                print("[ANALYSE] current_pred_df est None")

            rows = []

            # 1) IA vs Réel
            if current_pred_df is not None and not current_pred_df.empty:
                m = _compute_metrics(
                    label="Prévision IA vs Réel",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    # tu peux remonter le seuil ici plus tard si tu veux
                    min_abs_value=10.0,
                    ecart_seuil=40.0,
                )
                print("[ANALYSE] IA vs Réel ->", m)
                if m is not None:
                    rows.append(m)

            # 2) Profils cochés vs Réel
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    m = _compute_metrics(
                        label=f"Profil '{name}' vs Réel",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=ia_profils_dates,
                        cmp_values=serie,
                        min_abs_value=10.0,
                        ecart_seuil=40.0,
                    )
                    print(f"[ANALYSE] Profil {name} vs Réel ->", m)
                    if m is not None:
                        rows.append(m)

            ctk.CTkLabel(
                analysis_table_frame,
                text=f"Analyse des écarts (seuil 40 %, ref/prev ≥ 0) année {current_target_year or (int(annees_var.get())+1)}",
                font=("Segoe UI", 13, "bold"),
                text_color="white"
            ).pack(anchor="w", pady=(0, 6))

            if not rows:
                ctk.CTkLabel(
                    analysis_table_frame,
                    text="Aucune comparaison disponible (vérifie que le modèle est entraîné et que des profils sont cochés).",
                    font=("Segoe UI", 11),
                    text_color="#c9defa"
                ).pack(anchor="w", pady=(0, 8))
                print("[ANALYSE] Aucune ligne de comparaison construite.")
                return

            cols = [
                "Comparaison",
                "Nb écarts (> 40 %)",
                "Taux d'écarts (%)",
                "Valeur totale des écarts (K€)"
            ]
            tree = ttk.Treeview(
                analysis_table_frame, columns=cols, show="headings",
                height=len(rows)
            )
            for col in cols:
                tree.heading(col, text=col)

            tree.column("Comparaison", anchor="w", width=260)
            tree.column("Nb écarts (> 40 %)", anchor="center", width=150)
            tree.column("Taux d'écarts (%)", anchor="center", width=150)
            tree.column("Valeur totale des écarts (K€)", anchor="center", width=180)

            for r in rows:
                tree.insert(
                    "", "end",
                    values=(
                        r["Comparaison"],
                        r["Nb écarts (> 40 %)"],
                        r["Taux d'écarts (%)"],
                        r["Valeur totale des écarts (K€)"],
                    )
                )

            tree.pack(fill="x", expand=True, pady=(0, 8))
            analysis_tree = tree

            # bouton export Excel du tableau d'analyse
            def _export_analysis_excel():
                if not rows:
                    messagebox.showinfo("Export", "Aucune donnée à exporter.")
                    return
                df_export = pd.DataFrame(rows)
                try:
                    file_path = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Fichiers Excel", "*.xlsx")],
                        title="Exporter le tableau d'analyse en Excel"
                    )
                    if not file_path:
                        return
                    df_export.to_excel(file_path, index=False)
                    messagebox.showinfo("Export", f"Tableau d'analyse exporté vers :\n{file_path}")
                except Exception as e:
                    messagebox.showerror("Erreur export", str(e))

            analysis_export_button = ctk.CTkButton(
                analysis_table_frame,
                text="📤 Exporter le tableau d'analyse (Excel)",
                width=260, height=32,
                corner_radius=10,
                fg_color="#2563eb", hover_color="#1d4ed8",
                text_color="white",
                command=_export_analysis_excel
            )
            analysis_export_button.pack(anchor="w", pady=(6, 4))

        # ---------- export Excel des prédictions N+1 ----------

        def _export_predictions_to_excel():
            nonlocal exported_pred_df, current_target_year, current_filiale_name

            if exported_pred_df is None or exported_pred_df.empty:
                messagebox.showinfo(
                    "Export",
                    "Aucune prédiction à exporter. Lance d'abord l'entraînement / la prévision."
                )
                return

            try:
                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Fichiers Excel", "*.xlsx")],
                    title="Exporter les prédictions N+1 en Excel"
                )
                if not file_path:
                    return

                df_export = exported_pred_df.copy()
                df_export.insert(0, "Filiale", current_filiale_name or "")
                df_export.insert(1, "Année", current_target_year or "")

                df_export.to_excel(file_path, index=False)
                messagebox.showinfo(
                    "Export",
                    f"Prédictions N+1 exportées vers :\n{file_path}"
                )
            except Exception as e:
                messagebox.showerror("Erreur export", str(e))

        # ---------- PROFILS UI ----------

        def _build_ia_profils_ui(filiale, base_year):
            """
            Construit les cases à cocher de profils pour l'année N+1 de la filiale sélectionnée.
            Dès qu'on coche/décoche, _redraw_graph2() et _rebuild_analysis_table() sont appelés.
            """
            nonlocal ia_profils_vars, ia_profils_names, ia_profils_dates, ia_profils_series

            # Nettoyage du frame
            for w in ia_profils_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            ia_profils_vars = []
            ia_profils_names = []
            ia_profils_dates = []
            ia_profils_series = []

            if not filiale:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Aucune filiale sélectionnée.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                base_year_int = int(base_year)
            except Exception:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Année N invalide.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            target_year = base_year_int + 1
            feuille = sections.get(filiale)
            if not feuille:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune feuille trouvée pour {filiale}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees (profils N+1) pour {feuille} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Erreur de chargement des données N+1.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                lbl = tk.Label(
                    ia_profils_frame,
                    text="Flux 'Enc. Autres Produits' introuvable pour cette filiale.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            _, col_start = cible[0]

            try:
                dates_p, reel_p, previsions_p, noms_profils_p = extraire_valeurs(
                    ws, col_start, nb_prev, annee=target_year
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs (profils N+1) {filiale}/{flux_cible}/{target_year} : {e}")
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée de profils pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            if not dates_p:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            # filtrer les profils actifs (au moins une valeur non nulle)
            actifs = []
            for serie in previsions_p:
                exist = any(v not in (None, 0, 0.0, "") for v in serie)
                actifs.append(exist)

            noms_actifs = [n for n, ok in zip(noms_profils_p, actifs) if ok]
            series_actives = [s for s, ok in zip(previsions_p, actifs) if ok]

            if not noms_actifs:
                lbl = tk.Label(
                    ia_profils_frame,
                    text=f"Aucun profil actif pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                )
                lbl.pack(anchor="w")
                return

            ia_profils_dates = dates_p
            ia_profils_names = noms_actifs
            ia_profils_series = series_actives

            # Label d'info
            lbl = tk.Label(
                ia_profils_frame,
                text=f"Profils N+1 ({target_year}) : coche pour afficher dans le graphe détaillé et l'analyse",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
            )
            lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 6))

            nb_lignes = 4
            for i, name in enumerate(noms_actifs):
                var = tk.BooleanVar(value=False)

                def _on_toggle(v=var):
                    _redraw_graph2()
                    _redraw_monthly_graph()
                    _rebuild_analysis_table()


                cb = tk.Checkbutton(
                    ia_profils_frame, text=name, variable=var,
                    bg="#00122e", fg="white", font=('Segoe UI', 10),
                    selectcolor="#00aced", activebackground="#003366",
                    activeforeground="white",
                    command=_on_toggle
                )

                row = 1 + (i % nb_lignes)   # +1 car première ligne = label
                col = i // nb_lignes

                cb.grid(row=row, column=col, sticky="w", padx=12, pady=4)
                cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
                cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

                ia_profils_vars.append(var)

        def _on_filiale_or_year_change(_event=None):
            """Quand filiale ou année N change, on recharge les profils N+1."""
            filiale = selected_filiale.get()
            base_year = annees_var.get()
            if not filiale or not base_year:
                return
            _build_ia_profils_ui(filiale, base_year)

        filiale_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        annees_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        _on_filiale_or_year_change()

        # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------

                # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------

        def _train_model():
            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame, export_button

            _clear_graph_widgets()
            print("======================== ENTRAINEMENT DU MODELE (RF + XGB résidus + WEEKEND=0) ========================")

            filiale = selected_filiale.get()
            print(f"Filiale sélectionnée : {filiale}")

            df_filiale = df[df["section"] == filiale].copy()
            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée trouvée pour la filiale {filiale}.")
                return

            # ----------------------- ANNÉE N -----------------------
            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["year"].max())
            print(f"Base year for training {filiale} : {base_year}")

            try:
                # ----------------------- HISTORIQUE -----------------------
                if df_filiale["year"].nunique() < 2:
                    messagebox.showinfo(
                        "Information",
                        f"Pas assez d'historique pour {filiale} (au moins 2 années)."
                    )
                    return

                # ----------------------- FEATURES TEMPORELLES -----------------------
                df_filiale["dow"] = df_filiale["date"].dt.weekday
                df_filiale["month"] = df_filiale["date"].dt.month
                df_filiale["dom"] = df_filiale["date"].dt.day
                df_filiale["is_eom"] = df_filiale["date"].dt.is_month_end.astype(int)
                df_filiale["lag_1"] = df_filiale["y"].shift(1)
                df_filiale["lag_7"] = df_filiale["y"].shift(7)

                # ========== PAIRS YEAR-OVER-YEAR ==========
                s = df_filiale.copy()
                s_prev = s.copy()
                s_prev["year_target"] = s_prev["year"] + 1

                s_prev = s_prev.rename(columns={
                    "y": "y_prev_year",
                    "roll_mean_7": "roll_prev_7",
                    "roll_mean_30": "roll_prev_30",
                    "dow": "dow_prev",
                    "lag_1": "lag_1_prev",
                    "lag_7": "lag_7_prev",
                    "month": "month_prev",
                    "dom": "dom_prev",
                    "is_eom": "is_eom_prev"
                })

                merged = pd.merge(
                    s_prev,
                    s[["section", "year", "dayofyear", "y"]],
                    left_on=["section", "year_target", "dayofyear"],
                    right_on=["section", "year", "dayofyear"],
                    how="inner",
                )

                if merged.empty:
                    messagebox.showinfo("Information", f"Aucune paire année→année pour {filiale}.")
                    return

                merged = merged.rename(columns={"y": "y_target"})
                df_pairs = merged[[  
                    "section", "year_target", "dayofyear",
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev",
                    "y_target"
                ]].copy()
                df_pairs["section_id"] = 0

                df_train_pairs = df_pairs[df_pairs["year_target"] <= base_year]
                if df_train_pairs.shape[0] < 30:
                    messagebox.showinfo(
                        "Information",
                        f"Pas assez de données pour entraîner le modèle jusqu'à {base_year}."
                    )
                    return

                # ----------------------- FEATURES UTILISÉES -----------------------
                features = [
                    "y_prev_year", "roll_prev_7", "roll_prev_30",
                    "dayofyear", "section_id",
                    "dow_prev", "lag_1_prev", "lag_7_prev",
                    "month_prev", "dom_prev", "is_eom_prev"
                ]

                df_train_pairs = df_train_pairs.dropna(subset=features + ["y_target"])
                X = df_train_pairs[features].values
                y_target = df_train_pairs["y_target"].values

                # ----------------------- SPLIT -----------------------
                from sklearn.model_selection import train_test_split
                X_train, X_valid, y_train, y_valid = train_test_split(
                    X, y_target, test_size=0.2, random_state=42
                )

                from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
                from sklearn.ensemble import RandomForestRegressor
                from xgboost import XGBRegressor

                base_n = int(n_estimators_var.get())
                base_d = int(max_depth_var.get()) if use_depth_var.get() else None

                # ========================================================================================
                #                      NIVEAU 1 : RANDOM FOREST (tendance)
                # ========================================================================================
                print("\n===== NIVEAU 1 : RANDOM FOREST (tendance) =====")

                rf = RandomForestRegressor(
                    n_estimators=base_n,
                    max_depth=base_d,
                    max_features="sqrt",
                    min_samples_leaf=2,
                    random_state=42,
                    n_jobs=-1
                )
                rf.fit(X_train, y_train)

                rf_pred_train = rf.predict(X_train)
                rf_pred_valid = rf.predict(X_valid)

                # ========================================================================================
                #                 NIVEAU 2 : XGBOOST sur les résidus
                # ========================================================================================
                print("\n===== NIVEAU 2 : XGBOOST (résidus) =====")

                res_train = y_train - rf_pred_train

                xgb = XGBRegressor(
                    n_estimators=base_n,
                    max_depth=base_d if base_d is not None else 6,
                    learning_rate=0.05,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    objective="reg:squarederror",
                    tree_method="hist",
                    random_state=42
                )
                xgb.fit(X_train, res_train)

                res_valid = xgb.predict(X_valid)

                # ========================================================================================
                #                    PRÉDICTION FINALE = RF + XGB(resid)
                # ========================================================================================
                ensemble_pred = rf_pred_valid + res_valid

                mae = mean_absolute_error(y_valid, ensemble_pred)
                rmse = mean_squared_error(y_valid, ensemble_pred)**0.5
                r2 = r2_score(y_valid, ensemble_pred)

                print(f"[MODEL FINAL] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")

                # ========================================================================================
                #                                    KPI
                # ========================================================================================
                kpi_frame = ctk.CTkFrame(scrollable_frame, fg_color="#0f1b31", corner_radius=18)
                kpi_frame.pack(fill="x", padx=10, pady=(0, 10))
                graph_widgets.append(kpi_frame)

                for i in range(3):
                    kpi_frame.grid_columnconfigure(i, weight=1)

                def _kpi(parent, title, value, subtitle, col):
                    card = ctk.CTkFrame(
                        parent, fg_color="#142544", corner_radius=16,
                        border_width=1, border_color="#223658"
                    )
                    card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
                    ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                        .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
                    ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 22, "bold"),
                                text_color="white") \
                        .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
                    ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                                text_color="#7ea2d8") \
                        .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

                _kpi(kpi_frame, "MAE", f"{mae:.1f}", "Erreur moyenne", 0)
                _kpi(kpi_frame, "RMSE", f"{rmse:.1f}", "Racine MSE", 1)
                _kpi(kpi_frame, "R²", f"{r2:.3f}", "Pouvoir explicatif", 2)

                # ========================================================================================
                #                     PRÉVISION N+1 (WEEK-END = 0 + rattrapage lundi)
                # ========================================================================================

                target_year = base_year + 1
                df_prev_year = df_filiale[df_filiale["year"] == base_year].copy()
                df_prev_year = df_prev_year.sort_values("date")  # important pour gérer le carryover

                future_rows = []
                carryover = 0.0  # somme des prédictions des jours de week-end à reporter

                for _, row in df_prev_year.iterrows():
                    d_prev = row["date"]
                    d_next = d_prev + pd.DateOffset(years=1)
                    dow_next = d_next.weekday()  # jour de la semaine dans l'année N+1

                    feat_vec = [[
                        float(row["y"]),
                        float(row["roll_mean_7"]),
                        float(row["roll_mean_30"]),
                        int(row["dayofyear"]),
                        0,
                        int(d_prev.weekday()),
                        float(row["lag_1"]),
                        float(row["lag_7"]),
                        int(row["month"]),
                        int(row["dom"]),
                        int(row["is_eom"])
                    ]]

                    # Prédiction brute (sans logique week-end)
                    pred_rf = float(rf.predict(feat_vec)[0])
                    pred_res = float(xgb.predict(feat_vec)[0])
                    raw_pred = pred_rf + pred_res

                    # WEEK-END = 0 mais on accumule pour le lundi suivant (ou prochain jour ouvré)
                    if dow_next >= 5:  # samedi (5) ou dimanche (6)
                        carryover += raw_pred
                        pred_final = 0.0
                    else:
                        # Jour ouvré : on ajoute ce qu'on a accumulé le week-end précédent
                        pred_final = raw_pred + carryover
                        carryover = 0.0

                    future_rows.append({
                        "section": filiale,
                        "date": d_next,
                        "year": d_next.year,
                        "month": d_next.month,
                        "dayofyear": d_next.dayofyear,
                        "pred_value": pred_final
                    })

                df_future_all = pd.DataFrame(future_rows)

                # ========================================================================================
                #     AJUSTEMENT HEBDO : sur chaque semaine, le plus haut pic est mis le lundi
                # ========================================================================================
                if not df_future_all.empty:
                    # semaine ISO + année ISO pour éviter les problèmes autour du nouvel an
                    iso = df_future_all["date"].dt.isocalendar()
                    df_future_all["iso_year"] = iso.year.astype(int)
                    df_future_all["iso_week"] = iso.week.astype(int)
                    df_future_all["dow"] = df_future_all["date"].dt.weekday

                    for (y_iso, w_iso), g in df_future_all.groupby(["iso_year", "iso_week"]):
                        idxs = g.index
                        # lundi de la semaine
                        monday_idx = g[g["dow"] == 0].index
                        if monday_idx.empty:
                            continue
                        monday_idx = monday_idx[0]

                        # jour avec le max de la semaine
                        max_idx = g["pred_value"].idxmax()

                        # si déjà lundi, rien à faire
                        if max_idx == monday_idx:
                            continue

                        # swap des valeurs
                        val_monday = df_future_all.at[monday_idx, "pred_value"]
                        val_max = df_future_all.at[max_idx, "pred_value"]
                        df_future_all.at[monday_idx, "pred_value"] = val_max
                        df_future_all.at[max_idx, "pred_value"] = val_monday

                    # nettoyage des colonnes techniques
                    df_future_all = df_future_all.drop(columns=["iso_year", "iso_week", "dow"])

                # ========================================================================================
                #                                    GRAPHE 1 (jours)
                # ========================================================================================
                fig1, ax1 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")

                color_real = "#1f77b4"   # bleu
                color_pred = "#ff7f0e"   # orange

                df_hist_plot = df_filiale[df_filiale["year"] <= base_year]

                # Réel historique
                ax1.plot(
                    df_hist_plot["date"],
                    df_hist_plot["y"],
                    label=f"Réel (≤ {base_year})",
                    linewidth=2,
                    color=color_real
                )

                # Prévision N+1
                ax1.plot(
                    df_future_all["date"],
                    df_future_all["pred_value"],
                    label=f"Prévision Modèle 2 niveaux {target_year}",
                    linewidth=2,
                    linestyle="--",
                    color=color_pred
                )

                ax1.set_title(f"Prévision N+1 – Modèle RF + XGB résidus (Weekend=0 + pic hebdo lundi)", color="white")
                ax1.tick_params(axis='x', colors="white", rotation=30)
                ax1.tick_params(axis='y', colors="white")

                # Bordures
                for spine in ax1.spines.values():
                    spine.set_color("#00122e")

                leg1 = ax1.legend(facecolor="#00122e", edgecolor="white")
                for text in leg1.get_texts():
                    text.set_color("white")

                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                w_fig1 = canvas_fig1.get_tk_widget()
                w_fig1.pack(pady=10, fill="both", expand=True)
                graph_widgets.append(w_fig1)
                plt.close(fig1)

                # ========================================================================================
                #                   MISE À JOUR GLOBAL + ANALYSE
                # ========================================================================================
                current_pred_df = df_future_all
                current_real_target_df = df_filiale[df_filiale["year"] == target_year]
                current_target_year = target_year
                current_filiale_name = filiale
                exported_pred_df = df_future_all.copy()

                _redraw_monthly_graph()
                _redraw_graph2()

                if analysis_table_frame is not None:
                    analysis_table_frame.destroy()

                analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
                analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
                graph_widgets.append(analysis_table_frame)

                _rebuild_analysis_table()

            except Exception:
                messagebox.showerror("Erreur", traceback.format_exc())
                print("[IA] ERREUR:\n", traceback.format_exc())

        # 🔗 On connecte le bouton à la fonction d'entraînement
        bouton_train.configure(command=_train_model)

        # ============ SCROLL MOLETTE ============

        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # ============ BOUTON RETOUR ============

        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)





if __name__ == "__main__":
    app = Application()
    app.mainloop()
    
 