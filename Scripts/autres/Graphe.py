# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# === Utilitaires divers ===
import os
import re
from datetime import datetime
from collections import defaultdict
import mplcursors  # à mettre en haut de ton fichier
from openpyxl import load_workbook
import re

fichier_excel = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Résultats\Historique_prev_reel_filiales3.xlsx"

pattern = re.compile(r"Prévision \d{2}/\d{2}")

# Charger le classeur
wb = load_workbook(fichier_excel, data_only=True)

# Set comprehension pour récupérer toutes les cellules uniques correspondant au motif
previsions_triees = sorted(
    {
        cell.strip()
        for ws in wb.worksheets
        for row in ws.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str) and pattern.search(cell)
    }
)

sections = {
    "SNCF_SA": "SNCF_SA",
    "RESEAU": "RESEAU",
    "SA_VOYAGEURS": "SA_VOYAGEURS",
    "GEODIS": "GEODIS",
    "G&C": "G&C"
}

nb_prev = len(previsions_triees)
taille_bloc = 2 + 2 * nb_prev + 1
print(nb_prev)

def charger_donnees(feuille, taille_bloc):
    ws = wb[feuille]
    noms = []
    noms_a_exclure = ["Trésorerie de fin", "Cashpool", "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
                    "Créances CDP", "Placements", "CC financiers",
                    "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement", "Endettement Net"
                    ]
    for col in range(3, ws.max_column + 1, taille_bloc):
        cell = ws.cell(row=2, column=col)
        if cell.value and cell.value not in noms_a_exclure:
            noms.append((cell.value, col))
    return ws, noms

def extraire_valeurs(ws, col_start, nb_prev):
    """
    Extrait dates, réel et prévisions depuis la feuille `ws` à partir de `col_start`.
    Gère correctement :
      - datetime/date Python
      - strings "JJ/MM/AAAA" (avec dayfirst)
      - numéros Excel (origin '1899-12-30')
    Retourne : dates (Timestamp normalisées à minuit), reel (list), previsions (list[list]), noms_previsions (list)
    """
    import pandas as pd
    from datetime import datetime, date

    # 1) Noms des profils de prévision
    noms_previsions = []
    for i in range(nb_prev):
        col_prev = col_start + 1 + i * 2
        nom = ws.cell(row=3, column=col_prev).value
        if nom is None:
            nom_propre = f"Profil {i+1}"
        else:
            nom_propre = str(nom).replace("(K€)", "").replace("Prévision", "Profil").strip()
        noms_previsions.append(nom_propre)

    previsions = [[] for _ in range(len(noms_previsions))]
    dates, reel = [], []

    def parse_excel_date(v):
        # Normalise TOUT en pandas.Timestamp (sans heure)
        if v is None:
            return None
        # Déjà datetime/date Python
        if isinstance(v, (datetime, date)):
            ts = pd.Timestamp(v)
            return ts.normalize()  # 00:00:00
        # Serial Excel (nombre de jours depuis 1899-12-30)
        if isinstance(v, (int, float)):
            # int() protège contre les flottants "propres"
            ts = pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")
            return ts.normalize()
        # Chaîne — on force dayfirst pour JJ/MM/AAAA
        if isinstance(v, str):
            ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
            if pd.isna(ts):
                return None
            return ts.normalize()
        # Dernier recours
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.normalize()

    # 2) Parcours des lignes
    row = 4
    while True:
        date_val = ws.cell(row=row, column=col_start - 1).value
        if date_val is None:
            break
        d = parse_excel_date(date_val)
        if d is None:
            # si la cellule n’est pas interprétable, on sort (ou on continue selon ton besoin)
            break

        # Réel
        r = ws.cell(row=row, column=col_start).value
        reel.append(r if r is not None else 0)
        dates.append(d)

        # Prévisions (profil par profil)
        for i in range(len(noms_previsions)):
            col_prev = col_start + 1 + i * 2
            v = ws.cell(row=row, column=col_prev).value
            previsions[i].append(v if v is not None else None)

        row += 1

    return dates, reel, previsions, noms_previsions


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        self.attributes("-fullscreen", True)
        self.geometry("1300x900")
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

    GREEN = "\033[32m"
    RESET = "\033[0m"
    BOLD = "\033[1m"
    FUTURISTIC_PREFIX = f"{BOLD}{GREEN}[⚡ FUTURE APP]{RESET}"

    def print_info(self, message: str):
        print(f"{self.FUTURISTIC_PREFIX} {self.GREEN}{message}{self.RESET}")

    def creer_accueil(self):
        self.vider_fenetre()
        import customtkinter as ctk
        from PIL import Image, ImageTk
        import tkinter as tk
        from customtkinter import CTkImage

        self.print_info("Initialisation de l'apparence et du thème...")
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.print_info("Création de l'entête...")
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - ANALYSE DE L'EXISTANT", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        self.print_info("Chargement du logo...")
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)

            # Utiliser CTkImage (ça gère tout seul DPI scaling)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo  # garder une référence

            logo_label.pack(side="right", anchor="e", padx=(10, 0))
            self.print_info("Logo chargé avec succès ✅")

        except Exception as e:
            print(f"{self.FUTURISTIC_PREFIX} Erreur chargement du logo: {e}")

        self.print_info("Création de la barre blanche...")
        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        self.print_info("Création de la frame principale...")
        self.frame_accueil = ctk.CTkFrame(self, corner_radius=15, fg_color="#00122e")
        self.frame_accueil.place(relx=0.5, rely=0.5, anchor='center')
        self.frame_accueil.configure(width=900, height=650)
        self.frame_accueil.pack_propagate(False)

        self.print_info("Ajout des titres et sous-titres...")
        titre_style = {"font": ("Segoe UI Semibold", 22, "bold"), "text_color": "white"}
        sous_titre_style = {"font": ("Arial", 14), "text_color": "white"}

        ctk.CTkLabel(self.frame_accueil, text="Bienvenue dans le Menu", **titre_style).pack(pady=(40, 10), padx=60)
        ctk.CTkLabel(self.frame_accueil,
                    text="Analyse quantitative des données réelles et prévisions, et les écarts importants.",
                    wraplength=700, justify="center", **sous_titre_style).pack(pady=(0, 40), padx=60)

        self.print_info("Création des boutons...")


        # --- FRAME BOUTONS ---
        btn_frame = ctk.CTkFrame(self.frame_accueil, fg_color="#00122e")
        btn_frame.pack(pady=(0, 40))

        # Styles boutons
        style_btn_blue = {
            "width": 200,
            "height": 40,
            "corner_radius": 20,
            "fg_color": "#005288",
            "hover_color": "#00396B",
            "text_color": "white",
            "font": ("Segoe UI", 13, "bold")
        }

        style_btn_red = {
            "width": 200,
            "height": 40,
            "corner_radius": 20,
            
            "fg_color": "#C21515",
            "hover_color": "#A00000",
            "text_color": "white",
            "font": ("Segoe UI", 13, "bold")
        }

       # Création boutons (avec emojis)
        btn1 = ctk.CTkButton(btn_frame, text="📊 Voir Graphique (Filiale + flux)", command=self.afficher_page_graphique, **style_btn_blue)
        btn2 = ctk.CTkButton(btn_frame, text="⚠️ Voir Écarts Importants", command=self.afficher_ecarts, **style_btn_blue)
        btn3 = ctk.CTkButton(btn_frame, text="📈 Écarts par Filiale", command=self.afficher_repartition, **style_btn_blue)
        btn4 = ctk.CTkButton(btn_frame, text="❌ Fermer l'application", command=self.demander_confirmation_quit, **style_btn_red)
        btn5 = ctk.CTkButton(btn_frame, text="📊 Écarts par Profil", command=self.afficher_repartition_par_prevision, **style_btn_blue)
        btn6 = ctk.CTkButton(btn_frame, text="📅 Catégorisation des écarts", command=self.analyser_ecarts_ml, **style_btn_blue)
        btn7 = ctk.CTkButton(btn_frame, text="📊 Écarts par Flux", command=self.afficher_repartition_flux, **style_btn_blue)
        btn8 = ctk.CTkButton(btn_frame, text="🔥 Heatmap des écarts", command=self.afficher_heatmap_ecarts, **style_btn_blue)
        btn9 = ctk.CTkButton(btn_frame, text="📁 Importer les profils", command=self.exporter_les_profils, **style_btn_blue)
        btn10 = ctk.CTkButton(btn_frame, text="📊 Analyse Notes", **style_btn_blue)
        # btn11 = ctk.CTkButton(btn_frame, text="📊 Voir Graphique (Filiale)", command=self.creer_page_graphique_total, **style_btn_blue)

        # Placement avec grid, 3 colonnes pour un bon équilibre
        boutons = [
            btn9, btn1,  btn2,   
            btn7,  btn5, btn3,   
            btn8,  btn10,  btn6,  
            None,  btn4,  None,   
            None,  None,  None    
        ]

        positions = [
            (0, 0), (0, 1), (0, 2),
            (1, 0), (1, 1), (1, 2),
            (2, 0), (2, 1), (2, 2),
            (3, 0), (3, 1), (3, 2),
            (4, 0), (4, 1), (4, 2)
        ]

        # Placement dans la grille
        for btn, (r, c) in zip(boutons, positions):
            if btn is not None:
                btn.grid(row=r, column=c, padx=20, pady=15, sticky="ew")
        
        self.print_info("Interface prête 🚀")

    def vider_fenetre(self):
        for widget in self.winfo_children():
            widget.destroy()

    def afficher_page_graphique(self):
        self.vider_fenetre()
        self.creer_page_graphique()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.quit()
    
    def exporter_les_profils(self):
        import customtkinter as ctk
        from tkinter import filedialog, ttk
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import os

        self.vider_fenetre()

        # --- HEADER ---
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 28, "bold")
        titre_label = ctk.CTkLabel(
            header_frame,
            text="🚀 PROJET PULSE - IMPORTER LES PROFILS",
            font=titre_font,
            text_color="white"
        )
        titre_label.pack(side="left", anchor="w", padx=(10, 0))

        # Logo SNCF
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(20, 10))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # --- CONTAINER PRINCIPAL (arrondi) ---
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=20)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        # configure grid pour que le canvas prenne tout l'espace
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        # frame qui va contenir le canvas (facilite sticky/expand)
        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")

        # canvas principal
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")

        # faire en sorte que le canvas prenne tout l'espace du canvas_frame
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        # scrollbars dans le container :
        # - verticale à droite (sticky ns) prend toute la hauteur
        # - horizontale en bas qui span les 2 colonnes -> prend toute la largeur
        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")

        # frame scrollable (contenu) placé à l'intérieur du canvas
        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        # mise à jour de la scrollregion quand le contenu change
        def _on_configure(e):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_configure)

        # lier les scrollbars au canvas
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # ------------------------------------------------------------------
        # === CONTENU (identique à ta version originale, placé dans scrollable_frame)
        # ------------------------------------------------------------------

        # Deux colonnes : gauche = tableau, droite = boutons
        left_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        left_frame.pack(side="left", fill="both", expand=True, padx=(30, 15), pady=30)

        right_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        right_frame.pack(side="right", fill="y", padx=(15, 30), pady=30)

        # --- Tableau fichiers trouvés ---
        colonnes = ("Nom des fichiers profil",)
        style = ttk.Style()
        style.configure("Treeview", rowheight=30, font=("Segoe UI", 12))
        style.configure("Treeview.Heading", font=("Segoe UI Semibold", 13))

        self.tableau = ttk.Treeview(left_frame, columns=colonnes, show="headings", height=20)
        for col in colonnes:
            self.tableau.heading(col, text=col)
            self.tableau.column(col, width=500, anchor="w")
        self.tableau.pack(fill="both", expand=True)

        # --- Fonctions recherche ---
        def chercher_excel_commence_par(root_folder, prefix):
            resultats = []
            for dossier, _, fichiers in os.walk(root_folder):
                for fichier in fichiers:
                    if fichier.endswith(".xlsx") and fichier.startswith(prefix):
                        resultats.append(os.path.join(dossier, fichier))
            return resultats

        def choisir_dossier():
            chemin = filedialog.askdirectory(title="Sélectionner un dossier où les profils sont stockés")
            if chemin:
                self.chemin_dossier = chemin
                label_chemin.configure(text=f"✅ Dossier choisi :\n{chemin}")

                # vider tableau
                for row in self.tableau.get_children():
                    self.tableau.delete(row)

                fichiers = chercher_excel_commence_par(chemin, "Profil Tréso SNCF")
                for f in fichiers:
                    self.tableau.insert("", "end", values=(os.path.basename(f),))

        def vider_tableau():
            for row in self.tableau.get_children():
                self.tableau.delete(row)
            label_chemin.configure(text="❌ Tableau vidé.")

        # --- Boutons (droite) ---
        bouton_parcourir = ctk.CTkButton(
            right_frame, text="📁 Parcourir...",
            command=choisir_dossier,
            width=200, height=45, corner_radius=20,
            fg_color="#005288", hover_color="#00396B",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_parcourir.pack(pady=15)

        bouton_nettoyer = ctk.CTkButton(
            right_frame, text="🧹 Vider tableau",
            command=vider_tableau,
            width=200, height=45, corner_radius=20,
            fg_color="#888888", hover_color="#555555",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_nettoyer.pack(pady=15)

        bouton_creer = ctk.CTkButton(
            right_frame, text="📁 Créer fichier",
            command=self.creer_fichier_excel,
            width=200, height=45, corner_radius=20,
            fg_color="#005288", hover_color="#00396B",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_creer.pack(pady=15)

        bouton_valider = ctk.CTkButton(
            right_frame, text="✅ Lancer import",
            command=self.lancer_export_profils,
            width=200, height=45, corner_radius=20,
            fg_color="#008C4B", hover_color="#006C39",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_valider.pack(pady=15)

        bouton_retour = ctk.CTkButton(
            right_frame, text="⬅️ Retour menu",
            command=self.creer_accueil,
            width=200, height=45, corner_radius=20,
            fg_color="#444",
            hover_color="#666",
            text_color="white", font=("Segoe UI", 14, "bold")
        )
        bouton_retour.pack(pady=(40, 0))

        # --- Barre de progression (MASQUÉE par défaut) ---
        # (je garde la même logique que l'original : widgets attachés à self)
        self.progress_label = ctk.CTkLabel(
            self, text="", font=("Segoe UI", 13), text_color="white"
        )
        self.progress_label.pack_forget()

        self.progress_bar = ctk.CTkProgressBar(
            self, width=400, height=20, progress_color="#00B050"
        )
        self.progress_bar.set(0)
        self.progress_bar.pack_forget()  # masqué tant qu’on n’a pas lancé

        # --- Label chemin (sous tableau à gauche) ---
        label_chemin = ctk.CTkLabel(
            left_frame, text="",
            font=("Segoe UI", 13), text_color="white",
            wraplength=500, justify="center"
        )
        label_chemin.pack(pady=(15, 0))

        # (optionnel) bind molette souris pour scroller le canvas
        def _on_mousewheel(event):
            # Windows / Mac compatibility
            delta = -1 * int(event.delta / 120) if hasattr(event, 'delta') else 0
            main_canvas.yview_scroll(delta, "units")
        main_canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def lancer_export_profils(self):
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
        import re
        from datetime import datetime
        import os
        from collections import defaultdict

        # === Fonctions internes ===
        def chercher_excel_commence_par(root_folder, prefix):
            resultats = []
            for dossier, _, fichiers in os.walk(root_folder):
                for fichier in fichiers:
                    if fichier.endswith(".xlsx") and fichier.startswith(prefix):
                        resultats.append(os.path.join(dossier, fichier))
            return resultats

        def lire_date_c6(fichier, feuille="SA_SNCF"):
            try:
                wb = load_workbook(fichier, read_only=True, data_only=True)
                if feuille not in wb.sheetnames:
                    return None
                ws = wb[feuille]
                valeur = ws["C6"].value
                wb.close()
                if isinstance(valeur, datetime):
                    return valeur.date()
                elif isinstance(valeur, str):
                    try:
                        return datetime.strptime(valeur, "%d/%m/%Y").date()
                    except:
                        return None
                return None
            except:
                return None

        def choisir_meilleure_version(fichiers):
            fichiers = sorted(fichiers)
            v3 = [f for f in fichiers if "V3" in os.path.basename(f)]
            if v3:
                return v3[0]
            v2 = [f for f in fichiers if "V2" in os.path.basename(f)]
            if v2:
                return v2[0]
            return fichiers[0] if fichiers else None

        # === Activation de la barre de progression ===
        self.progress_label.configure(text="⏳ Export en cours...")
        self.progress_label.pack(pady=(10, 0))
        self.progress_bar.pack(pady=(5, 20))
        self.progress_bar.set(0)
        self.update_idletasks()

        # --- Utilisation ---
        dossier_depart = self.chemin_dossier
        fichiers_trouves = chercher_excel_commence_par(dossier_depart, "Profil Tréso SNCF")

        # Récupérer dates par fichier
        dates_par_fichier = []
        for f in fichiers_trouves:
            date_c6 = lire_date_c6(f)
            if date_c6:
                dates_par_fichier.append((f, date_c6))

        # Grouper par date
        groupes = defaultdict(list)
        for fichier, date in dates_par_fichier:
            groupes[date].append(fichier)

        # Choisir une seule version par date
        dates_uniques = []
        for date, fichiers in groupes.items():
            meilleur_fichier = choisir_meilleure_version(fichiers)
            if meilleur_fichier:
                dates_uniques.append((meilleur_fichier, date))

        # Trier les résultats par date
        dates_uniques.sort(key=lambda x: x[1])

        # Calcul des différences
        deltas = []
        for i in range(1, len(dates_uniques)):
            _, date1 = dates_uniques[i-1]
            _, date2 = dates_uniques[i]
            delta = (date2 - date1).days
            deltas.append(delta)

        # Calcul cumulatif
        cumul = [0]
        somme = 0
        for d in deltas:
            somme += d
            cumul.append(somme)

        # === Chemins des fichiers ===
        fichier_source = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Données\Données Cashsolve 0101 310525.xlsx"
        fichier_prev = [fichier for fichier, _ in dates_uniques]
        fichier_destination = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Résultats\Historique_prev_reel_filiales3.xlsx"

        # === Feuilles à traiter ===
        sections = [ {"source": "1_-_SA_SNCF", "prev": "SA_SNCF", "dest": "SNCF_SA"}, 
                    {"source": "3_-_SA_VOYAGEURS", "prev": "SA_VOYAGEURS", "dest": "SA_VOYAGEURS"},
                    {"source": "2_-_RESEAU", "prev": "RESEAU", "dest": "RESEAU"}, 
                    {"source": "4_-_G&C", "prev": "G&C", "dest": "G&C"},
                    {"source": "5_-_GEODIS", "prev": "GEODIS", "dest": "GEODIS"}, ]

        # === Styles ===
        fill_jaune = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        fill_bleu = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
        align_center = Alignment(horizontal='center', vertical='center')
        font_bold = Font(bold=True)

        # === Extraire les dates des chemins ===
        dates_prev = []
        for path in fichier_prev:
            match = re.search(r"\d{4}-\d{2}-\d{2}", path)
            if match:
                date_obj = datetime.strptime(match.group(), "%Y-%m-%d")
                dates_prev.append(date_obj.strftime("%d/%m"))  # JJ/MM
            else:
                dates_prev.append("N/A")

        nb_prev = len(fichier_prev)
        taille_bloc = 2 + 2 * nb_prev

        # === Boucle principale ===
        total = len(sections)
        for idx_section, section in enumerate(sections, start=1):
            feuille_source = section["source"]
            feuille_prev = section["prev"]
            feuille_dest = section["dest"]

            # --- Ton traitement inchangé ---
            df = pd.read_excel(fichier_source, sheet_name=feuille_source, header=None, skiprows=4)
            dates = df.iloc[0, 4:].dropna().reset_index(drop=True)
            lignes_valides = df.iloc[4:, [2] + list(range(4, df.shape[1]))].dropna(subset=[2])
            noms = lignes_valides.iloc[:, 0].astype(str).str.strip().reset_index(drop=True)
            valeurs_reelles = lignes_valides.iloc[:, 1:].reset_index(drop=True)

            wb = load_workbook(fichier_destination)
            ws = wb[feuille_dest]

            start_col = 3
            for idx, nom in enumerate(noms):
                col1 = get_column_letter(start_col)
                col10 = get_column_letter(start_col + 9)
                ws.merge_cells(f"{col1}2:{col10}2")
                titre_cell = ws[f"{col1}2"]
                titre_cell.value = nom
                titre_cell.fill = fill_jaune
                titre_cell.font = font_bold
                titre_cell.alignment = align_center
                titre_cell.border = border

                sous_titres = ["Dates", "Réel (K€)"]
                for date_str in dates_prev:
                    sous_titres.append(f"Prévision {date_str} (K€)")
                    sous_titres.append(f"Écart {date_str} (K€)")

                for i, titre in enumerate(sous_titres):
                    cell = ws.cell(row=3, column=start_col - 1 + i)
                    cell.value = titre
                    cell.fill = fill_bleu
                    cell.font = font_bold
                    cell.alignment = align_center
                    cell.border = border

                for i, date in enumerate(dates):
                    row = 4 + i
                    ws.cell(row=row, column=start_col - 1, value=pd.to_datetime(date, dayfirst=True).date()).alignment = align_center
                    ws.cell(row=row, column=start_col - 1).border = border

                    valeur = valeurs_reelles.iloc[idx, i]
                    val_k = round(valeur / 1000) if pd.notna(valeur) else None
                    ws.cell(row=row, column=start_col, value=val_k).alignment = align_center
                    ws.cell(row=row, column=start_col).border = border

                start_col += taille_bloc + 1

            # --- Écriture des prévisions ---
            valeurs_prev_all = []
            for f in fichier_prev:
                df_prev = pd.read_excel(f, sheet_name=feuille_prev, header=None, skiprows=5)
                raw_noms = df_prev.iloc[0, 5:]
                mask_valid = raw_noms.notna() & (raw_noms.astype(str).str.strip() != "")
                noms_prev = raw_noms[mask_valid].astype(str).str.strip().reset_index(drop=True)
                valeurs = df_prev.iloc[:, 5:].loc[:, mask_valid.values].reset_index(drop=True)
                valeurs.columns = noms_prev
                valeurs_prev_all.append(valeurs.reset_index(drop=True))

            nb_lignes = len(dates)
            start_col = 3
            for idx in range(len(noms)):
                for i in range(nb_lignes):
                    row = 4 + i
                    val_reel = ws.cell(row=row, column=start_col).value

                    for j, decal in enumerate(cumul):
                        prev_col = start_col + 1 + j * 2
                        ecart_col = start_col + 2 + j * 2
                        if i < decal or j >= len(valeurs_prev_all):
                            ws.cell(row=row, column=prev_col, value=None)
                            ws.cell(row=row, column=ecart_col, value=None)
                        else:
                            prev_data = valeurs_prev_all[j]
                            val = prev_data.iloc[i - decal, idx] if (i - decal < len(prev_data)) else None
                            val_k = round(val / 1000) if pd.notna(val) else None
                            ecart = val_k - val_reel if val_k is not None and val_reel is not None else None

                            ws.cell(row=row, column=prev_col, value=val_k).alignment = align_center
                            ws.cell(row=row, column=prev_col).border = border

                            ws.cell(row=row, column=ecart_col, value=ecart).alignment = align_center
                            ws.cell(row=row, column=ecart_col).border = border

                            col_letter = get_column_letter(ecart_col)
                            plage = f"{col_letter}4:{col_letter}{3 + nb_lignes}"
                            rule_rouge = FormulaRule(formula=[f"${col_letter}4<0"], font=Font(color="FF0000", bold=True))
                            rule_vert = FormulaRule(formula=[f"${col_letter}4>0"], font=Font(color="00B050", bold=True))
                            ws.conditional_formatting.add(plage, rule_rouge)
                            ws.conditional_formatting.add(plage, rule_vert)

                start_col += taille_bloc + 1

            wb.save(fichier_destination)
            print(f"✅ Feuille {feuille_dest} traitée avec succès.")

            # === Mise à jour PROGRESS BAR ===
            progress = idx_section / total
            self.progress_bar.set(progress)
            self.progress_label.configure(text=f"⏳ Export en cours... {int(progress*100)}%")
            self.update_idletasks()

        # === Fin ===
        self.progress_bar.set(1)
        self.progress_label.configure(text="✅ Export terminé !")

    def creer_fichier_excel(self):
        # === Demande de confirmation ===
        confirmer = messagebox.askyesno(
            "Confirmation",
            "Voulez-vous vraiment créer le fichier Excel ?\n"
            "Le fichier existant sera remplacé si présent."
        )
        if not confirmer:
            return  # Sort de la fonction si l'utilisateur clique sur Non

        # === Chemin du fichier à créer ===
        fichier_destination = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Résultats\Historique_prev_reel_filiales3.xlsx"

        # === Sections avec noms de feuilles ===
        sections = [
            {"source": "1_-_SA_SNCF", "prev": "SA_SNCF", "dest": "SNCF_SA"},
            {"source": "3_-_SA_VOYAGEURS", "prev": "SA_VOYAGEURS", "dest": "SA_VOYAGEURS"},
            {"source": "2_-_RESEAU", "prev": "RESEAU", "dest": "RESEAU"},
            {"source": "4_-_G&C", "prev": "G&C", "dest": "G&C"},
            {"source": "5_-_GEODIS", "prev": "GEODIS", "dest": "GEODIS"},
        ]

        # === Crée le dossier s'il n'existe pas ===
        dossier_resultats = os.path.dirname(fichier_destination)
        os.makedirs(dossier_resultats, exist_ok=True)

        # === Supprime le fichier s'il existe déjà ===
        if os.path.exists(fichier_destination):
            os.remove(fichier_destination)
            print(f"🗑️ Fichier existant supprimé : {fichier_destination}")

        # === Crée un nouveau classeur ===
        wb = Workbook()

        # Supprime la feuille par défaut
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Crée toutes les feuilles définies dans sections
        for s in sections:
            wb.create_sheet(title=s["dest"])

        # Sauvegarde le fichier
        wb.save(fichier_destination)

        print(f"✅ Nouveau fichier Excel créé avec les feuilles : {[s['dest'] for s in sections]}")
         # === Message visuel à l'utilisateur ===
        messagebox.showinfo(
            "Fichier créé",
            f"Le fichier Excel a été créé avec succès.\n\nChemin :\n{fichier_destination}"
        )
   
    def afficher_ecarts(self):
        import customtkinter as ctk
        from tkinter import ttk
        import tkinter as tk
        from PIL import Image
        from customtkinter import CTkImage

        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - ÉCARTS IMPORTANTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === Colonnes du tableau ===
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (k€)", "Écart (%)"]

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral", "Créances CDP",
            "Placements", "CC financiers", "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]

        decaissements = [
            "Péages", "Charges de personnel", "ACE & Investissements"
        ]

        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital",
            "Sous total financier", "Free cash Flow", "Emprunts",
            "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cash flow de financement",
            "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
            "Sous total dépenses"
        ]

        # === Fonction pour convertir en flux ===
        def en_flux(values):
            values = [float(v) if v is not None else None for v in values]
            if not values or all(v is None for v in values):
                return values
            flux = [0 if values[0] is not None else None]
            for i in range(1, len(values)):
                v, v_prev = values[i], values[i - 1]
                flux.append(v - v_prev if v is not None and v_prev is not None else None)
            return flux

        # === Récupération et calcul des écarts ===
        ecarts_data = []
        repartition = {feuille: 0 for feuille in sections.values()}


        for feuille in sections.values():
            ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
            for nom, col_start in noms_colonnes:
                dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                for i, date in enumerate(dates):
                    if i >= len(reel) or reel[i] is None:
                        continue
                    for idx, prev_list in enumerate(previsions):
                        if i >= len(prev_list) or prev_list[i] is None:
                            continue

                        r = reel[i]
                        prev_val = prev_list[i]

                        if r == 0 and prev_val == 0:
                            continue
                        elif prev_val == 0:
                            prev_val = 1

                        ecart = (r - prev_val) / prev_val


                        if abs(ecart) >= 0.4:
                            profil_label = noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}"
                            repartition[feuille] += 1

                            ecarts_data.append((
                                date,                    # Date
                                profil_label,            # Profil
                                feuille,                 # Filiale / Feuille
                                nom,                     # Flux
                                round(reel[i], 2),       # Réel
                                round(prev_val, 2),      # Prévision
                                round(reel[i]-prev_val, 2),  # Écart k€
                                round(ecart * 100, 1)    # Écart %
                            ))


        # === PIE CHART : % d'écarts par rapport au total des écarts toutes filiales confondues ===
        feuilles = list(repartition.keys())
        total_ecarts = sum(repartition.values())
        
        valeurs = []
        for f in feuilles:
            if total_ecarts > 0:
                pourcentage = (repartition[f] / total_ecarts) * 100
                print(f"{pourcentage:.1f}%, nb écarts: {repartition[f]}, total écarts : {total_ecarts}")
            else:
                pourcentage = 0
            valeurs.append(pourcentage)
            
        ecarts_data.sort(key=lambda x: abs(x[7]), reverse=True)

        # === Frame filtres et boutons ===
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(padx=30, pady=(10, 0), fill="x")

        filtre_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        filtre_frame.pack(side="left", fill="x", expand=True)

        colonnes_filtrables = ["Date", "Profil", "Filiales", "Flux"]
        filtres = {}
        valeurs_uniques = {col: set() for col in colonnes_filtrables}
        for row in ecarts_data:
            for i, col in enumerate(colonnes):
                if col in colonnes_filtrables:
                    valeurs_uniques[col].add(str(row[i]))

        for col in colonnes_filtrables:
            label = ctk.CTkLabel(filtre_frame, text=f"{col} :", font=("Segoe UI", 11, "bold"))
            label.pack(side="left", padx=(0,2), pady=5)
            combo_frame = ctk.CTkFrame(filtre_frame, fg_color="#0084ff", corner_radius=8)
            combo_frame.pack(side="left", padx=(0,8), pady=5)
            valeurs = ["Tous"] + sorted(valeurs_uniques[col])
            combo = ttk.Combobox(combo_frame, values=valeurs, state="readonly", width=15, font=("Segoe UI", 11, "bold"))
            combo.set("Tous")
            combo.pack(padx=5, pady=2, fill="x")
            filtres[col] = combo

        btn_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        btn_frame.pack(side="right")

        btn_retour = ctk.CTkButton(
            btn_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_retour.pack(side="top", pady=5)

        btn_export = ctk.CTkButton(
            btn_frame,
            text="📊 Exporter en Excel",
            command=lambda: self.exporter_ecarts_excel(ecarts_data),
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#0078D7",
            hover_color="#005A9E",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_export.pack(side="top", pady=5)

        # === Bouton pour visualiser graphiquement ===
        btn_graphique = ctk.CTkButton(
            btn_frame,
            text="📈 Visualiser graphiquement",
            command=lambda: self.analyser_ecarts_ml(),  # Appelle ta fonction ML/graphique
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#FC7100", 
            hover_color="#6C4100",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_graphique.pack(side="top", pady=5)


        # === Treeview ===
        tree = ttk.Treeview(self, columns=colonnes, show="headings", height=25)
        for col in colonnes:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=130)
        tree.pack(pady=10, padx=30, fill="both", expand=True)

        tree.tag_configure("neg", foreground="red")
        tree.tag_configure("pos", foreground="green")

        # === Fonctions de formatage ===
        def format_milliers(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ")
                return str(val)
            except Exception:
                return str(val)

        def format_pourcentage(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ") + "%"
                return str(val)
            except Exception:
                return str(val)

        # === Fonction pour déterminer si un écart est favorable ===
        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                # Corrigé pour prendre en compte les valeurs négatives
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                if prev_val >= 0:
                    return reel_val >= prev_val
                else:
                    return abs(reel_val) <= abs(prev_val)
            else:
                return reel_val >= prev_val


        # === Fonction pour afficher les données ===
        def afficher_donnees(donnees):
            for row in tree.get_children():
                tree.delete(row)
            for data in donnees:
                date_str = data[0].strftime("%Y-%m-%d")
                reel_str = format_milliers(data[4])
                prev_str = format_milliers(data[5])
                ecart_k_str = format_milliers(data[6])
                ecart_pct_str = format_pourcentage(data[7])

                flux_nom = data[3]
                reel_val, prev_val = data[4], data[5]
                favorable = est_favorable(flux_nom, reel_val, prev_val)
                tags = ("pos",) if favorable else ("neg",)

                tree.insert(
                    "",
                    "end",
                    values=(date_str, data[1], data[2], flux_nom, reel_str, prev_str, ecart_k_str, ecart_pct_str),
                    tags=tags
                )

        # Premier affichage
        afficher_donnees(ecarts_data)

        # === Filtrage ===
        def appliquer_filtre(event=None):
            filtred = ecarts_data
            for i, col in enumerate(colonnes):
                if col in filtres:
                    val = filtres[col].get()
                    if val != "Tous":
                        filtred = [row for row in filtred if str(row[i]) == val]
            afficher_donnees(filtred)

        for combo in filtres.values():
            combo.bind("<<ComboboxSelected>>", appliquer_filtre)

    def analyser_ecarts_ml(self):
        import tkinter as tk
        import customtkinter as ctk
        import pandas as pd
        import numpy as np
        from sklearn.cluster import KMeans
        from sklearn.preprocessing import StandardScaler
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox, ttk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors
        import matplotlib.colors as mcolors
        import matplotlib.cm as cm

        try:
            self.vider_fenetre()

            # === HEADER ===
            header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
            header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)
            titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - ANALYSE ML DES ÉCARTS",
                                        font=("Segoe UI Semibold", 26, "bold"))
            titre_label.pack(side="left", anchor="w")

            # Logo
            try:
                image_path = r"C:\Users\0304336A\...\logo_Pulse.png"
                logo_img = Image.open(image_path)
                resized_logo = logo_img.resize((90, 40), Image.Resampling.LANCZOS)
                ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo)
                logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
                logo_label.image = ctk_logo
                logo_label.pack(side="right", anchor="e", padx=(10, 0))
            except Exception as e:
                print(f"Erreur chargement logo: {e}")

            barre = ctk.CTkFrame(self, height=2, fg_color="#00aced")
            barre.pack(side="top", fill="x", pady=(0, 15))

            # === FRAME PRINCIPALE SCROLLABLE ===
            container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
            container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

            main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
            scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
            scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
            scrollable_frame.bind("<Configure>", lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all")))
            main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            main_canvas.configure(yscrollcommand=scrollbar.set)
            main_canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # === CONTROLES HORIZONTAUX ===
            control_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=15)
            control_frame.pack(padx=20, pady=10, fill="x")

            ttk.Label(control_frame, text="Sélectionnez une filiale :", background="#00122e",
                    foreground="white", font=('Segoe UI', 12)).pack(side="left", padx=(0, 10))

            valeurs_combo = ["Toutes les filiales"] + list(sections.values())
            feuille_combo = ttk.Combobox(control_frame, values=valeurs_combo, state="readonly", width=35)
            feuille_combo.pack(side="left", padx=(0, 10))

            btn_afficher = ctk.CTkButton(control_frame, text="Afficher l'analyse ML", width=200,
                                        fg_color="#00aced", hover_color="#0099e6")
            btn_afficher.pack(side="left", padx=(0, 10))

            btn_retour_tableau = ctk.CTkButton(control_frame, text="Retour au tableau", command=self.afficher_ecarts,
                                            width=200, fg_color="#FC7100", hover_color="#6C4100")
            btn_retour_tableau.pack(side="left", padx=(0, 10))

            btn_retour_menu = ctk.CTkButton(control_frame, text="Retour au menu", command=self.retour_menu,
                                            width=200, fg_color="#444", hover_color="#666")
            btn_retour_menu.pack(side="left", padx=(0, 10))

            # === FRAME GRAPHIQUE ===
            graph_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=15)
            graph_frame.pack(padx=20, pady=20, fill="both", expand=True)

            def afficher_graphique():
                feuille = feuille_combo.get()
                if not feuille:
                    messagebox.showinfo("Sélection", "Veuillez sélectionner une filiale.")
                    return

                # === COLLECTE DES DONNÉES ===
                data = []
                feuilles_a_traiter = list(sections.values()) if feuille == "Toutes les filiales" else [feuille]

                valorisation_ecarts = {f: 0 for f in sections.values()}

                for ws_feuille in feuilles_a_traiter:
                    ws, noms_colonnes = charger_donnees(ws_feuille, taille_bloc)
                    for nom, col_start in noms_colonnes:
                        dates, reel, previsions, profils = extraire_valeurs(ws, col_start, nb_prev)
                        for i, date in enumerate(dates):
                            if i >= len(reel) or reel[i] is None:
                                continue
                            for prev_list in previsions:
                                if i >= len(prev_list) or prev_list[i] is None:
                                    continue
                                r = reel[i]
                                prev_val = prev_list[i]
                                if r == 0 and prev_val == 0:
                                    continue
                                elif r == 0:
                                    r = 1
                                ecart_pct = (prev_val - r) / r * 100
                                if abs(ecart_pct) < 40:
                                    continue
                                data.append({"ecart_pct": ecart_pct, "filiale": ws_feuille})
                                valorisation_ecarts[ws_feuille] += abs(prev_val - r)

                if not data:
                    messagebox.showinfo("Analyse ML", "Aucune donnée exploitable (écarts trop faibles).")
                    return

                df = pd.DataFrame(data)
                X = df[['ecart_pct']].values
                scaler = StandardScaler()
                X_scaled = scaler.fit_transform(X)
                kmeans = KMeans(n_clusters=4, random_state=42, n_init=20)
                df['cluster'] = kmeans.fit_predict(X_scaled)
                centroids_scaled = kmeans.cluster_centers_
                centroids = scaler.inverse_transform(centroids_scaled)
                df['y_val'] = np.random.uniform(0, 1, size=len(df))

                # Nettoyage du graph_frame
                for widget in graph_frame.winfo_children():
                    widget.destroy()

                # === SUMMARY CLUSTER ===
                summary = df.groupby('cluster')['ecart_pct'].agg(['count', 'mean', 'std']).round(2)
                summary = summary.sort_values('mean')
                cluster_order = {old: new for new, old in enumerate(summary.index)}
                df['cluster_ordered'] = df['cluster'].map(cluster_order)

                palette = ["#FC7100", "#000DFF", "#00CC66", "#FF0000"]
                colors = {i: palette[i] for i in range(4)}

                plt.style.use("seaborn-v0_8-darkgrid")
                fig, ax = plt.subplots(figsize=(11, 6))
                ax.scatter(df['ecart_pct'], df['y_val'], c=df['cluster_ordered'].map(colors),
                        s=50, alpha=0.85, edgecolors='none')

                # Centroids dynamiques
                for i, (cluster_id, row) in enumerate(summary.iterrows()):
                    centroid_val = centroids[cluster_id][0]
                    ax.scatter(centroid_val, 0.5, c='black', marker='X', s=160)

                # Légende simplifiée
                for i, (cluster_id, row) in enumerate(summary.iterrows()):
                    ax.scatter([], [], color=colors[i], label=f"Cluster {i+1}", s=70)
                ax.scatter([], [], color='black', label="Centroid", s=70)
                ax.legend(title="Clusters", loc='center left', bbox_to_anchor=(1, 0.5))

                ax.set_xlabel("Écart (%)", fontsize=12)
                ax.get_yaxis().set_visible(False)
                titre = f"Visualisation de la dispersion des écarts (%) - {feuille if feuille != 'Toutes les filiales' else 'Ensemble des filiales'}"
                ax.set_title(titre, fontsize=15, fontweight="bold")
                ax.grid(True, linestyle="--", alpha=0.4)

                # Curseur interactif
                cursor = mplcursors.cursor(ax.collections[0], hover=True)
                @cursor.connect("add")
                def on_hover(sel):
                    idx = sel.index
                    ecart = df.iloc[idx]["ecart_pct"]
                    filiale = df.iloc[idx]["filiale"]
                    sel.annotation.set_text(f"Filiale: {filiale}\nÉcart: {ecart:.2f} %")
                    sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

                # Intégration Tkinter
                canvas_fig = FigureCanvasTkAgg(fig, master=graph_frame)
                canvas_fig.draw()
                canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)

                toolbar_frame = tk.Frame(graph_frame)
                toolbar_frame.pack()
                toolbar = NavigationToolbar2Tk(canvas_fig, toolbar_frame)
                toolbar.update()
                plt.close(fig)

                # === TABLEAU RÉCAPITULATIF SOUS LE GRAPHIQUE (VALO BRUTE) ===
                table_frame = ctk.CTkFrame(graph_frame, fg_color="#00122e")
                table_frame.pack(pady=10, fill="x")

                ttk.Label(table_frame, text="Résumé des clusters :",
                        background="#00122e", foreground="white", font=('Segoe UI Semibold', 13)
                        ).pack(anchor="w", padx=5, pady=(0,5))

                tree = ttk.Treeview(table_frame, columns=["Cluster", "Nb écarts", "Valorisation totale (k€)"],
                                    show='headings', height=4)
                for col in ["Cluster", "Nb écarts", "Valorisation totale (k€)"]:
                    tree.heading(col, text=col)
                    tree.column(col, width=150, anchor="center")

                # Calcul de la valorisation brute par cluster
                valorisation_par_cluster = {}
                for cluster_id in summary.index:
                    idx_cluster = df[df['cluster'] == cluster_id].index
                    val_cluster = df.loc[idx_cluster, 'filiale'].map(valorisation_ecarts).sum()
                    valorisation_par_cluster[cluster_id] = val_cluster

                # Remplissage du tableau
                for i, cluster_id in enumerate(summary.index):
                    cluster_num = i + 1
                    nb_ecarts = f"{int(summary.loc[cluster_id, 'count']):,}".replace(",", " ")
                    val_brute = f"{int(valorisation_par_cluster.get(cluster_id, 0)):,}".replace(",", " ")
                    tree.insert("", "end", values=[cluster_num, nb_ecarts, val_brute])


                # Ligne totale
                total_ecarts = f"{int(summary['count'].sum()):,}".replace(",", " ")
                total_valo = f"{int(sum(valorisation_par_cluster.values())):,}".replace(",", " ")
                tree.insert("", "end", values=["Total", total_ecarts, total_valo], tags=("total",))
                tree.tag_configure("total", background="#444", foreground="white", font=('Segoe UI', 12, 'bold'))

                tree.pack(pady=5, padx=5, fill="x")


            btn_afficher.configure(command=afficher_graphique)

        except Exception as e:
            messagebox.showerror("Erreur ML", f"Une erreur est survenue : {e}")

    def exporter_ecarts_excel(self, ecarts_data):
        # Création du fichier Excel
        wb_out = Workbook()
        ws = wb_out.active
        ws.title = "Écarts Importants"

        # Titres
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]

        # Style pour titres
        titre_font = Font(bold=True, color="000000")
        titre_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Ajout des titres
        for col_num, titre in enumerate(colonnes, 1):
            cell = ws.cell(row=1, column=col_num, value=titre)
            cell.font = titre_font
            cell.fill = titre_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajout des données
        for row_num, data in enumerate(ecarts_data, start=2):
            date_str = data[0].strftime("%Y-%m-%d")
            row_values = [date_str, data[1], data[2], data[3], data[4], data[5], f"{data[6]}%"]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Sauvegarde
        from tkinter import filedialog
        fichier_sortie = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Enregistrer le fichier Excel"
        )
        if fichier_sortie:
            wb_out.save(fichier_sortie)
            messagebox.showinfo("Export réussi", f"Fichier exporté : {fichier_sortie}")
    
    def afficher_repartition(self):
        import customtkinter as ctk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from collections import defaultdict
        import mplcursors
        import numpy as np
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors

        # --- helpers: compatibilité des signatures (lecture) ---
        def _charger(feuille):
            """Wrapper tolérant: essaye charger_donnees(feuille) puis charger_donnees(feuille, taille_bloc)."""
            try:
                return charger_donnees(feuille)
            except TypeError:
                return charger_donnees(feuille, taille_bloc)

        def _extraire(ws, col_start):
            """Wrapper tolérant: essaye extraire_valeurs(ws, col_start) puis avec nb_prev."""
            try:
                return extraire_valeurs(ws, col_start)
            except TypeError:
                return extraire_valeurs(ws, col_start, nb_prev)

        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - RÉPARTITION DES ÉCARTS PAR FILIALE",
            font=titre_font
        )
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE ...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="#00aced")
        barre.pack(side="top", fill="x", pady=(0, 15))

        # === FRAME PRINCIPALE SCROLLABLE AVEC MARGES ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        scrollable_frame = tk.Frame(main_canvas, bg="#00122e")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )

        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=v_scrollbar.set)

        main_canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")

        # === TYPES DE FLUX POUR L'ANALYSE DES ÉCARTS ===
        encaissements = ["Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
                        "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"]
        decaissements = ["Péages", "Charges de personnel", "ACE & Investissements"]
        mixtes = ["Sous total Investissements nets et ACE", "Charges et produits financiers",
                "Dividendes reçus et versés", "Augmentations de capital",
                "Sous total financier", "Free cash Flow", "Emprunts",
                "Tirages Lignes CT", "Change", "Variation de collatéral",
                "Créances CDP", "Placements", "CC financiers",
                "Emprunts / Prêts - Groupe", "Cash flow de financement",
                "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
                "Sous total dépenses"]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                return (reel_val >= prev_val) if prev_val >= 0 else (abs(reel_val) <= abs(prev_val))
            else:
                return reel_val >= prev_val

        # === CALCUL DES ÉCARTS ET VALORISATION ===
        feuilles_all = list(sections.values())
        repartition = {feuille: 0 for feuille in feuilles_all}
        valorisation_ecarts = {feuille: 0 for feuille in feuilles_all}
        details_ecarts = {feuille: [] for feuille in feuilles_all}

        # Lecture + calculs
        for feuille in feuilles_all:
            ws, noms_colonnes = _charger(feuille)
            for nom, col_start in noms_colonnes:
                dates, reel, previsions, noms_profils = _extraire(ws, col_start)
                for i, date in enumerate(dates):
                    if i >= len(reel) or reel[i] is None:
                        continue
                    r = reel[i]
                    for idx, prev_list in enumerate(previsions):
                        if i >= len(prev_list) or prev_list[i] is None:
                            continue
                        prev_val = prev_list[i]

                        # éviter division par zéro
                        if r == 0 and prev_val == 0:
                            continue
                        elif r == 0:
                            r = 1

                        ecart = (prev_val - r) / r  # signe: positif si prev > reel
                        if abs(ecart) >= 0.4:
                            repartition[feuille] += 1
                            valorisation_ecarts[feuille] += (r - prev_val)  # valorisation signée en k€
                            details_ecarts[feuille].append({
                                "date": date,
                                "profil": noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}",
                                "filiale": feuille,
                                "flux": nom,
                                "reel": round(reel[i], 2),
                                "prevision": round(prev_val, 2),
                                "ecart_pct": round(ecart * 100, 1)
                            })

        # === PIE CHART : % d'écarts par rapport au total des écarts toutes filiales confondues ===
        feuilles = feuilles_all
        total_ecarts = sum(repartition.values())
        valeurs = [(repartition[f] / total_ecarts) * 100 if total_ecarts > 0 else 0 for f in feuilles]

        # Filtrer les filiales avec au moins un écart (pour le camembert)
        feuilles_filtrees = [f for i, f in enumerate(feuilles) if valeurs[i] > 0]
        valeurs_filtrees = [v for v in valeurs if v > 0]

        fig, ax = plt.subplots(figsize=(8, 6), facecolor="#00122e")
        ax.set_facecolor("#00122e")
        palette_pie = ["#5D5F83", "#34495E", "#5D6D7E", "#85929E", "#AAB7B8", "#D5DBDB"]

        if not valeurs_filtrees:
            ax.text(0.5, 0.5, "Aucun écart important détecté", ha='center', va='center', fontsize=12, color="white")
            ax.axis('off')
            wedges = []
        else:
            wedges, texts, autotexts = ax.pie(
                valeurs_filtrees,
                labels=feuilles_filtrees,
                autopct='%1.1f%%',
                startangle=140,
                colors=palette_pie[:len(valeurs_filtrees)],
                textprops={'color': 'white', 'fontsize': 10}
            )
            for t in texts + autotexts:
                t.set_color("white")

        ax.set_title(
            "Répartition des écarts significatifs par filiale – en nombre d’occurrences",
            fontsize=14,
            color="white"
        )

        canvas_fig = FigureCanvasTkAgg(fig, master=scrollable_frame)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(pady=20, padx=30, fill="both", expand=True)

        # === BAR CHART : pourcentage d'écarts par filiale ===
        # Pour chaque filiale, % = nb écarts / nb points de prévision non-None
        pourcentage_ecarts_filiales = {}
        nombre_previsions_par_feuille = {}

        for feuille in feuilles_all:
            ecarts = details_ecarts.get(feuille, [])
            nombre_ecarts = len(ecarts)

            ws, noms_colonnes_local = _charger(feuille)
            nombre_previsions = 0
            for nom_colonne, col_start in noms_colonnes_local:
                dates_local, reel_local, previsions_local, noms_profils_local = _extraire(ws, col_start)
                for prev_list in previsions_local:
                    for prev_val in prev_list:
                        if prev_val is not None:
                            nombre_previsions += 1

            nombre_previsions_par_feuille[feuille] = nombre_previsions
            pourcentage = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0
            pourcentage_ecarts_filiales[feuille] = pourcentage

        feuilles_bar = list(pourcentage_ecarts_filiales.keys())
        valeurs_bar = list(pourcentage_ecarts_filiales.values())

        fig_bar, ax_bar = plt.subplots(figsize=(10, 5), facecolor="#00122e")
        ax_bar.set_facecolor("#00122e")

        if valeurs_bar:
            norm = mcolors.Normalize(vmin=min(valeurs_bar), vmax=max(valeurs_bar) if max(valeurs_bar) != min(valeurs_bar) else min(valeurs_bar)+1)
            cmap = cm.Blues
            colors_bar = [cmap(norm(v)) for v in valeurs_bar]
        else:
            colors_bar = []

        bars = ax_bar.bar(feuilles_bar, valeurs_bar, color=colors_bar, alpha=0.9)

        ax_bar.set_title("Fréquence des écarts significatifs par filiale",
                        fontsize=14, color="white")
        ax_bar.set_ylabel("% Écarts", color="white")
        ax_bar.tick_params(axis='x', rotation=45, colors="white")
        ax_bar.tick_params(axis='y', colors="white")
        ax_bar.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
        fig_bar.tight_layout(pad=2.0)

        cursor_bar = mplcursors.cursor(bars, hover=True)
        @cursor_bar.connect("add")
        def on_hover_bar(sel):
            idx = sel.index
            feuille = feuilles_bar[idx]
            sel.annotation.set_text(
                f"{feuille}\n"
                f"Nombre d'écarts: {len(details_ecarts.get(feuille, []))}\n"
                f"Nombre de Prévisions: {nombre_previsions_par_feuille.get(feuille, 0)}\n"
                f"Pourcentage: {valeurs_bar[idx]:.1f}%"
            )
            sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

        canvas_bar = FigureCanvasTkAgg(fig_bar, master=scrollable_frame)
        canvas_bar.draw()
        canvas_bar.get_tk_widget().pack(pady=20, padx=30, fill="both", expand=True)

        # === BAR CHART : valorisation des écarts (toutes filiales, +/- et 0) ===
        graph_valorisation_frame = tk.Frame(scrollable_frame, bg="#00122e")
        graph_valorisation_frame.pack(fill="both", expand=True, padx=10, pady=(10, 20))

        feuilles_val = list(valorisation_ecarts.keys())
        valeurs_val = [0 if valorisation_ecarts[f] is None else valorisation_ecarts[f] for f in feuilles_val]

        fig_val, ax_val = plt.subplots(figsize=(10, 5), facecolor="#00122e")
        ax_val.set_facecolor("#00122e")

        if any(v != 0 for v in valeurs_val):
            max_abs = max(abs(v) for v in valeurs_val)
            if max_abs == 0:
                max_abs = 1
            norm_val = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
            cmap_val = cm.RdBu_r
            colors_val = [cmap_val(norm_val(v)) for v in valeurs_val]
        else:
            # toutes les valeurs = 0 → couleur neutre
            colors_val = ["#8395a7"] * len(valeurs_val)

        bars_val = ax_val.bar(feuilles_val, valeurs_val, color=colors_val, alpha=0.95)

        ax_val.axhline(0, color="white", linewidth=1, alpha=0.7)
        ax_val.set_title("Valorisation totale des écarts pour chaque filiale (k€)",
                        fontsize=14, color="white")
        ax_val.set_ylabel("Écarts cumulés (k€)", color="white")
        ax_val.tick_params(axis='x', rotation=45, colors="white")
        ax_val.tick_params(axis='y', colors="white")
        ax_val.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
        fig_val.tight_layout(pad=2.0)

        cursor_val = mplcursors.cursor(bars_val, hover=True)
        @cursor_val.connect("add")
        def on_add(sel):
            idx = sel.index
            feuille = feuilles_val[idx]
            val = valeurs_val[idx]
            nb_ecarts = repartition.get(feuille, 0)
            nb_prevs = nombre_previsions_par_feuille.get(feuille, 0)
            val_str = f"{int(val):,}".replace(",", " ")
            sel.annotation.set_text(
                f"{feuille}\nÉcarts: {nb_ecarts}\nPrévisions: {nb_prevs}\nValorisation: {val_str} k€"
            )
            sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

        canvas_val = FigureCanvasTkAgg(fig_val, master=graph_valorisation_frame)
        canvas_val.draw()
        canvas_val.get_tk_widget().pack(pady=10, fill="both", expand=True)

        # === HOVER & CLIC SUR PIE (mise en avant + détails) ===
        if wedges:
            original_colors = [w.get_facecolor() for w in wedges]

            def reset_colors():
                for i, w in enumerate(wedges):
                    w.set_facecolor(original_colors[i])
                canvas_fig.draw_idle()

            def on_hover(event):
                if event.inaxes == ax and wedges:
                    found = False
                    for i, w in enumerate(wedges):
                        if w.contains_point((event.x, event.y)):
                            w.set_facecolor("blue")
                            for j, w2 in enumerate(wedges):
                                if j != i:
                                    w2.set_facecolor(original_colors[j])
                            canvas_fig.draw_idle()
                            found = True
                            break
                    if not found:
                        reset_colors()

            def afficher_details_feuille(feuille):
                if hasattr(self, 'frame_details') and self.frame_details is not None:
                    self.frame_details.destroy()

                self.frame_details = tk.Frame(scrollable_frame, bg="#00122e")
                self.frame_details.pack(fill='x', padx=30, pady=(10, 5))

                ecarts = details_ecarts.get(feuille, [])
                nombre_ecarts = len(ecarts)

                # recompute nombre_previsions (au cas où)
                ws, noms_colonnes_local = _charger(feuille)
                nombre_previsions = 0
                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = _extraire(ws, col_start)
                    for prev_list in previsions_local:
                        for prev_val in prev_list:
                            if prev_val is not None:
                                nombre_previsions += 1

                pourcentage_ecarts = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0

                ctk.CTkLabel(
                    self.frame_details,
                    text=(f"Détails pour {feuille} : "
                        f"{nombre_ecarts} écarts sur {nombre_previsions} prévisions "
                        f"({pourcentage_ecarts:.1f}%)"),
                    font=("Arial", 12, "bold")
                ).pack(anchor='w', pady=(0, 10))

                colonnes = ["Date", "Profil", "Filiale", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]
                tree = ttk.Treeview(self.frame_details, columns=colonnes, show="headings", height=5)
                for col in colonnes:
                    tree.heading(col, text=col)
                    tree.column(col, anchor="center", width=120)
                tree.pack(pady=10, fill="x")

                cinq_plus_gros = sorted(ecarts, key=lambda x: abs(x["ecart_pct"]), reverse=True)[:5]
                for e in cinq_plus_gros:
                    date_str = e["date"].strftime("%Y-%m-%d") if hasattr(e["date"], 'strftime') else str(e["date"])
                    ecart_str = f"{e['ecart_pct']}%"
                    tags = ("pos",) if est_favorable(e["flux"], e["reel"], e["prevision"]) else ("neg",)
                    tree.insert("", "end", values=(date_str, e["profil"], e["filiale"], e["flux"],
                                                e["reel"], e["prevision"], ecart_str), tags=tags)
                tree.tag_configure("neg", foreground="red")
                tree.tag_configure("pos", foreground="green")

                self.update_idletasks()
                try:
                    main_canvas.yview_moveto(self.frame_details.winfo_y() / max(1, scrollable_frame.winfo_height()))
                except Exception:
                    pass

            def on_click(event):
                if event.inaxes == ax and wedges:
                    for i, w in enumerate(wedges):
                        if w.contains_point((event.x, event.y)):
                            afficher_details_feuille(feuilles_filtrees[i])
                            break

            canvas_fig.mpl_connect("motion_notify_event", on_hover)
            canvas_fig.mpl_connect("button_press_event", on_click)

        # === Bouton retour ===
        bouton_retour = ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=220,
            height=40,
            corner_radius=12,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_retour.pack(pady=20)

    def afficher_repartition_par_prevision(self):
        from collections import defaultdict
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors

        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - RÉPARTITION PAR PROFIL", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) ...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_height = 40
            new_width = int(new_height * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === FRAME PRINCIPALE (container arrondi) ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        # === Canvas + scrollbars ===
        canvas_container = tk.Frame(container, bg="#00122e")
        canvas_container.pack(fill="both", expand=True)

        main_canvas = tk.Canvas(canvas_container, bg="#00122e", highlightthickness=0)
        v_scrollbar = tk.Scrollbar(canvas_container, orient="vertical", command=main_canvas.yview)
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)

        scrollable_frame = tk.Frame(main_canvas, bg="#00122e")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        )
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        main_canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        # === TITRE ===
        ctk.CTkLabel(
            scrollable_frame,
            text="Répartition des Écarts Importants par Profil (en % et k€)",
            font=("Arial", 18, "bold"),
            text_color="white",
            fg_color="#00122e"
        ).pack(pady=15)

        # === Select filiale ===
        ctk.CTkLabel(
            scrollable_frame,
            text="Sélectionnez une filiale :",
            font=("Arial", 12, "bold"),
            text_color="white",
            fg_color="#00122e"
        ).pack(pady=(10, 2), anchor="center")

        filiales = list(sections.values())
        filiales.insert(0, "Toutes filiales")
        selected_filiale = tk.StringVar(value=filiales[0])

        select_box = ttk.Combobox(
            scrollable_frame,
            textvariable=selected_filiale,
            values=filiales,
            state="readonly",
            width=25
        )
        select_box.pack(pady=(0, 15), anchor="center")

        # === Zones graphique et tableau ===
        graph_frame = tk.Frame(scrollable_frame, bg="#00122e")
        graph_frame.pack(fill="both", expand=True, padx=10, pady=(5, 20))

        table_frame = tk.Frame(scrollable_frame, bg="#00122e")
        table_frame.pack(fill="x", padx=10, pady=(10, 30))

        colonnes = ("Profil", "Nb prévisions", "Nb écarts >=40%", "Taux (%)", "Valorisation (k€)")
        table = ttk.Treeview(table_frame, columns=colonnes, show="headings", height=6)
        for col in colonnes:
            table.heading(col, text=col)
            table.column(col, anchor="center", width=160)
        table.pack(pady=5, fill="x", expand=True)


        # === Fonction maj graphique avec suppression des anciens graphes ===
        def maj_graphique(filiale):
            # Supprimer anciens graphes
            for widget in graph_frame.winfo_children():
                widget.destroy()

            compteur_ecarts = defaultdict(int)
            compteur_total = defaultdict(int)
            valorisation_ecarts = defaultdict(float)

            if filiale == "Toutes filiales":
                filiales_calc = list(sections.values())
            else:
                filiales_calc = [filiale]

            for f in filiales_calc:
                ws, noms_colonnes = charger_donnees(f, taille_bloc)
                for nom_colonne, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                    for p_idx, nom_profil in enumerate(noms_profils):
                        if p_idx >= len(previsions):
                            continue
                        prev_list = previsions[p_idx]
                        for i, r in enumerate(reel):
                            if i >= len(prev_list):
                                continue
                            prev_val = prev_list[i]
                            if prev_val is None:
                                continue

                            compteur_total[nom_profil] += 1

                            if r is None or r == 0:
                                if prev_val == 0:
                                    continue
                                r = 1

                            ecart = (prev_val - r) / r
                            if abs(ecart) >= 0.4:
                                compteur_ecarts[nom_profil] += 1
                                valorisation_ecarts[nom_profil] += abs(prev_val - r)

            noms_final = list(compteur_total.keys())
            pourcentages = [
                (compteur_ecarts[nom] / compteur_total[nom]) * 100 if compteur_total[nom] > 0 else 0
                for nom in noms_final
            ]
            valorisations = [valorisation_ecarts[nom] for nom in noms_final]

            # --- FIGURE 1 : Taux d'écarts ---
            fig1, ax1 = plt.subplots(figsize=(14, 5), facecolor="#00122e")
            ax1.set_facecolor("#00122e")
            if noms_final:
                bars1 = ax1.bar(noms_final, pourcentages, color="#1f77b4", alpha=0.9)
                ax1.plot(noms_final, pourcentages, color="white", marker="o", linewidth=2)
                cursor = mplcursors.cursor(bars1, hover=True)

                @cursor.connect("add")
                def on_add(sel):
                    idx = sel.index
                    sel.annotation.set_text(f"Profil : {noms_final[idx]}\nTaux : {pourcentages[idx]:.2f}%")
                    sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

            titre1 = "Toutes filiales - Taux d'Écarts" if filiale == "Toutes filiales" else f"{filiale} - Taux d'Écarts"
            ax1.set_title(titre1, fontsize=14, color="white")
            ax1.set_ylabel("Taux d'écarts (%)", color="white")
            ax1.tick_params(axis='y', colors="white")
            ax1.tick_params(axis='x', rotation=60, labelcolor="white")
            ax1.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig1.tight_layout(pad=2.0)

            canvas1 = FigureCanvasTkAgg(fig1, master=graph_frame)
            canvas1.draw()
            canvas1.get_tk_widget().pack(pady=(10, 20), fill="both", expand=True)

            # --- FIGURE 2 : Valorisation des écarts (sans texte) ---
            fig2, ax2 = plt.subplots(figsize=(14, 5), facecolor="#00122e")
            ax2.set_facecolor("#00122e")
            if noms_final:
                bars2 = ax2.bar(noms_final, valorisations, color="#28B463", alpha=0.8)
                # Ne pas mettre les valeurs sur les barres pour éviter illisibilité

            titre2 = "Toutes filiales - Valorisation des écarts" if filiale == "Toutes filiales" else f"{filiale} - Valorisation des écarts"
            ax2.set_title(titre2, fontsize=14, color="white")
            ax2.set_ylabel("Valorisation (k€)", color="white")
            ax2.tick_params(axis='y', colors="white")
            ax2.tick_params(axis='x', rotation=60, labelcolor="white")
            ax2.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            fig2.tight_layout(pad=2.0)

            canvas2 = FigureCanvasTkAgg(fig2, master=graph_frame)
            canvas2.draw()
            canvas2.get_tk_widget().pack(pady=(10, 40), fill="both", expand=True)

            # --- Remplissage tableau ---
            for row in table.get_children():
                table.delete(row)

            total_valorisation = 0
            for nom, taux, valo in zip(noms_final, pourcentages, valorisations):
                total_valorisation += valo
                table.insert(
                    "",
                    "end",
                    values=(
                        nom,
                        compteur_total[nom],
                        compteur_ecarts[nom],
                        f"{taux:.2f}%",
                        f"{valo:,.0f}".replace(",", " ")  # ajout formatage
                    )
                )

            total_previsions = sum(compteur_total.values())
            total_ecarts = sum(compteur_ecarts.values())
            taux_total = (total_ecarts / total_previsions * 100) if total_previsions > 0 else 0
            table.insert(
                "",
                "end",
                values=(
                    "TOTAL",
                    total_previsions,
                    total_ecarts,
                    f"{taux_total:.2f}%",
                    f"{total_valorisation:,.0f}".replace(",", " ")
                )
            )


        # === Lier la combobox ===
        select_box.bind("<<ComboboxSelected>>", lambda e: maj_graphique(selected_filiale.get()))
        maj_graphique(filiales[0])

        # === Bouton retour ===
        bouton_retour = ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_retour.pack(pady=20)

    def afficher_repartition_flux(self):
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.colors as mcolors
        import matplotlib.cm as cm

        self.vider_fenetre()

        # === HEADER (titre + logo) ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - RÉPARTITION DES FLUX", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === CONTAINER PRINCIPAL (scrollable) ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_configure(e):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_configure)
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        # === Titre ===
        ctk.CTkLabel(
            scrollable_frame,
            text="Top flux avec le plus d'écarts importants sur l’ensemble des filiales",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        ).pack(pady=15)

        # === Sélecteur de filiale ===
        ctk.CTkLabel(
            scrollable_frame,
            text="Sélectionnez une filiale :",
            font=("Arial", 12, "bold"),
            text_color="white",
            fg_color="#00122e"
        ).pack(pady=(10, 2), anchor="center")

        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = tk.StringVar(value=filiales[0])

        select_box = ttk.Combobox(
            scrollable_frame,
            textvariable=selected_filiale,
            values=filiales,
            state="readonly",
            width=25
        )
        select_box.pack(pady=(0, 5), anchor="center")

        # === Frame pour boutons Top 10 / Top 15 ===
        btn_frame = ctk.CTkFrame(scrollable_frame, fg_color="transparent")
        btn_frame.pack(pady=(0, 15))

        # === Variable pour stocker les widgets graphiques existants ===
        self.graph_widgets = []

        def afficher_graph(top_n=10):
            # Nettoyage anciens graphiques
            for widget in self.graph_widgets:
                widget.destroy()
            self.graph_widgets.clear()

            # === Flux à exclure des graphiques ===
            flux_a_exclure = [
                "Cash flow de financement",
                "Cash flow net",
                "Sous total financier",
                "Sous total Investissements nets et ACE",
                "Free cash Flow",
                "Sous total recettes",
                "Sous total dépenses"
            ]

            filiale_actuelle = selected_filiale.get()
            feuilles = sections.values() if filiale_actuelle == "Toute filiale" else [sections[filiale_actuelle]]

            # --- Calcul du nombre total de prévisions et d'écarts ---
            nb_previsions_filiale = 0
            nb_ecarts_filiale = 0
            repartition_flux = {}

            for feuille in feuilles:
                ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)

                # Nombre de prévisions
                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = extraire_valeurs(ws, col_start, nb_prev)
                    for prev_list in previsions_local:
                        for prev_val in prev_list:
                            if prev_val is not None:
                                nb_previsions_filiale += 1

                # Nombre d'écarts par flux
                for nom, col_start in noms_colonnes_local:
                    if nom in flux_a_exclure:
                      continue
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                    for i, date in enumerate(dates):
                        if i >= len(reel) or reel[i] is None:
                            continue
                        for idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list) or prev_list[i] is None:
                                continue
                            r = reel[i]
                            prev_val = prev_list[i]
                            if r == 0 and prev_val == 0:
                                continue
                            elif r == 0:
                                r = 1
                            ecart = (prev_val - r) / r
                            if abs(ecart) >= 0.4:
                                nb_ecarts_filiale += 1
                                repartition_flux[nom] = repartition_flux.get(nom, 0) + 1

            # --- Label résumé ---
            if hasattr(self, "label_resume") and self.label_resume is not None:
                self.label_resume.destroy()

            self.label_resume = ctk.CTkLabel(
                scrollable_frame,
                text=f"Filiale : {filiale_actuelle} | Nombre total de prévisions : {nb_previsions_filiale} | Nombre d'écarts ≥ 40% : {nb_ecarts_filiale}",
                font=("Segoe UI", 14, "bold"),
                text_color="white"
            )
            self.label_resume.pack(pady=(0, 15))
            self.graph_widgets.append(self.label_resume)

            # Si aucun écart
            if not repartition_flux:
                fig, ax = plt.subplots(figsize=(8, 6), facecolor="#00122e", constrained_layout=True)
                ax.text(0.5, 0.5, "Aucun écart important détecté", ha='center', va='center', fontsize=12, color="white")
                ax.axis("off")
                canvas_fig = FigureCanvasTkAgg(fig, master=scrollable_frame)
                canvas_fig.draw()
                canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)
                self.graph_widgets.append(canvas_fig.get_tk_widget())
                plt.close(fig)
                return

            # --- Top flux par nombre d'écarts ---
            flux_tries = sorted(repartition_flux.items(), key=lambda x: x[1], reverse=True)[:top_n]
            noms_flux = [f for f, _ in flux_tries]
            valeurs = [v for _, v in flux_tries]

            # Valorisation des écarts (signée) par flux
            valeur_ecarts = []
            for flux in noms_flux:
                total_ecart = 0
                for feuille in feuilles:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for nom, col_start in noms_colonnes:
                        if nom in flux_a_exclure:
                            continue
                        if nom != flux:
                            continue
                        dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                        for i, date in enumerate(dates):
                            if i >= len(reel) or reel[i] is None:
                                continue
                            for idx, prev_list in enumerate(previsions):
                                if i >= len(prev_list) or prev_list[i] is None:
                                    continue
                                r = reel[i]
                                prev_val = prev_list[i]
                                if r == 0 and prev_val == 0:
                                    continue
                                elif r == 0:
                                    r = 1
                                ecart = (prev_val - r) / r
                                if abs(ecart) >= 0.4:
                                    total_ecart += (r - prev_val)  # signé
                valeur_ecarts.append(total_ecart)

            # --- Graphique 1 : Nombre d'écarts ---
            fig1, ax1 = plt.subplots(figsize=(12, max(6, len(noms_flux) * 0.5)), facecolor="#00122e", constrained_layout=True)
            ax1.set_facecolor("#00122e")
            bars1 = ax1.barh(noms_flux, valeurs, color="#5DADE2")
            ax1.set_xlabel("Nombre d'écarts importants", color="white", fontsize=12)
            ax1.set_ylabel("Flux", color="white", fontsize=12)
            ax1.tick_params(axis='x', colors="white")
            ax1.tick_params(axis='y', colors="white")
            ax1.set_title(f"Top {top_n} flux – Nombre d'écarts ({filiale_actuelle})", color="white", fontsize=14)
            for bar in bars1:
                ax1.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
                        str(int(bar.get_width())), va='center', color="white")
            canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
            canvas_fig1.draw()
            canvas_fig1.get_tk_widget().pack(pady=10, fill="both", expand=True)
            self.graph_widgets.append(canvas_fig1.get_tk_widget())
            plt.close(fig1)

            # --- Graphique 2 : Valorisation des écarts (gère négatifs/0 + marges dynamiques) ---
            fig2, ax2 = plt.subplots(figsize=(12, max(6, len(noms_flux) * 0.5)), facecolor="#00122e", constrained_layout=True)
            ax2.set_facecolor("#00122e")

            # Palette divergente centrée sur 0
            if valeur_ecarts:
                max_abs = max(abs(v) for v in valeur_ecarts)
            else:
                max_abs = 1
            if max_abs == 0:
                max_abs = 1

            norm = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
            cmap = cm.RdBu_r
            colors2 = [cmap(norm(v)) for v in valeur_ecarts]

            bars2 = ax2.barh(noms_flux, valeur_ecarts, color=colors2, alpha=0.9)

            ax2.axvline(0, color="white", linewidth=1, alpha=0.8)
            ax2.set_xlabel("Valorisation cash des écarts (k€)", color="white", fontsize=12)
            ax2.set_ylabel("Flux", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white")
            ax2.tick_params(axis='y', colors="white")
            ax2.set_title(f"Top {top_n} flux – Valorisation des écarts ({filiale_actuelle})", color="white", fontsize=14)

            # --- Placement intelligent des labels + calcul des marges nécessaires ---
            # Seuil "place suffisante" à l'intérieur: 20% de l'amplitude max
            inside_threshold = 0.20 * max_abs
            # Padding pour labels externes en unités "données"
            pad_data = 0.03 * max_abs

            # On mesure la longueur (en caractères) des libellés qui seront posés à l'extérieur
            longest_out_right_chars = 0
            longest_out_left_chars = 0

            # 1ère passe: décider inside/outside + mémoriser longueur max à droite/gauche
            labels_info = []  # (bar, value, place_inside, sign)
            for bar, v in zip(bars2, valeur_ecarts):
                place_inside = abs(v) >= inside_threshold
                sign = 1 if v >= 0 else -1
                labels_info.append((bar, v, place_inside, sign))

                # Texte affiché
                txt = f"{int(v):,} k€".replace(",", " ")

                if not place_inside:
                    if sign > 0:
                        longest_out_right_chars = max(longest_out_right_chars, len(txt))
                    else:
                        longest_out_left_chars  = max(longest_out_left_chars, len(txt))

            # Estimation marges en fonction des longueurs (1 caractère ≈ 0,012 * max_abs)
            char_unit = 0.012 * max_abs
            extra_right = longest_out_right_chars * char_unit + (pad_data if longest_out_right_chars > 0 else 0)
            extra_left  = longest_out_left_chars  * char_unit + (pad_data if longest_out_left_chars  > 0 else 0)

            # Limites dynamiques
            ax2.set_xlim(-max_abs - extra_left, max_abs + extra_right)

            # 2e passe: dessiner les labels
            for bar, v, inside, sign in labels_info:
                txt = f"{int(v):,} k€".replace(",", " ")
                y = bar.get_y() + bar.get_height()/2

                if inside:
                    # placement DANS la barre, proche de l'extrémité
                    if sign > 0:
                        x = v - pad_data
                        ha = 'right'
                    else:
                        x = v + pad_data
                        ha = 'left'
                    ax2.text(x, y, txt, va='center', ha=ha, color="white", clip_on=True)
                else:
                    # placement HORS de la barre, au-delà de l'extrémité
                    if sign > 0:
                        x = v + pad_data
                        ha = 'left'
                    else:
                        x = v - pad_data
                        ha = 'right'
                    ax2.text(x, y, txt, va='center', ha=ha, color="white", clip_on=True)

            fig2.tight_layout(pad=2.0)

            canvas_fig2 = FigureCanvasTkAgg(fig2, master=scrollable_frame)
            canvas_fig2.draw()
            canvas_fig2.get_tk_widget().pack(pady=10, fill="both", expand=True)
            self.graph_widgets.append(canvas_fig2.get_tk_widget())
            plt.close(fig2)


            # --- Boutons Top 10 / Top 15 ---
            for child in btn_frame.winfo_children():
                child.destroy()
            ctk.CTkButton(
                btn_frame, text=" Afficher le Top 10", command=lambda: afficher_graph(10),
                width=120, corner_radius=15, fg_color="#0078D7", hover_color="#005A9E"
            ).pack(side="left", padx=5, pady=5)
            ctk.CTkButton(
                btn_frame, text=" Afficher le Top 15", command=lambda: afficher_graph(15),
                width=120, corner_radius=15, fg_color="#0078D7", hover_color="#005A9E"
            ).pack(side="left", padx=5, pady=5)

        # ⚡ Lier la combobox au recalcul
        selected_filiale.trace_add("write", lambda *args: afficher_graph())

        # ⚡ Affichage initial
        afficher_graph(15)

        # --- Bouton retour ---
        ctk.CTkButton(
            scrollable_frame, text="⬅️ Retour au menu", command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)

    def afficher_heatmap_ecarts(self):
        import customtkinter as ctk
        from tkinter import ttk, messagebox
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import numpy as np
        import pandas as pd
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.text import Annotation
        from sklearn.ensemble import IsolationForest
        import seaborn as sns
        import matplotlib.patches as patches

        self.vider_fenetre()

        # === HEADER ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - DÉTECTION D’ANOMALIES", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        # Logo
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365)\Projet PULSE\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)
            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()
            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Logo ignoré: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === CONTROLES HORIZONTAUX ===
        control_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        control_frame.pack(side="top", fill="x", pady=10, padx=30)

        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = tk.StringVar(value=filiales[0])
        ttk.Label(control_frame, text="Filiale :", background="#001f3f", foreground="white",
                font=('Segoe UI', 12, 'bold')).pack(side="left", padx=(0, 10))
        select_box = ttk.Combobox(control_frame, textvariable=selected_filiale, values=filiales, state="readonly", width=25)
        select_box.pack(side="left", padx=(0, 20))

        bouton_retour = ctk.CTkButton(control_frame, text="⬅️ Retour au menu", command=self.retour_menu,
                                    width=180, height=40, corner_radius=15,
                                    fg_color="#444", hover_color="#666",
                                    text_color="white", font=("Segoe UI", 13, "bold"))
        bouton_retour.pack(side="left")

        # === CONTENEUR SCROLLABLE ===
        container = tk.Frame(self, bg="#001f3f")
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg="#001f3f", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=canvas.xview)

        scrollable_frame = tk.Frame(canvas, bg="#001f3f")
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")

        # === Tableau dynamique ===
        cols = ["Date", "Flux", "Profil", "Réel (k€)", "Prévision (k€)", "Écart (k€)"]
        tree = ttk.Treeview(scrollable_frame, columns=cols, show="headings", height=10)
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=150, anchor="center")
        tree.pack(pady=10, fill="x")

        tree.tag_configure("neg", foreground="red")
        tree.tag_configure("pos", foreground="green")

        self.canvas_fig = None
        self.fig = None

        # === Fonction pour déterminer si un écart est favorable ===
        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]

        decaissements = [
            "Péages", "Charges de personnel", "ACE & Investissements"
        ]

        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital",
            "Sous total financier", "Free cash Flow", "Emprunts",
            "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cash flow de financement",
            "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
            "Sous total dépenses"
        ]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                if prev_val >= 0:
                    return reel_val >= prev_val
                else:
                    return abs(reel_val) <= abs(prev_val)
            else:
                return reel_val >= prev_val

        # === Fonction principale optimisée ===
        def afficher_heatmap():
            if self.canvas_fig is not None:
                self.canvas_fig.get_tk_widget().destroy()
                self.canvas_fig = None
                self.fig = None

            filiale_actuelle = selected_filiale.get()
            df_all = []

            feuilles = sections.values() if filiale_actuelle == "Toute filiale" else [sections[filiale_actuelle]]
             # === Flux à exclure des graphiques ===
            flux_a_exclure = [
                "Cash flow de financement",
                "Cash flow net",
                "Sous total financier",
                "Sous total Investissements nets et ACE",
                "Free cash Flow",
                "Sous total recettes",
                "Sous total dépenses"
            ]
            for feuille in feuilles:
                ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes_local:
                    if nom_flux in flux_a_exclure:
                        continue
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                    for i, date in enumerate(dates):
                        if i >= len(reel) or reel[i] is None:
                            continue
                        for idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list) or prev_list[i] is None:
                                continue
                            r = reel[i]
                            prev_val = prev_list[i]
                            if r == 0: r = 1
                            ecart = prev_val - r
                            df_all.append({
                                "Date": date,
                                "Flux": nom_flux,
                                "Profil": noms_profils[idx],
                                "Réel": reel[i],
                                "Prévision": prev_val,
                                "Écart": ecart
                            })

            if not df_all:
                messagebox.showinfo("Info", "Aucune donnée exploitable trouvée.")
                return

            df_all = pd.DataFrame(df_all)
            df_all["Écart_abs"] = df_all["Écart"].abs()

            seuil = 2 * df_all["Écart"].std()
            contamination_dyn = min(0.05, max(0.01, (df_all["Écart"].abs() > seuil).mean()))
            iso = IsolationForest(contamination=contamination_dyn, random_state=42)
            df_all["Anomalie"] = iso.fit_predict(df_all[["Écart"]])
            df_all["Anomalie"] = df_all["Anomalie"].map({1: 0, -1: 1})

            heatmap_data = df_all.pivot_table(index="Profil", columns="Flux", values="Anomalie", aggfunc="sum", fill_value=0)
            mean_data = df_all.pivot_table(index="Profil", columns="Flux", values="Écart_abs", aggfunc="mean", fill_value=0)
            heatmap_data = heatmap_data.reindex(df_all['Profil'].drop_duplicates()[::-1])
            mean_data = mean_data.reindex(heatmap_data.index)

            # === Graphique ===
            plt.close("all")
            self.fig, ax = plt.subplots(figsize=(18, max(6, heatmap_data.shape[0]*0.5)), facecolor="#001f3f")

            # Création du heatmap
            sns.heatmap(
                heatmap_data,
                cmap="Reds",
                annot=True,
                fmt=".0f",
                linewidths=0.5,
                linecolor="#444",
                ax=ax,
                cbar=True
            )

            # === Axes et titre en blanc ===
            ax.set_xlabel("Flux", color="white", fontsize=12, fontweight="bold")
            ax.set_ylabel("Profil", color="white", fontsize=12, fontweight="bold")
            ax.set_title(f"Heatmap des anomalies ({filiale_actuelle})", color="white", fontsize=16, fontweight="bold")

            # Ticks des axes en blanc
            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=10, color="white")
            plt.setp(ax.get_yticklabels(), rotation=0, fontsize=10, color="white")

            # === Colorbar ===
            cbar = ax.collections[0].colorbar
            cbar.set_label("Nombre d'anomalies", color="white", fontsize=12, fontweight="bold")
            cbar.ax.yaxis.set_tick_params(color='white')
            plt.setp(cbar.ax.yaxis.get_ticklabels(), color='white')
            cbar.outline.set_edgecolor('white')

            self.fig.tight_layout(pad=2.0)
            self.fig.subplots_adjust(bottom=0.25)


            # === Tooltip ===
            tooltip = Annotation("", xy=(0, 0), xytext=(15, 15), textcoords="offset points",
                                ha="left", va="bottom",
                                bbox=dict(boxstyle="round", fc="black", ec="white", lw=1, alpha=0.8),
                                color="white", fontsize=9)
            tooltip.set_visible(False)
            ax.add_artist(tooltip)

            import matplotlib.patches as patches
            hover_rect = patches.Rectangle((0, 0), 1, 1, fill=True, edgecolor="black", linewidth=2, facecolor="blue", alpha=0.3)
            hover_rect.set_visible(False)
            ax.add_patch(hover_rect)



            def get_cell(event):
                if event.inaxes != ax or event.xdata is None or event.ydata is None:
                    return None, None
                x, y = int(event.xdata), int(event.ydata)
                if x < 0 or y < 0 or x >= heatmap_data.shape[1] or y >= heatmap_data.shape[0]:
                    return None, None
                return x, y

            def on_hover(event):
                x, y = get_cell(event)
                if x is None:
                    hover_rect.set_visible(False)
                    tooltip.set_visible(False)
                    self.fig.canvas.draw_idle()
                    return
                hover_rect.set_xy((x, y))
                hover_rect.set_visible(True)

                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]
                n_anomalies = heatmap_data.iloc[y, x]
                avg_ecart = mean_data.iloc[y, x]

                tooltip.xy = (event.xdata, event.ydata)
                tooltip.set_text(f"{profil} / {flux}\nAnomalies : {n_anomalies}\n")
                tooltip.set_visible(True)
                self.fig.canvas.draw_idle()

            def on_click(event):
                x, y = get_cell(event)
                if x is None: return
                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]
                tree.delete(*tree.get_children())

                filtered = df_all[(df_all["Flux"] == flux) & (df_all["Profil"] == profil) & (df_all["Anomalie"] == 1)].copy()
                filtered["Écart_abs"] = filtered["Écart"].abs()
                filtered = filtered.sort_values(by="Écart_abs", ascending=False)

                for _, row in filtered.iterrows():
                    date_str = row["Date"].strftime("%d/%m/%Y") if hasattr(row["Date"], "strftime") else str(row["Date"])
                    reel_val, prev_val = row["Réel"], row["Prévision"]
                    favorable = est_favorable(row["Flux"], reel_val, prev_val)
                    tag = "pos" if favorable else "neg"

                    tree.insert("", "end", values=[
                        date_str,
                        row.get("Flux", ""),
                        row.get("Profil", ""),
                        f"{row.get('Réel', 0):,.0f}".replace(",", " "),
                        f"{row.get('Prévision', 0):,.0f}".replace(",", " "),
                        f"{row.get('Écart', 0):,.0f}".replace(",", " ")
                    ], tags=(tag,))

            self.fig.canvas.mpl_connect("motion_notify_event", on_hover)
            self.fig.canvas.mpl_connect("button_press_event", on_click)

            self.canvas_fig = FigureCanvasTkAgg(self.fig, master=scrollable_frame)
            self.canvas_fig.draw()
            self.canvas_fig.get_tk_widget().pack(pady=10, fill="both", expand=True)

        selected_filiale.trace_add("write", lambda *args: afficher_heatmap())
        afficher_heatmap()

    def creer_page_graphique(self):
        import tkinter as tk
        from tkinter import ttk
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage

        self.vider_fenetre()

        # === HEADER avec logo et titre ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - VISUALISATION GRAPHIQUE DES ÉCARTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\...\logo_Pulse.png"
            logo_img = Image.open(image_path)

            # Mesure de la hauteur de la police pour harmoniser la taille du logo
            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="#00aced")
        barre.pack(side="top", fill="x", pady=(0, 15))

        # 🔹 Extraction des profils (inchangé)
        for feuille in sections.values():
            ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
            for nom, col_start in noms_colonnes:
                dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)

        # === FRAME PRINCIPALE SCROLLABLE AVEC SCROLLBAR HORIZONTALE ET VERTICALE ===
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)  # marges extérieures

        # Canvas principal (pour le scroll)
        self.main_canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)

        # Scrollbars
        self.scrollbar_y = tk.Scrollbar(container, orient="vertical", command=self.main_canvas.yview)
        # ⬇️ Scrollbar horizontale positionnée en bas de la fenêtre principale (prend toute la largeur)
        self.scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=self.main_canvas.xview)

        # Frame scrollable à l'intérieur du canvas
        self.scrollable_frame = ctk.CTkFrame(self.main_canvas, fg_color="#00122e", corner_radius=0)

        # Quand la taille du contenu change, on met à jour la zone de scroll
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )

        # Ajout de la frame dans le canvas (ne PAS forcer la largeur -> permet le scroll horizontal si contenu large)
        self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")

        # Liaison des scrollbars
        self.main_canvas.configure(yscrollcommand=self.scrollbar_y.set, xscrollcommand=self.scrollbar_x.set)

        # Placement des éléments
        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar_y.pack(side="right", fill="y")
        self.scrollbar_x.pack(side="bottom", fill="x")  # ✅ visible en bas, pleine largeur

        # === CONTENU DANS LE SCROLLABLE FRAME ===
        self.center_frame = ctk.CTkFrame(self.scrollable_frame, fg_color="#00122e", corner_radius=15)
        # Astuce : on évite de forcer fill="x" ici pour ne pas étirer à la largeur totale et laisser le H-scroll jouer son rôle
        self.center_frame.pack(pady=20, padx=30, fill="both", expand=True)

        # 🔹 Sélection filiale
        ttk.Label(self.center_frame, text="Sélectionnez une filiale :", background="#00122e", foreground="white",
                font=('Segoe UI', 12)).pack(pady=(10, 5))
        self.feuille_combo = ttk.Combobox(self.center_frame, values=list(sections.values()), state="readonly", width=35)
        self.feuille_combo.pack(pady=(0, 15))
        self.feuille_combo.bind("<<ComboboxSelected>>", self.charger_noms)

        # 🔹 Sélection flux
        ttk.Label(self.center_frame, text="Sélectionnez un flux :", background="#00122e", foreground="white",
                font=('Segoe UI', 12)).pack(pady=(5, 5))
        self.nom_combo = ttk.Combobox(self.center_frame, state="readonly", width=35)
        self.nom_combo.pack(pady=(0, 15))

        # 🔹 Checkbox Réel
        self.var_reel = tk.BooleanVar(value=True)
        chk_reel = tk.Checkbutton(self.center_frame, text="Réel", variable=self.var_reel,
                                bg="#00122e", fg="white", font=('Segoe UI', 11),
                                selectcolor="#00aced", activebackground="#003366",
                                activeforeground="white")
        chk_reel.pack(anchor="w", padx=20, pady=5)
        chk_reel.bind("<Enter>", lambda e, c=chk_reel: c.config(bg="#003366"))
        chk_reel.bind("<Leave>", lambda e, c=chk_reel: c.config(bg="#00122e"))

        # 🔹 Checkbox Profils avec disposition sur 3 lignes max
        self.vars_prev = []
        profils_frame = tk.Frame(self.center_frame, bg="#00122e")
        profils_frame.pack(pady=15, padx=15, fill="x")

        nb_cols = max(1, len(noms_profils) // 3 + (1 if len(noms_profils) % 3 else 0))

        for i, nom_profil in enumerate(noms_profils):
            var = tk.BooleanVar(value=False)
            cb = tk.Checkbutton(profils_frame, text=nom_profil, variable=var,
                                bg="#00122e", fg="white", font=('Segoe UI', 10),
                                selectcolor="#00aced", activebackground="#003366",
                                activeforeground="white")
            row = i % 3
            col = i // 3
            cb.grid(row=row, column=col, sticky="w", padx=12, pady=6)
            self.vars_prev.append(var)

            # Hover effet
            cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
            cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

        # 🔹 Boutons
        self.btn_afficher = ctk.CTkButton(self.center_frame, text="Afficher Graphique",
                                        command=self.afficher_graphique, width=240,
                                        fg_color="#00aced", hover_color="#0099e6",
                                        corner_radius=12)
        self.btn_afficher.pack(pady=(25, 10))

        self.btn_afficher_cumule = ctk.CTkButton(
            self.center_frame, text="Afficher Graphique Cumulé",
            command=self.afficher_graphique_cumule, width=240,
            fg_color="#0078D7", hover_color="#005A9E",
            corner_radius=12
        )
        self.btn_afficher_cumule.pack(pady=(10, 10))

        self.btn_retour = ctk.CTkButton(self.center_frame, text="Retour au menu",
                                        command=self.retour_menu, width=240,
                                        fg_color="#444", hover_color="#666",
                                        corner_radius=12)
        self.btn_retour.pack(pady=(5, 20))

        # Canvas matplotlib (sera créé dans afficher_graphique / afficher_graphique_cumule)
        self.canvas = None

    def charger_noms(self, event):
        feuille = self.feuille_combo.get()
        try:
            self.ws, self.noms_colonnes = charger_donnees(feuille, taille_bloc)
            noms_affichables = [nom for nom, _ in self.noms_colonnes]
            self.nom_combo['values'] = noms_affichables
            self.nom_combo.set('')
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors du chargement de la feuille : {e}")

    def afficher_graphique(self):
        import matplotlib.pyplot as plt
        import mplcursors
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import matplotlib.ticker as mticker
        import tkinter as tk

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un nom.")
            return

        col_start = dict(self.noms_colonnes).get(nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Nom sélectionné invalide.")
            return

        dates, reel, previsions, noms_profils = extraire_valeurs(self.ws, col_start, nb_prev)

        # Conversion en flux si nécessaire
        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i-1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux

            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        # Nettoyage ancien graphique
        if hasattr(self, "canvas") and self.canvas:
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
        if hasattr(self, "toolbar_frame") and self.toolbar_frame:
            self.toolbar_frame.destroy()
            self.toolbar_frame = None

        # === Création de la figure (plus grande) ===
        plt.style.use("seaborn-v0_8-darkgrid")
        fig, ax = plt.subplots(figsize=(16, 8))
        palette = plt.cm.tab10.colors

        # Données réelles
        if self.var_reel.get():
            ax.plot(dates, reel, label="Réel", color="black", linewidth=2, marker="o")

        # Prévisions
        for i, var in enumerate(self.vars_prev):
            if var.get() and i < len(previsions):
                y = previsions[i]
                ax.plot(
                    dates, y,
                    label=noms_profils[i] if i < len(noms_profils) else f"Profil {i+1}",
                    alpha=0.8,
                    linewidth=1.8,
                    marker=".",
                    color=palette[i % len(palette)]
                )

        # Personnalisation
        feuille = self.feuille_combo.get()
        ax.set_title(f"{feuille} - {nom_selectionne}", fontsize=18, fontweight="bold")
        ax.set_xlabel("Date", fontsize=12)
        ax.set_ylabel("Valeur (k€)" if nom_selectionne not in noms_a_convertir_flux else "Flux", fontsize=12)
        ax.legend(loc="upper left", bbox_to_anchor=(1, 1))
        ax.grid(True, linestyle="--", alpha=0.6)

        # Format des nombres
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " ")))

        # Interaction
        mplcursors.cursor(ax, hover=True)

        # === Intégration du graphique dans Tkinter (plein largeur) ===
        self.canvas = FigureCanvasTkAgg(fig, master=self.center_frame)
        self.canvas.draw()
        # Occupe toute la largeur disponible du frame scrollable
        self.canvas.get_tk_widget().pack(pady=20)

        # === Barre d’outils ===
        self.toolbar_frame = tk.Frame(self.center_frame, bg="#00122e")
        self.toolbar_frame.pack(pady=(5, 10))
        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()

    def afficher_graphique_cumule(self):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import tkinter as tk
        import matplotlib.ticker as mticker
        import pandas as pd
        import numpy as np
        import mplcursors

        # --- Sélection / validation ---
        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un nom.")
            return

        col_start = dict(self.noms_colonnes).get(nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Nom sélectionné invalide.")
            return

        # --- Lecture des séries ---
        dates, reel, previsions, noms_profils = extraire_valeurs(self.ws, col_start, nb_prev)

        # --- Conversion en flux si nécessaire ---
        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement", "Endettement Net"
        ]
        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i-1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux
            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        # --- Nettoyage ancien rendu ---
        if hasattr(self, "canvas") and self.canvas:
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
        if hasattr(self, "toolbar_frame") and self.toolbar_frame:
            self.toolbar_frame.destroy()
            self.toolbar_frame = None

        # --- DataFrame ---
        df = pd.DataFrame({"Date": dates, "Réel": reel})
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Mois"] = df["Date"].dt.strftime("%Y-%m")

        # --- Colonnes de prévision pour profils cochés ---
        profils_selectionnes = []
        for i, var in enumerate(self.vars_prev):
            if var.get() and i < len(previsions):
                nom_prof = noms_profils[i]
                df[nom_prof] = previsions[i]
                profils_selectionnes.append(nom_prof)

        print("\n=== CALCUL CUMULÉ (par profil & par mois) ===")
        print(f"Flux : {nom_selectionne}")
        print(f"Profils cochés : {profils_selectionnes}\n")

        # --- Réel de base (impression par mois) ---
        print("--- Réel de base ---")
        for mois, g in df.groupby("Mois", dropna=True):
            nb_j_reel = int(g["Réel"].notna().sum())
            val_reel = float(g["Réel"].sum()) if nb_j_reel > 0 else 0.0
            print(f"Mois {mois} | Réel total = {val_reel} (jours={nb_j_reel})")
        print()

        # --- Combinaison par profil : Prévision sinon Réel
        #     + Masquage des mois sans AUCUNE prévision pour ce profil ---
        series_combinees = {}
        for nom_prof in profils_selectionnes:
            # Série quotidienne : prev sinon réel
            comb = pd.Series(df[nom_prof])
            comb = comb.where(~comb.isna(), df["Réel"])

            # Détecter les mois contenant au moins 1 prévision
            tmp_prev = pd.DataFrame({"Mois": df["Mois"], "Prev": df[nom_prof]})
            mois_avec_prev = set(tmp_prev.loc[tmp_prev["Prev"].notna(), "Mois"].dropna().unique())
            mois_tous = set(df["Mois"].dropna().unique())
            mois_masques = sorted(mois_tous - mois_avec_prev)

            # Masquer complètement les mois sans prévision pour ce profil
            mask_keep = df["Mois"].isin(mois_avec_prev)
            comb = comb.where(mask_keep, np.nan)

            # Debug : lister les mois masqués
            if mois_masques:
                print(f"[{nom_prof}] Mois masqués (0 prévision) : {', '.join(mois_masques)}")

            # Print demandés (seulement sur les mois gardés = contenant ≥1 prev)
            print(f"--- Profil : {nom_prof} ---")
            tmp = pd.DataFrame({
                "Mois": df["Mois"],
                "Reel": df["Réel"],
                "Prev": df[nom_prof],
            })
            tmp = tmp[tmp["Mois"].isin(mois_avec_prev)]  # ne garder que les mois avec prev
            if tmp.empty:
                print("(Aucun mois avec prévision — ce profil n'apparaîtra pas sur le graphe.)\n")
            else:
                # Indicateurs journaliers
                tmp["has_prev"] = tmp["Prev"].notna()
                tmp["use_reel_equil"] = (~tmp["has_prev"]) & (tmp["Reel"].notna())
                for mois, g in tmp.groupby("Mois", dropna=True):
                    # prévision
                    nb_j_prev = int(g["has_prev"].sum())
                    val_prev = float(g.loc[g["has_prev"], "Prev"].sum()) if nb_j_prev > 0 else 0.0
                    # réel d'équilibrage (jours sans prev mais avec réel)
                    nb_j_reel_eq = int(g["use_reel_equil"].sum())
                    val_reel_eq = float(g.loc[g["use_reel_equil"], "Reel"].sum()) if nb_j_reel_eq > 0 else 0.0
                    # réel pur du mois (informatif)
                    nb_j_reel = int(g["Reel"].notna().sum())
                    val_reel = float(g["Reel"].sum()) if nb_j_reel > 0 else 0.0
                    total = val_prev + val_reel_eq

                    print(
                        f"Mois {mois} | Réel pur = {val_reel} (jours={nb_j_reel}) | "
                        f"Prévision = {val_prev} (jours={nb_j_prev}) | "
                        f"Réel ajouté (équilibrage) = {val_reel_eq} (jours={nb_j_reel_eq}) | "
                        f"Total combiné = {total}"
                    )
                print()

            # Enregistrer la série (avec mois sans prev => NaN)
            series_combinees[f"{nom_prof} (Prévision sinon Réel)"] = comb

        # --- Passage au mensuel ---
        df_comb = pd.DataFrame({"Mois": df["Mois"], "Réel": df["Réel"]})
        for nom_serie, s in series_combinees.items():
            df_comb[nom_serie] = s
        df_cumule = df_comb.groupby("Mois", as_index=True).sum(min_count=1)

        # --- Tracé du graphique ---
        plt.style.use("seaborn-v0_8-darkgrid")
        fig, ax = plt.subplots(figsize=(14, 7))

        n_mois = len(df_cumule)
        n_series = len(series_combinees) + 1  # +1 pour le Réel
        positions = np.arange(n_mois)
        largeur_barre = 0.8 / max(1, n_series)
        palette = plt.cm.tab10.colors
        bar_containers = []

        # Réel pur en noir
        bars_reel = ax.bar(
            positions,
            df_cumule["Réel"].fillna(0).values,
            width=largeur_barre,
            label="Réel",
            color="black",
        )
        bar_containers.append(bars_reel)

        # Profils (seules barres des mois avec prev seront dessinées)
        for i, (nom_serie, _) in enumerate(series_combinees.items()):
            offset = (i + 1) * largeur_barre
            bars = ax.bar(
                positions + offset,
                df_cumule[nom_serie].fillna(0).values,  # les mois sans prev ont été NaN → somme = 0
                width=largeur_barre,
                label=nom_serie,
                color=palette[i % len(palette)],
            )
            bar_containers.append(bars)

        # Axe X / style
        ax.set_xticks(positions + largeur_barre * (n_series - 1) / 2)
        ax.set_xticklabels(df_cumule.index, rotation=45)
        ax.set_title(
            f"{self.feuille_combo.get()} - {nom_selectionne}",
            fontsize=16, fontweight="bold"
        )
        ax.set_xlabel("Mois", fontsize=12)
        ax.set_ylabel("Valeur cumulée (k€)", fontsize=12)
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " ")))
        ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))

        # Hover interactif
        cursor = mplcursors.cursor(bar_containers, hover=True)
        @cursor.connect("add")
        def on_add(sel):
            bars = sel.artist
            idx = sel.index
            serie = bars.get_label()
            val = bars.datavalues[idx]
            mois = df_cumule.index[idx]
            sel.annotation.set_text(
                f"Série : {serie}\nMois : {mois}\nValeur : {int(val):,} K€".replace(",", " ")
            )
            sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

        # Intégration Tkinter
        self.canvas = FigureCanvasTkAgg(fig, master=self.center_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(pady=10, fill="x", expand=True)
        if hasattr(self, "_bind_responsive_matplotlib"):
            self._bind_responsive_matplotlib(self.canvas, self.center_frame, height_ratio=0.5, min_height=420)

        self.toolbar_frame = tk.Frame(self.center_frame)
        self.toolbar_frame.pack()
        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()

    def afficher_heatmap_ecarts2(self):
        import customtkinter as ctk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        from collections import defaultdict
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import tkinter as tk

        self.vider_fenetre()

        # === HEADER avec titre et logo ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - VISUALISATION GRAPHIQUE DES ÉCARTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        # --- Logo ---
        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365)\Projet PULSE\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)
            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === CONTENEUR PRINCIPAL AVEC SCROLLBAR HORIZONTALE ET VERTICALE ===
        container = tk.Frame(self, bg="#001f3f")
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container, bg="#001f3f", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollbar_x = tk.Scrollbar(self, orient="horizontal", command=canvas.xview)

        scrollable_frame = tk.Frame(canvas, bg="#001f3f")

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)

        # Placement
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")  # occupe toute la largeur de la fenêtre

        # === Titre secondaire ===
        ttk.Label(
            scrollable_frame,
            text="Carte thermique des écarts par mois et profil",
            font=("Segoe UI Semibold", 20, "bold"),
            foreground="white",
            background="#001f3f"
        ).pack(pady=15)

        # -----------------------
        # Partie calcul des données
        # -----------------------
        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement",
            "Endettement Net"
        ]

        data_dict = defaultdict(lambda: defaultdict(float))

        for feuille in sections.values():
            ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
            for nom, col_start in noms_colonnes:
                if nom in noms_a_convertir_flux:
                    continue
                dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                for i, date in enumerate(dates):
                    if i >= len(reel) or reel[i] in (None, 0):
                        continue
                    mois = pd.to_datetime(date).strftime("%Y-%m")
                    for p_idx, prev_list in enumerate(previsions):
                        if i >= len(prev_list):
                            continue
                        prev_val = prev_list[i]
                        if prev_val is None:
                            continue
                        ecart = (prev_val - reel[i]) / reel[i]
                        if abs(ecart) >= 0.4:
                            profil_name = noms_profils[p_idx]
                            data_dict[profil_name][mois] += 1

        heatmap_data = pd.DataFrame(data_dict).T.fillna(0)

        # -----------------------
        # Partie affichage graphique
        # -----------------------
        fig, ax = plt.subplots(figsize=(16, 10), facecolor="#001f3f")
        sns.heatmap(
            heatmap_data,
            cmap="coolwarm",
            annot=True,
            fmt=".1f",
            linewidths=0.5,
            linecolor="#444",
            ax=ax,
            cbar_kws={'label': "Nombre d'écarts significatifs"}
        )

        cbar = ax.collections[0].colorbar
        cbar.ax.yaxis.label.set_color("white")
        cbar.ax.tick_params(colors="white")

        ax.set_title("", fontsize=16, color="white")
        ax.set_xlabel("Mois", color="white")
        ax.set_ylabel("Profil", color="white")
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right", color="white")
        plt.setp(ax.get_yticklabels(), color="white")

        canvas_fig = FigureCanvasTkAgg(fig, master=scrollable_frame)
        canvas_fig.draw()
        canvas_fig.get_tk_widget().pack(pady=10)

        # === Bouton retour ===
        bouton_retour = ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_retour.pack(pady=20)

if __name__ == "__main__":
    app = Application()
    app.mainloop()

 