from spire.pdf.common import *
from spire.pdf import *
import os
import sys
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from copy import copy
import zipfile
import re

# --- Fonction pour nettoyer les fichiers Excel corrompus ---
def clean_xlsx(path):
    """
    Supprime les caractères XML interdits dans sharedStrings.xml
    et recrée un fichier réparé.
    """
    temp_path = path.replace(".xlsx", "_clean.xlsx")
    with zipfile.ZipFile(path, 'r') as zin, zipfile.ZipFile(temp_path, 'w') as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/sharedStrings.xml":
                # Supprimer caractères interdits (0x00-0x1F sauf tab, LF, CR)
                data = re.sub(rb"[\x00-\x08\x0B\x0C\x0E-\x1F]", b"", data)
            zout.writestr(item, data)
    return temp_path


# 📂 Dossier contenant les PDF découpés
pdf_folder = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\pdf à convertir\pdf_en_parties"

# 📂 Dossier pour stocker les fichiers Excel intermédiaires
excel_parts_folder = os.path.join(pdf_folder, "Excels_parts")
os.makedirs(excel_parts_folder, exist_ok=True)

# 📂 Dossier pour stocker les fichiers Excel finaux (fusionnés)
excel_final_folder = os.path.join(pdf_folder, "Excels_final")
os.makedirs(excel_final_folder, exist_ok=True)

# 📄 Liste des fichiers PDF découpés
pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith(".pdf")]

if not pdf_files:
    print("⚠️ Aucun fichier PDF trouvé dans le dossier.")
    sys.exit(0)


# --- Étape 1 : Conversion PDF → Excel (chaque partie)
for pdf_file in pdf_files:
    pdf_path = os.path.abspath(os.path.join(pdf_folder, pdf_file))
    excel_file = os.path.abspath(os.path.join(excel_parts_folder, os.path.splitext(pdf_file)[0] + ".xlsx"))

    print(f"🔄 Conversion en cours : {pdf_file} ...")
    pdf = PdfDocument()
    try:
        pdf.LoadFromFile(pdf_path)
        if pdf.Pages.Count == 0:
            print(f"⚠️ Fichier vide ignoré : {pdf_file}")
            continue

        options = XlsxLineLayoutOptions(True, True, True, True, False)
        pdf.ConvertOptions.SetPdfToXlsxOptions(options)

        pdf.SaveToFile(excel_file, FileFormat.XLSX)
        print(f"✅ Converti avec succès : {pdf_file} -> {excel_file}")

    except Exception as e:
        print(f"❌ Erreur lors de la conversion de {pdf_file} : {e}")
        traceback.print_exc()
    finally:
        pdf.Close()


# --- Étape 2 : Regrouper les fichiers Excel par document original
excel_files = [f for f in os.listdir(excel_parts_folder) if f.lower().endswith(".xlsx")]

# Exemple : cc_part1.xlsx, cc_part2.xlsx → regroupés dans cc.xlsx
grouped = {}
for f in excel_files:
    base_name = f.split("_part")[0]  # on enlève "_partX"
    grouped.setdefault(base_name, []).append(f)


# --- Étape 3 : Fusionner les feuilles Excel (avec styles + nettoyage)
for base_name, files in grouped.items():
    print(f"📚 Fusion des parties pour : {base_name}")

    final_wb = Workbook()
    final_ws = final_wb.active
    final_ws.title = "Feuille1"
    first = True

    for f in sorted(files):
        part_path = os.path.join(excel_parts_folder, f)

        try:
            cleaned_path = clean_xlsx(part_path)
            wb = load_workbook(cleaned_path, data_only=False)
        except Exception as e:
            print(f"❌ Impossible d’ouvrir {f} : {e}")
            continue

        for sheet in wb.worksheets:
            # Nom de feuille max 31 caractères
            sheet_name = sheet.title[:31]

            if first and final_ws.max_row == 1 and final_ws.max_column == 1 and not final_ws["A1"].value:
                ws_target = final_ws
                ws_target.title = sheet_name
                first = False
            else:
                ws_target = final_wb.create_sheet(title=f"{sheet_name}_{f}"[:31])

            # --- Copier les cellules (valeurs + styles), en ignorant les 2 premières lignes
            for row in sheet.iter_rows(min_row=3):  # commence à la ligne 3
                for cell in row:
                    if isinstance(cell, Cell):  # ignorer les MergedCell
                        new_cell = ws_target.cell(row=cell.row - 2, column=cell.column, value=cell.value)

                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.fill = copy(cell.fill)
                            new_cell.border = copy(cell.border)
                            new_cell.alignment = copy(cell.alignment)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)

            # --- Copier les colonnes & lignes
            for col_letter, col_dim in sheet.column_dimensions.items():
                ws_target.column_dimensions[col_letter].width = col_dim.width

            for row_idx, row_dim in sheet.row_dimensions.items():
                if row_idx > 2:  # ignorer les 2 premières lignes supprimées
                    ws_target.row_dimensions[row_idx - 2].height = row_dim.height

            # --- Copier aussi les zones fusionnées
            for merged_range in sheet.merged_cells.ranges:
                coords = str(merged_range)
                m = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", coords)
                if m:
                    start_col, start_row, end_col, end_row = m.groups()
                    start_row, end_row = int(start_row) - 2, int(end_row) - 2
                    if start_row > 0 and end_row > 0:
                        ws_target.merge_cells(f"{start_col}{start_row}:{end_col}{end_row}")

        wb.close()

    # --- Ajustement automatique des colonnes dans le fichier final
    for sheet in final_wb.worksheets:
        for col_cells in sheet.columns:
            max_length = 0
            first_real_cell = None

            # Cherche une cellule "réelle" pour identifier la colonne
            for cell in col_cells:
                if isinstance(cell, Cell):
                    first_real_cell = cell
                    break

            if not first_real_cell:
                continue  # colonne vide → skip

            col_letter = get_column_letter(first_real_cell.column)

            for cell in col_cells:
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except:
                    pass

            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

    final_path = os.path.join(excel_final_folder, f"{base_name}.xlsx")
    final_wb.save(final_path)
    print(f"✅ Fichier final créé (avec styles conservés et colonnes ajustées) : {final_path}")


print("\n🎯 Processus terminé : PDF découpés → Excels convertis → Excels fusionnés.")
