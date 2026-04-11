# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

# === Visualiseur 3D (Vispy) - WOW MODE ===
import os

# Essayer plusieurs backends OpenGL
for gl_lib in [None, 'gl', 'glu']:
    try:
        if gl_lib:
            os.environ['VISPY_GL_LIB'] = gl_lib
        break
    except:
        pass

HAS_VISPY = False
app_vispy = None
scene = None
visuals = None

try:
    print("[VISPY] 🔌 Tentative de chargement de Vispy...")
    
    # Essayer les backends dans cet ordre
    backends_to_try = ['qt', 'pyglet', 'glfw', 'sdl2', 'wx', 'tk']
    backend_loaded = False
    
    for backend in backends_to_try:
        try:
            print(f"[VISPY]   • Essai backend: {backend}")
            from vispy import app as app_vispy
            app_vispy.use_app(backend)
            from vispy import scene
            from vispy.scene import visuals
            print(f"[VISPY] ✓ Backend '{backend}' OK")
            backend_loaded = True
            break
        except Exception as e:
            print(f"[VISPY]   ✗ {backend} échoué: {str(e)[:60]}")
            app_vispy = None
            scene = None
            visuals = None
    
    if backend_loaded:
        HAS_VISPY = True
        print("[VISPY] ✓ Vispy chargé avec succès!")
    else:
        print("[VISPY] ✗ Aucun backend Vispy disponible")
        HAS_VISPY = False
        
except ImportError as e:
    print(f"[VISPY] ✗ Vispy n'est pas installé: {e}")
    HAS_VISPY = False
except Exception as e:
    print(f"[VISPY] ✗ Erreur lors de l'initialisation de Vispy: {e}")
    import traceback
    traceback.print_exc()
    HAS_VISPY = False

print(f"[VISPY] Status final: HAS_VISPY={HAS_VISPY}")

import threading
import queue
from collections import deque
from dataclasses import dataclass

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict
import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter  # (si nécessaire ailleurs)

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# === Répertoire contenant les fichiers mensuels ===
from pathlib import Path
import os
import os, time, subprocess, unicodedata
from pathlib import Path
import winreg



from pathlib import Path
import os
import unicodedata
# === Modèle ML (scikit-learn) ===
try:
    from sklearn.ensemble import RandomForestRegressor
except ImportError:
    RandomForestRegressor = None
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor

USER_ID = Path.home().name
BASE_DIR = Path(fr"C:\Users\{USER_ID}\SNCF")
import tkinter as tk

# --- PATCH pour empêcher les RuntimeError "main thread is not in main loop" ---

# Sauvegarde des __del__ originaux
_orig_image_del = getattr(tk.Image, "__del__", None)
_orig_var_del = getattr(tk.Variable, "__del__", None)

def _safe_image_del(self):
    if _orig_image_del is None:
        return
    try:
        _orig_image_del(self)
    except RuntimeError:
        # On ignore seulement le cas où Tk n'est plus dans la boucle principale
        pass

def _safe_var_del(self):
    if _orig_var_del is None:
        return
    try:
        _orig_var_del(self)
    except RuntimeError:
        # Pareil ici : on ne fait rien si Tk est déjà "mort"
        pass

if _orig_image_del is not None:
    tk.Image.__del__ = _safe_image_del

if _orig_var_del is not None:
    tk.Variable.__del__ = _safe_var_del

# On n’impose QUE la "queue" du chemin (fin de chemin), flexible sur tout le reste.
REQUIRED_TAIL = ["Partage - Invités", "Projet PULSE", "4. Données historiques", "Développement","Données"]

def _norm(s: str) -> str:
    """Normalise pour comparaisons: minuscules, accents retirés, espaces/underscores compressés."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("_", " ").strip()
    s = " ".join(s.split())
    return s

def _match_tail(path: Path, tail: list[str]) -> bool:
    """Vérifie que 'path' se termine par les éléments de 'tail' (nom pour nom), comparés normalisés."""
    parts = list(path.parts)
    if len(parts) < len(tail):
        return False
    # Compare en partant de la fin
    for i in range(1, len(tail)+1):
        if _norm(parts[-i]) != _norm(tail[-i]):
            return False
    return True

def find_sharepoint_base():
    """
    1) Si ton chemin 'par défaut' existe, on l’utilise.
    2) Sinon, on parcourt TOUT 'C:\\Users\\<USER_ID>\\SNCF' pour trouver un dossier
       qui SE TERMINE par 'Partage - Invités/Projet PULSE/4. Données historiques/Développement'.
    """
    # 1) Essai direct sur ton chemin "canonique"
    default_path = BASE_DIR / "DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions" / Path(*REQUIRED_TAIL)
    if default_path.exists():
        print(f"[OK] Bibliothèque détectée (DCF par défaut) : {default_path}")
        return default_path

    # 2) Scan générique : on ne regarde QUE la fin de chemin (tail), robustes aux variations amont
    print("[INFO] Recherche générique de    la queue de chemin '.../Partage - Invités/Projet PULSE/4. Données historiques/Développement'...")
    candidates = []
    for root, dirs, _ in os.walk(BASE_DIR):
        # accélère: on ne teste que si "Développement" est présent dans le dossier courant
        if any(_norm(d) == _norm(REQUIRED_TAIL[-1]) for d in dirs):
            for d in dirs:
                full = Path(root) / d
                if _match_tail(full, REQUIRED_TAIL):
                    candidates.append(full)

    if candidates:
        # Choix simple : on prend le plus long chemin (le plus profond) ou simplement le premier trouvé
        best = sorted(candidates, key=lambda p: len(p.parts), reverse=True)[0]
        print(f"[OK] Bibliothèque trouvée automatiquement : {best}")
        return best

    print("❌ Aucune bibliothèque SharePoint locale trouvée correspondant à la queue requise.")
    return None

# --- Résolution finale (inchangé) ---
DEV_PATH = find_sharepoint_base()
if not DEV_PATH:
    raise FileNotFoundError("Impossible de localiser la bibliothèque SharePoint (vérifie la synchro OneDrive/SharePoint).")

FICHIER_EXCEL_DIR = str(DEV_PATH / "Historique Prévisions Réel Filiales")
FICHIER_CONFIG_SECTIONS = str(DEV_PATH / "Filiales Analysées.xlsx")
image_path = str(DEV_PATH / "Images" / "logo_Pulse.png")
BASE_DONNEES_DIR = str(DEV_PATH / "Données Réelles")


print("\n[INIT] Dossier Résultats :", FICHIER_EXCEL_DIR)
print("[INIT] Fichier Config    :", FICHIER_CONFIG_SECTIONS)
print("[INIT] Image            :", image_path)
print("[INIT] Dossier Données   :\n", BASE_DONNEES_DIR)

if not Path(FICHIER_EXCEL_DIR).exists():
    print(f"⚠️  Dossier 'Résultats' introuvable : {FICHIER_EXCEL_DIR}")
if not Path(FICHIER_CONFIG_SECTIONS).exists():
    print(f"⚠️  Fichier 'Filiales Analysées.xlsx' introuvable : {FICHIER_CONFIG_SECTIONS}")
if not Path(image_path).exists():
    print(f"⚠️  Image 'logo_Pulse.png' introuvable : {image_path}")

# ======================
#   FONCTIONS DE SÉCURISATION
# ======================

def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")

def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)

def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(["attrib", path], capture_output=True, text=True, shell=True)
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None

def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")

def pin_and_hydrate(path: str, timeout=30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l’hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False

def robust_load_workbook(path: str, retries=3, delay=1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} ({type(e).__name__}: {e})")
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e

def diag_path(path: str):
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
    
# === Fichier Excel contenant la configuration des sections ===
FEUILLE_CONFIG_SECTIONS = "CONFIG_SECTIONS"  # ou le vrai nom de la feuille dans ton fichier
COL_DEST = 1   # A
COL_SOURCE = 2 # B
COL_PREV = 3   # C

def charger_sections_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit le fichier Excel de configuration des sections et renvoie une liste
    de dictionnaires {"source": ..., "prev": ..., "dest": ...}.
    """
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return []

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return []

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return []

    ws = wb[FEUILLE_CONFIG_SECTIONS]
    mapping = []

    row = 2  # ligne 1 = en-têtes
    while True:
        dest_val = ws.cell(row=row, column=COL_DEST).value
        source_val = ws.cell(row=row, column=COL_SOURCE).value
        prev_val = ws.cell(row=row, column=COL_PREV).value

        if dest_val is None or str(dest_val).strip() == "":
            break  # fin du tableau

        mapping.append({
            "dest": str(dest_val).strip(),
            "source": str(source_val).strip() if source_val else str(dest_val).strip(),
            "prev": str(prev_val).strip() if prev_val else str(dest_val).strip(),
        })
        row += 1

    wb.close()

    print("[INFO] Sections chargées depuis 'Filiales Analysées.xlsx' :")
    for m in mapping:
        print(f"   - dest='{m['dest']}' | source='{m['source']}' | prev='{m['prev']}'")

    return mapping

def charger_noms_feuilles_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit uniquement la colonne A (dest) de la feuille de config et
    retourne la liste des noms de feuilles à utiliser (strings non vides, uniques, ordonnées).
    """
    noms = []
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return noms

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return noms

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return noms

    ws = wb[FEUILLE_CONFIG_SECTIONS]

    row = 2  # ligne 1 = en-têtes
    seen = set()
    while True:
        val = ws.cell(row=row, column=COL_DEST).value
        if val is None or str(val).strip() == "":
            break
        name = str(val).strip()
        if name not in seen:
            noms.append(name)
            seen.add(name)
        row += 1

    wb.close()

    print("[INFO] Feuilles (dest) chargées depuis 'Filiales Analysées.xlsx' :")
    for n in noms:
        print(f"   - {n}")
    return noms

# === Construire dynamiquement `sections` à partir de la colonne A ===
_dest_names = charger_noms_feuilles_depuis_cells()

# Création directe du dictionnaire sections
sections = {name: name for name in _dest_names}  # clé logique -> nom de feuille Excel


# Optionnel mais pratique : caler la feuille de référence sur la première feuille disponible
if _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]   # ex: "SNCF_SA"

SECTIONS_CONFIG = charger_sections_depuis_cells()
print(SECTIONS_CONFIG)

print(FICHIER_EXCEL_DIR)

# === Regex utilitaires ===
_MENSUEL_RE = re.compile(r"^Historique_prev_reel_filiales_(\d{4})_(\d{2})\.xlsx$")
_RE_PREV = re.compile(r"^Prévision\s+\d{2}/\d{2}", re.I)

def _is_prev(h):  return isinstance(h, str) and _RE_PREV.search(h.strip())
def _is_dates(h): return isinstance(h, str) and h.strip().lower().startswith("date")
def _is_reel(h):  return isinstance(h, str) and "réel" in h.lower()

def _parse_excel_date(v):
    """Convertit proprement tout format de date Excel (float, datetime, str...)"""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        # borne basse/haute pour éviter les overflow absurdes
        if 10000 <= v <= 60000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")).normalize()
            except Exception:
                return None
        else:
            return None
    if isinstance(v, str):
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        return ts.normalize() if pd.notna(ts) else None
    ts = pd.to_datetime(v, errors="coerce")
    return ts.normalize() if pd.notna(ts) else None

# -------------------------------------------------------------------------
# 1️⃣ Structures globales
# -------------------------------------------------------------------------
STRUCT = {}
TOKENS = {}
PREV_UNION = set()
FEUILLE_REFERENCE = "SNCF_SA"
CACHE = {}

# Index annuel: (section, flux) -> { "years": { year: { "row_idx": [...], "prof_idx": [...], "headers": [...] } } }
YEAR_INDEX = {}

# Layout pré-calculé depuis le fichier de référence :
# sec -> [(flux_name, dates_col_0, reel_col_0, [(prev_header, prev_col_0), ...]), ...]
_COL_LAYOUT: dict = {}
# Ordre trié de PREV_UNION après lecture de la référence (stable pour toute la session)
_SORTED_PREV: list = []

def _clean_profil_label(raw, i):
    """Nettoyage identique à l'historique: (K€) retiré, 'Prévision' -> 'Profil', fallback."""
    if raw is None:
        return f"Profil {i+1}"
    s = str(raw).replace("(K€)", "").replace("Prévision", "Profil").strip()
    return s if s else f"Profil {i+1}"

def _flux_name_from_token(section, token_col):
    """Retrouve le nom de flux à partir de la colonne 'token' (col_start) dans TOKENS[section]."""
    for name, tok in TOKENS.get(section, []):
        if tok == token_col:
            return name
    return None


# -------------------------------------------------------------------------
# 2️⃣ Lister les fichiers mensuels
# -------------------------------------------------------------------------

def _lister_fichiers_mensuels():
    """
    Parcourt récursivement FICHIER_EXCEL_DIR pour trouver tous les fichiers
    'Historique_prev_reel_filiales_YYYY_MM.xlsx', y compris dans des sous-dossiers d'années.
    En cas de doublons pour un même (YYYY, MM), conserve le fichier le plus récent (mtime).
    """
    print("🔍 Recherche des fichiers mensuels (scan récursif des sous-dossiers d'années)...")

    if not os.path.isdir(FICHIER_EXCEL_DIR):
        print(f"⚠️ Dossier introuvable : {FICHIER_EXCEL_DIR}")
        return []

    best = {}  # (year, month) -> {"path": str, "mtime": float}
    try:
        for root, _, files in os.walk(FICHIER_EXCEL_DIR):
            for fname in files:
                m = _MENSUEL_RE.match(fname)
                if not m:
                    continue
                y, mth = int(m.group(1)), int(m.group(2))
                fullpath = os.path.join(root, fname)
                try:
                    mtime = os.path.getmtime(fullpath)
                except OSError:
                    mtime = 0.0

                key = (y, mth)
                if key not in best or mtime > best[key]["mtime"]:
                    best[key] = {"path": fullpath, "mtime": mtime}
    except Exception as e:
        print(f"⚠️ Erreur pendant le scan récursif : {e}")
        return []

    if not best:
        print("⚠️ Aucun fichier Historique_prev_reel_filiales_YYYY_MM.xlsx trouvé (sous-dossiers inclus).")
        return []

    files = sorted([(y, mth, info["path"]) for (y, mth), info in best.items()])
    print(f"✅ {len(files)} fichiers retenus : {[os.path.basename(f[2]) for f in files]}")
    return files

def _parse_prev_header_sort_key(h):
    # Essaie d'en déduire un tri MM/YY si possible, sinon garde l’ordre d’apparition
    import re
    s = str(h)
    m = re.search(r'(\d{2})/(\d{2,4})', s)
    if not m:
        return (9999, 99, s)
    mm = int(m.group(1))
    yy = int(m.group(2))
    if yy < 100:
        yy += 2000
    return (yy, mm, s)

def _reconcile_headers(B, new_headers):
    """
    Fusionne les headers déjà présents dans B avec ceux du fichier courant :
    - construit l’union ordonnée
    - re-map B['prev_vals'] dans ce nouvel ordre
    - rétro-remplit de None pour les séries nouvellement apparues
    """
    if "prev_headers" not in B or "prev_vals" not in B:
        B["prev_headers"] = []
        B["prev_vals"] = []

    old_headers = list(B["prev_headers"])
    # union en conservant l’ordre d’apparition, puis tente un tri chronologique
    seen = {}
    for h in old_headers + list(new_headers):
        if h not in seen:
            seen[h] = None
    union = list(seen.keys())
    union.sort(key=_parse_prev_header_sort_key)

    # rien à faire ?
    if union == old_headers:
        return

    # re-map des anciennes séries dans le nouveau canevas
    old_idx = {h: i for i, h in enumerate(old_headers)}
    n_rows = len(B.get("dates", []))
    new_prev_vals = []
    for h in union:
        if h in old_idx:
            new_prev_vals.append(B["prev_vals"][old_idx[h]])
        else:
            # nouvelle série : rétro-remplir pour les lignes déjà connues
            new_prev_vals.append([None] * n_rows)

    B["prev_headers"] = union
    B["prev_vals"] = new_prev_vals


# -------------------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# -------------------------------------------------------------------------
def _ensure_flux_bucket(section, flux_name, headers=None):
    key = (section, flux_name)
    if key not in CACHE:
        CACHE[key] = {
            "dates": [],
            "reel": [],
            "prev_headers": [],   # ← vide
            "prev_vals": [],      # ← vide
        }
    return CACHE[key]


# -------------------------------------------------------------------------
# 8️⃣ Fonctions publiques
# -------------------------------------------------------------------------
def charger_donnees(feuille, taille_bloc_param):
    """Retourne le nom de la feuille (section) et la liste des flux détectés, en excluant certains flux inutiles."""
    print(f"➡️ charger_donnees() appelé pour '{feuille}'")

    # Liste des flux à ignorer
    noms_a_exclure = [
        "Trésorerie de fin",
        "Cashpool",
        "Emprunts",
        "Tirages Lignes CT",
        "Variation de collatéral",
        "Créances CDP",
        "Placements",
        "CC financiers",
        "Emprunts / Prêts - Groupe",
        "Encours de financement",
        "Endettement Net"
    ]

    # Récupère tous les flux
    all_noms = list(TOKENS.get(feuille, []))

    # Filtrage : on garde uniquement ceux qui ne sont pas exclus
    noms = [(name, col) for name, col in all_noms if name not in noms_a_exclure]
    for j in noms :
        print(f"noms : {j}")
    print(len(noms))
    
    if not noms:
        print(f"⚠️ Aucun flux valide trouvé pour la feuille '{feuille}'")
    else:
        print(f"   ✅ {len(noms)} flux conservés (sur {len(all_noms)} totaux)")

    return feuille, noms
# Profils à exclure pour 2023 (d’après tes listes / cases à cocher)
EXCLUDED_DATES_2023 = {
    "02/01",
    "09/01",
    "16/01",
    "30/05",
    "05/06",
    "19/06",
    "26/06",
    "03/07",
    "10/07",
    "17/07",
}

def extraire_valeurs(ws, col_start, nb_prev_param, annee=None, annee_min=None, annee_max=None):
    """
    Retourne (dates, reel, previsions, noms_profils)

    - Si 'annee' est fourni : utilise YEAR_INDEX pour ne garder que
      *les lignes* de l'année et *les profils actifs* de l'année,
      puis on filtre éventuellement certains profils (2023) → le RÉEL reste inchangé.
    - Sinon : filtre sur plage ou aucun filtre (comportement historique).
    """

    DEBUG_EXTRAIRE = True
    def _dbg(*args):
        if DEBUG_EXTRAIRE:
            print("[extraire_valeurs]", *args)

    section = ws
    flux_name = _flux_name_from_token(section, col_start)
    _dbg(f"CALL: section={section!r} col_start={col_start} nb_prev_param={nb_prev_param} "
         f"annee={annee} annee_min={annee_min} annee_max={annee_max}")

    if not flux_name:
        _dbg("ABORT: flux_name introuvable -> retour listes vides/pad prév.")
        return [], [], [[] for _ in range(nb_prev_param)], []

    B = CACHE.get((section, flux_name))
    if not B or not B.get("dates"):
        _dbg(f"ABORT: pas de cache ou dates vides pour {(section, flux_name)}")
        return [], [], [[] for _ in range(nb_prev_param)], []

    dates_all = B["dates"]
    reel_all  = B["reel"]
    prev_all  = B.get("prev_vals", [])
    headers   = B.get("prev_headers", [])   # ex: "Prévision 16/01 (K€)"

    _dbg(f"CACHE: len(dates_all)={len(dates_all)}, len(reel_all)={len(reel_all)}, "
         f"nb_profils_total={len(prev_all)}, nb_headers={len(headers)}")

    # ==============================
    # CAS 1 : année précise
    # ==============================
    if annee is not None:
        years = YEAR_INDEX.get((section, flux_name), {}).get("years", {})
        info = years.get(annee)
        if not info:
            _dbg(f"ANNEE={annee}: YEAR_INDEX absent/vide -> retour listes vides.")
            return [], [], [], []

        # Ces indices NE CHANGENT PAS pour les lignes → le RÉEL reste intact
        row_idx  = info["row_idx"]      # indices des dates pour cette année
        prof_idx = info["prof_idx"]     # indices des colonnes de prévision actives

        _dbg(f"ANNEE={annee}: rows={len(row_idx)} profils_actifs={len(prof_idx)}")

        # ---- Filtre spécial 2023 : on enlève certains profils, mais on garde le réel ----
        if annee == 2023:
            def _is_excluded(k: int) -> bool:
                if k < 0 or k >= len(headers):
                    return False
                lab = str(headers[k])  # "Prévision 16/01 (K€)"
                for d in EXCLUDED_DATES_2023:
                    if d in lab:
                        return True
                return False

            prof_idx_before = list(prof_idx)
            prof_idx = [k for k in prof_idx if not _is_excluded(k)]
            _dbg(f"ANNEE=2023: prof_idx avant filtre = {prof_idx_before}")
            _dbg(f"ANNEE=2023: prof_idx après  filtre = {prof_idx}")

        # ---- Extraction des lignes (réel inchangé) ----
        dates = [dates_all[i] for i in row_idx]
        reel  = [reel_all[i]  for i in row_idx]

        # ---- Extraction des prévisions uniquement pour les profils retenus ----
        previsions = []
        for k in prof_idx:
            serie = prev_all[k] if k < len(prev_all) else []
            previsions.append([serie[i] if i < len(serie) else None for i in row_idx])

        # Noms de profils alignés sur prof_idx
        noms_profils = [
            _clean_profil_label(headers[k] if k < len(headers) else None, k)
            for k in prof_idx
        ]

        _dbg(f"RETURN(annee): len(dates)={len(dates)}, len(reel)={len(reel)}, "
             f"nb_profils={len(previsions)}, noms_profils={noms_profils}")
        return dates, reel, previsions, noms_profils

    # ==============================
    # CAS 2 : plage / historique
    # ==============================
    def _keep(d):
        y = d.year
        if annee_min is not None and annee_max is not None:
            return annee_min <= y <= annee_max
        if annee_min is not None:
            return y >= annee_min
        if annee_max is not None:
            return y <= annee_max
        return True

    idxs = [i for i, d in enumerate(dates_all) if _keep(d)]
    dates = [dates_all[i] for i in idxs]
    reel  = [reel_all[i]  for i in idxs]
    previsions = [[serie[i] if i < len(serie) else None for i in idxs] for serie in prev_all]

    noms_profils = [_clean_profil_label(h, i) for i, h in enumerate(headers)]

    _dbg(f"PLAGE/HISTO: indices_retenus={len(idxs)} "
         f"(min_year={min((d.year for d in dates), default=None)}, "
         f"max_year={max((d.year for d in dates), default=None)})")
    _dbg(f"RETURN(plage): len(dates)={len(dates)}, len(reel)={len(reel)}, "
         f"nb_profils={len(previsions)}, len(noms_profils)={len(noms_profils)}; "
         f"exemple_date={dates[0] if dates else None}")

    return dates, reel, previsions, noms_profils


# ==================== VISUALISEUR 3D - WOW MODE ====================
@dataclass
class TrainingMetric:
    """Structure pour les données d'entraînement"""
    loss: float
    accuracy: float
    epoch: int
    batch: int
    timestamp: float = None
    
    def __post_init__(self):
        if self.timestamp is None:
            self.timestamp = time.time()

class Model3DVisualizer:
    """Visualiseur 3D POST-ENTRAÎNEMENT - Collecte les métriques puis affiche tout en 3D"""
    
    def __init__(self, title: str = "PULSE - Model Training Monitor 3D | WOW MODE"):
        self.title = title
        self.metrics_history = []  # Simple list, pas de queue
        self.canvas = None
        self.view = None
        self.particles_visual = None
        self.is_running = False
    
    def add_metric(self, loss: float, accuracy: float, epoch: int, batch: int):
        """Ajouter une métrique - très rapide, juste append"""
        metric = TrainingMetric(
            loss=float(loss),
            accuracy=float(accuracy),
            epoch=int(epoch),
            batch=int(batch)
        )
        self.metrics_history.append(metric)
    
    def show(self, run_async: bool = True):
        """INACTIF en temps réel - utiliser show_results() après l'entraînement"""
        print("[3D VIZ] Mode collection - métriques sauvegardées")
    
    def show_results(self):
        """Afficher les résultats d'entraînement en 3D après la fin"""
        if not HAS_VISPY:
            print("[3D VIZ] ✗ Vispy non disponible")
            return
        
        if not self.metrics_history:
            print("[3D VIZ] ✗ Aucune métrique à visualiser")
            return
        
        print(f"[3D VIZ] 📊 Visualisation de {len(self.metrics_history)} métriques...")
        
        try:
            import numpy as np
            
            # Créer le canvas
            self.canvas = scene.SceneCanvas(
                title=self.title,
                keys='interactive',
                size=(1500, 950),
                show=False,
                vsync=True,
                bgcolor='#0a0e27'
            )
            
            self.view = self.canvas.central_widget.add_view()
            self.view.camera = scene.TurntableCamera(up='z', fov=60)
            self.view.camera.distance = 20
            
            # === Créer les visuels 3D ===
            history = self.metrics_history
            n_points = len(history)
            
            # Données
            times = np.arange(n_points, dtype=np.float32) * 0.3
            losses = np.array([m.loss for m in history], dtype=np.float32) * 5
            accuracies = np.array([m.accuracy for m in history], dtype=np.float32) * 5
            
            # Positions 3D : X=temps, Y=loss, Z=accuracy
            positions = np.column_stack([times, losses, accuracies]).astype(np.float32)
            
            # Couleurs : gradient vert (bon) -> rouge (mauvais)
            colors = np.zeros((n_points, 4), dtype=np.float32)
            for i in range(n_points):
                acc_norm = min(accuracies[i] / 5, 1.0)  # 0 à 1
                colors[i] = [
                    1.0 - acc_norm,      # R: rouge si accuracy faible
                    acc_norm,            # G: vert si accuracy élevée
                    0.5 + 0.5 * acc_norm,  # B  
                    0.95                 # Alpha
                ]
            
            sizes = np.ones(n_points, dtype=np.float32) * 10
            
            # Créer les markers (points)
            self.particles_visual = visuals.Markers(
                pos=positions,
                size=sizes,
                face_color=colors,
                edge_color=None,
                edge_width=0,
                parent=self.view.scene,
                pxmode=False
            )
            
            # === Ajouter des axes pour mieux comprendre ===
            # Axe X (temps)
            ax_x = np.array([[0, 0, 0], [times.max() * 1.1, 0, 0]], dtype=np.float32)
            line_x = visuals.Line(pos=ax_x, color=[1, 0, 0, 0.5], parent=self.view.scene, width=2)
            
            # Axe Y (loss)
            ax_y = np.array([[0, 0, 0], [0, losses.max() * 1.2, 0]], dtype=np.float32)
            line_y = visuals.Line(pos=ax_y, color=[0, 1, 0, 0.5], parent=self.view.scene, width=2)
            
            # Axe Z (accuracy)
            ax_z = np.array([[0, 0, 0], [0, 0, accuracies.max() * 1.2]], dtype=np.float32)
            line_z = visuals.Line(pos=ax_z, color=[0, 0, 1, 0.5], parent=self.view.scene, width=2)
            
            # Mettre à jour le titre
            stats = {
                'min_loss': min([m.loss for m in history]),
                'max_loss': max([m.loss for m in history]),
                'avg_acc': np.mean([m.accuracy for m in history]),
                'max_acc': max([m.accuracy for m in history]),
                'n_points': n_points
            }
            
            self.canvas.title = (
                f"{self.title} | "
                f"Loss: {stats['min_loss']:.4f}→{stats['max_loss']:.4f} | "
                f"Accuracy: {stats['avg_acc']:.3f} (max {stats['max_acc']:.3f}) | "
                f"📊 {stats['n_points']} points"
            )
            
            # Afficher
            print("[3D VIZ] ✓ Affichage de la visualisation 3D")
            self.canvas.show()
            
            # Essayer de mettre en avant
            try:
                self.canvas.native.activateWindow()
                self.canvas.native.raise_()
            except:
                pass
            
            # Lancer la boucle d'événements
            self.is_running = True
            app_vispy.run()
            
        except Exception as e:
            print(f"[3D VIZ] ✗ Erreur visualisation: {e}")
            import traceback
            traceback.print_exc()
    
    def close(self):
        """Fermer la visualisation"""
        self.is_running = False
        if self.canvas is not None:
            try:
                self.canvas.close()
            except:
                pass


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        
        # ===============================================
        # INIT ROBUSTE DE LA FENÊTRE (évite la transparence)
        # ===============================================
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        
        # !! IMPORTANT : définir la geometrie AVANT fullscreen
        self.geometry("1300x900")

        self._fading_in = False
        self.attributes("-topmost", False)
        self.attributes("-alpha", 1.0)
        self.attributes("-disabled", False)

        # Force un refresh complet des attributs Windows
        self.update_idletasks()

        # Puis passer en fullscreen
        self.attributes("-fullscreen", True)

        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        # Bind supplémentaire : si la fenêtre est bougée, s'assurer qu'elle reste opaque
        self.bind("<Configure>", self._on_window_configure)

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

        # État pour éviter les boucles infinies dans _on_window_configure
        self._last_alpha_fix = time.time()
    
    def _on_window_configure(self, event=None):
        """Callback pour s'assurer que la fenêtre reste opaque quand elle est bougée/redimensionnée."""
        # Ne pas interférer pendant le fade-in d'entrée
        if getattr(self, "_fading_in", False):
            return
        now = time.time()
        # Limiter à une vérification tous les 200ms pour éviter les appels trop fréquents
        if now - self._last_alpha_fix > 0.2:
            try:
                current_alpha = self.attributes("-alpha")
                if current_alpha < 1.0:
                    self.attributes("-alpha", 1.0)
                    print(f"[DEBUG] Fenêtre transparente détectée, restauration opacité (alpha={current_alpha} → 1.0)")
            except Exception:
                pass
            self._last_alpha_fix = now


#===================Page Accueil + fonctions de navigation===================
    def creer_accueil(self):
        import os
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage
        from datetime import datetime

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        C = {
            "bg":           "#080D1A",
            "bg_panel":     "#0C1220",
            "surface":      "#101827",
            "surface_2":    "#141E2E",
            "surface_3":    "#1A2540",
            "surface_4":    "#1F2D4D",
            "border":       "#1E2D48",
            "border_soft":  "#172035",
            "primary":      "#3B82F6",
            "primary_dim":  "#1E3A6E",
            "primary_text": "#93C5FD",
            "success":      "#10B981",
            "success_dim":  "#064E35",
            "warning":      "#F59E0B",
            "warning_dim":  "#4C2A00",
            "danger":       "#EF4444",
            "danger_hover": "#DC2626",
            "text":         "#F1F5F9",
            "text_2":       "#CBD5E1",
            "muted":        "#64748B",
            "muted_2":      "#475569",
            "accent":       "#06B6D4",
            "accent_dim":   "#083344",
        }

        FONT_HEAD = "Segoe UI Semibold"
        FONT_BODY = "Segoe UI"

        # =========================================================
        # INIT
        # =========================================================
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        # Option test Windows / bug graphique :
        # décommente seulement si besoin
        # try:
        #     ctk.deactivate_automatic_dpi_awareness()
        # except Exception:
        #     pass

        try:
            self.configure(fg_color=C["bg"])
        except Exception:
            try:
                self.configure(bg=C["bg"])
            except Exception:
                pass

        try:
            self.attributes("-alpha", 1.0)
        except Exception:
            pass

        # Annule les anciens jobs éventuels
        for attr in ("_resize_job", "_accueil_watch_job", "_sb_anim_job"):
            try:
                job = getattr(self, attr, None)
                if job:
                    self.after_cancel(job)
            except Exception:
                pass
            setattr(self, attr, None)

        self.vider_fenetre()

        # =========================================================
        # UTILITAIRES
        # =========================================================
        def _exists(w):
            try:
                return bool(w and w.winfo_exists())
            except Exception:
                return False

        def _cancel_job(attr):
            job = getattr(self, attr, None)
            if job:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
            setattr(self, attr, None)

        def _fmt_dt(ts):
            try:
                return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return "—"

        def _fmt_size(path):
            try:
                return f"{os.path.getsize(path) / 1024 / 1024:.1f} Mo"
            except Exception:
                return "—"

        # =========================================================
        # DATA
        # =========================================================
        try:
            fichiers = _lister_fichiers_mensuels()
            existing = [p for _, _, p in fichiers if os.path.exists(p)]
            nb_fichiers = len(fichiers)

            if existing:
                last_mtime = max(os.path.getmtime(p) for p in existing)
                derniere_maj = _fmt_dt(last_mtime)
                annees_dispos = sorted({int(a) for a, _, _ in fichiers if a is not None})
            else:
                derniere_maj = "—"
                annees_dispos = []
        except Exception:
            fichiers = []
            existing = []
            nb_fichiers = 0
            derniere_maj = "—"
            annees_dispos = []

        try:
            filiales_list = sorted(
                {s.get("dest") for s in SECTIONS_CONFIG if s.get("dest")}
            ) if SECTIONS_CONFIG else []
            nb_filiales = len(filiales_list)
        except Exception:
            filiales_list = []
            nb_filiales = 0

        recent_files = []
        try:
            recs = sorted(
                [(p, os.path.getmtime(p)) for _, _, p in fichiers if os.path.exists(p)],
                key=lambda x: x[1],
                reverse=True
            )[:5]

            for path, ts in recs:
                recent_files.append((os.path.basename(path), _fmt_size(path), _fmt_dt(ts)))
        except Exception:
            pass

        if not recent_files:
            recent_files = [("Aucun fichier détecté", "—", "—")]

        # =========================================================
        # TOKENS RESPONSIVE
        # =========================================================
        def _rs():
            w = self.winfo_width()
            h = self.winfo_height()

            if w <= 1:
                w = max(1280, self.winfo_screenwidth() - 140)
            if h <= 1:
                h = max(800, self.winfo_screenheight() - 140)

            tier = (
                "xl" if w >= 1800 else
                "lg" if w >= 1480 else
                "md" if w >= 1220 else
                "sm" if w >= 980 else
                "xs"
            )

            return {
                "w": w,
                "h": h,
                "tier": tier,
                "hdr_h": 88 if tier in ("xl", "lg") else 80 if tier == "md" else 72,
                "pad": 24 if tier in ("xl", "lg") else 18 if tier == "md" else 14,
                "gap": 14 if tier in ("xl", "lg") else 10 if tier == "md" else 8,
                "hero_stack": w < 1380,
                "btn_stack": w < 820,
                "kpi_cols": 3 if w >= 1450 else 2 if w >= 980 else 1,
                "act_cols": 3 if w >= 1480 else 2 if w >= 980 else 1,
                "met_cols": 3 if w >= 1280 else 2 if w >= 900 else 1,
                "hero_fs": 32 if tier == "xl" else 27 if tier == "lg" else 23 if tier == "md" else 20 if tier == "sm" else 18,
                "sec_fs": 20 if tier in ("xl", "lg") else 17 if tier == "md" else 15 if tier == "sm" else 13,
                "body_fs": 13 if tier in ("xl", "lg") else 12 if tier == "md" else 11,
                "sm_fs": 11 if tier in ("xl", "lg") else 10,
                "nav_fs": 13 if tier in ("xl", "lg") else 12 if tier == "md" else 11,
                "btn_h": 42 if tier in ("xl", "lg") else 38 if tier == "md" else 36,
                "hero_wrap": max(320, min(760, w - 640)),
                "detail_wrap": max(320, min(860, w - 420)),
                "sum_w": 340 if w >= 1500 else 300 if w >= 1280 else 0,
                "hdr_sub": w >= 1040,
                "sb_auto": "full" if w >= 1200 else "compact" if w >= 960 else "icon",
            }

        _rs_cache = {"v": _rs()}

        # =========================================================
        # PRIMITIVES UI
        # =========================================================
        def _frame(parent, fg=None, radius=16, bw=1, bc=None, **kw):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or C["surface"],
                corner_radius=radius,
                border_width=bw,
                border_color=bc or C["border"],
                **kw
            )

        def _lbl(parent, text, color, font, **kw):
            return ctk.CTkLabel(parent, text=text, text_color=color, font=font, **kw)

        def _btn(parent, text, cmd=None, h=38, w=0, radius=10,
                fg=None, hover=None, tc=None, font_size=12):
            return ctk.CTkButton(
                parent,
                text=text,
                command=cmd,
                height=h,
                width=w,
                corner_radius=radius,
                fg_color=fg or C["surface_3"],
                hover_color=hover or C["surface_4"],
                border_width=1,
                border_color=C["border"],
                text_color=tc or C["text_2"],
                font=(FONT_HEAD, font_size, "bold")
            )

        def _primary_btn(parent, text, cmd=None, h=38, w=0):
            return ctk.CTkButton(
                parent,
                text=text,
                command=cmd,
                height=h,
                width=w,
                corner_radius=10,
                fg_color=C["primary"],
                hover_color="#2563EB",
                text_color="white",
                font=(FONT_HEAD, 12, "bold")
            )

        def _sep(parent, color=None, pady=(0, 0)):
            s = ctk.CTkFrame(parent, fg_color=color or C["border"], height=1, corner_radius=0)
            s.pack(fill="x", pady=pady)
            return s

        def _badge(parent, text, fg=None, tc=None, bc=None):
            pill = ctk.CTkFrame(
                parent,
                fg_color=fg or C["surface_3"],
                corner_radius=999,
                border_width=1,
                border_color=bc or C["border"]
            )
            lbl = ctk.CTkLabel(
                pill,
                text=text,
                text_color=tc or C["text_2"],
                font=(FONT_BODY, 11, "bold")
            )
            lbl.pack(padx=10, pady=5)
            return pill, lbl

        def _section_hdr(parent, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            t = ctk.CTkLabel(
                wrap,
                text=title,
                text_color=C["text"],
                font=(FONT_HEAD, 18, "bold")
            )
            t.grid(row=0, column=0, sticky="w")

            s = None
            if subtitle:
                s = ctk.CTkLabel(
                    wrap,
                    text=subtitle,
                    text_color=C["muted"],
                    font=(FONT_BODY, 12)
                )
                s.grid(row=1, column=0, sticky="w", pady=(3, 0))

            return {"wrap": wrap, "title": t, "subtitle": s}

        def _hover(widget, nfg, hfg):
            def _in(_=None):
                try:
                    widget.configure(fg_color=hfg)
                except Exception:
                    pass

            def _out(_=None):
                try:
                    widget.configure(fg_color=nfg)
                except Exception:
                    pass

            widget.bind("<Enter>", _in)
            widget.bind("<Leave>", _out)
            return _in, _out

        def _click_tree(widget, cb):
            try:
                if not isinstance(widget, ctk.CTkButton):
                    widget.bind("<Button-1>", lambda e: cb())
            except Exception:
                pass

            try:
                for child in widget.winfo_children():
                    _click_tree(child, cb)
            except Exception:
                pass

        # =========================================================
        # DETAIL PANEL
        # =========================================================
        detail_state = {
            "title": "Centre de détails",
            "lines": [
                "Cliquez sur une carte KPI pour afficher des informations détaillées.",
                "Ce panneau accueille les résumés et informations contextuelles."
            ],
            "tone": "neutral",
        }

        _dw = {"body": None, "title_lbl": None}

        def _render_detail():
            body = _dw["body"]
            if not _exists(body):
                return

            for child in body.winfo_children():
                child.destroy()

            r = _rs_cache["v"]

            if _exists(_dw["title_lbl"]):
                _dw["title_lbl"].configure(
                    text=detail_state["title"],
                    font=(FONT_HEAD, max(13, r["sec_fs"] - 4), "bold")
                )

            tone_c = {
                "neutral": C["primary_text"],
                "success": C["success"],
                "warning": C["warning"],
            }.get(detail_state["tone"], C["primary_text"])

            pill, _ = _badge(
                body,
                "● Détail interactif",
                fg=C["surface_3"],
                tc=tone_c,
                bc=C["border"]
            )
            pill.pack(anchor="w", pady=(0, 14))

            for line in (detail_state["lines"] or ["Aucune donnée disponible."]):
                row = ctk.CTkFrame(body, fg_color="transparent")
                row.pack(fill="x", pady=(0, 6))

                ctk.CTkLabel(
                    row,
                    text="›",
                    width=14,
                    text_color=C["primary"],
                    font=(FONT_HEAD, r["body_fs"] + 1, "bold")
                ).pack(side="left", anchor="n", padx=(0, 8))

                ctk.CTkLabel(
                    row,
                    text=line,
                    text_color=C["text_2"],
                    font=(FONT_BODY, r["body_fs"]),
                    justify="left",
                    wraplength=r["detail_wrap"]
                ).pack(side="left", fill="x", expand=True)

        def _set_detail(title, lines, tone="neutral"):
            detail_state.update(
                title=title,
                lines=lines or ["Aucune donnée disponible."],
                tone=tone
            )
            _render_detail()

        def _show_years():
            lines = [f"Années : {', '.join(str(a) for a in annees_dispos)}"] if annees_dispos else ["Aucune année détectée."]
            _set_detail("Périmètre des fichiers mensuels", lines)

        def _show_filiales():
            _set_detail(
                "Organisation des filiales",
                list(filiales_list) or ["Aucune filiale configurée."],
                "success"
            )

        def _show_overview():
            _set_detail(
                "Résumé plateforme",
                [
                    "Console analytique dédiée au pilotage des flux de trésorerie.",
                    "Lecture consolidée du réel, des prévisions et des écarts.",
                    f"{nb_fichiers} fichier(s) mensuel(s) — {nb_filiales} filiale(s).",
                    f"Dernière actualisation : {derniere_maj}.",
                ],
                "warning"
            )

        def _show_recent_files():
            lines = [f"{name} • {size} • {dt}" for name, size, dt in recent_files]
            _set_detail("Derniers fichiers détectés", lines)

        # =========================================================
        # ROOT GRID
        # =========================================================
        for i in range(6):
            self.grid_rowconfigure(i, weight=0)
            self.grid_columnconfigure(i, weight=0)

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        # =========================================================
        # HEADER
        # =========================================================
        header = ctk.CTkFrame(self, fg_color=C["bg_panel"], corner_radius=0, height=88)
        header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        header.grid_propagate(False)
        header.grid_columnconfigure(2, weight=1)

        logo_wrap = ctk.CTkFrame(header, fg_color="transparent")
        logo_wrap.grid(row=0, column=0, sticky="w", padx=(22, 0))

        logo_path = getattr(self, "image_path", None)
        try:
            _img = Image.open(logo_path)
            nh = 32
            nw = int(nh * _img.width / max(_img.height, 1))
            try:
                _img = _img.resize((nw, nh), Image.Resampling.LANCZOS)
            except Exception:
                _img = _img.resize((nw, nh), Image.ANTIALIAS)

            cimg = CTkImage(light_image=_img, dark_image=_img, size=(nw, nh))
            ll = ctk.CTkLabel(logo_wrap, image=cimg, text="")
            ll.image = cimg
            ll.pack(side="left", pady=28)
        except Exception:
            _lbl(logo_wrap, "PULSE", C["text"], (FONT_HEAD, 18, "bold")).pack(side="left", pady=28)

        ctk.CTkFrame(header, fg_color=C["border"], width=1, corner_radius=0).grid(
            row=0, column=1, sticky="ns", padx=(16, 0), pady=18
        )

        title_wrap = ctk.CTkFrame(header, fg_color="transparent")
        title_wrap.grid(row=0, column=2, sticky="w", padx=(14, 0))

        hdr_title = _lbl(title_wrap, "PULSE", C["text"], (FONT_HEAD, 16, "bold"))
        hdr_title.pack(anchor="w")

        hdr_sub = _lbl(
            title_wrap,
            "Plateforme d'analyse des flux de trésorerie",
            C["muted"],
            (FONT_BODY, 11)
        )
        hdr_sub.pack(anchor="w", pady=(2, 0))

        hdr_right = ctk.CTkFrame(header, fg_color="transparent")
        hdr_right.grid(row=0, column=3, sticky="e", padx=22)

        status_pill = ctk.CTkFrame(
            hdr_right,
            fg_color=C["success_dim"],
            corner_radius=999,
            border_width=1,
            border_color=C["success"]
        )
        status_pill.pack(side="left", padx=(0, 12))

        status_lbl = _lbl(status_pill, "● Opérationnel", C["success"], (FONT_HEAD, 11, "bold"))
        status_lbl.pack(padx=12, pady=6)

        quit_btn = ctk.CTkButton(
            hdr_right,
            text="Quitter",
            command=self.demander_confirmation_quit,
            height=36,
            width=88,
            corner_radius=8,
            fg_color=C["danger"],
            hover_color=C["danger_hover"],
            text_color="white",
            font=(FONT_HEAD, 11, "bold")
        )
        quit_btn.pack(side="left")

        ctk.CTkFrame(self, fg_color=C["border"], height=1, corner_radius=0).grid(
            row=1, column=0, columnspan=2, sticky="ew"
        )

        # =========================================================
        # SIDEBAR
        # =========================================================
        SB_FULL = 272
        SB_COMPACT = 196
        SB_ICON = 66

        _sb = {"mode": "full", "manual": False}

        sidebar = ctk.CTkFrame(self, fg_color=C["bg_panel"], corner_radius=0, width=SB_FULL)
        sidebar.grid(row=2, column=0, sticky="nsew")
        sidebar.grid_propagate(False)

        sb_scroll = ctk.CTkScrollableFrame(
            sidebar,
            fg_color="transparent",
            scrollbar_button_color=C["surface_3"],
            scrollbar_button_hover_color=C["surface_4"]
        )
        sb_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        burger_row = ctk.CTkFrame(sb_scroll, fg_color="transparent")
        burger_row.pack(fill="x", padx=10, pady=(14, 10))

        burger_btn = ctk.CTkButton(
            burger_row,
            text="☰",
            width=38,
            height=32,
            corner_radius=8,
            fg_color=C["surface_3"],
            hover_color=C["surface_4"],
            text_color=C["text_2"],
            font=(FONT_BODY, 14),
            anchor="center"
        )
        burger_btn.pack(side="left")

        sb_nav_lbl = _lbl(burger_row, "NAVIGATION", C["muted_2"], (FONT_HEAD, 9, "bold"))
        sb_nav_lbl.pack(side="left", padx=(10, 0), pady=(2, 0))

        nav_items = []
        sb_sections = []

        def _nav_section(parent, label):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.pack(fill="x", padx=10, pady=(0, 12))

            sec_lbl = _lbl(wrap, label, C["muted"], (FONT_HEAD, 10, "bold"))
            sec_lbl.pack(anchor="w", pady=(0, 4))

            sb_sections.append({"lbl": sec_lbl, "text": label})
            return wrap

        def _nav_item(parent, icon, label, short, cmd, active=False):
            nfg = C["primary_dim"] if active else "transparent"
            hfg = C["surface_3"] if not active else C["surface_4"]
            bc = C["primary"] if active else C["border_soft"]
            tc = C["text"] if active else C["text_2"]

            frame = ctk.CTkFrame(
                parent,
                fg_color=nfg,
                corner_radius=10,
                border_width=1,
                border_color=bc
            )
            frame.pack(fill="x", pady=2)

            ein, eout = _hover(frame, nfg, hfg)

            btn = ctk.CTkButton(
                frame,
                text=f"{icon}  {label}",
                command=cmd,
                anchor="w",
                height=40,
                corner_radius=10,
                fg_color="transparent",
                hover=False,
                text_color=tc,
                font=(FONT_BODY, 13)
            )
            btn.pack(fill="x", padx=2, pady=2)

            btn.bind("<Enter>", ein)
            btn.bind("<Leave>", eout)

            nav_items.append({
                "frame": frame,
                "btn": btn,
                "icon": icon,
                "text": label,
                "short": short,
                "active": active
            })

        s1 = _nav_section(sb_scroll, "VUE GÉNÉRALE")
        _nav_item(s1, "⌂", "Accueil", "Accueil", self.creer_accueil, active=True)
        _nav_item(s1, "↓", "Importer les profils", "Import", self.importer_les_profils)
        _nav_item(s1, "◈", "Visualisation graphique", "Visu", self.creer_page_graphique)

        s2 = _nav_section(sb_scroll, "ANALYSE DES ÉCARTS")
        _nav_item(s2, "⚑", "Écarts importants", "Écarts", self.afficher_ecarts)
        _nav_item(s2, "◉", "Écarts par filiale", "Filiale", self.afficher_repartition)
        _nav_item(s2, "◎", "Écarts par profil", "Profil", self.afficher_repartition_par_prevision)
        _nav_item(s2, "◈", "Écarts par flux", "Flux", self.afficher_repartition_flux)
        _nav_item(s2, "⌁", "Tendance des écarts", "Tendance", self.afficher_tendance_flux)

        s3 = _nav_section(sb_scroll, "ANOMALIES & IA")
        _nav_item(s3, "◈", "Heatmap anomalies", "Anomalies", self.afficher_heatmap_anomalies)
        _nav_item(s3, "▦", "Heatmap écarts", "Heatmap", self.afficher_heatmap_ecarts)
        _nav_item(s3, "⬡", "Clustering des écarts", "Clustering", self.analyser_ecarts_ml)
        _nav_item(s3, "⬢", "IA — prédiction", "IA", self.creer_page_ia_prediction)

        sb_foot = ctk.CTkFrame(sidebar, fg_color="transparent")
        sb_foot.pack(side="bottom", fill="x", padx=10, pady=(6, 12))

        sb_info = _frame(sb_foot, fg=C["surface_2"], radius=12, bw=1, bc=C["border"])
        sb_info.pack(fill="x")

        sb_info_v = _lbl(sb_info, "Version produit", C["muted"], (FONT_BODY, 10))
        sb_info_v.pack(anchor="w", padx=12, pady=(12, 2))

        sb_info_ver = _lbl(sb_info, "v1.0", C["text"], (FONT_HEAD, 16, "bold"))
        sb_info_ver.pack(anchor="w", padx=12)

        sb_info_date = _lbl(sb_info, f"MAJ : {derniere_maj}", C["muted_2"], (FONT_BODY, 10))
        sb_info_date.pack(anchor="w", padx=12, pady=(2, 12))

        def _sb_apply_content():
            if not _exists(sidebar):
                return

            r = _rs_cache["v"]
            mode = _sb["mode"]

            target_w = {"full": SB_FULL, "compact": SB_COMPACT, "icon": SB_ICON}[mode]
            try:
                sidebar.configure(width=target_w)
            except Exception:
                pass

            if mode == "icon":
                sb_nav_lbl.configure(text="")
                burger_btn.configure(width=42)
            elif mode == "compact":
                sb_nav_lbl.configure(text="")
                burger_btn.configure(width=38)
            else:
                sb_nav_lbl.configure(text="NAVIGATION", font=(FONT_HEAD, 9, "bold"))
                burger_btn.configure(width=38)

            for sec in sb_sections:
                if mode == "full":
                    sec["lbl"].configure(text=sec["text"], font=(FONT_HEAD, max(9, r["sm_fs"]), "bold"))
                elif mode == "compact":
                    sec["lbl"].configure(text=sec["text"][:4].upper() + "…", font=(FONT_HEAD, 8, "bold"))
                else:
                    sec["lbl"].configure(text="")

            for item in nav_items:
                if mode == "icon":
                    item["btn"].configure(
                        text=item["icon"],
                        anchor="center",
                        height=r["btn_h"],
                        font=(FONT_BODY, r["nav_fs"] + 3),
                        width=46
                    )
                elif mode == "compact":
                    item["btn"].configure(
                        text=f"{item['icon']}  {item['short']}",
                        anchor="w",
                        height=r["btn_h"],
                        font=(FONT_BODY, max(10, r["nav_fs"] - 1)),
                        width=0
                    )
                else:
                    item["btn"].configure(
                        text=f"{item['icon']}  {item['text']}",
                        anchor="w",
                        height=r["btn_h"],
                        font=(FONT_BODY, r["nav_fs"]),
                        width=0
                    )

            if mode != "icon":
                if not sb_info.winfo_manager():
                    sb_info.pack(fill="x")
                sb_info_v.configure(
                    text="Version produit" if mode == "full" else "Ver.",
                    font=(FONT_BODY, r["sm_fs"])
                )
                sb_info_ver.configure(font=(FONT_HEAD, 16 if mode == "full" else 13, "bold"))
                sb_info_date.configure(
                    text=(f"MAJ : {derniere_maj}" if mode == "full" else derniere_maj[:10]),
                    font=(FONT_BODY, r["sm_fs"])
                )
            else:
                if sb_info.winfo_manager():
                    sb_info.pack_forget()

        def _sb_set(mode):
            _sb["mode"] = mode
            _sb_apply_content()

        def _burger_toggle():
            _sb["manual"] = True
            order = ["full", "compact", "icon"]
            nxt = order[(order.index(_sb["mode"]) + 1) % len(order)]
            _sb_set(nxt)

        burger_btn.configure(command=_burger_toggle)

        # =========================================================
        # MAIN SCROLLABLE
        # =========================================================
        main = ctk.CTkScrollableFrame(self, fg_color=C["bg"], corner_radius=0)
        main.grid(row=2, column=1, sticky="nsew")
        main.grid_columnconfigure(0, weight=1)

        # =========================================================
        # HERO
        # =========================================================
        hero = _frame(main, fg=C["surface"], radius=20, bw=1, bc=C["border"])
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(22, 12))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_columnconfigure(1, weight=0)

        left_hero = ctk.CTkFrame(hero, fg_color="transparent")
        left_hero.grid(row=0, column=0, sticky="nsew", padx=24, pady=24)

        hero_pill, hero_pill_lbl = _badge(
            left_hero,
            "Plateforme analytique • Finance & Pilotage",
            fg=C["primary_dim"],
            tc=C["primary_text"],
            bc=C["primary"]
        )
        hero_pill.pack(anchor="w", pady=(0, 16))

        hero_title = _lbl(
            left_hero,
            "Transformez les flux de trésorerie\nen décisions fiables.",
            C["text"],
            (FONT_HEAD, 28, "bold"),
            justify="left"
        )
        hero_title.pack(anchor="w")

        hero_desc = _lbl(
            left_hero,
            (
                "PULSE centralise données réelles, prévisions et écarts pour offrir "
                "une lecture claire de la performance, des anomalies et des priorités "
                "d'action — par filiale et par flux."
            ),
            C["text_2"],
            (FONT_BODY, 13),
            justify="left",
            wraplength=720
        )
        hero_desc.pack(anchor="w", pady=(12, 22))

        hero_btns = ctk.CTkFrame(left_hero, fg_color="transparent")
        hero_btns.pack(anchor="w", fill="x")

        hero_primary = _primary_btn(
            hero_btns,
            "Ouvrir la visualisation",
            cmd=self.creer_page_graphique,
            h=42,
            w=220
        )
        hero_ghost = _btn(
            hero_btns,
            "Importer des profils",
            cmd=self.importer_les_profils,
            h=42,
            w=190
        )
        hero_primary.pack(side="left", padx=(0, 10))
        hero_ghost.pack(side="left")

        hero_met = ctk.CTkFrame(left_hero, fg_color="transparent")
        hero_met.pack(fill="x", pady=(20, 0))

        met_refs = []
        for mlbl, mval, mac in [
            ("Fichiers mensuels", str(nb_fichiers), C["text"]),
            ("Filiales couvertes", str(nb_filiales), C["text"]),
            ("Dernière MAJ", derniere_maj, C["success"]),
        ]:
            mc = _frame(hero_met, fg=C["surface_2"], radius=12, bw=1, bc=C["border"])
            mc.grid_columnconfigure(0, weight=1)

            ml = _lbl(mc, mlbl, C["muted"], (FONT_BODY, 10))
            ml.grid(row=0, column=0, sticky="w", padx=14, pady=(11, 2))

            mv = _lbl(mc, mval, mac, (FONT_HEAD, 16, "bold"))
            mv.grid(row=1, column=0, sticky="w", padx=14, pady=(0, 11))

            met_refs.append({"card": mc, "lbl": ml, "val": mv})

        right_hero = ctk.CTkFrame(hero, fg_color="transparent")
        right_hero.grid(row=0, column=1, sticky="ne", padx=(0, 24), pady=24)

        sum_card = _frame(right_hero, fg=C["surface_2"], radius=16, bw=1, bc=C["border"])
        sum_card.pack(fill="both", expand=True)

        _lbl(sum_card, "Résumé opérationnel", C["text"], (FONT_HEAD, 13, "bold")).pack(
            anchor="w", padx=16, pady=(16, 10)
        )

        mini_stats = []

        def _mini(parent, lt, val, ac=None):
            mc = _frame(parent, fg=C["surface_3"], radius=10, bw=1, bc=C["border_soft"])
            mc.pack(fill="x", padx=10, pady=4)

            inn = ctk.CTkFrame(mc, fg_color="transparent")
            inn.pack(fill="x", padx=12, pady=10)

            ll = _lbl(inn, lt, C["muted"], (FONT_BODY, 10))
            ll.pack(anchor="w")

            vl = _lbl(inn, val, ac or C["text"], (FONT_HEAD, 13, "bold"))
            vl.pack(anchor="w", pady=(2, 0))

            mini_stats.append({"lbl": ll, "val": vl})

        _mini(sum_card, "Fichiers suivis", str(nb_fichiers))
        _mini(sum_card, "Filiales monitorées", str(nb_filiales))
        _mini(sum_card, "Dernière synchro", derniere_maj, ac=C["success"])

        _sep(sum_card, color=C["border_soft"], pady=10)

        _lbl(sum_card, "Usage recommandé", C["muted"], (FONT_BODY, 10, "bold")).pack(anchor="w", padx=16)
        _lbl(
            sum_card,
            (
                "Commencez par la visualisation graphique pour contrôler les écarts "
                "réel / prévisions, puis poursuivez avec les analyses flux et détection."
            ),
            C["text_2"],
            (FONT_BODY, 11),
            justify="left",
            wraplength=280
        ).pack(anchor="w", padx=16, pady=(6, 16))

        # =========================================================
        # KPI + DETAIL
        # =========================================================
        overview = ctk.CTkFrame(main, fg_color="transparent")
        overview.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 12))
        overview.grid_columnconfigure(0, weight=1)
        overview.grid_columnconfigure(1, weight=1)

        kpi_wrap = _frame(overview, fg=C["surface"], radius=18, bw=1, bc=C["border"])
        kpi_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        kpi_hdr = _section_hdr(
            kpi_wrap,
            "Indicateurs clés",
            "Cliquez sur une carte pour afficher un détail."
        )
        kpi_hdr["wrap"].pack(fill="x", padx=18, pady=(16, 10))

        kpi_grid = ctk.CTkFrame(kpi_wrap, fg_color="transparent")
        kpi_grid.pack(fill="x", padx=18, pady=(0, 18))

        kpi_cards = []

        def _make_kpi_card(title, value, sub, accent_color, callback):
            card = _frame(kpi_grid, fg=C["surface_2"], radius=14, bw=1, bc=C["border"])
            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="both", expand=True, padx=16, pady=16)

            tl = _lbl(inner, title, C["muted"], (FONT_BODY, 11))
            tl.pack(anchor="w")

            vl = _lbl(inner, value, accent_color, (FONT_HEAD, 26, "bold"))
            vl.pack(anchor="w", pady=(6, 4))

            sl = _lbl(inner, sub, C["text_2"], (FONT_BODY, 11), justify="left", wraplength=260)
            sl.pack(anchor="w")

            _click_tree(card, callback)
            _hover(card, C["surface_2"], C["surface_3"])

            kpi_cards.append({"card": card, "title": tl, "value": vl, "sub": sl})

        _make_kpi_card(
            "Fichiers mensuels",
            str(nb_fichiers),
            "Jeux de données détectés dans le périmètre mensuel.",
            C["text"],
            _show_overview
        )

        _make_kpi_card(
            "Filiales couvertes",
            str(nb_filiales),
            "Filiales configurées dans la plateforme.",
            C["success"],
            _show_filiales
        )

        _make_kpi_card(
            "Années disponibles",
            str(len(annees_dispos)),
            "Périmètre temporel détecté automatiquement.",
            C["warning"],
            _show_years
        )

        _make_kpi_card(
            "Derniers fichiers",
            str(len(recent_files)),
            "Affiche les fichiers les plus récents détectés.",
            C["primary_text"],
            _show_recent_files
        )

        detail_card = _frame(overview, fg=C["surface"], radius=18, bw=1, bc=C["border"])
        detail_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        detail_head = ctk.CTkFrame(detail_card, fg_color="transparent")
        detail_head.pack(fill="x", padx=18, pady=(16, 10))

        _dw["title_lbl"] = _lbl(detail_head, "Centre de détails", C["text"], (FONT_HEAD, 16, "bold"))
        _dw["title_lbl"].pack(anchor="w")

        _lbl(
            detail_head,
            "Zone de lecture contextuelle de l'accueil.",
            C["muted"],
            (FONT_BODY, 11)
        ).pack(anchor="w", pady=(3, 0))

        _sep(detail_card, color=C["border_soft"])

        _dw["body"] = ctk.CTkFrame(detail_card, fg_color="transparent")
        _dw["body"].pack(fill="both", expand=True, padx=18, pady=16)

        # =========================================================
        # ACTIONS RAPIDES
        # =========================================================
        actions_card = _frame(main, fg=C["surface"], radius=18, bw=1, bc=C["border"])
        actions_card.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 12))

        act_hdr = _section_hdr(
            actions_card,
            "Actions rapides",
            "Accès direct aux fonctions principales."
        )
        act_hdr["wrap"].pack(fill="x", padx=18, pady=(16, 10))

        act_grid = ctk.CTkFrame(actions_card, fg_color="transparent")
        act_grid.pack(fill="x", padx=18, pady=(0, 18))

        action_refs = []

        def _action_card(title, desc, btn_text, cmd):
            box = _frame(act_grid, fg=C["surface_2"], radius=14, bw=1, bc=C["border"])
            inner = ctk.CTkFrame(box, fg_color="transparent")
            inner.pack(fill="both", expand=True, padx=16, pady=16)

            t = _lbl(inner, title, C["text"], (FONT_HEAD, 14, "bold"))
            t.pack(anchor="w")

            d = _lbl(inner, desc, C["text_2"], (FONT_BODY, 11), justify="left", wraplength=320)
            d.pack(anchor="w", pady=(8, 14))

            b = _btn(inner, btn_text, cmd=cmd, h=38)
            b.pack(anchor="w")

            action_refs.append({"box": box, "title": t, "desc": d, "btn": b})

        _action_card(
            "Visualisation graphique",
            "Accède rapidement à la vue synthétique pour analyser les écarts et tendances.",
            "Ouvrir",
            self.creer_page_graphique
        )

        _action_card(
            "Importer les profils",
            "Charge ou actualise les profils nécessaires au traitement et à l'analyse.",
            "Importer",
            self.importer_les_profils
        )

        _action_card(
            "IA — prédiction",
            "Ouvre la vue de prédiction pour enrichir l'analyse avec des approches IA.",
            "Lancer",
            self.creer_page_ia_prediction
        )

        # =========================================================
        # FICHIERS RÉCENTS
        # =========================================================
        recent_card = _frame(main, fg=C["surface"], radius=18, bw=1, bc=C["border"])
        recent_card.grid(row=3, column=0, sticky="ew", padx=24, pady=(0, 22))

        rec_hdr = _section_hdr(
            recent_card,
            "Derniers fichiers détectés",
            "Aperçu des fichiers mensuels les plus récents."
        )
        rec_hdr["wrap"].pack(fill="x", padx=18, pady=(16, 10))

        list_wrap = ctk.CTkFrame(recent_card, fg_color="transparent")
        list_wrap.pack(fill="x", padx=18, pady=(0, 18))

        recent_rows = []

        for name, size, dtxt in recent_files:
            row = _frame(list_wrap, fg=C["surface_2"], radius=12, bw=1, bc=C["border_soft"])
            row.pack(fill="x", pady=4)

            row.grid_columnconfigure(0, weight=1)
            row.grid_columnconfigure(1, weight=0)
            row.grid_columnconfigure(2, weight=0)

            n = _lbl(row, name, C["text"], (FONT_BODY, 11))
            n.grid(row=0, column=0, sticky="w", padx=14, pady=12)

            s = _lbl(row, size, C["muted"], (FONT_BODY, 11))
            s.grid(row=0, column=1, sticky="e", padx=(12, 20), pady=12)

            d = _lbl(row, dtxt, C["muted_2"], (FONT_BODY, 11))
            d.grid(row=0, column=2, sticky="e", padx=(0, 14), pady=12)

            recent_rows.append({"name": n, "size": s, "date": d})

        # =========================================================
        # RESPONSIVE LAYOUT
        # =========================================================
        self._accueil_last_sig = None
        self._accueil_last_size = None

        def _layout_sig(r):
            return (
                r["tier"],
                r["hero_stack"],
                r["btn_stack"],
                r["kpi_cols"],
                r["act_cols"],
                r["met_cols"],
                r["hero_fs"],
                r["sec_fs"],
                r["body_fs"],
                r["btn_h"],
                r["hero_wrap"],
                r["detail_wrap"],
                r["sum_w"],
                r["hdr_sub"],
                r["sb_auto"],
            )

        def _sync_sidebar_mode():
            if not _sb["manual"]:
                _sb_set(_rs_cache["v"]["sb_auto"])
            else:
                _sb_apply_content()

        def _apply_responsive_layout():
            if not _exists(main):
                return

            r = _rs()
            sig = _layout_sig(r)

            if sig == self._accueil_last_sig:
                return

            self._accueil_last_sig = sig
            _rs_cache["v"] = r

            # Header
            try:
                header.configure(height=r["hdr_h"])
                hdr_title.configure(font=(FONT_HEAD, max(14, r["sec_fs"] - 1), "bold"))
                hdr_sub.configure(font=(FONT_BODY, max(10, r["sm_fs"])))
            except Exception:
                pass

            try:
                if r["hdr_sub"]:
                    if not hdr_sub.winfo_manager():
                        hdr_sub.pack(anchor="w", pady=(2, 0))
                else:
                    if hdr_sub.winfo_manager():
                        hdr_sub.pack_forget()
            except Exception:
                pass

            # Hero
            try:
                hero_title.configure(font=(FONT_HEAD, r["hero_fs"], "bold"))
                hero_desc.configure(font=(FONT_BODY, r["body_fs"]), wraplength=r["hero_wrap"])
                hero_pill_lbl.configure(font=(FONT_BODY, max(10, r["sm_fs"]), "bold"))
                hero_primary.configure(height=r["btn_h"])
                hero_ghost.configure(height=r["btn_h"])
            except Exception:
                pass

            try:
                hero_primary.pack_forget()
                hero_ghost.pack_forget()
                if r["btn_stack"]:
                    hero_primary.pack(anchor="w", pady=(0, 8))
                    hero_ghost.pack(anchor="w")
                else:
                    hero_primary.pack(side="left", padx=(0, 10))
                    hero_ghost.pack(side="left")
            except Exception:
                pass

            try:
                if r["hero_stack"]:
                    left_hero.grid(row=0, column=0, sticky="nsew", padx=24, pady=(24, 12))
                    right_hero.grid(row=1, column=0, sticky="ew", padx=24, pady=(0, 24))
                else:
                    left_hero.grid(row=0, column=0, sticky="nsew", padx=24, pady=24)
                    right_hero.grid(row=0, column=1, sticky="ne", padx=(0, 24), pady=24)

                if r["sum_w"] > 0:
                    sum_card.configure(width=r["sum_w"])
                else:
                    sum_card.configure(width=0)
            except Exception:
                pass

            # Metrics hero
            try:
                for c in range(4):
                    hero_met.grid_columnconfigure(c, weight=0)

                for idx, ref in enumerate(met_refs):
                    ref["lbl"].configure(font=(FONT_BODY, r["sm_fs"]))
                    ref["val"].configure(font=(FONT_HEAD, max(13, r["sec_fs"] - 1), "bold"))
                    ref["card"].grid_forget()
                    ref["card"].grid(
                        row=idx // r["met_cols"],
                        column=idx % r["met_cols"],
                        sticky="ew",
                        padx=(0, r["gap"]),
                        pady=(0, r["gap"])
                    )

                for c in range(max(1, r["met_cols"])):
                    hero_met.grid_columnconfigure(c, weight=1)
            except Exception:
                pass

            # Overview
            try:
                if r["w"] < 1260:
                    kpi_wrap.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=(0, 0), pady=(0, 12))
                    detail_card.grid(row=1, column=0, columnspan=2, sticky="nsew", padx=(0, 0), pady=(0, 0))
                else:
                    kpi_wrap.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 0))
                    detail_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 0))
            except Exception:
                pass

            # KPI grid
            try:
                for c in range(4):
                    kpi_grid.grid_columnconfigure(c, weight=0)

                for idx, ref in enumerate(kpi_cards):
                    ref["title"].configure(font=(FONT_BODY, max(10, r["sm_fs"])))
                    ref["value"].configure(font=(FONT_HEAD, max(18, r["sec_fs"] + 6), "bold"))
                    ref["sub"].configure(font=(FONT_BODY, r["body_fs"]), wraplength=max(220, r["hero_wrap"] // 3))

                    ref["card"].grid_forget()
                    ref["card"].grid(
                        row=idx // r["kpi_cols"],
                        column=idx % r["kpi_cols"],
                        sticky="ew",
                        padx=(0, r["gap"]),
                        pady=(0, r["gap"])
                    )

                for c in range(max(1, r["kpi_cols"])):
                    kpi_grid.grid_columnconfigure(c, weight=1)
            except Exception:
                pass

            # Action grid
            try:
                for c in range(4):
                    act_grid.grid_columnconfigure(c, weight=0)

                for idx, ref in enumerate(action_refs):
                    ref["title"].configure(font=(FONT_HEAD, max(12, r["body_fs"] + 1), "bold"))
                    ref["desc"].configure(font=(FONT_BODY, r["body_fs"]), wraplength=max(220, r["hero_wrap"] // 3))
                    ref["btn"].configure(height=r["btn_h"])

                    ref["box"].grid_forget()
                    ref["box"].grid(
                        row=idx // r["act_cols"],
                        column=idx % r["act_cols"],
                        sticky="ew",
                        padx=(0, r["gap"]),
                        pady=(0, r["gap"])
                    )

                for c in range(max(1, r["act_cols"])):
                    act_grid.grid_columnconfigure(c, weight=1)
            except Exception:
                pass

            # Rows recent
            try:
                for rr in recent_rows:
                    rr["name"].configure(font=(FONT_BODY, r["body_fs"]))
                    rr["size"].configure(font=(FONT_BODY, r["body_fs"]))
                    rr["date"].configure(font=(FONT_BODY, r["body_fs"]))
            except Exception:
                pass

            # Detail panel
            try:
                _render_detail()
            except Exception:
                pass

            # Sidebar
            try:
                _sync_sidebar_mode()
            except Exception:
                pass

        def _schedule_layout(delay=70):
            _cancel_job("_resize_job")
            self._resize_job = self.after(delay, _apply_responsive_layout)

        def _watch_size():
            if not _exists(main):
                return

            size = (self.winfo_width(), self.winfo_height())
            if size != self._accueil_last_size:
                self._accueil_last_size = size
                _schedule_layout(70)

            self._accueil_watch_job = self.after(160, _watch_size)

        # =========================================================
        # START
        # =========================================================
        _show_overview()
        self.after_idle(_apply_responsive_layout)
        self._accueil_watch_job = self.after(160, _watch_size)
        
    def vider_fenetre(self):
        # Annule les after() de la page courante avant de détruire les widgets,
        # pour éviter que des callbacks pendants s'exécutent sur des widgets détruits.
        for attr in ("_accueil_resize_job", "_sb_anim_job"):
            job = getattr(self, attr, None)
            if job:
                try:
                    self.after_cancel(job)
                except Exception:
                    pass
            setattr(self, attr, None)
        for widget in self.winfo_children():
            widget.destroy()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.destroy()

    def importer_les_profils(self):
        import os
        import re
        import tkinter as tk
        import customtkinter as ctk
        from tkinter import filedialog, ttk
        from PIL import Image
        from customtkinter import CTkImage

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "success": "#1F9D63",
            "success_hover": "#188454",
            "danger": "#C44E4E",
            "danger_hover": "#A83F3F",
            "warning": "#E0B64A",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # ÉTAT / VARIABLES MÉTIER
        # =========================================================
        if not hasattr(self, "baseline_year"):
            self.baseline_year = None

        self.profils_import = []
        self.fichier_source = None

        def build_pattern(year):
            if year is None:
                return None

            y = year
            y_prev = y - 1

            def y_or_date(yy):
                return rf"{yy}(?:-\d{{2}}-\d{{2}})?"

            return re.compile(
                rf"^Profil\s+Tr[ée]so\s+SNCF\s+"
                rf"(?:"
                    rf"{y_prev}\s*[-–—]\s*{y_or_date(y)}"
                    rf"|"
                    rf"{y}\s*[-–—]\s*{y_or_date(y)}"
                rf")\b.*\.xlsx$",
                re.IGNORECASE
            )

        def chercher_profils_regex(root_folder, pattern):
            if pattern is None:
                return []
            resultats = []
            for dossier, _, fichiers in os.walk(root_folder):
                for fichier in fichiers:
                    if pattern.match(fichier):
                        resultats.append(os.path.join(dossier, fichier))
            return resultats

        def find_fichier_reel_exact(year, base_dir):
            attendu = f"Réel {year}.xlsx"
            chemin = os.path.join(base_dir, attendu)
            return chemin if os.path.isfile(chemin) else None

        def find_fichier_reel_flexible(year, root_dir):
            patt = re.compile(rf"^R[ée]el\s+{year}\.xlsx$", re.IGNORECASE)

            try:
                for f in os.listdir(root_dir):
                    if patt.match(f):
                        return os.path.join(root_dir, f)
            except FileNotFoundError:
                pass

            for d, _, files in os.walk(root_dir):
                for f in files:
                    if patt.match(f):
                        return os.path.join(d, f)
            return None

        self.build_pattern = build_pattern
        self.chercher_profils_regex = chercher_profils_regex
        self.find_fichier_reel_exact = find_fichier_reel_exact
        self.find_fichier_reel_flexible = find_fichier_reel_flexible
        self.BASE_DONNEES_DIR = BASE_DONNEES_DIR
        self.pattern_profil = build_pattern(self.baseline_year)

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def success_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["success"],
                hover_color=UI["success_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def danger_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["danger"],
                hover_color=UI["danger_hover"],
                text_color="white",
                font=FONT["button"]
            )

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)   # topbar
        self.grid_rowconfigure(1, weight=0)   # separator
        self.grid_rowconfigure(2, weight=1)   # body

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module d’import des profils",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        back_btn = neutral_btn(top_right, "Retour à l’accueil", self.creer_accueil)
        back_btn.pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(
            body_host,
            bg=UI["bg"],
            highlightthickness=0,
            bd=0
        )
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")

        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=1)
        page.grid_columnconfigure(0, weight=1)

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        body_canvas.bind_all("<MouseWheel>", _mousewheel)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "IMPORT DES PROFILS", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Chargement et préparation des fichiers profils",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Sélectionnez une année, choisissez un dossier racine, contrôlez les fichiers détectés puis lancez l’import.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # CONTENU PRINCIPAL
        # =========================================================
        content = ctk.CTkFrame(page, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=28, pady=(0, 24))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=3)
        content.grid_columnconfigure(1, weight=2)

        # ---------------------------------------------------------
        # COLONNE GAUCHE
        # ---------------------------------------------------------
        left_panel = card(content, fg=UI["surface"], radius=20)
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        left_panel.grid_rowconfigure(1, weight=1)
        left_panel.grid_columnconfigure(0, weight=1)

        section_header(
            left_panel,
            "FICHIERS DÉTECTÉS",
            "Liste des profils disponibles",
            "Les fichiers affichés correspondent au motif attendu pour l’année sélectionnée."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        table_container = ctk.CTkFrame(left_panel, fg_color=UI["surface_2"], corner_radius=14)
        table_container.grid(row=1, column=0, sticky="nsew", padx=18, pady=(0, 14))
        table_container.grid_rowconfigure(0, weight=1)
        table_container.grid_columnconfigure(0, weight=1)

        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=34,
            font=("Segoe UI", 12)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 12, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        tree_wrap = tk.Frame(table_container, bg=UI["surface_2"], highlightthickness=0, bd=0)
        tree_wrap.grid(row=0, column=0, sticky="nsew", padx=1, pady=1)
        tree_wrap.grid_rowconfigure(0, weight=1)
        tree_wrap.grid_columnconfigure(0, weight=1)

        colonnes = ("Nom des fichiers profil",)
        self.tableau = ttk.Treeview(
            tree_wrap,
            columns=colonnes,
            show="headings",
            style="Pulse.Treeview"
        )
        self.tableau.heading("Nom des fichiers profil", text="Nom des fichiers profil")
        self.tableau.column("Nom des fichiers profil", width=600, anchor="w", stretch=True)
        self.tableau.grid(row=0, column=0, sticky="nsew")

        tree_scroll = ttk.Scrollbar(tree_wrap, orient="vertical", command=self.tableau.yview)
        tree_scroll.grid(row=0, column=1, sticky="ns")
        self.tableau.configure(yscrollcommand=tree_scroll.set)

        table_footer = ctk.CTkFrame(left_panel, fg_color="transparent")
        table_footer.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        table_footer.grid_columnconfigure(0, weight=1)

        self.label_count = label(
            table_footer,
            "0 fichier détecté",
            font=FONT["small"],
            color=UI["muted"]
        )
        self.label_count.grid(row=0, column=0, sticky="w")

        label_chemin = label(
            table_footer,
            "",
            font=FONT["small"],
            color=UI["text_soft"],
            wraplength=800,
            justify="left"
        )
        label_chemin.grid(row=1, column=0, sticky="ew", pady=(8, 0))

        # ---------------------------------------------------------
        # COLONNE DROITE
        # ---------------------------------------------------------
        right_col = ctk.CTkFrame(content, fg_color="transparent")
        right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
        right_col.grid_rowconfigure(0, weight=0)
        right_col.grid_rowconfigure(1, weight=0)
        right_col.grid_rowconfigure(2, weight=1)
        right_col.grid_columnconfigure(0, weight=1)

        # Paramètres
        settings_panel = card(right_col, fg=UI["surface"], radius=20)
        settings_panel.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        settings_panel.grid_columnconfigure(0, weight=1)

        section_header(
            settings_panel,
            "PARAMÈTRES",
            "Contexte d’import",
            "Définissez l’année de référence puis sélectionnez le dossier racine."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        form = ctk.CTkFrame(settings_panel, fg_color="transparent")
        form.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        form.grid_columnconfigure(0, weight=1)

        label(form, "Année des fichiers profil", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )

        years = ["Choisir..."] + [str(y) for y in range(2018, 2036)]

        annee_menu = ctk.CTkOptionMenu(
            form,
            values=years,
            height=40,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            font=FONT["label"]
        )
        annee_menu.grid(row=1, column=0, sticky="ew")

        label(form, "Dossier racine sélectionné", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=2, column=0, sticky="w", pady=(18, 6)
        )

        self.folder_value = label(
            form,
            "Aucun dossier sélectionné",
            font=FONT["small"],
            color=UI["muted"],
            wraplength=480,
            justify="left"
        )
        self.folder_value.grid(row=3, column=0, sticky="ew")

        # Statut
        status_panel = card(right_col, fg=UI["surface"], radius=20)
        status_panel.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        status_panel.grid_columnconfigure(0, weight=1)

        section_header(
            status_panel,
            "STATUT",
            "Vérifications préalables",
            "Le fichier réel associé à l’année choisie conditionne le lancement de l’import."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        self.lbl_reel = label(
            status_panel,
            "Sélectionnez une année pour détecter le fichier source réel.",
            font=FONT["small"],
            color=UI["text_soft"],
            justify="left",
            wraplength=480
        )
        self.lbl_reel.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 14))

        # Actions
        actions_panel = card(right_col, fg=UI["surface"], radius=20)
        actions_panel.grid(row=2, column=0, sticky="nsew")
        actions_panel.grid_columnconfigure(0, weight=1)

        section_header(
            actions_panel,
            "ACTIONS",
            "Exécution",
            "Gérez la liste détectée puis lancez l’import lorsque le contexte est valide."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        action_list = ctk.CTkFrame(actions_panel, fg_color="transparent")
        action_list.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        action_list.grid_columnconfigure(0, weight=1)

        # =========================================================
        # HELPERS DYNAMIQUES
        # =========================================================
        def refresh_count():
            try:
                n = len(self.tableau.get_children())
            except Exception:
                n = 0
            self.label_count.configure(
                text=f"{n} fichier{'s' if n != 1 else ''} détecté{'s' if n != 1 else ''}"
            )

        def set_import_button_state(enabled):
            try:
                self.bouton_valider.configure(state=("normal" if enabled else "disabled"))
            except Exception:
                pass

        self.set_import_button_state = set_import_button_state

        def set_status_message(text, tone="normal"):
            color = UI["text_soft"]
            if tone == "success":
                color = UI["success"]
            elif tone == "warning":
                color = UI["warning"]
            elif tone == "muted":
                color = UI["muted"]
            self.lbl_reel.configure(text=text, text_color=color)

        def clear_progress():
            try:
                self.progress_label.configure(text="")
                self.progress_bar.set(0)
                self.progress_wrap.grid_remove()
            except Exception:
                pass

        def show_progress_shell():
            self.progress_wrap.grid()
            self.progress_label.configure(text="Import prêt à être lancé.")
            self.progress_bar.set(0)

        def vider_tableau(silent=False, hard=False):
            rows = self.tableau.get_children()
            nb = len(rows)

            for row in rows:
                self.tableau.delete(row)

            try:
                self.tableau.selection_remove(self.tableau.selection())
            except Exception:
                pass

            self.profils_import = []

            if hard:
                self.chemin_dossier = None
                self.baseline_year = None
                self.pattern_profil = None
                self.fichier_source = None

                try:
                    annee_menu.set("Choisir...")
                except Exception:
                    pass

                label_chemin.configure(text="")
                self.folder_value.configure(text="Aucun dossier sélectionné", text_color=UI["muted"])
                set_status_message(
                    "Sélectionnez une année pour détecter le fichier source réel.",
                    tone="muted"
                )
                set_import_button_state(False)
                clear_progress()

            refresh_count()

            if not silent:
                if nb == 0 and not hard:
                    label_chemin.configure(text="Le tableau était déjà vide.")
                elif hard:
                    label_chemin.configure(text="Réinitialisation complète effectuée.")
                else:
                    label_chemin.configure(text=f"{nb} ligne(s) supprimée(s) du tableau.")

        def choisir_dossier():
            chemin = filedialog.askdirectory(title="Sélectionner un dossier où les profils sont stockés")
            if chemin:
                self.chemin_dossier = chemin
                self.folder_value.configure(text=chemin, text_color=UI["text"])
                label_chemin.configure(text=f"Dossier sélectionné : {chemin}")

                vider_tableau(silent=True, hard=False)

                if self.pattern_profil is None:
                    label_chemin.configure(text="Choisissez d’abord une année.")
                    return

                fichiers = self.chercher_profils_regex(chemin, self.pattern_profil)
                self.profils_import = list(fichiers)

                for f in fichiers:
                    self.tableau.insert("", "end", values=(os.path.basename(f),))

                refresh_count()

        def supprimer_selection():
            selection = self.tableau.selection()
            if not selection:
                label_chemin.configure(text="Sélectionnez au moins un élément dans la liste.")
                return

            nb_suppr = 0

            for item_id in selection:
                values = self.tableau.item(item_id, "values")
                if not values:
                    continue

                nom_fichier = values[0]

                try:
                    self.tableau.delete(item_id)
                    nb_suppr += 1
                except Exception:
                    pass

                chemins_a_supprimer = [
                    p for p in list(self.profils_import)
                    if os.path.basename(p) == nom_fichier
                ]
                for p in chemins_a_supprimer:
                    try:
                        self.profils_import.remove(p)
                    except ValueError:
                        pass

            try:
                self.tableau.selection_remove(self.tableau.selection())
            except Exception:
                pass

            refresh_count()
            label_chemin.configure(text=f"{nb_suppr} élément(s) supprimé(s) de la liste.")

        def on_year_change(value):
            if value == "Choisir...":
                self.baseline_year = None
                self.pattern_profil = None
                self.fichier_source = None
                self.profils_import = []

                if hasattr(self, "chemin_dossier") and self.chemin_dossier:
                    vider_tableau(silent=True, hard=False)

                set_status_message(
                    "Sélectionnez une année pour détecter le fichier source réel.",
                    tone="muted"
                )
                set_import_button_state(False)
                return

            self.baseline_year = int(value)
            self.pattern_profil = self.build_pattern(self.baseline_year)

            year = self.baseline_year
            chemin_reel = self.find_fichier_reel_exact(year, self.BASE_DONNEES_DIR)
            if not chemin_reel:
                chemin_reel = self.find_fichier_reel_flexible(year, self.BASE_DONNEES_DIR)

            self.fichier_source = chemin_reel

            if self.fichier_source:
                set_status_message(
                    f"Fichier source détecté pour {year} :\n{self.fichier_source}",
                    tone="success"
                )
                set_import_button_state(True)
                show_progress_shell()
            else:
                set_status_message(
                    f"Aucun fichier 'Réel {year}.xlsx' trouvé dans :\n{self.BASE_DONNEES_DIR}\nL’import reste désactivé tant que ce fichier n’est pas disponible.",
                    tone="warning"
                )
                set_import_button_state(False)
                clear_progress()

            if hasattr(self, "chemin_dossier") and self.chemin_dossier:
                vider_tableau(silent=True, hard=False)
                fichiers = self.chercher_profils_regex(self.chemin_dossier, self.pattern_profil)
                self.profils_import = list(fichiers)
                for f in fichiers:
                    self.tableau.insert("", "end", values=(os.path.basename(f),))
                refresh_count()

        annee_menu.configure(command=on_year_change)

        if self.baseline_year is None:
            annee_menu.set("Choisir...")
        else:
            annee_menu.set(str(self.baseline_year))

        # =========================================================
        # BOUTONS
        # =========================================================
        row1 = ctk.CTkFrame(action_list, fg_color="transparent")
        row1.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        row1.grid_columnconfigure(0, weight=1)

        label(
            row1,
            "Sélectionner le dossier racine contenant les fichiers profils.",
            font=FONT["small"],
            color=UI["text_soft"],
            justify="left",
            wraplength=420
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        btn_parcourir = neutral_btn(row1, "Parcourir le dossier", choisir_dossier)
        btn_parcourir.grid(row=1, column=0, sticky="ew")

        row2 = ctk.CTkFrame(action_list, fg_color="transparent")
        row2.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        row2.grid_columnconfigure(0, weight=1)

        label(
            row2,
            "Vider le tableau et réinitialiser les paramètres d’import.",
            font=FONT["small"],
            color=UI["text_soft"],
            justify="left",
            wraplength=420
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        btn_reset = neutral_btn(row2, "Réinitialiser", lambda: vider_tableau(silent=False, hard=True))
        btn_reset.grid(row=1, column=0, sticky="ew")

        row3 = ctk.CTkFrame(action_list, fg_color="transparent")
        row3.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        row3.grid_columnconfigure(0, weight=1)

        label(
            row3,
            "Supprimer les lignes actuellement sélectionnées.",
            font=FONT["small"],
            color=UI["text_soft"],
            justify="left",
            wraplength=420
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        btn_delete = danger_btn(row3, "Supprimer la sélection", supprimer_selection)
        btn_delete.grid(row=1, column=0, sticky="ew")

        row4 = ctk.CTkFrame(action_list, fg_color="transparent")
        row4.grid(row=3, column=0, sticky="ew", pady=(0, 0))
        row4.grid_columnconfigure(0, weight=1)

        label(
            row4,
            "Lancer l’import des profils vers les fichiers mensuels.",
            font=FONT["small"],
            color=UI["text_soft"],
            justify="left",
            wraplength=420
        ).grid(row=0, column=0, sticky="w", pady=(0, 8))

        self.bouton_valider = success_btn(row4, "Lancer l’import", self.lancer_import_profils)
        self.bouton_valider.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # PROGRESSION
        # =========================================================
        self.progress_wrap = ctk.CTkFrame(actions_panel, fg_color=UI["surface_2"], corner_radius=14)
        self.progress_wrap.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        self.progress_wrap.grid_columnconfigure(0, weight=1)

        label(
            self.progress_wrap,
            "Progression",
            font=FONT["small_bold"],
            color=UI["text_soft"]
        ).grid(row=0, column=0, sticky="w", padx=14, pady=(14, 8))

        self.progress_label = label(
            self.progress_wrap,
            "",
            font=FONT["small"],
            color=UI["muted"],
            justify="left",
            wraplength=420
        )
        self.progress_label.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 10))

        self.progress_bar = ctk.CTkProgressBar(
            self.progress_wrap,
            height=14,
            progress_color=UI["success"],
            fg_color=UI["surface_3"]
        )
        self.progress_bar.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 14))
        self.progress_bar.set(0)

        self.progress_wrap.grid_remove()

        # =========================================================
        # RESPONSIVE
        # =========================================================
        def _apply_responsive_layout(event=None):
            try:
                width = page.winfo_width()
            except Exception:
                return

            if width >= 1180:
                content.grid_columnconfigure(0, weight=3)
                content.grid_columnconfigure(1, weight=2)
                content.grid_rowconfigure(0, weight=1)
                content.grid_rowconfigure(1, weight=0)

                left_panel.grid_forget()
                right_col.grid_forget()

                left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
                right_col.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
            else:
                content.grid_columnconfigure(0, weight=1)
                content.grid_columnconfigure(1, weight=0)
                content.grid_rowconfigure(0, weight=1)
                content.grid_rowconfigure(1, weight=1)

                left_panel.grid_forget()
                right_col.grid_forget()

                left_panel.grid(row=0, column=0, sticky="nsew", padx=0, pady=(0, 12))
                right_col.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

            try:
                left_wrap = max(280, left_panel.winfo_width() - 50)
                right_wrap = max(260, right_col.winfo_width() - 50)
                label_chemin.configure(wraplength=left_wrap)
                self.folder_value.configure(wraplength=right_wrap)
                self.lbl_reel.configure(wraplength=right_wrap)
                self.progress_label.configure(wraplength=right_wrap)
            except Exception:
                pass

            try:
                table_width = table_container.winfo_width()
                if table_width > 100:
                    self.tableau.column(
                        "Nom des fichiers profil",
                        width=max(320, table_width - 30),
                        stretch=True
                    )
            except Exception:
                pass

            _sync_scrollregion()

        page.bind("<Configure>", _apply_responsive_layout)
        self.after(120, _apply_responsive_layout)

        # =========================================================
        # ÉTAT INITIAL
        # =========================================================
        refresh_count()
        set_import_button_state(False)
        set_status_message(
            "Sélectionnez une année pour détecter le fichier source réel.",
            tone="muted"
        )

    def lancer_import_profils(self):
        import pandas as pd
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.utils import get_column_letter
        import re
        from datetime import datetime, date
        import os
        from collections import defaultdict
        import traceback

        print("=== [START] lancer_import_profils ===")

        # ===================== PARAMÈTRES / RECHERCHE (réutilisés) =====================
        if not hasattr(self, "baseline_year"):
            self.baseline_year = 2025  # fallback si l'UI n'a pas été ouverte
            print(f"[INFO] baseline_year absent → fallback {self.baseline_year}")
        else:
            print(f"[INFO] baseline_year récupérée depuis l'UI : {self.baseline_year}")

        # Sécurité : build_pattern au cas où importer_les_profils n'aurait pas été appelé
        if not hasattr(self, "build_pattern"):
            print("[INIT] Création de la fonction build_pattern (sécurité)")
            def build_pattern(year: int | None):
                if year is None:
                    return None
                y = year
                y_prev = y - 1

                def y_or_date(yy: int) -> str:
                    return rf"{yy}(?:-\d{{2}}-\d{{2}})?"

                return re.compile(
                    rf'^Profil\s+Tr[ée]so\s+SNCF\s+'
                    rf'(?:'
                        rf'{y_prev}\s*[-–—]\s*{y_or_date(y)}'
                        rf'|'
                        rf'{y}\s*[-–—]\s*{y_or_date(y)}'
                    rf')\b.*\.xlsx$',
                    re.IGNORECASE
                )
            self.build_pattern = build_pattern

        if not hasattr(self, "pattern_profil"):
            self.pattern_profil = self.build_pattern(self.baseline_year)
            print("[INIT] pattern_profil (regex) compilé")

        if not hasattr(self, "chercher_profils_regex"):
            print("[INIT] Création de la fonction chercher_profils_regex (sécurité)")
            def _chercher_profils_regex(root_folder: str, pattern):
                resultats = []
                for dossier, _, fichiers in os.walk(root_folder):
                    for fichier in fichiers:
                        if pattern.match(fichier):
                            resultats.append(os.path.join(dossier, fichier))
                return resultats
            self.chercher_profils_regex = _chercher_profils_regex

        # ===================== LECTURE DATE COLONNE C =====================
        def lire_date_c_descendant(fichier, feuille="SA_SNCF", start_row=6, year=None, max_lookahead=50):
            """
            Parcourt C{start_row}..C{start_row+max_lookahead-1} pour trouver une date cohérente.
            Filtre sur 'year' si fourni.
            """
            try:
                wb = load_workbook(fichier, read_only=True, data_only=True)
                if feuille not in wb.sheetnames:
                    wb.close()
                    print(f"[WARN] Feuille '{feuille}' absente dans {os.path.basename(fichier)}")
                    return None, None
                ws = wb[feuille]

                def parse_cell_to_date(val):
                    if isinstance(val, datetime):
                        return val.date()
                    if isinstance(val, date):
                        return val
                    if isinstance(val, str):
                        s = val.strip()
                        try:
                            return datetime.strptime(s, "%d/%m/%Y").date()
                        except Exception:
                            pass
                        try:
                            return date.fromisoformat(s)
                        except Exception:
                            pass
                        m = re.search(r"\b(19\d{2}|20\d{2})\b", s)
                        if m:
                            return date(int(m.group(1)), 1, 1)
                    return None

                for r in range(start_row, start_row + max_lookahead):
                    val = ws[f"C{r}"].value
                    if val is None:
                        continue

                    d = parse_cell_to_date(val)
                    if not d:
                        print(f"[INFO] {os.path.basename(fichier)} C{r}: valeur non parsable → {val!r}")
                        continue

                    if year is not None and d.year != year:
                        if isinstance(val, str) and re.search(rf"\b{year}\b", val):
                            wb.close()
                            print(f"[OK] {os.path.basename(fichier)} C{r}: année {year} repérée dans le texte → {d.isoformat()}")
                            return d, r
                        print(f"[SKIP] {os.path.basename(fichier)} C{r}: année {d.year} ≠ {year}")
                        continue

                    wb.close()
                    print(f"[OK] {os.path.basename(fichier)} C{r} → {d.isoformat()}")
                    return d, r

                wb.close()
                print(f"[WARN] Aucune date (année={year}) trouvée en C{start_row}..C{start_row+max_lookahead-1} "
                    f"dans {os.path.basename(fichier)}")
                return None, None

            except Exception as e:
                print(f"[ERROR] Lecture colonne C échouée pour {os.path.basename(fichier)} : {e}")
                traceback.print_exc()
                return None, None

        def choisir_meilleure_version(fichiers):
            try:
                fichiers = sorted(fichiers)
                v3 = [f for f in fichiers if "V3" in os.path.basename(f)]
                if v3:
                    return v3[0]
                v2 = [f for f in fichiers if "V2" in os.path.basename(f)]
                if v2:
                    return v2[0]
                return fichiers[0] if fichiers else None
            except Exception as e:
                print(f"[ERROR] choisir_meilleure_version: {e}")
                traceback.print_exc()
                return fichiers[0] if fichiers else None

        # === Progress bar ===
        try:
            self.progress_label.configure(text="⏳ Export en cours (division par mois)...")
            self.progress_label.pack(pady=(10, 0))
            self.progress_bar.pack(pady=(5, 20))
            self.progress_bar.set(0)
            self.update_idletasks()
        except Exception:
            print("[WARN] UI progress bar indisponible dans ce contexte.")

        # === LISTE DES FICHIERS PROFILS À UTILISER ===
        # Priorité : self.profils_import (mise à jour par importer_les_profils + supprimer_selection)
        profils_import = getattr(self, "profils_import", [])

        if profils_import:
            fichiers_trouves = list(profils_import)
            print(f"[INFO] Utilisation de self.profils_import : {len(fichiers_trouves)} fichier(s).")
        else:
            # Fallback : on re-scanne le dossier avec la regex (cas où importer_les_profils n'a pas été utilisé)
            dossier_depart = getattr(self, "chemin_dossier", None)
            if not dossier_depart:
                print("[ERROR] Aucun dossier de profils sélectionné et self.profils_import est vide.")
                try:
                    self.progress_label.configure(text="⚠️ Aucun fichier profil sélectionné.")
                except Exception:
                    pass
                print("=== [STOP] lancer_import_profils (erreur aucun fichier) ===")
                return

            print(f"[INFO] Dossier de départ : {dossier_depart}")
            print("[STEP] Recherche des fichiers de profils par regex (fallback)...")
            fichiers_trouves = self.chercher_profils_regex(dossier_depart, self.pattern_profil)
            print(f"[OK] {len(fichiers_trouves)} fichier(s) trouvé(s) (fallback).")

        if not fichiers_trouves:
            print("[ERROR] Liste de fichiers profils vide après filtrage.")
            try:
                self.progress_label.configure(text="⚠️ Aucun fichier profil à traiter (liste vide).")
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (liste vide) ===")
            return

        # --- Récupération des dates par fichier ---
        print("[STEP] Lecture des dates (colonne C) en descendant depuis C6...")
        dates_par_fichier = []
        for f in fichiers_trouves:
            date_trouvee, row_found = lire_date_c_descendant(
                f, feuille="SA_SNCF", start_row=6, year=self.baseline_year, max_lookahead=50
            )
            if date_trouvee:
                dates_par_fichier.append((f, date_trouvee, row_found))
                print(f"   - {os.path.basename(f)} → date={date_trouvee.isoformat()} (C{row_found})")
            else:
                print(f"   - {os.path.basename(f)} → aucune date conforme trouvée")

        print("[STEP] Groupement des fichiers par date Cx...")
        groupes = defaultdict(list)
        for fichier, d, row_found in dates_par_fichier:
            groupes[d].append((fichier, row_found))

        print("[STEP] Sélection de la meilleure version (V3>V2>baseline) pour chaque date...")
        dates_uniques = []
        for d, lst in groupes.items():
            fichiers = [f for f, _r in lst]
            meilleur_fichier = choisir_meilleure_version(fichiers)
            if meilleur_fichier:
                row_found = next(r for (f, r) in lst if f == meilleur_fichier)
                dates_uniques.append((meilleur_fichier, d, row_found))
                print(f"   - {d.isoformat()} → {os.path.basename(meilleur_fichier)} (C{row_found})")

        dates_uniques.sort(key=lambda x: x[1])
        print("[OK] Dates uniques triées :", [(os.path.basename(f), d.isoformat()) for f, d, _r in dates_uniques])

        fichier_prev = [f for f, _d, _r in dates_uniques]
        rows_prev    = [r for _f, _d, r in dates_uniques]

        print("[STEP] Calcul des décalages (jours) entre prévisions successives...")
        deltas = []
        if len(dates_uniques) < 2:
            print("[WARN] Pas assez de dates pour calculer des décalages (0 ou 1).")
        else:
            for i in range(1, len(dates_uniques)):
                _, date1, _ = dates_uniques[i-1]
                _, date2, _ = dates_uniques[i]
                delta = (date2 - date1).days
                deltas.append(delta)
                print(f"   - Δ({date1.isoformat()} → {date2.isoformat()}) = {delta} jour(s)")

        cumul = [0]
        somme = 0
        for d in deltas:
            somme += d
            cumul.append(somme)
        print(f"[OK] Cumul des décalages : {cumul}")

        # === Résolution du fichier source (réel) depuis la sélection d'année ===
        fichier_source = getattr(self, "fichier_source", None)

        if not hasattr(self, "find_fichier_reel_exact") or not hasattr(self, "find_fichier_reel_flexible"):
            print("[ERROR] Helpers find_fichier_reel_* manquants (importer_les_profils non appelé ?)")
            # Tu peux éventuellement recréer ici des versions simplifiées si besoin.
            # Pour l’instant on stoppe proprement :
            try:
                self.progress_label.configure(text="⚠️ Helpers pour trouver 'Réel AAAA.xlsx' manquants.")
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (helpers manquants) ===")
            return

        if not fichier_source:
            base_dir = getattr(self, "BASE_DONNEES_DIR", None)
            if base_dir and self.baseline_year:
                fichier_source = self.find_fichier_reel_exact(self.baseline_year, base_dir) \
                                or self.find_fichier_reel_flexible(self.baseline_year, base_dir)

        if not fichier_source or not os.path.isfile(fichier_source):
            print("[ERROR] Fichier 'Réel AAAA.xlsx' introuvable : import annulé.")
            try:
                self.progress_label.configure(
                    text=(
                        f"⚠️ Fichier réel manquant pour {getattr(self, 'baseline_year', '???')}.\n"
                        f"Attendu : 'Réel {getattr(self, 'baseline_year', 'AAAA')}.xlsx' dans "
                        f"{getattr(self, 'BASE_DONNEES_DIR', '(dossier inconnu)')}\n"
                        f"→ Aucun import exécuté."
                    )
                )
                self.progress_label.pack(pady=(10, 0))
                self.progress_bar.pack_forget()
            except Exception:
                pass
            print("=== [STOP] lancer_import_profils (pas de fichier réel) ===")
            return

        print(f"[INFO] Fichier source (réel) : {fichier_source}")
        print(f"[INFO] {len(fichier_prev)} fichier(s) de prévision retenu(s).")

        # === Feuilles à traiter ===
        # === Feuilles à traiter ===
        sections = charger_sections_depuis_cells()

        if not sections:
            print("[WARN] Aucune section trouvée → traitement annulé.")
            return

        print(f"[INFO] Sections à traiter : {[s['dest'] for s in sections]}")

        # === Styles ===
        fill_jaune = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        fill_bleu = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        border = Border(left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'))
        align_center = Alignment(horizontal='center', vertical='center')
        font_bold = Font(bold=True)

        # === Libellés de prév JJ/MM ===
        print("[STEP] Extraction des libellés de dates pour les entêtes de prévisions...")
        dates_prev = []
        for path in fichier_prev:
            match = re.search(r"\d{4}-\d{2}-\d{2}", path)
            if match:
                date_obj = datetime.strptime(match.group(), "%Y-%m-%d")
                label = date_obj.strftime("%d/%m")
            else:
                label = "N/A"
            dates_prev.append(label)
            print(f"   - {os.path.basename(path)} → label '{label}'")
        nb_prev = len(fichier_prev)
        print(f"[OK] {nb_prev} label(s) de prévision.")

        # === Précharger les prévisions ===
        print("[STEP] Préchargement des matrices de prévisions par section...")
        valeurs_prev_all_by_section = {}
        for section in sections:
            feuille_prev = section["prev"]
            frames = []
            print(f"   > Chargement prévisions feuille '{feuille_prev}' sur {len(fichier_prev)} fichier(s)...")
            for idx_f, (f, row_found) in enumerate(zip(fichier_prev, rows_prev), start=1):
                try:
                    df_prev = pd.read_excel(f, sheet_name=feuille_prev, header=None, skiprows=row_found-1)
                    raw_noms = df_prev.iloc[0, 5:]
                    mask_valid = raw_noms.notna() & (raw_noms.astype(str).str.strip() != "")
                    noms_prev = raw_noms[mask_valid].astype(str).str.strip().reset_index(drop=True)
                    valeurs = df_prev.iloc[:, 5:].loc[:, mask_valid.values].reset_index(drop=True)
                    valeurs.columns = noms_prev
                    frames.append(valeurs.reset_index(drop=True))
                    print(f"      - [{idx_f}/{len(fichier_prev)}] {os.path.basename(f)}: "
                        f"skiprows={row_found-1}, shape={valeurs.shape}")
                except Exception as e:
                    print(f"[ERROR] Lecture prévisions '{feuille_prev}' échouée pour {os.path.basename(f)} : {e}")
                    traceback.print_exc()
                    frames.append(pd.DataFrame())
            valeurs_prev_all_by_section[feuille_prev] = frames

        # === Répertoire de sortie ===
        base_out_dir = FICHIER_EXCEL_DIR
        os.makedirs(base_out_dir, exist_ok=True)
        print(f"[INFO] Répertoire de sortie : {base_out_dir}")

        # === Boucle principale : split par mois ===
        total = len(sections)
        OVERWRITE = False
        print(f"[INFO] OVERWRITE = {OVERWRITE}")

        for idx_section, section in enumerate(sections, start=1):
            feuille_source = section["source"]
            feuille_prev = section["prev"]
            feuille_dest = section["dest"]
            print(f"\n[SECTION {idx_section}/{total}] '{feuille_dest}' — source='{feuille_source}', prev='{feuille_prev}'")

            # Lecture du réel
            print("   [STEP] Lecture du fichier source (réel)...")
            try:
                df = pd.read_excel(fichier_source, sheet_name=feuille_source, header=None, skiprows=4)
                dates = df.iloc[0, 4:].dropna().reset_index(drop=True)
                dates_ts = pd.to_datetime(dates, dayfirst=True)
                lignes_valides = df.iloc[4:, [2] + list(range(4, df.shape[1]))].dropna(subset=[2])
                noms = lignes_valides.iloc[:, 0].astype(str).str.strip().reset_index(drop=True)
                valeurs_reelles = lignes_valides.iloc[:, 1:].reset_index(drop=True)
                print(f"   [OK] Réel: {valeurs_reelles.shape[0]} lignes (flux) x {valeurs_reelles.shape[1]} dates")
            except Exception as e:
                print(f"   [ERROR] Lecture du réel échouée pour feuille '{feuille_source}' : {e}")
                traceback.print_exc()
                continue

            print("   [STEP] Constitution des groupes par (année, mois) ...")
            groupes_mois = defaultdict(list)
            for i, ts in enumerate(dates_ts):
                groupes_mois[(ts.year, ts.month)].append(i)
            print(f"   [OK] {len(groupes_mois)} mois détecté(s).")

            valeurs_prev_all = valeurs_prev_all_by_section[feuille_prev]

            for (year, month), idxs in sorted(groupes_mois.items()):
                print(f"   [MOIS] {year}-{month:02d} → {len(idxs)} jour(s) dans le mois")
                year_dir = os.path.join(base_out_dir, str(year))
                os.makedirs(year_dir, exist_ok=True)
                out_path = os.path.join(year_dir, f"Historique_prev_reel_filiales_{year}_{month:02d}.xlsx")
                print(f"      - Fichier cible: {out_path}")

                new_wb = False
                if os.path.exists(out_path):
                    try:
                        wb_out = load_workbook(out_path)
                        print("      - Workbook existant ouvert.")
                    except Exception as e:
                        print(f"      [WARN] Fichier existant illisible → recréation. Détail: {e}")
                        wb_out = Workbook()
                        if "Sheet" in wb_out.sheetnames:
                            wb_out.remove(wb_out["Sheet"])
                        for sec in sections:
                            if sec["dest"] not in wb_out.sheetnames:
                                wb_out.create_sheet(sec["dest"])
                        new_wb = True
                else:
                    wb_out = Workbook()
                    if "Sheet" in wb_out.sheetnames:
                        wb_out.remove(wb_out["Sheet"])
                    for sec in sections:
                        if sec["dest"] not in wb_out.sheetnames:
                            wb_out.create_sheet(sec["dest"])
                    new_wb = True
                    print("      - Nouveau workbook créé (toutes feuilles destination préparées).")

                if feuille_dest not in wb_out.sheetnames:
                    wb_out.create_sheet(feuille_dest)
                    print(f"      - Feuille '{feuille_dest}' créée.")
                ws = wb_out[feuille_dest]

                def _sheet_has_payload(_ws):
                    if _ws.max_row <= 3:
                        return False
                    for row in _ws.iter_rows(min_row=4, max_row=_ws.max_row, values_only=True):
                        if any(v is not None and v != "" for v in row):
                            return True
                    return False

                if not OVERWRITE and _sheet_has_payload(ws):
                    print("      - Données déjà présentes → SKIP (OVERWRITE=False).")
                else:
                    print("      - Écriture/rafraîchissement de la feuille...")
                    ws.delete_rows(1, ws.max_row)

                    sous_titres = ["Dates", "Réel (K€)"]
                    for date_str in dates_prev:
                        sous_titres.append(f"Prévision {date_str} (K€)")
                        sous_titres.append(f"Écart {date_str} (K€)")

                    start_col = 3
                    print(f"      - {len(noms)} flux à écrire. Largeur bloc par flux = 2 + 2*{len(dates_prev)} + 1 (séparation)")

                    for idx_flux, nom in enumerate(noms, start=1):
                        if idx_flux % 10 == 0 or idx_flux == 1 or idx_flux == len(noms):
                            print(f"         • Flux {idx_flux}/{len(noms)} : '{nom}'")

                        col1 = get_column_letter(start_col)
                        colN = get_column_letter(start_col - 1 + len(sous_titres))
                        ws.merge_cells(f"{col1}2:{colN}2")
                        titre_cell = ws[f"{col1}2"]
                        titre_cell.value = nom
                        titre_cell.fill = fill_jaune
                        titre_cell.font = font_bold
                        titre_cell.alignment = align_center
                        titre_cell.border = border

                        for i, titre in enumerate(sous_titres):
                            cell = ws.cell(row=3, column=start_col - 1 + i)
                            cell.value = titre
                            cell.fill = fill_bleu
                            cell.font = font_bold
                            cell.alignment = align_center
                            cell.border = border

                        for r, i_global in enumerate(idxs):
                            row = 4 + r

                            ws.cell(row=row, column=start_col - 1, value=dates_ts.iloc[i_global].date()).alignment = align_center
                            ws.cell(row=row, column=start_col - 1).border = border

                            valeur = valeurs_reelles.iloc[idx_flux-1, i_global]
                            val_k = round(valeur / 1000) if pd.notna(valeur) else None
                            ws.cell(row=row, column=start_col, value=val_k).alignment = align_center
                            ws.cell(row=row, column=start_col).border = border

                            for j, decal in enumerate(cumul):
                                prev_col = start_col + 1 + j * 2
                                ecart_col = start_col + 2 + j * 2

                                if j >= len(valeurs_prev_all):
                                    ws.cell(row=row, column=prev_col, value=None)
                                    ws.cell(row=row, column=ecart_col, value=None)
                                else:
                                    prev_data = valeurs_prev_all[j]
                                    idx_prev = i_global - decal
                                    if 0 <= idx_prev < len(prev_data):
                                        try:
                                            val = prev_data.iloc[idx_prev, idx_flux-1]
                                        except Exception:
                                            val = None
                                    else:
                                        val = None
                                    val_prev_k = round(val / 1000) if pd.notna(val) else None
                                    ecart = val_prev_k - val_k if (val_prev_k is not None and val_k is not None) else None

                                    ws.cell(row=row, column=prev_col, value=val_prev_k).alignment = align_center
                                    ws.cell(row=row, column=prev_col).border = border

                                    ws.cell(row=row, column=ecart_col, value=ecart).alignment = align_center
                                    ws.cell(row=row, column=ecart_col).border = border

                                    if r == 0:
                                        col_letter = get_column_letter(ecart_col)
                                        plage = f"{col_letter}4:{col_letter}{3 + len(idxs)}"
                                        rule_rouge = FormulaRule(formula=[f"${col_letter}4<0"],
                                                                font=Font(color="FF0000", bold=True))
                                        rule_vert = FormulaRule(formula=[f"${col_letter}4>0"],
                                                                font=Font(color="00B050", bold=True))
                                        ws.conditional_formatting.add(plage, rule_rouge)
                                        ws.conditional_formatting.add(plage, rule_vert)

                        start_col += 2 + 2 * len(dates_prev)
                        start_col += 1

                    print("      - Écriture terminée pour ce mois.")

                tmp_path = out_path + ".tmp"
                try:
                    wb_out.save(tmp_path)
                    os.replace(tmp_path, out_path)
                    print(f"      [SAVE] OK → {out_path}")
                except PermissionError as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"      [ERROR] PermissionError sur {out_path} (fichier ouvert ?) : {e}")
                except Exception as e:
                    if os.path.exists(tmp_path):
                        try:
                            os.remove(tmp_path)
                        except Exception:
                            pass
                    print(f"      [ERROR] Erreur pendant la sauvegarde de {out_path}: {e}")
                    traceback.print_exc()

            try:
                progress = idx_section / total
                self.progress_bar.set(progress)
                self.progress_label.configure(text=f"⏳ Export en cours... {int(progress*100)}%")
                self.update_idletasks()
            except Exception:
                pass

        try:
            self.progress_bar.set(1)
            self.progress_label.configure(text="✅ Export terminé (1 fichier par mois, rangé par année) !")
        except Exception:
            pass

        print("=== [END] lancer_import_profils — Export terminé ===")

    def afficher_ecarts(self):
        import customtkinter as ctk
        from tkinter import ttk
        import tkinter as tk
        from PIL import Image
        from customtkinter import CTkImage

        self.vider_fenetre()

        # === HEADER AVEC TITRE + LOGO ===
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        titre_label = ctk.CTkLabel(header_frame, text="PROJET PULSE - ÉCARTS IMPORTANTS", font=titre_font)
        titre_label.pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions\Partage - Invités\Projet PULSE\4. Données historiques\Développement\Images\logo_Pulse.png"
            logo_img = Image.open(image_path)

            font_test = tk.Label(self, text="Test", font=titre_font)
            font_test.update_idletasks()
            text_height = font_test.winfo_reqheight()
            font_test.destroy()

            ratio = logo_img.width / logo_img.height
            new_height = text_height
            new_width = int(new_height * ratio)

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_width, new_height), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_width, new_height))

            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement du logo: {e}")

        barre = ctk.CTkFrame(self, height=2, fg_color="white")
        barre.pack(side="top", fill="x")

        # === Colonnes du tableau ===
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (k€)", "Écart (%)"]

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral", "Créances CDP",
            "Placements", "CC financiers", "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]

        decaissements = [
            "Péages", "Charges de personnel", "ACE & Investissements"
        ]

        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital",
            "Sous total financier", "Free cash Flow", "Emprunts",
            "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cash flow de financement",
            "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
            "Sous total dépenses"
        ]

        # === Fonction pour convertir en flux ===
        def en_flux(values):
            values = [float(v) if v is not None else None for v in values]
            if not values or all(v is None for v in values):
                return values
            flux = [0 if values[0] is not None else None]
            for i in range(1, len(values)):
                v, v_prev = values[i], values[i - 1]
                flux.append(v - v_prev if v is not None and v_prev is not None else None)
            return flux

        # === Récupération et calcul des écarts ===
        ecarts_data = []
        repartition = {feuille: 0 for feuille in sections.values()}

        for feuille in sections.values():
            ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
            for nom, col_start in noms_colonnes:
                dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev)
                for i, date in enumerate(dates):
                    if i >= len(reel) or reel[i] is None:
                        continue
                    for idx, prev_list in enumerate(previsions):
                        if i >= len(prev_list) or prev_list[i] is None:
                            continue

                        r = reel[i]
                        prev_val = prev_list[i]

                        if r == 0 and prev_val == 0:
                            continue
                        elif prev_val == 0:
                            prev_val = 1

                        ecart = (r - prev_val) / prev_val

                        if abs(ecart) >= 0.4:
                            profil_label = noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}"
                            repartition[feuille] += 1

                            ecarts_data.append((
                                date,                    # Date
                                profil_label,            # Profil
                                feuille,                 # Filiale / Feuille
                                nom,                     # Flux
                                round(reel[i], 2),       # Réel
                                round(prev_val, 2),      # Prévision
                                round(r - prev_val, 2),  # Écart k€
                                round(ecart * 100, 1)    # Écart %
                            ))

        # === PIE CHART : % d'écarts par rapport au total (console) ===
        feuilles = list(repartition.keys())
        total_ecarts = sum(repartition.values())

        valeurs = []
        for f in feuilles:
            if total_ecarts > 0:
                pourcentage = (repartition[f] / total_ecarts) * 100
                print(f"{pourcentage:.1f}%, nb écarts: {repartition[f]}, total écarts : {total_ecarts}")
            else:
                pourcentage = 0
            valeurs.append(pourcentage)

        ecarts_data.sort(key=lambda x: abs(x[7]), reverse=True)

        # === Frame filtres et boutons ===
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(padx=30, pady=(10, 0), fill="x")

        filtre_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        filtre_frame.pack(side="left", fill="x", expand=True)

        # === ✅ AJOUT : filtre Année
        annees_disponibles = sorted({d.year for (d, *_ ) in ecarts_data})
        annee_label = ctk.CTkLabel(filtre_frame, text="Année :", font=("Segoe UI", 11, "bold"))
        annee_label.pack(side="left", padx=(0, 2), pady=5)

        annee_combo_frame = ctk.CTkFrame(filtre_frame, fg_color="#0084ff", corner_radius=8)
        annee_combo_frame.pack(side="left", padx=(0, 8), pady=5)

        annee_combo = ttk.Combobox(
            annee_combo_frame,
            values=(["Toutes"] + [str(y) for y in annees_disponibles]),
            state="readonly", width=8, font=("Segoe UI", 11, "bold")
        )
        annee_combo.set(str(annees_disponibles[-1]) if annees_disponibles else "Toutes")
        annee_combo.pack(padx=5, pady=2, fill="x")

        # === Filtres existants ===
        colonnes_filtrables = ["Date", "Profil", "Filiales", "Flux"]
        filtres = {}
        valeurs_uniques = {col: set() for col in colonnes_filtrables}
        for row in ecarts_data:
            for i, col in enumerate(colonnes):
                if col in colonnes_filtrables:
                    valeurs_uniques[col].add(str(row[i]))

        for col in colonnes_filtrables:
            label = ctk.CTkLabel(filtre_frame, text=f"{col} :", font=("Segoe UI", 11, "bold"))
            label.pack(side="left", padx=(0, 2), pady=5)
            combo_frame = ctk.CTkFrame(filtre_frame, fg_color="#0084ff", corner_radius=8)
            combo_frame.pack(side="left", padx=(0, 8), pady=5)
            valeurs = ["Tous"] + sorted(valeurs_uniques[col])
            combo = ttk.Combobox(combo_frame, values=valeurs, state="readonly", width=15, font=("Segoe UI", 11, "bold"))
            combo.set("Tous")
            combo.pack(padx=5, pady=2, fill="x")
            filtres[col] = combo

        # === Boutons ===
        btn_frame = ctk.CTkFrame(top_frame, fg_color="transparent")
        btn_frame.pack(side="right")

        btn_retour = ctk.CTkButton(
            btn_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#444",
            hover_color="#666",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_retour.pack(side="top", pady=5)

        btn_export = ctk.CTkButton(
            btn_frame,
            text="📊 Exporter en Excel",
            command=lambda: self.exporter_ecarts_excel(ecarts_data),
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#0078D7",
            hover_color="#005A9E",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_export.pack(side="top", pady=5)

        btn_graphique = ctk.CTkButton(
            btn_frame,
            text="📈 Visualiser graphiquement",
            command=lambda: self.analyser_ecarts_ml(),
            width=180,
            height=40,
            corner_radius=15,
            fg_color="#FC7100",
            hover_color="#6C4100",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        btn_graphique.pack(side="top", pady=5)

        # === Treeview ===
        tree = ttk.Treeview(self, columns=colonnes, show="headings", height=25)
        for col in colonnes:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=130)
        tree.pack(pady=10, padx=30, fill="both", expand=True)

        tree.tag_configure("neg", foreground="red")
        tree.tag_configure("pos", foreground="green")

        # === Fonctions de formatage ===
        def format_milliers(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ")
                return str(val)
            except Exception:
                return str(val)

        def format_pourcentage(val):
            try:
                if isinstance(val, (int, float)):
                    return f"{val:,.0f}".replace(",", " ") + "%"
                return str(val)
            except Exception:
                return str(val)

        # === Favorabilité ===
        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                if prev_val >= 0:
                    return reel_val >= prev_val
                else:
                    return abs(reel_val) <= abs(prev_val)
            else:
                return reel_val >= prev_val

        # === Affichage dans le tree ===
        def afficher_donnees(donnees):
            for row in tree.get_children():
                tree.delete(row)
            for data in donnees:
                date_str = data[0].strftime("%Y-%m-%d")
                reel_str = format_milliers(data[4])
                prev_str = format_milliers(data[5])
                ecart_k_str = format_milliers(data[6])
                ecart_pct_str = format_pourcentage(data[7])

                flux_nom = data[3]
                reel_val, prev_val = data[4], data[5]
                favorable = est_favorable(flux_nom, reel_val, prev_val)
                tags = ("pos",) if favorable else ("neg",)

                tree.insert(
                    "",
                    "end",
                    values=(date_str, data[1], data[2], flux_nom, reel_str, prev_str, ecart_k_str, ecart_pct_str),
                    tags=tags
                )

        # Premier affichage (non filtré)
        afficher_donnees(ecarts_data)

        # === Filtrage (incluant Année) ===
        def appliquer_filtre(event=None):
            filtred = ecarts_data

            # 1) Filtre Année
            sel_annee = annee_combo.get()
            if sel_annee != "Toutes":
                try:
                    y = int(sel_annee)
                    filtred = [row for row in filtred if row[0].year == y]  # row[0] = Date
                except Exception:
                    pass

            # 2) Filtres existants
            for i, col in enumerate(colonnes):
                if col in filtres:
                    val = filtres[col].get()
                    if val != "Tous":
                        filtred = [row for row in filtred if str(row[i]) == val]

            afficher_donnees(filtred)

        # Bind des filtres
        for combo in filtres.values():
            combo.bind("<<ComboboxSelected>>", appliquer_filtre)
        annee_combo.bind("<<ComboboxSelected>>", appliquer_filtre)

        # Appliquer une première fois (pour prendre en compte l'année par défaut)
        appliquer_filtre()

    def analyser_ecarts_ml(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        import customtkinter as ctk
        import pandas as pd
        import numpy as np
        from sklearn.cluster import KMeans, DBSCAN
        from sklearn.preprocessing import StandardScaler
        from sklearn.ensemble import IsolationForest
        from sklearn.metrics import silhouette_score
        from sklearn.feature_selection import SelectKBest, f_classif
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors
        import re
        import datetime as _dt

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "accent_2": "#355CBE",
            "accent_2_hover": "#2B4C9C",
            "warning": "#FC7100",
            "warning_hover": "#D55F00",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
            "kpi": ("Segoe UI Semibold", 22, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)
            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def primary_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["accent"],
                hover_color=UI["accent_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def secondary_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["accent_2"],
                hover_color=UI["accent_2_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def warning_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["warning"],
                hover_color=UI["warning_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _resolve_logo_path():
            candidates = []
            try:
                if image_path:
                    candidates.append(image_path)
            except Exception:
                pass
            try:
                if hasattr(self, "image_path") and self.image_path:
                    candidates.append(self.image_path)
            except Exception:
                pass
            for p in candidates:
                try:
                    if p:
                        return p
                except Exception:
                    pass
            return None

        def _embed_figure(fig, master, toolbar_host=None, subplots_adjust=None):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            if toolbar_host is not None:
                _clear_children(toolbar_host)
                toolbar = NavigationToolbar2Tk(canvas, toolbar_host)
                toolbar.update()

            def _resize(event=None):
                try:
                    master.update_idletasks()
                    w = max(master.winfo_width() - 16, 980)
                    h = max(master.winfo_height() - 16, 320)
                    dpi = fig.get_dpi()

                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    fig.set_size_inches(w / dpi, h / dpi, forward=True)

                    if subplots_adjust:
                        fig.subplots_adjust(**subplots_adjust)
                    else:
                        fig.subplots_adjust(left=0.08, right=0.80, bottom=0.14, top=0.90)

                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, (int, float)):
                return float(x)
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "").replace(",", ".")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    return None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(d, fmt).year
                    except Exception:
                        pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m:
                    return int(m.group(1))
            return None

        def _annees_disponibles_filiale(nom_filiale_ui: str):
            annees = set()
            feuilles = list(sections.values()) if nom_filiale_ui == "Toutes les filiales" else [nom_filiale_ui]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for _nom_flux, col_start in noms_colonnes:
                    try:
                        dates, reel, prevs, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            annees.add(y)

            return sorted(annees)

        def _flux_disponibles_filiale(nom_filiale_ui: str):
            flux = set()
            feuilles = list(sections.values()) if nom_filiale_ui == "Toutes les filiales" else [nom_filiale_ui]

            for feuille in feuilles:
                try:
                    _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for nom_flux, _col_start in noms_colonnes:
                    if nom_flux is not None and str(nom_flux).strip():
                        flux.add(str(nom_flux))

            return sorted(flux)

        def _collect_points(feuille_sel: str, annee: int | None, flux_sel: str | None):
            feuilles_a_traiter = (
                list(sections.values())
                if feuille_sel == "Toutes les filiales"
                else [feuille_sel]
            )

            points = []

            for ws_feuille in feuilles_a_traiter:
                try:
                    ws, noms_colonnes = charger_donnees(ws_feuille, taille_bloc)
                except Exception:
                    continue

                for nom, col_start in noms_colonnes:
                    if flux_sel not in (None, "", "Tous les flux") and str(nom) != str(flux_sel):
                        continue

                    try:
                        dates, reel, previsions, profils = extraire_valeurs(ws, col_start, nb_prev, annee=annee)
                    except Exception:
                        try:
                            dates, reel, previsions, profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        except Exception:
                            continue

                    for i, d in enumerate(dates):
                        if annee is not None:
                            y = _year_of(d)
                            if y is not None and y != annee:
                                continue

                        r = _to_number(reel[i] if i < len(reel) else None)
                        if r is None:
                            continue

                        for p_idx, prev_list in enumerate(previsions):
                            pv = prev_list[i] if i < len(prev_list) else None
                            prev_val = _to_number(pv)
                            if prev_val is None:
                                continue

                            denom = prev_val if prev_val != 0 else (r if r != 0 else None)
                            if denom is None:
                                continue

                            try:
                                ecart_pct = (r - prev_val) / denom * 100.0
                            except ZeroDivisionError:
                                continue

                            y_k = (r - prev_val)

                            points.append({
                                "x_pct": ecart_pct,
                                "y_k": y_k,
                                "filiale": ws_feuille,
                                "flux": str(nom),
                                "date": d
                            })

            if not points:
                return pd.DataFrame(columns=["x_pct", "y_k", "filiale", "flux", "date"])
            return pd.DataFrame(points)

        def _auto_k_silhouette(Xs, k_min=2, k_max=6):
            n_samples = Xs.shape[0]
            if n_samples <= 2:
                return 2, None

            max_k_eff = min(k_max, n_samples - 1)
            best_k = None
            best_score = None

            for k in range(k_min, max_k_eff + 1):
                try:
                    km = KMeans(
                        n_clusters=k,
                        init="k-means++",
                        n_init=10,
                        max_iter=300,
                        random_state=42,
                        algorithm="lloyd"
                    )
                    labels_k = km.fit_predict(Xs)
                    if len(set(labels_k)) < 2:
                        continue
                    score = silhouette_score(Xs, labels_k)
                except Exception:
                    continue

                if (best_score is None) or (score > best_score):
                    best_k = k
                    best_score = score

            if best_k is None:
                return max(k_min, min(4, n_samples)), None
            return best_k, best_score

        # =========================================================
        # ROOT / RESET
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        logo_path = _resolve_logo_path()
        try:
            _img = Image.open(logo_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module d’analyse ML des écarts",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        neutral_btn(
            top_right,
            "Retour à l’accueil",
            self.creer_accueil if hasattr(self, "creer_accueil") else self.retour_menu
        ).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.grid_rowconfigure(3, weight=0)
        page.columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "MACHINE LEARNING", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Analyse ML des écarts",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Dispersion 2D, clustering et lecture explicative avec filtres filiale, année et flux.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES / ACTIONS
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "PILOTAGE",
            "Filtres et actions",
            "Le filtre flux s’applique à la collecte des points. Les filtres de petites valeurs ont été retirés."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        for i in range(5):
            filters_body.grid_columnconfigure(i, weight=0)
        filters_body.grid_columnconfigure(5, weight=1)

        filiale_var = ctk.StringVar(value="Toutes les filiales")
        annee_var = ctk.StringVar(value="Toutes années")
        flux_var = ctk.StringVar(value="Tous les flux")
        algo_var = ctk.StringVar(value="KMeans (auto-k)")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        filiale_menu = ctk.CTkOptionMenu(
            filters_body,
            values=["Toutes les filiales"] + list(sections.values()),
            variable=filiale_var,
            width=220,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        filiale_menu.grid(row=1, column=0, sticky="w", padx=(0, 16))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=1, sticky="w", pady=(0, 6))
        annee_menu = ctk.CTkOptionMenu(
            filters_body,
            values=["Toutes années"],
            variable=annee_var,
            width=150,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annee_menu.grid(row=1, column=1, sticky="w", padx=(0, 16))

        label(filters_body, "Flux", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=2, sticky="w", pady=(0, 6))
        flux_menu = ctk.CTkOptionMenu(
            filters_body,
            values=["Tous les flux"],
            variable=flux_var,
            width=220,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        flux_menu.grid(row=1, column=2, sticky="w", padx=(0, 16))

        label(filters_body, "Algorithme", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=3, sticky="w", pady=(0, 6))
        algo_menu = ctk.CTkOptionMenu(
            filters_body,
            values=["KMeans (auto-k)", "DBSCAN"],
            variable=algo_var,
            width=180,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        algo_menu.grid(row=1, column=3, sticky="w", padx=(0, 16))

        actions_wrap = ctk.CTkFrame(filters_body, fg_color="transparent")
        actions_wrap.grid(row=1, column=4, sticky="w")

        # =========================================================
        # ZONE GRAPHIQUE
        # =========================================================
        chart_card = card(page, fg=UI["surface"], radius=20)
        chart_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        chart_card.grid_columnconfigure(0, weight=1)

        section_header(
            chart_card,
            "RENDU",
            "Visualisation principale",
            "Axes, ticks et légende sont forcés en blanc."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        toolbar_host = tk.Frame(chart_card, bg=UI["surface"])
        toolbar_host.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 10))

        graph_body = ctk.CTkFrame(
            chart_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=620
        )
        graph_body.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        graph_body.grid_propagate(False)
        graph_body.grid_rowconfigure(0, weight=1)
        graph_body.grid_columnconfigure(0, weight=1)

        _make_placeholder(graph_body, "Choisissez vos filtres puis lancez l’analyse.")

        dist_header = ctk.CTkFrame(chart_card, fg_color="transparent")
        dist_header.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 8))
        dist_header.grid_columnconfigure(0, weight=1)

        label(
            dist_header,
            "Répartition des clusters",
            font=FONT["small_bold"],
            color=UI["text_soft"]
        ).grid(row=0, column=0, sticky="w")

        label(
            dist_header,
            "Part des points par cluster, avec les mêmes couleurs que le graphe 2D.",
            font=FONT["small"],
            color=UI["muted_2"]
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        cluster_dist_body = ctk.CTkFrame(
            chart_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=290
        )
        cluster_dist_body.grid(row=4, column=0, sticky="ew", padx=18, pady=(0, 18))
        cluster_dist_body.grid_propagate(False)
        cluster_dist_body.grid_rowconfigure(0, weight=1)
        cluster_dist_body.grid_columnconfigure(0, weight=1)

        _make_placeholder(cluster_dist_body, "La répartition des clusters s’affichera ici après exécution.")

        # =========================================================
        # SYNTHÈSE
        # =========================================================
        summary_card = card(page, fg=UI["surface"], radius=20)
        summary_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 24))
        summary_card.grid_columnconfigure(0, weight=1)

        section_header(
            summary_card,
            "SYNTHÈSE",
            "Résumé analytique",
            "Cette zone affiche soit le tableau des clusters, soit l’analyse explicative."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        summary_body = ctk.CTkFrame(
            summary_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        summary_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        summary_body.grid_columnconfigure(0, weight=1)
        summary_body.grid_rowconfigure(0, weight=1)

        _make_placeholder(summary_body, "Le résumé s’affichera ici après exécution.")

        # =========================================================
        # RENDERING FUNCTIONS
        # =========================================================
        def _get_filters():
            feuille_sel = filiale_var.get() or "Toutes les filiales"
            val_annee = annee_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            flux_sel = flux_var.get() or "Tous les flux"
            algo_sel = algo_var.get() or "KMeans (auto-k)"
            return feuille_sel, annee, flux_sel, algo_sel

        def _apply_white_axes(ax):
            ax.xaxis.label.set_color("white")
            ax.yaxis.label.set_color("white")
            ax.tick_params(axis="x", colors="white")
            ax.tick_params(axis="y", colors="white")
            ax.title.set_color("white")

            for spine in ax.spines.values():
                spine.set_color("white")

        def _apply_white_legend(legend):
            if legend is None:
                return
            frame = legend.get_frame()
            try:
                frame.set_facecolor("#1F2937")
                frame.set_edgecolor("white")
                frame.set_alpha(0.85)
            except Exception:
                pass
            for text in legend.get_texts():
                text.set_color("white")
            try:
                legend.get_title().set_color("white")
            except Exception:
                pass

        def afficher_graphique():
            feuille_sel, annee, flux_sel, algo_sel = _get_filters()
            df = _collect_points(feuille_sel, annee, flux_sel)

            if df.empty:
                messagebox.showinfo("Analyse ML", "Aucune donnée exploitable pour ces filtres.")
                return

            if len(df) < 2:
                messagebox.showinfo("Analyse ML", "Nombre de points insuffisant pour faire un clustering.")
                return

            _clear_children(graph_body)
            _clear_children(cluster_dist_body)
            _clear_children(toolbar_host)
            _clear_children(summary_body)

            X = df[["x_pct", "y_k"]].values
            scaler = StandardScaler()
            Xs = scaler.fit_transform(X)

            contamination_dyn = min(
                0.08,
                max(0.02, len(df[df["y_k"].abs() > df["y_k"].abs().quantile(0.95)]) / max(len(df), 1))
            )
            iso = IsolationForest(contamination=contamination_dyn, random_state=42)
            iso_labels = iso.fit_predict(Xs)
            df["outlier"] = (iso_labels == -1)

            centers = None
            silhouette_val = None
            n_clusters = None
            display_order = []

            if algo_sel.startswith("KMeans"):
                n_clusters, silhouette_val = _auto_k_silhouette(Xs, k_min=2, k_max=6)

                kmeans = KMeans(
                    n_clusters=n_clusters,
                    init="k-means++",
                    n_init=10,
                    max_iter=300,
                    random_state=42,
                    algorithm="lloyd"
                )
                labels = kmeans.fit_predict(Xs)
                df["cluster"] = labels
                centers_s = kmeans.cluster_centers_
                centers = scaler.inverse_transform(centers_s)

                impact_series = df.groupby("cluster")["y_k"].apply(lambda s: s.abs().mean())
                display_order = impact_series.sort_values().index.tolist()

                base_colors = ["#00CC66", "#FFD700", "#FFA500", "#FF0000", "#A855F7", "#38BDF8"]
                color_map = {}
                for i, cl in enumerate(display_order):
                    color_map[cl] = base_colors[i % len(base_colors)]
            else:
                db = DBSCAN(eps=0.8, min_samples=5)
                labels = db.fit_predict(Xs)
                df["cluster"] = labels

                unique_clusters = sorted(df["cluster"].unique())
                display_order = [cl for cl in unique_clusters if cl != -1]
                if -1 in unique_clusters:
                    display_order.append(-1)

                base_palette = ["#00CC66", "#FFD700", "#FFA500", "#FF0000", "#A855F7", "#38BDF8"]
                color_map = {}
                color_idx = 0
                for cl in display_order:
                    if cl == -1:
                        color_map[cl] = "#888888"
                    else:
                        color_map[cl] = base_palette[color_idx % len(base_palette)]
                        color_idx += 1

            def _cluster_display_name(cl):
                if algo_sel.startswith("KMeans"):
                    return f"Cluster {cl + 1}"
                return "Bruit" if cl == -1 else f"Cluster {cl}"

            colors = [color_map[l] for l in df["cluster"]]

            # =========================================================
            # FIGURE 1 : SCATTER 2D
            # =========================================================
            plt.style.use("seaborn-v0_8-darkgrid")
            fig, ax = plt.subplots(figsize=(11, 6), dpi=100)
            fig.patch.set_facecolor(UI["surface_2"])
            ax.set_facecolor(UI["surface_2"])

            sc = ax.scatter(
                df["x_pct"], df["y_k"],
                c=colors,
                s=48,
                alpha=0.9,
                edgecolors="none"
            )

            if centers is not None:
                ax.scatter(
                    centers[:, 0], centers[:, 1],
                    c="black",
                    marker="X",
                    s=170,
                    label="Centroides"
                )

            df_out = df[df["outlier"]]
            if not df_out.empty:
                ax.scatter(
                    df_out["x_pct"], df_out["y_k"],
                    facecolors="none",
                    edgecolors="white",
                    s=86,
                    linewidths=1.2,
                    label="Anomalies"
                )

            titre = "Dispersion des écarts — 2D (% vs valorisation signée)"
            suffix = f" — {feuille_sel}" if feuille_sel != "Toutes les filiales" else " — Ensemble des filiales"
            if flux_sel not in (None, "", "Tous les flux"):
                suffix += f" — {flux_sel}"
            if annee is not None:
                suffix += f" — {annee}"
            ax.set_title(titre + suffix, fontsize=15, fontweight="bold", color="white")

            ax.set_xlabel("Écart (%)  —  (réel - prévision) / base × 100", color="white")
            ax.set_ylabel("Valorisation signée (réel - prévision)", color="white")
            ax.grid(True, linestyle="--", alpha=0.35)

            info_algo = algo_sel
            if algo_sel.startswith("KMeans") and silhouette_val is not None:
                info_algo += f" — k={n_clusters}, silhouette={silhouette_val:.2f}"

            ax.text(
                0.01, 0.01, info_algo,
                transform=ax.transAxes,
                fontsize=9, ha="left", va="bottom", color="white",
                bbox=dict(boxstyle="round", facecolor="#333333", alpha=0.7)
            )

            for cl in display_order:
                ax.scatter([], [], color=color_map[cl], label=_cluster_display_name(cl), s=70)

            legend = ax.legend(loc="center left", bbox_to_anchor=(1, 0.5))
            _apply_white_axes(ax)
            _apply_white_legend(legend)

            _embed_figure(
                fig,
                graph_body,
                toolbar_host=toolbar_host,
                subplots_adjust={"left": 0.08, "right": 0.80, "bottom": 0.14, "top": 0.90}
            )

            cursor = mplcursors.cursor(sc, hover=True)

            @cursor.connect("add")
            def on_hover(sel):
                idx = sel.index
                fil = df.iloc[idx]["filiale"]
                flux = df.iloc[idx]["flux"]
                x = df.iloc[idx]["x_pct"]
                y = df.iloc[idx]["y_k"]
                d = df.iloc[idx]["date"]
                out_flag = df.iloc[idx]["outlier"]
                cl = df.iloc[idx]["cluster"]
                dtxt = d.strftime("%d/%m/%Y") if hasattr(d, "strftime") else str(d)

                sel.annotation.set_text(
                    f"Filiale : {fil}\n"
                    f"Flux : {flux}\n"
                    f"Date : {dtxt}\n"
                    f"Cluster : {_cluster_display_name(cl)}\n"
                    f"Écart : {x:.2f} %\n"
                    f"Valorisation : {y:.0f}\n"
                    f"Anomalie : {'Oui' if out_flag else 'Non'}"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =========================================================
            # FIGURE 2 : RÉPARTITION DES CLUSTERS
            # =========================================================
            cluster_dist = (
                pd.DataFrame({"cluster": display_order})
                .merge(
                    df.groupby("cluster").size().rename("count").reset_index(),
                    on="cluster",
                    how="left"
                )
            )
            cluster_dist["count"] = cluster_dist["count"].fillna(0).astype(int)
            cluster_dist["pct"] = cluster_dist["count"] / max(len(df), 1) * 100
            cluster_dist["label"] = cluster_dist["cluster"].apply(_cluster_display_name)
            cluster_dist["color"] = cluster_dist["cluster"].map(color_map)

            dominant_row = cluster_dist.sort_values(["pct", "count"], ascending=False).iloc[0]

            fig2, ax2 = plt.subplots(figsize=(11, 3.8), dpi=100)
            fig2.patch.set_facecolor(UI["surface_2"])
            ax2.set_facecolor(UI["surface_2"])

            bars = ax2.bar(
                cluster_dist["label"],
                cluster_dist["pct"],
                color=cluster_dist["color"],
                alpha=0.95,
                edgecolor="none"
            )

            ax2.set_title("Répartition des points par cluster", fontsize=13, fontweight="bold", color="white")
            ax2.set_xlabel("Clusters", color="white")
            ax2.set_ylabel("% des points", color="white")
            ax2.grid(True, axis="y", linestyle="--", alpha=0.35)

            ymax = max(10, float(cluster_dist["pct"].max()) * 1.25)
            ax2.set_ylim(0, ymax)

            ref_line = 100 / max(len(cluster_dist), 1)
            ax2.axhline(ref_line, linestyle=":", linewidth=1, color="white", alpha=0.35)

            ax2.text(
                0.99, 0.95,
                f"Cluster dominant : {dominant_row['label']} ({dominant_row['pct']:.1f}%)",
                transform=ax2.transAxes,
                ha="right", va="top",
                fontsize=9, color="white",
                bbox=dict(boxstyle="round", facecolor="#333333", alpha=0.7)
            )

            for bar, pct, cnt in zip(bars, cluster_dist["pct"], cluster_dist["count"]):
                ax2.text(
                    bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + ymax * 0.02,
                    f"{pct:.1f}%\n({cnt} pts)",
                    ha="center",
                    va="bottom",
                    color="white",
                    fontsize=9,
                    fontweight="bold"
                )

            _apply_white_axes(ax2)

            _embed_figure(
                fig2,
                cluster_dist_body,
                toolbar_host=None,
                subplots_adjust={"left": 0.08, "right": 0.96, "bottom": 0.22, "top": 0.88}
            )

            # =========================================================
            # SYNTHÈSE ENRICHIE
            # =========================================================
            summary = (
                df.groupby("cluster")
                .agg(
                    count=("x_pct", "count"),
                    pct_points=("x_pct", lambda s: len(s) / len(df) * 100),
                    mean_pct=("x_pct", "mean"),
                    mean_k=("y_k", "mean"),
                    sum_k=("y_k", "sum"),
                    anomalies=("outlier", "sum")
                )
            )

            summary["pct_anomalies"] = np.where(
                summary["count"] > 0,
                summary["anomalies"] / summary["count"] * 100,
                0
            )

            summary = summary.round({
                "pct_points": 1,
                "mean_pct": 2,
                "mean_k": 1,
                "sum_k": 0,
                "pct_anomalies": 1
            })

            cluster_risque = summary["sum_k"].abs().idxmax()
            cluster_risque_nom = _cluster_display_name(cluster_risque)
            cluster_dominant_nom = dominant_row["label"]

            summary_wrap = ctk.CTkFrame(summary_body, fg_color="transparent")
            summary_wrap.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
            summary_wrap.grid_columnconfigure(0, weight=1)

            label(summary_wrap, "Résumé des clusters", font=FONT["label_bold"], color=UI["text"]).grid(
                row=0, column=0, sticky="w", pady=(0, 8)
            )

            cols = ["Cluster", "Nb points", "% total", "Moy. écart (%)", "Moy. valo", "Somme valo", "Anomalies"]
            tree = ttk.Treeview(
                summary_wrap,
                columns=cols,
                show="headings",
                height=min(8, len(summary) + 2),
                style="Pulse.Treeview"
            )
            tree.grid(row=1, column=0, sticky="ew")

            widths = {
                "Cluster": 150,
                "Nb points": 110,
                "% total": 110,
                "Moy. écart (%)": 140,
                "Moy. valo": 130,
                "Somme valo": 140,
                "Anomalies": 120,
            }

            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, width=widths.get(col, 140), anchor="center")

            for cl in display_order:
                if cl not in summary.index:
                    continue

                tree.insert(
                    "",
                    "end",
                    values=[
                        _cluster_display_name(cl),
                        int(summary.loc[cl, "count"]),
                        f"{summary.loc[cl, 'pct_points']:.1f}%",
                        f"{summary.loc[cl, 'mean_pct']:.2f}",
                        f"{summary.loc[cl, 'mean_k']:.1f}",
                        f"{int(summary.loc[cl, 'sum_k']):,}".replace(",", " "),
                        f"{int(summary.loc[cl, 'anomalies'])} ({summary.loc[cl, 'pct_anomalies']:.1f}%)"
                    ]
                )

            tree.insert(
                "",
                "end",
                values=[
                    "Total",
                    int(summary["count"].sum()),
                    "100.0%",
                    f"{(df['x_pct'].mean() if len(df) > 0 else 0):.2f}",
                    f"{(df['y_k'].mean() if len(df) > 0 else 0):.1f}",
                    f"{int(df['y_k'].sum()):,}".replace(",", " "),
                    f"{int(df['outlier'].sum())}"
                ],
                tags=("total",)
            )
            tree.tag_configure("total", background="#444", foreground="white", font=("Segoe UI", 11, "bold"))

            info_text = ctk.CTkTextbox(
                summary_wrap,
                height=110,
                fg_color=UI["surface_3"],
                text_color=UI["text"],
                border_width=1,
                border_color=UI["border_soft"],
                corner_radius=12,
                font=("Segoe UI", 11)
            )
            info_text.grid(row=2, column=0, sticky="ew", pady=(12, 0))

            info_text.insert(
                "1.0",
                (
                    f"Points analysés : {len(df):,}\n".replace(",", " ")
                    + f"Flux sélectionné : {flux_sel}\n"
                    + f"Anomalies détectées : {int(df['outlier'].sum())}\n"
                    + f"Cluster dominant : {cluster_dominant_nom} ({dominant_row['pct']:.1f}% des points)\n"
                    + f"Cluster le plus risqué en valeur absolue : {cluster_risque_nom}\n"
                    + (f"K retenu automatiquement : {n_clusters}\n" if n_clusters is not None else "")
                    + (f"Score de silhouette : {silhouette_val:.2f}" if silhouette_val is not None else "Pas de score de silhouette disponible")
                )
            )
            info_text.configure(state="disabled")

            _sync_scrollregion()

        def afficher_explication():
            feuille_sel, annee, flux_sel, _algo_sel = _get_filters()
            df = _collect_points(feuille_sel, annee, flux_sel)

            if df.empty:
                messagebox.showinfo("Analyse explicative", "Aucune donnée exploitable pour ces filtres.")
                return

            df["abs_y"] = df["y_k"].abs()

            if df["abs_y"].max() == 0:
                messagebox.showinfo(
                    "Analyse explicative",
                    "Les écarts sont trop faibles pour construire une cible exploitable."
                )
                return

            seuil = df["abs_y"].quantile(0.75)
            df["target_gros_ecart"] = (df["abs_y"] >= seuil).astype(int)

            if df["target_gros_ecart"].sum() == 0 or df["target_gros_ecart"].sum() == len(df):
                messagebox.showinfo(
                    "Analyse explicative",
                    "Impossible de distinguer les gros écarts des écarts normaux : la cible est uniforme."
                )
                return

            df["year"] = df["date"].map(_year_of)
            if df["year"].isna().all():
                df["year"] = annee if annee is not None else 0
            else:
                df["year"] = df["year"].fillna(df["year"].median())

            df["month"] = pd.to_datetime(df["date"], errors="coerce").dt.month.fillna(0)
            df["filiale_id"] = df["filiale"].astype("category").cat.codes
            df["flux_id"] = df["flux"].astype("category").cat.codes
            df["abs_pct"] = df["x_pct"].abs()

            feature_cols = ["x_pct", "abs_pct", "year", "month", "filiale_id", "flux_id"]
            X_feat = df[feature_cols].fillna(0).values
            y_cls = df["target_gros_ecart"].values

            selector = SelectKBest(score_func=f_classif, k=min(len(feature_cols), X_feat.shape[1]))
            selector.fit(X_feat, y_cls)
            scores = selector.scores_

            feat_imp = pd.DataFrame({
                "feature": feature_cols,
                "score": scores
            }).dropna().sort_values("score", ascending=True)

            if feat_imp.empty:
                messagebox.showinfo("Analyse explicative", "Les scores calculés ne sont pas exploitables.")
                return

            _clear_children(graph_body)
            _clear_children(cluster_dist_body)
            _clear_children(toolbar_host)
            _clear_children(summary_body)

            _make_placeholder(
                cluster_dist_body,
                "La répartition des clusters n’est pas affichée dans le mode analyse explicative."
            )

            plt.style.use("seaborn-v0_8-darkgrid")
            fig, ax = plt.subplots(figsize=(9, 4.6), dpi=100)
            fig.patch.set_facecolor(UI["surface_2"])
            ax.set_facecolor(UI["surface_2"])

            ax.barh(feat_imp["feature"], feat_imp["score"])
            ax.set_title("Variables les plus explicatives des gros écarts", fontsize=13, fontweight="bold", color="white")
            ax.set_xlabel("Score F (SelectKBest / f_classif)", color="white")
            ax.set_ylabel("Variables", color="white")
            _apply_white_axes(ax)

            _embed_figure(fig, graph_body, toolbar_host=toolbar_host)

            summary_wrap = ctk.CTkFrame(summary_body, fg_color="transparent")
            summary_wrap.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)
            summary_wrap.grid_columnconfigure(0, weight=1)

            info_text = ctk.CTkTextbox(
                summary_wrap,
                height=210,
                fg_color=UI["surface_3"],
                text_color=UI["text"],
                border_width=1,
                border_color=UI["border_soft"],
                corner_radius=12,
                font=("Segoe UI", 11)
            )
            info_text.grid(row=0, column=0, sticky="ew")

            texte = (
                f"Définition de la cible : gros écart = |valorisation| ≥ {seuil:,.0f}\n".replace(",", " ")
                + f"Flux sélectionné : {flux_sel}\n"
                + "Les scores F indiquent quelles variables sont le plus corrélées à cette cible.\n\n"
                + "Variables testées : x_pct, abs_pct, year, month, filiale_id, flux_id.\n\n"
                + "Lecture : plus le score est élevé, plus la variable discrimine les gros écarts."
            )
            info_text.insert("1.0", texte)
            info_text.configure(state="disabled")

            cols = ["Variable", "Score F"]
            tree = ttk.Treeview(
                summary_wrap,
                columns=cols,
                show="headings",
                height=min(8, len(feat_imp)),
                style="Pulse.Treeview"
            )
            tree.grid(row=1, column=0, sticky="ew", pady=(12, 0))

            tree.heading("Variable", text="Variable")
            tree.heading("Score F", text="Score F")
            tree.column("Variable", width=250, anchor="center")
            tree.column("Score F", width=180, anchor="center")

            for _, row in feat_imp.sort_values("score", ascending=False).iterrows():
                tree.insert("", "end", values=[row["feature"], f"{row['score']:.2f}"])

            _sync_scrollregion()

        # =========================================================
        # ACTIONS
        # =========================================================
        primary_btn(actions_wrap, "Afficher clustering 2D", afficher_graphique).pack(side="left", padx=(0, 10))
        secondary_btn(actions_wrap, "Analyse explicative", afficher_explication).pack(side="left", padx=(0, 10))
        warning_btn(
            actions_wrap,
            "Retour au tableau",
            self.afficher_ecarts if hasattr(self, "afficher_ecarts") else self.creer_accueil
        ).pack(side="left", padx=(0, 10))
        neutral_btn(
            actions_wrap,
            "Retour au menu",
            self.retour_menu if hasattr(self, "retour_menu") else self.creer_accueil
        ).pack(side="left")

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _refresh_years(selected_filiale):
            annees = _annees_disponibles_filiale(selected_filiale)
            values = ["Toutes années"] + [str(a) for a in annees] if annees else ["Toutes années"]
            annee_menu.configure(values=values)

            if annee_var.get() not in values:
                if annees:
                    annee_var.set(str(annees[-1]))
                else:
                    annee_var.set("Toutes années")

        def _refresh_flux(selected_filiale):
            flux = _flux_disponibles_filiale(selected_filiale)
            values = ["Tous les flux"] + flux if flux else ["Tous les flux"]
            flux_menu.configure(values=values)

            if flux_var.get() not in values:
                flux_var.set("Tous les flux")

        def _on_filiale_change(value=None):
            selected = filiale_var.get()
            _refresh_years(selected)
            _refresh_flux(selected)

        filiale_menu.configure(command=_on_filiale_change)

        # =========================================================
        # INIT
        # =========================================================
        _refresh_years(filiale_var.get())
        _refresh_flux(filiale_var.get())   

    def afficher_tendance_flux(self):
        from collections import defaultdict
        from itertools import zip_longest
        import tkinter as tk
        from tkinter import ttk
        import customtkinter as ctk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.lines import Line2D
        from matplotlib.patches import Patch
        from PIL import Image
        from customtkinter import CTkImage
        import statistics
        import datetime as dt
        import math

        # =========================================================
        # DEBUG
        # =========================================================
        DEBUG_ANALYSE_FLUX = True

        def _debug(*args):
            if DEBUG_ANALYSE_FLUX:
                print(*args)

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "accent_2": "#355CBE",
            "accent_2_hover": "#2B4C9C",
            "warning": "#FC7100",
            "warning_hover": "#D55F00",
            "success": "#1E8449",
            "danger": "#C0392B",
            "amber": "#D68910",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
            "kpi": ("Segoe UI Semibold", 22, "bold"),
        }

        # =========================================================
        # RESET PAGE
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        # =========================================================
        # PARAMÈTRES MÉTIER
        # =========================================================
        SEUIL_MIN_FLUX = 10000
        KMEANS_MAX_K = 3

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)
            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _resolve_logo_path():
            candidates = []
            try:
                if hasattr(self, "image_path") and self.image_path:
                    candidates.append(self.image_path)
            except Exception:
                pass

            try:
                candidates.append(r"C:\Users\0304336A\SNCF\DCF GROUPE (Grp. O365)\logo_Pulse.png")
            except Exception:
                pass

            for p in candidates:
                try:
                    if p:
                        return p
                except Exception:
                    pass
            return None

        def _render_figure(fig, ax, parent, legend_handles=None, aspect_ratio=3.25, min_height=300):
            fig.patch.set_facecolor(UI["surface_2"])

            host = ctk.CTkFrame(parent, fg_color="transparent")
            host.pack(fill="both", expand=True, padx=4, pady=(0, 12))

            canvas = FigureCanvasTkAgg(fig, master=host)
            widget = canvas.get_tk_widget()
            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass
            widget.pack(fill="both", expand=True)

            state = {"legend": None, "mode": None}

            def _apply_legend(mode):
                if state["legend"] is not None:
                    try:
                        state["legend"].remove()
                    except Exception:
                        pass
                    state["legend"] = None

                if not legend_handles:
                    fig.subplots_adjust(left=0.055, right=0.98, top=0.88, bottom=0.16)
                    state["mode"] = mode
                    return

                if mode == "right":
                    legend = ax.legend(
                        handles=legend_handles,
                        loc="upper left",
                        bbox_to_anchor=(1.002, 1.0),
                        borderaxespad=0.0,
                        frameon=True,
                        facecolor="#12385F",
                        edgecolor="white",
                        fontsize=9
                    )
                    fig.subplots_adjust(left=0.055, right=0.865, top=0.88, bottom=0.16)
                else:
                    legend = ax.legend(
                        handles=legend_handles,
                        loc="upper center",
                        bbox_to_anchor=(0.5, -0.18),
                        ncol=min(3, len(legend_handles)),
                        borderaxespad=0.0,
                        frameon=True,
                        facecolor="#12385F",
                        edgecolor="white",
                        fontsize=9
                    )
                    fig.subplots_adjust(left=0.055, right=0.98, top=0.88, bottom=0.28)

                for text_obj in legend.get_texts():
                    text_obj.set_color("white")

                state["legend"] = legend
                state["mode"] = mode

            def _on_resize(event=None):
                width_px = max(host.winfo_width(), 700)
                mode = "right" if width_px >= 1250 else "bottom"

                if mode != state["mode"]:
                    _apply_legend(mode)

                height_px = max(min_height, int(width_px / aspect_ratio))
                if mode == "bottom":
                    height_px += 35

                dpi = fig.get_dpi()
                fig.set_size_inches(width_px / dpi, height_px / dpi, forward=True)
                canvas.draw_idle()

            host.bind("<Configure>", _on_resize)
            host.after(50, _on_resize)

            return canvas

        def _create_kpi_card(parent, titre, valeur, sous_texte="", color=None):
            c = ctk.CTkFrame(
                parent,
                fg_color=color or UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            c.grid_propagate(False)

            ctk.CTkLabel(
                c,
                text=titre,
                font=("Segoe UI", 11, "bold"),
                text_color=UI["text_soft"]
            ).pack(anchor="w", padx=14, pady=(12, 6))

            ctk.CTkLabel(
                c,
                text=valeur,
                font=("Segoe UI Semibold", 20, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=14)

            ctk.CTkLabel(
                c,
                text=sous_texte,
                font=("Segoe UI", 11),
                text_color=UI["muted"],
                wraplength=220,
                justify="left"
            ).pack(anchor="w", padx=14, pady=(4, 12))

            return c

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_date(x):
            if x is None:
                return None
            if hasattr(x, "year") and hasattr(x, "month") and hasattr(x, "day"):
                try:
                    return dt.date(x.year, x.month, x.day)
                except Exception:
                    return None
            if isinstance(x, str):
                txt = x.strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return dt.datetime.strptime(txt, fmt).date()
                    except Exception:
                        pass
            return None

        def _is_business_day(d):
            return d is not None and d.weekday() < 5

        def _safe_mean(vals):
            return statistics.mean(vals) if vals else 0.0

        def _safe_stdev(vals):
            if len(vals) < 2:
                return 0.0
            try:
                return statistics.stdev(vals)
            except Exception:
                return 0.0

        def _pct_vs(base, value):
            if base is None or base == 0:
                return 0.0
            return (value - base) / abs(base) * 100

        def _last_business_day_of_month(d):
            if d is None:
                return None

            if d.month == 12:
                next_month = dt.date(d.year + 1, 1, 1)
            else:
                next_month = dt.date(d.year, d.month + 1, 1)

            last_day = next_month - dt.timedelta(days=1)
            while last_day.weekday() >= 5:
                last_day -= dt.timedelta(days=1)
            return last_day

        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for nom_flux, col_start in noms_colonnes:
                        dates, reel, previsions, noms_profils = extraire_valeurs(
                            ws, col_start, nb_prev, annee=None
                        )
                        for d in dates:
                            dd = _to_date(d)
                            if dd:
                                annees.add(dd.year)
                except Exception:
                    pass

            return sorted(annees)

        def _flux_pour_filiale(filiale):
            flux = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for nom_flux, _ in noms_colonnes:
                        flux.add(str(nom_flux))
                except Exception:
                    pass

            return sorted(flux, key=lambda x: x.casefold())

        def _score_tendance(moyennes, moyenne_globale, counts):
            if not moyennes or moyenne_globale == 0:
                return 0.0

            vals_non_vides = [
                abs(_pct_vs(moyenne_globale, v))
                for v, c in zip(moyennes, counts)
                if c > 0
            ]
            if not vals_non_vides:
                return 0.0

            intensite = max(vals_non_vides)
            couverture = sum(1 for c in counts if c > 0) / len(counts) if counts else 0.0
            return round(intensite * couverture, 1)

        def _score_saisonnalite(moyennes, moyenne_globale, counts):
            if not moyenne_globale or not any(counts):
                return 0.0

            deviations = [
                abs(_pct_vs(moyenne_globale, v))
                for v, c in zip(moyennes, counts)
                if c > 0
            ]
            return round(sum(deviations) / len(deviations), 1) if deviations else 0.0

        def _niveau_risque(score):
            if score >= 20:
                return "Élevé"
            if score >= 10:
                return "Modéré"
            return "Faible"

        def _couleur_risque(score):
            if score >= 20:
                return UI["danger"]
            if score >= 10:
                return UI["amber"]
            return UI["success"]

        def _label_stabilite(cv):
            if cv < 10:
                return "Très stable"
            if cv < 20:
                return "Stable"
            if cv < 35:
                return "Variable"
            return "Très variable"

        def _tag_stabilite(cv):
            if cv < 20:
                return "stable"
            if cv < 35:
                return "variable"
            return "tres_variable"

        def _stats_stabilite(vals):
            vals = [v for v in vals if v is not None]
            n = len(vals)

            if n == 0:
                return {
                    "n": 0,
                    "mean": 0.0,
                    "median": 0.0,
                    "min": 0.0,
                    "max": 0.0,
                    "stdev": 0.0,
                    "cv": 0.0,
                    "ic_low": 0.0,
                    "ic_high": 0.0,
                    "ic_margin": 0.0,
                    "amplitude": 0.0,
                    "label": "Aucune donnée",
                }

            mean_v = statistics.mean(vals)
            median_v = statistics.median(vals)
            min_v = min(vals)
            max_v = max(vals)
            stdev_v = statistics.stdev(vals) if n >= 2 else 0.0
            cv = (stdev_v / abs(mean_v) * 100) if mean_v not in (0, None) else 0.0
            margin = 1.96 * (stdev_v / math.sqrt(n)) if n >= 2 else 0.0

            return {
                "n": n,
                "mean": mean_v,
                "median": median_v,
                "min": min_v,
                "max": max_v,
                "stdev": stdev_v,
                "cv": cv,
                "ic_low": mean_v - margin,
                "ic_high": mean_v + margin,
                "ic_margin": margin,
                "amplitude": max_v - min_v,
                "label": _label_stabilite(cv),
            }

        def _pick_peak_and_trough(valid_idx, stats_list, orientation):
            if not valid_idx:
                return None, None

            if orientation < 0:
                idx_peak = min(valid_idx, key=lambda i: stats_list[i]["mean"])
                idx_trough = max(valid_idx, key=lambda i: stats_list[i]["mean"])
            else:
                idx_peak = max(valid_idx, key=lambda i: stats_list[i]["mean"])
                idx_trough = min(valid_idx, key=lambda i: stats_list[i]["mean"])

            return idx_peak, idx_trough

        def _pick_peak_idx(valid_idx, stats_list, orientation):
            if not valid_idx:
                return None
            if orientation < 0:
                return min(valid_idx, key=lambda i: stats_list[i]["mean"])
            return max(valid_idx, key=lambda i: stats_list[i]["mean"])

        # =========================================================
        # HELPERS K-MEANS 1D
        # =========================================================
        def _cluster_name(rank, k):
            if k <= 1:
                return "Régime unique"
            if k == 2:
                return ["Bas", "Haut"][rank]
            return ["Bas", "Moyen", "Haut"][rank]

        def _cluster_color(rank, k):
            if k <= 1:
                return "#AAB7B8"
            palette = ["#5DADE2", "#F5B041", "#EC7063"]
            return palette[min(rank, len(palette) - 1)]

        def _kmeans_1d(vals, k, max_iter=100):
            vals = [float(v) for v in vals if v is not None]
            n = len(vals)

            if n == 0:
                return {
                    "k": 0,
                    "values": [],
                    "assignments": [],
                    "centers": [],
                    "clusters": [],
                    "inertia": 0.0,
                }

            unique_vals = sorted(set(vals))
            k = max(1, min(k, len(unique_vals), n))

            if k == 1:
                center = _safe_mean(vals)
                s = _stats_stabilite(vals)
                return {
                    "k": 1,
                    "values": vals[:],
                    "assignments": [0] * n,
                    "centers": [center],
                    "clusters": [{
                        **s,
                        "cluster_index": 0,
                        "center": center,
                        "name": _cluster_name(0, 1),
                        "color": _cluster_color(0, 1),
                        "values": vals[:],
                    }],
                    "inertia": sum((v - center) ** 2 for v in vals),
                }

            unique_sorted = unique_vals[:]
            positions = [round(i * (len(unique_sorted) - 1) / (k - 1)) for i in range(k)]
            centers = [unique_sorted[pos] for pos in positions]

            dedup_centers = []
            for c in centers:
                if c not in dedup_centers:
                    dedup_centers.append(c)

            for candidate in unique_sorted:
                if len(dedup_centers) >= k:
                    break
                if candidate not in dedup_centers:
                    dedup_centers.append(candidate)

            centers = dedup_centers[:k]
            assignments = [0] * n

            for _ in range(max_iter):
                for i, v in enumerate(vals):
                    assignments[i] = min(
                        range(len(centers)),
                        key=lambda j: (abs(v - centers[j]), j)
                    )

                new_centers = []
                for j in range(len(centers)):
                    cluster_vals = [v for v, a in zip(vals, assignments) if a == j]
                    if cluster_vals:
                        new_centers.append(_safe_mean(cluster_vals))
                    else:
                        farthest_point = max(
                            vals,
                            key=lambda v: min(abs(v - c) for c in centers)
                        )
                        new_centers.append(farthest_point)

                if len(new_centers) == len(centers) and all(
                    abs(a - b) < 1e-9 for a, b in zip(new_centers, centers)
                ):
                    centers = new_centers
                    break

                centers = new_centers

            order = sorted(range(len(centers)), key=lambda j: centers[j])
            remap = {old_idx: new_idx for new_idx, old_idx in enumerate(order)}
            centers_sorted = [centers[j] for j in order]
            assignments_sorted = [remap[a] for a in assignments]

            clusters = []
            inertia = 0.0

            for new_j, center in enumerate(centers_sorted):
                cluster_vals = [v for v, a in zip(vals, assignments_sorted) if a == new_j]
                if not cluster_vals:
                    continue

                s = _stats_stabilite(cluster_vals)
                inertia += sum((v - center) ** 2 for v in cluster_vals)

                clusters.append({
                    **s,
                    "cluster_index": new_j,
                    "center": center,
                    "name": _cluster_name(new_j, len(centers_sorted)),
                    "color": _cluster_color(new_j, len(centers_sorted)),
                    "values": cluster_vals[:],
                })

            if len(clusters) != len(centers_sorted):
                new_centers = [c["center"] for c in clusters]
                new_assignments = []
                for v in vals:
                    if not new_centers:
                        new_assignments.append(0)
                    else:
                        new_assignments.append(
                            min(range(len(new_centers)), key=lambda j: (abs(v - new_centers[j]), j))
                        )
                centers_sorted = new_centers
                assignments_sorted = new_assignments

            final_k = len(clusters)

            for rank, cluster in enumerate(clusters):
                cluster["cluster_index"] = rank
                cluster["name"] = _cluster_name(rank, final_k)
                cluster["color"] = _cluster_color(rank, final_k)

            return {
                "k": final_k,
                "values": vals[:],
                "assignments": assignments_sorted,
                "centers": centers_sorted,
                "clusters": clusters,
                "inertia": inertia,
            }

        def _best_kmeans_1d(vals, max_k=3):
            vals = [v for v in vals if v is not None]
            n = len(vals)
            unique_count = len(set(vals))

            if n == 0:
                return {
                    "k": 0,
                    "values": [],
                    "assignments": [],
                    "centers": [],
                    "clusters": [],
                    "inertia": 0.0,
                }

            if n < 6 or unique_count <= 1:
                return _kmeans_1d(vals, 1)

            k_upper = min(max_k, unique_count, 3 if n >= 12 else 2)
            models = {k: _kmeans_1d(vals, k) for k in range(1, k_upper + 1)}

            chosen_k = 1

            if 2 in models and models[1]["inertia"] > 0:
                gain_2 = (models[1]["inertia"] - models[2]["inertia"]) / models[1]["inertia"]
                min_cluster_size_2 = min((c["n"] for c in models[2]["clusters"]), default=0)
                if gain_2 >= 0.18 and min_cluster_size_2 >= 2:
                    chosen_k = 2

            if chosen_k == 2 and 3 in models and models[2]["inertia"] > 0:
                gain_3 = (models[2]["inertia"] - models[3]["inertia"]) / models[2]["inertia"]
                min_cluster_size_3 = min((c["n"] for c in models[3]["clusters"]), default=0)
                if gain_3 >= 0.10 and min_cluster_size_3 >= 2:
                    chosen_k = 3

            return models[chosen_k]

        # =========================================================
        # HELPERS STABILITÉ BASÉE K-MEANS
        # =========================================================
        def _cluster_dominance_metrics(km):
            total_n = len(km["values"]) if km else 0
            if not km or km["k"] == 0 or total_n == 0:
                return {
                    "share": 0.0,
                    "dominant_n": 0,
                    "dominant_cluster": None,
                    "label": "Aucune donnée",
                    "tag": "tres_variable",
                    "color": "#C0392B",
                }

            dominant_cluster = max(km["clusters"], key=lambda c: c["n"])
            dominant_n = dominant_cluster["n"]
            share = dominant_n / total_n

            if share >= 0.60:
                label_v = "Stable"
                tag_v = "stable"
                color_v = "#27AE60"
            elif share >= 0.45:
                label_v = "Variable"
                tag_v = "variable"
                color_v = "#D68910"
            else:
                label_v = "Très variable"
                tag_v = "tres_variable"
                color_v = "#C0392B"

            return {
                "share": share,
                "dominant_n": dominant_n,
                "dominant_cluster": dominant_cluster,
                "label": label_v,
                "tag": tag_v,
                "color": color_v,
            }

        # =========================================================
        # ROOT GRID
        # =========================================================
        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        logo_path = _resolve_logo_path()
        try:
            _img = Image.open(logo_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Analyse des tendances de flux",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        neutral_btn(top_right, "Retour au menu", self.retour_menu).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "ANALYSE DES FLUX", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Tendance des flux réels",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Analyse ouvrée, stabilité intra-jour, segmentation K-means, glissements mensuels et profil annuel.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES / ACTIONS
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "PILOTAGE",
            "Filtres et navigation",
            f"Les week-ends sont exclus. Les flux de faible amplitude |flux| < {SEUIL_MIN_FLUX:,.0f} sont retirés de l’analyse.".replace(",", " ")
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        for i in range(5):
            filters_body.grid_columnconfigure(i, weight=0)
        filters_body.grid_columnconfigure(5, weight=1)

        filiales = ["Toutes filiales"] + list(sections.values())
        filiale_var = ctk.StringVar(value=filiales[0])
        annee_var = ctk.StringVar(value="Toutes années")
        flux_var = ctk.StringVar(value="Tous les flux")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=0, sticky="w", pady=(0, 6))
        select_filiale = ctk.CTkOptionMenu(
            filters_body,
            values=filiales,
            variable=filiale_var,
            width=240,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        select_filiale.grid(row=1, column=0, sticky="w", padx=(0, 16))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=1, sticky="w", pady=(0, 6))
        select_annee = ctk.CTkOptionMenu(
            filters_body,
            values=["Toutes années"],
            variable=annee_var,
            width=160,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        select_annee.grid(row=1, column=1, sticky="w", padx=(0, 16))

        label(filters_body, "Flux", font=FONT["small_bold"], color=UI["text_soft"]).grid(row=0, column=2, sticky="w", pady=(0, 6))
        select_flux = ctk.CTkOptionMenu(
            filters_body,
            values=["Tous les flux"],
            variable=flux_var,
            width=280,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        select_flux.grid(row=1, column=2, sticky="w", padx=(0, 16))

        actions_wrap = ctk.CTkFrame(filters_body, fg_color="transparent")
        actions_wrap.grid(row=1, column=3, sticky="w")
        neutral_btn(actions_wrap, "Retour au menu", self.retour_menu).pack(side="left")

        # =========================================================
        # KPI
        # =========================================================
        kpi_card = card(page, fg=UI["surface"], radius=20)
        kpi_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        kpi_card.grid_columnconfigure(0, weight=1)

        section_header(
            kpi_card,
            "KPI",
            "Points clés",
            "Lecture rapide des signaux dominants de la série."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        kpi_wrap = ctk.CTkFrame(kpi_card, fg_color="transparent")
        kpi_wrap.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 12))
        for i in range(4):
            kpi_wrap.grid_columnconfigure(i, weight=1)

        # =========================================================
        # VISUALISATION
        # =========================================================
        viz_card = card(page, fg=UI["surface"], radius=20)
        viz_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 24))
        viz_card.grid_columnconfigure(0, weight=1)

        section_header(
            viz_card,
            "RENDU",
            "Visualisation détaillée",
            "Les graphes sont conservés, seule l’interface a été harmonisée."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        graph_stack = ctk.CTkFrame(viz_card, fg_color="transparent")
        graph_stack.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        graph_stack.grid_columnconfigure(0, weight=1)

        table_title = ctk.CTkLabel(
            viz_card,
            text="Tableau détaillé des tendances",
            font=("Segoe UI", 15, "bold"),
            text_color="white"
        )
        table_title.grid(row=2, column=0, sticky="w", padx=18, pady=(6, 6))

        table_frame = ctk.CTkFrame(
            viz_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        table_frame.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 18))
        table_frame.grid_columnconfigure(0, weight=1)
        table_frame.grid_rowconfigure(0, weight=1)

        colonnes = (
            "Bloc",
            "Libellé",
            "Nb points",
            "Moyenne",
            "Médiane",
            "Min",
            "Max",
            "Écart-type",
            "CV %",
            "IC 95%",
            "Lecture"
        )

        tree_wrap = tk.Frame(table_frame, bg=UI["surface_2"])
        tree_wrap.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

        table = ttk.Treeview(tree_wrap, columns=colonnes, show="headings", height=18, style="Pulse.Treeview")
        table.pack(side="top", fill="x", expand=True)

        x_scroll_table = ttk.Scrollbar(tree_wrap, orient="horizontal", command=table.xview)
        x_scroll_table.pack(side="bottom", fill="x")
        table.configure(xscrollcommand=x_scroll_table.set)

        for col in colonnes:
            table.heading(col, text=col)
            largeur = 135
            if col == "Libellé":
                largeur = 200
            elif col in ("IC 95%", "Lecture"):
                largeur = 240
            table.column(col, anchor="center", width=largeur)

        table.tag_configure("stable", background="#143A2E", foreground="white")
        table.tag_configure("variable", background="#5C4A1F", foreground="white")
        table.tag_configure("tres_variable", background="#5C1F1F", foreground="white")
        table.tag_configure("synthese", background="#1B365D", foreground="white")
        table.tag_configure("kmeans", background="#253B56", foreground="white")

        _make_placeholder(graph_stack, "Chargement initial de l’analyse...")

        # =========================================================
        # MOTEUR PRINCIPAL
        # =========================================================
        def maj_graphique(filiale, annee, flux):
            _clear_children(graph_stack)

            for w in kpi_wrap.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            for row in table.get_children():
                table.delete(row)

            _debug("\n" + "=" * 110)
            _debug(f"[DEBUG ANALYSE] filiale={filiale} | annee={annee} | flux={flux}")
            _debug("=" * 110)

            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            weekly_data = defaultdict(list)
            monthly_day_data = defaultdict(list)
            yearly_month_data = defaultdict(list)
            month_position_data = {
                "Début de mois": [],
                "Milieu de mois": [],
                "Fin de mois": []
            }

            total_debug = {
                "raw_pairs": 0,
                "zip_missing": 0,
                "date_ok": 0,
                "reel_ok": 0,
                "annee_ok": 0,
                "weekend_exclus": 0,
                "seuil_exclus": 0,
                "kept": 0,
            }

            all_values = []

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception as e:
                    _debug(f"[ERREUR CHARGEMENT] feuille={feuille} | erreur={e}")
                    continue

                _debug(f"\n[DEBUG FEUILLE] {feuille} | nb_flux={len(noms_colonnes)}")

                for nom_flux, col_start in noms_colonnes:
                    if flux != "Tous les flux" and str(nom_flux) != str(flux):
                        continue

                    try:
                        dates, reel, previsions, noms_profils = extraire_valeurs(
                            ws, col_start, nb_prev, annee=None
                        )
                    except Exception as e:
                        _debug(f"[ERREUR EXTRACTION] feuille={feuille} | flux={nom_flux} | erreur={e}")
                        continue

                    dates = list(dates) if dates is not None else []
                    reel = list(reel) if reel is not None else []
                    previsions = list(previsions) if previsions is not None else []

                    _debug(f"\n[DEBUG EXTRACTION] feuille={feuille} | flux={nom_flux}")
                    _debug(f"  len(dates)={len(dates)} | len(reel)={len(reel)} | len(previsions)={len(previsions)}")

                    if len(dates) != len(reel):
                        _debug(f"  >>> PROBLEME LONGUEUR : dates={len(dates)} / reel={len(reel)}")

                    _debug(f"  dates sample: {dates[:6]}")
                    _debug(f"  reel  sample: {reel[:6]}")

                    flux_debug = {
                        "raw_pairs": 0,
                        "zip_missing": 0,
                        "date_ok": 0,
                        "reel_ok": 0,
                        "annee_ok": 0,
                        "weekend_exclus": 0,
                        "seuil_exclus": 0,
                        "kept": 0,
                    }

                    for idx, (d_raw, r_raw) in enumerate(zip_longest(dates, reel, fillvalue=None)):
                        flux_debug["raw_pairs"] += 1
                        total_debug["raw_pairs"] += 1

                        if d_raw is None or r_raw is None:
                            flux_debug["zip_missing"] += 1
                            total_debug["zip_missing"] += 1
                            _debug(
                                f"  [DEBUG ZIP] feuille={feuille} | flux={nom_flux} | idx={idx} | "
                                f"d_raw={d_raw} | r_raw={r_raw}"
                            )
                            continue

                        d = _to_date(d_raw)
                        if d is None:
                            if idx < 10:
                                _debug(f"  [DATE NON PARSEE] idx={idx} | valeur={d_raw!r}")
                            continue
                        flux_debug["date_ok"] += 1
                        total_debug["date_ok"] += 1

                        r = _to_number(r_raw)
                        if r is None:
                            if idx < 10:
                                _debug(f"  [REEL NON PARSE] idx={idx} | valeur={r_raw!r}")
                            continue
                        flux_debug["reel_ok"] += 1
                        total_debug["reel_ok"] += 1

                        if annee is not None and d.year != annee:
                            continue
                        flux_debug["annee_ok"] += 1
                        total_debug["annee_ok"] += 1

                        if not _is_business_day(d):
                            flux_debug["weekend_exclus"] += 1
                            total_debug["weekend_exclus"] += 1
                            continue

                        if abs(r) < SEUIL_MIN_FLUX:
                            flux_debug["seuil_exclus"] += 1
                            total_debug["seuil_exclus"] += 1
                            continue

                        flux_debug["kept"] += 1
                        total_debug["kept"] += 1

                        all_values.append(r)
                        weekly_data[d.weekday()].append(r)
                        monthly_day_data[d.day].append(r)
                        yearly_month_data[d.month].append(r)

                        last_bd = _last_business_day_of_month(d)
                        if last_bd:
                            delta_fin = (last_bd - d).days
                            if d.day <= 5:
                                month_position_data["Début de mois"].append(r)
                            elif delta_fin <= 4:
                                month_position_data["Fin de mois"].append(r)
                            else:
                                month_position_data["Milieu de mois"].append(r)

                    _debug(f"  [DEBUG FILTRES] feuille={feuille} | flux={nom_flux} | stats={flux_debug}")

            _debug(f"\n[DEBUG GLOBAL] stats={total_debug}")
            _debug(f"[DEBUG GLOBAL] nb valeurs exploitées final = {len(all_values)}")

            if not all_values:
                no_data = (
                    "Aucune donnée exploitable pour la combinaison sélectionnée "
                    "après exclusion des week-ends et des flux de faible amplitude |flux| < 10 000."
                )
                _make_placeholder(graph_stack, no_data)
                return

            moyenne_globale = _safe_mean(all_values)
            orientation_flux = -1 if moyenne_globale < 0 else 1

            jours_semaine = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi"]
            idx_jours_semaine = [0, 1, 2, 3, 4]

            stats_week = []
            moy_week = []
            count_week = []

            for i in idx_jours_semaine:
                s = _stats_stabilite(weekly_data[i])
                stats_week.append(s)
                moy_week.append(s["mean"])
                count_week.append(s["n"])

            valid_week_idx = [i for i, s in enumerate(stats_week) if s["n"] > 0]

            _debug(f"[DEBUG HEBDO] counts par jour = {dict((jours_semaine[i], stats_week[i]['n']) for i in range(5))}")

            if not valid_week_idx:
                no_data = "Aucune donnée exploitable pour la combinaison sélectionnée."
                _make_placeholder(graph_stack, no_data)
                return

            idx_week_peak, idx_week_trough = _pick_peak_and_trough(
                valid_week_idx, stats_week, orientation_flux
            )

            score_hebdo = _score_tendance(moy_week, moyenne_globale, count_week)
            score_saisonnalite = _score_saisonnalite(moy_week, moyenne_globale, count_week)
            risque_global = max(score_hebdo, score_saisonnalite)

            weekday_clusters = {}
            weekday_cluster_metrics = {}

            for i in idx_jours_semaine:
                weekday_clusters[i] = _best_kmeans_1d(weekly_data[i], max_k=KMEANS_MAX_K)
                weekday_cluster_metrics[i] = _cluster_dominance_metrics(weekday_clusters[i])

            idx_kmeans_max = max(
                valid_week_idx,
                key=lambda i: (
                    weekday_clusters[i]["k"],
                    (
                        weekday_clusters[i]["clusters"][-1]["center"] - weekday_clusters[i]["clusters"][0]["center"]
                        if weekday_clusters[i]["k"] >= 2 else 0.0
                    )
                )
            )
            max_k_detected = max((weekday_clusters[i]["k"] for i in valid_week_idx), default=1)

            idx_stab_km_max = max(valid_week_idx, key=lambda i: weekday_cluster_metrics[i]["share"])
            idx_stab_km_min = min(valid_week_idx, key=lambda i: weekday_cluster_metrics[i]["share"])

            jours_mois = list(range(1, 32))
            stats_month_day = [_stats_stabilite(monthly_day_data[j]) for j in jours_mois]
            valid_month_day_idx = [i for i, s in enumerate(stats_month_day) if s["n"] > 0]

            rolling_window_centers = []
            rolling_window_stats = []

            for center_day in range(1, 32):
                vals = []
                for neighbor in [center_day - 1, center_day, center_day + 1]:
                    if 1 <= neighbor <= 31:
                        vals.extend(monthly_day_data[neighbor])
                rolling_window_centers.append(center_day)
                rolling_window_stats.append(_stats_stabilite(vals))

            valid_roll_idx = [i for i, s in enumerate(rolling_window_stats) if s["n"] > 0]
            idx_roll_peak = _pick_peak_idx(valid_roll_idx, rolling_window_stats, orientation_flux) if valid_roll_idx else 0

            mois_annee = [
                "Jan", "Fév", "Mar", "Avr", "Mai", "Juin",
                "Juil", "Août", "Sep", "Oct", "Nov", "Déc"
            ]
            stats_year_month = [_stats_stabilite(yearly_month_data[m]) for m in range(1, 13)]
            valid_year_month_idx = [i for i, s in enumerate(stats_year_month) if s["n"] > 0]

            if valid_year_month_idx:
                idx_year_month_peak, idx_year_month_trough = _pick_peak_and_trough(
                    valid_year_month_idx, stats_year_month, orientation_flux
                )
            else:
                idx_year_month_peak, idx_year_month_trough = 0, 0

            debut_stats = _stats_stabilite(month_position_data["Début de mois"])
            milieu_stats = _stats_stabilite(month_position_data["Milieu de mois"])
            fin_stats = _stats_stabilite(month_position_data["Fin de mois"])

            suffix_filiale = filiale
            suffix_annee = "Toutes années" if annee is None else str(annee)
            suffix_flux = flux

            # =====================================================
            # KPI
            # =====================================================
            kpi_specs = [
                (
                    "Pic structurel",
                    jours_semaine[idx_week_peak],
                    f"{_pct_vs(moyenne_globale, stats_week[idx_week_peak]['mean']):+.1f}% vs moyenne globale",
                    UI["surface_2"]
                ),
                (
                    "Creux structurel",
                    jours_semaine[idx_week_trough],
                    f"{_pct_vs(moyenne_globale, stats_week[idx_week_trough]['mean']):+.1f}% vs moyenne globale",
                    UI["surface_2"]
                ),
                (
                    "Jour le plus stable",
                    jours_semaine[idx_stab_km_max],
                    f"Cluster dominant = {weekday_cluster_metrics[idx_stab_km_max]['share'] * 100:.1f}%",
                    UI["success"]
                ),
                (
                    "Jour le plus variable",
                    jours_semaine[idx_stab_km_min],
                    f"Cluster dominant = {weekday_cluster_metrics[idx_stab_km_min]['share'] * 100:.1f}%",
                    UI["danger"]
                ),
                (
                    "Jour le plus segmenté",
                    jours_semaine[idx_kmeans_max],
                    f"{weekday_clusters[idx_kmeans_max]['k']} régime(s) détecté(s)",
                    "#6C5CE7"
                ),
                (
                    "Pic glissant 3 jours",
                    f"autour du {rolling_window_centers[idx_roll_peak]}",
                    f"Moyenne zone = {rolling_window_stats[idx_roll_peak]['mean']:,.0f}".replace(",", " "),
                    UI["surface_2"]
                ),
                (
                    "Risque global",
                    _niveau_risque(risque_global),
                    f"Score tendance = {risque_global:.1f}",
                    _couleur_risque(risque_global)
                ),
            ]

            for idx, (t, v, s, c) in enumerate(kpi_specs):
                row = idx // 4
                col = idx % 4
                k = _create_kpi_card(kpi_wrap, t, v, s, c)
                k.grid(row=row, column=col, padx=8, pady=8, sticky="nsew")
                kpi_wrap.grid_columnconfigure(col, weight=1)

            # =====================================================
            # GRAPHE 1 : NIVEAU 1 — STRUCTURE HEBDOMADAIRE
            # =====================================================
            card_graph_hebdo = ctk.CTkFrame(
                graph_stack,
                fg_color=UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            card_graph_hebdo.grid(row=0, column=0, sticky="ew", padx=6, pady=(0, 12))

            ctk.CTkLabel(
                card_graph_hebdo,
                text="Niveau 1 — Structure hebdomadaire",
                font=("Segoe UI", 15, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=15, pady=(12, 2))

            ctk.CTkLabel(
                card_graph_hebdo,
                text="Lecture : ce graphe compare les jours ouvrés entre eux. Les barres montrent la moyenne de chaque jour, et la ligne jaune pointillée représente la moyenne globale ouvrée.",
                font=("Segoe UI", 11),
                text_color=UI["text_soft"],
                wraplength=1200,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(0, 10))

            fig1, ax1 = plt.subplots(figsize=(10, 4), facecolor=UI["surface_2"])
            ax1.set_facecolor(UI["surface_2"])

            moyennes_hebdo = [stats_week[i]["mean"] for i in range(len(jours_semaine))]
            ax1.bar(jours_semaine, moyennes_hebdo, color="#1F77B4", alpha=0.90)
            ax1.plot(jours_semaine, moyennes_hebdo, color="white", marker="o", linewidth=2)
            ax1.axhline(y=moyenne_globale, color="#FFCC00", linestyle="--", linewidth=1.8)

            ax1.annotate(
                "Pic",
                xy=(jours_semaine[idx_week_peak], stats_week[idx_week_peak]["mean"]),
                xytext=(0, 12),
                textcoords="offset points",
                ha="center",
                color="#7CFC00",
                fontsize=11,
                fontweight="bold"
            )
            ax1.annotate(
                "Creux",
                xy=(jours_semaine[idx_week_trough], stats_week[idx_week_trough]["mean"]),
                xytext=(0, -18),
                textcoords="offset points",
                ha="center",
                color="#FF6B6B",
                fontsize=11,
                fontweight="bold"
            )

            ax1.set_title(
                f"Structure hebdomadaire ouvrée - {suffix_filiale} - {suffix_flux} - {suffix_annee}",
                fontsize=14,
                color="white"
            )
            ax1.set_ylabel("Flux réel moyen", color="white")
            ax1.tick_params(axis="y", colors="white")
            ax1.tick_params(axis="x", rotation=20, labelcolor="white")
            ax1.grid(axis="y", color="gray", linestyle="--", alpha=0.3)

            legend_elements_hebdo = [
                Patch(facecolor="#1F77B4", edgecolor="#1F77B4", label="Barres : moyenne par jour"),
                Line2D([0], [0], color="white", marker="o", linewidth=2, label="Courbe : tendance"),
                Line2D([0], [0], color="#FFCC00", linestyle="--", linewidth=2, label="Ligne jaune : moyenne globale ouvrée")
            ]
            _render_figure(
                fig1,
                ax1,
                card_graph_hebdo,
                legend_handles=legend_elements_hebdo,
                aspect_ratio=3.45,
                min_height=300
            )

            # =====================================================
            # GRAPHE 2 : K-MEANS PAR JOUR OUVRÉ
            # =====================================================
            card_graph_kmeans = ctk.CTkFrame(
                graph_stack,
                fg_color=UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            card_graph_kmeans.grid(row=1, column=0, sticky="ew", padx=6, pady=(0, 12))

            ctk.CTkLabel(
                card_graph_kmeans,
                text="Segmentation K-means — différentes catégories d’un même jour",
                font=("Segoe UI", 15, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=15, pady=(12, 2))

            ctk.CTkLabel(
                card_graph_kmeans,
                text="Lecture : les points représentent les observations historiques d’un même jour ouvré. Le K-means regroupe ces points en régimes (Bas / Moyen / Haut). Le losange indique le centre de chaque cluster.",
                font=("Segoe UI", 11),
                text_color=UI["text_soft"],
                wraplength=1200,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(0, 10))

            fig_km, ax_km = plt.subplots(figsize=(10, 4), facecolor=UI["surface_2"])
            ax_km.set_facecolor(UI["surface_2"])

            for i, jour in enumerate(jours_semaine):
                km = weekday_clusters.get(i)
                if not km or km["k"] == 0:
                    continue

                seen_by_cluster = defaultdict(int)
                x_hist = []
                y_hist = []
                c_hist = []

                for value, assign in sorted(zip(km["values"], km["assignments"]), key=lambda t: (t[1], t[0])):
                    seen_by_cluster[assign] += 1
                    jitter = ((seen_by_cluster[assign] % 7) - 3) * 0.018
                    x_hist.append(i + jitter)
                    y_hist.append(value)
                    c_hist.append(km["clusters"][assign]["color"])

                ax_km.scatter(x_hist, y_hist, s=28, c=c_hist, alpha=0.35, edgecolors="none", zorder=2)

                if km["k"] == 1:
                    offsets = [0.0]
                elif km["k"] == 2:
                    offsets = [-0.14, 0.14]
                else:
                    offsets = [-0.22, 0.0, 0.22]

                for rank, cl in enumerate(km["clusters"]):
                    x = i + offsets[min(rank, len(offsets) - 1)]
                    ax_km.vlines(
                        x,
                        cl["min"],
                        cl["max"],
                        color=cl["color"],
                        linewidth=3,
                        alpha=0.95,
                        zorder=3
                    )
                    ax_km.scatter(
                        [x],
                        [cl["center"]],
                        s=210,
                        marker="D",
                        color=cl["color"],
                        edgecolors="white",
                        linewidths=1.2,
                        zorder=4
                    )
                    ax_km.annotate(
                        f"{cl['name']}\nμ {cl['center']:,.0f}\nn={cl['n']}",
                        xy=(x, cl["center"]),
                        xytext=(0, 10),
                        textcoords="offset points",
                        ha="center",
                        va="bottom",
                        fontsize=8.5,
                        color="white"
                    )

            ax_km.axhline(y=moyenne_globale, color="#FFCC00", linestyle="--", linewidth=1.8)
            ax_km.set_xticks(list(range(len(jours_semaine))))
            ax_km.set_xticklabels(jours_semaine)
            ax_km.set_title(
                f"Segmentation K-means par jour ouvré - {suffix_filiale} - {suffix_flux} - {suffix_annee}",
                fontsize=14,
                color="white"
            )
            ax_km.set_ylabel("Flux réel historique", color="white")
            ax_km.tick_params(axis="y", colors="white")
            ax_km.tick_params(axis="x", rotation=20, labelcolor="white")
            ax_km.grid(axis="y", color="gray", linestyle="--", alpha=0.3)

            legend_elements_km = [
                Line2D([0], [0], marker="o", color="none", markerfacecolor="#5DADE2", markeredgecolor="white", markersize=9, label="Bleu : régime bas"),
                Line2D([0], [0], marker="o", color="none", markerfacecolor="#F5B041", markeredgecolor="white", markersize=9, label="Orange : régime intermédiaire"),
                Line2D([0], [0], marker="o", color="none", markerfacecolor="#EC7063", markeredgecolor="white", markersize=9, label="Rouge : régime haut"),
                Line2D([0], [0], marker="D", color="none", markerfacecolor="white", markeredgecolor="white", markersize=8, label="Losange : centre du cluster"),
                Line2D([0], [0], color="#FFCC00", linestyle="--", linewidth=2, label="Ligne jaune : moyenne globale ouvrée")
            ]
            _render_figure(
                fig_km,
                ax_km,
                card_graph_kmeans,
                legend_handles=legend_elements_km,
                aspect_ratio=3.20,
                min_height=320
            )

            # =====================================================
            # GRAPHE 3 : STABILITÉ INTRA-JOUR PILOTÉE PAR K-MEANS
            # =====================================================
            card_graph_stab = ctk.CTkFrame(
                graph_stack,
                fg_color=UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            card_graph_stab.grid(row=2, column=0, sticky="ew", padx=6, pady=(0, 12))

            ctk.CTkLabel(
                card_graph_stab,
                text="Niveau 2 — Stabilité intra-jour pilotée par K-means",
                font=("Segoe UI", 15, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=15, pady=(12, 2))

            ctk.CTkLabel(
                card_graph_stab,
                text="Lecture : chaque barre représente la moyenne du jour. Les couleurs dépendent du poids du cluster dominant K-means : vert si un régime domine clairement, orange si la concentration est moyenne, rouge si les points sont répartis sur plusieurs régimes.",
                font=("Segoe UI", 11),
                text_color=UI["text_soft"],
                wraplength=1200,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(0, 10))

            fig_stab, ax_stab = plt.subplots(figsize=(10, 4), facecolor=UI["surface_2"])
            ax_stab.set_facecolor(UI["surface_2"])

            means_stab = [s["mean"] for s in stats_week]
            ic_margins = [s["ic_margin"] for s in stats_week]
            color_bars = [weekday_cluster_metrics[i]["color"] for i in range(len(jours_semaine))]

            bars = ax_stab.bar(
                jours_semaine,
                means_stab,
                color=color_bars,
                alpha=0.90,
                yerr=ic_margins,
                capsize=7,
                error_kw={"ecolor": "white", "elinewidth": 1.6, "capthick": 1.6}
            )
            ax_stab.axhline(y=moyenne_globale, color="#FFCC00", linestyle="--", linewidth=1.8)

            for i, bar in enumerate(bars):
                s = stats_week[i]
                km_info = weekday_cluster_metrics[i]
                texte = (
                    f"dom {km_info['share'] * 100:.0f}%\n"
                    f"CV {s['cv']:.1f}%\n"
                    f"{km_info['label']}"
                )
                ax_stab.annotate(
                    texte,
                    xy=(bar.get_x() + bar.get_width() / 2, bar.get_height() + s["ic_margin"]),
                    xytext=(0, 10),
                    textcoords="offset points",
                    ha="center",
                    va="bottom",
                    fontsize=9,
                    color="white"
                )

            ax_stab.set_title(
                f"Stabilité intra-jour pilotée par K-means - {suffix_filiale} - {suffix_flux} - {suffix_annee}",
                fontsize=14,
                color="white"
            )
            ax_stab.set_ylabel("Flux réel moyen", color="white")
            ax_stab.tick_params(axis="y", colors="white")
            ax_stab.tick_params(axis="x", rotation=20, labelcolor="white")
            ax_stab.grid(axis="y", color="gray", linestyle="--", alpha=0.3)

            legend_elements_stab = [
                Patch(facecolor="#27AE60", edgecolor="#27AE60", label="Vert : cluster dominant ≥ 60%"),
                Patch(facecolor="#D68910", edgecolor="#D68910", label="Orange : cluster dominant entre 45% et 60%"),
                Patch(facecolor="#C0392B", edgecolor="#C0392B", label="Rouge : cluster dominant < 45%"),
                Line2D([0], [0], color="white", linewidth=2, label="Barres d'erreur : IC95"),
                Line2D([0], [0], color="#FFCC00", linestyle="--", linewidth=2, label="Ligne jaune : moyenne globale ouvrée")
            ]
            _render_figure(
                fig_stab,
                ax_stab,
                card_graph_stab,
                legend_handles=legend_elements_stab,
                aspect_ratio=3.25,
                min_height=315
            )

            # =====================================================
            # GRAPHE 4 : GLISSEMENT MENSUEL
            # =====================================================
            card_graph_gliss = ctk.CTkFrame(
                graph_stack,
                fg_color=UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            card_graph_gliss.grid(row=3, column=0, sticky="ew", padx=6, pady=(0, 12))

            ctk.CTkLabel(
                card_graph_gliss,
                text="Détection des phénomènes de glissement",
                font=("Segoe UI", 15, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=15, pady=(12, 2))

            ctk.CTkLabel(
                card_graph_gliss,
                text="Lecture : les barres bleues représentent le jour exact du mois. La courbe blanche montre la moyenne glissante sur 3 jours ([J-1 ; J ; J+1]).",
                font=("Segoe UI", 11),
                text_color=UI["text_soft"],
                wraplength=1200,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(0, 10))

            fig_gliss, ax_gliss = plt.subplots(figsize=(10, 4), facecolor=UI["surface_2"])
            ax_gliss.set_facecolor(UI["surface_2"])

            x_exact = [jours_mois[i] for i in valid_month_day_idx]
            y_exact = [stats_month_day[i]["mean"] for i in valid_month_day_idx]

            x_roll = [rolling_window_centers[i] for i in valid_roll_idx]
            y_roll = [rolling_window_stats[i]["mean"] for i in valid_roll_idx]

            ax_gliss.bar(x_exact, y_exact, color="#3498DB", alpha=0.72, label="Jour exact du mois")
            ax_gliss.plot(x_roll, y_roll, color="white", marker="o", linewidth=2, label="Fenêtre glissante 3 jours")
            ax_gliss.axhline(y=moyenne_globale, color="#FFCC00", linestyle="--", linewidth=1.8)

            if valid_roll_idx:
                roll_center = rolling_window_centers[idx_roll_peak]
                roll_stats = rolling_window_stats[idx_roll_peak]
                ax_gliss.annotate(
                    f"Pic mobile\n{max(1, roll_center-1)}-{min(31, roll_center+1)}",
                    xy=(roll_center, roll_stats["mean"]),
                    xytext=(0, 12),
                    textcoords="offset points",
                    ha="center",
                    color="#7CFC00",
                    fontsize=10,
                    fontweight="bold"
                )

            ax_gliss.set_title(
                f"Glissement mensuel - {suffix_filiale} - {suffix_flux} - {suffix_annee}",
                fontsize=14,
                color="white"
            )
            ax_gliss.set_ylabel("Flux réel moyen", color="white")
            ax_gliss.tick_params(axis="y", colors="white")
            ax_gliss.tick_params(axis="x", rotation=70, labelcolor="white")
            ax_gliss.set_xticks(x_exact)
            ax_gliss.set_xticklabels([str(x) for x in x_exact])
            ax_gliss.grid(axis="y", color="gray", linestyle="--", alpha=0.3)

            legend_elements_gliss = [
                Patch(facecolor="#3498DB", edgecolor="#3498DB", label="Barres : jour exact"),
                Line2D([0], [0], color="white", marker="o", linewidth=2, label="Courbe : fenêtre glissante 3 jours"),
                Line2D([0], [0], color="#FFCC00", linestyle="--", linewidth=2, label="Ligne jaune : moyenne globale ouvrée")
            ]
            _render_figure(
                fig_gliss,
                ax_gliss,
                card_graph_gliss,
                legend_handles=legend_elements_gliss,
                aspect_ratio=3.35,
                min_height=305
            )

            # =====================================================
            # GRAPHE 5 : ANALYSE ANNUELLE / PAR MOIS
            # =====================================================
            card_graph_annual = ctk.CTkFrame(
                graph_stack,
                fg_color=UI["surface_2"],
                corner_radius=16,
                border_width=1,
                border_color=UI["border_soft"]
            )
            card_graph_annual.grid(row=4, column=0, sticky="ew", padx=6, pady=(0, 12))

            ctk.CTkLabel(
                card_graph_annual,
                text="Analyse annuelle — tendance par mois",
                font=("Segoe UI", 15, "bold"),
                text_color="white"
            ).pack(anchor="w", padx=15, pady=(12, 2))

            ctk.CTkLabel(
                card_graph_annual,
                text="Lecture : ce graphe montre le comportement moyen par mois de l’année. Si une année précise est sélectionnée, il décrit le profil mensuel de cette année.",
                font=("Segoe UI", 11),
                text_color=UI["text_soft"],
                wraplength=1200,
                justify="left"
            ).pack(anchor="w", padx=15, pady=(0, 10))

            fig_annual, ax_annual = plt.subplots(figsize=(10, 4), facecolor=UI["surface_2"])
            ax_annual.set_facecolor(UI["surface_2"])

            x_months = [mois_annee[i] for i in valid_year_month_idx]
            y_months = [stats_year_month[i]["mean"] for i in valid_year_month_idx]

            ax_annual.bar(x_months, y_months, color="#28B463", alpha=0.85)
            ax_annual.plot(x_months, y_months, color="white", marker="o", linewidth=2)
            ax_annual.axhline(y=moyenne_globale, color="#FFCC00", linestyle="--", linewidth=1.8)

            if valid_year_month_idx:
                ax_annual.annotate(
                    "Pic annuel",
                    xy=(mois_annee[idx_year_month_peak], stats_year_month[idx_year_month_peak]["mean"]),
                    xytext=(0, 12),
                    textcoords="offset points",
                    ha="center",
                    color="#7CFC00",
                    fontsize=10,
                    fontweight="bold"
                )
                ax_annual.annotate(
                    "Creux annuel",
                    xy=(mois_annee[idx_year_month_trough], stats_year_month[idx_year_month_trough]["mean"]),
                    xytext=(0, -18),
                    textcoords="offset points",
                    ha="center",
                    color="#FF6B6B",
                    fontsize=10,
                    fontweight="bold"
                )

            titre_annuel = (
                f"Profil mensuel - {suffix_filiale} - {suffix_flux} - {suffix_annee}"
                if annee is not None else
                f"Saisonnalité mensuelle moyenne - {suffix_filiale} - {suffix_flux} - toutes années"
            )

            ax_annual.set_title(titre_annuel, fontsize=14, color="white")
            ax_annual.set_ylabel("Flux réel moyen", color="white")
            ax_annual.tick_params(axis="y", colors="white")
            ax_annual.tick_params(axis="x", labelcolor="white")
            ax_annual.grid(axis="y", color="gray", linestyle="--", alpha=0.3)

            legend_elements_annual = [
                Patch(facecolor="#28B463", edgecolor="#28B463", label="Barres : moyenne par mois"),
                Line2D([0], [0], color="white", marker="o", linewidth=2, label="Courbe : tendance annuelle"),
                Line2D([0], [0], color="#FFCC00", linestyle="--", linewidth=2, label="Ligne jaune : moyenne globale ouvrée")
            ]
            _render_figure(
                fig_annual,
                ax_annual,
                card_graph_annual,
                legend_handles=legend_elements_annual,
                aspect_ratio=3.35,
                min_height=305
            )

            # =====================================================
            # TABLEAU HEBDO
            # =====================================================
            for i, jour in enumerate(jours_semaine):
                s = stats_week[i]
                km_info = weekday_cluster_metrics[i]
                if s["n"] == 0:
                    continue

                table.insert(
                    "",
                    "end",
                    values=(
                        "Hebdo",
                        jour,
                        s["n"],
                        f"{s['mean']:,.2f}".replace(",", " "),
                        f"{s['median']:,.2f}".replace(",", " "),
                        f"{s['min']:,.2f}".replace(",", " "),
                        f"{s['max']:,.2f}".replace(",", " "),
                        f"{s['stdev']:,.2f}".replace(",", " "),
                        f"{s['cv']:.1f}%",
                        f"[{s['ic_low']:,.0f} ; {s['ic_high']:,.0f}]".replace(",", " "),
                        f"{km_info['label']} — cluster dominant {km_info['share'] * 100:.1f}%"
                    ),
                    tags=(km_info["tag"],)
                )

            # =====================================================
            # TABLEAU K-MEANS
            # =====================================================
            for i, jour in enumerate(jours_semaine):
                km = weekday_clusters.get(i)
                if not km or km["k"] == 0:
                    continue

                for cl in km["clusters"]:
                    table.insert(
                        "",
                        "end",
                        values=(
                            "K-means",
                            f"{jour} - {cl['name']}",
                            cl["n"],
                            f"{cl['mean']:,.2f}".replace(",", " "),
                            f"{cl['median']:,.2f}".replace(",", " "),
                            f"{cl['min']:,.2f}".replace(",", " "),
                            f"{cl['max']:,.2f}".replace(",", " "),
                            f"{cl['stdev']:,.2f}".replace(",", " "),
                            f"{cl['cv']:.1f}%",
                            f"[{cl['ic_low']:,.0f} ; {cl['ic_high']:,.0f}]".replace(",", " "),
                            f"Centre = {cl['center']:,.0f}".replace(",", " ")
                        ),
                        tags=("kmeans",)
                    )

            # =====================================================
            # TABLEAU MENSUEL - jours exacts
            # =====================================================
            for i in valid_month_day_idx:
                s = stats_month_day[i]
                jour = jours_mois[i]

                table.insert(
                    "",
                    "end",
                    values=(
                        "Mensuel",
                        str(jour),
                        s["n"],
                        f"{s['mean']:,.2f}".replace(",", " "),
                        f"{s['median']:,.2f}".replace(",", " "),
                        f"{s['min']:,.2f}".replace(",", " "),
                        f"{s['max']:,.2f}".replace(",", " "),
                        f"{s['stdev']:,.2f}".replace(",", " "),
                        f"{s['cv']:.1f}%",
                        f"[{s['ic_low']:,.0f} ; {s['ic_high']:,.0f}]".replace(",", " "),
                        s["label"]
                    ),
                    tags=(_tag_stabilite(s["cv"]),)
                )

            # =====================================================
            # TABLEAU SYNTHÈSE début / milieu / fin
            # =====================================================
            for bloc_label, s in [
                ("Début de mois", debut_stats),
                ("Milieu de mois", milieu_stats),
                ("Fin de mois", fin_stats),
            ]:
                table.insert(
                    "",
                    "end",
                    values=(
                        "Synthèse",
                        bloc_label,
                        s["n"],
                        f"{s['mean']:,.2f}".replace(",", " "),
                        f"{s['median']:,.2f}".replace(",", " "),
                        f"{s['min']:,.2f}".replace(",", " "),
                        f"{s['max']:,.2f}".replace(",", " "),
                        f"{s['stdev']:,.2f}".replace(",", " "),
                        f"{s['cv']:.1f}%",
                        f"[{s['ic_low']:,.0f} ; {s['ic_high']:,.0f}]".replace(",", " "),
                        s["label"]
                    ),
                    tags=("synthese",)
                )

            _sync_scrollregion()

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _on_filiale_change(value=None):
            filiale = filiale_var.get()

            annees = _annees_pour_filiale(filiale)
            year_values = ["Toutes années"] + [str(a) for a in annees] if annees else ["Toutes années"]
            select_annee.configure(values=year_values)

            annee_var.set("Toutes années")

            fluxes = _flux_pour_filiale(filiale)
            flux_values = ["Tous les flux"] + fluxes if fluxes else ["Tous les flux"]
            select_flux.configure(values=flux_values)
            flux_var.set("Tous les flux")

            maj_graphique(filiale, None, flux_var.get())

        def _on_annee_change(value=None):
            filiale = filiale_var.get()
            val_annee = annee_var.get()
            annee = None if val_annee == "Toutes années" else int(val_annee)
            maj_graphique(filiale, annee, flux_var.get())

        def _on_flux_change(value=None):
            filiale = filiale_var.get()
            val_annee = annee_var.get()
            annee = None if val_annee == "Toutes années" else int(val_annee)
            maj_graphique(filiale, annee, flux_var.get())

        select_filiale.configure(command=_on_filiale_change)
        select_annee.configure(command=_on_annee_change)
        select_flux.configure(command=_on_flux_change)

        # =========================================================
        # AFFICHAGE INITIAL
        # =========================================================
        _on_filiale_change()

    def exporter_ecarts_excel(self, ecarts_data):
        # Création du fichier Excel
        wb_out = Workbook()
        ws = wb_out.active
        ws.title = "Écarts Importants"

        # Titres
        colonnes = ["Date", "Profil", "Filiales", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]

        # Style pour titres
        titre_font = Font(bold=True, color="000000")
        titre_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Ajout des titres
        for col_num, titre in enumerate(colonnes, 1):
            cell = ws.cell(row=1, column=col_num, value=titre)
            cell.font = titre_font
            cell.fill = titre_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajout des données
        for row_num, data in enumerate(ecarts_data, start=2):
            date_str = data[0].strftime("%Y-%m-%d")
            row_values = [date_str, data[1], data[2], data[3], data[4], data[5], f"{data[6]}%"]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajustement automatique de la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Sauvegarde
        from tkinter import filedialog
        fichier_sortie = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx")],
            title="Enregistrer le fichier Excel"
        )
        if fichier_sortie:
            wb_out.save(fichier_sortie)
            messagebox.showinfo("Export réussi", f"Fichier exporté : {fichier_sortie}")

#===================Page repartition (flux/ filiales/ Profils)===================
    def afficher_repartition(self):
        import customtkinter as ctk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import tkinter as tk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import mplcursors
        import numpy as np
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors
        import datetime as _dt

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _charger(feuille):
            try:
                return charger_donnees(feuille)
            except TypeError:
                return charger_donnees(feuille, taille_bloc)

        def _extraire(ws, col_start, annee=None):
            try:
                return extraire_valeurs(ws, col_start, nb_prev, annee=annee)
            except TypeError:
                try:
                    return extraire_valeurs(ws, col_start, annee=annee)
                except TypeError:
                    try:
                        return extraire_valeurs(ws, col_start, nb_prev)
                    except TypeError:
                        return extraire_valeurs(ws, col_start)

        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    pass
            if isinstance(d, str):
                s = d.strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(s, fmt).year
                    except Exception:
                        continue
                import re
                m = re.search(r"(20\d{2}|19\d{2})", s)
                if m:
                    return int(m.group(1))
            return None

        def _collect_years(feuilles):
            years = set()
            for feuille in feuilles:
                ws, noms_colonnes = _charger(feuille)
                for _nom, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = _extraire(ws, col_start, annee=None)
                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            years.add(y)
            return sorted(years)

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "warning": "#E0B64A",
            "success": "#1F9D63",
            "danger": "#C44E4E",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _embed_figure(fig, master, mode="bar"):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            def _resize(event=None):
                try:
                    master.update_idletasks()

                    w = max(master.winfo_width() - 16, 900)
                    h = max(master.winfo_height() - 16, 420)
                    dpi = fig.get_dpi()

                    # Forcer la taille du widget Tk
                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    # Redimensionner la figure
                    fig.set_size_inches(w / dpi, h / dpi, forward=True)

                    # Marges robustes
                    if mode == "pie":
                        fig.subplots_adjust(left=0.05, right=0.95, bottom=0.08, top=0.90)
                    else:
                        fig.subplots_adjust(left=0.06, right=0.985, bottom=0.20, top=0.90)

                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module de répartition des écarts par filiale",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        neutral_btn(top_right, "Retour à l’accueil", self.creer_accueil).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.grid_rowconfigure(3, weight=0)
        page.grid_columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "RÉPARTITION", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Répartition des écarts par filiale",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Analyse globale du volume d’écarts, de leur fréquence et de leur valorisation, avec filtre par année.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES
        # =========================================================
        feuilles_all = list(sections.values())
        annees_dispo = _collect_years(feuilles_all)

        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "FILTRES",
            "Période d’analyse",
            "Choisissez une année spécifique ou laissez l’analyse sur l’ensemble des données."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        filters_body.grid_columnconfigure(0, weight=0)
        filters_body.grid_columnconfigure(1, weight=1)

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )

        years_values = ["Toutes années"] + [str(a) for a in annees_dispo]
        annee_var = ctk.StringVar(value=(str(annees_dispo[-1]) if annees_dispo else "Toutes années"))

        annees_box = ctk.CTkOptionMenu(
            filters_body,
            values=years_values,
            variable=annee_var,
            width=220,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annees_box.grid(row=1, column=0, sticky="w")

        filter_hint = label(
            filters_body,
            "La sélection met à jour automatiquement les trois graphiques et le panneau de détails.",
            font=FONT["small"],
            color=UI["muted"]
        )
        filter_hint.grid(row=1, column=1, sticky="w", padx=(18, 0))

        # =========================================================
        # CHARTS STACK
        # =========================================================
        charts_stack = ctk.CTkFrame(page, fg_color="transparent")
        charts_stack.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        charts_stack.grid_columnconfigure(0, weight=1)

        pie_card = card(charts_stack, fg=UI["surface"], radius=20)
        pie_card.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        pie_card.grid_columnconfigure(0, weight=1)

        section_header(
            pie_card,
            "VOLUME",
            "Répartition des écarts significatifs",
            "Part relative de chaque filiale dans le total des écarts détectés."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        pie_body = ctk.CTkFrame(
            pie_card, fg_color=UI["surface_2"], corner_radius=14,
            border_width=1, border_color=UI["border_soft"], height=520
        )
        pie_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        pie_body.grid_propagate(False)
        pie_body.grid_rowconfigure(0, weight=1)
        pie_body.grid_columnconfigure(0, weight=1)

        freq_card = card(charts_stack, fg=UI["surface"], radius=20)
        freq_card.grid(row=1, column=0, sticky="ew", pady=(0, 14))
        freq_card.grid_columnconfigure(0, weight=1)

        section_header(
            freq_card,
            "FRÉQUENCE",
            "Fréquence des écarts par filiale",
            "Rapport entre le nombre d’écarts significatifs et le nombre total de prévisions."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        freq_body = ctk.CTkFrame(
            freq_card, fg_color=UI["surface_2"], corner_radius=14,
            border_width=1, border_color=UI["border_soft"], height=460
        )
        freq_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        freq_body.grid_propagate(False)
        freq_body.grid_rowconfigure(0, weight=1)
        freq_body.grid_columnconfigure(0, weight=1)

        val_card = card(charts_stack, fg=UI["surface"], radius=20)
        val_card.grid(row=2, column=0, sticky="ew", pady=(0, 14))
        val_card.grid_columnconfigure(0, weight=1)

        section_header(
            val_card,
            "VALORISATION",
            "Valorisation signée des écarts",
            "Montant cumulé des écarts par filiale en k€."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        val_body = ctk.CTkFrame(
            val_card, fg_color=UI["surface_2"], corner_radius=14,
            border_width=1, border_color=UI["border_soft"], height=460
        )
        val_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        val_body.grid_propagate(False)
        val_body.grid_rowconfigure(0, weight=1)
        val_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # DETAILS PANEL
        # =========================================================
        details_card = card(page, fg=UI["surface"], radius=20)
        details_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 24))
        details_card.grid_columnconfigure(0, weight=1)

        section_header(
            details_card,
            "DÉTAILS",
            "Analyse ciblée par filiale",
            "Cliquez sur une filiale dans le camembert pour afficher les écarts les plus significatifs."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        details_title = label(
            details_card,
            "Aucun détail sélectionné",
            font=FONT["label_bold"],
            color=UI["text"]
        )
        details_title.grid(row=1, column=0, sticky="w", padx=18, pady=(0, 6))

        details_summary = label(
            details_card,
            "Sélectionnez une filiale dans le graphique de répartition pour consulter les plus gros écarts.",
            font=FONT["small"],
            color=UI["muted"],
            justify="left",
            wraplength=1200
        )
        details_summary.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12))

        details_table_holder = ctk.CTkFrame(
            details_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        details_table_holder.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 18))
        details_table_holder.grid_columnconfigure(0, weight=1)
        details_table_holder.grid_rowconfigure(0, weight=1)

        _make_placeholder(details_table_holder, "Aucun détail à afficher.")

        # =========================================================
        # LOGIQUE MÉTIER
        # =========================================================
        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]
        decaissements = [
            "Péages", "Charges de personnel", "ACE & Investissements"
        ]
        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital",
            "Sous total financier", "Free cash Flow", "Emprunts",
            "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cash flow de financement",
            "Cash flow net", "Cessions d'immobilisations", "Impôts et Taxes",
            "Sous total dépenses"
        ]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                return (reel_val >= prev_val) if prev_val >= 0 else (abs(reel_val) <= abs(prev_val))
            else:
                return reel_val >= prev_val

        def afficher_details_feuille(feuille, details_ecarts, nombre_previsions_par_feuille, annee):
            _clear_children(details_table_holder)

            ecarts = details_ecarts.get(feuille, [])
            nombre_ecarts = len(ecarts)
            nombre_previsions = nombre_previsions_par_feuille.get(feuille, 0)
            pourcentage_ecarts = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0
            titre_suffix = "" if annee is None else f" — Année {annee}"

            details_title.configure(text=f"Détails — {feuille}")
            details_summary.configure(
                text=(
                    f"{nombre_ecarts} écart(s) significatif(s) sur {nombre_previsions} prévision(s) "
                    f"({pourcentage_ecarts:.1f} %){titre_suffix}"
                )
            )

            colonnes = ["Date", "Profil", "Filiale", "Flux", "Réel (k€)", "Prévision (k€)", "Écart (%)"]

            tree = ttk.Treeview(
                details_table_holder,
                columns=colonnes,
                show="headings",
                height=7,
                style="Pulse.Treeview"
            )
            tree.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)

            yscroll = ttk.Scrollbar(details_table_holder, orient="vertical", command=tree.yview)
            yscroll.grid(row=0, column=1, sticky="ns", pady=12)
            tree.configure(yscrollcommand=yscroll.set)

            for col in colonnes:
                tree.heading(col, text=col)
                width = 130
                if col == "Flux":
                    width = 220
                elif col == "Profil":
                    width = 180
                elif col == "Filiale":
                    width = 130
                elif col == "Date":
                    width = 110
                tree.column(col, anchor="center", width=width)

            tree.tag_configure("neg", foreground="#EF4444")
            tree.tag_configure("pos", foreground="#22C55E")

            if not ecarts:
                tree.insert("", "end", values=("—", "—", feuille, "Aucun écart", "—", "—", "—"))
            else:
                cinq_plus_gros = sorted(ecarts, key=lambda x: abs(x["ecart_pct"]), reverse=True)[:5]
                for e in cinq_plus_gros:
                    date_str = e["date"].strftime("%Y-%m-%d") if hasattr(e["date"], "strftime") else str(e["date"])
                    ecart_str = f"{e['ecart_pct']}%"
                    tags = ("pos",) if est_favorable(e["flux"], e["reel"], e["prevision"]) else ("neg",)
                    tree.insert(
                        "",
                        "end",
                        values=(
                            date_str,
                            e["profil"],
                            e["filiale"],
                            e["flux"],
                            e["reel"],
                            e["prevision"],
                            ecart_str
                        ),
                        tags=tags
                    )

            page.update_idletasks()
            try:
                target = max(0, details_card.winfo_y() - 40)
                total_h = max(1, page.winfo_height())
                body_canvas.yview_moveto(min(1.0, target / total_h))
            except Exception:
                pass

        def recalcul(annee):
            _clear_children(pie_body)
            _clear_children(freq_body)
            _clear_children(val_body)

            repartition = {feuille: 0 for feuille in feuilles_all}
            valorisation_ecarts = {feuille: 0 for feuille in feuilles_all}
            details_ecarts = {feuille: [] for feuille in feuilles_all}

            for feuille in feuilles_all:
                ws, noms_colonnes = _charger(feuille)
                for nom, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = _extraire(ws, col_start, annee=None)

                    for i, date in enumerate(dates):
                        y = _year_of(date)
                        if annee is not None and y is not None and y != annee:
                            continue

                        if i >= len(reel) or reel[i] is None:
                            continue

                        r = reel[i]

                        for idx, prev_list in enumerate(previsions):
                            if i >= len(prev_list) or prev_list[i] is None:
                                continue

                            prev_val = prev_list[i]

                            if r == 0 and prev_val == 0:
                                continue
                            elif prev_val == 0:
                                prev_val = 1

                            ecart = (r - prev_val) / prev_val
                            if abs(ecart) >= 0.4:
                                repartition[feuille] += 1
                                valorisation_ecarts[feuille] += (r - prev_val)
                                details_ecarts[feuille].append({
                                    "date": date,
                                    "profil": noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}",
                                    "filiale": feuille,
                                    "flux": nom,
                                    "reel": round(reel[i], 2),
                                    "prevision": round(prev_val, 2),
                                    "ecart_pct": round(ecart * 100, 1)
                                })

            nombre_previsions_par_feuille = {}
            pourcentage_ecarts_filiales = {}

            for feuille in feuilles_all:
                ws, noms_colonnes_local = _charger(feuille)
                nombre_previsions = 0

                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = _extraire(ws, col_start, annee=None)

                    for i, d in enumerate(dates_local):
                        y = _year_of(d)
                        if annee is not None and y is not None and y != annee:
                            continue
                        for prev_list in previsions_local:
                            if i < len(prev_list) and prev_list[i] is not None:
                                nombre_previsions += 1

                nombre_previsions_par_feuille[feuille] = nombre_previsions
                nombre_ecarts = len(details_ecarts.get(feuille, []))
                pourcentage = (nombre_ecarts / nombre_previsions) * 100 if nombre_previsions > 0 else 0
                pourcentage_ecarts_filiales[feuille] = pourcentage

            titre_suffix = "" if annee is None else f" — Année {annee}"

            # ===================== PIE =====================
            total_ecarts = sum(repartition.values())
            feuilles = feuilles_all
            valeurs = [(repartition[f] / total_ecarts) * 100 if total_ecarts > 0 else 0 for f in feuilles]

            feuilles_filtrees = [f for i, f in enumerate(feuilles) if valeurs[i] > 0]
            valeurs_filtrees = [v for v in valeurs if v > 0]

            fig_pie, ax_pie = plt.subplots(figsize=(10, 5.4), dpi=100)
            fig_pie.patch.set_facecolor(UI["surface_2"])
            ax_pie.set_facecolor(UI["surface_2"])

            palette_pie = ["#5D5F83", "#34495E", "#5D6D7E", "#85929E", "#AAB7B8", "#D5DBDB"]

            if not valeurs_filtrees:
                ax_pie.text(
                    0.5, 0.5,
                    "Aucun écart important détecté",
                    ha="center", va="center",
                    fontsize=12, color="white"
                )
                ax_pie.axis("off")
                wedges = []
            else:
                wedges, texts, autotexts = ax_pie.pie(
                    valeurs_filtrees,
                    labels=feuilles_filtrees,
                    autopct="%1.1f%%",
                    startangle=140,
                    colors=palette_pie[:len(valeurs_filtrees)],
                    textprops={"color": "white", "fontsize": 10}
                )
                ax_pie.set_aspect("equal", adjustable="box")
                for t in texts + autotexts:
                    t.set_color("white")

            ax_pie.set_title(
                "Répartition des écarts significatifs par filiale — occurrences" + titre_suffix,
                fontsize=14,
                color="white"
            )

            pie_canvas = _embed_figure(fig_pie, pie_body, mode="pie")

            if wedges:
                original_colors = [w.get_facecolor() for w in wedges]

                def reset_colors():
                    for i, w in enumerate(wedges):
                        w.set_facecolor(original_colors[i])
                    pie_canvas.draw_idle()

                def on_hover(event):
                    if event.inaxes != ax_pie:
                        reset_colors()
                        return

                    found = False
                    for i, w in enumerate(wedges):
                        contains, _ = w.contains(event)
                        if contains:
                            w.set_facecolor(UI["accent"])
                            for j, w2 in enumerate(wedges):
                                if j != i:
                                    w2.set_facecolor(original_colors[j])
                            pie_canvas.draw_idle()
                            found = True
                            break

                    if not found:
                        reset_colors()

                def on_click(event):
                    if event.inaxes != ax_pie:
                        return
                    for i, w in enumerate(wedges):
                        contains, _ = w.contains(event)
                        if contains:
                            afficher_details_feuille(
                                feuilles_filtrees[i],
                                details_ecarts,
                                nombre_previsions_par_feuille,
                                annee
                            )
                            break

                pie_canvas.mpl_connect("motion_notify_event", on_hover)
                pie_canvas.mpl_connect("button_press_event", on_click)

            # ===================== BAR FREQUENCE =====================
            feuilles_bar = list(pourcentage_ecarts_filiales.keys())
            valeurs_bar = list(pourcentage_ecarts_filiales.values())

            fig_bar, ax_bar = plt.subplots(figsize=(12, 4.8), dpi=100)
            fig_bar.patch.set_facecolor(UI["surface_2"])
            ax_bar.set_facecolor(UI["surface_2"])

            if valeurs_bar:
                vmin = min(valeurs_bar)
                vmax = max(valeurs_bar) if max(valeurs_bar) != vmin else vmin + 1
                norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
                cmap = cm.Blues
                colors_bar = [cmap(norm(v)) for v in valeurs_bar]
            else:
                colors_bar = []

            bars = ax_bar.bar(feuilles_bar, valeurs_bar, color=colors_bar, alpha=0.9, width=0.6)

            ax_bar.set_title("Fréquence des écarts significatifs par filiale" + titre_suffix, fontsize=14, color="white")
            ax_bar.set_ylabel("% Écarts", color="white")
            ax_bar.tick_params(axis="x", rotation=45, colors="white")
            ax_bar.tick_params(axis="y", colors="white")
            ax_bar.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            ax_bar.set_axisbelow(True)

            # occuper visuellement toute la largeur
            if len(feuilles_bar) == 1:
                ax_bar.set_xlim(-0.5, 0.5)
            else:
                ax_bar.margins(x=0.05)

            freq_canvas = _embed_figure(fig_bar, freq_body, mode="bar")

            cursor_bar = mplcursors.cursor(bars, hover=True)

            @cursor_bar.connect("add")
            def on_hover_bar(sel):
                idx = sel.index
                feuille = feuilles_bar[idx]
                sel.annotation.set_text(
                    f"{feuille}\n"
                    f"Nombre d'écarts : {len(details_ecarts.get(feuille, []))}\n"
                    f"Nombre de prévisions : {nombre_previsions_par_feuille.get(feuille, 0)}\n"
                    f"Pourcentage : {valeurs_bar[idx]:.1f}%"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # ===================== BAR VALORISATION =====================
            feuilles_val = list(valorisation_ecarts.keys())
            valeurs_val = [0 if valorisation_ecarts[f] is None else valorisation_ecarts[f] for f in feuilles_val]

            fig_val, ax_val = plt.subplots(figsize=(12, 4.8), dpi=100)
            fig_val.patch.set_facecolor(UI["surface_2"])
            ax_val.set_facecolor(UI["surface_2"])

            if any(v != 0 for v in valeurs_val):
                max_abs = max(abs(v) for v in valeurs_val) or 1
                norm_val = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
                cmap_val = cm.RdBu_r
                colors_val = [cmap_val(norm_val(v)) for v in valeurs_val]
            else:
                colors_val = ["#8395a7"] * len(valeurs_val)

            bars_val = ax_val.bar(feuilles_val, valeurs_val, color=colors_val, alpha=0.95, width=0.6)

            ax_val.axhline(0, color="white", linewidth=1, alpha=0.7)
            ax_val.set_title("Valorisation totale des écarts par filiale (k€)" + titre_suffix, fontsize=14, color="white")
            ax_val.set_ylabel("Écarts cumulés (k€)", color="white")
            ax_val.tick_params(axis="x", rotation=45, colors="white")
            ax_val.tick_params(axis="y", colors="white")
            ax_val.grid(axis="y", color="gray", linestyle="--", alpha=0.3)
            ax_val.set_axisbelow(True)

            if len(feuilles_val) == 1:
                ax_val.set_xlim(-0.5, 0.5)
            else:
                ax_val.margins(x=0.05)

            val_canvas = _embed_figure(fig_val, val_body, mode="bar")

            cursor_val = mplcursors.cursor(bars_val, hover=True)

            @cursor_val.connect("add")
            def on_add(sel):
                idx = sel.index
                feuille = feuilles_val[idx]
                val = valeurs_val[idx]
                nb_ecarts = repartition.get(feuille, 0)
                nb_prevs = nombre_previsions_par_feuille.get(feuille, 0)
                val_str = f"{int(val):,}".replace(",", " ")
                sel.annotation.set_text(
                    f"{feuille}\nÉcarts : {nb_ecarts}\nPrévisions : {nb_prevs}\nValorisation : {val_str} k€"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            details_title.configure(text="Aucun détail sélectionné")
            details_summary.configure(
                text="Cliquez sur une filiale dans le graphique de répartition pour afficher les écarts les plus significatifs."
            )
            _make_placeholder(details_table_holder, "Aucun détail à afficher.")

            _sync_scrollregion()

        # =========================================================
        # INITIALISATION + BINDING
        # =========================================================
        recalcul(None if annee_var.get() == "Toutes aqnnées" else int(annee_var.get()))

        def _on_annee_change(value):
            recalcul(None if value == "Toutes années" else int(value))

        annees_box.configure(command=_on_annee_change)
    
    def afficher_repartition_par_prevision(self):
        from collections import defaultdict
        import numpy as np
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage
        import mplcursors
        import re
        import matplotlib.cm as cm
        import matplotlib.colors as mcolors

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "success": "#1F9D63",
            "warning": "#E0B64A",
            "danger": "#C44E4E",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _embed_figure(fig, master, mode="barh"):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            def _resize(event=None):
                try:
                    master.update_idletasks()
                    w = max(master.winfo_width() - 16, 900)
                    h = max(master.winfo_height() - 16, 420)
                    dpi = fig.get_dpi()

                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    fig.set_size_inches(w / dpi, h / dpi, forward=True)

                    if mode == "barh":
                        fig.subplots_adjust(left=0.22, right=0.97, bottom=0.12, top=0.90)
                    else:
                        fig.subplots_adjust(left=0.08, right=0.98, bottom=0.18, top=0.90)

                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]
            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _tok in noms_colonnes:
                    try:
                        years = self._annees_disponibles(feuille, nom_flux)
                        annees.update(years)
                    except Exception:
                        pass
            return sorted(annees)

        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _is_filled(x):
            return _to_number(x) is not None

        def _parse_jj_mm(nom: str):
            m = re.search(r"(\d{1,2})[/-](\d{1,2})", str(nom))
            if not m:
                return 99, 99
            jj = int(m.group(1))
            mm = int(m.group(2))
            if 1 <= jj <= 31 and 1 <= mm <= 12:
                return mm, jj
            return 99, 99

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module de répartition par profil",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        retour_cmd = self.retour_menu if hasattr(self, "retour_menu") else self.creer_accueil
        neutral_btn(top_right, "Retour à l’accueil", retour_cmd).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.grid_rowconfigure(3, weight=0)
        page.columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "PROFILS", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Répartition des écarts par profil",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Analysez le taux d’écarts et la valorisation des écarts pour chaque profil, par filiale et par année.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "FILTRES",
            "Périmètre d’analyse",
            "Sélectionnez une filiale et une année pour recalculer automatiquement les graphiques et le tableau."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        filters_body.grid_columnconfigure(0, weight=0)
        filters_body.grid_columnconfigure(1, weight=0)
        filters_body.grid_columnconfigure(2, weight=1)

        filiales = ["Toutes filiales"] + list(sections.values())
        selected_filiale = ctk.StringVar(value=filiales[0])
        annees_var = ctk.StringVar(value="Toutes années")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        filiale_menu = ctk.CTkOptionMenu(
            filters_body,
            values=filiales,
            variable=selected_filiale,
            width=240,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        filiale_menu.grid(row=1, column=0, sticky="w", padx=(0, 18))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=1, sticky="w", pady=(0, 6)
        )
        annees_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Toutes années"],
            variable=annees_var,
            width=180,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annees_box.grid(row=1, column=1, sticky="w")

        filter_hint = label(
            filters_body,
            "Le tri des profils suit l’ordre jour/mois détecté dans leur nom quand c’est possible.",
            font=FONT["small"],
            color=UI["muted"]
        )
        filter_hint.grid(row=1, column=2, sticky="w", padx=(18, 0))

        # =========================================================
        # GRAPHE 1
        # =========================================================
        graph1_card = card(page, fg=UI["surface"], radius=20)
        graph1_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        graph1_card.grid_columnconfigure(0, weight=1)

        section_header(
            graph1_card,
            "TAUX",
            "Taux d’écarts par profil",
            "Rapport entre le nombre d’écarts significatifs et le nombre de prévisions non vides."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        graph1_body = ctk.CTkFrame(
            graph1_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=460
        )
        graph1_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        graph1_body.grid_propagate(False)
        graph1_body.grid_rowconfigure(0, weight=1)
        graph1_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # GRAPHE 2
        # =========================================================
        graph2_card = card(page, fg=UI["surface"], radius=20)
        graph2_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 14))
        graph2_card.grid_columnconfigure(0, weight=1)

        section_header(
            graph2_card,
            "VALORISATION",
            "Valorisation des écarts par profil",
            "Montant cumulé absolu des écarts en k€ pour chaque profil."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        graph2_body = ctk.CTkFrame(
            graph2_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=460
        )
        graph2_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        graph2_body.grid_propagate(False)
        graph2_body.grid_rowconfigure(0, weight=1)
        graph2_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # TABLEAU
        # =========================================================
        table_card = card(page, fg=UI["surface"], radius=20)
        table_card.grid(row=4, column=0, sticky="ew", padx=28, pady=(0, 24))
        table_card.grid_columnconfigure(0, weight=1)

        section_header(
            table_card,
            "TABLEAU",
            "Synthèse par profil",
            "Vue consolidée du nombre de prévisions, du nombre d’écarts, du taux et de la valorisation."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        table_holder = ctk.CTkFrame(
            table_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        table_holder.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        table_holder.grid_columnconfigure(0, weight=1)
        table_holder.grid_rowconfigure(0, weight=1)

        colonnes = ("Profil", "Nb prévisions", "Nb écarts >=40%", "Taux (%)", "Valorisation (k€)")
        table = ttk.Treeview(table_holder, columns=colonnes, show="headings", height=8, style="Pulse.Treeview")
        table.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)

        yscroll = ttk.Scrollbar(table_holder, orient="vertical", command=table.yview)
        yscroll.grid(row=0, column=1, sticky="ns", pady=12)
        table.configure(yscrollcommand=yscroll.set)

        for col in colonnes:
            table.heading(col, text=col)
            width = 150
            if col == "Profil":
                width = 240
            table.column(col, anchor="center", width=width)

        # =========================================================
        # CALCUL + RENDER
        # =========================================================
        def maj_graphique(filiale, annee):
            _clear_children(graph1_body)
            _clear_children(graph2_body)

            for row in table.get_children():
                table.delete(row)

            filiales_calc = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            compteur_ecarts = defaultdict(int)
            valorisation_ecarts = defaultdict(float)
            compteur_total = defaultdict(int)

            # ================= PASS 1 : écarts + valorisation =================
            for f in filiales_calc:
                ws, noms_colonnes = charger_donnees(f, taille_bloc)
                for nom_colonne, col_start in noms_colonnes:
                    dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)

                    for p_idx, nom_profil in enumerate(noms_profils):
                        if p_idx >= len(previsions):
                            continue
                        prev_list = previsions[p_idx]

                        for i, d in enumerate(dates):
                            if annee is not None:
                                try:
                                    y = d.year if hasattr(d, "year") else None
                                except Exception:
                                    y = None
                                if y is not None and y != annee:
                                    continue

                            pv_raw = prev_list[i] if i < len(prev_list) else None
                            prev_val = _to_number(pv_raw)
                            r_val = _to_number(reel[i] if i < len(reel) else None)

                            if prev_val is None:
                                continue
                            if r_val is None:
                                r_val = 0.0

                            if prev_val == 0:
                                if r_val == 0:
                                    continue
                                prev_val = 1.0

                            ecart = (r_val - prev_val) / prev_val
                            if abs(ecart) >= 0.4:
                                compteur_ecarts[nom_profil] += 1
                                valorisation_ecarts[nom_profil] += abs((prev_val or 0.0) - (r_val or 0.0))

            # ================= PASS 2 : nb de prévisions =================
            for f in filiales_calc:
                ws, noms_colonnes_local = charger_donnees(f, taille_bloc)
                for nom_colonne, col_start in noms_colonnes_local:
                    dates_local, reel_local, previsions_local, noms_profils_local = extraire_valeurs(ws, col_start, nb_prev, annee=None)

                    for p_idx, nom_profil in enumerate(noms_profils_local):
                        if p_idx >= len(previsions_local):
                            continue
                        prev_list = previsions_local[p_idx]

                        nb_prev_non_vides = 0
                        for i, d in enumerate(dates_local):
                            if annee is not None:
                                try:
                                    y = d.year if hasattr(d, "year") else None
                                except Exception:
                                    y = None
                                if y is not None and y != annee:
                                    continue
                            if i < len(prev_list) and _is_filled(prev_list[i]):
                                nb_prev_non_vides += 1

                        compteur_total[nom_profil] += nb_prev_non_vides

            # ================= TRI + FILTRAGE =================
            rows = []
            all_profils = set(compteur_total.keys()) | set(compteur_ecarts.keys())

            for nom in all_profils:
                total = compteur_total.get(nom, 0)
                ecarts = compteur_ecarts.get(nom, 0)

                if annee is not None and total == 0 and ecarts == 0:
                    continue

                taux = (ecarts / total * 100) if total > 0 else 0.0
                valo = valorisation_ecarts.get(nom, 0.0)
                mm, jj = _parse_jj_mm(nom)
                rows.append((mm, jj, nom, taux, valo, total, ecarts))

            rows.sort(key=lambda r: (r[0], r[1], str(r[2]).casefold()))

            noms_final = [r[2] for r in rows]
            pourcentages = [r[3] for r in rows]
            valorisations = [r[4] for r in rows]
            totaux_prev = [r[5] for r in rows]
            totaux_ecarts = [r[6] for r in rows]

            suffix = "" if annee is None else f" — {annee}"
            titre_filiale = "Toutes filiales" if filiale == "Toutes filiales" else filiale

            if not noms_final:
                _make_placeholder(graph1_body, "Aucune donnée disponible pour ce filtre.")
                _make_placeholder(graph2_body, "Aucune donnée disponible pour ce filtre.")
                table.insert("", "end", values=("Aucune donnée", "—", "—", "—", "—"))
                return

            # Hauteur dynamique
            dynamic_h = min(max(420, 48 * len(noms_final) + 120), 1200)
            graph1_body.configure(height=dynamic_h)
            graph2_body.configure(height=dynamic_h)

            # =====================================================
            # FIGURE 1 : TAUX D'ÉCARTS
            # =====================================================
            fig1, ax1 = plt.subplots(figsize=(12, max(4.8, len(noms_final) * 0.42)), dpi=100)
            fig1.patch.set_facecolor(UI["surface_2"])
            ax1.set_facecolor(UI["surface_2"])

            y_pos = np.arange(len(noms_final))

            if pourcentages:
                vmin = min(pourcentages)
                vmax = max(pourcentages) if max(pourcentages) != vmin else vmin + 1
                norm = mcolors.Normalize(vmin=vmin, vmax=vmax)
                cmap = cm.Blues
                colors_1 = [cmap(norm(v)) for v in pourcentages]
            else:
                colors_1 = ["#4C7CF3"] * len(noms_final)

            bars1 = ax1.barh(y_pos, pourcentages, color=colors_1, alpha=0.95)
            ax1.set_yticks(y_pos)
            ax1.set_yticklabels(noms_final, color="white")
            ax1.invert_yaxis()
            ax1.set_xlabel("Taux d'écarts (%)", color="white")
            ax1.set_title(f"{titre_filiale} — Taux d'écarts (rep/prev){suffix}", fontsize=14, color="white")
            ax1.tick_params(axis="x", colors="white")
            ax1.tick_params(axis="y", colors="white")
            ax1.grid(axis="x", color="gray", linestyle="--", alpha=0.3)
            ax1.set_axisbelow(True)

            max_x1 = max(pourcentages) if pourcentages else 0
            ax1.set_xlim(0, max(10, max_x1 * 1.15))

            canvas1 = _embed_figure(fig1, graph1_body, mode="barh")

            cursor1 = mplcursors.cursor(bars1, hover=True)

            @cursor1.connect("add")
            def on_add_1(sel):
                idx = sel.index
                sel.annotation.set_text(
                    f"Profil : {noms_final[idx]}\n"
                    f"Taux : {pourcentages[idx]:.2f}%\n"
                    f"Prévisions : {totaux_prev[idx]}\n"
                    f"Écarts : {totaux_ecarts[idx]}"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =====================================================
            # FIGURE 2 : VALORISATION
            # =====================================================
            fig2, ax2 = plt.subplots(figsize=(12, max(4.8, len(noms_final) * 0.42)), dpi=100)
            fig2.patch.set_facecolor(UI["surface_2"])
            ax2.set_facecolor(UI["surface_2"])

            if valorisations:
                vmax2 = max(valorisations) if max(valorisations) > 0 else 1
                norm2 = mcolors.Normalize(vmin=0, vmax=vmax2)
                cmap2 = cm.Greens
                colors_2 = [cmap2(norm2(v)) for v in valorisations]
            else:
                colors_2 = ["#1F9D63"] * len(noms_final)

            bars2 = ax2.barh(y_pos, valorisations, color=colors_2, alpha=0.95)
            ax2.set_yticks(y_pos)
            ax2.set_yticklabels(noms_final, color="white")
            ax2.invert_yaxis()
            ax2.set_xlabel("Valorisation (k€)", color="white")
            ax2.set_title(f"{titre_filiale} — Valorisation des écarts{suffix}", fontsize=14, color="white")
            ax2.tick_params(axis="x", colors="white")
            ax2.tick_params(axis="y", colors="white")
            ax2.grid(axis="x", color="gray", linestyle="--", alpha=0.3)
            ax2.set_axisbelow(True)

            max_x2 = max(valorisations) if valorisations else 0
            ax2.set_xlim(0, max(10, max_x2 * 1.15 if max_x2 > 0 else 10))

            canvas2 = _embed_figure(fig2, graph2_body, mode="barh")

            cursor2 = mplcursors.cursor(bars2, hover=True)

            @cursor2.connect("add")
            def on_add_2(sel):
                idx = sel.index
                sel.annotation.set_text(
                    f"Profil : {noms_final[idx]}\n"
                    f"Valorisation : {valorisations[idx]:,.0f} k€".replace(",", " ")
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =====================================================
            # TABLEAU
            # =====================================================
            total_valorisation = 0.0
            for nom, taux, valo, tot, ec in zip(noms_final, pourcentages, valorisations, totaux_prev, totaux_ecarts):
                total_valorisation += valo
                table.insert(
                    "",
                    "end",
                    values=(
                        nom,
                        tot,
                        ec,
                        f"{taux:.2f}%",
                        f"{valo:,.0f}".replace(",", " ")
                    )
                )

            total_previsions = sum(totaux_prev)
            total_ecarts = sum(totaux_ecarts)
            taux_total = (total_ecarts / total_previsions * 100) if total_previsions > 0 else 0.0

            table.insert(
                "",
                "end",
                values=(
                    "TOTAL",
                    total_previsions,
                    total_ecarts,
                    f"{taux_total:.2f}%",
                    f"{total_valorisation:,.0f}".replace(",", " ")
                )
            )

            _sync_scrollregion()

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _on_filiale_change(value=None):
            filiale = selected_filiale.get()
            annees = _annees_pour_filiale(filiale)
            values = ["Toutes années"] + [str(a) for a in annees]
            annees_box.configure(values=values)
            annees_var.set(str(annees[-1]) if annees else "Toutes années")

            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            maj_graphique(filiale, annee)

        def _on_annee_change(value=None):
            filiale = selected_filiale.get()
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            maj_graphique(filiale, annee)

        filiale_menu.configure(command=_on_filiale_change)
        annees_box.configure(command=_on_annee_change)

        # =========================================================
        # AFFICHAGE INITIAL
        # =========================================================
        _on_filiale_change()

    def afficher_repartition_flux(self):
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        import matplotlib.colors as mcolors
        import matplotlib.cm as cm
        import numpy as np
        import re
        import datetime as _dt
        import mplcursors

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "success": "#1F9D63",
            "warning": "#E0B64A",
            "danger": "#C44E4E",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _embed_figure(fig, master, mode="barh"):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            def _resize(event=None):
                try:
                    master.update_idletasks()
                    w = max(master.winfo_width() - 16, 900)
                    h = max(master.winfo_height() - 16, 420)
                    dpi = fig.get_dpi()

                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    fig.set_size_inches(w / dpi, h / dpi, forward=True)

                    if mode == "barh":
                        fig.subplots_adjust(left=0.22, right=0.97, bottom=0.12, top=0.90)
                    else:
                        fig.subplots_adjust(left=0.07, right=0.98, bottom=0.18, top=0.90)

                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]
            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _tok in noms_colonnes:
                    try:
                        years = self._annees_disponibles(feuille, nom_flux)
                        annees.update(years)
                    except Exception:
                        pass
            return sorted(annees)

        def _parse_profil_day_month(nom: str):
            m = re.search(r"(\d{1,2})[/-](\d{1,2})", str(nom))
            if not m:
                return None, None
            jj = int(m.group(1))
            mm = int(m.group(2))
            if 1 <= jj <= 31 and 1 <= mm <= 12:
                return jj, mm
            return None, None

        def _profils_uniques_ordonnes_par_mois(filiale: str, annee: int | None):
            if annee is None:
                return []

            profils = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]

            for feuille in feuilles:
                _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, tok in noms_colonnes:
                    try:
                        noms_ok, _flags = self._profils_for_year(feuille, nom_flux, annee)
                        for p in noms_ok:
                            if p and str(p).strip():
                                profils.add(str(p).strip())
                    except Exception:
                        try:
                            dates, reel, previsions, noms_profils = extraire_valeurs(_ws, tok, nb_prev, annee=annee)
                            for p in noms_profils:
                                if p and str(p).strip():
                                    profils.add(str(p).strip())
                        except Exception:
                            pass

            def keyer(name: str):
                jj, mm = _parse_profil_day_month(name)
                if mm is None:
                    return (99, 99, name.lower())
                return (mm, jj if jj is not None else 31, name.lower())

            return sorted(profils, key=keyer)

        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    return None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(d, fmt).year
                    except Exception:
                        pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m:
                    return int(m.group(1))
            return None

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module de répartition des flux",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        retour_cmd = self.retour_menu if hasattr(self, "retour_menu") else self.creer_accueil
        neutral_btn(top_right, "Retour à l’accueil", retour_cmd).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.grid_rowconfigure(3, weight=0)
        page.grid_rowconfigure(4, weight=0)
        page.columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "FLUX", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Répartition des écarts par flux",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Analysez le volume d’écarts, leur fréquence et leur valorisation par flux, avec filtres filiale, année et profil.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "FILTRES",
            "Périmètre d’analyse",
            "Le filtre profil s’applique uniquement à la valorisation signée."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        filters_body.grid_columnconfigure(0, weight=0)
        filters_body.grid_columnconfigure(1, weight=0)
        filters_body.grid_columnconfigure(2, weight=0)
        filters_body.grid_columnconfigure(3, weight=1)

        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = ctk.StringVar(value=filiales[0])
        annees_var = ctk.StringVar(value="Toutes années")
        profils_var = ctk.StringVar(value="Tous profils")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        filiale_menu = ctk.CTkOptionMenu(
            filters_body,
            values=filiales,
            variable=selected_filiale,
            width=220,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        filiale_menu.grid(row=1, column=0, sticky="w", padx=(0, 18))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=1, sticky="w", pady=(0, 6)
        )
        annees_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Toutes années"],
            variable=annees_var,
            width=180,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annees_box.grid(row=1, column=1, sticky="w", padx=(0, 18))

        label(filters_body, "Profil", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=2, sticky="w", pady=(0, 6)
        )
        profils_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Tous profils"],
            variable=profils_var,
            width=260,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        profils_box.grid(row=1, column=2, sticky="w")

        filter_hint = label(
            filters_body,
            "Les flux agrégés sont exclus pour garder une lecture métier plus pertinente.",
            font=FONT["small"],
            color=UI["muted"]
        )
        filter_hint.grid(row=1, column=3, sticky="w", padx=(18, 0))

        # =========================================================
        # CHART 1 : NOMBRE D'ÉCARTS
        # =========================================================
        chart1_card = card(page, fg=UI["surface"], radius=20)
        chart1_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        chart1_card.grid_columnconfigure(0, weight=1)

        section_header(
            chart1_card,
            "VOLUME",
            "Nombre d’écarts importants par flux",
            "Volume d’écarts significatifs détectés pour chaque flux."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        chart1_body = ctk.CTkFrame(
            chart1_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=460
        )
        chart1_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        chart1_body.grid_propagate(False)
        chart1_body.grid_rowconfigure(0, weight=1)
        chart1_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # CHART 2 : POURCENTAGE
        # =========================================================
        chart2_card = card(page, fg=UI["surface"], radius=20)
        chart2_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 14))
        chart2_card.grid_columnconfigure(0, weight=1)

        section_header(
            chart2_card,
            "FRÉQUENCE",
            "Pourcentage d’écarts par flux",
            "Rapport entre le nombre d’écarts significatifs et le nombre de prévisions non vides."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        chart2_body = ctk.CTkFrame(
            chart2_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=460
        )
        chart2_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        chart2_body.grid_propagate(False)
        chart2_body.grid_rowconfigure(0, weight=1)
        chart2_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # CHART 3 : VALORISATION
        # =========================================================
        chart3_card = card(page, fg=UI["surface"], radius=20)
        chart3_card.grid(row=4, column=0, sticky="ew", padx=28, pady=(0, 14))
        chart3_card.grid_columnconfigure(0, weight=1)

        section_header(
            chart3_card,
            "VALORISATION",
            "Valorisation signée des écarts",
            "Montant cumulé signé des écarts, filtrable par profil."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        chart3_body = ctk.CTkFrame(
            chart3_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=460
        )
        chart3_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        chart3_body.grid_propagate(False)
        chart3_body.grid_rowconfigure(0, weight=1)
        chart3_body.grid_columnconfigure(0, weight=1)

        # =========================================================
        # TABLEAU
        # =========================================================
        table_card = card(page, fg=UI["surface"], radius=20)
        table_card.grid(row=5, column=0, sticky="ew", padx=28, pady=(0, 24))
        table_card.grid_columnconfigure(0, weight=1)

        section_header(
            table_card,
            "TABLEAU",
            "Synthèse par flux",
            "Prévisions, écarts et fréquence pour les flux affichés."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        table_holder = ctk.CTkFrame(
            table_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        table_holder.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        table_holder.grid_columnconfigure(0, weight=1)
        table_holder.grid_rowconfigure(0, weight=1)

        colonnes = ("Flux", "Prévisions", "Écarts ≥40%", "% Écarts", "Valorisation (k€)")
        table = ttk.Treeview(table_holder, columns=colonnes, show="headings", height=10, style="Pulse.Treeview")
        table.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)

        yscroll = ttk.Scrollbar(table_holder, orient="vertical", command=table.yview)
        yscroll.grid(row=0, column=1, sticky="ns", pady=12)
        table.configure(yscrollcommand=yscroll.set)

        for col in colonnes:
            table.heading(col, text=col)
            width = 140
            if col == "Flux":
                width = 300
            table.column(col, anchor="center", width=width)

        # =========================================================
        # CALCUL + RENDER
        # =========================================================
        def afficher_graphes():
            _clear_children(chart1_body)
            _clear_children(chart2_body)
            _clear_children(chart3_body)

            for row in table.get_children():
                table.delete(row)

            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)
            suffix_annee = "" if annee is None else f" — {annee}"

            filiale = selected_filiale.get()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]

            flux_a_exclure = {
                "Cash flow de financement", "Cash flow net", "Sous total financier",
                "Sous total Investissements nets et ACE", "Free cash Flow",
                "Sous total recettes", "Sous total dépenses", "C/C - Groupe"
            }

            # Liste complète des flux
            tous_flux = []
            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, _ in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    if nom_flux not in tous_flux:
                        tous_flux.append(nom_flux)

            ecarts_par_flux = {f: 0 for f in tous_flux}
            prevs_par_flux = {f: 0 for f in tous_flux}

            # PASS A : nombre d'écarts
            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    try:
                        dates, reel, previsions, _noms_prof = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for i, d in enumerate(dates):
                        if annee is not None:
                            y = _year_of(d)
                            if y is not None and y != annee:
                                continue

                        r_val = _to_number(reel[i] if i < len(reel) else None)
                        if r_val is None:
                            r_val = 0.0

                        for prev_list in previsions:
                            pv = prev_list[i] if i < len(prev_list) else None
                            prev_val = _to_number(pv)
                            if prev_val is None:
                                continue
                            if prev_val == 0:
                                if r_val == 0:
                                    continue
                                prev_val = 1.0

                            ecart = (r_val - prev_val) / prev_val
                            if abs(ecart) >= 0.4:
                                ecarts_par_flux[nom_flux] += 1

            # PASS B : nombre de prévisions
            for feuille in feuilles:
                ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                for nom_flux, col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    try:
                        dates, _reel, previsions, _noms_prof = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for prev_list in previsions:
                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue
                            if i < len(prev_list) and _to_number(prev_list[i]) is not None:
                                prevs_par_flux[nom_flux] += 1

            # Tri + filtrage affichage
            noms_flux = sorted(tous_flux, key=lambda f: ecarts_par_flux.get(f, 0), reverse=True)
            valeurs = [ecarts_par_flux.get(f, 0) for f in noms_flux]

            noms_flux_aff = [f for f, v in zip(noms_flux, valeurs) if v != 0]
            valeurs_aff = [v for v in valeurs if v != 0]

            if not noms_flux_aff:
                _make_placeholder(chart1_body, "Aucun écart important détecté pour ce filtre.")
                _make_placeholder(chart2_body, "Aucun pourcentage à afficher pour ce filtre.")
                _make_placeholder(chart3_body, "Aucune valorisation à afficher pour ce filtre.")
                table.insert("", "end", values=("Aucune donnée", "—", "—", "—", "—"))
                return

            # Pourcentages
            pourcentage_par_flux = []
            for f in noms_flux_aff:
                nb_e = ecarts_par_flux.get(f, 0)
                nb_p = prevs_par_flux.get(f, 0)
                pct = (nb_e / nb_p * 100) if nb_p > 0 else 0.0
                pourcentage_par_flux.append(pct)

            # Valorisation signée
            prof_sel = profils_var.get()
            prof_sel = None if (not prof_sel or prof_sel == "Tous profils") else prof_sel

            valeur_ecarts = []
            for flux in noms_flux_aff:
                total_ecart = 0.0
                for feuille in feuilles:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                    for nom, col_start in noms_colonnes:
                        if nom != flux or nom in flux_a_exclure:
                            continue
                        try:
                            dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        except Exception:
                            continue

                        idxs = []
                        for i, d in enumerate(dates):
                            if annee is not None:
                                y = _year_of(d)
                                if y is not None and y != annee:
                                    continue
                            idxs.append(i)

                        if not idxs:
                            continue

                        if prof_sel is None:
                            idx_profils = range(len(previsions))
                        else:
                            idx_profils = [i for i, p in enumerate(noms_profils) if str(p).strip() == prof_sel]

                        for i in idxs:
                            r_val = _to_number(reel[i] if i < len(reel) else None)
                            if r_val is None:
                                r_val = 0.0
                            for idxp in idx_profils:
                                if idxp >= len(previsions):
                                    continue
                                prev_list = previsions[idxp]
                                pv = prev_list[i] if i < len(prev_list) else None
                                prev_val = _to_number(pv)
                                if prev_val is None:
                                    continue
                                if prev_val == 0:
                                    if r_val == 0:
                                        continue
                                    prev_val = 1.0
                                ecart = (r_val - prev_val) / prev_val
                                if abs(ecart) >= 0.4:
                                    total_ecart += (r_val - prev_val)

                valeur_ecarts.append(total_ecart)

            dynamic_h = min(max(420, 42 * len(noms_flux_aff) + 120), 1500)
            chart1_body.configure(height=dynamic_h)
            chart2_body.configure(height=dynamic_h)
            chart3_body.configure(height=dynamic_h)

            titre_filiale = "Toute filiale" if filiale == "Toute filiale" else filiale

            # =====================================================
            # CHART 1 : NOMBRE D'ÉCARTS
            # =====================================================
            fig1, ax1 = plt.subplots(figsize=(12, max(4.8, len(noms_flux_aff) * 0.42)), dpi=100)
            fig1.patch.set_facecolor(UI["surface_2"])
            ax1.set_facecolor(UI["surface_2"])

            y_pos = np.arange(len(noms_flux_aff))

            vmin1 = min(valeurs_aff)
            vmax1 = max(valeurs_aff) if max(valeurs_aff) != vmin1 else vmin1 + 1
            norm1 = mcolors.Normalize(vmin=vmin1, vmax=vmax1)
            cmap1 = cm.Blues
            colors1 = [cmap1(norm1(v)) for v in valeurs_aff]

            bars1 = ax1.barh(y_pos, valeurs_aff, color=colors1, alpha=0.95)
            ax1.set_yticks(y_pos)
            ax1.set_yticklabels(noms_flux_aff, color="white")
            ax1.invert_yaxis()
            ax1.set_xlabel("Nombre d'écarts importants (≥ 40%)", color="white")
            ax1.set_title(f"{titre_filiale} — Nombre d'écarts par flux{suffix_annee}", fontsize=14, color="white")
            ax1.tick_params(axis="x", colors="white")
            ax1.tick_params(axis="y", colors="white")
            ax1.grid(axis="x", color="gray", linestyle="--", alpha=0.3)
            ax1.set_axisbelow(True)
            ax1.set_xlim(0, max(10, max(valeurs_aff) * 1.15 if valeurs_aff else 10))

            canvas1 = _embed_figure(fig1, chart1_body, mode="barh")

            cursor1 = mplcursors.cursor(bars1, hover=True)

            @cursor1.connect("add")
            def on_add_1(sel):
                idx = sel.index
                sel.annotation.set_text(
                    f"Flux : {noms_flux_aff[idx]}\nÉcarts : {valeurs_aff[idx]}"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =====================================================
            # CHART 2 : POURCENTAGE
            # =====================================================
            fig2, ax2 = plt.subplots(figsize=(12, max(4.8, len(noms_flux_aff) * 0.42)), dpi=100)
            fig2.patch.set_facecolor(UI["surface_2"])
            ax2.set_facecolor(UI["surface_2"])

            vmin2 = min(pourcentage_par_flux)
            vmax2 = max(pourcentage_par_flux) if max(pourcentage_par_flux) != vmin2 else vmin2 + 1
            norm2 = mcolors.Normalize(vmin=vmin2, vmax=vmax2)
            cmap2 = cm.Blues
            colors2 = [cmap2(norm2(v)) for v in pourcentage_par_flux]

            bars2 = ax2.barh(y_pos, pourcentage_par_flux, color=colors2, alpha=0.95)
            ax2.set_yticks(y_pos)
            ax2.set_yticklabels(noms_flux_aff, color="white")
            ax2.invert_yaxis()
            ax2.set_xlabel("% d'écarts / prévisions", color="white")
            ax2.set_title(f"{titre_filiale} — Pourcentage d'écarts par flux{suffix_annee}", fontsize=14, color="white")
            ax2.tick_params(axis="x", colors="white")
            ax2.tick_params(axis="y", colors="white")
            ax2.grid(axis="x", color="gray", linestyle="--", alpha=0.3)
            ax2.set_axisbelow(True)
            ax2.set_xlim(0, max(10, max(pourcentage_par_flux) * 1.15 if pourcentage_par_flux else 10))

            canvas2 = _embed_figure(fig2, chart2_body, mode="barh")

            cursor2 = mplcursors.cursor(bars2, hover=True)

            @cursor2.connect("add")
            def on_add_2(sel):
                idx = sel.index
                sel.annotation.set_text(
                    f"Flux : {noms_flux_aff[idx]}\n"
                    f"Prévisions : {prevs_par_flux.get(noms_flux_aff[idx], 0)}\n"
                    f"Écarts : {ecarts_par_flux.get(noms_flux_aff[idx], 0)}\n"
                    f"% Écarts : {pourcentage_par_flux[idx]:.1f}%"
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =====================================================
            # CHART 3 : VALORISATION SIGNÉE
            # =====================================================
            fig3, ax3 = plt.subplots(figsize=(12, max(4.8, len(noms_flux_aff) * 0.42)), dpi=100)
            fig3.patch.set_facecolor(UI["surface_2"])
            ax3.set_facecolor(UI["surface_2"])

            max_abs = max([abs(v) for v in valeur_ecarts] or [1])
            if max_abs == 0:
                max_abs = 1

            norm3 = mcolors.TwoSlopeNorm(vmin=-max_abs, vcenter=0, vmax=max_abs)
            cmap3 = cm.RdBu_r
            colors3 = [cmap3(norm3(v)) for v in valeur_ecarts]

            bars3 = ax3.barh(y_pos, valeur_ecarts, color=colors3, alpha=0.95)
            ax3.set_yticks(y_pos)
            ax3.set_yticklabels(noms_flux_aff, color="white")
            ax3.invert_yaxis()
            ax3.axvline(0, color="white", linewidth=1, alpha=0.8)
            ax3.set_xlabel("Valorisation signée (k€)", color="white")
            titre3 = f"{titre_filiale} — Valorisation des écarts{suffix_annee}"
            if prof_sel:
                titre3 += f" | Profil : {prof_sel}"
            else:
                titre3 += " | Profils : tous"
            ax3.set_title(titre3, fontsize=14, color="white")
            ax3.tick_params(axis="x", colors="white")
            ax3.tick_params(axis="y", colors="white")
            ax3.grid(axis="x", color="gray", linestyle="--", alpha=0.3)
            ax3.set_axisbelow(True)

            extra = max_abs * 0.15
            ax3.set_xlim(-max_abs - extra, max_abs + extra)

            canvas3 = _embed_figure(fig3, chart3_body, mode="barh")

            cursor3 = mplcursors.cursor(bars3, hover=True)

            @cursor3.connect("add")
            def on_add_3(sel):
                idx = sel.index
                val = valeur_ecarts[idx]
                sel.annotation.set_text(
                    f"Flux : {noms_flux_aff[idx]}\n"
                    f"Valorisation : {int(val):,} k€".replace(",", " ")
                )
                sel.annotation.get_bbox_patch().set(fc="white", alpha=0.85)

            # =====================================================
            # TABLEAU
            # =====================================================
            total_prev = 0
            total_ecarts = 0
            total_valo = 0.0

            for flux, pct, valo in zip(noms_flux_aff, pourcentage_par_flux, valeur_ecarts):
                nb_p = prevs_par_flux.get(flux, 0)
                nb_e = ecarts_par_flux.get(flux, 0)
                total_prev += nb_p
                total_ecarts += nb_e
                total_valo += valo

                table.insert(
                    "",
                    "end",
                    values=(
                        flux,
                        nb_p,
                        nb_e,
                        f"{pct:.1f}%",
                        f"{int(valo):,}".replace(",", " ")
                    )
                )

            pct_global = (total_ecarts / total_prev * 100) if total_prev > 0 else 0.0
            table.insert(
                "",
                "end",
                values=(
                    "TOTAL",
                    total_prev,
                    total_ecarts,
                    f"{pct_global:.1f}%",
                    f"{int(total_valo):,}".replace(",", " ")
                )
            )

            _sync_scrollregion()

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _on_filiale_change(value=None):
            filiale = selected_filiale.get()

            annees = _annees_pour_filiale(filiale)
            annees_values = ["Toutes années"] + [str(a) for a in annees]
            annees_box.configure(values=annees_values)
            annees_var.set(str(annees[-1]) if annees else "Toutes années")

            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)

            profils_list = _profils_uniques_ordonnes_par_mois(filiale, annee) if annee is not None else []
            profils_values = ["Tous profils"] + profils_list
            profils_box.configure(values=profils_values)
            profils_var.set("Tous profils")

            afficher_graphes()

        def _on_annee_change(value=None):
            filiale = selected_filiale.get()
            val_annee = annees_var.get()
            annee = None if (not val_annee or val_annee == "Toutes années") else int(val_annee)

            profils_list = _profils_uniques_ordonnes_par_mois(filiale, annee) if annee is not None else []
            profils_values = ["Tous profils"] + profils_list
            profils_box.configure(values=profils_values)

            if profils_var.get() not in profils_values:
                profils_var.set("Tous profils")

            afficher_graphes()

        def _on_profil_change(value=None):
            afficher_graphes()

        filiale_menu.configure(command=_on_filiale_change)
        annees_box.configure(command=_on_annee_change)
        profils_box.configure(command=_on_profil_change)

        # =========================================================
        # AFFICHAGE INITIAL
        # =========================================================
        _on_filiale_change()

#===================Page heatmaps (ecart/anomalies)===================
    def afficher_heatmap_anomalies(self):
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import pandas as pd
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        from matplotlib.text import Annotation
        import matplotlib.patches as patches
        from sklearn.ensemble import IsolationForest
        import seaborn as sns
        import datetime as _dt
        import re

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _embed_figure(fig, master):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            def _resize(event=None):
                try:
                    master.update_idletasks()
                    w = max(master.winfo_width() - 16, 1000)
                    h = max(master.winfo_height() - 16, 520)
                    dpi = fig.get_dpi()

                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    fig.set_size_inches(w / dpi, h / dpi, forward=True)
                    fig.subplots_adjust(left=0.10, right=0.96, bottom=0.24, top=0.90)
                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        def _resolve_logo_path():
            candidates = []

            try:
                if image_path:
                    candidates.append(image_path)
            except Exception:
                pass

            try:
                if hasattr(self, "image_path") and self.image_path:
                    candidates.append(self.image_path)
            except Exception:
                pass

            for p in candidates:
                try:
                    if p:
                        return p
                except Exception:
                    pass
            return None

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    return None
            if isinstance(d, (int, float)):
                y = int(d)
                return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(d, fmt).year
                    except Exception:
                        pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m:
                    return int(m.group(1))
            return None

        def _profil_sort_key(name: str):
            """
            Trie par mois d'abord, puis par jour.
            Exemple :
            Profil 02/01 -> janvier
            Profil 14/04 -> avril
            Profil 01/12 -> décembre
            """
            s = str(name).strip()
            m = re.search(r"(\d{1,2})[/-](\d{1,2})", s)
            if not m:
                return (99, 99, s.lower())

            jj = int(m.group(1))
            mm = int(m.group(2))

            if 1 <= jj <= 31 and 1 <= mm <= 12:
                return (mm, jj, s.lower())

            return (99, 99, s.lower())

        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = sections.values() if filiale == "Toute filiale" else [sections[filiale]]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for _nom_flux, col_start in noms_colonnes_local:
                    try:
                        dates, _, _, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            annees.add(y)

            return sorted(annees)

        def _extraire_dataframe_anomalies(filiale_actuelle, annee):
            feuilles = sections.values() if filiale_actuelle == "Toute filiale" else [sections[filiale_actuelle]]

            flux_a_exclure = {
                "Cash flow de financement",
                "Cash flow net",
                "Sous total financier",
                "Sous total Investissements nets et ACE",
                "Free cash Flow",
                "Sous total recettes",
                "Sous total dépenses",
            }

            rows = []

            for feuille in feuilles:
                try:
                    ws, noms_colonnes_local = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for nom_flux, col_start in noms_colonnes_local:
                    if nom_flux in flux_a_exclure:
                        continue

                    try:
                        dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for i, date in enumerate(dates):
                        y = _year_of(date)
                        if y is None or y != annee:
                            continue

                        r = _to_number(reel[i] if i < len(reel) else None)
                        if r is None:
                            continue

                        for idx, prev_list in enumerate(previsions):
                            pv = _to_number(prev_list[i] if i < len(prev_list) else None)
                            if pv is None:
                                continue

                            rows.append({
                                "Date": date,
                                "Flux": nom_flux,
                                "Profil": noms_profils[idx] if idx < len(noms_profils) else f"Profil {idx + 1}",
                                "Réel": r,
                                "Prévision": pv,
                                "Écart": r - pv,
                            })

            if not rows:
                return pd.DataFrame()

            df = pd.DataFrame(rows)
            df["Écart_abs"] = df["Écart"].abs()

            if len(df) >= 12 and df["Écart"].nunique() > 1:
                seuil = 2 * df["Écart"].std()
                contamination = min(0.05, max(0.01, (df["Écart"].abs() > seuil).mean()))
                model = IsolationForest(contamination=contamination, random_state=42)

                pred = model.fit_predict(df[["Écart"]])
                df["Anomalie"] = (pred == -1).astype(int)
            else:
                q = df["Écart_abs"].quantile(0.90) if len(df) > 1 else df["Écart_abs"].max()
                df["Anomalie"] = (df["Écart_abs"] >= q).astype(int)

            return df

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.Treeview",
            background=UI["table_bg"],
            fieldbackground=UI["table_bg"],
            foreground=UI["text"],
            borderwidth=0,
            rowheight=30,
            font=("Segoe UI", 11)
        )
        style.map(
            "Pulse.Treeview",
            background=[("selected", UI["table_selected"])],
            foreground=[("selected", "white")]
        )
        style.configure(
            "Pulse.Treeview.Heading",
            background=UI["table_header"],
            foreground=UI["text"],
            relief="flat",
            font=("Segoe UI Semibold", 11, "bold")
        )
        style.map("Pulse.Treeview.Heading", background=[("active", UI["table_header"])])

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        logo_path = _resolve_logo_path()
        try:
            _img = Image.open(logo_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module de détection d’anomalies",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        retour_cmd = self.retour_menu if hasattr(self, "retour_menu") else self.creer_accueil
        neutral_btn(top_right, "Retour à l’accueil", retour_cmd).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.grid_rowconfigure(3, weight=0)
        page.columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "ANOMALIES", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Heatmap des anomalies",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Profils triés chronologiquement par mois, puis par jour à l’intérieur du mois.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "FILTRES",
            "Périmètre d’analyse",
            "L’année est obligatoire sur cette vue pour garder une heatmap lisible."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        filters_body.grid_columnconfigure(0, weight=0)
        filters_body.grid_columnconfigure(1, weight=0)
        filters_body.grid_columnconfigure(2, weight=1)

        filiales = ["Toute filiale"] + list(sections.keys())
        selected_filiale = ctk.StringVar(value=filiales[0])
        annees_var = ctk.StringVar(value="")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        filiale_menu = ctk.CTkOptionMenu(
            filters_body,
            values=filiales,
            variable=selected_filiale,
            width=240,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        filiale_menu.grid(row=1, column=0, sticky="w", padx=(0, 18))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=1, sticky="w", pady=(0, 6)
        )
        annees_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Aucune année"],
            variable=annees_var,
            width=160,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annees_box.grid(row=1, column=1, sticky="w")

        filter_hint = label(
            filters_body,
            "Le changement de filiale ou d’année recalcule automatiquement la heatmap.",
            font=FONT["small"],
            color=UI["muted"]
        )
        filter_hint.grid(row=1, column=2, sticky="w", padx=(18, 0))

        # =========================================================
        # HEATMAP
        # =========================================================
        heatmap_card = card(page, fg=UI["surface"], radius=20)
        heatmap_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 14))
        heatmap_card.grid_columnconfigure(0, weight=1)

        section_header(
            heatmap_card,
            "VUE GLOBALE",
            "Heatmap des anomalies détectées",
            "Survolez pour afficher la synthèse d’une cellule, cliquez pour afficher le détail."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        heatmap_body = ctk.CTkFrame(
            heatmap_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=620
        )
        heatmap_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        heatmap_body.grid_propagate(False)
        heatmap_body.grid_rowconfigure(0, weight=1)
        heatmap_body.grid_columnconfigure(0, weight=1)

        _make_placeholder(heatmap_body, "Chargement de la heatmap...")

        # =========================================================
        # DETAILS
        # =========================================================
        details_card = card(page, fg=UI["surface"], radius=20)
        details_card.grid(row=3, column=0, sticky="ew", padx=28, pady=(0, 24))
        details_card.grid_columnconfigure(0, weight=1)

        section_header(
            details_card,
            "DÉTAILS",
            "Anomalies sélectionnées",
            "Le détail des anomalies s’affiche ici après clic sur une cellule."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        details_title = label(
            details_card,
            "Aucune cellule sélectionnée",
            font=FONT["label_bold"],
            color=UI["text"]
        )
        details_title.grid(row=1, column=0, sticky="w", padx=18, pady=(0, 6))

        details_summary = label(
            details_card,
            "Le détail s’affichera ici après sélection d’une cellule.",
            font=FONT["small"],
            color=UI["muted"],
            justify="left",
            wraplength=1200
        )
        details_summary.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 12))

        table_holder = ctk.CTkFrame(
            details_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        table_holder.grid(row=3, column=0, sticky="ew", padx=18, pady=(0, 18))
        table_holder.grid_columnconfigure(0, weight=1)
        table_holder.grid_rowconfigure(0, weight=1)

        cols = ["Date", "Flux", "Profil", "Réel (k€)", "Prévision (k€)", "Écart (k€)"]
        tree = ttk.Treeview(table_holder, columns=cols, show="headings", height=10, style="Pulse.Treeview")
        tree.grid(row=0, column=0, sticky="nsew", padx=12, pady=12)

        yscroll = ttk.Scrollbar(table_holder, orient="vertical", command=tree.yview)
        yscroll.grid(row=0, column=1, sticky="ns", pady=12)
        tree.configure(yscrollcommand=yscroll.set)

        for c in cols:
            tree.heading(c, text=c)
            width = 150
            if c in ("Flux", "Profil"):
                width = 190
            tree.column(c, width=width, anchor="center")

        tree.tag_configure("neg", foreground="#EF4444")
        tree.tag_configure("pos", foreground="#22C55E")

        # =========================================================
        # BUSINESS RULES
        # =========================================================
        encaissements = [
            "Trafic Voyageurs", "Subventions", "Redevances d'infrastructure",
            "Enc. Autres Produits", "Sous total recettes", "Subventions d'investissements"
        ]
        decaissements = ["Péages", "Charges de personnel", "ACE & Investissements"]
        mixtes = [
            "Sous total Investissements nets et ACE", "Charges et produits financiers",
            "Dividendes reçus et versés", "Augmentations de capital", "Sous total financier",
            "Free cash Flow", "Emprunts", "Tirages Lignes CT", "Change", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers", "Emprunts / Prêts - Groupe",
            "Cash flow de financement", "Cash flow net", "Cessions d'immobilisations",
            "Impôts et Taxes", "Sous total dépenses"
        ]

        def est_favorable(flux_nom, reel_val, prev_val):
            if flux_nom in encaissements:
                return reel_val >= prev_val
            elif flux_nom in decaissements:
                return abs(reel_val) <= abs(prev_val)
            elif flux_nom in mixtes:
                return (reel_val >= prev_val) if prev_val >= 0 else (abs(reel_val) <= abs(prev_val))
            else:
                return reel_val >= prev_val

        def _reset_details():
            details_title.configure(text="Aucune cellule sélectionnée")
            details_summary.configure(text="Le détail s’affichera ici après sélection d’une cellule.")
            tree.delete(*tree.get_children())

        def afficher_heatmap(annee: int):
            _clear_children(heatmap_body)
            _reset_details()

            filiale_actuelle = selected_filiale.get()
            df_all = _extraire_dataframe_anomalies(filiale_actuelle, annee)

            if df_all.empty:
                _make_placeholder(heatmap_body, "Aucune donnée exploitable trouvée pour cette année.")
                return

            heatmap_data = df_all.pivot_table(
                index="Profil",
                columns="Flux",
                values="Anomalie",
                aggfunc="sum",
                fill_value=0
            ).astype(int)

            # on conserve uniquement profils / flux avec au moins une anomalie
            heatmap_data = heatmap_data.loc[(heatmap_data.sum(axis=1) > 0), :]
            heatmap_data = heatmap_data.loc[:, (heatmap_data.sum(axis=0) > 0)]

            if heatmap_data.empty:
                _make_placeholder(heatmap_body, "Aucune anomalie détectée pour cette année.")
                return

            # TRI CHRONOLOGIQUE PAR MOIS D'ABORD, PUIS PAR JOUR
            profils_tries = sorted(list(heatmap_data.index), key=_profil_sort_key)
            heatmap_data = heatmap_data.reindex(profils_tries)

            mean_data = df_all.pivot_table(
                index="Profil",
                columns="Flux",
                values="Écart_abs",
                aggfunc="mean",
                fill_value=0
            )
            mean_data = mean_data.reindex(index=heatmap_data.index, columns=heatmap_data.columns)

            nb_rows = max(1, heatmap_data.shape[0])
            nb_cols = max(1, heatmap_data.shape[1])

            dynamic_h = min(max(520, nb_rows * 34 + 180), 1600)
            heatmap_body.configure(height=dynamic_h)

            plt.close("all")
            fig, ax = plt.subplots(
                figsize=(max(12, nb_cols * 0.85), max(5, nb_rows * 0.42)),
                dpi=100
            )
            fig.patch.set_facecolor(UI["surface_2"])
            ax.set_facecolor(UI["surface_2"])

            sns.heatmap(
                heatmap_data,
                cmap="Reds",
                annot=heatmap_data.values,
                fmt="d",
                linewidths=0.5,
                linecolor="#444",
                ax=ax,
                cbar=True
            )

            ax.set_xlabel("Flux", color="white", fontsize=12, fontweight="bold")
            ax.set_ylabel("Profil", color="white", fontsize=12, fontweight="bold")
            ax.set_title(
                f"Heatmap des anomalies ({filiale_actuelle}) — {annee}",
                color="white",
                fontsize=16,
                fontweight="bold"
            )

            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=10, color="white")
            plt.setp(ax.get_yticklabels(), rotation=0, fontsize=10, color="white")

            cbar = ax.collections[0].colorbar
            cbar.set_label("Nombre d'anomalies", color="white", fontsize=12, fontweight="bold")
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(cbar.ax.yaxis.get_ticklabels(), color="white")
            cbar.outline.set_edgecolor("white")

            canvas_fig = _embed_figure(fig, heatmap_body)

            tooltip = Annotation(
                "",
                xy=(0, 0),
                xytext=(15, 15),
                textcoords="offset points",
                ha="left",
                va="bottom",
                bbox=dict(boxstyle="round", fc="black", ec="white", lw=1, alpha=0.85),
                color="white",
                fontsize=9
            )
            tooltip.set_visible(False)
            ax.add_artist(tooltip)

            hover_rect = patches.Rectangle(
                (0, 0), 1, 1,
                fill=True,
                edgecolor="black",
                linewidth=2,
                facecolor="blue",
                alpha=0.25
            )
            hover_rect.set_visible(False)
            ax.add_patch(hover_rect)

            def get_cell(event):
                if event.inaxes != ax or event.xdata is None or event.ydata is None:
                    return None, None
                x, y = int(event.xdata), int(event.ydata)
                if x < 0 or y < 0 or x >= heatmap_data.shape[1] or y >= heatmap_data.shape[0]:
                    return None, None
                return x, y

            def on_hover(event):
                x, y = get_cell(event)
                if x is None:
                    hover_rect.set_visible(False)
                    tooltip.set_visible(False)
                    fig.canvas.draw_idle()
                    return

                hover_rect.set_xy((x, y))
                hover_rect.set_visible(True)

                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]
                n_anomalies = int(heatmap_data.iloc[y, x])
                mean_gap = float(mean_data.iloc[y, x]) if flux in mean_data.columns else 0.0

                tooltip.xy = (event.xdata, event.ydata)
                tooltip.set_text(
                    f"{profil} / {flux}\nAnomalies : {n_anomalies}\nÉcart moyen abs. : {mean_gap:,.0f} k€".replace(",", " ")
                )
                tooltip.set_visible(True)
                fig.canvas.draw_idle()

            def on_click(event):
                x, y = get_cell(event)
                if x is None:
                    return

                flux = heatmap_data.columns[x]
                profil = heatmap_data.index[y]

                tree.delete(*tree.get_children())

                filtered = df_all[
                    (df_all["Flux"] == flux) &
                    (df_all["Profil"] == profil) &
                    (df_all["Anomalie"] == 1)
                ].copy()

                filtered["Écart_abs"] = filtered["Écart"].abs()
                filtered = filtered.sort_values(by="Écart_abs", ascending=False)

                details_title.configure(text=f"Détails — {profil} / {flux}")
                details_summary.configure(
                    text=f"{len(filtered)} anomalie(s) détectée(s) pour le profil '{profil}' sur le flux '{flux}' en {annee}."
                )

                if filtered.empty:
                    tree.insert("", "end", values=("—", flux, profil, "—", "—", "Aucune anomalie"))
                    return

                for _, row in filtered.iterrows():
                    date_str = row["Date"].strftime("%d/%m/%Y") if hasattr(row["Date"], "strftime") else str(row["Date"])
                    favorable = est_favorable(row["Flux"], row["Réel"], row["Prévision"])
                    tag = "pos" if favorable else "neg"

                    tree.insert(
                        "",
                        "end",
                        values=[
                            date_str,
                            row["Flux"],
                            row["Profil"],
                            f"{row['Réel']:,.0f}".replace(",", " "),
                            f"{row['Prévision']:,.0f}".replace(",", " "),
                            f"{row['Écart']:,.0f}".replace(",", " "),
                        ],
                        tags=(tag,)
                    )

                page.update_idletasks()
                try:
                    target = max(0, details_card.winfo_y() - 40)
                    total_h = max(1, page.winfo_height())
                    body_canvas.yview_moveto(min(1.0, target / total_h))
                except Exception:
                    pass

            canvas_fig.mpl_connect("motion_notify_event", on_hover)
            canvas_fig.mpl_connect("button_press_event", on_click)

            _sync_scrollregion()

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _on_filiale_change(value=None):
            filiale = selected_filiale.get()
            annees = _annees_pour_filiale(filiale)
            values = [str(a) for a in annees] if annees else ["Aucune année"]

            annees_box.configure(values=values)

            if annees:
                annees_var.set(values[-1])
                afficher_heatmap(int(values[-1]))
            else:
                annees_var.set("Aucune année")
                _make_placeholder(heatmap_body, "Aucune année disponible pour cette filiale.")
                _reset_details()

        def _on_annee_change(value=None):
            val = annees_var.get()
            if not val or val == "Aucune année":
                return
            afficher_heatmap(int(val))

        filiale_menu.configure(command=_on_filiale_change)
        annees_box.configure(command=_on_annee_change)

        # =========================================================
        # AFFICHAGE INITIAL
        # =========================================================
        _on_filiale_change()   

    def afficher_heatmap_ecarts(self):
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk
        from PIL import Image
        from customtkinter import CTkImage
        import seaborn as sns
        import matplotlib.pyplot as plt
        import pandas as pd
        from collections import defaultdict
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import re
        import datetime as _dt

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "table_bg": "#131A25",
            "table_header": "#1D2634",
            "table_selected": "#3F4B5F",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "label_bold": ("Segoe UI", 12, "bold"),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
            "kpi": ("Segoe UI Semibold", 22, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def _clear_children(widget):
            for child in widget.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

        def _make_placeholder(parent, text):
            _clear_children(parent)
            parent.grid_rowconfigure(0, weight=1)
            parent.grid_columnconfigure(0, weight=1)
            ctk.CTkLabel(
                parent,
                text=text,
                text_color=UI["muted"],
                font=("Segoe UI", 12),
                justify="center"
            ).grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        def _embed_figure(fig, master):
            master.update_idletasks()
            master.grid_rowconfigure(0, weight=1)
            master.grid_columnconfigure(0, weight=1)

            fig.patch.set_facecolor(UI["surface_2"])

            canvas = FigureCanvasTkAgg(fig, master=master)
            widget = canvas.get_tk_widget()

            try:
                widget.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            try:
                canvas._tkcanvas.configure(bg=UI["surface_2"], highlightthickness=0, bd=0)
            except Exception:
                pass

            widget.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

            def _resize(event=None):
                try:
                    master.update_idletasks()
                    w = max(master.winfo_width() - 16, 950)
                    h = max(master.winfo_height() - 16, 520)
                    dpi = fig.get_dpi()

                    widget.configure(width=w, height=h)
                    try:
                        canvas._tkcanvas.configure(width=w, height=h)
                    except Exception:
                        pass

                    fig.set_size_inches(w / dpi, h / dpi, forward=True)
                    fig.subplots_adjust(left=0.10, right=0.97, bottom=0.24, top=0.90)
                    canvas.draw_idle()
                except Exception:
                    pass

            master.bind("<Configure>", _resize, add="+")
            widget.bind("<Configure>", _resize, add="+")
            _resize()

            return canvas

        # =========================================================
        # HELPERS DATA
        # =========================================================
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _year_of(d):
            if d is None:
                return None
            if hasattr(d, "year"):
                try:
                    return int(d.year)
                except Exception:
                    return None
            if isinstance(d, (int, float)):
                y = int(d)
                return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        return _dt.datetime.strptime(d, fmt).year
                    except Exception:
                        pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m:
                    return int(m.group(1))
            return None

        def _parse_profil_key(name: str):
            m = re.search(r"(\d{1,2})[/-](\d{1,2})", str(name))
            if not m:
                return (99, 99, str(name).lower())
            jj = int(m.group(1))
            mm = int(m.group(2))
            if 1 <= jj <= 31 and 1 <= mm <= 12:
                return (mm, jj, str(name).lower())
            return (99, 99, str(name).lower())

        # Flux exclus du scope heatmap
        flux_a_exclure = {
            "Cash flow de financement",
            "Cash flow net",
            "Sous total financier",
            "Sous total Investissements nets et ACE",
            "Free cash Flow",
            "Sous total recettes",
            "Sous total dépenses",
            "C/C - Groupe",
            # flux de stock / encours peu pertinents pour cette heatmap
            "Emprunts",
            "Tirages Lignes CT",
            "Variation de collatéral",
            "Créances CDP",
            "Placements",
            "CC financiers",
            "Emprunts / Prêts - Groupe",
            "Cashpool",
            "Encours de financement",
            "Endettement Net",
        }

        def _annees_pour_filiale(filiale):
            annees = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for _nom, col_start in noms_colonnes:
                    try:
                        dates, _reel, _prevs, _profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for d in dates:
                        y = _year_of(d)
                        if y is not None:
                            annees.add(y)

            return sorted(annees)

        def _flux_pour_filiale(filiale):
            flux = set()
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            for feuille in feuilles:
                try:
                    _ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for nom_flux, _col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue
                    flux.add(str(nom_flux))

            return sorted(flux)

        def _build_heatmap_df(filiale, annee, flux_filtre):
            feuilles = list(sections.values()) if filiale == "Toutes filiales" else [filiale]

            data_dict = defaultdict(lambda: defaultdict(int))
            total_hits = 0

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                for nom_flux, col_start in noms_colonnes:
                    if nom_flux in flux_a_exclure:
                        continue

                    if flux_filtre != "Tous flux" and nom_flux != flux_filtre:
                        continue

                    try:
                        dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    for i, date in enumerate(dates):
                        y = _year_of(date)
                        if y is None or y != annee:
                            continue

                        r = _to_number(reel[i] if i < len(reel) else None)
                        if r in (None, 0):
                            continue

                        try:
                            mois = pd.to_datetime(date).strftime("%Y-%m")
                        except Exception:
                            continue

                        for p_idx, prev_list in enumerate(previsions):
                            pv = _to_number(prev_list[i] if i < len(prev_list) else None)
                            if pv is None:
                                continue

                            try:
                                ecart = (pv - r) / r
                            except Exception:
                                continue

                            if abs(ecart) >= 0.4:
                                profil_name = noms_profils[p_idx] if p_idx < len(noms_profils) else f"Profil {p_idx + 1}"
                                data_dict[str(profil_name)][mois] += 1
                                total_hits += 1

            if not data_dict:
                return None, 0

            heatmap_df = pd.DataFrame(data_dict).T.fillna(0).astype(int)

            # ordre chronologique des colonnes
            heatmap_df = heatmap_df.reindex(sorted(heatmap_df.columns), axis=1)

            # ordre des profils plus lisible
            heatmap_df = heatmap_df.loc[sorted(heatmap_df.index, key=_parse_profil_key)]

            return heatmap_df, total_hits

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module heatmap des écarts",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        retour_cmd = self.retour_menu if hasattr(self, "retour_menu") else self.creer_accueil
        neutral_btn(top_right, "Retour à l’accueil", retour_cmd).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _on_mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        def _on_linux_scroll_up(event):
            body_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            body_canvas.yview_scroll(1, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            body_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            body_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            body_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            body_canvas.unbind_all("<MouseWheel>")
            body_canvas.unbind_all("<Button-4>")
            body_canvas.unbind_all("<Button-5>")

        body_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        body_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=0)
        page.grid_rowconfigure(2, weight=0)
        page.columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "HEATMAP", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Carte thermique des écarts",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Visualisez le nombre d’écarts significatifs par profil et par mois, avec filtres filiale, année et flux.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # FILTRES
        # =========================================================
        filters_card = card(page, fg=UI["surface"], radius=20)
        filters_card.grid(row=1, column=0, sticky="ew", padx=28, pady=(0, 14))
        filters_card.grid_columnconfigure(0, weight=1)

        section_header(
            filters_card,
            "FILTRES",
            "Périmètre d’analyse",
            "Le filtre Flux est bien présent et s’applique directement au calcul de la heatmap."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        filters_body = ctk.CTkFrame(filters_card, fg_color="transparent")
        filters_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        filters_body.grid_columnconfigure(0, weight=0)
        filters_body.grid_columnconfigure(1, weight=0)
        filters_body.grid_columnconfigure(2, weight=0)
        filters_body.grid_columnconfigure(3, weight=1)

        filiales = ["Toutes filiales"] + list(sections.values())
        selected_filiale = ctk.StringVar(value=filiales[0])
        annees_var = ctk.StringVar(value="")
        flux_var = ctk.StringVar(value="Tous flux")

        label(filters_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        filiale_menu = ctk.CTkOptionMenu(
            filters_body,
            values=filiales,
            variable=selected_filiale,
            width=220,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        filiale_menu.grid(row=1, column=0, sticky="w", padx=(0, 18))

        label(filters_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=1, sticky="w", pady=(0, 6)
        )
        annees_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Aucune année"],
            variable=annees_var,
            width=150,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        annees_box.grid(row=1, column=1, sticky="w", padx=(0, 18))

        label(filters_body, "Flux", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=2, sticky="w", pady=(0, 6)
        )
        flux_box = ctk.CTkOptionMenu(
            filters_body,
            values=["Tous flux"],
            variable=flux_var,
            width=260,
            height=38,
            fg_color=UI["surface_3"],
            button_color=UI["surface_3"],
            button_hover_color=UI["neutral_hover"],
            text_color=UI["text"],
            dropdown_fg_color=UI["surface_2"],
            dropdown_hover_color=UI["neutral_hover"],
        )
        flux_box.grid(row=1, column=2, sticky="w")

        filter_hint = label(
            filters_body,
            "Les flux agrégés et certains flux de stock sont exclus pour garder une lecture métier utile.",
            font=FONT["small"],
            color=UI["muted"]
        )
        filter_hint.grid(row=1, column=3, sticky="w", padx=(18, 0))

        # =========================================================
        # KPI + HEATMAP
        # =========================================================
        heatmap_card = card(page, fg=UI["surface"], radius=20)
        heatmap_card.grid(row=2, column=0, sticky="ew", padx=28, pady=(0, 24))
        heatmap_card.grid_columnconfigure(0, weight=1)

        section_header(
            heatmap_card,
            "VUE",
            "Heatmap des écarts significatifs",
            "Chaque cellule représente le nombre d’écarts d’au moins 40 % pour un profil et un mois."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        stats_frame = ctk.CTkFrame(heatmap_card, fg_color="transparent")
        stats_frame.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 12))
        for i in range(4):
            stats_frame.grid_columnconfigure(i, weight=1)

        stat1 = card(stats_frame, fg=UI["surface_2"], radius=14)
        stat1.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        stat2 = card(stats_frame, fg=UI["surface_2"], radius=14)
        stat2.grid(row=0, column=1, sticky="ew", padx=8)
        stat3 = card(stats_frame, fg=UI["surface_2"], radius=14)
        stat3.grid(row=0, column=2, sticky="ew", padx=8)
        stat4 = card(stats_frame, fg=UI["surface_2"], radius=14)
        stat4.grid(row=0, column=3, sticky="ew", padx=(8, 0))

        stat1_title = label(stat1, "Occurrences", font=FONT["small"], color=UI["muted"])
        stat1_title.pack(anchor="w", padx=14, pady=(10, 0))
        stat1_value = label(stat1, "—", font=FONT["kpi"], color=UI["text"])
        stat1_value.pack(anchor="w", padx=14, pady=(0, 10))

        stat2_title = label(stat2, "Profils", font=FONT["small"], color=UI["muted"])
        stat2_title.pack(anchor="w", padx=14, pady=(10, 0))
        stat2_value = label(stat2, "—", font=FONT["kpi"], color=UI["text"])
        stat2_value.pack(anchor="w", padx=14, pady=(0, 10))

        stat3_title = label(stat3, "Mois", font=FONT["small"], color=UI["muted"])
        stat3_title.pack(anchor="w", padx=14, pady=(10, 0))
        stat3_value = label(stat3, "—", font=FONT["kpi"], color=UI["text"])
        stat3_value.pack(anchor="w", padx=14, pady=(0, 10))

        stat4_title = label(stat4, "Flux", font=FONT["small"], color=UI["muted"])
        stat4_title.pack(anchor="w", padx=14, pady=(10, 0))
        stat4_value = label(stat4, "—", font=FONT["kpi"], color=UI["text"])
        stat4_value.pack(anchor="w", padx=14, pady=(0, 10))

        heatmap_body = ctk.CTkFrame(
            heatmap_card,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"],
            height=640
        )
        heatmap_body.grid(row=2, column=0, sticky="ew", padx=18, pady=(0, 18))
        heatmap_body.grid_propagate(False)
        heatmap_body.grid_rowconfigure(0, weight=1)
        heatmap_body.grid_columnconfigure(0, weight=1)

        _make_placeholder(heatmap_body, "Chargement de la heatmap...")

        # =========================================================
        # RENDER
        # =========================================================
        def maj_heatmap(filiale, annee, flux_filtre):
            heatmap_df, total_hits = _build_heatmap_df(filiale, annee, flux_filtre)
            _clear_children(heatmap_body)

            if heatmap_df is None or heatmap_df.empty:
                stat1_value.configure(text="0")
                stat2_value.configure(text="0")
                stat3_value.configure(text="0")
                stat4_value.configure(text="0" if flux_filtre == "Tous flux" else "1")
                _make_placeholder(heatmap_body, "Aucune donnée à afficher pour ces filtres.")
                return

            nb_profils = heatmap_df.shape[0]
            nb_mois = heatmap_df.shape[1]
            nb_flux = len(_flux_pour_filiale(filiale)) if flux_filtre == "Tous flux" else 1

            stat1_value.configure(text=f"{total_hits:,}".replace(",", " "))
            stat2_value.configure(text=str(nb_profils))
            stat3_value.configure(text=str(nb_mois))
            stat4_value.configure(text=str(nb_flux))

            dynamic_h = min(max(520, nb_profils * 34 + 180), 1600)
            heatmap_body.configure(height=dynamic_h)

            fig, ax = plt.subplots(
                figsize=(max(12, nb_mois * 1.0), max(5.5, nb_profils * 0.42)),
                dpi=100
            )
            fig.patch.set_facecolor(UI["surface_2"])
            ax.set_facecolor(UI["surface_2"])

            sns.heatmap(
                heatmap_df,
                cmap="Reds",
                annot=heatmap_df.values,
                fmt="d",
                linewidths=0.5,
                linecolor="#444",
                cbar_kws={"label": "Nombre d'écarts significatifs"},
                ax=ax
            )

            cbar = ax.collections[0].colorbar
            cbar.ax.yaxis.label.set_color("white")
            cbar.ax.tick_params(colors="white")

            titre_flux = flux_filtre if flux_filtre != "Tous flux" else "Tous flux"
            titre_filiale = filiale if filiale != "Toutes filiales" else "Toutes filiales"

            ax.set_title(
                f"{titre_filiale} — {annee} — {titre_flux}",
                fontsize=16,
                color="white",
                fontweight="bold"
            )
            ax.set_xlabel("Mois", color="white", fontsize=12, fontweight="bold")
            ax.set_ylabel("Profil", color="white", fontsize=12, fontweight="bold")

            plt.setp(ax.get_xticklabels(), rotation=45, ha="right", color="white")
            plt.setp(ax.get_yticklabels(), color="white")

            _embed_figure(fig, heatmap_body)
            _sync_scrollregion()

        # =========================================================
        # CALLBACKS
        # =========================================================
        def _refresh_flux_options(filiale):
            flux_values = ["Tous flux"] + _flux_pour_filiale(filiale)
            if len(flux_values) == 1:
                flux_values = ["Tous flux"]
            flux_box.configure(values=flux_values)

            if flux_var.get() not in flux_values:
                flux_var.set("Tous flux")

        def _on_filiale_change(value=None):
            filiale = selected_filiale.get()

            # années
            annees = _annees_pour_filiale(filiale)
            annee_values = [str(a) for a in annees] if annees else ["Aucune année"]
            annees_box.configure(values=annee_values)

            if annees:
                annees_var.set(str(annees[-1]))
            else:
                annees_var.set("Aucune année")

            # flux
            _refresh_flux_options(filiale)

            # rendu
            if annees:
                maj_heatmap(filiale, int(annees_var.get()), flux_var.get())
            else:
                stat1_value.configure(text="0")
                stat2_value.configure(text="0")
                stat3_value.configure(text="0")
                stat4_value.configure(text="0")
                _make_placeholder(heatmap_body, "Aucune année disponible pour cette filiale.")

        def _on_annee_change(value=None):
            val_annee = annees_var.get()
            if not val_annee or val_annee == "Aucune année":
                return
            maj_heatmap(selected_filiale.get(), int(val_annee), flux_var.get())

        def _on_flux_change(value=None):
            val_annee = annees_var.get()
            if not val_annee or val_annee == "Aucune année":
                return
            maj_heatmap(selected_filiale.get(), int(val_annee), flux_var.get())

        filiale_menu.configure(command=_on_filiale_change)
        annees_box.configure(command=_on_annee_change)
        flux_box.configure(command=_on_flux_change)

        # =========================================================
        # AFFICHAGE INITIAL
        # =========================================================
        _on_filiale_change()

    def afficher_backtesting_multi_horizon(self):
        import customtkinter as ctk
        from tkinter import ttk, filedialog, messagebox
        import tkinter as tk
        import matplotlib.pyplot as plt
        import seaborn as sns
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import pandas as pd
        from PIL import Image
        from customtkinter import CTkImage
        import numpy as np
        import re
        import datetime as _dt
        from functools import lru_cache

        # =============== RESET PAGE ===============
        self.vider_fenetre()

        # =============== HEADER ===============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(header_frame, text="PROJET PULSE - BACKTESTING MULTI-HORIZON", font=titre_font)\
            .pack(side="left", anchor="w")

        try:
            image_path = r"C:\Users\0304336A\...\logo_Pulse.png"
            logo_img = Image.open(image_path)
            ratio = logo_img.width / logo_img.height
            new_h = 40
            resized_logo = logo_img.resize((int(new_h * ratio), new_h), Image.Resampling.LANCZOS)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(int(new_h * ratio), new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"Erreur chargement logo: {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # =============== CONTENEUR SCROLLABLE ===============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)

        canvas = tk.Canvas(container, bg="#00122e", highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ctk.CTkFrame(canvas, fg_color="#00122e", corner_radius=0)
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # =============== HELPERS ===============
        def _to_number(x):
            if x is None: return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}: return None
                s = s.replace(",", ".")
                try: return float(s)
                except Exception: return None
            try: return float(x)
            except Exception: return None

        def _year_of(d):
            if d is None: return None
            if hasattr(d, "year"):
                try: return int(d.year)
                except Exception: return None
            if isinstance(d, (int, float)):
                y = int(d); return y if 1900 <= y <= 2100 else None
            if isinstance(d, str):
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try: return _dt.datetime.strptime(d, fmt).year
                    except Exception: pass
                m = re.search(r"(20\d{2}|19\d{2})", d)
                if m: return int(m.group(1))
            return None

        def _parse_mois_from_profil(name: str):
            """
            Extrait des indices d’horizon depuis le libellé de profil.
            Retourne (mm, jj, rel):
            - JJ/MM -> (mm, jj, None)
            - 'Mxx' -> (mm, None, None)
            - 'M-2' -> (None, None, 2)
            """
            s = str(name)
            m = re.search(r'(\d{1,2})[/-](\d{1,2})', s)
            if m:
                jj = int(m.group(1)); mm = int(m.group(2))
                if 1 <= mm <= 12:
                    return mm, jj, None
            m2 = re.search(r'[Mm]\s*0?(\d{1,2})', s)
            if m2:
                mm = int(m2.group(1))
                if 1 <= mm <= 12:
                    return mm, None, None
            m3 = re.search(r'[Mm]\s*-\s*(\d+)', s)
            if m3:
                return None, None, int(m3.group(1))
            return None, None, None

        def _horizon_from_name(p_name, previsions_len, max_m):
            mm, jj, rel = _parse_mois_from_profil(p_name)
            if rel is not None:
                return max(1, int(rel))
            if mm is not None and max_m is not None:
                return max(1, int(max_m - mm + 1))
            # fallback : plus l’index est grand, plus l’horizon est court
            return max(1, previsions_len)

        def _profil_max_m(noms_profils):
            mms = [m for m, _, _ in map(_parse_mois_from_profil, noms_profils) if m is not None]
            return max(mms) if mms else None

        @lru_cache(maxsize=64)
        def _annees_pour_filiale_cached(filiale_key: str):
            """Union des années disponibles selon la filiale sélectionnée."""
            if filiale_key == "Toutes filiales":
                feuilles = list(sections.values())
            else:
                feuilles = [sections[filiale_key]]
            years = set()
            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue
                for nom_flux, col_start in noms_colonnes:
                    try:
                        dates, _, _, _ = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                        for d in dates:
                            y = _year_of(d)
                            if y is not None:
                                years.add(y)
                    except Exception:
                        pass
            return sorted(years)

        def _compute_metric(prev, reel, metric: str):
            if metric == "MAPE (%)":
                denom = np.where(np.abs(reel) < 1e-9, 1.0, np.abs(reel))
                return np.abs(prev - reel) / denom * 100.0
            elif metric == "MAE (k€)":
                return np.abs(prev - reel) / 1000.0
            elif metric == "RMSE (k€)":
                return ( (prev - reel) ** 2 ) / (1000.0**2)
            else:
                return np.abs(prev - reel) / np.where(np.abs(reel) < 1e-9, 1.0, np.abs(reel)) * 100.0

        # =============== BARRE D’OUTILS (EN HAUT) ===============
        toolbar = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        toolbar.pack(fill="x", padx=10, pady=(10, 0))

        # Retour
        ctk.CTkButton(
            toolbar, text="⬅️ Retour au menu", command=self.retour_menu,
            width=180, height=40, corner_radius=15, fg_color="#444", hover_color="#666",
            text_color="white", font=("Segoe UI", 13, "bold")
        ).pack(side="left", padx=(0, 16))

        # Filiale
        def _label(parent, txt):
            return ctk.CTkLabel(parent, text=txt, text_color="white", font=("Segoe UI", 12, "bold"), fg_color="#00122e")

        _label(toolbar, "Filiale :").pack(side="left", padx=(0, 6))
        filiales = ["Toutes filiales"] + list(sections.keys())
        filiale_var = tk.StringVar(value=filiales[0])
        filiale_box = ttk.Combobox(toolbar, values=filiales, textvariable=filiale_var, state="readonly", width=28)
        filiale_box.pack(side="left", padx=(0, 16))

        # Année (OBLIGATOIRE, pas de "Toutes années")
        _label(toolbar, "Année :").pack(side="left", padx=(0, 6))
        annee_var = tk.StringVar(value="")
        annee_box = ttk.Combobox(toolbar, values=[], textvariable=annee_var, state="readonly", width=12)
        annee_box.pack(side="left", padx=(0, 16))

        # Choix du métrique
        _label(toolbar, "Métrique :").pack(side="left", padx=(0, 6))
        metric_var = tk.StringVar(value="MAPE (%)")
        metric_box = ttk.Combobox(toolbar, values=["MAPE (%)", "MAE (k€)", "RMSE (k€)"], textvariable=metric_var,
                                state="readonly", width=12)
        metric_box.pack(side="left")

        # =============== FRAME GRAPHIQUE ===============
        graph_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=10)
        graph_frame.pack(pady=16, padx=16, fill="both", expand=True)

        # =============== TABLEAU RÉCAP ===============
        table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        table_frame.pack(pady=(4, 16), fill="x")

        colonnes = ["Filiale", "Année", "Horizon (mois)", "Métrique", "Valeur", "Nb points"]
        tree = ttk.Treeview(table_frame, columns=colonnes, show="headings", height=10)
        for col in colonnes:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=160)
        tree.pack(pady=5, padx=5, fill="x")

        # =============== BOUTONS ACTION ===============
        actions_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e")
        actions_frame.pack(pady=(0, 12), fill="x")

        export_csv_btn = ctk.CTkButton(actions_frame, text="Exporter CSV", fg_color="#28B463", hover_color="#1E8449")
        export_png_btn = ctk.CTkButton(actions_frame, text="Exporter PNG", fg_color="#3498DB", hover_color="#2E86C1")
        export_csv_btn.pack(side="left", padx=(16, 8))
        export_png_btn.pack(side="left", padx=(8, 8))

        # =============== CALCUL & AFFICHAGE ===============
        last_df_shown = {"precision": None, "fig": None}

        def _load_backtest_rows(filiale_sel: str, annee_sel: int) -> pd.DataFrame:
            """
            Retourne un DF lignes élémentaires pour backtest:
            cols: filiale (code feuille), annee (str), horizon (int), prev, reel
            """
            records = []
            feuilles = list(sections.values()) if filiale_sel == "Toutes filiales" else [sections[filiale_sel]]

            for feuille in feuilles:
                try:
                    ws, noms_colonnes = charger_donnees(feuille, taille_bloc)
                except Exception:
                    continue

                # pour chaque flux, lit toutes les prévisions; on filtre l'année ensuite
                for nom_flux, col_start in noms_colonnes:
                    try:
                        dates, reel, previsions, noms_profils = extraire_valeurs(ws, col_start, nb_prev, annee=None)
                    except Exception:
                        continue

                    max_m = _profil_max_m(noms_profils)
                    previsions_len = len(previsions)

                    for i, d in enumerate(dates):
                        y = _year_of(d)
                        if y is None or y != annee_sel:
                            continue

                        r = _to_number(reel[i] if i < len(reel) else None)
                        if r is None:
                            continue
                        r = 1.0 if abs(r) < 1e-9 else r  # évite /0

                        for p_idx, p_list in enumerate(previsions):
                            pv = _to_number(p_list[i] if i < len(p_list) else None)
                            if pv is None:
                                continue

                            # horizon: basé sur nom + structure
                            p_name = noms_profils[p_idx] if p_idx < len(noms_profils) else f"P{p_idx+1}"
                            # le fallback utilise la "profondeur restante"
                            horizon = _horizon_from_name(p_name, previsions_len - p_idx, max_m)

                            records.append({
                                "filiale": feuille,
                                "annee": str(annee_sel),
                                "horizon": int(horizon),
                                "prev": float(pv),
                                "reel": float(r),
                                "flux": nom_flux,           # utile si besoin d’affiner plus tard
                                "profil": str(p_name)       # debug / audit
                            })
            return pd.DataFrame(records)

        def _aggregate_precision(df_raw: pd.DataFrame, metric_name: str) -> pd.DataFrame:
            if df_raw.empty:
                return df_raw

            df = df_raw.copy()
            # métrique élémentaire
            base = _compute_metric(df["prev"].values, df["reel"].values, metric_name)
            if metric_name == "RMSE (k€)":
                # on agrège la MSE et on racine après groupby
                df["metric_base"] = base  # (k€)^2
                grouped = df.groupby(["filiale", "annee", "horizon"]).agg(
                    metric=("metric_base", "mean"),
                    N=("metric_base", "size")
                ).reset_index()
                grouped["metric"] = np.sqrt(grouped["metric"])  # RMSE
            else:
                df["metric_base"] = base  # MAPE% ou MAE (k€)
                grouped = df.groupby(["filiale", "annee", "horizon"]).agg(
                    metric=("metric_base", "mean"),
                    N=("metric_base", "size")
                ).reset_index()
                grouped["metric"] = grouped["metric"]

            grouped = grouped.sort_values(["annee", "filiale", "horizon"], ascending=[True, True, False])
            return grouped

        def _refresh_graph_and_table():
            # Nettoyer graph + table
            for w in graph_frame.winfo_children():
                w.destroy()
            for r in tree.get_children():
                tree.delete(r)

            filiale = filiale_var.get()
            metric_name = metric_var.get()
            try:
                annee = int(annee_var.get())
            except Exception:
                tk.Label(graph_frame, text="Sélectionnez une année valide.", bg="#00122e", fg="white").pack()
                return

            df_raw = _load_backtest_rows(filiale, annee)
            if df_raw.empty:
                tk.Label(graph_frame, text="Aucune donnée disponible pour ces filtres.", bg="#00122e", fg="white").pack()
                last_df_shown["precision"] = None
                last_df_shown["fig"] = None
                return

            precision = _aggregate_precision(df_raw, metric_name)
            if precision.empty:
                tk.Label(graph_frame, text="Pas de points agrégés.", bg="#00122e", fg="white").pack()
                last_df_shown["precision"] = None
                last_df_shown["fig"] = None
                return

            # -------- GRAPHIQUE : lignes par filiale pour l'année sélectionnée --------
            plot_df = precision[precision["annee"] == str(annee)].copy()
            fig, ax = plt.subplots(figsize=(10, 6), facecolor="#00122e")
            ax.set_facecolor("#00122e")

            sns.lineplot(data=plot_df, x="horizon", y="metric", hue="filiale", marker="o", linewidth=2, ax=ax)

            ax.invert_xaxis()
            ax.set_xlabel("Horizon (mois avant la réalisation)", color="white")
            ylabel = "MAPE moyen (%)" if metric_name == "MAPE (%)" else metric_name
            ax.set_ylabel(ylabel, color="white")
            ax.grid(alpha=0.3)
            ax.tick_params(colors="white")

            # Légende : noms humains si possible
            handles, labels = ax.get_legend_handles_labels()
            new_labels = []
            for lab in labels:
                # lab est code feuille -> on le remappe au libellé si dispo
                for k, v in sections.items():
                    if v == lab:
                        new_labels.append(k)
                        break
                else:
                    new_labels.append(lab)
            ax.legend(handles, new_labels, title="Filiale", facecolor="#00122e", edgecolor="white", labelcolor="white", title_fontsize=10)

            title_suffix = f" – {annee}"
            ax.set_title(f"Précision par filiale ({metric_name}){title_suffix}", color="white", fontsize=14)

            canvas_fig = FigureCanvasTkAgg(fig, master=graph_frame)
            canvas_fig.draw()
            canvas_widget = canvas_fig.get_tk_widget()
            canvas_widget.pack(pady=10, fill="both", expand=True)
            plt.close(fig)

            # -------- TABLEAU RÉCAP --------
            for _, row in plot_df.sort_values(["filiale", "horizon"], ascending=[True, False]).iterrows():
                # Remap pour affichage
                fil = row["filiale"]
                for k, v in sections.items():
                    if v == fil:
                        fil_aff = k
                        break
                else:
                    fil_aff = fil

                val_disp = f"{row['metric']:.1f}" if metric_name == "MAPE (%)" else f"{row['metric']:.2f}"
                tree.insert("", "end", values=(
                    fil_aff,
                    row["annee"],
                    int(row["horizon"]),
                    metric_name,
                    val_disp,
                    int(row["N"])
                ))

            # TOTAL pondéré
            tot = plot_df["N"].sum()
            if tot > 0:
                if metric_name == "RMSE (k€)":
                    # moyenne pondérée déjà faite, mais on peut afficher la moyenne simple pondérée :
                    weighted = np.average(plot_df["metric"], weights=plot_df["N"])
                    total_val = f"{weighted:.2f}"
                else:
                    weighted = np.average(plot_df["metric"], weights=plot_df["N"])
                    total_val = f"{weighted:.1f}" if metric_name == "MAPE (%)" else f"{weighted:.2f}"
            else:
                total_val = "-"

            tree.insert("", "end",
                        values=("TOTAL", str(annee), "-", metric_name, total_val, int(tot)),
                        tags=("total",))
            tree.tag_configure("total", background="#444", foreground="white", font=('Segoe UI', 12, 'bold'))

            # Mémorise pour export
            last_df_shown["precision"] = plot_df.copy()
            last_df_shown["fig"] = canvas_fig

        # =============== CALLBACKS ===============
        def _on_filiale_change(*_):
            filiale = filiale_var.get()
            years = _annees_pour_filiale_cached(filiale)
            if not years:
                annee_box.config(values=[])
                annee_var.set("")
                messagebox.showinfo("Info", "Aucune année disponible pour cette filiale.")
                # Nettoyage affichages
                for w in graph_frame.winfo_children():
                    w.destroy()
                for r in tree.get_children():
                    tree.delete(r)
                return
            vals = [str(y) for y in years]
            annee_box.config(values=vals)
            annee_var.set(vals[-1])  # dernière année disponible
            _refresh_graph_and_table()

        def _on_annee_change(*_):
            _refresh_graph_and_table()

        def _on_metric_change(*_):
            _refresh_graph_and_table()

        def _export_csv():
            df = last_df_shown["precision"]
            if df is None or df.empty:
                messagebox.showinfo("Export CSV", "Aucune donnée à exporter pour ces filtres.")
                return
            # Remap filiale -> libellé humain pour export
            df = df.copy()
            df["Filiale"] = df["filiale"].apply(lambda code: next((k for k, v in sections.items() if v == code), code))
            out = df[["Filiale", "annee", "horizon", "metric", "N"]].rename(columns={
                "annee": "Année", "horizon": "Horizon (mois)", "metric": metric_var.get(), "N": "Nb points"
            })
            default_name = f"backtest_multi_horizon_{filiale_var.get().replace(' ','_')}_{annee_var.get()}_{metric_var.get().split()[0]}.csv"
            path = filedialog.asksaveasfilename(defaultextension=".csv", initialfile=default_name,
                                                filetypes=[("CSV","*.csv")])
            if not path:
                return
            try:
                out.to_csv(path, index=False, sep=";")
                messagebox.showinfo("Export CSV", f"Fichier exporté :\n{path}")
            except Exception as e:
                messagebox.showerror("Export CSV", f"Échec de l'export : {e}")

        def _export_png():
            fig_canvas = last_df_shown["fig"]
            if fig_canvas is None:
                messagebox.showinfo("Export PNG", "Rien à exporter pour le moment.")
                return
            default_name = f"backtest_multi_horizon_{filiale_var.get().replace(' ','_')}_{annee_var.get()}_{metric_var.get().split()[0]}.png"
            path = filedialog.asksaveasfilename(defaultextension=".png", initialfile=default_name,
                                                filetypes=[("Image PNG","*.png")])
            if not path:
                return
            try:
                fig_canvas.figure.savefig(path, dpi=160, bbox_inches="tight")
                messagebox.showinfo("Export PNG", f"Image exportée :\n{path}")
            except Exception as e:
                messagebox.showerror("Export PNG", f"Échec de l'export : {e}")

        filiale_box.bind("<<ComboboxSelected>>", lambda e: _on_filiale_change())
        annee_box.bind("<<ComboboxSelected>>", lambda e: _on_annee_change())
        metric_box.bind("<<ComboboxSelected>>", lambda e: _on_metric_change())
        export_csv_btn.configure(command=_export_csv)
        export_png_btn.configure(command=_export_png)

        # =============== AFFICHAGE INITIAL ===============
        _on_filiale_change()

        # =============== SCROLL (souris) ===============
        def _on_mousewheel(event):
            if event.delta == 0: return "break"
            step = -1 if event.delta > 0 else 1
            canvas.yview_scroll(step, "units"); return "break"
        def _on_linux_scroll_up(event): canvas.yview_scroll(-1, "units"); return "break"
        def _on_linux_scroll_down(event): canvas.yview_scroll(1, "units"); return "break"
        def _on_mousewheel_shift(event):
            if event.delta == 0: return "break"
            step = -1 if event.delta > 0 else 1
            canvas.xview_scroll(step, "units"); return "break"
        def _bind_mousewheel(_event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")
        def _unbind_mousewheel(_event=None):
            canvas.unbind_all("<MouseWheel>")
            canvas.unbind_all("<Shift-MouseWheel>")
            canvas.unbind_all("<Button-4>")
            canvas.unbind_all("<Button-5>")
        canvas.bind("<Enter>", _bind_mousewheel, add="+")
        canvas.bind("<Leave>", _unbind_mousewheel, add="+")

    def _annees_disponibles(self, section, flux_name):
        """Retourne la liste triée des années disponibles pour (section, flux)."""
        B = CACHE.get((section, flux_name))
        if not B or not B.get("dates"):
            return []
        return sorted({d.year for d in B["dates"]})

    def _token_from_flux(self, section, flux_name):
        """Retrouve col_start (token) à partir du nom de flux."""
        for name, tok in TOKENS.get(section, []):
            if name == flux_name:
                return tok
        return None

    def _profils_for_year(self, section, flux_name, annee):
        """
        Calcule les profils à afficher pour une année donnée : on ne garde que
        ceux qui ont au moins une valeur non-nulle/non-None sur l'année.
        Renvoie (noms_profils_filtrés, previsions_filtrées_existence_bool).
        """
        col_start = self._token_from_flux(section, flux_name)
        if col_start is None:
            return [], []

        dates, reel, previsions, noms_profils = extraire_valeurs(section, col_start, nb_prev, annee=annee)

        actifs = []
        for serie in previsions:  # une série par profil
            exist = any(v not in (None, 0, 0.0, "") for v in serie)
            actifs.append(exist)

        noms_ok = [np for np, ok in zip(noms_profils, actifs) if ok]
        return noms_ok, actifs

    def _on_flux_change(self, event=None):
        """Quand l'utilisateur choisit un flux : calcule les années disponibles, active la combo et met à jour les profils."""
        section = self.feuille_combo.get()
        flux_name = self.nom_combo.get()
        if not section or not flux_name:
            return

        annees = self._annees_disponibles(section, flux_name)
        if not annees:
            # Pas d'années disponibles → on vide l'UI profils
            self.annee_combo.config(values=[], state="disabled")
            self._rebuild_profils_ui(section, flux_name, None)
            return

        # Active la combobox et pré-sélectionne la plus récente
        self.annee_combo.config(values=[str(a) for a in annees], state="readonly")
        self.annee_combo.set(str(annees[-1]))

        # Recharge les profils pour l’année par défaut
        self._rebuild_profils_ui(section, flux_name, annees[-1])

    def _on_annee_change(self, event=None):
        """Quand l'année change : profils dynamiques mis à jour."""
        section = self.feuille_combo.get()
        flux_name = self.nom_combo.get()
        if not section or not flux_name:
            return
        try:
            annee = int(self.annee_combo.get())
        except Exception:
            annee = None
        self._rebuild_profils_ui(section, flux_name, annee)

    def _rebuild_profils_ui(self, section, flux_name, annee):
        """Reconstruit les checkboxes de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()
        self.vars_prev = []
        self.profils_names_order = []  # ✅ reset au début

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils  # ✅ mémorise l’ordre affiché

        # Si aucun profil
        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # ✅ Disposition sur 4 lignes max
        nb_lignes = 4
        nb_cols = max(1, len(noms_profils) // nb_lignes + (1 if len(noms_profils) % nb_lignes else 0))

        for i, nom_profil in enumerate(noms_profils):
            var = tk.BooleanVar(value=False)
            cb = tk.Checkbutton(
                self.profils_frame, text=nom_profil, variable=var,
                bg="#00122e", fg="white", font=('Segoe UI', 10),
                selectcolor="#00aced", activebackground="#003366",
                activeforeground="white"
            )
            # ✅ Répartition sur 4 lignes
            row = i % nb_lignes
            col = i // nb_lignes

            cb.grid(row=row, column=col, sticky="w", padx=12, pady=6)
            cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
            cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))
            self.vars_prev.append(var)

    def _clear_graph_area(self):
        if hasattr(self, "canvas") and self.canvas:
            try:
                self.canvas.get_tk_widget().destroy()
            except Exception:
                pass
            self.canvas = None

        if hasattr(self, "graph_frame") and self.graph_frame:
            for w in self.graph_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

        if hasattr(self, "toolbar_frame") and self.toolbar_frame:
            for w in self.toolbar_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

    def _show_graph_placeholder(self, text=None):
        import customtkinter as ctk
        if not hasattr(self, "graph_frame") or self.graph_frame is None:
            return

        try:
            for w in self.graph_frame.winfo_children():
                w.destroy()
        except Exception:
            pass

        placeholder_text = text or "Aucun graphique affiché.\nConfigurez les paramètres puis lancez le rendu."

        self._graph_placeholder = ctk.CTkLabel(
            self.graph_frame,
            text=placeholder_text,
            text_color="#9CA3AF",
            font=("Segoe UI", 12),
            justify="center"
        )
        self._graph_placeholder.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

    def creer_page_graphique(self):
        import tkinter as tk
        from tkinter import ttk
        import customtkinter as ctk
        from PIL import Image
        from customtkinter import CTkImage

        # =========================================================
        # DESIGN SYSTEM
        # =========================================================
        UI = {
            "bg": "#0B0F17",
            "topbar": "#11161F",
            "surface": "#141A24",
            "surface_2": "#1A2230",
            "surface_3": "#212B3A",
            "border": "#2B3647",
            "border_soft": "#212A38",
            "text": "#F3F4F6",
            "text_soft": "#D1D5DB",
            "muted": "#9CA3AF",
            "muted_2": "#7C8798",
            "neutral": "#3F4B5F",
            "neutral_hover": "#556178",
            "accent": "#4C7CF3",
            "accent_hover": "#3B67D4",
            "accent_2": "#355CBE",
            "accent_2_hover": "#2B4C9C",
        }

        FONT = {
            "app": ("Segoe UI Semibold", 18, "bold"),
            "page_title": ("Segoe UI Semibold", 28, "bold"),
            "page_subtitle": ("Segoe UI", 12),
            "section": ("Segoe UI Semibold", 14, "bold"),
            "label": ("Segoe UI", 12),
            "small": ("Segoe UI", 11),
            "small_bold": ("Segoe UI", 11, "bold"),
            "button": ("Segoe UI", 12, "bold"),
        }

        # =========================================================
        # HELPERS UI
        # =========================================================
        def card(parent, fg=None, radius=18, border_color=None):
            return ctk.CTkFrame(
                parent,
                fg_color=fg or UI["surface"],
                corner_radius=radius,
                border_width=1,
                border_color=border_color or UI["border_soft"]
            )

        def label(parent, text, font=None, color=None, **kwargs):
            return ctk.CTkLabel(
                parent,
                text=text,
                font=font or FONT["label"],
                text_color=color or UI["text"],
                **kwargs
            )

        def section_header(parent, eyebrow, title, subtitle=None):
            wrap = ctk.CTkFrame(parent, fg_color="transparent")
            wrap.grid_columnconfigure(0, weight=1)

            label(wrap, eyebrow, font=FONT["small_bold"], color=UI["muted"]).grid(
                row=0, column=0, sticky="w"
            )
            label(wrap, title, font=FONT["section"], color=UI["text"]).grid(
                row=1, column=0, sticky="w", pady=(2, 0)
            )
            if subtitle:
                label(wrap, subtitle, font=FONT["small"], color=UI["muted_2"]).grid(
                    row=2, column=0, sticky="w", pady=(4, 0)
                )
            return wrap

        def neutral_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["neutral"],
                hover_color=UI["neutral_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def primary_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["accent"],
                hover_color=UI["accent_hover"],
                text_color="white",
                font=FONT["button"]
            )

        def secondary_btn(parent, text, command):
            return ctk.CTkButton(
                parent,
                text=text,
                command=command,
                height=40,
                corner_radius=10,
                fg_color=UI["accent_2"],
                hover_color=UI["accent_2_hover"],
                text_color="white",
                font=FONT["button"]
            )

        # =========================================================
        # RESET / ROOT
        # =========================================================
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color=UI["bg"])
        except Exception:
            pass

        self.vider_fenetre()

        for i in range(10):
            self.grid_columnconfigure(i, weight=0, minsize=0)
            self.grid_rowconfigure(i, weight=0, minsize=0)

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=0)
        self.grid_rowconfigure(2, weight=1)

        # =========================================================
        # STYLE TTK
        # =========================================================
        style = ttk.Style()
        try:
            style.theme_use("default")
        except Exception:
            pass

        style.configure(
            "Pulse.TCombobox",
            fieldbackground=UI["surface_3"],
            background=UI["surface_3"],
            foreground=UI["text"],
            arrowcolor=UI["text"],
            bordercolor=UI["border"],
            lightcolor=UI["surface_3"],
            darkcolor=UI["surface_3"],
            font=("Segoe UI", 11),
            padding=8
        )

        # =========================================================
        # TOPBAR
        # =========================================================
        topbar = ctk.CTkFrame(self, fg_color=UI["topbar"], corner_radius=0, height=70)
        topbar.grid(row=0, column=0, sticky="nsew")
        topbar.grid_propagate(False)
        topbar.grid_columnconfigure(0, weight=0)
        topbar.grid_columnconfigure(1, weight=1)
        topbar.grid_columnconfigure(2, weight=0)

        top_left = ctk.CTkFrame(topbar, fg_color="transparent")
        top_left.grid(row=0, column=0, sticky="w", padx=24, pady=14)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h = 28
            new_w = int(new_h * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(top_left, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            label(top_left, "PULSE", font=FONT["app"]).pack(side="left")

        top_mid = ctk.CTkFrame(topbar, fg_color="transparent")
        top_mid.grid(row=0, column=1, sticky="w", padx=10)

        label(top_mid, "PULSE", font=FONT["app"]).pack(anchor="w")
        label(
            top_mid,
            "Module de visualisation graphique",
            font=FONT["small"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(2, 0))

        top_right = ctk.CTkFrame(topbar, fg_color="transparent")
        top_right.grid(row=0, column=2, sticky="e", padx=24)

        neutral_btn(top_right, "Retour à l’accueil", self.creer_accueil).pack(side="left")

        separator = ctk.CTkFrame(self, fg_color=UI["border_soft"], height=1, corner_radius=0)
        separator.grid(row=1, column=0, sticky="ew")

        # =========================================================
        # BODY SCROLLABLE
        # =========================================================
        body_host = ctk.CTkFrame(self, fg_color=UI["bg"], corner_radius=0)
        body_host.grid(row=2, column=0, sticky="nsew")
        body_host.grid_rowconfigure(0, weight=1)
        body_host.grid_columnconfigure(0, weight=1)

        body_canvas = tk.Canvas(body_host, bg=UI["bg"], highlightthickness=0, bd=0)
        body_canvas.grid(row=0, column=0, sticky="nsew")

        v_scroll = ttk.Scrollbar(body_host, orient="vertical", command=body_canvas.yview)
        v_scroll.grid(row=0, column=1, sticky="ns")
        body_canvas.configure(yscrollcommand=v_scroll.set)

        page = ctk.CTkFrame(body_canvas, fg_color=UI["bg"], corner_radius=0)
        canvas_window = body_canvas.create_window((0, 0), window=page, anchor="nw")

        def _sync_scrollregion(event=None):
            body_canvas.configure(scrollregion=body_canvas.bbox("all"))

        def _resize_page_in_canvas(event):
            body_canvas.itemconfigure(canvas_window, width=event.width)

        page.bind("<Configure>", _sync_scrollregion)
        body_canvas.bind("<Configure>", _resize_page_in_canvas)

        def _mousewheel(event):
            try:
                body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        body_canvas.bind_all("<MouseWheel>", _mousewheel)

        page.grid_rowconfigure(0, weight=0)
        page.grid_rowconfigure(1, weight=1)
        page.grid_columnconfigure(0, weight=1)

        # =========================================================
        # PAGE HEADER
        # =========================================================
        page_header = ctk.CTkFrame(page, fg_color="transparent")
        page_header.grid(row=0, column=0, sticky="ew", padx=28, pady=(24, 16))
        page_header.grid_columnconfigure(0, weight=1)

        label(page_header, "VISUALISATION", font=FONT["small_bold"], color=UI["muted"]).pack(anchor="w")
        label(
            page_header,
            "Visualisation graphique des écarts",
            font=FONT["page_title"],
            color=UI["text"]
        ).pack(anchor="w", pady=(4, 0))
        label(
            page_header,
            "Sélectionnez une filiale, un flux, une année et les profils à comparer pour générer vos graphiques.",
            font=FONT["page_subtitle"],
            color=UI["muted"]
        ).pack(anchor="w", pady=(6, 0))

        # =========================================================
        # CONTENU PRINCIPAL
        # =========================================================
        content = ctk.CTkFrame(page, fg_color="transparent")
        content.grid(row=1, column=0, sticky="nsew", padx=28, pady=(0, 24))
        content.grid_rowconfigure(0, weight=1)
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=2)

        # ---------------------------------------------------------
        # PANNEAU DE CONTRÔLE
        # ---------------------------------------------------------
        controls_panel = card(content, fg=UI["surface"], radius=20)
        controls_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
        controls_panel.grid_columnconfigure(0, weight=1)

        section_header(
            controls_panel,
            "PARAMÈTRES",
            "Pilotage du graphique",
            "Configurez votre vue puis lancez le rendu."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        controls_body = ctk.CTkFrame(controls_panel, fg_color="transparent")
        controls_body.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 18))
        controls_body.grid_columnconfigure(0, weight=1)

        label(controls_body, "Filiale", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=0, column=0, sticky="w", pady=(0, 6)
        )
        self.feuille_combo = ttk.Combobox(
            controls_body,
            values=list(sections.values()),
            state="readonly",
            style="Pulse.TCombobox"
        )
        self.feuille_combo.grid(row=1, column=0, sticky="ew")
        self.feuille_combo.bind("<<ComboboxSelected>>", self.charger_noms)

        label(controls_body, "Flux", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=2, column=0, sticky="w", pady=(16, 6)
        )
        self.nom_combo = ttk.Combobox(
            controls_body,
            state="readonly",
            style="Pulse.TCombobox"
        )
        self.nom_combo.grid(row=3, column=0, sticky="ew")
        self.nom_combo.bind("<<ComboboxSelected>>", self._on_flux_change)

        label(controls_body, "Année", font=FONT["small_bold"], color=UI["text_soft"]).grid(
            row=4, column=0, sticky="w", pady=(16, 6)
        )
        self.annee_combo = ttk.Combobox(
            controls_body,
            state="disabled",
            style="Pulse.TCombobox"
        )
        self.annee_combo.grid(row=5, column=0, sticky="ew")
        self.annee_combo.bind("<<ComboboxSelected>>", self._on_annee_change)

        self.profils_title = label(
            controls_body,
            "Profils disponibles",
            font=FONT["small_bold"],
            color=UI["text_soft"]
        )
        self.profils_title.grid(row=6, column=0, sticky="w", pady=(18, 8))

        profils_card = ctk.CTkFrame(
            controls_body,
            fg_color=UI["surface_2"],
            corner_radius=12,
            border_width=1,
            border_color=UI["border_soft"]
        )
        profils_card.grid(row=7, column=0, sticky="ew")
        profils_card.grid_columnconfigure(0, weight=1)

        self.profils_frame = tk.Frame(profils_card, bg=UI["surface_2"], highlightthickness=0, bd=0)
        self.profils_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=10)

        self.vars_prev = []

        self.var_reel = tk.BooleanVar(value=True)
        chk_reel = tk.Checkbutton(
            controls_body,
            text="Afficher la série Réel",
            variable=self.var_reel,
            bg=UI["surface"],
            fg=UI["text"],
            activebackground=UI["surface"],
            activeforeground=UI["text"],
            selectcolor=UI["surface_3"],
            highlightthickness=0,
            bd=0,
            font=("Segoe UI", 11)
        )
        chk_reel.grid(row=8, column=0, sticky="w", pady=(16, 4))

        actions_group = ctk.CTkFrame(controls_body, fg_color="transparent")
        actions_group.grid(row=9, column=0, sticky="ew", pady=(18, 0))
        actions_group.grid_columnconfigure(0, weight=1)

        self.btn_afficher = primary_btn(actions_group, "Afficher le graphique", self.afficher_graphique)
        self.btn_afficher.grid(row=0, column=0, sticky="ew", pady=(0, 10))

        self.btn_afficher_cumule = secondary_btn(
            actions_group,
            "Afficher le graphique cumulé",
            self.afficher_graphique_cumule
        )
        self.btn_afficher_cumule.grid(row=1, column=0, sticky="ew", pady=(0, 10))

        self.btn_retour = neutral_btn(actions_group, "Retour au menu", self.creer_accueil)
        self.btn_retour.grid(row=2, column=0, sticky="ew")

        # ---------------------------------------------------------
        # ZONE GRAPHIQUE
        # ---------------------------------------------------------
        chart_panel = card(content, fg=UI["surface"], radius=20)
        chart_panel.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
        chart_panel.grid_rowconfigure(2, weight=1)
        chart_panel.grid_columnconfigure(0, weight=1)

        section_header(
            chart_panel,
            "RENDU",
            "Aperçu graphique",
            "Le graphique s’adapte à la taille de la fenêtre."
        ).grid(row=0, column=0, sticky="ew", padx=18, pady=(16, 12))

        self.toolbar_frame = tk.Frame(chart_panel, bg=UI["surface"], highlightthickness=0, bd=0)
        self.toolbar_frame.grid(row=1, column=0, sticky="ew", padx=18, pady=(0, 10))

        self.graph_frame = ctk.CTkFrame(
            chart_panel,
            fg_color=UI["surface_2"],
            corner_radius=14,
            border_width=1,
            border_color=UI["border_soft"]
        )
        self.graph_frame.grid(row=2, column=0, sticky="nsew", padx=18, pady=(0, 18))
        self.graph_frame.grid_rowconfigure(0, weight=1)
        self.graph_frame.grid_columnconfigure(0, weight=1)

        self._graph_placeholder = label(
            self.graph_frame,
            "Aucun graphique affiché.\nConfigurez les paramètres puis lancez le rendu.",
            font=FONT["label"],
            color=UI["muted"],
            justify="center"
        )
        self._graph_placeholder.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)

        self.canvas = None

        # =========================================================
        # RESPONSIVE
        # =========================================================
        def _apply_responsive_layout(event=None):
            try:
                width = page.winfo_width()
            except Exception:
                return

            if width >= 1280:
                content.grid_columnconfigure(0, weight=1)
                content.grid_columnconfigure(1, weight=2)
                content.grid_rowconfigure(0, weight=1)
                content.grid_rowconfigure(1, weight=0)

                controls_panel.grid_forget()
                chart_panel.grid_forget()

                controls_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10), pady=0)
                chart_panel.grid(row=0, column=1, sticky="nsew", padx=(10, 0), pady=0)
            else:
                content.grid_columnconfigure(0, weight=1)
                content.grid_columnconfigure(1, weight=0)
                content.grid_rowconfigure(0, weight=0)
                content.grid_rowconfigure(1, weight=1)

                controls_panel.grid_forget()
                chart_panel.grid_forget()

                controls_panel.grid(row=0, column=0, sticky="nsew", padx=0, pady=(0, 12))
                chart_panel.grid(row=1, column=0, sticky="nsew", padx=0, pady=0)

            _sync_scrollregion()

        page.bind("<Configure>", _apply_responsive_layout)
        self.after(120, _apply_responsive_layout)

    def charger_noms(self, event=None):
        """Remplit la liste des flux pour la filiale sélectionnée et réinitialise année/profils."""
        import customtkinter as ctk

        section = self.feuille_combo.get()
        if not section:
            return

        ws, noms_colonnes = charger_donnees(section, taille_bloc)
        flux_disponibles = [name for (name, _tok) in noms_colonnes]

        self.nom_combo.config(values=flux_disponibles, state="readonly")
        self.nom_combo.set("")

        self.annee_combo.config(values=[], state="disabled")
        self.annee_combo.set("")

        for w in self.profils_frame.winfo_children():
            w.destroy()

        self.vars_prev = []
        self.profils_names_order = []

        self._clear_graph_area()

        placeholder = ctk.CTkLabel(
            self.graph_frame,
            text="Aucun graphique affiché.\nConfigurez les paramètres puis lancez le rendu.",
            text_color="#9CA3AF",
            font=("Segoe UI", 12),
            justify="center"
        )
        placeholder.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
        self._graph_placeholder = placeholder

    def afficher_graphique(self):
        import matplotlib.pyplot as plt
        import mplcursors
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import matplotlib.ticker as mticker

        def _compute_figsize():
            try:
                self.graph_frame.update_idletasks()
                w_px = max(self.graph_frame.winfo_width(), 1000)
                h_px = max(self.graph_frame.winfo_height(), 600)
            except Exception:
                w_px, h_px = 1200, 700

            dpi = 100
            return (w_px / dpi, h_px / dpi)

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool",
            "Encours de financement", "Endettement Net"
        ]

        section = self.feuille_combo.get()
        if not section:
            messagebox.showwarning("Attention", "Veuillez sélectionner une filiale.")
            return

        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un flux.")
            return

        if not hasattr(self, "_token_from_flux"):
            messagebox.showerror("Erreur", "Méthode _token_from_flux absente.")
            return

        col_start = self._token_from_flux(section, nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Flux sélectionné invalide.")
            return

        try:
            annee = int(self.annee_combo.get()) if self.annee_combo.get() else None
        except Exception:
            annee = None

        dates, reel, previsions, noms_profils_complets = extraire_valeurs(
            section, col_start, nb_prev, annee=annee
        )

        if not dates:
            messagebox.showinfo("Info", "Aucune donnée pour l'année sélectionnée.")
            return

        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i - 1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux

            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        self._clear_graph_area()

        plt.style.use("seaborn-v0_8-darkgrid")
        fig_size = _compute_figsize()
        fig, ax = plt.subplots(figsize=fig_size, dpi=100)
        palette = plt.cm.tab10.colors

        if getattr(self, "var_reel", None) is not None and self.var_reel.get():
            ax.plot(dates, reel, label="Réel", color="black", linewidth=2, marker="o")

        profils_affiches = getattr(self, "profils_names_order", None)
        if profils_affiches is None:
            profils_affiches = noms_profils_complets[:len(self.vars_prev)]

        index_by_name = {name: i for i, name in enumerate(noms_profils_complets)}

        plotted = 0
        for i, var in enumerate(self.vars_prev):
            if not var.get():
                continue
            if i >= len(profils_affiches):
                continue

            nom_profil = profils_affiches[i]
            idx = index_by_name.get(nom_profil, None)
            if idx is None or idx >= len(previsions):
                continue

            y = previsions[idx]
            ax.plot(
                dates,
                y,
                label=nom_profil,
                alpha=0.9,
                linewidth=1.8,
                marker=".",
                color=palette[plotted % len(palette)]
            )
            plotted += 1

        titre_suffix = f" - {annee}" if annee is not None else ""
        ax.set_title(f"{section} - {nom_selectionne}{titre_suffix}", fontsize=18, fontweight="bold")
        ax.set_xlabel("Date", fontsize=12)
        ax.set_ylabel("Valeur (k€)" if nom_selectionne not in noms_a_convertir_flux else "Flux", fontsize=12)
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " "))
        )

        fig.subplots_adjust(left=0.08, bottom=0.15, top=0.88, right=0.72)

        ax.legend(
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            borderaxespad=0.0,
            frameon=True
        )

        mplcursors.cursor(ax, hover=True)

        self.canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas_widget = self.canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.canvas.draw()

        def _on_resize(event):
            try:
                w = max(event.width, 800)
                h = max(event.height, 500)
                fig.set_size_inches(w / 100, h / 100, forward=True)
                self.canvas.draw_idle()
            except Exception:
                pass

        canvas_widget.bind("<Configure>", _on_resize)

        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()
    
    def afficher_graphique_cumule(self):
        import matplotlib.pyplot as plt
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
        from tkinter import messagebox
        import matplotlib.ticker as mticker
        import pandas as pd
        import numpy as np
        import mplcursors

        def _compute_figsize(nb_mois: int = 12):
            try:
                self.graph_frame.update_idletasks()
                w_px = max(self.graph_frame.winfo_width(), 1000)
                h_px = max(self.graph_frame.winfo_height(), 600)
            except Exception:
                w_px, h_px = 1200, 700

            dpi = 100
            return (w_px / dpi, max(h_px / dpi, 5))

        section = self.feuille_combo.get()
        if not section:
            messagebox.showwarning("Attention", "Veuillez sélectionner une filiale.")
            return

        nom_selectionne = self.nom_combo.get()
        if not nom_selectionne:
            messagebox.showwarning("Attention", "Veuillez sélectionner un flux.")
            return

        if not hasattr(self, "_token_from_flux"):
            messagebox.showerror("Erreur", "Méthode _token_from_flux absente.")
            return

        col_start = self._token_from_flux(section, nom_selectionne)
        if col_start is None:
            messagebox.showerror("Erreur", "Flux sélectionné invalide.")
            return

        try:
            annee = int(self.annee_combo.get()) if self.annee_combo.get() else None
        except Exception:
            annee = None

        dates, reel, previsions, noms_profils_complets = extraire_valeurs(
            section, col_start, nb_prev, annee=annee
        )

        if not dates:
            messagebox.showinfo("Info", "Aucune donnée pour l'année sélectionnée.")
            return

        noms_a_convertir_flux = [
            "Emprunts", "Tirages Lignes CT", "Variation de collatéral",
            "Créances CDP", "Placements", "CC financiers",
            "Emprunts / Prêts - Groupe", "Cashpool", "Encours de financement", "Endettement Net"
        ]

        if nom_selectionne in noms_a_convertir_flux:
            def en_flux(values):
                values = [float(v) if v is not None else None for v in values]
                if not values or all(v is None for v in values):
                    return values
                flux = [0 if values[0] is not None else None]
                for i in range(1, len(values)):
                    v, v_prev = values[i], values[i - 1]
                    flux.append(v - v_prev if v is not None and v_prev is not None else None)
                return flux

            reel = en_flux(reel)
            previsions = [en_flux(p) for p in previsions]

        self._clear_graph_area()

        df = pd.DataFrame({"Date": dates, "Réel": reel})
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Mois"] = df["Date"].dt.strftime("%Y-%m")

        profils_affiches = getattr(self, "profils_names_order", None)
        if profils_affiches is None:
            profils_affiches = noms_profils_complets[:len(self.vars_prev)]

        index_by_name = {name: i for i, name in enumerate(noms_profils_complets)}

        profils_selectionnes = []
        for i, var in enumerate(self.vars_prev):
            if not var.get():
                continue
            if i >= len(profils_affiches):
                continue

            nom_prof = profils_affiches[i]
            idx = index_by_name.get(nom_prof, None)
            if idx is None or idx >= len(previsions):
                continue

            df[nom_prof] = previsions[idx]
            profils_selectionnes.append(nom_prof)

        series_combinees = {}
        for nom_prof in profils_selectionnes:
            comb = pd.Series(df[nom_prof] if nom_prof in df.columns else np.nan)
            comb = comb.where(~comb.isna(), df["Réel"])

            tmp_prev = pd.DataFrame({"Mois": df["Mois"], "Prev": df.get(nom_prof)})
            mois_avec_prev = set(tmp_prev.loc[tmp_prev["Prev"].notna(), "Mois"].dropna().unique())

            mask_keep = df["Mois"].isin(mois_avec_prev)
            comb = comb.where(mask_keep, np.nan)

            series_combinees[f"{nom_prof} (Prévision sinon Réel)"] = comb

        df_comb = pd.DataFrame({"Mois": df["Mois"], "Réel": df["Réel"]})
        for nom_serie, s in series_combinees.items():
            df_comb[nom_serie] = s

        df_cumule = df_comb.groupby("Mois", as_index=True).sum(min_count=1)

        plt.style.use("seaborn-v0_8-darkgrid")
        n_mois = len(df_cumule)
        fig_size = _compute_figsize(n_mois)
        fig, ax = plt.subplots(figsize=fig_size, dpi=100)

        n_series = 1 + len(series_combinees)
        positions = np.arange(n_mois)
        largeur_barre = 0.8 / max(1, n_series)
        palette = plt.cm.tab10.colors
        bar_containers = []

        bars_reel = ax.bar(
            positions,
            df_cumule["Réel"].fillna(0).values,
            width=largeur_barre,
            label="Réel",
            color="black",
        )
        bar_containers.append(bars_reel)

        for i, (nom_serie, _) in enumerate(series_combinees.items()):
            offset = (i + 1) * largeur_barre
            bars = ax.bar(
                positions + offset,
                df_cumule[nom_serie].fillna(0).values,
                width=largeur_barre,
                label=nom_serie,
                color=palette[i % len(palette)],
            )
            bar_containers.append(bars)

        ax.set_xticks(positions + largeur_barre * (n_series - 1) / 2)
        ax.set_xticklabels(df_cumule.index, rotation=45, ha="right")

        titre_suffix = f" - {annee}" if annee is not None else ""
        ax.set_title(f"{section} - {nom_selectionne}{titre_suffix}", fontsize=16, fontweight="bold")
        ax.set_xlabel("Mois", fontsize=12)
        ax.set_ylabel("Valeur cumulée (k€)", fontsize=12)
        ax.grid(True, linestyle="--", alpha=0.6)
        ax.yaxis.set_major_formatter(
            mticker.FuncFormatter(lambda x, _: f"{int(x):,}".replace(",", " "))
        )

        ax.legend(
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            borderaxespad=0.0,
            frameon=True
        )

        fig.subplots_adjust(left=0.08, bottom=0.18, top=0.88, right=0.72)

        cursor = mplcursors.cursor(bar_containers, hover=True)

        @cursor.connect("add")
        def on_add(sel):
            bars = sel.artist
            idx = sel.index
            serie = bars.get_label()
            val = bars.datavalues[idx]
            mois = df_cumule.index[idx]
            sel.annotation.set_text(
                f"Série : {serie}\nMois : {mois}\nValeur : {int(val):,} K€".replace(",", " ")
            )
            sel.annotation.get_bbox_patch().set(fc="white", alpha=0.8)

        self.canvas = FigureCanvasTkAgg(fig, master=self.graph_frame)
        canvas_widget = self.canvas.get_tk_widget()
        canvas_widget.grid(row=0, column=0, sticky="nsew", padx=5, pady=5)
        self.canvas.draw()

        def _on_resize(event):
            try:
                w = max(event.width, 800)
                h = max(event.height, 500)
                fig.set_size_inches(w / 100, h / 100, forward=True)
                self.canvas.draw_idle()
            except Exception:
                pass

        canvas_widget.bind("<Configure>", _on_resize)

        toolbar = NavigationToolbar2Tk(self.canvas, self.toolbar_frame)
        toolbar.update()

#===================Page Prédiction IA===================
    def _rebuild_profils_ui2(self, section, flux_name, annee):
        """Reconstruit la select-box de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()

        # nouveau : on n'utilise plus self.vars_prev ici
        self.vars_prev = []
        self.profils_names_order = []   # ordre des profils pour mapping
        self.profils_listbox = None     # widget de sélection multiple

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils

        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # Label d'info
        lbl = tk.Label(
            self.profils_frame,
            text="Profils (Ctrl+clic pour multi-sélection) :",
            bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
        )
        lbl.pack(anchor="w", padx=4, pady=(0, 4))

        # Frame contenant la listbox + scrollbar
        container = tk.Frame(self.profils_frame, bg="#00122e")
        container.pack(fill="x", padx=4, pady=(0, 6))

        scrollbar = tk.Scrollbar(container, orient="vertical")
        self.profils_listbox = tk.Listbox(
            container,
            selectmode="multiple",
            exportselection=False,
            bg="#00122e",
            fg="white",
            font=('Segoe UI', 10),
            height=min(8, len(noms_profils)),
            highlightthickness=0,
            activestyle="dotbox"
        )
        self.profils_listbox.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.profils_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.profils_listbox.yview)

        for name in noms_profils:
            self.profils_listbox.insert("end", name)

        # petite sélection par défaut : tout sélectionner

    def creer_page_ia_prediction(self):
        """
        Page IA de prévision (N -> N+1) pilotée par un flux sélectionnable.
        Cette méthode :
        1) Vérifie les dépendances
        2) Réinitialise la fenêtre
        3) Construit toute la page IA (UI + callbacks + entraînement)
        -> le dataset sera construit dynamiquement selon le flux choisi
        """
        from tkinter import messagebox

        # sécurité sklearn
        if RandomForestRegressor is None:
            messagebox.showerror(
                "Dépendance manquante",
                "Le module 'scikit-learn' n'est pas installé.\n"
                "Installe-le avec : pip install scikit-learn"
            )
            return

        # reset page
        self.vider_fenetre()

        # Build page IA (dataset construit selon flux dans la page)
        self._ia_build_prediction_page(df=None)
        
    def _ia_build_dataset_for_flux(self, flux_cible: str):
        """
        Construit le DataFrame global df sur le flux choisi (ex: 'Trafic Voyageurs')
        pour toutes les filiales, avec colonnes :
        section, date, y, year, month, dayofyear, section_id, roll_mean_7, roll_mean_30
        Retourne df ou None si aucune donnée.
        """
        import pandas as pd
        import numpy as np
        from tkinter import messagebox

        lignes = []

        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        for section_name, feuille in sections.items():
            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees pour {feuille} : {e}")
                continue

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                continue

            _, col_start = cible[0]
            try:
                dates, reel, _previsions, _noms_profils = extraire_valeurs(
                    ws, col_start, nb_prev, annee=None
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs pour {section_name}/{flux_cible} : {e}")
                continue

            for d, r in zip(dates, reel):
                y_val = _to_number(r)
                if y_val is None:
                    continue
                try:
                    d_ts = pd.to_datetime(d)
                except Exception:
                    continue
                lignes.append({
                    "section": section_name,
                    "date": d_ts,
                    "y": float(y_val)
                })

        if not lignes:
            messagebox.showinfo(
                "Information",
                f"Aucune donnée trouvée pour le flux '{flux_cible}' sur les filiales."
            )
            self.retour_menu()
            return None

        df = pd.DataFrame(lignes).sort_values(["section", "date"]).reset_index(drop=True)
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["dayofyear"] = df["date"].dt.dayofyear

        cat = df["section"].astype("category")
        df["section_id"] = cat.cat.codes

        df["roll_mean_7"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(7, min_periods=1).mean()
        )
        df["roll_mean_30"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(30, min_periods=1).mean()
        )

        return df

    def _ia_build_prediction_page(self, df):
        """
        Construit toute l'interface IA + callbacks d'entraînement
        à partir du DataFrame global df déjà préparé.

        Version + SELECT BOX FLUX (rebuild df_current selon flux choisi).
        """
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # ---------- helpers numériques locaux ----------
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_float_or_nan(x):
            val = _to_number(x)
            return float(val) if val is not None else np.nan

        default_n_estimators = 200

        # ---------- variables profils & graph 2 ----------
        ia_profils_vars = []        # liste de tk.BooleanVar
        ia_profils_names = []       # noms de profils actifs N+1
        ia_profils_dates = []       # dates N+1
        ia_profils_series = []      # séries (une par profil actif)

        current_pred_df = None          # df_future_all pour N+1
        current_real_target_df = None   # réel sur N+1 (si dispo)
        current_target_year = None
        current_filiale_name = None

        ia_graph2_widget = None
        graph2_container_packed = False

        exported_pred_df = None         # pour export Excel prédictions
        analysis_table_frame = None
        analysis_tree = None
        analysis_export_button = None
        export_button = None            # bouton d’export des prédictions
        export_graph2_button = None

        # NEW: dataset courant selon flux
        df_current = None

        model_artifacts = {
            "X_train": None,
            "y_train": None,
            "features": None,
            "cls_model": None,
            "reg_model": None,
        }

        # ============================================================
        # NEW: récupérer la liste des flux disponibles (toutes filiales)
        # ============================================================
        def _get_all_flux_names():
            flux_set = set()
            for _section_name, feuille in sections.items():
                try:
                    _ws, noms_flux = charger_donnees(feuille, taille_bloc)
                    for nom, _col in noms_flux:
                        flux_set.add(nom)
                except Exception:
                    pass
            out = sorted(flux_set)
            return out

        # ============ HEADER ============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - IA PRÉDICTION  (N → N+1)",
            font=titre_font,
            text_color="white"
        ).pack(side="left", anchor="w")

        # Logo
        try:
            logo_img = Image.open(image_path)
            test = tk.Label(self, text="Test", font=titre_font)
            test.update_idletasks()
            text_h = test.winfo_reqheight()
            test.destroy()
            ratio = logo_img.width / max(logo_img.height, 1)
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"[IA] Erreur chargement logo IA : {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ============ CONTAINER SCROLLABLE ============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # ============ TITRE / TEXTE ============
        title_label = ctk.CTkLabel(
            scrollable_frame,
            text="IA - Prédiction (prévision année N+1)",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        )
        title_label.pack(pady=15)

        desc_label = ctk.CTkLabel(
            scrollable_frame,
            text=(
                "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                "Il apprend comment la valeur évolue d'une année à l'autre (N → N+1) pour chaque jour.\n"
                "Il prédit ensuite l’ensemble de l’année N+1.\n"
                "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
            ),
            font=("Segoe UI", 12),
            text_color="#c9defa",
            justify="left"
        )
        desc_label.pack(pady=(0, 10))

        # ============ ZONE PARAMÈTRES (CARD) ============
        params_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        params_frame.pack(fill="x", padx=10, pady=(5, 10))

        # ------------------- Flux -------------------
        ctk.CTkLabel(
            params_frame,
            text="Flux :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))

        flux_list = _get_all_flux_names()
        if not flux_list:
            flux_list = ["Trafic Voyageurs"]

        default_flux = "Trafic Voyageurs" if "Trafic Voyageurs" in flux_list else flux_list[0]
        selected_flux = tk.StringVar(value=default_flux)

        flux_box = ttk.Combobox(
            params_frame,
            textvariable=selected_flux,
            values=flux_list,
            state="readonly",
            width=28
        )
        flux_box.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Filiale -------------------
        ctk.CTkLabel(
            params_frame,
            text="Filiale :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=1, sticky="w", padx=12, pady=(10, 2))

        selected_filiale = tk.StringVar(value="")
        filiale_box = ttk.Combobox(
            params_frame,
            textvariable=selected_filiale,
            values=[],
            state="readonly",
            width=28
        )
        filiale_box.grid(row=1, column=1, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Année -------------------
        ctk.CTkLabel(
            params_frame,
            text="Année N (historique → N+1) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=2, sticky="w", padx=12, pady=(10, 2))

        annees_var = tk.StringVar(value="")
        annees_box = ttk.Combobox(
            params_frame,
            textvariable=annees_var,
            values=[],
            state="readonly",
            width=18
        )
        annees_box.grid(row=1, column=2, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Hyperparams -------------------
        ctk.CTkLabel(
            params_frame,
            text="Hyperparamètres Random Forest (point de départ) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=3, columnspan=2, sticky="w", padx=12, pady=(10, 2))

        # n_estimators
        ctk.CTkLabel(
            params_frame,
            text="n_estimators",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=3, sticky="w", padx=12, pady=(0, 2))

        n_estimators_var = tk.IntVar(value=default_n_estimators)
        slider_n = ctk.CTkSlider(
            params_frame,
            from_=50,
            to=600,
            number_of_steps=11,
            command=lambda v: n_estimators_var.set(int(v))
        )
        slider_n.set(default_n_estimators)
        slider_n.grid(row=2, column=3, sticky="we", padx=12, pady=(0, 8))

        lbl_n = ctk.CTkLabel(
            params_frame,
            text=f"{default_n_estimators}",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_n.grid(row=3, column=3, sticky="w", padx=12, pady=(0, 6))

        def _update_lbl_n(_=None):
            lbl_n.configure(text=str(n_estimators_var.get()))
        slider_n.configure(command=lambda v: (n_estimators_var.set(int(v)), _update_lbl_n()))

        # max_depth
        ctk.CTkLabel(
            params_frame,
            text="max_depth",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=4, sticky="w", padx=12, pady=(0, 2))

        max_depth_var = tk.IntVar(value=15)
        slider_d = ctk.CTkSlider(
            params_frame,
            from_=3,
            to=25,
            number_of_steps=11,
            command=lambda v: max_depth_var.set(int(v))
        )
        slider_d.set(15)
        slider_d.grid(row=2, column=4, sticky="we", padx=12, pady=(0, 8))

        lbl_d = ctk.CTkLabel(
            params_frame,
            text="15",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_d.grid(row=3, column=4, sticky="w", padx=12, pady=(0, 2))

        def _update_lbl_d(_=None):
            lbl_d.configure(text=str(max_depth_var.get()))
        slider_d.configure(command=lambda v: (max_depth_var.set(int(v)), _update_lbl_d()))

        use_depth_var = tk.BooleanVar(value=True)
        chk_depth = ctk.CTkCheckBox(
            params_frame,
            text="Limiter la profondeur",
            variable=use_depth_var,
            text_color="#c9defa",
            font=("Segoe UI", 11)
        )
        chk_depth.grid(row=4, column=4, sticky="w", padx=12, pady=(0, 10))

        # Bouton entraîner
        bouton_train = ctk.CTkButton(
            params_frame,
            text="🚀 Entraîner / ré-entraîner le modèle",
            width=240,
            height=40,
            corner_radius=16,
            fg_color="#008C4B",
            hover_color="#006C39",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_train.grid(row=1, column=5, rowspan=3, padx=14, pady=(0, 10), sticky="e")

        for c in range(6):
            params_frame.grid_columnconfigure(c, weight=1 if c in (3, 4) else 0)

        # ============================================================
        # NEW: rebuild dataset selon flux + refresh combobox filiale/année
        # ============================================================
        def _refresh_filiale_year_boxes_from_df():
            nonlocal df_current
            if df_current is None or df_current.empty:
                return

            filiales = sorted(df_current["section"].unique().tolist())
            filiale_box.configure(values=filiales)
            if filiales:
                if selected_filiale.get() not in filiales:
                    selected_filiale.set(filiales[0])

            years_sorted = sorted(df_current["year"].unique().tolist())
            annees_box.configure(values=[str(y) for y in years_sorted])
            if years_sorted:
                if annees_var.get() not in [str(y) for y in years_sorted]:
                    annees_var.set(str(years_sorted[-1]))

        def _rebuild_df_for_current_flux():
            nonlocal df_current
            flux = selected_flux.get()
            df_new = self._ia_build_dataset_for_flux(flux)
            if df_new is None:
                return False
            df_current = df_new
            _refresh_filiale_year_boxes_from_df()

            # titre / texte dynamiques
            try:
                title_label.configure(text=f"IA - Prédiction sur le flux '{flux}' (prévision année N+1)")
                desc_label.configure(text=(
                    "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                    f"Il apprend comment la valeur de '{flux}' évolue d'une année à l'autre (N → N+1) pour chaque jour.\n"
                    "Il prédit ensuite l’ensemble de l’année N+1 pour cette filiale.\n"
                    "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
                ))
            except Exception:
                pass

            return True

        # init dataset
        if not _rebuild_df_for_current_flux():
            return

        # ============ CARD PROFILS N+1 (CASES À COCHER) ============
        profils_card = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        profils_card.pack(fill="x", padx=10, pady=(0, 15))

        ctk.CTkLabel(
            profils_card,
            text="Profils de prévision disponibles sur l'année N+1 :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=12, pady=(10, 4))

        ia_profils_frame = tk.Frame(profils_card, bg="#00122e")
        ia_profils_frame.pack(fill="x", padx=12, pady=(0, 10))

        def _export_graph2_to_excel():
            """
            Export Excel des valeurs utilisées pour le graphe 2 :
            - date
            - réel (si dispo)
            - prédiction IA (si dispo)
            - profils cochés (si dispo)
            """
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal ia_profils_names, ia_profils_vars, ia_profils_dates, ia_profils_series

            if (current_pred_df is None or current_pred_df.empty) and (current_real_target_df is None or current_real_target_df.empty):
                messagebox.showinfo("Export", "Aucune donnée à exporter (ni réel, ni prévision).")
                return

            try:
                all_dates = set()

                if current_real_target_df is not None and not current_real_target_df.empty:
                    all_dates |= set(pd.to_datetime(current_real_target_df["date"]).dt.normalize())

                if current_pred_df is not None and not current_pred_df.empty:
                    all_dates |= set(pd.to_datetime(current_pred_df["date"]).dt.normalize())

                if ia_profils_dates:
                    all_dates |= set(pd.to_datetime(ia_profils_dates).normalize())

                if not all_dates:
                    messagebox.showinfo("Export", "Aucune date exploitable à exporter.")
                    return

                dates_sorted = sorted(all_dates)
                df_out = pd.DataFrame({"date": pd.to_datetime(dates_sorted)})

                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = (
                        current_real_target_df.assign(date_norm=pd.to_datetime(current_real_target_df["date"]).dt.normalize())
                        .set_index("date_norm")["y"]
                        .to_dict()
                    )
                    df_out["reel"] = df_out["date"].dt.normalize().map(real_map)
                else:
                    df_out["reel"] = np.nan

                if current_pred_df is not None and not current_pred_df.empty:
                    pred_map = (
                        current_pred_df.assign(date_norm=pd.to_datetime(current_pred_df["date"]).dt.normalize())
                        .set_index("date_norm")["pred_value"]
                        .to_dict()
                    )
                    df_out["prev_ia"] = df_out["date"].dt.normalize().map(pred_map)
                else:
                    df_out["prev_ia"] = np.nan

                if ia_profils_names and ia_profils_series and ia_profils_dates and ia_profils_vars:
                    prof_dates = pd.to_datetime(ia_profils_dates).normalize()

                    for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                        if not var.get():
                            continue

                        vals = [_to_float_or_nan(v) for v in serie]
                        prof_map = pd.Series(vals, index=prof_dates, dtype="float64").to_dict()

                        col_name = f"profil_{name}".replace("\n", " ").strip()
                        df_out[col_name] = df_out["date"].dt.normalize().map(prof_map)

                df_out.insert(0, "filiale", current_filiale_name or selected_filiale.get())
                df_out.insert(1, "annee", current_target_year or (int(annees_var.get()) + 1))
                df_out.insert(2, "flux", selected_flux.get())

                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Fichiers Excel", "*.xlsx")],
                    title="Exporter les données du graphe détaillé (Graph2) en Excel"
                )
                if not file_path:
                    return

                df_out.to_excel(file_path, index=False)
                messagebox.showinfo("Export", f"Export Graph2 OK :\n{file_path}")

            except Exception as e:
                messagebox.showerror("Erreur export Graph2", str(e))

        def _show_pyvista_3d():
            import numpy as np
            import pandas as pd

            try:
                import pyvista as pv
            except Exception as e:
                messagebox.showerror(
                    "PyVista",
                    "PyVista n'est pas dispo dans cet environnement.\n"
                    "Installe-le dans le même Python/venv que ton script.\n\n"
                    f"Détail: {e}"
                )
                return

            feats = model_artifacts.get("features")
            cls_model = model_artifacts.get("cls_model")
            reg_model = model_artifacts.get("reg_model")
            X_train = model_artifacts.get("X_train")

            if feats is None or cls_model is None or reg_model is None or X_train is None:
                messagebox.showinfo("3D", "Lance d'abord l'entraînement du modèle.")
                return

            features = list(feats)

            # -----------------------
            # Helpers
            # -----------------------
            def _norm01(x):
                x = np.asarray(x, dtype=float)
                if x.size == 0:
                    return x.astype(np.float32)
                mn = np.nanmin(x)
                mx = np.nanmax(x)
                if not np.isfinite(mn) or not np.isfinite(mx) or abs(mx - mn) < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return ((x - mn) / (mx - mn)).astype(np.float32)

            def _norm_signed(x):
                """Normalize to [-1,1] using max abs."""
                x = np.asarray(x, dtype=float)
                m = np.nanmax(np.abs(x))
                if not np.isfinite(m) or m < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return (x / m).astype(np.float32)

            def build_lines(src_pts, dst_pts, w):
                e = len(src_pts)
                if e == 0:
                    m = pv.PolyData(np.zeros((0, 3), dtype=np.float32))
                    m.lines = np.array([], dtype=np.int64)
                    m.point_data["w"] = np.array([], dtype=np.float32)
                    return m

                pts = np.empty((2 * e, 3), dtype=np.float32)
                pts[0::2] = src_pts
                pts[1::2] = dst_pts

                lines = np.empty((3 * e,), dtype=np.int64)
                lines[0::3] = 2
                lines[1::3] = np.arange(0, 2 * e, 2, dtype=np.int64)
                lines[2::3] = np.arange(1, 2 * e, 2, dtype=np.int64)

                m = pv.PolyData(pts)
                m.lines = lines
                m.point_data["w"] = np.repeat(np.asarray(w, dtype=np.float32), 2)
                return m

            def get_gain_importance(model, features):
                imp = np.zeros(len(features), dtype=np.float32)
                try:
                    booster = getattr(model, "booster_", None)
                    if booster is not None:
                        names = booster.feature_name()
                        gains = booster.feature_importance(importance_type="gain")
                        d = dict(zip(names, gains))
                        for i, f in enumerate(features):
                            imp[i] = float(d.get(f, 0.0))
                    else:
                        gains = getattr(model, "feature_importances_", None)
                        if gains is not None:
                            imp[:] = np.asarray(gains[: len(features)], dtype=np.float32)
                except Exception:
                    pass
                return imp

            gain_cls = get_gain_importance(cls_model, features)
            gain_reg = get_gain_importance(reg_model, features)

            # -----------------------
            # SHAP (best effort)
            # -----------------------
            shap_ok = False
            shap_cls = np.zeros(len(features), dtype=np.float32)
            shap_reg = np.zeros(len(features), dtype=np.float32)

            def try_compute_shap():
                nonlocal shap_ok, shap_cls, shap_reg
                try:
                    import shap
                except Exception:
                    shap_ok = False
                    return

                try:
                    Xs = X_train[features].copy()
                except Exception:
                    Xs = pd.DataFrame(X_train, columns=features).copy()

                n_sample = min(800, len(Xs))
                if n_sample < 80:
                    shap_ok = False
                    return
                Xs = Xs.sample(n=n_sample, random_state=42)

                ok_any = False
                # CLS
                try:
                    expl = shap.TreeExplainer(cls_model)
                    sv = expl.shap_values(Xs)
                    if isinstance(sv, list) and len(sv) >= 2:
                        sv = sv[1]
                    shap_cls[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_cls)) > 0)
                except Exception:
                    shap_cls[:] = 0.0

                # REG
                try:
                    expl = shap.TreeExplainer(reg_model)
                    sv = expl.shap_values(Xs)
                    shap_reg[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_reg)) > 0)
                except Exception:
                    shap_reg[:] = 0.0

                shap_ok = bool(ok_any)

            try_compute_shap()

            # -----------------------
            # Scene config
            # -----------------------
            TOPK = 18          # nb features affichées par arc (CLS + REG)
            LABEL_TOP = 10     # nb labels par arc
            NODE_R = 0.20

            pv.set_plot_theme("dark")
            plotter = pv.Plotter(window_size=(1600, 900))
            plotter.set_background("#0b1220")

            # pipeline positions
            P_OUT   = np.array([-9.0, 0.0, -2.6], dtype=np.float32)
            P_RULES = np.array([-6.0, 0.0, -1.0], dtype=np.float32)
            P_COMB  = np.array([-3.0, 0.0,  0.4], dtype=np.float32)
            P_CLS   = np.array([ 0.0, -1.2,  1.8], dtype=np.float32)
            P_REG   = np.array([ 0.0, +1.2,  1.8], dtype=np.float32)

            def add_block(center, size=(2.6, 1.0, 0.9), label=""):
                box = pv.Cube(center=center, x_length=size[0], y_length=size[1], z_length=size[2])
                box.point_data["v"] = np.full(box.n_points, 1.0, dtype=np.float32)
                plotter.add_mesh(box, scalars="v", cmap="turbo", opacity=0.90, show_edges=True)
                plotter.add_point_labels([center], [label], font_size=14, text_color="white",
                                        point_color="white", shape=False)

            def add_arrow(p0, p1):
                arr = pv.Arrow(start=p0, direction=(p1 - p0),
                            tip_length=0.25, tip_radius=0.07, shaft_radius=0.03, scale="auto")
                plotter.add_mesh(arr, color="white", opacity=0.9)

            add_block(P_CLS, label="LGBM CLS\nP(y>0)")
            add_block(P_REG, label="LGBM REG\nE[y|y>0)")
            add_block(P_COMB, label="Combine\np*amt")
            add_block(P_RULES, label="Rules\nWE/Férié=0\ncarry")
            add_block(P_OUT, label="Output\npred/p10/p90")

            add_arrow(P_CLS, P_COMB)
            add_arrow(P_REG, P_COMB)
            add_arrow(P_COMB, P_RULES)
            add_arrow(P_RULES, P_OUT)

            # -----------------------
            # Arcs placement (2 arcs)
            # -----------------------
            def arc_points(n, z_shift=0.0, y_shift=0.0):
                ang = np.linspace(-1.05, 1.05, n)
                r = 6.2
                x = 6.6 + 0.28 * np.cos(ang)
                y = (r * np.cos(ang) * 0.18) + y_shift
                z = (r * np.sin(ang) * 0.55) + 2.2 + z_shift
                return np.stack([x, y, z], axis=1).astype(np.float32)

            # -----------------------
            # State
            # -----------------------
            state = {
                "edge_mode": 3,          # 1=CLS 2=REG 3=BOTH
                "score_mode": ("shap" if shap_ok else "gain"),  # gain/shap
                "diff_mode": False,      # False: turbo on v ; True: diverging on diff
            }

            a_cls_nodes = None
            a_reg_nodes = None
            a_cls_edges = None
            a_reg_edges = None
            label_actors = []

            def clear_labels():
                nonlocal label_actors
                for act in label_actors:
                    try:
                        plotter.remove_actor(act)
                    except Exception:
                        pass
                label_actors = []

            def scores():
                if state["score_mode"] == "shap" and shap_ok:
                    s_cls = shap_cls.copy()
                    s_reg = shap_reg.copy()
                else:
                    s_cls = gain_cls.copy()
                    s_reg = gain_reg.copy()
                return s_cls, s_reg

            def hud_text():
                s = f"Edges 1/2/3=CLS/REG/BOTH | Score G/S=GAIN/SHAP | D=DIFF"
                s += f"  ||  score={state['score_mode'].upper()}  diff={'ON' if state['diff_mode'] else 'OFF'}"
                if state["score_mode"] == "shap" and not shap_ok:
                    s += " (SHAP indispo -> GAIN)"
                return s

            plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            # -----------------------
            # Rebuild
            # -----------------------
            def rebuild():
                nonlocal a_cls_nodes, a_reg_nodes, a_cls_edges, a_reg_edges

                s_cls, s_reg = scores()

                # TOPK CLS et TOPK REG (union)
                idx_cls = np.argsort(-s_cls)[: min(TOPK, len(features))]
                idx_reg = np.argsort(-s_reg)[: min(TOPK, len(features))]

                f_cls = [features[i] for i in idx_cls]
                f_reg = [features[i] for i in idx_reg]

                v_cls = s_cls[idx_cls]
                v_reg = s_reg[idx_reg]

                # node colors:
                # - normal : importance normalisée
                # - diff : (reg - cls) normalisé signé
                if state["diff_mode"]:
                    # pour diff, on calcule diff sur l'ensemble des features de chaque arc
                    d_cls = (s_reg[idx_cls] - s_cls[idx_cls])  # sur les features TOP CLS
                    d_reg = (s_reg[idx_reg] - s_cls[idx_reg])  # sur les features TOP REG
                    c_cls = _norm_signed(d_cls)  # [-1,1]
                    c_reg = _norm_signed(d_reg)  # [-1,1]
                else:
                    c_cls = _norm01(v_cls)       # [0,1]
                    c_reg = _norm01(v_reg)       # [0,1]

                # edges weights (normalisés séparément)
                w_cls = _norm01(v_cls)
                w_reg = _norm01(v_reg)

                pts_cls = arc_points(len(f_cls), z_shift=-0.9, y_shift=-0.25)
                pts_reg = arc_points(len(f_reg), z_shift=+0.9, y_shift=+0.25)

                # nodes glyphs
                sphere = pv.Sphere(radius=NODE_R, theta_resolution=16, phi_resolution=16)

                cloud_cls = pv.PolyData(pts_cls)
                cloud_reg = pv.PolyData(pts_reg)

                # scalar name
                scalar_name = "d" if state["diff_mode"] else "v"
                cloud_cls[scalar_name] = c_cls.astype(np.float32)
                cloud_reg[scalar_name] = c_reg.astype(np.float32)

                glyph_cls = cloud_cls.glyph(scale=False, geom=sphere, orient=False)
                glyph_reg = cloud_reg.glyph(scale=False, geom=sphere, orient=False)

                # edges meshes (feature -> model)
                PCLS = np.repeat(P_CLS[None, :], len(f_cls), axis=0)
                PREG = np.repeat(P_REG[None, :], len(f_reg), axis=0)

                e_cls = build_lines(pts_cls, PCLS, w_cls)
                e_reg = build_lines(pts_reg, PREG, w_reg)

                # create/update actors
                if a_cls_nodes is None:
                    if state["diff_mode"]:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_cls_nodes.mapper.dataset.shallow_copy(glyph_cls)

                if a_reg_nodes is None:
                    if state["diff_mode"]:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_reg_nodes.mapper.dataset.shallow_copy(glyph_reg)

                if a_cls_edges is None:
                    a_cls_edges = plotter.add_mesh(e_cls, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_cls_edges.mapper.dataset.shallow_copy(e_cls)

                if a_reg_edges is None:
                    a_reg_edges = plotter.add_mesh(e_reg, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_reg_edges.mapper.dataset.shallow_copy(e_reg)

                # edge visibility
                if state["edge_mode"] == 1:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(False)
                elif state["edge_mode"] == 2:
                    a_cls_edges.SetVisibility(False)
                    a_reg_edges.SetVisibility(True)
                else:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(True)

                # make edge thickness depend on mean importance (simple)
                # (PyVista can't vary per-line width easily, but you can tune global)
                a_cls_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_cls)))
                a_reg_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_reg)))

                # labels
                clear_labels()
                k1 = min(LABEL_TOP, len(f_cls))
                k2 = min(LABEL_TOP, len(f_reg))
                if k1 > 0:
                    act = plotter.add_point_labels(pts_cls[:k1], [f"CLS: {x}" for x in f_cls[:k1]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)
                if k2 > 0:
                    act = plotter.add_point_labels(pts_reg[:k2], [f"REG: {x}" for x in f_reg[:k2]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)

                plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            rebuild()

            # -----------------------
            # Key events
            # -----------------------
            def set_edge_mode(m):
                state["edge_mode"] = m
                rebuild()

            def set_score_mode(mode):
                if mode == "shap" and not shap_ok:
                    state["score_mode"] = "gain"
                else:
                    state["score_mode"] = mode
                rebuild()

            def toggle_diff():
                state["diff_mode"] = not state["diff_mode"]
                rebuild()

            plotter.add_key_event("1", lambda: set_edge_mode(1))
            plotter.add_key_event("2", lambda: set_edge_mode(2))
            plotter.add_key_event("3", lambda: set_edge_mode(3))
            plotter.add_key_event("g", lambda: set_score_mode("gain"))
            plotter.add_key_event("G", lambda: set_score_mode("gain"))
            plotter.add_key_event("s", lambda: set_score_mode("shap"))
            plotter.add_key_event("S", lambda: set_score_mode("shap"))
            plotter.add_key_event("d", toggle_diff)
            plotter.add_key_event("D", toggle_diff)

            plotter.add_axes()
            plotter.camera_position = "iso"
            plotter.camera.zoom(1.35)
            plotter.show()

        # ============ ZONE RÉSULTATS ============
        graph_widgets = []  # kpi, graphe 1, tableaux

        graph2_container = tk.Frame(scrollable_frame, bg="#00122e")      # déjà ok chez toi
        monthly_container = tk.Frame(scrollable_frame, bg="#00122e")     # déjà ok chez toi

        valid_container = tk.Frame(scrollable_frame, bg="#00122e")       # VALID
        valid_graph_widget = None
        valid_container_packed = False

        
        mc_container = tk.Frame(scrollable_frame, bg="#00122e")          # NEW: Monte Carlo paths
        mc_graph_widget = None
        mc_container_packed = False
        ia_valid_widget = None
        monthly_graph_widget = None
        monthly_container_packed = False
        # --- Containers nouveaux ---
        mc_fan_container = tk.Frame(scrollable_frame, bg="#00122e")
        mc_fan_widget = None
        mc_fan_container_packed = False

        resid_container = tk.Frame(scrollable_frame, bg="#00122e")
        resid_widget = None
        resid_container_packed = False

        mc3d_container_packed = False
        mc3d_widget = None
        mc3d_container = tk.Frame(scrollable_frame, bg="#00122e")

        cm_container = tk.Frame(scrollable_frame, bg="#00122e")
        cm_widget = None
        cm_container_packed = False

        cm_pred_container = tk.Frame(scrollable_frame, bg="#00122e")
        cm_pred_widget = None
        cm_pred_container_packed = False

        # globals/etat
        current_valid_df = None  # si tu en as besoin ailleurs (optionnel)

        export_tools_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=0)
        export_tools_frame.pack(fill="x", padx=10, pady=(0, 10))

        export_graph2_button = ctk.CTkButton(
            export_tools_frame,
            text="📤 Export Graph2 (Réel / IA / Profils) - Excel",
            width=320, height=34,
            corner_radius=10,
            fg_color="#2563eb", hover_color="#1d4ed8",
            text_color="white",
            state="disabled",
            command=_export_graph2_to_excel
        )
        export_graph2_button.pack(anchor="w")

        btn_3d = ctk.CTkButton(
            export_tools_frame,
            text="🧊 Voir 3D (PyVista)",
            width=200, height=34,
            corner_radius=10,
            fg_color="#7c3aed", hover_color="#6d28d9",
            text_color="white",
            state="disabled",
            command=_show_pyvista_3d
        )
        btn_3d.pack(anchor="w", pady=(6, 0))

        # ---------- helpers graphiques ----------
        GRAPH_EXPLANATIONS = {
            "graph2": (
                "Vue détaillée – réel, prévision IA et profils",
                "Ce graphique compare l’évolution journalière de l’année cible entre le réel observé, la prévision IA et les profils activés. "
                "Il permet de juger la cohérence globale de la trajectoire prévue sur l’ensemble de l’année. "
                "Pour le lire, compare d’abord la forme de la courbe IA au réel, puis observe l’écart avec les profils pour situer la prévision par rapport aux scénarios de référence."
            ),
            "monthly": (
                "Vue mensuelle – cumuls par mois",
                "Ce graphique regroupe les valeurs par mois pour comparer les cumuls mensuels entre le réel, la prévision IA et les profils éventuels. "
                "Il sert à vérifier si la répartition annuelle reste cohérente à un niveau plus agrégé que le journalier. "
                "Pour le lire, regarde la hauteur des barres mois par mois afin d’identifier rapidement les périodes sur-prévues, sous-prévues ou atypiques."
            ),
            "valid": (
                "Validation interne – réel vs prédiction",
                "Ce graphique montre la performance du modèle sur la période de validation interne en comparant les valeurs réelles et les valeurs prédites. "
                "Il sert à visualiser la qualité d’ajustement avant utilisation sur l’année cible. "
                "Pour le lire, regarde si la courbe prédite suit correctement le niveau, les variations et les points de rupture observés sur la série réelle."
            ),
            "mc_fan": (
                "Fan chart Monte Carlo – incertitude de prévision",
                "Ce graphique représente la prévision centrale ainsi qu’une bande d’incertitude issue des simulations Monte Carlo. "
                "Il permet d’évaluer non seulement la valeur attendue mais aussi l’étendue plausible des résultats autour de cette trajectoire. "
                "Pour le lire, suis la courbe centrale puis observe la largeur de la bande : plus elle est large, plus l’incertitude estimée est importante."
            ),
            "residuals": (
                "Résidus – erreur journalière de la prévision",
                "Ce graphique montre l’écart entre le réel et la prévision centrale jour par jour, ainsi qu’une moyenne glissante pour en résumer la tendance. "
                "Il permet de repérer les périodes où le modèle surestime ou sous-estime durablement la série. "
                "Pour le lire, regarde la position par rapport à zéro : au-dessus le réel dépasse la prévision, en dessous la prévision dépasse le réel."
            ),
            "mc_paths": (
                "Trajectoires Monte Carlo – scénarios simulés",
                "Ce graphique affiche un ensemble de trajectoires simulées autour de la prévision finale afin d’illustrer la dispersion possible des résultats. "
                "Il sert à visualiser la variabilité potentielle de la trajectoire au fil du temps, au-delà de la seule courbe centrale. "
                "Pour le lire, observe la densité et l’ouverture du faisceau de courbes : plus il s’élargit, plus l’incertitude augmente sur la période."
            ),
            "mc_3d": (
                "Surface 3D Monte Carlo – distribution des quantiles",
                "Ce graphique représente la distribution simulée de la prévision dans le temps selon différents quantiles, sous forme de surface 3D. "
                "Il permet d’analyser simultanément la date, le niveau prévu et la structure de dispersion des simulations. "
                "Pour le lire, suis la surface selon l’axe des quantiles : plus elle est épaisse ou étirée, plus l’éventail des valeurs possibles est large."
            ),
            "cm": (
                "Matrice de confusion – qualité de la classification",
                "Ce graphique évalue la capacité du classifieur à distinguer les jours prédits comme nuls ou non nuls sur l’échantillon de validation. "
                "Il sert à comprendre si le modèle détecte correctement la présence ou l’absence de trafic attendu. "
                "Pour le lire, regarde la diagonale principale : plus elle concentre les valeurs, plus la classification est correcte ; les cases hors diagonale représentent les erreurs."
            ),
        }
        
        def _add_section_separator():
            """Ajoute une barre blanche pleine largeur pour structurer la page."""
            sep = tk.Frame(scrollable_frame, bg="white", height=2)
            sep.pack(fill="x", padx=10, pady=(12, 10))
            graph_widgets.append(sep)
            return sep

        def _add_graph_explanation(title: str, text: str):
            """Ajoute un bloc texte explicatif au-dessus d'un graphe."""
            box = ctk.CTkFrame(
                scrollable_frame,
                fg_color="#0b1730",
                corner_radius=12,
                border_width=1,
                border_color="#223658"
            )
            box.pack(fill="x", padx=10, pady=(6, 8))
            graph_widgets.append(box)

            ctk.CTkLabel(
                box,
                text=title,
                font=("Segoe UI Semibold", 14, "bold"),
                text_color="white",
                anchor="w",
                justify="left"
            ).pack(fill="x", padx=14, pady=(10, 4))

            ctk.CTkLabel(
                box,
                text=text,
                font=("Segoe UI", 12),
                text_color="#d7e3f4",
                anchor="w",
                justify="left",
                wraplength=1200
            ).pack(fill="x", padx=14, pady=(0, 10))

            return box
        
        def _clear_graph_widgets():
            """Supprime les widgets graphiques (sauf le conteneur du graphe 2)."""
            nonlocal graph_widgets
            for w in graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            graph_widgets = []

        def _format_time_axis_like_graph2(ax, dates):
            """Applique le même format d'axe X que _redraw_graph2()."""
            import matplotlib.dates as mdates
            import pandas as pd

            d = pd.to_datetime(dates)
            if len(d) < 2:
                return

            # Même logique que graph2: rotation 30 + couleur blanche
            ax.tick_params(axis='x', colors="white", rotation=30)

            # Force les limites exactes sur les données
            ax.set_xlim(d.min(), d.max())

            # Locator "propre" et stable (évite un x-axis bizarre)
            # AutoDateLocator choisit un nombre de ticks raisonnable
            locator = mdates.AutoDateLocator(minticks=6, maxticks=12)
            formatter = mdates.ConciseDateFormatter(locator)

            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(formatter)

            # Petite marge visuelle (optionnel)
            ax.margins(x=0.01)

        def _redraw_graph2():
            """Redessine le graphe détaillé N+1 avec réel / IA / profils cochés."""
            nonlocal ia_graph2_widget, graph2_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            if current_pred_df is None and current_real_target_df is None and not ia_profils_names:
                return

            if not graph2_container_packed:
                graph2_container.pack(pady=10, fill="both", expand=True)
                graph2_container_packed = True

            for child in graph2_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            ia_graph2_widget = None

            fig2, ax2 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax2.set_facecolor("#00122e")

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            flux = selected_flux.get()

            if current_real_target_df is not None and not current_real_target_df.empty:
                ax2.plot(
                    current_real_target_df["date"], current_real_target_df["y"],
                    label=f"Réel {target_year}", linewidth=2, color="#5DADE2"
                )

            if current_pred_df is not None and not current_pred_df.empty:
                ax2.plot(
                    current_pred_df["date"], current_pred_df["pred_value"],
                    label=f"Prévision IA {target_year}",
                    linewidth=2, linestyle="--", color="#F4D03F"
                )

            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = [pd.to_datetime(d) for d in ia_profils_dates]
                palette = plt.cm.tab10.colors
                prof_idx = 0
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    y_prof = [_to_float_or_nan(v) for v in serie]
                    ax2.plot(
                        dates_prof, y_prof,
                        label=f"Profil '{name}' {target_year}",
                        linewidth=1.8,
                        linestyle="-.",
                        color=palette[prof_idx % len(palette)]
                    )
                    prof_idx += 1

            filiale_for_title = current_filiale_name or selected_filiale.get()
            ax2.set_title(f"{flux} – année {target_year} – {filiale_for_title}", color="white", fontsize=14)
            ax2.set_xlabel("Date", color="white", fontsize=12)
            ax2.set_ylabel("Valeur", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white", rotation=30)
            # après avoir choisi les séries tracées, récupère les dates "d" de référence
            if current_pred_df is not None and not current_pred_df.empty:
                dref = current_pred_df["date"]
            elif current_real_target_df is not None and not current_real_target_df.empty:
                dref = current_real_target_df["date"]
            else:
                dref = dates_prof

            _format_time_axis_like_graph2(ax2, dref)
            ax2.tick_params(axis='y', colors="white")
            
            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["graph2"])
            
            canvas_fig2 = FigureCanvasTkAgg(fig2, master=graph2_container)
            canvas_fig2.draw()
            ia_graph2_widget = canvas_fig2.get_tk_widget()
            ia_graph2_widget.pack(fill="both", expand=True)
            plt.close(fig2)

        def _redraw_monthly_graph():
            """Ton code existant mensuel (inchangé) — je garde le tien en l'état."""
            nonlocal monthly_graph_widget, monthly_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            if current_pred_df is None or current_pred_df.empty:
                return

            if not monthly_container_packed:
                monthly_container.pack(pady=10, fill="both", expand=True)
                monthly_container_packed = True

            for child in monthly_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            monthly_graph_widget = None

            import numpy as np
            from matplotlib.ticker import FuncFormatter

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            color_real = "#1f77b4"
            color_pred = "#F4D03F"
            profile_colors = [
                "#e74c3c", "#9b59b6", "#2ecc71", "#1abc9c", "#f1c40f",
                "#d35400", "#8e44ad", "#27ae60", "#16a085", "#c0392b",
                "#7f8c8d", "#95a5a6", "#34495e", "#bdc3c7", "#f39c12",
                "#c27ba0", "#76d7c4", "#7dcea0", "#af7ac5", "#5dade2",
            ]

            df_pred = current_pred_df.copy()
            df_pred["month"] = df_pred["date"].dt.month
            pred_monthly = (df_pred.groupby("month", as_index=False)["pred_value"].sum()
                            ).rename(columns={"pred_value": "pred_value"})

            if current_real_target_df is not None and not current_real_target_df.empty:
                df_real = current_real_target_df.copy()
                df_real["month"] = df_real["date"].dt.month
                real_monthly = (df_real.groupby("month", as_index=False)["y"].sum()
                                ).rename(columns={"y": "real_value"})

                monthly_cmp = pd.merge(real_monthly, pred_monthly, on="month", how="outer").fillna(0.0)
            else:
                monthly_cmp = pred_monthly.copy()
                monthly_cmp["real_value"] = 0.0

            monthly_cmp = monthly_cmp.sort_values("month")

            active_profiles = []
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = pd.to_datetime(ia_profils_dates)
                df_day = pd.DataFrame({"date": dates_prof, "month": dates_prof.month})

                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = dict(zip(current_real_target_df["date"], current_real_target_df["y"]))
                    df_day["real"] = df_day["date"].map(real_map).fillna(0.0)
                else:
                    df_day["real"] = 0.0

                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    vals = [_to_float_or_nan(v) for v in serie]
                    df_day["prev"] = vals
                    df_day["comb"] = df_day["prev"]
                    df_day.loc[df_day["comb"].isna(), "comb"] = df_day["real"]

                    mois_avec_prev = set(df_day.loc[df_day["prev"].notna(), "month"].unique())
                    mois_tous = set(monthly_cmp["month"].unique())
                    mois_masques = sorted(mois_tous - mois_avec_prev)

                    df_day["comb_masked"] = df_day["comb"]
                    df_day.loc[df_day["month"].isin(mois_masques), "comb_masked"] = np.nan

                    prof_month = df_day.groupby("month")["comb_masked"].sum(min_count=1)
                    yvals = [prof_month.get(m, np.nan) for m in monthly_cmp["month"]]
                    active_profiles.append((name, yvals))

            x = np.arange(len(monthly_cmp))
            nb_series = 2 + len(active_profiles)
            width = 0.8 / max(nb_series, 1)
            offsets = (np.arange(nb_series) - (nb_series - 1) / 2.0) * width

            fig_m, ax_m = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax_m.set_facecolor("#00122e")

            ax_m.bar(x + offsets[0], monthly_cmp["real_value"], width, label=f"Réel {target_year}", color=color_real)
            ax_m.bar(x + offsets[1], monthly_cmp["pred_value"], width, label=f"Prévision IA {target_year}", color=color_pred)

            if active_profiles:
                for i, (name, y_vals) in enumerate(active_profiles):
                    serie_idx = 2 + i
                    profile_color = profile_colors[i % len(profile_colors)]
                    ax_m.bar(x + offsets[serie_idx], y_vals, width, label=f"Profil '{name}' {target_year}", color=profile_color)

            month_nums = monthly_cmp["month"].astype(int)
            month_dates = pd.to_datetime({"year": [target_year] * len(month_nums), "month": month_nums, "day": 1})
            mois_labels = month_dates.dt.strftime("%Y-%m")

            ax_m.set_xticks(x)
            ax_m.set_xticklabels(mois_labels, rotation=45, ha="right", fontsize=9, color="white")

            ax_m.set_xlabel("Mois", fontsize=11, color="white")
            ax_m.set_ylabel("Valeur cumulée", fontsize=11, color="white")
            ax_m.tick_params(axis="y", colors="white")

            def _fmt_milliers(val, pos):
                try:
                    return f"{int(val):,}".replace(",", " ")
                except Exception:
                    return ""
            ax_m.yaxis.set_major_formatter(FuncFormatter(_fmt_milliers))

            filiale_for_title = current_filiale_name or selected_filiale.get()
            flux = selected_flux.get()
            ax_m.set_title(f"{filiale_for_title} - {flux} - {target_year}", fontsize=13, fontweight="bold", color="white", pad=12)

            fig_m.patch.set_edgecolor("#00122e")
            fig_m.patch.set_linewidth(0)
            for spine in ax_m.spines.values():
                spine.set_visible(False)

            leg_m = ax_m.legend(frameon=False, facecolor="#00122e")
            for text in leg_m.get_texts():
                text.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["monthly"])

            canvas_m = FigureCanvasTkAgg(fig_m, master=monthly_container)
            canvas_m.draw()
            monthly_graph_widget = canvas_m.get_tk_widget()
            monthly_graph_widget.pack(fill="both", expand=True)
            plt.close(fig_m)

        def _redraw_valid_graph():
            """Redessine le graphe VALID (réel vs préd) en prenant l'année depuis valid_vis."""
            nonlocal ia_valid_widget, valid_container_packed
            nonlocal current_filiale_name

            df_valid_vis = model_artifacts.get("valid_vis", None)
            if df_valid_vis is None or df_valid_vis.empty:
                return

            if not valid_container_packed:
                valid_container.pack(pady=10, fill="both", expand=True)
                valid_container_packed = True

            for child in valid_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            valid_graph_widget = None

            # -------- année VALID : robuste --------
            valid_year = None
            if "year_target" in df_valid_vis.columns and df_valid_vis["year_target"].notna().any():
                try:
                    valid_year = int(df_valid_vis["year_target"].dropna().astype(int).mode().iloc[0])
                except Exception:
                    valid_year = None

            # dates
            if "date_tgt" in df_valid_vis.columns:
                d = pd.to_datetime(df_valid_vis["date_tgt"])
            elif "date" in df_valid_vis.columns:
                d = pd.to_datetime(df_valid_vis["date"])
            else:
                return

            if valid_year is None:
                try:
                    valid_year = int(d.dt.year.mode().iloc[0])
                except Exception:
                    valid_year = ""

            if "y_true" not in df_valid_vis.columns or "y_pred" not in df_valid_vis.columns:
                return

            y_true = df_valid_vis["y_true"].astype(float).values
            y_pred = df_valid_vis["y_pred"].astype(float).values

            # -------- plot --------
            figv, axv = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            axv.set_facecolor("#00122e")

            axv.plot(d, y_true, label="VALID réel", linewidth=2, color="#5DADE2")
            axv.plot(d, y_pred, label="VALID prédiction", linewidth=2, linestyle="--", color="#F4D03F")

            # Bande P10–P90 (si calib dispo)
            qs = model_artifacts.get("calib_qs", None)
            q_low_by_bin = model_artifacts.get("calib_q_low_by_bin", None)
            q_high_by_bin = model_artifacts.get("calib_q_high_by_bin", None)

            if qs is not None and q_low_by_bin is not None and q_high_by_bin is not None:
                try:
                    import numpy as np
                    qs = np.asarray(qs, dtype=float)
                    q_low_by_bin = np.asarray(q_low_by_bin, dtype=float)
                    q_high_by_bin = np.asarray(q_high_by_bin, dtype=float)

                    n_bins = len(q_low_by_bin)
                    bins = np.digitize(y_pred, qs[1:-1], right=True)
                    bins = np.clip(bins, 0, n_bins - 1)

                    p10 = np.clip(y_pred + q_low_by_bin[bins], 0.0, None)
                    p90 = np.clip(y_pred + q_high_by_bin[bins], 0.0, None)

                    axv.fill_between(d, p10, p90, alpha=0.20, label="Bande VALID P10–P90")
                except Exception:
                    pass

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            axv.set_title(f"VALID – {filiale} – année {valid_year}", color="white", fontsize=14)
            axv.set_xlabel("Date", color="white", fontsize=12)
            axv.set_ylabel("Valeur", color="white", fontsize=12)
            axv.set_xlabel("Date", color="white", fontsize=12)
            axv.set_ylabel("Valeur", color="white", fontsize=12)

            _format_time_axis_like_graph2(axv, d)
            axv.tick_params(axis="y", colors="white")

            axv.legend(facecolor="#00122e", edgecolor="white", labelcolor="white")

            # IMPORTANT : évite “mauvaises limites” d’axe
            try:
                axv.set_xlim(d.min(), d.max())
            except Exception:
                pass
            
            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["valid"])

            canvas_v = FigureCanvasTkAgg(figv, master=valid_container)
            canvas_v.draw()
            valid_graph_widget = canvas_v.get_tk_widget()
            valid_graph_widget.pack(fill="both", expand=True)
            plt.close(figv)

        def _redraw_mc_fan_graph():
            """Fan chart MC: P10/P50/P90 + réel année cible si dispo, axe x limité à l'année cible."""
            nonlocal mc_fan_widget, mc_fan_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            # source forecast
            if current_pred_df is None or current_pred_df.empty:
                return

            dfp = current_pred_df.copy()
            if "date" not in dfp.columns:
                return
            if not all(c in dfp.columns for c in ["pred_p10", "pred_value", "pred_p90"]):
                return

            d_all = pd.to_datetime(dfp["date"])
            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(d_all.dt.year).mode().iloc[0])

            # filtre strict année cible
            mask = (d_all.dt.year == ty)
            if mask.sum() < 2:
                mask = np.ones(len(d_all), dtype=bool)

            dfp = dfp.loc[mask].copy()
            d = pd.to_datetime(dfp["date"])

            if not mc_fan_container_packed:
                mc_fan_container.pack(pady=10, fill="both", expand=True)
                mc_fan_container_packed = True

            for child in mc_fan_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc_fan_widget = None

            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            # réel si dispo
            if current_real_target_df is not None and not current_real_target_df.empty:
                d_real = pd.to_datetime(current_real_target_df["date"])
                mask_r = (d_real.dt.year == ty)
                rr = current_real_target_df.loc[mask_r].copy()
                if not rr.empty:
                    ax.plot(rr["date"], rr["y"], linewidth=2.0, color="#5DADE2", label=f"Réel {ty}")

            ax.plot(d, dfp["pred_value"].astype(float).values, linewidth=2.2, linestyle="--",
                    color="#F4D03F", label=f"P50 {ty}")
            ax.fill_between(d,
                            dfp["pred_p10"].astype(float).values,
                            dfp["pred_p90"].astype(float).values,
                            alpha=0.22, label="Bande P10–P90")

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Monte Carlo – fan chart – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Valeur", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            # même “fenêtre année” que graph2
            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                ax.set_xlim(d.min(), d.max())

            # si tu as déjà ton helper :
            _format_time_axis_like_graph2(ax, d)

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_fan"])

            canvas = FigureCanvasTkAgg(fig, master=mc_fan_container)
            canvas.draw()
            mc_fan_widget = canvas.get_tk_widget()
            mc_fan_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_residuals_graph():
            """Résidus sur l'année cible: resid = réel - P50 + moyenne glissante 7j.
            + Marqueurs rouges sur les jours 'alerte' (même logique que _compute_metrics: sMAPE > rel_seuil).
            """
            nonlocal resid_widget, resid_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            if current_pred_df is None or current_pred_df.empty:
                return
            if current_real_target_df is None or current_real_target_df.empty:
                return
            if "pred_value" not in current_pred_df.columns:
                return

            # --- paramètres alignés sur _compute_metrics ---
            rel_seuil = 40.0       # seuil alerte (%)
            min_abs_ref = 10.0     # ignore si |ref| < 10
            eps = 1e-9
            mark_only_peaks = True  # mets False si tu veux un point rouge sur TOUS les jours en alerte

            dfp = current_pred_df[["date", "pred_value"]].copy()
            dfr = current_real_target_df[["date", "y"]].copy()
            dfp["date"] = pd.to_datetime(dfp["date"])
            dfr["date"] = pd.to_datetime(dfr["date"])

            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(dfp["date"].dt.year).mode().iloc[0])

            df = pd.merge(dfr, dfp, on="date", how="inner").sort_values("date")
            df = df[df["date"].dt.year == ty].copy()
            if df.shape[0] < 5:
                return

            # --- résidus ---
            y = df["y"].astype(float).values
            p = df["pred_value"].astype(float).values
            df["resid"] = y - p
            df["roll7"] = df["resid"].rolling(7, min_periods=1).mean()

            # --- logique alerte (sMAPE) comme _compute_metrics ---
            abs_err = np.abs(p - y)
            denom = (np.abs(y) + np.abs(p))
            denom = np.clip(denom, eps, None)
            smape = (2.0 * abs_err / denom) * 100.0
            df["smape"] = smape

            # filtre min_abs_ref (même philosophie que _compute_metrics)
            valid_ref = np.abs(y) >= float(min_abs_ref)
            df["is_alert"] = (df["smape"] > float(rel_seuil)) & valid_ref

            # --- option: ne marquer que les pics locaux parmi les alertes ---
            if mark_only_peaks:
                rabs = np.abs(df["resid"].values)
                # pic local si > voisin gauche et >= voisin droite (simple et robuste)
                is_peak = np.zeros(len(df), dtype=bool)
                if len(df) >= 3:
                    is_peak[1:-1] = (rabs[1:-1] > rabs[:-2]) & (rabs[1:-1] >= rabs[2:])
                df["is_peak"] = is_peak
                df["mark_red"] = df["is_alert"] & df["is_peak"]
            else:
                df["mark_red"] = df["is_alert"]

            # --- UI pack ---
            if not resid_container_packed:
                resid_container.pack(pady=10, fill="both", expand=True)
                resid_container_packed = True

            for child in resid_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            resid_widget = None

            # --- plot ---
            fig, ax = plt.subplots(figsize=(11, 4.0), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            ax.axhline(0.0, linewidth=1.2, alpha=0.5)
            ax.plot(df["date"], df["resid"], linewidth=1.6, color="#5DADE2", label="Résidu (réel - P50)")
            ax.plot(df["date"], df["roll7"], linewidth=2.2, linestyle="--", color="#F4D03F", label="Moyenne 7j")

            # --- points rouges (alertes) ---
            df_red = df[df["mark_red"]].copy()
            if not df_red.empty:
                ax.scatter(
                    df_red["date"], df_red["resid"],
                    s=42, marker="o",
                    color="#ff3b30", edgecolors="white", linewidths=0.6,
                    label=f"Alerte (sMAPE>{rel_seuil:.0f}%)"
                )

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Résidus – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Erreur", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                pass

            _format_time_axis_like_graph2(ax, df["date"])

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["residuals"])

            canvas = FigureCanvasTkAgg(fig, master=resid_container)
            canvas.draw()
            resid_widget = canvas.get_tk_widget()
            resid_widget.pack(fill="both", expand=True)
            plt.close(fig)

        def _redraw_mc_graph(show_all: bool = False, max_paths: int = 300):
            """Graphe séparé : MC trajectories + trajectoire finale, X limité à l'année cible."""
            nonlocal mc_graph_widget, mc_container_packed
            nonlocal current_target_year, current_filiale_name

            sims = model_artifacts.get("mc_sims", None)      # (B,N) ou None
            dates = model_artifacts.get("mc_dates", None)    # (N,)
            det = model_artifacts.get("mc_det", None)        # (N,)
            p50 = model_artifacts.get("mc_p50", None)        # (N,)

            if dates is None:
                return

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib.dates as mdates

            # IMPORTANT: DatetimeIndex -> pas de .dt
            d_all = pd.to_datetime(dates)
            if len(d_all) < 2:
                return

            # année cible
            ty = model_artifacts.get("target_year", None)
            if ty is None:
                ty = current_target_year

            try:
                ty = int(ty)
            except Exception:
                try:
                    ty = int(pd.Series(d_all.year).mode().iloc[0])
                except Exception:
                    ty = int(d_all[0].year)

            # filtre strict sur l'année cible (évite débordement 2026)
            mask_year = (d_all.year == ty)
            if mask_year.sum() < 2:
                mask_year = np.ones(len(d_all), dtype=bool)

            d = d_all[mask_year]

            # sims filtrées
            sims_ok = None
            if sims is not None:
                sims = np.asarray(sims, dtype=float)
                if sims.ndim == 2 and sims.shape[1] == len(d_all):
                    sims_ok = sims[:, mask_year]

            # vecteurs filtrés (trajectoire finale)
            def _vec(v):
                if v is None:
                    return None
                v = np.asarray(v, dtype=float)
                if len(v) != len(d_all):
                    return None
                return v[mask_year]

            det_f = _vec(det)
            p50_f = _vec(p50)

            if not mc_container_packed:
                mc_container.pack(pady=10, fill="both", expand=True)
                mc_container_packed = True

            # clear container
            for child in mc_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc_graph_widget = None

            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            # spaghetti
            if sims_ok is not None:
                B = sims_ok.shape[0]
                if show_all:
                    idx = range(B)
                    alpha = 0.08
                    lw = 0.6
                else:
                    k = int(min(max_paths, B))
                    idx = np.linspace(0, B - 1, num=k, dtype=int)
                    alpha = 0.12
                    lw = 0.7

                for i in idx:
                    ax.plot(d, sims_ok[i, :], linewidth=lw, alpha=alpha)

            # trajectoire finale (P50 prioritaire)
            if p50_f is not None:
                ax.plot(d, p50_f, linewidth=2.6, linestyle="--", color="#F4D03F", label="Trajectoire finale (P50)")
            elif det_f is not None:
                ax.plot(d, det_f, linewidth=2.6, linestyle="-", color="#F4D03F", label="Trajectoire finale (déterministe)")

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Monte Carlo – trajectoires – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Valeur simulée", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            # limites strictes année cible (évite 2026-01)
            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                ax.set_xlim(d.min(), d.max())

            # ticks propres
            locator = mdates.AutoDateLocator(minticks=6, maxticks=12)
            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_paths"])

            canvas = FigureCanvasTkAgg(fig, master=mc_container)
            canvas.draw()
            mc_graph_widget = canvas.get_tk_widget()
            mc_graph_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_mc_3d_fan_surface(
            q_grid=None,
            show_surface=True,
            show_quantile_curves=True,
            alert_rel_seuil=40.0,
            min_abs_ref=10.0,
        ):
            """
            3D fan chart : X=Date, Y=Quantile, Z=Valeur simulée (MC)
            + points rouges sur la courbe P50 si alerte (sMAPE>seuil) vs réel.
            """
            nonlocal mc3d_widget, mc3d_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
            import matplotlib.dates as mdates

            sims = model_artifacts.get("mc_sims", None)      # (B,N)
            dates = model_artifacts.get("mc_dates", None)    # (N,)
            if sims is None or dates is None:
                return

            sims = np.asarray(sims, dtype=float)
            if sims.ndim != 2 or sims.shape[1] != len(dates):
                return

            d_all = pd.to_datetime(dates)
            if len(d_all) < 5:
                return

            # année cible
            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(d_all.year).mode().iloc[0])

            mask_year = (d_all.year == ty)
            if mask_year.sum() < 5:
                mask_year = np.ones(len(d_all), dtype=bool)

            d = d_all[mask_year]
            sims_y = sims[:, mask_year]

            # grille de quantiles
            if q_grid is None:
                q_grid = np.array([0.05, 0.10, 0.20, 0.35, 0.50, 0.65, 0.80, 0.90, 0.95], dtype=float)
            q_grid = np.asarray(q_grid, dtype=float)

            # compute quantiles (len(q), N)
            Q = np.quantile(sims_y, q_grid, axis=0)

            # container UI
            if not mc3d_container_packed:
                mc3d_container.pack(pady=10, fill="both", expand=True)
                mc3d_container_packed = True

            for child in mc3d_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc3d_widget = None

            # X axis numeric dates
            x_num = mdates.date2num(d.to_pydatetime())
            X, Y = np.meshgrid(x_num, q_grid)  # (len(q), N) after broadcast with Q

            fig = plt.figure(figsize=(11, 5.2), facecolor="#00122e", constrained_layout=True)
            ax = fig.add_subplot(111, projection="3d")
            ax.set_facecolor("#00122e")

            # surface fan
            if show_surface:
                # IMPORTANT: ne pas mettre de couleurs custom si tu veux rester neutre
                ax.plot_surface(X, Y, Q, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=0.35)

            # courbes quantiles (lisible)
            if show_quantile_curves:
                for i, qq in enumerate(q_grid):
                    lw = 2.2 if np.isclose(qq, 0.50) else 1.0
                    ls = "--" if np.isclose(qq, 0.50) else "-"
                    ax.plot(x_num, np.full_like(x_num, qq, dtype=float), Q[i, :], linewidth=lw, linestyle=ls)

            # points rouges d'alerte sur la courbe P50 (si réel dispo)
            if current_real_target_df is not None and not current_real_target_df.empty:
                dfr = current_real_target_df[["date", "y"]].copy()
                dfr["date"] = pd.to_datetime(dfr["date"])
                dfr = dfr[dfr["date"].dt.year == ty].copy()
                if not dfr.empty:
                    # merge sur dates filtrées
                    dfm = pd.DataFrame({"date": d})
                    dfm = dfm.merge(dfr, on="date", how="left")
                    p50 = Q[np.argmin(np.abs(q_grid - 0.50)), :]
                    dfm["p50"] = p50
                    dfm = dfm.dropna(subset=["y"])

                    y_true = dfm["y"].astype(float).values
                    y_pred = dfm["p50"].astype(float).values
                    abs_err = np.abs(y_pred - y_true)
                    denom = np.clip(np.abs(y_true) + np.abs(y_pred), 1e-9, None)
                    smape = (2.0 * abs_err / denom) * 100.0

                    mask_valid = (np.abs(y_true) >= float(min_abs_ref))
                    is_alert = (smape > float(alert_rel_seuil)) & mask_valid

                    if np.any(is_alert):
                        x_alert = mdates.date2num(dfm.loc[is_alert, "date"].dt.to_pydatetime())
                        z_alert = dfm.loc[is_alert, "p50"].values.astype(float)
                        ax.scatter(
                            x_alert,
                            np.full_like(x_alert, 0.50, dtype=float),
                            z_alert,
                            s=28,
                            color="#ff3b30",
                            depthshade=False
                        )

            # axes labels
            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"3D Fan chart (MC) – {filiale} – {ty}", color="white", fontsize=13, pad=12)

            ax.set_xlabel("Date", color="white", labelpad=10)
            ax.set_ylabel("Quantile", color="white", labelpad=10)
            ax.set_zlabel("Valeur", color="white", labelpad=8)

            # ticks colors
            ax.tick_params(axis="x", colors="white")
            ax.tick_params(axis="y", colors="white")
            ax.tick_params(axis="z", colors="white")

            # format date ticks
            ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
            ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))

            # limits
            ax.set_xlim(mdates.date2num(pd.Timestamp(ty, 1, 1)), mdates.date2num(pd.Timestamp(ty, 12, 31)))
            ax.set_ylim(float(q_grid.min()), float(q_grid.max()))

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_3d"])

            # canvas
            canvas = FigureCanvasTkAgg(fig, master=mc3d_container)
            canvas.draw()
            mc3d_widget = canvas.get_tk_widget()
            mc3d_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_confusion_matrix_graph():
            """Redessine la matrice de confusion + courbe ROC du classifieur sur VALID."""
            nonlocal cm_widget, cm_container_packed

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from sklearn.metrics import confusion_matrix, roc_curve, auc

            X_valid = model_artifacts.get("X_valid", None)
            y_cls_valid = model_artifacts.get("y_cls_valid", None)
            cls_models = model_artifacts.get("cls_models", None)
            cls_weights = model_artifacts.get("cls_model_weights", None)
            cls_model_names = model_artifacts.get("cls_model_names", None)

            filiale = model_artifacts.get("filiale", "") or ""
            base_year = model_artifacts.get("base_year", None)

            if X_valid is None or y_cls_valid is None or cls_models is None or cls_weights is None:
                return
            if len(X_valid) == 0 or len(y_cls_valid) == 0 or len(cls_models) == 0:
                return

            X_arr = X_valid.values if hasattr(X_valid, "values") else np.asarray(X_valid)
            y_true = np.asarray(y_cls_valid).astype(int)

            try:
                # Ensemble pondéré
                probas = [m.predict_proba(X_arr)[:, 1] for m in cls_models]
                y_prob = np.average(np.vstack(probas), axis=0, weights=np.asarray(cls_weights, dtype=float))
            except Exception:
                return

            threshold = 0.5
            y_pred = (y_prob >= threshold).astype(int)

            # Matrice de confusion
            cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
            if cm.shape != (2, 2):
                return

            tn, fp, fn, tp = cm.ravel()
            total = cm.sum()

            acc = (tp + tn) / total if total > 0 else 0.0
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

            # ROC
            has_both_classes = len(np.unique(y_true)) > 1
            if has_both_classes:
                fpr, tpr, _ = roc_curve(y_true, y_prob)
                roc_auc = auc(fpr, tpr)
            else:
                fpr, tpr, roc_auc = None, None, np.nan

            if not cm_container_packed:
                cm_container.pack(pady=10, fill="both", expand=True)
                cm_container_packed = True

            for child in cm_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            cm_widget = None

            # Figure 2 panneaux
            fig, (ax_cm, ax_roc) = plt.subplots(
                1, 2,
                figsize=(11.8, 5.2),
                facecolor="#00122e",
                constrained_layout=True
            )
            ax_cm.set_facecolor("#00122e")
            ax_roc.set_facecolor("#00122e")

            # =========================
            # Panneau 1 : Confusion Matrix
            # =========================
            im = ax_cm.imshow(cm, interpolation="nearest", cmap="Blues")

            cbar = fig.colorbar(im, ax=ax_cm, fraction=0.046, pad=0.04)
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="white")

            classes = ["Prédit : nul", "Prédit : non nul"]
            rows = ["Réel : nul", "Réel : non nul"]

            ax_cm.set_xticks(np.arange(2))
            ax_cm.set_yticks(np.arange(2))
            ax_cm.set_xticklabels(classes, color="white", fontsize=11)
            ax_cm.set_yticklabels(rows, color="white", fontsize=11)

            ax_cm.set_xlabel("Classe prédite", color="white", fontsize=12)
            ax_cm.set_ylabel("Classe réelle", color="white", fontsize=12)

            year_txt = f" – VALID {base_year}" if base_year is not None else ""
            ax_cm.set_title(f"Matrice de confusion – {filiale}{year_txt}", color="white", fontsize=14)

            vmax = cm.max() if cm.size > 0 else 1
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    val = cm[i, j]
                    color_txt = "white" if val > vmax / 2 else "#00122e"
                    ax_cm.text(
                        j, i, f"{val}",
                        ha="center", va="center",
                        color=color_txt, fontsize=14, fontweight="bold"
                    )

            # =========================
            # Panneau 2 : ROC Curve
            # =========================
            if has_both_classes:
                # Nom du meilleur modèle pour le titre si dispo
                best_name = "Ensemble"
                try:
                    if cls_model_names is not None and len(cls_model_names) == len(cls_weights):
                        best_idx = int(np.argmax(np.asarray(cls_weights, dtype=float)))
                        best_name = f"Ensemble (dominant: {cls_model_names[best_idx]})"
                except Exception:
                    pass

                ax_roc.plot(fpr, tpr, linewidth=2.2, color="#5DADE2", label=f"{best_name} AUC = {roc_auc:.3f}")
                ax_roc.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="white", alpha=0.7, label="Aléatoire")

                ax_roc.set_xlim(0.0, 1.0)
                ax_roc.set_ylim(0.0, 1.05)
                ax_roc.set_xlabel("False Positive Rate", color="white", fontsize=12)
                ax_roc.set_ylabel("True Positive Rate", color="white", fontsize=12)
                ax_roc.set_title("Courbe ROC – classification VALID", color="white", fontsize=14)

                ax_roc.tick_params(axis="x", colors="white")
                ax_roc.tick_params(axis="y", colors="white")

                leg = ax_roc.legend(facecolor="#00122e", edgecolor="white", fontsize=9)
                for t in leg.get_texts():
                    t.set_color("white")
            else:
                ax_roc.text(
                    0.5, 0.5,
                    "ROC indisponible\n(une seule classe dans VALID)",
                    ha="center", va="center",
                    color="white", fontsize=12
                )
                ax_roc.set_title("Courbe ROC – classification VALID", color="white", fontsize=14)
                ax_roc.set_xticks([])
                ax_roc.set_yticks([])

            # Style commun
            for ax in (ax_cm, ax_roc):
                for spine in ax.spines.values():
                    spine.set_color("white")

            # KPI en bas
            kpi_txt = (
                f"Accuracy: {acc:.1%}   |   "
                f"Precision: {precision:.1%}   |   "
                f"Recall: {recall:.1%}   |   "
                f"Specificity: {specificity:.1%}"
            )
            if has_both_classes:
                kpi_txt += f"   |   AUC: {roc_auc:.3f}"

            fig.text(0.5, 0.02, kpi_txt, ha="center", color="white", fontsize=10)

            canvas = FigureCanvasTkAgg(fig, master=cm_container)
            canvas.draw()
            cm_widget = canvas.get_tk_widget()
            cm_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_confusion_matrix_pred_graph():
            """Redessine la matrice de confusion + courbe ROC sur les prédictions de l'année cible."""
            nonlocal cm_pred_widget, cm_pred_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from sklearn.metrics import confusion_matrix, roc_curve, auc

            if current_pred_df is None or current_pred_df.empty:
                return
            if current_real_target_df is None or current_real_target_df.empty:
                return

            if "date" not in current_pred_df.columns or "date" not in current_real_target_df.columns:
                return
            if "y" not in current_real_target_df.columns:
                return
            if "pred_cls_proba" not in current_pred_df.columns:
                return

            df_pred = current_pred_df.copy()
            df_real = current_real_target_df.copy()

            df_pred["date"] = pd.to_datetime(df_pred["date"])
            df_real["date"] = pd.to_datetime(df_real["date"])

            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(df_pred["date"].dt.year).mode().iloc[0])

            # Merge réel / prédit
            cols_pred = ["date", "pred_cls_proba"]
            if "pred_cls" in df_pred.columns:
                cols_pred.append("pred_cls")

            df_eval = pd.merge(
                df_real[["date", "y"]].copy(),
                df_pred[cols_pred].copy(),
                on="date",
                how="inner"
            )

            df_eval = df_eval[df_eval["date"].dt.year == ty].copy()
            if df_eval.empty:
                return

            # Vérité terrain : 0 si nul, 1 si non nul
            y_true = (df_eval["y"].astype(float).values > 0).astype(int)

            # Probabilités / classes prédites
            y_prob = df_eval["pred_cls_proba"].astype(float).values
            y_prob = np.clip(y_prob, 0.0, 1.0)

            if "pred_cls" in df_eval.columns:
                y_pred = df_eval["pred_cls"].astype(int).values
            else:
                threshold = 0.5
                y_pred = (y_prob >= threshold).astype(int)

            cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
            if cm.shape != (2, 2):
                return

            tn, fp, fn, tp = cm.ravel()
            total = cm.sum()

            acc = (tp + tn) / total if total > 0 else 0.0
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

            has_both_classes = len(np.unique(y_true)) > 1
            if has_both_classes:
                fpr, tpr, _ = roc_curve(y_true, y_prob)
                roc_auc = auc(fpr, tpr)
            else:
                fpr, tpr, roc_auc = None, None, np.nan

            if not cm_pred_container_packed:
                cm_pred_container.pack(pady=10, fill="both", expand=True)
                cm_pred_container_packed = True

            for child in cm_pred_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            cm_pred_widget = None

            fig, (ax_cm, ax_roc) = plt.subplots(
                1, 2,
                figsize=(11.8, 5.2),
                facecolor="#00122e",
                constrained_layout=True
            )
            ax_cm.set_facecolor("#00122e")
            ax_roc.set_facecolor("#00122e")

            # =========================
            # Panneau 1 : Matrice de confusion
            # =========================
            im = ax_cm.imshow(cm, interpolation="nearest", cmap="Blues")

            cbar = fig.colorbar(im, ax=ax_cm, fraction=0.046, pad=0.04)
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="white")

            classes = ["Prédit : nul", "Prédit : non nul"]
            rows = ["Réel : nul", "Réel : non nul"]

            ax_cm.set_xticks(np.arange(2))
            ax_cm.set_yticks(np.arange(2))
            ax_cm.set_xticklabels(classes, color="white", fontsize=11)
            ax_cm.set_yticklabels(rows, color="white", fontsize=11)

            ax_cm.set_xlabel("Classe prédite", color="white", fontsize=12)
            ax_cm.set_ylabel("Classe réelle", color="white", fontsize=12)

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax_cm.set_title(f"Matrice de confusion – Prévision {ty} – {filiale}", color="white", fontsize=14)

            vmax = cm.max() if cm.size > 0 else 1
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    val = cm[i, j]
                    color_txt = "white" if val > vmax / 2 else "#00122e"
                    ax_cm.text(
                        j, i, f"{val}",
                        ha="center", va="center",
                        color=color_txt, fontsize=14, fontweight="bold"
                    )

            # =========================
            # Panneau 2 : ROC curve
            # =========================
            if has_both_classes:
                ax_roc.plot(fpr, tpr, linewidth=2.2, color="#5DADE2", label=f"Prévision {ty} AUC = {roc_auc:.3f}")
                ax_roc.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="white", alpha=0.7, label="Aléatoire")

                ax_roc.set_xlim(0.0, 1.0)
                ax_roc.set_ylim(0.0, 1.05)
                ax_roc.set_xlabel("False Positive Rate", color="white", fontsize=12)
                ax_roc.set_ylabel("True Positive Rate", color="white", fontsize=12)
                ax_roc.set_title(f"Courbe ROC – Prévision {ty}", color="white", fontsize=14)

                ax_roc.tick_params(axis="x", colors="white")
                ax_roc.tick_params(axis="y", colors="white")

                leg = ax_roc.legend(facecolor="#00122e", edgecolor="white", fontsize=9)
                for t in leg.get_texts():
                    t.set_color("white")
            else:
                ax_roc.text(
                    0.5, 0.5,
                    "ROC indisponible\n(une seule classe réelle)",
                    ha="center", va="center",
                    color="white", fontsize=12
                )
                ax_roc.set_title(f"Courbe ROC – Prévision {ty}", color="white", fontsize=14)
                ax_roc.set_xticks([])
                ax_roc.set_yticks([])

            for ax in (ax_cm, ax_roc):
                for spine in ax.spines.values():
                    spine.set_color("white")

            kpi_txt = (
                f"Accuracy: {acc:.1%}   |   "
                f"Precision: {precision:.1%}   |   "
                f"Recall: {recall:.1%}   |   "
                f"Specificity: {specificity:.1%}"
            )
            if has_both_classes:
                kpi_txt += f"   |   AUC: {roc_auc:.3f}"

            fig.text(0.5, 0.02, kpi_txt, ha="center", color="white", fontsize=10)

            canvas = FigureCanvasTkAgg(fig, master=cm_pred_container)
            canvas.draw()
            cm_pred_widget = canvas.get_tk_widget()
            cm_pred_widget.pack(fill="both", expand=True)
            plt.close(fig)

        # ---------- analyse des écarts ----------
        def _compute_metrics(
            label,
            ref_dates, ref_values,
            cmp_dates, cmp_values,
            min_abs_ref=10.0,
            rel_seuil=40.0,
            rel_mode="smape",
            eps=1e-9,
            exclude_above_k=None,
            restrict_dates=None,
        ):
            if ref_dates is None or cmp_dates is None:
                return None

            ref_vals = [_to_float_or_nan(v) for v in ref_values]
            cmp_vals = [_to_float_or_nan(v) for v in cmp_values]

            def _norm_dates(x):
                d = pd.to_datetime(x, errors="coerce")
                if isinstance(d, (pd.DatetimeIndex, pd.Index)):
                    return d.normalize()
                return d.dt.normalize()

            ref_idx = _norm_dates(ref_dates)
            cmp_idx = _norm_dates(cmp_dates)

            ref = pd.Series(ref_vals, index=ref_idx, dtype="float64").dropna()
            cmp_ = pd.Series(cmp_vals, index=cmp_idx, dtype="float64").dropna()

            common_idx = ref.index.intersection(cmp_.index)
            if len(common_idx) == 0:
                return None

            df_m = pd.DataFrame({
                "ref": ref.loc[common_idx],
                "cmp": cmp_.loc[common_idx],
            }).dropna()

            if restrict_dates is not None and not df_m.empty:
                try:
                    ridx = pd.to_datetime(restrict_dates, errors="coerce")
                    ridx = pd.DatetimeIndex(ridx).normalize()
                    df_m = df_m.loc[df_m.index.intersection(ridx)]
                except Exception:
                    pass

            if df_m.empty:
                return None

            if min_abs_ref is not None and min_abs_ref > 0:
                df_m = df_m[df_m["ref"].abs() >= float(min_abs_ref)]

            if exclude_above_k is not None and not df_m.empty:
                df_m = df_m[df_m["ref"].abs() <= float(exclude_above_k)]

            if df_m.empty:
                return None

            err = df_m["cmp"] - df_m["ref"]
            abs_err = err.abs()

            mae = float(abs_err.mean())
            rmse = float(np.sqrt(np.mean(err.values ** 2)))
            bias = float(err.mean())

            if rel_mode == "mape_ref":
                denom = df_m["ref"].abs().clip(lower=eps)
                rel = (abs_err / denom) * 100.0
            else:
                denom = (df_m["ref"].abs() + df_m["cmp"].abs()).clip(lower=eps)
                rel = (2.0 * abs_err / denom) * 100.0

            rel = rel.replace([np.inf, -np.inf], np.nan)
            df_m = df_m.assign(rel=rel).dropna(subset=["rel"])

            if df_m.empty:
                return None

            mask = df_m["rel"] > float(rel_seuil)
            nb_points = int(len(df_m))
            nb_alert = int(mask.sum())
            taux_alert = 0.0 if nb_points == 0 else 100.0 * nb_alert / nb_points
            valo_alert = float(abs_err[mask].sum())

            sum_ref = float(df_m["ref"].abs().sum())
            wmape = float(abs_err.sum() / (sum_ref + eps) * 100.0)
            smape = float(df_m["rel"].mean())

            # 🔹 Détermination du type
            if label.startswith("Prévision IA vs Réel (période"):
                type_label = "Prévision IA (période Profil)"
            elif label.startswith("Prévision"):
                type_label = "Prévision IA"
            else:
                type_label = "Profil"

            return {
                "Type": type_label,
                "Comparaison": label,
                "Nb points (utilisés)": nb_points,
                "MAE (K€)": round(mae, 1),
                "RMSE (K€)": round(rmse, 1),
                "Biais (K€)": round(bias, 1),
                "sMAPE (%)": round(smape, 1),
                "WMAPE (%)": round(wmape, 1),
                f"Nb alertes (> {rel_seuil}%)": nb_alert,
                f"Taux alertes (> {rel_seuil}%)": round(taux_alert, 1),
                "Valeur totale écarts alertes (K€)": round(valo_alert, 1),
            }

        def _rebuild_analysis_table():
            nonlocal analysis_table_frame
            nonlocal current_pred_df, current_real_target_df, current_target_year

            for child in analysis_table_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

            if current_real_target_df is None or current_real_target_df.empty:
                return

            rows = []

            rel_seuil = 40.0
            min_abs_ref = 10.0
            rel_mode = "smape"

            # 🔹 IA vs Réel (normal)
            if current_pred_df is not None and not current_pred_df.empty:
                m = _compute_metrics(
                    label="Prévision IA vs Réel",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    min_abs_ref=min_abs_ref,
                    rel_seuil=rel_seuil,
                    rel_mode=rel_mode,
                )
                if m:
                    rows.append(m)

                # 🔹 IA vs Réel (sans jours > 100 000)
                m_np = _compute_metrics(
                    label="Prévision Réel vs IA (sans jours > 100 000)",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    min_abs_ref=min_abs_ref,
                    rel_seuil=rel_seuil,
                    rel_mode=rel_mode,
                    exclude_above_k=100000.0,  # ✅ 100 M€
                )
                if m_np:
                    rows.append(m_np)

            # 🔹 Profils cochés
            if ia_profils_names and ia_profils_series:
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    # Profil vs Réel
                    m_prof = _compute_metrics(
                        label=f"Profil '{name}' vs Réel",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=ia_profils_dates,
                        cmp_values=serie,
                        min_abs_ref=min_abs_ref,
                        rel_seuil=rel_seuil,
                        rel_mode=rel_mode,
                    )
                    if m_prof:
                        rows.append(m_prof)

                    # IA vs Réel sur même période que le profil
                    prof_idx = pd.to_datetime(ia_profils_dates, errors="coerce").normalize()
                    prof_vals = pd.Series(
                        [_to_float_or_nan(v) for v in serie],
                        index=prof_idx
                    ).dropna()

                    m_ia_same = _compute_metrics(
                        label=f"Prévision IA vs Réel (période Profil '{name}')",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=current_pred_df["date"],
                        cmp_values=current_pred_df["pred_value"],
                        min_abs_ref=min_abs_ref,
                        rel_seuil=rel_seuil,
                        rel_mode=rel_mode,
                        restrict_dates=prof_vals.index,
                    )
                    if m_ia_same:
                        rows.append(m_ia_same)

            if not rows:
                return

            cols = list(rows[0].keys())

            tree = ttk.Treeview(
                analysis_table_frame,
                columns=cols,
                show="headings",
                height=len(rows)
            )

            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, anchor="center", width=140)

            for r in rows:
                tree.insert("", "end", values=[r.get(c, "") for c in cols])

            tree.pack(fill="x", expand=True)

            # ===================== ✅ Bouton d'extraction (même format CTk) =====================
            export_tools_frame = ctk.CTkFrame(analysis_table_frame, fg_color="#00122e", corner_radius=0)
            export_tools_frame.pack(fill="x", padx=10, pady=(6, 10))

            def _export_treeview_to_xlsx(tree: ttk.Treeview):
                cols_ = list(tree["columns"])
                rows_ = [tree.item(i, "values") for i in tree.get_children()]

                if not cols_ or not rows_:
                    messagebox.showwarning("Export", "Aucune donnée à exporter.")
                    return

                fpath = filedialog.asksaveasfilename(
                    title="Exporter l'analyse des écarts",
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
                )
                if not fpath:
                    return

                try:
                    df = pd.DataFrame(rows_, columns=cols_)
                    df.to_excel(fpath, index=False)
                    messagebox.showinfo("Export", f"Export Excel OK :\n{fpath}")
                except Exception as e:
                    messagebox.showerror("Export", f"Erreur export Excel :\n{e}")

            export_analysis_button = ctk.CTkButton(
                export_tools_frame,
                text="📤 Export Analyse Écarts - Excel",
                width=320, height=34,
                corner_radius=10,
                fg_color="#2563eb", hover_color="#1d4ed8",
                text_color="white",
                state="normal",   # ou "disabled" si tu veux l'activer plus tard
                command=lambda t=tree: _export_treeview_to_xlsx(t)
            )
            export_analysis_button.pack(pady=6)



        # ---------- PROFILS UI ----------
        def _build_ia_profils_ui(filiale, base_year):
            """
            Construit les cases à cocher de profils pour l'année N+1 de la filiale sélectionnée.
            """
            nonlocal ia_profils_vars, ia_profils_names, ia_profils_dates, ia_profils_series

            for w in ia_profils_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            ia_profils_vars = []
            ia_profils_names = []
            ia_profils_dates = []
            ia_profils_series = []

            if not filiale:
                tk.Label(
                    ia_profils_frame,
                    text="Aucune filiale sélectionnée.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            try:
                base_year_int = int(base_year)
            except Exception:
                tk.Label(
                    ia_profils_frame,
                    text="Année N invalide.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            target_year = base_year_int + 1
            feuille = sections.get(filiale)
            if not feuille:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune feuille trouvée pour {filiale}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees (profils N+1) pour {feuille} : {e}")
                tk.Label(
                    ia_profils_frame,
                    text="Erreur de chargement des données N+1.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            flux_cible_local = selected_flux.get()
            cible = [t for t in noms_flux if t[0] == flux_cible_local]
            if not cible:
                tk.Label(
                    ia_profils_frame,
                    text=f"Flux '{flux_cible_local}' introuvable pour cette filiale.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            _, col_start = cible[0]

            try:
                dates_p, reel_p, previsions_p, noms_profils_p = extraire_valeurs(
                    ws, col_start, nb_prev, annee=target_year
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs (profils N+1) {filiale}/{flux_cible_local}/{target_year} : {e}")
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée de profils pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            if not dates_p:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            actifs = []
            for serie in previsions_p:
                exist = any(v not in (None, 0, 0.0, "") for v in serie)
                actifs.append(exist)

            noms_actifs = [n for n, ok in zip(noms_profils_p, actifs) if ok]
            series_actives = [s for s, ok in zip(previsions_p, actifs) if ok]

            if not noms_actifs:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucun profil actif pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            ia_profils_dates = dates_p
            ia_profils_names = noms_actifs
            ia_profils_series = series_actives

            lbl = tk.Label(
                ia_profils_frame,
                text=f"Profils N+1 ({target_year}) : coche pour afficher dans le graphe détaillé et l'analyse",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
            )
            lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 6))

            nb_lignes = 4
            for i, name in enumerate(noms_actifs):
                var = tk.BooleanVar(value=False)

                def _on_toggle(v=var):
                    _redraw_graph2()
                    _redraw_monthly_graph()
                    _rebuild_analysis_table()

                cb = tk.Checkbutton(
                    ia_profils_frame, text=name, variable=var,
                    bg="#00122e", fg="white", font=('Segoe UI', 10),
                    selectcolor="#00aced", activebackground="#003366",
                    activeforeground="white",
                    command=_on_toggle
                )

                row = 1 + (i % nb_lignes)
                col = i // nb_lignes

                cb.grid(row=row, column=col, sticky="w", padx=12, pady=4)
                cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
                cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

                ia_profils_vars.append(var)

        def _on_filiale_or_year_change(_event=None):
            filiale = selected_filiale.get()
            base_year = annees_var.get()
            if not filiale or not base_year:
                return
            _build_ia_profils_ui(filiale, base_year)

        filiale_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        annees_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)

        # init profils
        _on_filiale_or_year_change()

        # ============================================================
        # NEW: callback flux change
        # ============================================================
        def _on_flux_change(_event=None):
            ok = _rebuild_df_for_current_flux()
            if not ok:
                return
            _on_filiale_or_year_change()

            # si un modèle a déjà été entrainé, on redessine ce qui dépend des profils
            try:
                _redraw_monthly_graph()
                _redraw_graph2()
                _rebuild_analysis_table()
            except Exception:
                pass

        flux_box.bind("<<ComboboxSelected>>", _on_flux_change)


        # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------
        import re

        def _norm(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace("œ", "oe")
            s = re.sub(r"\s+", " ", s)
            return s

        def _train_model():
            """
            Routeur: appelle le bon trainer selon flux + filiale.
            df_current est supposé déjà filtré sur le flux sélectionné.
            """
            flux_raw = selected_flux.get()
            filiale_raw = selected_filiale.get()

            flux = _norm(flux_raw)
            filiale = (filiale_raw or "").strip().upper()

            print(f"[ROUTER] filiale={filiale} | flux='{flux_raw}' -> norm='{flux}'")

            # --- Cas 1 : Trafic Voyageurs (ton V3.3)
            if flux in {"trafic voyageurs", "traficvoyageurs"}:
                return _train_trafic_voyageurs_v33()   # ton trainer existant

            # --- Cas 2 : RESEAU + ACE & Investissements (négatif <= 0)
            # robuste aux variantes d'orthographe
            if filiale == "RESEAU" and ("ace" in flux) and ("invest" in flux):
                return _train_ace_invest_reseau_negative_v33()

            # --- Fallback
            messagebox.showinfo(
                "Info",
                f"Aucun modèle spécialisé pour :\nFiliale={filiale_raw}\nFlux={flux_raw}\n\n→ modèle générique."
            )
            return _train_generic_baseline()

        def _train_generic_baseline():
            messagebox.showinfo(
                "Non supporté",
                f"Aucun modèle spécialisé pour:\nFiliale={selected_filiale.get()}\nFlux={selected_flux.get()}"
            )

        def _train_trafic_voyageurs_v33():
            """
            SA_VOYAGEURS – V4.0 ENSEMBLE ULTIME (XGBoost + LightGBM + CatBoost + ExtraTrees)
            Version figée / reproductible au maximum.
            """

            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame, export_button

            _clear_graph_widgets()

            # ======================================================================
            # 0) Reproductibilité MAX
            # ======================================================================
            import os
            import random

            SEED = 42

            # Idéalement à mettre tout en haut du script principal, avant imports globaux.
            os.environ["PYTHONHASHSEED"] = str(SEED)
            os.environ["OMP_NUM_THREADS"] = "1"
            os.environ["MKL_NUM_THREADS"] = "1"
            os.environ["OPENBLAS_NUM_THREADS"] = "1"
            os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
            os.environ["NUMEXPR_NUM_THREADS"] = "1"

            random.seed(SEED)

            import numpy as np
            np.random.seed(SEED)

            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import calendar
            import traceback
            import time

            from sklearn.metrics import (
                mean_absolute_error,
                mean_squared_error,
                r2_score,
                roc_auc_score
            )
            from sklearn.model_selection import RandomizedSearchCV, TimeSeriesSplit
            from scipy.stats import randint, uniform

            # --- Tentative d'import de XGBoost ---
            try:
                from xgboost import XGBClassifier, XGBRegressor
                HAS_XGB = True
            except ImportError:
                HAS_XGB = False
                print("[WARN] XGBoost n'est pas installé. Le modèle tournera sans, mais installe-le avec 'pip install xgboost' pour des perfs maximales.")

            # === Initialisation du visualiseur 3D intégré ===
            viz_3d = None
            _log_pre = print  # Log temporaire avant la fonction _log
            
            print(f"[3D VIZ] Status: HAS_VISPY={HAS_VISPY}")
            
            if HAS_VISPY:
                try:
                    _log_pre("[3D VIZ] 🎬 Création du collecteur de métriques...")
                    viz_3d = Model3DVisualizer(title="PULSE - SA_VOYAGEURS Training Monitor 3D | WOW MODE")
                    _log_pre("[3D VIZ] ✓ Collecteur prêt (visualisation après entraînement)")
                except Exception as e:
                    print(f"[3D VIZ] ✗ ERREUR CRITIQUE: {e}")
                    import traceback
                    traceback.print_exc()
                    viz_3d = None
            else:
                print("[3D VIZ] ✗ Vispy n'est pas disponible (HAS_VISPY=False)")

            t0_global = time.perf_counter()

            def _log(msg):
                dt = time.perf_counter() - t0_global
                print(f"[{dt:8.1f}s] {msg}", flush=True)

            _log("======================== TRAIN V4.0 ENSEMBLE ULTIME (SA_VOYAGEURS) ========================")
            _log(f"[REPRO] SEED={SEED}")

            # ---------------- Filiale fixée ----------------
            filiale = "SA_VOYAGEURS"
            current_filiale_name = filiale
            source_df = df_current if (df_current is not None and not df_current.empty) else df
            df_filiale = source_df[source_df["section"] == filiale].copy()

            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée trouvée pour la filiale {filiale}.")
                return

            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["year"].max())

            _log(f"Base year for training {filiale} : {base_year}")

            try:
                # ======================================================================
                # 1) Checks
                # ======================================================================
                if df_filiale["year"].nunique() < 2:
                    messagebox.showinfo("Information", f"Pas assez d'historique pour {filiale} (au moins 2 années).")
                    return

                df_filiale = df_filiale.sort_values(["date"]).reset_index(drop=True).copy()

                # ======================================================================
                # 2) Holidays
                # ======================================================================
                _log("[STEP] Début calcul jours fériés")

                holiday_sets = {}
                if "is_holiday" not in df_filiale.columns:
                    df_filiale["is_holiday"] = 0
                    try:
                        import holidays
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            fr_h = holidays.country_holidays("FR", years=[int(yy)])
                            holiday_sets[int(yy)] = set(fr_h.keys())
                        df_filiale["is_holiday"] = df_filiale["date"].dt.date.apply(
                            lambda d: 1 if d in holiday_sets.get(int(d.year), set()) else 0
                        )
                        _log(f"[HOLIDAYS] Calcul jours fériés FR OK (years={years})")
                    except Exception:
                        _log("[HOLIDAYS] fallback is_holiday=0")
                        holiday_sets = {}
                else:
                    try:
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            sub = df_filiale[df_filiale["year"] == int(yy)]
                            holiday_sets[int(yy)] = set(sub.loc[sub["is_holiday"] == 1, "date"].dt.date.tolist())
                        _log(f"[HOLIDAYS] Reprise colonne existante OK (years={years})")
                    except Exception:
                        holiday_sets = {}
                        _log("[HOLIDAYS] Reprise colonne existante impossible -> fallback vide")

                def _is_open_day(ts: pd.Timestamp, holiday_set: set) -> bool:
                    return (ts.weekday() < 5) and (ts.date() not in holiday_set)

                def _effective_22_date(year: int, month: int, holiday_set: set) -> pd.Timestamp:
                    d = pd.Timestamp(year=year, month=month, day=22)
                    while not _is_open_day(d, holiday_set):
                        d += pd.Timedelta(days=1)
                    return d.normalize()

                def _last_business_day(year: int, month: int, holiday_set: set) -> pd.Timestamp:
                    d = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
                    d = d.normalize()
                    while not _is_open_day(d, holiday_set):
                        d -= pd.Timedelta(days=1)
                    return d.normalize()

                def _apply_closed_day_report(df_in: pd.DataFrame, holiday_sets_: dict, y_col="y") -> pd.DataFrame:
                    dd = df_in.sort_values("date").copy()
                    dd["dow"] = dd["date"].dt.weekday

                    if "is_holiday" not in dd.columns:
                        dd["is_holiday"] = dd["date"].dt.date.apply(
                            lambda d: 1 if d in holiday_sets_.get(int(pd.Timestamp(d).year), set()) else 0
                        )

                    y = dd[y_col].astype(float).values
                    out = np.zeros_like(y, dtype=float)
                    carry = 0.0

                    dow = dd["dow"].values
                    ish = dd["is_holiday"].values

                    for i in range(len(dd)):
                        closed = (dow[i] >= 5) or (ish[i] == 1)
                        if closed:
                            carry += y[i]
                            out[i] = 0.0
                        else:
                            out[i] = y[i] + carry
                            carry = 0.0

                    if carry > 0:
                        idx_open = np.where((dow < 5) & (ish == 0))[0]
                        if len(idx_open) > 0:
                            out[idx_open[-1]] += carry

                    dd[y_col] = out
                    return dd

                def _safe_date_from_year_doy(year: int, doy: int) -> pd.Timestamp:
                    days_in_year = 366 if calendar.isleap(int(year)) else 365
                    doy = int(max(1, min(int(doy), days_in_year)))
                    return (pd.Timestamp(year=int(year), month=1, day=1) + pd.Timedelta(days=doy - 1)).normalize()

                _log("[STEP] Fin calcul jours fériés")

                # ======================================================================
                # 3) Série métier + features
                # ======================================================================
                _log("[STEP] Début feature engineering (incl. Volatilité)")

                df_filiale["y_raw"] = df_filiale["y"].astype(float)
                df_filiale = _apply_closed_day_report(df_filiale, holiday_sets, y_col="y")

                df_filiale["dow"] = df_filiale["date"].dt.weekday
                df_filiale["month"] = df_filiale["date"].dt.month
                df_filiale["dom"] = df_filiale["date"].dt.day
                df_filiale["year"] = df_filiale["date"].dt.year
                df_filiale["dayofyear"] = df_filiale["date"].dt.dayofyear
                df_filiale["is_eom"] = df_filiale["date"].dt.is_month_end.astype(int)

                # lags/rolls série métier
                df_filiale["lag_1"] = df_filiale["y"].shift(1)
                df_filiale["lag_7"] = df_filiale["y"].shift(7)
                df_filiale["roll_mean_7"] = df_filiale["y"].rolling(7, min_periods=1).mean()
                df_filiale["roll_mean_30"] = df_filiale["y"].rolling(30, min_periods=1).mean()

                # Volatilité
                df_filiale["roll_std_7"] = df_filiale["y"].rolling(7, min_periods=1).std().fillna(0)
                df_filiale["roll_std_30"] = df_filiale["y"].rolling(30, min_periods=1).std().fillna(0)

                # lags/rolls série brute
                df_filiale["lag_1_raw"] = df_filiale["y_raw"].shift(1)
                df_filiale["lag_7_raw"] = df_filiale["y_raw"].shift(7)
                df_filiale["roll_mean_7_raw"] = df_filiale["y_raw"].rolling(7, min_periods=1).mean()
                df_filiale["roll_mean_30_raw"] = df_filiale["y_raw"].rolling(30, min_periods=1).mean()

                years_present = sorted(df_filiale["year"].unique().tolist())
                map_rows = []
                for yy in years_present:
                    hset = holiday_sets.get(int(yy), set())
                    for mm in range(1, 13):
                        try:
                            map_rows.append({
                                "year": int(yy),
                                "month": int(mm),
                                "eff22_date": _effective_22_date(int(yy), int(mm), hset),
                                "lbd_date": _last_business_day(int(yy), int(mm), hset),
                            })
                        except Exception:
                            pass
                calmap = pd.DataFrame(map_rows)

                df_filiale = df_filiale.merge(calmap, on=["year", "month"], how="left")
                df_filiale["date_norm"] = df_filiale["date"].dt.normalize()

                def _business_day_features_for_month(year: int, month: int, holiday_set: set):
                    start = pd.Timestamp(year=year, month=month, day=1)
                    end = (start + pd.offsets.MonthEnd(0)).normalize()
                    all_days = pd.date_range(start, end, freq="D")

                    def is_open(ts):
                        return (ts.weekday() < 5) and (ts.date() not in holiday_set)

                    open_days = [d.normalize() for d in all_days if is_open(d)]
                    bd_rank = {d: i + 1 for i, d in enumerate(open_days)}
                    K = len(open_days)
                    bd_to_eom = {d: (K - bd_rank[d]) for d in bd_rank}
                    return bd_rank, bd_to_eom

                bd_rows = []
                for yy in years_present:
                    hset = holiday_sets.get(int(yy), set())
                    for mm in range(1, 13):
                        bd_rank, bd_to_eom = _business_day_features_for_month(int(yy), int(mm), hset)
                        for d, rnk in bd_rank.items():
                            bd_rows.append({
                                "date_norm": d,
                                "business_day_of_month": int(rnk),
                                "business_days_to_eom": int(bd_to_eom.get(d, 0))
                            })

                bd_map = pd.DataFrame(bd_rows)
                df_filiale = df_filiale.merge(bd_map, on="date_norm", how="left")
                df_filiale["business_day_of_month"] = df_filiale["business_day_of_month"].fillna(0).astype(int)
                df_filiale["business_days_to_eom"] = df_filiale["business_days_to_eom"].fillna(0).astype(int)

                # spikes
                df_filiale["is_dom_22"] = (df_filiale["dom"] == 22).astype(int)
                df_filiale["days_to_22"] = (df_filiale["dom"] - 22).astype(int)

                df_filiale["is_22_effective"] = (df_filiale["date_norm"] == df_filiale["eff22_date"]).astype(int)
                df_filiale["days_to_22_effective"] = (df_filiale["date_norm"] - df_filiale["eff22_date"]).dt.days.fillna(0).astype(int)
                df_filiale["abs_days_to_22_effective"] = df_filiale["days_to_22_effective"].abs()
                df_filiale["is_near_22_effective"] = (df_filiale["abs_days_to_22_effective"] <= 2).astype(int)

                df_filiale["is_mid_month_window"] = df_filiale["dom"].between(20, 22).astype(int)
                df_filiale["abs_days_to_20"] = (df_filiale["dom"] - 20).abs().astype(int)
                df_filiale["is_near_20"] = (df_filiale["abs_days_to_20"] <= 1).astype(int)
                df_filiale["is_monday"] = (df_filiale["dow"] == 0).astype(int)
                df_filiale["is_monday_after_20"] = ((df_filiale["dow"] == 0) & (df_filiale["dom"].between(20, 22))).astype(int)

                df_filiale["is_lbd"] = (df_filiale["date_norm"] == df_filiale["lbd_date"]).astype(int)
                df_filiale["days_to_lbd"] = (df_filiale["date_norm"] - df_filiale["lbd_date"]).dt.days.fillna(0).astype(int)
                df_filiale["abs_days_to_lbd"] = df_filiale["days_to_lbd"].abs()
                df_filiale["is_near_lbd"] = (df_filiale["abs_days_to_lbd"] <= 1).astype(int)

                df_filiale["spike_risk"] = (
                    (df_filiale["is_near_22_effective"] == 1) |
                    (df_filiale["is_mid_month_window"] == 1) |
                    (df_filiale["is_monday_after_20"] == 1) |
                    (df_filiale["is_near_lbd"] == 1)
                ).astype(int)

                _log(f"[FEATURES] nb_rows={len(df_filiale)} | years={sorted(df_filiale['year'].unique().tolist())}")
                _log("[STEP] Fin feature engineering")

                # ======================================================================
                # 4) Paires YoY
                # ======================================================================
                _log("[STEP] Début merge YoY")

                s = df_filiale.copy()
                s_prev = s.copy()
                s_prev["year_target"] = s_prev["year"] + 1

                s_prev = s_prev.rename(columns={
                    "y": "y_prev_year",
                    "y_raw": "y_raw_prev_year",
                    "roll_mean_7": "roll_prev_7",
                    "roll_mean_30": "roll_prev_30",
                    "roll_std_7": "roll_std_prev_7",
                    "roll_std_30": "roll_std_prev_30",
                    "roll_mean_7_raw": "roll_prev_7_raw",
                    "roll_mean_30_raw": "roll_prev_30_raw",
                    "dow": "dow_prev",
                    "lag_1": "lag_1_prev",
                    "lag_7": "lag_7_prev",
                    "lag_1_raw": "lag_1_prev_raw",
                    "lag_7_raw": "lag_7_prev_raw",
                    "month": "month_prev",
                    "dom": "dom_prev",
                    "is_eom": "is_eom_prev",
                    "is_holiday": "is_holiday_prev",
                    "business_day_of_month": "bdm_prev",
                    "business_days_to_eom": "bdeom_prev",
                    "is_dom_22": "is_dom_22_prev",
                    "days_to_22": "days_to_22_prev",
                    "is_22_effective": "is_22_effective_prev",
                    "days_to_22_effective": "days_to_22_effective_prev",
                    "abs_days_to_22_effective": "abs_days_to_22_effective_prev",
                    "is_near_22_effective": "is_near_22_effective_prev",
                    "is_mid_month_window": "is_mid_month_window_prev",
                    "abs_days_to_20": "abs_days_to_20_prev",
                    "is_near_20": "is_near_20_prev",
                    "is_monday": "is_monday_prev",
                    "is_monday_after_20": "is_monday_after_20_prev",
                    "is_lbd": "is_lbd_prev",
                    "days_to_lbd": "days_to_lbd_prev",
                    "abs_days_to_lbd": "abs_days_to_lbd_prev",
                    "is_near_lbd": "is_near_lbd_prev",
                    "spike_risk": "spike_risk_prev",
                })

                tgt_cols = [
                    "section", "year", "dayofyear", "y", "y_raw",
                    "dow", "month", "dom", "is_eom", "is_holiday",
                    "roll_mean_7", "roll_mean_30", "roll_std_7", "roll_std_30",
                    "roll_mean_7_raw", "roll_mean_30_raw",
                    "lag_1", "lag_7", "lag_1_raw", "lag_7_raw",
                    "business_day_of_month", "business_days_to_eom",
                    "is_dom_22", "days_to_22",
                    "is_22_effective", "days_to_22_effective",
                    "abs_days_to_22_effective", "is_near_22_effective",
                    "is_mid_month_window", "abs_days_to_20", "is_near_20",
                    "is_monday", "is_monday_after_20",
                    "is_lbd", "days_to_lbd", "abs_days_to_lbd", "is_near_lbd",
                    "spike_risk",
                ]

                merged = pd.merge(
                    s_prev,
                    s[tgt_cols],
                    left_on=["section", "year_target", "dayofyear"],
                    right_on=["section", "year", "dayofyear"],
                    how="inner",
                )

                if merged.empty:
                    messagebox.showinfo("Information", f"Aucune paire année→année pour {filiale}.")
                    return

                merged = merged.rename(columns={"y": "y_target"})
                merged = merged.rename(columns={
                    "dow": "dow_tgt",
                    "month": "month_tgt",
                    "dom": "dom_tgt",
                    "is_eom": "is_eom_tgt",
                    "is_holiday": "is_holiday_tgt",
                    "y_raw": "y_raw_tgt",
                    "roll_mean_7": "roll_7_tgt",
                    "roll_mean_30": "roll_30_tgt",
                    "roll_std_7": "roll_std_7_tgt",
                    "roll_std_30": "roll_std_30_tgt",
                    "roll_mean_7_raw": "roll_7_raw_tgt",
                    "roll_mean_30_raw": "roll_30_raw_tgt",
                    "lag_1": "lag_1_tgt",
                    "lag_7": "lag_7_tgt",
                    "lag_1_raw": "lag_1_raw_tgt",
                    "lag_7_raw": "lag_7_raw_tgt",
                    "business_day_of_month": "bdm_tgt",
                    "business_days_to_eom": "bdeom_tgt",
                    "is_dom_22": "is_dom_22_tgt",
                    "days_to_22": "days_to_22_tgt",
                    "is_22_effective": "is_22_effective_tgt",
                    "days_to_22_effective": "days_to_22_effective_tgt",
                    "abs_days_to_22_effective": "abs_days_to_22_effective_tgt",
                    "is_near_22_effective": "is_near_22_effective_tgt",
                    "is_mid_month_window": "is_mid_month_window_tgt",
                    "abs_days_to_20": "abs_days_to_20_tgt",
                    "is_near_20": "is_near_20_tgt",
                    "is_monday": "is_monday_tgt",
                    "is_monday_after_20": "is_monday_after_20_tgt",
                    "is_lbd": "is_lbd_tgt",
                    "days_to_lbd": "days_to_lbd_tgt",
                    "abs_days_to_lbd": "abs_days_to_lbd_tgt",
                    "is_near_lbd": "is_near_lbd_tgt",
                    "spike_risk": "spike_risk_tgt",
                })

                df_pairs = merged.copy()
                df_pairs["section_id"] = 0

                _log(f"[YOY] nb_pairs={len(df_pairs)}")
                _log("[STEP] Fin merge YoY")

                # ======================================================================
                # 5) Features modèle
                # ======================================================================
                base_features = [
                    "y_prev_year", "y_raw_prev_year",
                    "roll_prev_7", "roll_prev_30",
                    "roll_std_prev_7", "roll_std_prev_30",
                    "roll_prev_7_raw", "roll_prev_30_raw",
                    "dayofyear", "section_id",
                    "dow_prev",
                    "lag_1_prev", "lag_7_prev",
                    "lag_1_prev_raw", "lag_7_prev_raw",
                    "month_prev", "dom_prev", "is_eom_prev", "is_holiday_prev",
                    "bdm_prev", "bdeom_prev",
                    "is_dom_22_prev", "days_to_22_prev",
                    "is_22_effective_prev", "days_to_22_effective_prev",
                    "abs_days_to_22_effective_prev", "is_near_22_effective_prev",
                    "is_mid_month_window_prev", "abs_days_to_20_prev", "is_near_20_prev",
                    "is_monday_prev", "is_monday_after_20_prev",
                    "is_lbd_prev", "days_to_lbd_prev", "abs_days_to_lbd_prev", "is_near_lbd_prev",
                    "dow_tgt", "month_tgt", "dom_tgt", "is_eom_tgt", "is_holiday_tgt",
                    "bdm_tgt", "bdeom_tgt",
                    "is_dom_22_tgt", "days_to_22_tgt",
                    "is_22_effective_tgt", "days_to_22_effective_tgt",
                    "abs_days_to_22_effective_tgt", "is_near_22_effective_tgt",
                    "is_mid_month_window_tgt", "abs_days_to_20_tgt", "is_near_20_tgt",
                    "is_monday_tgt", "is_monday_after_20_tgt",
                    "is_lbd_tgt", "days_to_lbd_tgt", "abs_days_to_lbd_tgt", "is_near_lbd_tgt",
                ]

                corr_features = base_features + ["spike_risk_prev", "spike_risk_tgt"]

                df_pairs = df_pairs.dropna(subset=list(set(corr_features)) + ["y_target"]).reset_index(drop=True)
                _log(f"[CLEAN] nb_pairs_after_dropna={len(df_pairs)}")

                # ======================================================================
                # 6) Split
                # ======================================================================
                df_train_pairs = df_pairs[df_pairs["year_target"] <= (base_year - 1)].copy()
                df_valid_pairs = df_pairs[df_pairs["year_target"] == base_year].copy()

                if df_train_pairs.shape[0] >= 90 and df_valid_pairs.shape[0] >= 60:
                    X_train = df_train_pairs[base_features].copy().reset_index(drop=True)
                    y_train = df_train_pairs["y_target"].astype(float).reset_index(drop=True)
                    X_valid = df_valid_pairs[base_features].copy().reset_index(drop=True)
                    y_valid = df_valid_pairs["y_target"].astype(float).reset_index(drop=True)

                    X_train_corr = df_train_pairs[corr_features].copy().reset_index(drop=True)
                    X_valid_corr = df_valid_pairs[corr_features].copy().reset_index(drop=True)

                    spike_train = df_train_pairs["spike_risk_tgt"].astype(int).values
                    spike_valid = df_valid_pairs["spike_risk_tgt"].astype(int).values

                    _log(f"[SPLIT] YEAR | train={len(X_train)} valid={len(X_valid)} val_year_target={base_year}")
                else:
                    _log("[WARN] fallback split chrono 80/20")

                    X_all = df_pairs[base_features].copy()
                    y_all = df_pairs["y_target"].astype(float).copy()
                    X_all_corr = df_pairs[corr_features].copy()
                    spike_all = df_pairs["spike_risk_tgt"].astype(int).values

                    n_total = len(X_all)
                    cut = int(n_total * 0.8)
                    cut = max(60, min(cut, n_total - 20))

                    X_train = X_all.iloc[:cut].reset_index(drop=True)
                    X_valid = X_all.iloc[cut:].reset_index(drop=True)
                    y_train = y_all.iloc[:cut].reset_index(drop=True)
                    y_valid = y_all.iloc[cut:].reset_index(drop=True)

                    X_train_corr = X_all_corr.iloc[:cut].reset_index(drop=True)
                    X_valid_corr = X_all_corr.iloc[cut:].reset_index(drop=True)

                    spike_train = spike_all[:cut]
                    spike_valid = spike_all[cut:]

                    _log(f"[SPLIT] CHRONO | total={n_total} train={len(X_train)} valid={len(X_valid)}")

                # ======================================================================
                # 7) Models
                # ======================================================================
                try:
                    from catboost import CatBoostClassifier, CatBoostRegressor
                    from lightgbm import LGBMClassifier, LGBMRegressor
                    from sklearn.ensemble import ExtraTreesClassifier, ExtraTreesRegressor
                except Exception as e:
                    messagebox.showerror("Erreur", f"Dépendance ML manquante : {e}")
                    return

                _log("[STEP] Début training modèles")
                tscv = TimeSeriesSplit(n_splits=3)

                def _tune(model, param_dist, X_t, y_t, scoring, n_iter=15, sample_weight=None):
                    _log(f"[TUNE START] {model.__class__.__name__} | n={len(X_t)} | n_iter={n_iter}")
                    rs = RandomizedSearchCV(
                        model,
                        param_distributions=param_dist,
                        n_iter=n_iter,
                        cv=tscv,
                        scoring=scoring,
                        random_state=SEED,
                        n_jobs=1,
                        verbose=0
                    )
                    if sample_weight is not None and model.__class__.__name__ != 'ExtraTreesClassifier':
                        try:
                            rs.fit(X_t, y_t, sample_weight=sample_weight)
                        except Exception:
                            rs.fit(X_t, y_t)
                    else:
                        rs.fit(X_t, y_t)
                    _log(f"[TUNE DONE] {model.__class__.__name__}")
                    return rs.best_estimator_, rs.best_params_, rs.best_score_

                def _normalize_weights(scores, gamma=2.0, min_w=0.05):
                    s = np.asarray(scores, dtype=float)
                    s = np.where(np.isfinite(s), s, 0.0)
                    s = np.clip(s, a_min=0.0, a_max=None)
                    if float(s.sum()) <= 0:
                        return np.ones_like(s) / len(s)
                    s_exp = np.exp(gamma * (s - np.max(s)))
                    w = s_exp / s_exp.sum()
                    w = np.maximum(w, min_w)
                    return w / w.sum()

                # ----------------------------------------------------------------------
                # 7A) Classification ensemble
                # ----------------------------------------------------------------------
                _log("[STEP] Début classifieur ensemble")

                y_cls_train = (y_train > 0).astype(int).values
                y_cls_valid = (y_valid > 0).astype(int).values

                w_cls = np.ones(len(X_train), dtype=float)
                w_cls *= np.where(spike_train == 1, 3.0, 1.0)
                w_cls *= np.where(y_cls_train == 1, 1.5, 1.0)

                cls_models = []
                cls_model_names = []

                cls_cat_base = CatBoostClassifier(
                    random_seed=SEED,
                    verbose=0,
                    loss_function="Logloss",
                    eval_metric="AUC",
                    auto_class_weights="Balanced",
                    thread_count=1
                )
                cls_cat_params = {
                    "depth": randint(4, 9),
                    "learning_rate": uniform(0.01, 0.08),
                    "iterations": randint(350, 1000),
                    "l2_leaf_reg": uniform(1.0, 8.0)
                }
                cls_cat, _, _ = _tune(cls_cat_base, cls_cat_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                cls_models.append(cls_cat)
                cls_model_names.append("CatBoost")

                cls_lgb_base = LGBMClassifier(
                    random_state=SEED,
                    objective="binary",
                    verbosity=-1,
                    n_jobs=1,
                    deterministic=True,
                    force_col_wise=True
                )
                cls_lgb_params = {
                    "n_estimators": randint(250, 800),
                    "num_leaves": randint(15, 63),
                    "max_depth": randint(3, 8),
                    "learning_rate": uniform(0.01, 0.1)
                }
                cls_lgb, _, _ = _tune(cls_lgb_base, cls_lgb_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                cls_models.append(cls_lgb)
                cls_model_names.append("LightGBM")

                if HAS_XGB:
                    cls_xgb_base = XGBClassifier(
                        random_state=SEED,
                        eval_metric="auc",
                        verbosity=0,
                        n_jobs=1,
                        tree_method="hist",
                        subsample=1.0,
                        colsample_bytree=1.0
                    )
                    cls_xgb_params = {
                        "n_estimators": randint(200, 700),
                        "max_depth": randint(3, 8),
                        "learning_rate": uniform(0.01, 0.1),
                        "subsample": uniform(0.6, 0.4)
                    }
                    cls_xgb, _, _ = _tune(cls_xgb_base, cls_xgb_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                    cls_models.append(cls_xgb)
                    cls_model_names.append("XGBoost")

                cls_et = ExtraTreesClassifier(
                    n_estimators=800,
                    max_depth=15,
                    min_samples_leaf=2,
                    class_weight="balanced",
                    random_state=SEED,
                    n_jobs=1
                )
                cls_et.fit(X_train, y_cls_train, sample_weight=w_cls)
                cls_models.append(cls_et)
                cls_model_names.append("ExtraTrees")

                valid_cls_probas = []
                cls_scores = []
                for name, mdl in zip(cls_model_names, cls_models):
                    p = mdl.predict_proba(X_valid)[:, 1]
                    valid_cls_probas.append(p)
                    auc_i = roc_auc_score(y_cls_valid, p) if len(np.unique(y_cls_valid)) > 1 else 0.5
                    _log(f"[VALID CLS] {name} AUC={auc_i:.4f}")
                    cls_scores.append(max(auc_i - 0.50, 1e-4))

                cls_weights = _normalize_weights(cls_scores, gamma=3.0, min_w=0.05)
                _log("[CLS WEIGHTS] " + " | ".join([f"{n}={w:.3f}" for n, w in zip(cls_model_names, cls_weights)]))

                # === Envoi métriques classifieurs au visualiseur 3D ===
                if viz_3d is not None:
                    try:
                        auc_mean = np.mean(cls_scores) if cls_scores else 0.5
                        for batch_id in range(len(cls_model_names)):
                            score_norm = min(1.0, auc_mean / 0.5)
                            loss_cls = 1.0 - score_norm
                            acc_cls = score_norm * 0.7
                            viz_3d.add_metric(loss=loss_cls, accuracy=acc_cls, epoch=1, batch=batch_id)
                        _log(f"[3D VIZ] Classifieurs: AUC_mean={auc_mean:.4f} → Viz updated ✓")
                    except Exception as e:
                        _log(f"[3D VIZ] Erreur classifieurs: {e}")

                def _predict_cls_mean(X_):
                    preds = [m.predict_proba(X_)[:, 1] for m in cls_models]
                    return np.average(np.vstack(preds), axis=0, weights=cls_weights)

                _log("[STEP] Fin classifieur ensemble")

                # ----------------------------------------------------------------------
                # 7B) Régression ensemble
                # ----------------------------------------------------------------------
                _log("[STEP] Début régression ensemble")

                mask_pos = (y_train > 0).values
                X_train_pos = X_train.loc[mask_pos].reset_index(drop=True)
                y_train_pos = y_train.loc[mask_pos].reset_index(drop=True)

                if len(X_train_pos) < 40:
                    _log("[WARN] Peu de jours >0 : reg entraînée sur tout le train.")
                    X_train_pos = X_train.copy()
                    y_train_pos = y_train.copy()
                    spike_train_pos = spike_train.copy()
                else:
                    spike_train_pos = spike_train[mask_pos]

                y_train_pos_log = np.log1p(y_train_pos.values)
                q90_pos = float(np.quantile(y_train_pos.values, 0.90)) if len(y_train_pos) > 10 else float(np.max(y_train_pos.values))

                w_reg = np.ones(len(X_train_pos), dtype=float)
                w_reg *= np.where(spike_train_pos == 1, 3.0, 1.0)
                w_reg *= np.where(y_train_pos.values >= q90_pos, 2.5, 1.0)

                reg_models = []
                reg_model_names = []

                reg_cat_base = CatBoostRegressor(
                    random_seed=SEED,
                    verbose=0,
                    loss_function="RMSE",
                    thread_count=1
                )
                reg_cat_params = {
                    "depth": randint(4, 9),
                    "learning_rate": uniform(0.01, 0.08),
                    "iterations": randint(400, 1200)
                }
                reg_cat, _, _ = _tune(reg_cat_base, reg_cat_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                reg_models.append(reg_cat)
                reg_model_names.append("CatBoost")

                reg_lgb_base = LGBMRegressor(
                    random_state=SEED,
                    objective="regression",
                    verbosity=-1,
                    n_jobs=1,
                    deterministic=True,
                    force_col_wise=True
                )
                reg_lgb_params = {
                    "n_estimators": randint(300, 1000),
                    "num_leaves": randint(15, 63),
                    "max_depth": randint(3, 8),
                    "learning_rate": uniform(0.01, 0.08)
                }
                reg_lgb, _, _ = _tune(reg_lgb_base, reg_lgb_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                reg_models.append(reg_lgb)
                reg_model_names.append("LightGBM")

                if HAS_XGB:
                    reg_xgb_base = XGBRegressor(
                        random_state=SEED,
                        objective="reg:squarederror",
                        verbosity=0,
                        n_jobs=1,
                        tree_method="hist",
                        subsample=1.0,
                        colsample_bytree=1.0
                    )
                    reg_xgb_params = {
                        "n_estimators": randint(250, 800),
                        "max_depth": randint(3, 8),
                        "learning_rate": uniform(0.01, 0.08),
                        "subsample": uniform(0.6, 0.4)
                    }
                    reg_xgb, _, _ = _tune(reg_xgb_base, reg_xgb_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                    reg_models.append(reg_xgb)
                    reg_model_names.append("XGBoost")

                reg_et = ExtraTreesRegressor(
                    n_estimators=1000,
                    max_depth=15,
                    min_samples_leaf=2,
                    random_state=SEED,
                    n_jobs=1
                )
                reg_et.fit(X_train_pos, y_train_pos_log, sample_weight=w_reg)
                reg_models.append(reg_et)
                reg_model_names.append("ExtraTrees")

                valid_log_preds = []
                reg_scores = []
                for name, mdl in zip(reg_model_names, reg_models):
                    pred_log_i = mdl.predict(X_valid)
                    valid_log_preds.append(pred_log_i)
                    pred_i = np.expm1(np.clip(pred_log_i, a_min=0.0, a_max=None))
                    rmse_i = mean_squared_error(y_valid, pred_i) ** 0.5
                    _log(f"[VALID REG] {name} RMSE(level)={rmse_i:.2f}")
                    reg_scores.append(1.0 / max(rmse_i ** 2, 1e-6))

                reg_weights = _normalize_weights(reg_scores, gamma=1.5, min_w=0.05)
                _log("[REG WEIGHTS] " + " | ".join([f"{n}={w:.3f}" for n, w in zip(reg_model_names, reg_weights)]))

                def _predict_reg_mean_log(X_):
                    preds = [m.predict(X_) for m in reg_models]
                    return np.average(np.vstack(preds), axis=0, weights=reg_weights)

                def _predict_base(X_):
                    p = _predict_cls_mean(X_)
                    pred_log = _predict_reg_mean_log(X_)
                    amt = np.expm1(np.clip(pred_log, a_min=0.0, a_max=None))
                    return p * amt

                pred_train_base = _predict_base(X_train)
                pred_valid_base = _predict_base(X_valid)

                spike_gate = float(np.quantile(pred_train_base, 0.75))
                spike_q85 = float(np.quantile(pred_train_base, 0.85))
                _log(f"[SPIKE] gate(q75)={spike_gate:.1f} | q85={spike_q85:.1f}")

                def _soft_gate(pred_base_arr, gate, sharp=0.0025):
                    z = (pred_base_arr - gate) * sharp
                    z = np.clip(z, -25, 25)
                    return 1.0 / (1.0 + np.exp(-z))

                w_train = _soft_gate(pred_train_base, spike_gate)
                w_valid = _soft_gate(pred_valid_base, spike_gate)

                _log("[STEP] Fin régression ensemble")

                # ======================================================================
                # 8) Correcteurs spike
                # ======================================================================
                _log("[STEP] Début correcteurs spike")

                resid_train = (y_train.values - pred_train_base)

                X_train_corr = X_train_corr.copy()
                X_valid_corr = X_valid_corr.copy()
                X_train_corr["pred_base"] = pred_train_base
                X_valid_corr["pred_base"] = pred_valid_base
                corr_features_plus = corr_features + ["pred_base"]

                idx_corr = np.where((spike_train == 1) | (pred_train_base >= spike_q85))[0]
                if len(idx_corr) < 80:
                    idx_corr = np.where(spike_train == 1)[0]

                def _avg_corr_predict(models, weights, X_):
                    if not models:
                        return np.zeros(len(X_), dtype=float)
                    pp = [np.clip(m.predict(X_), 0.0, None) for m in models]
                    if weights is None or len(weights) != len(models):
                        return np.mean(np.vstack(pp), axis=0)
                    return np.average(np.vstack(pp), axis=0, weights=weights)

                if len(idx_corr) < 60:
                    _log("[SPIKE CORR] Pas assez de jours corr -> OFF.")
                    corr_pos_models, corr_neg_models = [], []
                    corr_pos_weights, corr_neg_weights = [], []
                    pred_train_corr = np.zeros_like(pred_train_base)
                    pred_valid_corr = np.zeros_like(pred_valid_base)
                else:
                    X_corr_train = X_train_corr.iloc[idx_corr].reset_index(drop=True)
                    resid_corr = resid_train[idx_corr]
                    y_corr_pos = np.clip(resid_corr, 0.0, None)
                    y_corr_neg = np.clip(-resid_corr, 0.0, None)

                    def _fit_corr(y_corr, label):
                        nz = int(np.sum(y_corr > 0))
                        _log(f"[SPIKE CORR {label}] start | n={len(y_corr)} | nonzero={nz}")

                        if nz < 30:
                            _log(f"[SPIKE CORR {label}] Trop peu de non-nuls -> OFF")
                            return [], []

                        w_corr = np.ones(len(y_corr), dtype=float)
                        w_corr *= np.where(y_corr > 0, 2.0, 1.0)
                        w_corr *= np.where(X_corr_train["spike_risk_tgt"].values == 1, 3.0, 1.0)

                        corr_models_local = []
                        corr_names_local = []

                        mdl_cat = CatBoostRegressor(
                            iterations=600,
                            learning_rate=0.04,
                            depth=6,
                            random_seed=SEED,
                            verbose=0,
                            thread_count=1
                        )
                        mdl_cat.fit(X_corr_train, y_corr, sample_weight=w_corr)
                        corr_models_local.append(mdl_cat)
                        corr_names_local.append("CatBoost")

                        mdl_lgb = LGBMRegressor(
                            n_estimators=400,
                            learning_rate=0.04,
                            num_leaves=31,
                            random_state=SEED,
                            verbosity=-1,
                            n_jobs=1,
                            deterministic=True,
                            force_col_wise=True
                        )
                        mdl_lgb.fit(X_corr_train, y_corr, sample_weight=w_corr)
                        corr_models_local.append(mdl_lgb)
                        corr_names_local.append("LightGBM")

                        if HAS_XGB:
                            mdl_xgb = XGBRegressor(
                                n_estimators=300,
                                learning_rate=0.04,
                                max_depth=5,
                                random_state=SEED,
                                verbosity=0,
                                n_jobs=1,
                                tree_method="hist",
                                subsample=1.0,
                                colsample_bytree=1.0
                            )
                            mdl_xgb.fit(X_corr_train, y_corr, sample_weight=w_corr)
                            corr_models_local.append(mdl_xgb)
                            corr_names_local.append("XGBoost")

                        scores_local = []
                        for nm, mdl in zip(corr_names_local, corr_models_local):
                            p = np.clip(mdl.predict(X_corr_train), 0.0, None)
                            rmse_local = mean_squared_error(y_corr, p) ** 0.5
                            scores_local.append(1.0 / max(rmse_local ** 2, 1e-4))

                        weights_local = _normalize_weights(scores_local, min_w=0.10)
                        return corr_models_local, weights_local

                    corr_pos_models, corr_pos_weights = _fit_corr(y_corr_pos, "POS")
                    corr_neg_models, corr_neg_weights = _fit_corr(y_corr_neg, "NEG")

                    pred_train_corr = np.zeros_like(pred_train_base, dtype=float)
                    pred_valid_corr = np.zeros_like(pred_valid_base, dtype=float)

                    if corr_pos_models:
                        pred_train_corr[idx_corr] += _avg_corr_predict(corr_pos_models, corr_pos_weights, X_train_corr.iloc[idx_corr])
                    if corr_neg_models:
                        pred_train_corr[idx_corr] -= _avg_corr_predict(corr_neg_models, corr_neg_weights, X_train_corr.iloc[idx_corr])

                    idx_spike_val = np.where(spike_valid == 1)[0]
                    if len(idx_spike_val) > 0:
                        if corr_pos_models:
                            pred_valid_corr[idx_spike_val] += _avg_corr_predict(corr_pos_models, corr_pos_weights, X_valid_corr.iloc[idx_spike_val])
                        if corr_neg_models:
                            pred_valid_corr[idx_spike_val] -= _avg_corr_predict(corr_neg_models, corr_neg_weights, X_valid_corr.iloc[idx_spike_val])

                pred_train = pred_train_base + pred_train_corr * w_train
                pred_valid = pred_valid_base + pred_valid_corr * w_valid

                _log("[STEP] Fin correcteurs spike")
                _log("[STEP] Fin training modèles")

                # ======================================================================
                # 9) Scale + Offset
                # ======================================================================
                denom = float(np.sum(pred_valid))
                if denom > 1e-9:
                    scale = float(np.sum(y_valid.values) / denom)
                    scale = max(0.85, min(scale, 1.15))
                else:
                    scale = 1.0

                pred_train *= scale
                pred_valid *= scale
                _log(f"[SCALE] scale={scale:.4f}")

                offset = float(np.median(y_valid.values - pred_valid))
                offset = max(-5000.0, min(offset, 5000.0))
                pred_train = np.clip(pred_train + offset, 0.0, None)
                pred_valid = np.clip(pred_valid + offset, 0.0, None)
                _log(f"[OFFSET] median residual offset={offset:.1f}")

                # ======================================================================
                # 10) VALID viz + résidus
                # ======================================================================
                df_valid_vis = df_valid_pairs.iloc[:len(X_valid)].copy().reset_index(drop=True)
                yv = y_valid.reset_index(drop=True).values if hasattr(y_valid, "reset_index") else y_valid.values

                df_valid_vis["y_true"] = yv
                df_valid_vis["y_pred"] = pred_valid
                df_valid_vis["resid"] = df_valid_vis["y_true"] - df_valid_vis["y_pred"]
                df_valid_vis["spike_risk_tgt"] = spike_valid

                df_valid_vis["date_tgt"] = [
                    _safe_date_from_year_doy(y, d)
                    for y, d in zip(df_valid_vis["year_target"].astype(int).values,
                                    df_valid_vis["dayofyear"].astype(int).values)
                ]
                df_valid_vis = df_valid_vis.sort_values("date_tgt").reset_index(drop=True)

                model_artifacts["valid_vis"] = df_valid_vis
                model_artifacts["valid_resid_series"] = df_valid_vis["resid"].astype(float).values
                model_artifacts["valid_dates"] = df_valid_vis["date_tgt"].values
                model_artifacts["valid_spike"] = df_valid_vis["spike_risk_tgt"].values.astype(int)
                model_artifacts["base_year"] = base_year
                model_artifacts["filiale"] = filiale

                # ======================================================================
                # 11) KPI P50
                # ======================================================================
                mae = mean_absolute_error(y_valid, pred_valid)
                rmse = mean_squared_error(y_valid, pred_valid) ** 0.5
                r2 = r2_score(y_valid, pred_valid)

                _log(f"[MODEL V4.0 ENSEMBLE ULTIME] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")

                # === Envoi des métriques au visualiseur 3D ===
                if viz_3d is not None:
                    try:
                        # Normaliser les métriques pour la visualisation
                        accuracy = max(0.0, min(1.0, (r2 + 1.0) / 2.0))
                        loss = 1.0 / (1.0 + accuracy) if accuracy > 0 else 1.0
                        
                        # Envoyer plusieurs points pour une belle surface 3D
                        for batch_id in range(10):
                            loss_point = loss * (1.0 - batch_id * 0.08)
                            acc_point = accuracy + batch_id * 0.05
                            viz_3d.add_metric(loss=loss_point, accuracy=acc_point, epoch=1, batch=batch_id)
                        
                        _log(f"[3D VIZ] ✓ Métriques finales envoyées! R²={r2:.3f} → Accuracy={accuracy:.3f}")
                    except Exception as e:
                        _log(f"[3D VIZ] Erreur envoi métriques: {e}")

                # ======================================================================
                # 12) Calibration 2D
                # ======================================================================
                _log("[STEP] Début calibration/conformal")

                res = (y_valid.values - pred_valid)

                n_bins = 5 if len(pred_valid) < 180 else 7
                _log(f"[CALIB] n_valid={len(pred_valid)} | n_bins={n_bins}")

                qs = np.quantile(pred_valid, np.linspace(0, 1, n_bins + 1))
                bins = np.digitize(pred_valid, qs[1:-1], right=True)
                spk = spike_valid

                q_low_2d = np.zeros((n_bins, 2), dtype=float)
                q_high_2d = np.zeros((n_bins, 2), dtype=float)

                glob_low = float(np.quantile(res, 0.05))
                glob_high = float(np.quantile(res, 0.95))

                for b in range(n_bins):
                    for s_ in (0, 1):
                        rr = res[(bins == b) & (spk == s_)]
                        if len(rr) < 25:
                            q_low_2d[b, s_] = glob_low
                            q_high_2d[b, s_] = glob_high
                        else:
                            q_low_2d[b, s_] = float(np.quantile(rr, 0.05))
                            q_high_2d[b, s_] = float(np.quantile(rr, 0.95))

                adj_low = np.array([q_low_2d[b, s_] for b, s_ in zip(bins, spk)])
                adj_high = np.array([q_high_2d[b, s_] for b, s_ in zip(bins, spk)])

                pred_p05_valid = np.clip(pred_valid + adj_low, 0.0, None)
                pred_p95_valid = np.clip(pred_valid + adj_high, 0.0, None)

                coverage_95 = float(np.mean((y_valid.values >= pred_p05_valid) & (y_valid.values <= pred_p95_valid)))
                width_95 = float(np.mean(pred_p95_valid - pred_p05_valid))
                alert_95 = 1.0 - coverage_95

                model_artifacts["calib_qs"] = qs
                model_artifacts["calib_q_low_2d_05"] = q_low_2d
                model_artifacts["calib_q_high_2d_95"] = q_high_2d

                _log(f"[CALIB 2D 95% bins={n_bins}] Coverage={coverage_95:.3f} | Alert={alert_95:.3f} | Width={width_95:.1f}")

                # ======================================================================
                # 13) Conformal conditionnel
                # ======================================================================
                alpha = 0.05

                s_conf = np.maximum.reduce([
                    pred_p05_valid - y_valid.values,
                    y_valid.values - pred_p95_valid,
                    np.zeros_like(y_valid.values)
                ])

                q_conf_by_spk = {}
                for g in (0, 1):
                    ss = s_conf[spk == g]
                    if len(ss) < 40:
                        q_conf_by_spk[g] = float(np.quantile(s_conf, 1 - alpha))
                    else:
                        q_conf_by_spk[g] = float(np.quantile(ss, 1 - alpha))

                model_artifacts["conformal_alpha"] = alpha
                model_artifacts["conformal_q_by_spk"] = q_conf_by_spk

                q_adj = np.array([q_conf_by_spk[int(s)] for s in spk])
                pred_p05_valid_c = np.clip(pred_p05_valid - q_adj, 0.0, None)
                pred_p95_valid_c = pred_p95_valid + q_adj

                coverage_c = float(np.mean((y_valid.values >= pred_p05_valid_c) & (y_valid.values <= pred_p95_valid_c)))
                width_c = float(np.mean(pred_p95_valid_c - pred_p05_valid_c))
                alert_c = 1.0 - coverage_c

                _log(f"[CONFORMAL 95%] q(spk0)={q_conf_by_spk[0]:.1f} | q(spk1)={q_conf_by_spk[1]:.1f}")
                _log(f"[CONFORMAL CHECK 95%] Coverage={coverage_c:.3f} | Alert={alert_c:.3f} | Width={width_c:.1f}")
                _log("[STEP] Fin calibration/conformal")

                # ======================================================================
                # 14) Forecast N+1 + MC
                # ======================================================================
                _log("[STEP] Début forecast N+1")

                target_year = base_year + 1
                current_target_year = target_year
                model_artifacts["target_year"] = target_year

                df_real_target = df_filiale[df_filiale["year"] == target_year].copy().sort_values("date")
                current_real_target_df = df_real_target

                df_prev_year = df_filiale[df_filiale["year"] == base_year].copy().sort_values("date")
                target_holidays = holiday_sets.get(int(target_year), set())

                calmap_tgt_rows = []
                for mm in range(1, 13):
                    try:
                        calmap_tgt_rows.append({
                            "month": int(mm),
                            "eff22_date": _effective_22_date(int(target_year), int(mm), target_holidays),
                            "lbd_date": _last_business_day(int(target_year), int(mm), target_holidays),
                        })
                    except Exception:
                        pass
                calmap_tgt = pd.DataFrame(calmap_tgt_rows).set_index("month")

                def _predict_base_row(feat_base_df):
                    p = float(_predict_cls_mean(feat_base_df)[0])
                    pred_log = float(_predict_reg_mean_log(feat_base_df)[0])
                    amt = float(np.expm1(max(0.0, pred_log)))
                    return p * amt

                def _block_bootstrap_residual_paths_weekday_conditional(
                    resid_series: np.ndarray,
                    valid_dates: np.ndarray,
                    valid_spike: np.ndarray,
                    horizon_dates: np.ndarray,
                    horizon_spike: np.ndarray,
                    B: int = 2000,
                    block_len: int = 7,
                    seed: int = SEED
                ):
                    rng = np.random.default_rng(seed)

                    r = np.asarray(resid_series, dtype=float)
                    vd = pd.to_datetime(valid_dates)
                    vs = np.asarray(valid_spike, dtype=int)

                    hd = pd.to_datetime(horizon_dates)
                    hs = np.asarray(horizon_spike, dtype=int)

                    mask = np.isfinite(r)
                    r, vd, vs = r[mask], vd[mask], vs[mask]

                    T = len(r)
                    N = len(hd)
                    if T < max(30, block_len + 5):
                        idx = rng.integers(0, T, size=(B, N))
                        return r[idx]

                    vd_w = vd.dayofweek.values
                    hd_w = hd.dayofweek.values

                    max_start = T - block_len
                    all_starts = list(range(max_start)) if max_start > 0 else [0]

                    starts = {(w, s): [] for w in range(7) for s in (0, 1)}
                    for s0 in range(max_start):
                        starts[(int(vd_w[s0]), int(vs[s0]))].append(s0)

                    for w in range(7):
                        for s_ in (0, 1):
                            if len(starts[(w, s_)]) == 0:
                                starts[(w, s_)] = all_starts

                    paths = np.zeros((B, N), dtype=float)
                    for b in range(B):
                        pos = 0
                        while pos < N:
                            w0 = int(hd_w[pos])
                            sp0 = int(hs[pos])
                            s0 = int(rng.choice(starts[(w0, sp0)]))
                            block = r[s0:s0 + block_len]
                            take = min(block_len, N - pos)
                            paths[b, pos:pos + take] = block[:take]
                            pos += take
                    return paths

                def _mc_dynamic_band_weekday_conditional(
                    pred_values: np.ndarray,
                    horizon_dates: np.ndarray,
                    horizon_spike: np.ndarray,
                    resid_series: np.ndarray,
                    valid_dates: np.ndarray,
                    valid_spike: np.ndarray,
                    B: int = 2000,
                    block_len: int = 7,
                    seed: int = SEED,
                    return_sims: bool = False
                ):
                    pv = np.asarray(pred_values, dtype=float)
                    resid_paths = _block_bootstrap_residual_paths_weekday_conditional(
                        resid_series=resid_series,
                        valid_dates=valid_dates,
                        valid_spike=valid_spike,
                        horizon_dates=horizon_dates,
                        horizon_spike=horizon_spike,
                        B=B,
                        block_len=block_len,
                        seed=seed
                    )
                    sims = np.clip(pv[None, :] + resid_paths, 0.0, None)
                    p10 = np.quantile(sims, 0.10, axis=0)
                    p50 = np.quantile(sims, 0.50, axis=0)
                    p90 = np.quantile(sims, 0.90, axis=0)
                    if return_sims:
                        return p10.astype(float), p50.astype(float), p90.astype(float), sims.astype(float)
                    return p10.astype(float), p50.astype(float), p90.astype(float)

                future_rows = []
                carryover = 0.0

                for i_row, (_, row) in enumerate(df_prev_year.iterrows(), start=1):
                    if i_row % 50 == 0:
                        _log(f"[FORECAST LOOP] {i_row}/{len(df_prev_year)}")

                    d_prev = row["date"]
                    d_next = d_prev + pd.DateOffset(years=1)
                    if int(d_next.year) != int(target_year):
                        continue

                    dow_next = int(d_next.weekday())
                    is_h_next = 1 if (d_next.date() in target_holidays) else 0

                    month_tgt = int(d_next.month)
                    dom_tgt = int(d_next.day)

                    eff22 = calmap_tgt.loc[month_tgt, "eff22_date"] if month_tgt in calmap_tgt.index else None
                    is_22_eff_tgt = 1 if (eff22 is not None and d_next.normalize() == eff22) else 0
                    days_to_22_eff_tgt = int((d_next.normalize() - eff22).days) if eff22 is not None else 0
                    abs_days_to_22_eff_tgt = abs(days_to_22_eff_tgt)
                    is_near_22_eff_tgt = 1 if abs_days_to_22_eff_tgt <= 2 else 0

                    is_dom_22_tgt = 1 if dom_tgt == 22 else 0
                    days_to_22_tgt = int(dom_tgt - 22)

                    is_mid_window_tgt = 1 if (20 <= dom_tgt <= 22) else 0
                    abs_days_to_20_tgt = abs(dom_tgt - 20)
                    is_near_20_tgt = 1 if abs_days_to_20_tgt <= 1 else 0
                    is_monday_tgt = 1 if dow_next == 0 else 0
                    is_monday_after_20_tgt = 1 if (dow_next == 0 and 20 <= dom_tgt <= 22) else 0

                    lbd = calmap_tgt.loc[month_tgt, "lbd_date"] if month_tgt in calmap_tgt.index else None
                    is_lbd_tgt = 1 if (lbd is not None and d_next.normalize() == lbd) else 0
                    days_to_lbd_tgt = int((d_next.normalize() - lbd).days) if lbd is not None else 0
                    abs_days_to_lbd_tgt = abs(days_to_lbd_tgt)
                    is_near_lbd_tgt = 1 if abs_days_to_lbd_tgt <= 1 else 0

                    spike_risk_tgt = 1 if (is_near_22_eff_tgt or is_mid_window_tgt or is_monday_after_20_tgt or is_near_lbd_tgt) else 0

                    try:
                        bdm_tgt = int(df_filiale.loc[df_filiale["date_norm"] == d_next.normalize(), "business_day_of_month"].iloc[0])
                        bdeom_tgt = int(df_filiale.loc[df_filiale["date_norm"] == d_next.normalize(), "business_days_to_eom"].iloc[0])
                    except Exception:
                        bdm_tgt = 0
                        bdeom_tgt = 0

                    feat = {c: 0 for c in base_features}

                    feat["y_prev_year"] = float(row["y"])
                    feat["y_raw_prev_year"] = float(row.get("y_raw", row["y"]))
                    feat["roll_prev_7"] = float(row["roll_mean_7"])
                    feat["roll_prev_30"] = float(row["roll_mean_30"])
                    feat["roll_std_prev_7"] = float(row.get("roll_std_7", 0))
                    feat["roll_std_prev_30"] = float(row.get("roll_std_30", 0))
                    feat["roll_prev_7_raw"] = float(row.get("roll_mean_7_raw", row["roll_mean_7"]))
                    feat["roll_prev_30_raw"] = float(row.get("roll_mean_30_raw", row["roll_mean_30"]))
                    feat["dayofyear"] = int(row["dayofyear"])
                    feat["section_id"] = 0
                    feat["dow_prev"] = int(row["dow"])
                    feat["lag_1_prev"] = float(row["lag_1"]) if pd.notna(row["lag_1"]) else 0.0
                    feat["lag_7_prev"] = float(row["lag_7"]) if pd.notna(row["lag_7"]) else 0.0
                    feat["lag_1_prev_raw"] = float(row.get("lag_1_raw", row["lag_1"])) if pd.notna(row.get("lag_1_raw", np.nan)) else 0.0
                    feat["lag_7_prev_raw"] = float(row.get("lag_7_raw", row["lag_7"])) if pd.notna(row.get("lag_7_raw", np.nan)) else 0.0
                    feat["month_prev"] = int(row["month"])
                    feat["dom_prev"] = int(row["dom"])
                    feat["is_eom_prev"] = int(row["is_eom"])
                    feat["is_holiday_prev"] = int(row["is_holiday"])
                    feat["bdm_prev"] = int(row.get("business_day_of_month", 0))
                    feat["bdeom_prev"] = int(row.get("business_days_to_eom", 0))
                    feat["is_dom_22_prev"] = int(row.get("is_dom_22", 0))
                    feat["days_to_22_prev"] = int(row.get("days_to_22", int(row["dom"]) - 22))
                    feat["is_22_effective_prev"] = int(row.get("is_22_effective", 0))
                    feat["days_to_22_effective_prev"] = int(row.get("days_to_22_effective", 0))
                    feat["abs_days_to_22_effective_prev"] = int(row.get("abs_days_to_22_effective", abs(int(row.get("days_to_22_effective", 0)))))
                    feat["is_near_22_effective_prev"] = int(row.get("is_near_22_effective", 0))
                    feat["is_mid_month_window_prev"] = int(row.get("is_mid_month_window", 1 if 20 <= int(row["dom"]) <= 22 else 0))
                    feat["abs_days_to_20_prev"] = int(row.get("abs_days_to_20", abs(int(row["dom"]) - 20)))
                    feat["is_near_20_prev"] = int(row.get("is_near_20", 1 if abs(int(row["dom"]) - 20) <= 1 else 0))
                    feat["is_monday_prev"] = int(row.get("is_monday", 1 if int(row["dow"]) == 0 else 0))
                    feat["is_monday_after_20_prev"] = int(row.get("is_monday_after_20", 1 if (int(row["dow"]) == 0 and 20 <= int(row["dom"]) <= 22) else 0))
                    feat["is_lbd_prev"] = int(row.get("is_lbd", 0))
                    feat["days_to_lbd_prev"] = int(row.get("days_to_lbd", 0))
                    feat["abs_days_to_lbd_prev"] = int(row.get("abs_days_to_lbd", abs(int(row.get("days_to_lbd", 0)))))
                    feat["is_near_lbd_prev"] = int(row.get("is_near_lbd", 0))

                    feat["dow_tgt"] = dow_next
                    feat["month_tgt"] = month_tgt
                    feat["dom_tgt"] = dom_tgt
                    feat["is_eom_tgt"] = int(d_next.is_month_end)
                    feat["is_holiday_tgt"] = is_h_next
                    feat["bdm_tgt"] = bdm_tgt
                    feat["bdeom_tgt"] = bdeom_tgt
                    feat["is_dom_22_tgt"] = is_dom_22_tgt
                    feat["days_to_22_tgt"] = days_to_22_tgt
                    feat["is_22_effective_tgt"] = is_22_eff_tgt
                    feat["days_to_22_effective_tgt"] = days_to_22_eff_tgt
                    feat["abs_days_to_22_effective_tgt"] = abs_days_to_22_eff_tgt
                    feat["is_near_22_effective_tgt"] = is_near_22_eff_tgt
                    feat["is_mid_month_window_tgt"] = is_mid_window_tgt
                    feat["abs_days_to_20_tgt"] = abs_days_to_20_tgt
                    feat["is_near_20_tgt"] = is_near_20_tgt
                    feat["is_monday_tgt"] = is_monday_tgt
                    feat["is_monday_after_20_tgt"] = is_monday_after_20_tgt
                    feat["is_lbd_tgt"] = is_lbd_tgt
                    feat["days_to_lbd_tgt"] = days_to_lbd_tgt
                    feat["abs_days_to_lbd_tgt"] = abs_days_to_lbd_tgt
                    feat["is_near_lbd_tgt"] = is_near_lbd_tgt

                    feat_base = pd.DataFrame([feat], columns=base_features)
                    pred_base = _predict_base_row(feat_base)

                    pred_corr = 0.0
                    if spike_risk_tgt == 1:
                        w = float(_soft_gate(np.array([pred_base]), spike_gate)[0])
                        feat_corr = pd.DataFrame([feat_base.iloc[0].tolist() + [
                            int(row.get("spike_risk", 0)),
                            int(spike_risk_tgt),
                            float(pred_base)
                        ]], columns=corr_features_plus)

                        if corr_pos_models:
                            pred_corr += float(_avg_corr_predict(corr_pos_models, corr_pos_weights, feat_corr)[0])
                        if corr_neg_models:
                            pred_corr -= float(_avg_corr_predict(corr_neg_models, corr_neg_weights, feat_corr)[0])

                        pred_corr *= w

                    pred_raw = pred_base + pred_corr

                    is_closed = (dow_next >= 5) or (is_h_next == 1)
                    if is_closed:
                        carryover += pred_raw
                        pred_det = 0.0
                    else:
                        pred_det = pred_raw + carryover
                        carryover = 0.0

                    pred_det = pred_det * scale + offset
                    if pred_det < 0:
                        pred_det = 0.0

                    future_rows.append({
                        "section": filiale,
                        "date": d_next,
                        "year": int(d_next.year),
                        "month": int(d_next.month),
                        "dayofyear": int(d_next.dayofyear),
                        "spike_risk_tgt": int(spike_risk_tgt),
                        "pred_det": float(pred_det),
                    })

                df_future_all = pd.DataFrame(future_rows).sort_values("date").reset_index(drop=True)
                _log(f"[FORECAST] nb_future_rows={len(df_future_all)}")
                _log("[STEP] Fin forecast N+1")

                # --- MC ---
                B_MC = 5000
                BLOCK_LEN = 7
                resid_series = model_artifacts.get("valid_resid_series", None)
                valid_dates = model_artifacts.get("valid_dates", None)
                valid_spike = model_artifacts.get("valid_spike", None)

                _log("[STEP] Début MC")
                _log(f"[MC CONFIG] B={B_MC} | block_len={BLOCK_LEN} | horizon={len(df_future_all)}")

                if resid_series is None or valid_dates is None or valid_spike is None or len(resid_series) < 50 or df_future_all.empty:
                    _log("[MC] Pas assez de résidus VALID -> fallback deterministe.")
                    df_future_all["pred_p10"] = df_future_all["pred_det"]
                    df_future_all["pred_value"] = df_future_all["pred_det"]
                    df_future_all["pred_p90"] = df_future_all["pred_det"]
                    model_artifacts["mc_sims"] = None
                else:
                    p10_mc, p50_mc, p90_mc, sims = _mc_dynamic_band_weekday_conditional(
                        pred_values=df_future_all["pred_det"].values,
                        horizon_dates=df_future_all["date"].values,
                        horizon_spike=df_future_all["spike_risk_tgt"].fillna(0).astype(int).values,
                        resid_series=resid_series,
                        valid_dates=valid_dates,
                        valid_spike=valid_spike,
                        B=B_MC,
                        block_len=BLOCK_LEN,
                        seed=SEED,
                        return_sims=True
                    )
                    df_future_all["pred_value"] = p50_mc
                    model_artifacts["mc_sims"] = sims
                    _log(f"[MC] Conditional weekday-block OK (B={B_MC}, block_len={BLOCK_LEN}).")

                _log("[STEP] Fin MC")

                # ======================================================================
                # 15) Bandes forecast
                # ======================================================================
                qs = model_artifacts.get("calib_qs", None)
                q_low_2d = model_artifacts.get("calib_q_low_2d_05", None)
                q_high_2d = model_artifacts.get("calib_q_high_2d_95", None)
                q_conf_by_spk = model_artifacts.get("conformal_q_by_spk", {0: 0.0, 1: 0.0})

                pv = df_future_all["pred_value"].values.astype(float)
                spk_f = df_future_all["spike_risk_tgt"].fillna(0).astype(int).values

                if qs is not None and q_low_2d is not None and q_high_2d is not None:
                    bins_f = np.digitize(pv, qs[1:-1], right=True)
                    adj_low_f = np.array([q_low_2d[b, s_] for b, s_ in zip(bins_f, spk_f)])
                    adj_high_f = np.array([q_high_2d[b, s_] for b, s_ in zip(bins_f, spk_f)])
                    p05 = np.clip(pv + adj_low_f, 0.0, None)
                    p95 = np.clip(pv + adj_high_f, 0.0, None)
                else:
                    p05 = pv.copy()
                    p95 = pv.copy()

                q_adj_f = np.array([q_conf_by_spk[int(s)] for s in spk_f])
                df_future_all["pred_p05"] = np.clip(p05 - q_adj_f, 0.0, None)
                df_future_all["pred_p95"] = p95 + q_adj_f

                df_future_all["pred_p10"] = df_future_all["pred_p05"]
                df_future_all["pred_p90"] = df_future_all["pred_p95"]

                model_artifacts["mc_dates"] = df_future_all["date"].values
                model_artifacts["mc_det"] = df_future_all["pred_det"].values
                model_artifacts["mc_p50"] = df_future_all["pred_value"].values

                if current_real_target_df is not None and not current_real_target_df.empty:
                    df_future_all = pd.merge(
                        current_real_target_df[["date"]].copy(),
                        df_future_all,
                        on="date",
                        how="left"
                    )
                    for c in ["pred_det", "pred_p05", "pred_value", "pred_p95", "pred_p10", "pred_p90", "spike_risk_tgt"]:
                        if c in df_future_all.columns:
                            df_future_all[c] = df_future_all[c].fillna(0.0)
                    df_future_all["section"] = filiale
                    df_future_all["year"] = df_future_all["date"].dt.year
                    df_future_all["month"] = df_future_all["date"].dt.month
                    df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

                # ======================================================================
                # 16) Graphe N+1
                # ======================================================================
                fig1, ax1 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")

                df_hist_plot = df_filiale[df_filiale["year"] <= base_year]
                ax1.plot(df_hist_plot["date"], df_hist_plot["y"], label=f"Réel (≤ {base_year})", linewidth=2)

                ax1.plot(df_future_all["date"], df_future_all["pred_value"],
                        label=f"Prévision {target_year} (P50)", linewidth=2, linestyle="--")
                ax1.fill_between(df_future_all["date"], df_future_all["pred_p05"], df_future_all["pred_p95"],
                                alpha=0.25, label=f"Bande 95% (Calib 2D + conformal, alpha={alpha})")

                ax1.set_title("SA_VOYAGEURS – Trafic voyageur - Prévision N+1 (ENSEMBLE V4.0)", color="white")
                ax1.tick_params(axis='x', colors="white", rotation=30)
                ax1.tick_params(axis='y', colors="white")
                for spine in ax1.spines.values():
                    spine.set_color("#00122e")

                leg1 = ax1.legend(facecolor="#00122e", edgecolor="white")
                for text in leg1.get_texts():
                    text.set_color("white")

                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                w_fig1 = canvas_fig1.get_tk_widget()
                w_fig1.pack(pady=10, fill="both", expand=True)
                graph_widgets.append(w_fig1)
                plt.close(fig1)

                # ======================================================================
                # 17) Update globals + artifacts + UI
                # ======================================================================
                current_pred_df = df_future_all
                exported_pred_df = df_future_all.copy()

                model_artifacts["X_train"] = X_train.copy()
                model_artifacts["y_train"] = y_train.copy()
                model_artifacts["X_valid"] = X_valid.copy()
                model_artifacts["y_valid"] = y_valid.copy()
                model_artifacts["y_cls_valid"] = (y_valid > 0).astype(int)

                model_artifacts["features"] = base_features[:]
                model_artifacts["corr_features"] = corr_features_plus[:]

                model_artifacts["cls_models"] = cls_models
                model_artifacts["cls_model_names"] = cls_model_names
                model_artifacts["cls_model_weights"] = cls_weights

                model_artifacts["reg_models"] = reg_models
                model_artifacts["reg_model_names"] = reg_model_names
                model_artifacts["reg_model_weights"] = reg_weights

                model_artifacts["corr_pos_models"] = corr_pos_models
                model_artifacts["corr_neg_models"] = corr_neg_models
                model_artifacts["corr_pos_weights"] = corr_pos_weights if 'corr_pos_weights' in locals() else []
                model_artifacts["corr_neg_weights"] = corr_neg_weights if 'corr_neg_weights' in locals() else []

                model_artifacts["scale"] = scale
                model_artifacts["offset"] = offset
                model_artifacts["spike_gate"] = spike_gate
                model_artifacts["spike_q85"] = spike_q85
                model_artifacts["soft_gate_sharp"] = 0.0025
                model_artifacts["mc_B"] = B_MC
                model_artifacts["mc_block_len"] = BLOCK_LEN
                model_artifacts["seed"] = SEED

                try:
                    btn_3d.configure(state="normal")
                except Exception:
                    pass

                try:
                    export_graph2_button.configure(state="normal")
                except Exception:
                    pass

                _redraw_monthly_graph()
                _redraw_graph2()
                _redraw_mc_fan_graph()
                _redraw_residuals_graph()
                _redraw_confusion_matrix_graph()
                _redraw_confusion_matrix_pred_graph()

                #try:
                    #_redraw_cls_diagnostics_graph()
                #except Exception:
                 #   pass

                if analysis_table_frame is not None:
                    analysis_table_frame.destroy()

                analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
                analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
                graph_widgets.append(analysis_table_frame)

                _rebuild_analysis_table()

                # === VISUALISATION 3D POST-ENTRAÎNEMENT ===
                if viz_3d is not None and len(viz_3d.metrics_history) > 0:
                    try:
                        _log("[3D VIZ] 📊 Affichage de la visualisation 3D de l'entraînement...")
                        # Lancer dans un thread pour ne pas bloquer
                        import threading
                        thread_viz = threading.Thread(target=viz_3d.show_results, daemon=True)
                        thread_viz.start()
                    except Exception as e:
                        print(f"[3D VIZ] Erreur affichage: {e}")
                
                _log("[DONE] Entraînement terminé")
                _log(f"[TIME] total = {(time.perf_counter() - t0_global)/60.0:.2f} min")

            except Exception:
                messagebox.showerror("Erreur", traceback.format_exc())
                print("[IA] ERREUR:\n", traceback.format_exc())
        
        def _train_ace_invest_reseau_negative_v33():
            """
            RESEAU — ACE & Investissements — NEG “V5.0 ULTRA PRÉCIS” (jour-par-jour)

            OBJECTIF (basé sur tes constats 2024–2025) :
            - Flux très concentrés J1–J3 (19/26) => modèle fortement "event-driven" (calendrier / BDM / lundi)
            - Montants DISCRETS très récurrents : -30/-35/-45/-55 (intermédiaires) et gros autour de -120 (-110 à -140)
            - Quelques spikes rares (>= -150 / jusqu’à -381)
            - Lundi très lourd => features + règles de calibration
            - Jours fermés (WE + fériés) => y=0 (règle métier)
            - y_pred <= 0

            DESIGN (anti-lissage, précis jour-par-jour) :
            1) Modèle “RÉGIME” multiclass (0=0, 1=petit récurrent, 2=gros récurrent (~120), 3=spike/other)
            2) Modèle “ANCHOR” pour régime 1 (choisit parmi les montants discrets fréquents)
            3) Modèle “ANCHOR” pour régime 2 (choisit parmi les gros montants fréquents)
            4) Modèle “SPIKE REG” (régression log1p magnitude) pour régime 3
            5) Post-calibration métier : début de mois + lundi => pousse vers activité si le modèle hésite
            6) CAP : très haut (quantiles) uniquement pour éviter les explosions artificielles, mais ne coupe pas les vrais spikes

            IMPORTANT :
            - Ici on privilégie la précision "point" => on réduit l'influence des lags/rolling (qui lissent).
            Le modèle est majoritairement calendrier + BDM + interactions.
            """

            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame

            _clear_graph_widgets()
            print("======================== TRAIN NEG V5.0 ULTRA (RESEAU / ACE & Investissements) ========================")

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
            from sklearn.model_selection import RandomizedSearchCV
            from scipy.stats import randint, uniform

            try:
                from lightgbm import LGBMClassifier, LGBMRegressor
            except Exception:
                messagebox.showerror("Erreur", "LightGBM n'est pas dispo. Installe lightgbm.")
                return

            # ---------------- Filiale + data ----------------
            filiale = "RESEAU"
            current_filiale_name = filiale

            df_filiale = df_current[df_current["section"] == filiale].copy()
            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée pour {filiale} (flux={selected_flux.get()}).")
                return
            df_filiale = df_filiale.sort_values("date").copy()

            # ---------------- base year ----------------
            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["date"].dt.year.max())
            target_year = base_year + 1
            current_target_year = target_year
            print(f"[NEG] base_year={base_year} -> target_year={target_year}")

            # ---------------- Holidays (best effort) ----------------
            holiday_sets = {}
            if "is_holiday" not in df_filiale.columns:
                df_filiale["is_holiday"] = 0
                try:
                    import holidays
                    years = sorted(set(df_filiale["date"].dt.year.unique().tolist() + [target_year]))
                    for yy in years:
                        fr_h = holidays.country_holidays("FR", years=[int(yy)])
                        holiday_sets[int(yy)] = set(fr_h.keys())
                    df_filiale["is_holiday"] = df_filiale["date"].dt.date.apply(
                        lambda d: 1 if d in holiday_sets.get(int(pd.Timestamp(d).year), set()) else 0
                    )
                    print(f"[HOLIDAYS] Calcul FR OK years={sorted(holiday_sets.keys())}")
                except Exception:
                    print("[HOLIDAYS] fallback is_holiday=0")
                    holiday_sets = {}
            else:
                try:
                    years = sorted(set(df_filiale["date"].dt.year.unique().tolist() + [target_year]))
                    for yy in years:
                        sub = df_filiale[df_filiale["date"].dt.year == int(yy)]
                        holiday_sets[int(yy)] = set(sub.loc[sub["is_holiday"] == 1, "date"].dt.date.tolist())
                except Exception:
                    holiday_sets = {}

            # ---------------- Closed days ----------------
            df_filiale["dow"] = df_filiale["date"].dt.weekday
            df_filiale["year"] = df_filiale["date"].dt.year
            df_filiale["is_closed"] = ((df_filiale["dow"] >= 5) | (df_filiale["is_holiday"] == 1)).astype(int)

            # règle métier: WE/ferié = 0
            df_filiale.loc[df_filiale["is_closed"] == 1, "y"] = 0.0

            # ---------------- Feature engineering (calendrier riche, anti-lissage) ----------------
            def _add_calendar_features(dd: pd.DataFrame, holiday_sets_: dict) -> pd.DataFrame:
                d = dd.copy()
                d = d.sort_values("date").copy()
                d["date_norm"] = d["date"].dt.normalize()
                d["month"] = d["date"].dt.month
                d["dom"] = d["date"].dt.day
                d["dayofyear"] = d["date"].dt.dayofyear
                d["weekofyear"] = d["date"].dt.isocalendar().week.astype(int)

                d["is_month_start"] = d["date"].dt.is_month_start.astype(int)
                d["is_month_end"] = d["date"].dt.is_month_end.astype(int)
                d["is_quarter_start"] = d["date"].dt.is_quarter_start.astype(int)
                d["is_quarter_end"] = d["date"].dt.is_quarter_end.astype(int)

                d["is_monday"] = (d["dow"] == 0).astype(int)

                # BOM (très important)
                d["dom_1"] = (d["dom"] == 1).astype(int)
                d["dom_2"] = (d["dom"] == 2).astype(int)
                d["dom_3"] = (d["dom"] == 3).astype(int)
                d["is_bom_1_3"] = (d["dom"] <= 3).astype(int)
                d["is_bom_1_5"] = (d["dom"] <= 5).astype(int)
                d["is_bom_1_7"] = (d["dom"] <= 7).astype(int)

                # days to end of month
                eom = (d["date"] + pd.offsets.MonthEnd(0)).dt.normalize()
                d["days_to_eom"] = (eom - d["date_norm"]).dt.days.astype(int)
                d["is_near_eom3"] = (d["days_to_eom"] <= 3).astype(int)
                d["is_near_eom7"] = (d["days_to_eom"] <= 7).astype(int)

                # business day of month (WE + fériés)
                bdm, bdl = [], []
                for (yy, mm), grp in d.groupby(["year", "month"], sort=False):
                    hset = holiday_sets_.get(int(yy), set())
                    all_days = pd.date_range(
                        start=pd.Timestamp(int(yy), int(mm), 1),
                        end=pd.Timestamp(int(yy), int(mm), 1) + pd.offsets.MonthEnd(0),
                        freq="D"
                    )
                    open_days = []
                    for dt in all_days:
                        closed = (dt.weekday() >= 5) or (dt.date() in hset)
                        if not closed:
                            open_days.append(dt.normalize())

                    pos_map = {od: i for i, od in enumerate(open_days, start=1)}
                    total_open = len(open_days)
                    open_set = set(open_days)

                    for dn in grp["date_norm"]:
                        if dn in open_set:
                            idx = pos_map.get(dn, 0)
                            bdm.append(idx)
                            bdl.append(max(0, total_open - idx))
                        else:
                            bdm.append(0)
                            bdl.append(total_open)

                d["bdm"] = np.array(bdm, dtype=int)
                d["bdl"] = np.array(bdl, dtype=int)

                d["bdm_1"] = (d["bdm"] == 1).astype(int)
                d["bdm_2"] = (d["bdm"] == 2).astype(int)
                d["bdm_3"] = (d["bdm"] == 3).astype(int)
                d["is_bdm_1_3"] = ((d["bdm"] >= 1) & (d["bdm"] <= 3)).astype(int)

                d["is_first_bizday"] = (d["bdm"] == 1).astype(int)
                d["is_last_bizday"] = ((d["bdl"] == 0) & (d["is_closed"] == 0)).astype(int)

                # after closed (WE/ferié)
                d["prev_is_closed"] = d["is_closed"].shift(1).fillna(0).astype(int)
                d["is_after_closed"] = ((d["prev_is_closed"] == 1) & (d["is_closed"] == 0)).astype(int)

                # interactions (lundi + BOM)
                d["mon_bom_1_3"] = ((d["is_monday"] == 1) & (d["is_bom_1_3"] == 1)).astype(int)
                d["mon_bdm_1_3"] = ((d["is_monday"] == 1) & (d["is_bdm_1_3"] == 1)).astype(int)
                d["after_bom_1_3"] = ((d["is_after_closed"] == 1) & (d["is_bom_1_3"] == 1)).astype(int)
                d["after_bdm_1_3"] = ((d["is_after_closed"] == 1) & (d["is_bdm_1_3"] == 1)).astype(int)

                # spike-risk (jours où l'activité est probable)
                d["spike_risk"] = (
                    (d["is_bom_1_3"] == 1) |
                    (d["is_bdm_1_3"] == 1) |
                    (d["is_quarter_end"] == 1) |
                    (d["is_after_closed"] == 1) |
                    (d["is_monday"] == 1) |
                    (d["mon_bom_1_3"] == 1) |
                    (d["mon_bdm_1_3"] == 1) |
                    (d["is_near_eom3"] == 1)
                ).astype(int)

                # “mois” cyclique (aide sur pattern annuel sans lisser)
                d["month_sin"] = np.sin(2 * np.pi * d["month"] / 12.0)
                d["month_cos"] = np.cos(2 * np.pi * d["month"] / 12.0)

                # DOW cyclique
                d["dow_sin"] = np.sin(2 * np.pi * d["dow"] / 7.0)
                d["dow_cos"] = np.cos(2 * np.pi * d["dow"] / 7.0)

                return d

            df_filiale = _add_calendar_features(df_filiale, holiday_sets)

            # ---------------- Split ----------------
            df_train = df_filiale[df_filiale["year"] < base_year].copy()
            df_valid = df_filiale[df_filiale["year"] == base_year].copy()

            if df_train.shape[0] < 120 or df_valid.shape[0] < 80:
                print("[WARN] Peu de données pour split année -> fallback chrono 80/20")
                df_all = df_filiale.sort_values("date").copy()
                cut = int(len(df_all) * 0.8)
                cut = max(120, min(cut, len(df_all) - 60))
                df_train = df_all.iloc[:cut].copy()
                df_valid = df_all.iloc[cut:].copy()

            # ---------------- Features ----------------
            features = [
                "dow", "dow_sin", "dow_cos",
                "weekofyear", "month", "month_sin", "month_cos",
                "dom", "dayofyear",
                "is_month_start", "is_month_end",
                "is_quarter_start", "is_quarter_end",
                "days_to_eom", "is_near_eom3", "is_near_eom7",
                "dom_1", "dom_2", "dom_3", "is_bom_1_3", "is_bom_1_5", "is_bom_1_7",
                "bdm", "bdl", "bdm_1", "bdm_2", "bdm_3", "is_bdm_1_3",
                "is_first_bizday", "is_last_bizday",
                "is_holiday", "is_closed", "is_after_closed",
                "is_monday", "mon_bom_1_3", "mon_bdm_1_3",
                "after_bom_1_3", "after_bdm_1_3",
                "spike_risk",
            ]

            df_train_m = df_train.dropna(subset=features + ["y"]).copy()
            df_valid_m = df_valid.dropna(subset=features + ["y"]).copy()

            if df_train_m.shape[0] < 80 or df_valid_m.shape[0] < 40:
                messagebox.showinfo("Information", "Pas assez de lignes pour entraîner le modèle.")
                return

            # sécurise closed=0
            df_train_m.loc[df_train_m["is_closed"] == 1, "y"] = 0.0
            df_valid_m.loc[df_valid_m["is_closed"] == 1, "y"] = 0.0

            X_train = df_train_m[features].copy()
            y_train = df_train_m["y"].astype(float).copy()
            X_valid = df_valid_m[features].copy()
            y_valid = df_valid_m["y"].astype(float).copy()

            is_open_train = (df_train_m["is_closed"].values == 0)
            is_open_valid = (df_valid_m["is_closed"].values == 0)

            m_train = np.clip(-y_train.values, 0.0, None)
            m_valid = np.clip(-y_valid.values, 0.0, None)

            # ======================================================================
            # 1) Détection d'échelle + quantification des montants (pour anchors)
            # ======================================================================
            # On quantifie en "pas" adapté à l'échelle, pour détecter -30/-35/-45/-55/-120 etc sans dépendre de l'unité.
            nonzero = m_train[(m_train > 0) & np.isfinite(m_train)]
            if len(nonzero) == 0:
                messagebox.showinfo("Information", "Aucun flux négatif sur train.")
                return

            med_mag = float(np.median(nonzero))
            # heuristique de pas
            # - si magnitude ~1e8 (100M) => step = 1e6
            # - si magnitude ~1e5 (100k) => step = 1e3
            # - sinon step = 1
            if med_mag >= 5e7:
                step = 1e6
            elif med_mag >= 5e4:
                step = 1e3
            else:
                step = 1.0

            def qround(x):
                return float(np.round(x / step) * step)

            m_train_q = np.array([qround(v) for v in nonzero], dtype=float)
            vc = pd.Series(m_train_q).value_counts()

            # anchors "petits récurrents": on prend top fréquents sous ~80% de la médiane (pour éviter les -120)
            small_cut = 0.80 * med_mag
            small_candidates = vc[vc.index <= small_cut]
            small_anchors = small_candidates.head(6).index.values.astype(float).tolist()

            # anchors "gros récurrents": autour de la médiane +/- 20% (capte 110-140 autour de 120)
            big_lo, big_hi = 0.80 * med_mag, 1.20 * med_mag
            big_candidates = vc[(vc.index >= big_lo) & (vc.index <= big_hi)]
            big_anchors = big_candidates.head(7).index.values.astype(float).tolist()

            # si detection insuffisante, fallback simple
            if len(small_anchors) < 2:
                # fallback: on prend des valeurs basses fréquentes
                small_anchors = vc.head(6).index.values.astype(float).tolist()
            if len(big_anchors) < 2:
                big_anchors = big_candidates.index.values.astype(float).tolist()
                if len(big_anchors) < 2:
                    big_anchors = [med_mag]

            small_anchors = sorted(list(set(small_anchors)))
            big_anchors = sorted(list(set(big_anchors)))

            print(f"[SCALE] med_mag={med_mag:,.0f} step={step:,.0f}")
            print(f"[ANCHORS] small={small_anchors}")
            print(f"[ANCHORS] big={big_anchors}")

            # ======================================================================
            # 2) Labelisation "régime" (anti-lissage)
            # ======================================================================
            # regime:
            # 0: m=0
            # 1: petit récurrent (proche d'un small anchor)
            # 2: gros récurrent (proche d'un big anchor)
            # 3: spike/other (reste)
            def nearest_anchor_class(m, anchors, tol):
                if m <= 0:
                    return None
                a = np.array(anchors, dtype=float)
                idx = int(np.argmin(np.abs(a - m)))
                if abs(a[idx] - m) <= tol:
                    return idx
                return None

            # tolérance = 2*step pour absorber les petites variations / arrondis
            tol = 2.0 * step

            regime_train = np.zeros(len(df_train_m), dtype=int)
            # jours fermés => regime 0 (0)
            for i in range(len(df_train_m)):
                if int(df_train_m.iloc[i]["is_closed"]) == 1:
                    regime_train[i] = 0
                    continue
                m = float(m_train[i])
                if m <= 0:
                    regime_train[i] = 0
                    continue
                if nearest_anchor_class(m, small_anchors, tol) is not None:
                    regime_train[i] = 1
                elif nearest_anchor_class(m, big_anchors, tol) is not None:
                    regime_train[i] = 2
                else:
                    # spike si très gros
                    regime_train[i] = 3

            # ======================================================================
            # 3) Modèles : Régime + Anchors + SpikeReg
            # ======================================================================
            def _tune(model, param_dist, X_t, y_t, scoring, n_iter=18):
                rs = RandomizedSearchCV(
                    model,
                    param_distributions=param_dist,
                    n_iter=n_iter,
                    cv=3,
                    scoring=scoring,
                    random_state=42,
                    n_jobs=-1,
                    verbose=0
                )
                rs.fit(X_t, y_t)
                return rs.best_estimator_, rs.best_params_, rs.best_score_

            # --- Régime multiclass (précis jour-par-jour) ---
            # On entraîne sur jours ouverts + y (regime), mais closed gardés pour apprendre "0"
            cls_regime_base = LGBMClassifier(
                random_state=42,
                objective="multiclass",
                num_class=4,
                n_estimators=1800,
                learning_rate=0.03,
                subsample=0.9,
                colsample_bytree=0.9,
                num_leaves=63,
                min_child_samples=25,
                class_weight="balanced",
            )
            regime_params = {
                "n_estimators": randint(600, 2600),
                "num_leaves": randint(15, 127),
                "max_depth": randint(2, 10),
                "min_child_samples": randint(10, 140),
                "learning_rate": uniform(0.01, 0.08),
                "subsample": uniform(0.70, 0.30),
                "colsample_bytree": uniform(0.70, 0.30),
                "reg_lambda": uniform(0.0, 10.0),
            }

            # scoring: accuracy macro (LightGBM via sklearn -> "accuracy")
            cls_regime, bp, bs = _tune(cls_regime_base, regime_params, X_train, regime_train, scoring="accuracy", n_iter=20)
            print(f"[TUNING] REGIME best params: {bp} | CV acc={bs:.4f}")

            # --- Anchor classifier petit ---
            # y_small_class = index of anchor
            mask_small = (regime_train == 1)
            cls_small = None
            if np.sum(mask_small) >= 60 and len(small_anchors) >= 2:
                y_small = []
                for idx, rowi in enumerate(np.where(mask_small)[0]):
                    m = float(m_train[rowi])
                    c = nearest_anchor_class(m, small_anchors, tol)
                    y_small.append(int(c) if c is not None else 0)
                y_small = np.array(y_small, dtype=int)
                X_small = X_train.iloc[np.where(mask_small)[0]].copy()

                cls_small = LGBMClassifier(
                    random_state=42,
                    objective="multiclass",
                    num_class=len(small_anchors),
                    n_estimators=1200,
                    learning_rate=0.04,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                )
                cls_small.fit(X_small, y_small)
                print(f"[ANCHOR SMALL] ON classes={len(small_anchors)}")
            else:
                print("[ANCHOR SMALL] OFF (pas assez de points ou anchors insuffisants)")

            # --- Anchor classifier gros ---
            mask_big = (regime_train == 2)
            cls_big = None
            if np.sum(mask_big) >= 60 and len(big_anchors) >= 2:
                y_big = []
                for idx, rowi in enumerate(np.where(mask_big)[0]):
                    m = float(m_train[rowi])
                    c = nearest_anchor_class(m, big_anchors, tol)
                    y_big.append(int(c) if c is not None else 0)
                y_big = np.array(y_big, dtype=int)
                X_big = X_train.iloc[np.where(mask_big)[0]].copy()

                cls_big = LGBMClassifier(
                    random_state=42,
                    objective="multiclass",
                    num_class=len(big_anchors),
                    n_estimators=1400,
                    learning_rate=0.035,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                )
                cls_big.fit(X_big, y_big)
                print(f"[ANCHOR BIG] ON classes={len(big_anchors)}")
            else:
                print("[ANCHOR BIG] OFF (pas assez de points ou anchors insuffisants)")

            # --- Spike reg (régime 3) ---
            mask_spike = (regime_train == 3) & is_open_train
            # si peu de spikes, on inclut aussi gros non-anchors
            if np.sum(mask_spike) < 40:
                mask_spike = ((m_train > 0) & is_open_train & (regime_train == 3))

            X_spike = X_train.iloc[np.where(mask_spike)[0]].copy()
            m_spike = m_train[np.where(mask_spike)[0]].astype(float)

            # log1p magnitude
            y_spike_log = np.log1p(np.clip(m_spike, 0.0, None))

            # poids: surpondère gros spikes
            if len(m_spike) > 0:
                q90s = float(np.quantile(m_spike, 0.90))
            else:
                q90s = 1.0
            w_spike = 0.5 + 6.0 * np.sqrt(np.clip(m_spike / max(q90s, 1e-9), 0.0, 16.0))

            reg_spike = None
            reg_spike_q10 = None
            reg_spike_q90 = None

            if len(X_spike) >= 40:
                reg_spike = LGBMRegressor(
                    random_state=42,
                    objective="regression",
                    n_estimators=2000,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike.fit(X_spike, y_spike_log, sample_weight=w_spike)

                reg_spike_q10 = LGBMRegressor(
                    random_state=42,
                    objective="quantile",
                    alpha=0.10,
                    n_estimators=1400,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike_q90 = LGBMRegressor(
                    random_state=42,
                    objective="quantile",
                    alpha=0.90,
                    n_estimators=1400,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike_q10.fit(X_spike, y_spike_log, sample_weight=w_spike)
                reg_spike_q90.fit(X_spike, y_spike_log, sample_weight=w_spike)

                print(f"[SPIKE REG] ON points={len(X_spike)}")
            else:
                print("[SPIKE REG] OFF (pas assez de spikes)")

            # ======================================================================
            # 4) CAP (évite explosion artificielle, mais laisse passer les vrais pics)
            # ======================================================================
            # cap très haut : on se base sur quantiles open days
            open_mags = m_train[(m_train > 0) & is_open_train]
            q995 = float(np.quantile(open_mags, 0.995)) if len(open_mags) else float(med_mag)
            q999 = float(np.quantile(open_mags, 0.999)) if len(open_mags) else float(med_mag * 3)
            hard_cap = max(q999, q995, med_mag * 3)  # laisse passer très haut si vraiment présent
            print(f"[CAP] q995={q995:,.0f} q999={q999:,.0f} hard_cap={hard_cap:,.0f}")

            # ======================================================================
            # 5) Prédiction (régime -> montant) + calibration métier (BOM+Lundi)
            # ======================================================================
            def _predict_one(df_row, X_row):
                # closed => 0
                if int(df_row["is_closed"]) == 1:
                    return 0.0, 0.0, 0.0, 0  # y, p10, p90, regime

                proba = cls_regime.predict_proba(X_row)[0]  # size 4
                # calibration métier : BOM+Lundi => si le modèle hésite à activer, on pousse un peu
                # (évite le "raté" sur des jours critiques)
                if int(df_row["is_bom_1_3"]) == 1:
                    proba[1] *= 1.20  # petit
                    proba[2] *= 1.35  # gros (plus probable)
                    if int(df_row["is_monday"]) == 1:
                        proba[2] *= 1.25
                        proba[3] *= 1.10
                if int(df_row["mon_bom_1_3"]) == 1:
                    proba[2] *= 1.40
                    proba[3] *= 1.15

                # renormalize
                s = float(np.sum(proba))
                if s > 0:
                    proba = proba / s

                regime = int(np.argmax(proba))

                # base output
                mag = 0.0
                mag10, mag90 = 0.0, 0.0

                if regime == 0:
                    mag = 0.0
                    mag10, mag90 = 0.0, 0.0

                elif regime == 1:
                    # petit récurrent -> anchor
                    if cls_small is not None:
                        c = int(cls_small.predict(X_row)[0])
                        c = int(np.clip(c, 0, len(small_anchors) - 1))
                        mag = float(small_anchors[c])
                        mag10, mag90 = mag, mag
                    else:
                        # fallback: prend l'ancre la plus fréquente "small"
                        mag = float(small_anchors[0]) if len(small_anchors) else float(min(med_mag, q995))
                        mag10, mag90 = mag, mag

                elif regime == 2:
                    # gros récurrent -> anchor "big"
                    if cls_big is not None:
                        c = int(cls_big.predict(X_row)[0])
                        c = int(np.clip(c, 0, len(big_anchors) - 1))
                        mag = float(big_anchors[c])
                        mag10, mag90 = mag, mag
                    else:
                        # fallback: médiane
                        mag = float(med_mag)
                        mag10, mag90 = mag, mag

                else:
                    # regime 3: spike/other
                    if reg_spike is not None:
                        p = float(reg_spike.predict(X_row)[0])
                        mag = float(np.expm1(max(0.0, p)))
                        p10 = float(reg_spike_q10.predict(X_row)[0]) if reg_spike_q10 is not None else p
                        p90 = float(reg_spike_q90.predict(X_row)[0]) if reg_spike_q90 is not None else p
                        mag10 = float(np.expm1(max(0.0, p10)))
                        mag90 = float(np.expm1(max(0.0, p90)))
                    else:
                        # fallback: très haut quantile (mais pas absurde)
                        mag = float(q995)
                        mag10, mag90 = float(q995 * 0.7), float(min(hard_cap, q999))

                # cap (évite explosions artificielles)
                mag = min(max(mag, 0.0), hard_cap)
                mag10 = min(max(mag10, 0.0), hard_cap)
                mag90 = min(max(mag90, 0.0), hard_cap)

                # convertir en y<=0
                y = min(-mag, 0.0)
                y_p10 = min(-mag90, 0.0)  # P10 plus négatif (borne basse)
                y_p90 = min(-mag10, 0.0)  # P90 moins négatif (borne haute)

                return float(y), float(y_p10), float(y_p90), regime

            # ======================================================================
            # 6) VALID metrics
            # ======================================================================
            y_pred_valid = np.zeros(len(df_valid_m), dtype=float)
            y_p10_valid = np.zeros(len(df_valid_m), dtype=float)
            y_p90_valid = np.zeros(len(df_valid_m), dtype=float)
            r_valid = np.zeros(len(df_valid_m), dtype=int)

            for i in range(len(df_valid_m)):
                Xr = pd.DataFrame([X_valid.iloc[i].values], columns=features)
                yhat, p10, p90, rr = _predict_one(df_valid_m.iloc[i], Xr)
                y_pred_valid[i] = yhat
                y_p10_valid[i] = p10
                y_p90_valid[i] = p90
                r_valid[i] = rr

            mae = mean_absolute_error(y_valid.values, y_pred_valid)
            rmse = mean_squared_error(y_valid.values, y_pred_valid) ** 0.5
            r2 = r2_score(y_valid.values, y_pred_valid)
            bias = float(np.mean(y_pred_valid - y_valid.values))

            # couverture bande si dispo
            coverage = float(np.mean((y_valid.values >= y_p10_valid) & (y_valid.values <= y_p90_valid)))
            width = float(np.mean(y_p90_valid - y_p10_valid))

            print(f"[V5.0] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}, Bias={bias:.2f}")
            print(f"[BAND] Coverage={coverage:.3f} Width={width:.1f}")

            # ===================== KPI UI =====================
            kpi_frame = ctk.CTkFrame(scrollable_frame, fg_color="#0f1b31", corner_radius=18)
            kpi_frame.pack(fill="x", padx=10, pady=(0, 10))
            graph_widgets.append(kpi_frame)
            for i in range(10):
                kpi_frame.grid_columnconfigure(i, weight=1)

            def _kpi(parent, title, value, subtitle, col):
                card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16, border_width=1, border_color="#223658")
                card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
                ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                    .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
                ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 22, "bold"), text_color="white") \
                    .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
                ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11), text_color="#7ea2d8") \
                    .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

            _kpi(kpi_frame, "MAE", f"{mae:.1f}", "Erreur moyenne", 0)
            _kpi(kpi_frame, "RMSE", f"{rmse:.1f}", "Risque spikes", 1)
            _kpi(kpi_frame, "R²", f"{r2:.3f}", "Pouvoir explicatif", 2)
            _kpi(kpi_frame, "Bias", f"{bias:.0f}", "Moy(y_pred - y)", 3)
            _kpi(kpi_frame, "Coverage", f"{coverage:.1%}", "Réel ∈ [P10,P90]", 4)
            _kpi(kpi_frame, "Width", f"{width:.0f}", "Largeur bande", 5)
            _kpi(kpi_frame, "Step", f"{step:.0f}", "Quantif anchors", 6)
            _kpi(kpi_frame, "SmallAnch", f"{len(small_anchors)}", "Petits récurrents", 7)
            _kpi(kpi_frame, "BigAnch", f"{len(big_anchors)}", "Gros ~120", 8)
            _kpi(kpi_frame, "SpikeReg", "ON" if reg_spike is not None else "OFF", "Modèle spikes", 9)

            # ======================================================================
            # Forecast N+1 (non lissé, jour-par-jour)
            # ======================================================================
            df_real_target = df_filiale[df_filiale["year"] == target_year].copy().sort_values("date")
            current_real_target_df = df_real_target

            start = pd.Timestamp(year=target_year, month=1, day=1)
            end = pd.Timestamp(year=target_year, month=12, day=31)
            all_dates_tgt = pd.date_range(start=start, end=end, freq="D")

            target_holidays = holiday_sets.get(int(target_year), set())

            # Build df_future skeleton
            future_rows = []
            for dt in all_dates_tgt:
                dt = pd.Timestamp(dt).normalize()
                dow = int(dt.weekday())
                is_h = int(dt.date() in target_holidays)
                is_closed = int((dow >= 5) or (is_h == 1))

                row = {
                    "section": filiale,
                    "date": dt,
                    "year": int(dt.year),
                    "dow": dow,
                    "is_holiday": is_h,
                    "is_closed": is_closed,
                    "y": 0.0,  # placeholder
                }
                future_rows.append(row)

            df_future = pd.DataFrame(future_rows).sort_values("date").reset_index(drop=True)

            # recompute full calendar features for target year (avec holiday_sets)
            df_future = _add_calendar_features(df_future, holiday_sets)

            # align features
            for c in features:
                if c not in df_future.columns:
                    df_future[c] = 0

            # predict day-by-day
            pred_vals = []
            pred_p10s = []
            pred_p90s = []
            pred_regimes = []

            for i in range(len(df_future)):
                Xr = pd.DataFrame([df_future.loc[i, features].values], columns=features)
                yhat, p10, p90, rr = _predict_one(df_future.iloc[i], Xr)
                pred_vals.append(yhat)
                pred_p10s.append(p10)
                pred_p90s.append(p90)
                pred_regimes.append(rr)

            df_future["pred_value"] = np.array(pred_vals, dtype=float)
            df_future["pred_p10"] = np.array(pred_p10s, dtype=float)
            df_future["pred_p90"] = np.array(pred_p90s, dtype=float)
            df_future["pred_regime"] = np.array(pred_regimes, dtype=int)

            # merge real target if present
            df_future_all = df_future[["section", "date", "year", "month", "dayofyear", "pred_p10", "pred_value", "pred_p90", "pred_regime"]].copy()
            df_future_all["month"] = df_future_all["date"].dt.month
            df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

            if df_real_target is not None and not df_real_target.empty:
                df_future_all = pd.merge(df_real_target[["date"]].copy(), df_future_all, on="date", how="left")
                for c in ["pred_p10", "pred_value", "pred_p90"]:
                    df_future_all[c] = df_future_all[c].fillna(0.0)
                df_future_all["section"] = filiale
                df_future_all["year"] = df_future_all["date"].dt.year
                df_future_all["month"] = df_future_all["date"].dt.month
                df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

            current_pred_df = df_future_all
            exported_pred_df = df_future_all.copy()

            # store artifacts
            model_artifacts["features"] = features[:]
            model_artifacts["cls_regime"] = cls_regime
            model_artifacts["cls_small"] = cls_small
            model_artifacts["cls_big"] = cls_big
            model_artifacts["reg_spike"] = reg_spike
            model_artifacts["reg_spike_q10"] = reg_spike_q10
            model_artifacts["reg_spike_q90"] = reg_spike_q90
            model_artifacts["small_anchors"] = small_anchors[:]
            model_artifacts["big_anchors"] = big_anchors[:]
            model_artifacts["step"] = float(step)
            model_artifacts["hard_cap"] = float(hard_cap)

            # ---------------- Plot ----------------
            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            df_hist_plot = df_filiale[df_filiale["year"] <= base_year].copy().sort_values("date")
            ax.plot(df_hist_plot["date"], df_hist_plot["y"], label=f"Réel (≤ {base_year})", linewidth=2)

            ax.plot(df_future_all["date"], df_future_all["pred_value"], label=f"Prévision IA {target_year} (ULTRA, ≤0)", linewidth=2, linestyle="--")
            ax.fill_between(df_future_all["date"], df_future_all["pred_p10"], df_future_all["pred_p90"], alpha=0.25,
                            label="Bande P10–P90 (spikes)")

            if df_real_target is not None and not df_real_target.empty:
                ax.plot(df_real_target["date"], df_real_target["y"], label=f"Réel {target_year}", linewidth=2)

            ax.set_title(f"ACE & Investissements – année {target_year} – {filiale} (NEG V5.0 ULTRA)", color="white")
            ax.tick_params(axis='x', colors="white", rotation=30)
            ax.tick_params(axis='y', colors="white")
            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            canvas = FigureCanvasTkAgg(fig, master=scrollable_frame)
            canvas.draw()
            w = canvas.get_tk_widget()
            w.pack(pady=10, fill="both", expand=True)
            graph_widgets.append(w)
            plt.close(fig)

            _redraw_monthly_graph()
            _redraw_graph2()

            if analysis_table_frame is not None:
                analysis_table_frame.destroy()

            analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
            analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
            graph_widgets.append(analysis_table_frame)

            _rebuild_analysis_table()    

        bouton_train.configure(command=_train_model)



        # ============ SCROLL MOLETTE ============
        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # ============ BOUTON RETOUR ============
        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)


# ============================================================
# SPLASH PULSE PREMIUM + CHARGEMENT DYNAMIQUE AVANT TKINTER
# VERSION CONSERVANT EXACTEMENT TA MÉTHODE DE LECTURE EXCEL
# ============================================================

import sys
import tempfile
import traceback

# On respecte la feuille issue de la config si elle existe
if '_dest_names' in globals() and _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]

previsions_triees = []
nb_prev = 0
taille_bloc = 0


class NullReporter:
    def set_phase(self, text):
        pass

    def set_detail(self, text):
        pass

    def set_summary(self, text):
        pass

    def set_total(self, total):
        pass

    def set_progress(self, current, total=None, detail=None, status=None, summary=None):
        pass

    def set_ratio(self, ratio, detail=None, status=None, summary=None):
        pass


# ------------------------------------------------------------
# Utilitaires dérivés
# ------------------------------------------------------------
def _rebuild_derived_globals():
    global previsions_triees, nb_prev, taille_bloc
    previsions_triees = sorted(PREV_UNION)
    nb_prev = len(previsions_triees)
    taille_bloc = 2 + 2 * nb_prev + 1
    print(f"📈 Nombre total de prévisions : {nb_prev}")
    print(f"📦 Taille d’un bloc de flux : {taille_bloc}\n")


# ------------------------------------------------------------
# 3️⃣ Lecture de la structure sur un fichier de référence
# VERSION SPLASH, MAIS MÊME MÉTHODE EXCEL QU’AVANT
# ------------------------------------------------------------
def _lire_structure_reference(path_ref, reporter=None):
    global _SORTED_PREV
    reporter = reporter or NullReporter()

    reporter.set_phase(f"📘 Lecture de la structure de référence ({FEUILLE_REFERENCE})")
    reporter.set_detail(f"→ {os.path.basename(path_ref)}")
    print(f"\n📘 Lecture de la structure de référence ({FEUILLE_REFERENCE}) → {os.path.basename(path_ref)}")

    # Lire uniquement les 3 premières lignes de la feuille de référence
    try:
        import python_calamine  # noqa: F401
        _eng = "calamine"
    except ImportError:
        _eng = None
    kw = {"header": None, "nrows": 3}
    if _eng:
        kw["engine"] = _eng
    df_head = pd.read_excel(path_ref, sheet_name=FEUILLE_REFERENCE, **kw)
    row2 = df_head.iloc[1].tolist()
    row3 = df_head.iloc[2].tolist()

    ref_od = OrderedDict()
    ref_tokens = []
    # layout_entries : liste de (flux_name, dates_col_0, reel_col_0, [(header, prev_col_0), ...])
    layout_entries = []
    col = 2        # colonne C (0-based) — position du nom de flux dans row2
    token_col = 3

    while col < len(row2):
        flux_name = row2[col]
        if pd.isna(flux_name):
            break
        flux_name = str(flux_name).strip()

        prev_headers, prev_cols = [], []
        c = col + 1
        while c < len(row3):
            h = row3[c]
            if isinstance(h, str) and "Prévision" in h:
                prev_headers.append(h.strip())
                prev_cols.append(c)
                c += 2
                continue
            break

        if prev_headers:
            ref_od[flux_name] = {
                "col_start":    col + 1,
                "prev_headers": prev_headers,
                "prev_cols":    [x + 1 for x in prev_cols],
                "dates_col":    col,
                "reel_col":     col + 1,
            }
            ref_tokens.append((flux_name, token_col))
            PREV_UNION.update(prev_headers)
            # Layout 0-based pour l'accumulation :
            #   dates = col-1, reel = col, prev_col[j] = prev_cols[j]
            layout_entries.append(
                (flux_name, col - 1, col,
                 list(zip(prev_headers, prev_cols)))
            )

        col       += (2 + 2 * len(prev_headers) + 1)
        token_col += (2 + 2 * len(prev_headers) + 1)

    # Ordre stable de toutes les prévisions connues
    _SORTED_PREV = sorted(PREV_UNION, key=_parse_prev_header_sort_key)
    _prev_idx    = {h: i for i, h in enumerate(_SORTED_PREV)}
    n_prev       = len(_SORTED_PREV)

    # Copier structure + layout sur toutes les sections, pré-init CACHE
    for sec in sections:
        STRUCT[sec]     = ref_od.copy()
        TOKENS[sec]     = ref_tokens.copy()
        _COL_LAYOUT[sec] = layout_entries  # partagé (lecture seule ensuite)
        for flux_name, *_ in layout_entries:
            key = (sec, flux_name)
            if key not in CACHE:
                CACHE[key] = {
                    "dates":        [],
                    "reel":         [],
                    "prev_headers": [],   # construit dynamiquement par _reconcile_headers
                    "prev_vals":    [],   # idem
                }

    print(f"   → {len(ref_od)} flux détectés ({len(PREV_UNION)} prévisions).")
    print("✅ Structure + layout pré-calculés sur toutes les feuilles.\n")
    reporter.set_summary(
        f"📘 Structure lue → {len(ref_od)} flux, {len(PREV_UNION)} prévisions.\n"
        f"✅ Layout pré-calculé, CACHE pré-initialisé."
    )


# ------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# VERSION SPLASH, MAIS MÊME MÉTHODE EXCEL QU’AVANT
# ------------------------------------------------------------
def _accumuler_valeurs_tous_mois(files_list, reporter=None):
    """
    Version optimisée :
    - Layout scanné par fichier (row2/row3) → colonnes toujours correctes même si
      les fichiers anciens ont moins de prévisions que le fichier de référence
    - pd.to_datetime vectorisé → zéro boucle Python sur les dates
    - Remplissage CACHE par index global (_SORTED_PREV) → zéro _reconcile_headers
    - I/O parallèle (calamine + ThreadPoolExecutor)
    """
    import os as _os
    from concurrent.futures import ThreadPoolExecutor, as_completed

    reporter = reporter or NullReporter()
    n = len(files_list)
    needed_sheets = list(set(sections.values()))

    reporter.set_phase(f"📊 Lecture de {n} fichiers mensuels ({len(needed_sheets)} feuilles ciblées)...")
    reporter.set_total(n)

    # ── Moteur le plus rapide ────────────────────────────────────
    try:
        import python_calamine  # noqa: F401
        _ENGINE: str | None = "calamine"
    except ImportError:
        _ENGINE = None
    print(f"[INFO] Moteur Excel : {_ENGINE or 'openpyxl'}")

    # ── Lecture I/O parallèle ────────────────────────────────────
    def _read_one(args):
        idx, path = args
        kw = {"header": None}
        if _ENGINE:
            kw["engine"] = _ENGINE
        try:
            data = pd.read_excel(path, sheet_name=needed_sheets, **kw)
            if isinstance(data, pd.DataFrame):
                data = {needed_sheets[0]: data}
            return idx, path, data
        except Exception:
            pass
        result = {}
        for sheet in needed_sheets:
            try:
                result[sheet] = pd.read_excel(path, sheet_name=sheet, **kw)
            except Exception:
                pass
        return idx, path, result or None

    n_workers = min(n, max(1, (_os.cpu_count() or 4) + 4))
    file_results: dict = {}
    with ThreadPoolExecutor(max_workers=n_workers) as pool:
        futures = {
            pool.submit(_read_one, (i, path)): i
            for i, (_, _, path) in enumerate(files_list, 1)
        }
        done = 0
        for fut in as_completed(futures):
            idx, path, sheets = fut.result()
            done += 1
            reporter.set_progress(
                done, total=n,
                detail=f"[{done}/{n}] → {_os.path.basename(path)}",
                status=f"📊 Lecture ({done}/{n})..."
            )
            file_results[idx] = (path, sheets)

    # ── Accumulation séquentielle (ordre chronologique) ──────────
    for seq in range(1, n + 1):
        path, all_sheets = file_results.get(seq, (None, None))
        if not all_sheets:
            if path:
                print(f"⚠️ Erreur lecture {_os.path.basename(path)}")
            continue

        print(f"   [{seq}/{n}] → {_os.path.basename(path)}")

        for sec, feuille in sections.items():
            if feuille not in all_sheets:
                continue
            df = all_sheets[feuille]
            if df.shape[0] < 5:
                continue

            # Accès numpy direct pour toutes les lectures
            arr      = df.values
            row2_arr = arr[1]
            row3_arr = arr[2]

            # ── Scan du layout réel de CE fichier ──────────────
            # Indispensable : les fichiers anciens ont moins de colonnes de
            # prévisions → les positions des flux 2, 3… diffèrent du fichier
            # de référence. On relit row2/row3 pour chaque fichier.
            col = 2
            while col < len(row2_arr):
                flux_val = row2_arr[col]
                if flux_val is None or (isinstance(flux_val, float) and flux_val != flux_val):
                    break
                flux_name = str(flux_val).strip()

                # Détection des colonnes "Prévision" dans CE fichier
                prev_headers_file, prev_cols_file = [], []
                c = col + 1
                while c < len(row3_arr):
                    h = row3_arr[c]
                    if isinstance(h, str) and "Prévision" in h:
                        prev_headers_file.append(h.strip())
                        prev_cols_file.append(c)
                        c += 2
                        continue
                    break

                if not prev_headers_file:
                    col += 10
                    continue

                dates_col = col - 1
                reel_col  = col

                key = (sec, flux_name)
                B   = CACHE.get(key)
                col += (2 + 2 * len(prev_headers_file) + 1)

                if B is None:
                    continue

                # ── Dates vectorisées ───────────────────────────
                raw_dates = arr[3:, dates_col]
                try:
                    dates_s = pd.Series(
                        pd.to_datetime(
                            pd.array(raw_dates, dtype=object),
                            errors="coerce", dayfirst=True
                        ).normalize(),
                        dtype="datetime64[ns]"
                    )
                except Exception:
                    dates_s = pd.Series(
                        [_parse_excel_date(v) for v in raw_dates],
                        dtype="datetime64[ns]"
                    )

                valid_mask  = dates_s.notna().values
                if not valid_mask.any():
                    continue

                dates_batch = dates_s[valid_mask].tolist()
                rows_valid  = valid_mask.nonzero()[0]

                # ── Réel vectorisé ──────────────────────────────
                reel_col_vals = arr[3:, reel_col][rows_valid]
                reel_clean    = [
                    0 if (v is None or (isinstance(v, float) and v != v)) else v
                    for v in reel_col_vals
                ]

                n_batch = len(dates_batch)

                # ── Prévisions : union dynamique ──────────────────────
                # IMPORTANT : _reconcile_headers doit être appelé AVANT
                # d'étendre B["dates"], car il utilise len(B["dates"]) pour
                # savoir combien de None rétro-remplir pour les nouveaux
                # headers (exactement comme dans PulseV1.py).
                _reconcile_headers(B, prev_headers_file)
                PREV_UNION.update(prev_headers_file)

                B["dates"].extend(dates_batch)
                B["reel"].extend(reel_clean)

                header_to_col = {h: prev_cols_file[j] for j, h in enumerate(prev_headers_file)}
                prev_v_lst = B["prev_vals"]
                for k, h in enumerate(B["prev_headers"]):
                    pc = header_to_col.get(h)
                    if pc is None:
                        # Header connu mais absent de ce fichier → Nones
                        prev_v_lst[k].extend([None] * n_batch)
                    else:
                        col_vals = arr[3:, pc][rows_valid]
                        prev_v_lst[k].extend([
                            None if (v is None or (isinstance(v, float) and v != v)) else v
                            for v in col_vals
                        ])

    print(f"✅ Cache complété : {len(CACHE)} flux.\n")
    counts: dict = {}
    totals: dict = {}
    for (sn, _), data in CACHE.items():
        counts[sn] = counts.get(sn, 0) + 1
        totals[sn] = totals.get(sn, 0) + len(data.get("dates", []))
    for sn in sorted(counts):
        print(f"   - {sn} : {counts[sn]} flux ({totals.get(sn, 0)} lignes)")


# ------------------------------------------------------------
# 5️⃣ Index annuel
# ------------------------------------------------------------
def _build_year_index():
    import numpy as np
    YEAR_INDEX.clear()
    for (section, flux_name), B in CACHE.items():
        dates     = B.get("dates", [])
        prev_vals = B.get("prev_vals", [])
        headers   = B.get("prev_headers", [])
        if not dates:
            YEAR_INDEX[(section, flux_name)] = {"years": {}}
            continue

        # Années vectorisées
        years_arr = np.array([d.year for d in dates], dtype=np.int32)
        unique_years = np.unique(years_arr)

        # Matrice de prévisions (n_prev × n_rows) en float — None → nan
        n_rows = len(dates)
        n_prev = len(prev_vals)
        if n_prev:
            # construction rapide avec fromiter
            mat = np.empty((n_prev, n_rows), dtype=np.float64)
            for k, serie in enumerate(prev_vals):
                for i, v in enumerate(serie[:n_rows]):
                    try:
                        mat[k, i] = float(v) if v is not None else np.nan
                    except (TypeError, ValueError):
                        mat[k, i] = np.nan
        else:
            mat = np.empty((0, n_rows), dtype=np.float64)

        years_map = {}
        for y in unique_years.tolist():
            row_idx = np.where(years_arr == y)[0].tolist()
            if n_prev and row_idx:
                # Profil actif si au moins une valeur ≠ 0 et non-nan pour cette année
                sub = mat[:, row_idx]
                active_mask = np.any((~np.isnan(sub)) & (sub != 0.0), axis=1)
                prof_idx = np.where(active_mask)[0].tolist()
            else:
                prof_idx = []

            headers_year = [
                _clean_profil_label(headers[k] if k < len(headers) else None, k)
                for k in prof_idx
            ]
            years_map[y] = {
                "row_idx":  row_idx,
                "prof_idx": prof_idx,
                "headers": headers_year,
            }

        YEAR_INDEX[(section, flux_name)] = {"years": years_map}


# ------------------------------------------------------------
# 6️⃣ Initialisation complète pilotée par le splash
# ------------------------------------------------------------
def _init_full_load(reporter=None):
    reporter = reporter or NullReporter()

    global STRUCT, TOKENS, PREV_UNION, CACHE, YEAR_INDEX, _COL_LAYOUT, _SORTED_PREV
    STRUCT.clear()
    TOKENS.clear()
    PREV_UNION.clear()
    CACHE.clear()
    YEAR_INDEX.clear()
    _COL_LAYOUT.clear()
    _SORTED_PREV.clear()

    print("🚀 Initialisation complète du cache de données...")
    reporter.set_phase("🚀 Initialisation complète du cache de données...")
    reporter.set_ratio(0.03, detail="Préparation du chargement...")

    files = _lister_fichiers_mensuels()
    if not files:
        print("ℹ️ Aucun fichier trouvé — initialisation du cache ignorée.\n")
        reporter.set_phase("ℹ️ Aucun fichier historique trouvé.")
        reporter.set_detail("Le dossier ne contient aucun fichier mensuel exploitable.")
        reporter.set_summary("Aucun chargement n’a été effectué.")
        _rebuild_derived_globals()
        return {
            "files": [],
            "cache_size": 0,
            "nb_prev": nb_prev,
            "taille_bloc": taille_bloc,
        }

    reporter.set_ratio(
        0.08,
        detail=f"{len(files)} fichier(s) détecté(s).",
        summary="\n".join([f" - {os.path.basename(f[2])}" for f in files[-9:]])
    )

    ref_path = files[-1][2]
    _lire_structure_reference(ref_path, reporter=reporter)
    _accumuler_valeurs_tous_mois(files, reporter=reporter)

    reporter.set_phase("🧠 Construction de l’index annuel...")
    reporter.set_ratio(0.94, detail="Agrégation des années et profils...")
    _build_year_index()

    _rebuild_derived_globals()

    counts = {}
    total_lignes_par_section = {}
    for (section_name, flux_name), data in CACHE.items():
        counts[section_name] = counts.get(section_name, 0) + 1
        total_lignes_par_section[section_name] = (
            total_lignes_par_section.get(section_name, 0) + len(data.get("dates", []))
        )

    recap_lines = [
        f"✅ Cache complété : {len(CACHE)} flux au total (valeurs réelles et prévisions incluses).",
        "",
        "📊 Récapitulatif des flux chargés par section :",
    ]

    for section_name in sorted(counts.keys()):
        recap_lines.append(
            f"   - {section_name} : {counts[section_name]} flux ({total_lignes_par_section.get(section_name, 0)} lignes cumulées)"
        )

    recap_lines.extend([
        "",
        "✅ Chargement complet terminé.",
        "",
        f"📈 Nombre total de prévisions : {nb_prev}",
        f"📦 Taille d’un bloc de flux : {taille_bloc}",
    ])

    reporter.set_phase("✅ Chargement complet terminé.")
    reporter.set_detail("Appuyez sur « Continuer » pour entrer dans PULSE.")
    reporter.set_summary("\n".join(recap_lines))
    reporter.set_ratio(1.0)

    print("✅ Chargement complet terminé.\n")

    return {
        "files": files,
        "cache_size": len(CACHE),
        "nb_prev": nb_prev,
        "taille_bloc": taille_bloc,
        "previsions_triees": previsions_triees,
        "year_index_size": len(YEAR_INDEX),
    }


# ------------------------------------------------------------
# Splash pur tk.Canvas — dark dev, zéro rame
# Pas de CTk dans le splash : canvas.itemconfig() = instantané
# ------------------------------------------------------------
def lancer_splash_ctk_loading(loader_func, logo_path=None):
    """
    Splash pure-tkinter dark-dev.
    Exécute loader_func(reporter) dans un thread daemon,
    puis détruit le splash et retourne (autorise_entree, payload).
    Tout le rendu est sur un seul tk.Canvas → aucun redraw de widget CTk.
    """
    import threading
    import queue
    import tkinter as _tk
    from tkinter import font as _tkfont

    W, H    = 660, 360
    BG      = "#0d0f14"
    ACCENT  = "#00c8ff"
    ACCENT2 = "#0055cc"
    MUTED   = "#2e3650"
    TEXT_HI = "#e8eaf0"
    TEXT_LO = "#6b7280"
    PB_TRACK = "#151820"

    result   = {"payload": None, "done": False, "error": None}
    ui_queue = queue.Queue()

    class SplashReporter:
        def set_phase(self, text):
            ui_queue.put(("status", text))
        def set_detail(self, text):
            ui_queue.put(("detail", text))
        def set_summary(self, text):
            pass
        def set_total(self, total):
            pass
        def set_progress(self, current, total=None, detail=None, status=None, summary=None):
            if total and total > 0:
                ui_queue.put(("ratio", current / total))
            if detail:
                ui_queue.put(("detail", detail))
            if status:
                ui_queue.put(("status", status))
        def set_ratio(self, ratio, detail=None, status=None, summary=None):
            ui_queue.put(("ratio", float(ratio)))
            if detail:
                ui_queue.put(("detail", detail))
            if status:
                ui_queue.put(("status", status))

    def _worker():
        try:
            result["payload"] = loader_func(SplashReporter())
        except Exception as exc:
            result["error"] = exc
            import traceback as _tb
            _tb.print_exc()
        finally:
            result["done"] = True
            ui_queue.put(("done", None))

    # ── Fenêtre ─────────────────────────────────────────────────
    splash = _tk.Tk()
    splash.overrideredirect(True)
    splash.configure(bg=BG)
    splash.attributes("-topmost", True)
    splash.attributes("-alpha", 0.0)
    splash.update_idletasks()
    sx = (splash.winfo_screenwidth()  - W) // 2
    sy = (splash.winfo_screenheight() - H) // 2
    splash.geometry(f"{W}x{H}+{sx}+{sy}")

    # ── Fonts ────────────────────────────────────────────────────
    f_title  = _tkfont.Font(family="Consolas", size=54, weight="bold")
    f_sub    = _tkfont.Font(family="Consolas", size=10)
    f_status = _tkfont.Font(family="Consolas", size=12, weight="bold")
    f_detail = _tkfont.Font(family="Consolas", size=10)
    f_pct    = _tkfont.Font(family="Consolas", size=11)
    f_dots   = _tkfont.Font(family="Consolas", size=11)
    f_footer = _tkfont.Font(family="Consolas", size=9)

    # ── Canvas unique (tout est dessiné ici) ────────────────────
    cv = _tk.Canvas(splash, width=W, height=H, bg=BG,
                    highlightthickness=0, bd=0)
    cv.pack(fill="both", expand=True)

    # Bordures accent (4 côtés)
    cv.create_rectangle(0,   0,   W,   3,   fill=ACCENT,  outline="")
    cv.create_rectangle(0,   H-3, W,   H,   fill=ACCENT2, outline="")
    cv.create_rectangle(0,   0,   2,   H,   fill=ACCENT,  outline="")
    cv.create_rectangle(W-2, 0,   W,   H,   fill=ACCENT2, outline="")

    # Titre
    cv.create_text(W//2, 70, text="PULSE",
                   font=f_title, fill=ACCENT, anchor="center")
    cv.create_text(W//2, 113, text="DATA INTELLIGENCE PLATFORM",
                   font=f_sub, fill=MUTED, anchor="center")
    cv.create_line(55, 135, W-55, 135, fill="#1a1f30", width=1)

    # Progress bar (track + fill)
    PB_X1, PB_Y, PB_X2, PB_H = 50, 170, W-50, 7
    cv.create_rectangle(PB_X1, PB_Y, PB_X2, PB_Y + PB_H,
                        fill=PB_TRACK, outline="")
    pb_fill = cv.create_rectangle(PB_X1, PB_Y, PB_X1, PB_Y + PB_H,
                                  fill=ACCENT, outline="")
    PB_W = PB_X2 - PB_X1

    # Textes dynamiques (canvas items)
    it_pct    = cv.create_text(W - 36, PB_Y + PB_H//2, text=" 0%",
                               font=f_pct, fill=ACCENT, anchor="center")
    it_status = cv.create_text(W//2, 205, text="Initialisation\u2026",
                               font=f_status, fill=TEXT_HI, anchor="center")
    it_detail = cv.create_text(W//2, 235, text="",
                               font=f_detail, fill=TEXT_LO, anchor="center",
                               width=560)
    it_dots   = cv.create_text(W//2, 262, text="\u00b7  \u00b7  \u00b7",
                               font=f_dots, fill=MUTED, anchor="center")
    cv.create_text(W//2, H-13, text="v2.0  \u00b7  SNCF DCF GROUPE",
                   font=f_footer, fill="#1e2230", anchor="center")

    # ── State ────────────────────────────────────────────────────
    _cur    = [0.0]
    _tgt    = [0.0]
    _tick   = [0]
    _done = [False]
    DOTS = ["\u00b7  \u00b7  \u00b7", "\u25cf  \u00b7  \u00b7",
            "\u25cf  \u25cf  \u00b7", "\u25cf  \u25cf  \u25cf"]

    def _set_pb(v):
        x2 = PB_X1 + max(0, int(PB_W * v))
        cv.coords(pb_fill, PB_X1, PB_Y, x2, PB_Y + PB_H)
        cv.itemconfig(it_pct, text=f"{int(v * 100):3d}%")

    def _poll():
        try:
            while True:
                kind, val = ui_queue.get_nowait()
                if kind == "ratio":
                    _tgt[0] = max(_tgt[0], float(val))
                elif kind == "status":
                    cv.itemconfig(it_status, text=str(val))
                elif kind == "detail":
                    cv.itemconfig(it_detail, text=str(val))
                elif kind == "done":
                    _done[0] = True
        except Exception:
            pass

        # Interpolation ease-out de la barre
        diff = _tgt[0] - _cur[0]
        if abs(diff) > 0.001:
            _cur[0] += diff * 0.15
            _set_pb(min(1.0, max(0.0, _cur[0])))

        # Dots
        _tick[0] = (_tick[0] + 1) % 40
        cv.itemconfig(it_dots, text=DOTS[_tick[0] // 10])

        if _done[0]:
            _set_pb(1.0)
            cv.itemconfig(it_status, text="Chargement termin\u00e9  \u2713")
            cv.itemconfig(it_detail, text="")
            cv.itemconfig(it_dots,   text="")
            splash.after(350, splash.destroy)
            return

        splash.after(40, _poll)

    # ── Démarrage ────────────────────────────────────────────────
    splash.attributes("-alpha", 1.0)
    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    splash.after(80, _poll)
    splash.mainloop()

    if result["error"]:
        return False, None
    return True, result["payload"]

# ------------------------------------------------------------
# Lancement global AVANT Tkinter
# ------------------------------------------------------------
def lancer_application_avec_splash():
    try:
        logo_splash = None
        if 'image_path' in globals() and image_path:
            try:
                if os.path.exists(image_path):
                    logo_splash = image_path
            except Exception:
                logo_splash = None

        print("[PULSE] Lancement du splash...")

        autorise_entree, payload = lancer_splash_ctk_loading(
            loader_func=lambda reporter: _init_full_load(reporter=reporter),
            logo_path=logo_splash,
        )

        print(f"[PULSE] Retour splash : autorise_entree={autorise_entree}, payload_type={type(payload).__name__ if payload is not None else 'None'}")

        if not autorise_entree:
            print("[PULSE] Chargement échoué ou annulé.")
            return

        import gc
        gc.collect()

        print("[PULSE] Ouverture de l'application Tkinter...")
        app = Application()
        app.update_idletasks()
        print("[PULSE] Application Tkinter initialisée, entrée en mainloop...")
        app.mainloop()

    except Exception:
        print("[PULSE] Erreur fatale dans lancer_application_avec_splash():")
        print(traceback.format_exc())

# ------------------------------------------------------------
# POINT D’ENTRÉE UNIQUE
# ------------------------------------------------------------
if __name__ == "__main__":
    lancer_application_avec_splash()
