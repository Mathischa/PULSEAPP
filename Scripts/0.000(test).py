# === Interface graphique (Tkinter) ===
from ast import Continue
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

# === Manipulation et analyse de données ===
import pandas as pd

# === Visualisation (Matplotlib & Seaborn) ===
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation
import seaborn as sns

# === Gestion des fichiers Excel (OpenPyXL) ===
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict
import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# -*- coding: utf-8 -*-
from openpyxl.utils import get_column_letter  # (si nécessaire ailleurs)

import os
import re
import pandas as pd
from datetime import datetime, date
from collections import OrderedDict

# === Répertoire contenant les fichiers mensuels ===
from pathlib import Path
import os
import os, time, subprocess, unicodedata
from pathlib import Path
import winreg



from pathlib import Path
import os
import unicodedata
# === Modèle ML (scikit-learn) ===
try:
    from sklearn.ensemble import RandomForestRegressor
except ImportError:
    RandomForestRegressor = None
from sklearn.ensemble import RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor

USER_ID = Path.home().name
BASE_DIR = Path(fr"C:\Users\{USER_ID}\SNCF")
import tkinter as tk

# --- PATCH pour empêcher les RuntimeError "main thread is not in main loop" ---

# Sauvegarde des __del__ originaux
_orig_image_del = getattr(tk.Image, "__del__", None)
_orig_var_del = getattr(tk.Variable, "__del__", None)

def _safe_image_del(self):
    if _orig_image_del is None:
        return
    try:
        _orig_image_del(self)
    except RuntimeError:
        # On ignore seulement le cas où Tk n'est plus dans la boucle principale
        pass

def _safe_var_del(self):
    if _orig_var_del is None:
        return
    try:
        _orig_var_del(self)
    except RuntimeError:
        # Pareil ici : on ne fait rien si Tk est déjà "mort"
        pass

if _orig_image_del is not None:
    tk.Image.__del__ = _safe_image_del

if _orig_var_del is not None:
    tk.Variable.__del__ = _safe_var_del

# On n’impose QUE la "queue" du chemin (fin de chemin), flexible sur tout le reste.
REQUIRED_TAIL = ["Partage - Invités", "Projet PULSE", "4. Données historiques", "Développement","Données"]

def _norm(s: str) -> str:
    """Normalise pour comparaisons: minuscules, accents retirés, espaces/underscores compressés."""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower().replace("_", " ").strip()
    s = " ".join(s.split())
    return s

def _match_tail(path: Path, tail: list[str]) -> bool:
    """Vérifie que 'path' se termine par les éléments de 'tail' (nom pour nom), comparés normalisés."""
    parts = list(path.parts)
    if len(parts) < len(tail):
        return False
    # Compare en partant de la fin
    for i in range(1, len(tail)+1):
        if _norm(parts[-i]) != _norm(tail[-i]):
            return False
    return True

def find_sharepoint_base():
    """
    1) Si ton chemin 'par défaut' existe, on l’utilise.
    2) Sinon, on parcourt TOUT 'C:\\Users\\<USER_ID>\\SNCF' pour trouver un dossier
       qui SE TERMINE par 'Partage - Invités/Projet PULSE/4. Données historiques/Développement'.
    """
    # 1) Essai direct sur ton chemin "canonique"
    default_path = BASE_DIR / "DCF GROUPE (Grp. O365) GrpO365 - Reporting et prévisions" / Path(*REQUIRED_TAIL)
    if default_path.exists():
        print(f"[OK] Bibliothèque détectée (DCF par défaut) : {default_path}")
        return default_path

    # 2) Scan générique : on ne regarde QUE la fin de chemin (tail), robustes aux variations amont
    print("[INFO] Recherche générique de    la queue de chemin '.../Partage - Invités/Projet PULSE/4. Données historiques/Développement'...")
    candidates = []
    for root, dirs, _ in os.walk(BASE_DIR):
        # accélère: on ne teste que si "Développement" est présent dans le dossier courant
        if any(_norm(d) == _norm(REQUIRED_TAIL[-1]) for d in dirs):
            for d in dirs:
                full = Path(root) / d
                if _match_tail(full, REQUIRED_TAIL):
                    candidates.append(full)

    if candidates:
        # Choix simple : on prend le plus long chemin (le plus profond) ou simplement le premier trouvé
        best = sorted(candidates, key=lambda p: len(p.parts), reverse=True)[0]
        print(f"[OK] Bibliothèque trouvée automatiquement : {best}")
        return best

    print("❌ Aucune bibliothèque SharePoint locale trouvée correspondant à la queue requise.")
    return None

# --- Résolution finale (inchangé) ---
DEV_PATH = find_sharepoint_base()
if not DEV_PATH:
    raise FileNotFoundError("Impossible de localiser la bibliothèque SharePoint (vérifie la synchro OneDrive/SharePoint).")

FICHIER_EXCEL_DIR = str(DEV_PATH / "Historique Prévisions Réel Filiales")
FICHIER_CONFIG_SECTIONS = str(DEV_PATH / "Filiales Analysées.xlsx")
image_path = str(DEV_PATH / "Images" / "logo_Pulse.png")
BASE_DONNEES_DIR = str(DEV_PATH / "Données Réelles")


print("\n[INIT] Dossier Résultats :", FICHIER_EXCEL_DIR)
print("[INIT] Fichier Config    :", FICHIER_CONFIG_SECTIONS)
print("[INIT] Image            :", image_path)
print("[INIT] Dossier Données   :\n", BASE_DONNEES_DIR)

if not Path(FICHIER_EXCEL_DIR).exists():
    print(f"⚠️  Dossier 'Résultats' introuvable : {FICHIER_EXCEL_DIR}")
if not Path(FICHIER_CONFIG_SECTIONS).exists():
    print(f"⚠️  Fichier 'Filiales Analysées.xlsx' introuvable : {FICHIER_CONFIG_SECTIONS}")
if not Path(image_path).exists():
    print(f"⚠️  Image 'logo_Pulse.png' introuvable : {image_path}")

# ======================
#   FONCTIONS DE SÉCURISATION
# ======================

def _longpath(p: str) -> str:
    """Ajoute le préfixe \\?\ pour les chemins trop longs sous Windows."""
    p = str(Path(p))
    if p.startswith("\\\\?\\"):
        return p
    return "\\\\?\\" + p.replace("/", "\\")

def _nfc(p: str) -> str:
    """Normalise les accents dans le chemin (utile avec SharePoint)."""
    return unicodedata.normalize("NFC", p)

def _attrib_flags(path: str) -> str | None:
    """Retourne les attributs Windows du fichier (ex: O=Offline, P=Pinned)."""
    try:
        cp = subprocess.run(["attrib", path], capture_output=True, text=True, shell=True)
        if cp.returncode == 0 and cp.stdout:
            return cp.stdout.strip().splitlines()[-1]
    except Exception:
        pass
    return None

def is_cloud_only(path: str) -> bool:
    """Détecte si le fichier est en mode 'cloud-only' (non téléchargé localement)."""
    flags = _attrib_flags(path)
    if not flags:
        return False
    left = flags.split(path)[0] if path in flags else flags
    # O = Offline placeholder, P = Pinned (local)
    return (" O " in f" {left} ") or left.strip().endswith("O")

def pin_and_hydrate(path: str, timeout=30) -> bool:
    """Force la dispo locale du fichier cloud-only (OneDrive) et attend l’hydratation."""
    try:
        subprocess.run(f'attrib +P -U "{path}"', shell=True, check=False)
    except Exception:
        pass
    start = time.time()
    while time.time() - start < timeout:
        try:
            if os.path.getsize(path) > 0:
                return True
        except Exception:
            pass
        time.sleep(0.5)
    return False

def robust_load_workbook(path: str, retries=3, delay=1.0):
    """
    Ouvre un fichier Excel de façon robuste :
    - Support des chemins longs (\\?\\)
    - Gestion cloud-only OneDrive/SharePoint
    - Retry si sync en cours
    """
    if not path:
        raise FileNotFoundError("Chemin vide.")
    path_n = _nfc(os.path.normpath(path))
    path_lp = _longpath(path_n) if len(path_n) > 240 else path_n

    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if is_cloud_only(path_n):
                print(f"[INFO] Fichier cloud-only détecté → hydratation : {path_n}")
                pin_and_hydrate(path_n)

            if not os.path.exists(path_n) and not os.path.exists(path_lp):
                raise FileNotFoundError(f"Fichier introuvable (même après hydratation) : {path_n}")

            return load_workbook(path_lp, read_only=True, data_only=True)

        except Exception as e:
            last_err = e
            print(f"[WARN] Tentative {attempt}/{retries} échouée pour {path_n} ({type(e).__name__}: {e})")
            if attempt < retries:
                time.sleep(delay)
                continue
            raise FileNotFoundError(
                f"Échec d'ouverture finale : {path_n}\n"
                f" - LongPath: {path_lp}\n"
                f" - CloudOnly: {is_cloud_only(path_n)}\n"
                f" - Détails: {type(e).__name__}: {e}"
            ) from e

def diag_path(path: str):
    """Affiche un diagnostic complet sur un fichier problématique."""
    p = Path(path)
    print("=== DIAGNOSTIC PATH ===")
    print("Path :", path)
    print("Existe :", p.exists())
    print("Taille :", p.stat().st_size if p.exists() else "N/A")
    print("Attrib :", _attrib_flags(path))
    print("CloudOnly :", is_cloud_only(path))
    print("Long (>240) :", len(str(path)) > 240)
    print("Chemin long (\\?\\) :", _longpath(path))
    print("========================\n")
    
# === Fichier Excel contenant la configuration des sections ===
FEUILLE_CONFIG_SECTIONS = "CONFIG_SECTIONS"  # ou le vrai nom de la feuille dans ton fichier
COL_DEST = 1   # A
COL_SOURCE = 2 # B
COL_PREV = 3   # C

def charger_sections_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit le fichier Excel de configuration des sections et renvoie une liste
    de dictionnaires {"source": ..., "prev": ..., "dest": ...}.
    """
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return []

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return []

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return []

    ws = wb[FEUILLE_CONFIG_SECTIONS]
    mapping = []

    row = 2  # ligne 1 = en-têtes
    while True:
        dest_val = ws.cell(row=row, column=COL_DEST).value
        source_val = ws.cell(row=row, column=COL_SOURCE).value
        prev_val = ws.cell(row=row, column=COL_PREV).value

        if dest_val is None or str(dest_val).strip() == "":
            break  # fin du tableau

        mapping.append({
            "dest": str(dest_val).strip(),
            "source": str(source_val).strip() if source_val else str(dest_val).strip(),
            "prev": str(prev_val).strip() if prev_val else str(dest_val).strip(),
        })
        row += 1

    wb.close()

    print("[INFO] Sections chargées depuis 'Filiales Analysées.xlsx' :")
    for m in mapping:
        print(f"   - dest='{m['dest']}' | source='{m['source']}' | prev='{m['prev']}'")

    return mapping

def charger_noms_feuilles_depuis_cells(fichier_config=FICHIER_CONFIG_SECTIONS):
    """
    Lit uniquement la colonne A (dest) de la feuille de config et
    retourne la liste des noms de feuilles à utiliser (strings non vides, uniques, ordonnées).
    """
    noms = []
    if not os.path.isfile(fichier_config):
        print(f"[WARN] Fichier de configuration introuvable : {fichier_config}")
        return noms

    try:
        wb = robust_load_workbook(fichier_config)
    except Exception as e:
        print(f"[ERROR] Impossible d'ouvrir {fichier_config} : {e}")
        diag_path(fichier_config)
        return noms

    if FEUILLE_CONFIG_SECTIONS not in wb.sheetnames:
        print(f"[WARN] Feuille '{FEUILLE_CONFIG_SECTIONS}' absente dans {fichier_config}")
        wb.close()
        return noms

    ws = wb[FEUILLE_CONFIG_SECTIONS]

    row = 2  # ligne 1 = en-têtes
    seen = set()
    while True:
        val = ws.cell(row=row, column=COL_DEST).value
        if val is None or str(val).strip() == "":
            break
        name = str(val).strip()
        if name not in seen:
            noms.append(name)
            seen.add(name)
        row += 1

    wb.close()

    print("[INFO] Feuilles (dest) chargées depuis 'Filiales Analysées.xlsx' :")
    for n in noms:
        print(f"   - {n}")
    return noms

# === Construire dynamiquement `sections` à partir de la colonne A ===
_dest_names = charger_noms_feuilles_depuis_cells()

# Création directe du dictionnaire sections
sections = {name: name for name in _dest_names}  # clé logique -> nom de feuille Excel


# Optionnel mais pratique : caler la feuille de référence sur la première feuille disponible
if _dest_names:
    FEUILLE_REFERENCE = _dest_names[0]   # ex: "SNCF_SA"

SECTIONS_CONFIG = charger_sections_depuis_cells()
print(SECTIONS_CONFIG)

print(FICHIER_EXCEL_DIR)

# === Regex utilitaires ===
_MENSUEL_RE = re.compile(r"^Historique_prev_reel_filiales_(\d{4})_(\d{2})\.xlsx$")
_RE_PREV = re.compile(r"^Prévision\s+\d{2}/\d{2}", re.I)

def _is_prev(h):  return isinstance(h, str) and _RE_PREV.search(h.strip())
def _is_dates(h): return isinstance(h, str) and h.strip().lower().startswith("date")
def _is_reel(h):  return isinstance(h, str) and "réel" in h.lower()

def _parse_excel_date(v):
    """Convertit proprement tout format de date Excel (float, datetime, str...)"""
    if v is None or pd.isna(v):
        return None
    if isinstance(v, (datetime, date)):
        return pd.Timestamp(v).normalize()
    if isinstance(v, (int, float)):
        # borne basse/haute pour éviter les overflow absurdes
        if 10000 <= v <= 60000:
            try:
                return (pd.Timestamp("1899-12-30") + pd.to_timedelta(int(v), unit="D")).normalize()
            except Exception:
                return None
        else:
            return None
    if isinstance(v, str):
        ts = pd.to_datetime(v, dayfirst=True, errors="coerce")
        return ts.normalize() if pd.notna(ts) else None
    ts = pd.to_datetime(v, errors="coerce")
    return ts.normalize() if pd.notna(ts) else None

# -------------------------------------------------------------------------
# 1️⃣ Structures globales
# -------------------------------------------------------------------------
STRUCT = {}
TOKENS = {}
PREV_UNION = set()
FEUILLE_REFERENCE = "SNCF_SA"
CACHE = {}

# Index annuel: (section, flux) -> { "years": { year: { "row_idx": [...], "prof_idx": [...], "headers": [...] } } }
YEAR_INDEX = {}

def _clean_profil_label(raw, i):
    """Nettoyage identique à l'historique: (K€) retiré, 'Prévision' -> 'Profil', fallback."""
    if raw is None:
        return f"Profil {i+1}"
    s = str(raw).replace("(K€)", "").replace("Prévision", "Profil").strip()
    return s if s else f"Profil {i+1}"

def _flux_name_from_token(section, token_col):
    """Retrouve le nom de flux à partir de la colonne 'token' (col_start) dans TOKENS[section]."""
    for name, tok in TOKENS.get(section, []):
        if tok == token_col:
            return name
    return None


# -------------------------------------------------------------------------
# 2️⃣ Lister les fichiers mensuels
# -------------------------------------------------------------------------

def _lister_fichiers_mensuels():
    """
    Parcourt récursivement FICHIER_EXCEL_DIR pour trouver tous les fichiers
    'Historique_prev_reel_filiales_YYYY_MM.xlsx', y compris dans des sous-dossiers d'années.
    En cas de doublons pour un même (YYYY, MM), conserve le fichier le plus récent (mtime).
    """
    print("🔍 Recherche des fichiers mensuels (scan récursif des sous-dossiers d'années)...")

    if not os.path.isdir(FICHIER_EXCEL_DIR):
        print(f"⚠️ Dossier introuvable : {FICHIER_EXCEL_DIR}")
        return []

    best = {}  # (year, month) -> {"path": str, "mtime": float}
    try:
        for root, _, files in os.walk(FICHIER_EXCEL_DIR):
            for fname in files:
                m = _MENSUEL_RE.match(fname)
                if not m:
                    continue
                y, mth = int(m.group(1)), int(m.group(2))
                fullpath = os.path.join(root, fname)
                try:
                    mtime = os.path.getmtime(fullpath)
                except OSError:
                    mtime = 0.0

                key = (y, mth)
                if key not in best or mtime > best[key]["mtime"]:
                    best[key] = {"path": fullpath, "mtime": mtime}
    except Exception as e:
        print(f"⚠️ Erreur pendant le scan récursif : {e}")
        return []

    if not best:
        print("⚠️ Aucun fichier Historique_prev_reel_filiales_YYYY_MM.xlsx trouvé (sous-dossiers inclus).")
        return []

    files = sorted([(y, mth, info["path"]) for (y, mth), info in best.items()])
    print(f"✅ {len(files)} fichiers retenus : {[os.path.basename(f[2]) for f in files]}")
    return files

def _parse_prev_header_sort_key(h):
    # Essaie d'en déduire un tri MM/YY si possible, sinon garde l’ordre d’apparition
    import re
    s = str(h)
    m = re.search(r'(\d{2})/(\d{2,4})', s)
    if not m:
        return (9999, 99, s)
    mm = int(m.group(1))
    yy = int(m.group(2))
    if yy < 100:
        yy += 2000
    return (yy, mm, s)

def _reconcile_headers(B, new_headers):
    """
    Fusionne les headers déjà présents dans B avec ceux du fichier courant :
    - construit l’union ordonnée
    - re-map B['prev_vals'] dans ce nouvel ordre
    - rétro-remplit de None pour les séries nouvellement apparues
    """
    if "prev_headers" not in B or "prev_vals" not in B:
        B["prev_headers"] = []
        B["prev_vals"] = []

    old_headers = list(B["prev_headers"])
    # union en conservant l’ordre d’apparition, puis tente un tri chronologique
    seen = {}
    for h in old_headers + list(new_headers):
        if h not in seen:
            seen[h] = None
    union = list(seen.keys())
    union.sort(key=_parse_prev_header_sort_key)

    # rien à faire ?
    if union == old_headers:
        return

    # re-map des anciennes séries dans le nouveau canevas
    old_idx = {h: i for i, h in enumerate(old_headers)}
    n_rows = len(B.get("dates", []))
    new_prev_vals = []
    for h in union:
        if h in old_idx:
            new_prev_vals.append(B["prev_vals"][old_idx[h]])
        else:
            # nouvelle série : rétro-remplir pour les lignes déjà connues
            new_prev_vals.append([None] * n_rows)

    B["prev_headers"] = union
    B["prev_vals"] = new_prev_vals

# -------------------------------------------------------------------------
# 3️⃣ Lecture de la structure sur un fichier de référence
# -------------------------------------------------------------------------
def _lire_structure_reference(path_ref):
    print(f"\n📘 Lecture de la structure de référence ({FEUILLE_REFERENCE}) → {os.path.basename(path_ref)}")
    df_head = pd.read_excel(path_ref, sheet_name=FEUILLE_REFERENCE, header=None, nrows=3)
    row2 = df_head.iloc[1].tolist()
    row3 = df_head.iloc[2].tolist()

    ref_od = OrderedDict()
    ref_tokens = []
    col = 2  # colonne C (0-based)
    token_col = 3

    while col < len(row2):
        flux_name = row2[col]
        if pd.isna(flux_name):
            break
        flux_name = str(flux_name).strip()

        prev_headers, prev_cols = [], []
        c = col + 1
        while c < len(row3):
            h = row3[c]
            if isinstance(h, str) and "Prévision" in h:
                prev_headers.append(h.strip())
                prev_cols.append(c)
                c += 2
                continue
            break

        if prev_headers:
            ref_od[flux_name] = {
                "col_start": col + 1,
                "prev_headers": prev_headers,
                "prev_cols": [x + 1 for x in prev_cols],
                "dates_col": col,
                "reel_col": col + 1,
            }
            ref_tokens.append((flux_name, token_col))
            PREV_UNION.update(prev_headers)

        col += (2 + 2 * len(prev_headers) + 1)
        token_col += (2 + 2 * len(prev_headers) + 1)

    print(f"   → {len(ref_od)} flux détectés ({len(PREV_UNION)} prévisions).")
    for sec in sections:
        STRUCT[sec] = ref_od.copy()
        TOKENS[sec] = ref_tokens.copy()
    print("✅ Structure copiée sur toutes les feuilles.\n")

# -------------------------------------------------------------------------
# 4️⃣ Buckets & Accumulation
# -------------------------------------------------------------------------
def _ensure_flux_bucket(section, flux_name, headers=None):
    key = (section, flux_name)
    if key not in CACHE:
        CACHE[key] = {
            "dates": [],
            "reel": [],
            "prev_headers": [],   # ← vide
            "prev_vals": [],      # ← vide
        }
    return CACHE[key]

def _accumuler_valeurs_tous_mois(files_list):
    """
    Lit chaque fichier mensuel et alimente CACHE[(section, flux)] en alignant
    les séries de prévision sur l'union ordonnée des en-têtes (headers).
    - Réconciliation des headers à chaque fichier (ajout rétro-rempli de None).
    - Insertion directe ligne par ligne dans B["dates"], B["reel"], B["prev_vals"][k].
    """
    print(f"📊 Lecture de {len(files_list)} fichiers mensuels (toutes feuilles)...")
    for idx, (_, mois, path) in enumerate(files_list, 1):
        print(f"   [{idx}/{len(files_list)}] → {os.path.basename(path)}")

        try:
            all_sheets = pd.read_excel(path, sheet_name=None, header=None)
        except Exception as e:
            print(f"⚠️ Erreur lecture {path}: {e}")
            continue

        for sec, feuille in sections.items():
            if feuille not in all_sheets:
                continue

            df = all_sheets[feuille].copy()
            if df.shape[0] < 5:
                continue

            row2 = df.iloc[1].tolist()
            row3 = df.iloc[2].tolist()
            col = 2  # colonne C (0-based -> ici index pandas)

            while col < len(row2):
                flux_name = row2[col]
                if pd.isna(flux_name):
                    break
                flux_name = str(flux_name).strip()

                # Détection des colonnes "Prévision"
                prev_headers, prev_cols = [], []
                c = col + 1
                while c < len(row3):
                    h = row3[c]
                    if isinstance(h, str) and "Prévision" in h:
                        prev_headers.append(h.strip())
                        prev_cols.append(c)
                        c += 2  # on saute la colonne (k€) qui suit chaque prévision
                        continue
                    break

                if not prev_headers:
                    # Bloc sans prévisions (ex: espaces, colonnes diverses) → on saute "gros"
                    col += 10
                    continue

                # 1) Bucket pour ce flux
                B = _ensure_flux_bucket(sec, flux_name)

                # 2) Réconcilier les en-têtes avec ceux du fichier courant
                _reconcile_headers(B, prev_headers)

                # 3) Map header -> colonne pour CE fichier
                header_to_col = {h: prev_cols[j] for j, h in enumerate(prev_headers)}

                # 4) Pousser les lignes : date, réel, et séries alignées
                rows_appended = 0
                for r in range(3, df.shape[0]):
                    date_val = df.iat[r, col - 1]
                    if pd.isna(date_val):
                        continue
                    d = _parse_excel_date(date_val)
                    if d is None:
                        continue

                    B["dates"].append(d)

                    r_val = df.iat[r, col]
                    B["reel"].append(r_val if pd.notna(r_val) else 0)

                    # Pour chaque header du bucket, si la colonne existe ce mois-ci → valeur ; sinon → None
                    for k, h in enumerate(B["prev_headers"]):
                        if h in header_to_col:
                            cpr = header_to_col[h]
                            v = df.iat[r, cpr]
                            B["prev_vals"][k].append(v if pd.notna(v) else None)
                        else:
                            B["prev_vals"][k].append(None)

                    rows_appended += 1

                # Avance au bloc suivant : (dates + réel) + 2 colonnes par prévision + 1 de séparation
                col += (2 + 2 * len(prev_headers) + 1)

    print(f"✅ Cache complété : {len(CACHE)} flux au total (valeurs réelles et prévisions incluses).\n")

    # Récapitulatif par section (contrôle)
    print("📊 Récapitulatif des flux chargés par section :")
    counts = {}
    for (section, flux), data in CACHE.items():
        counts.setdefault(section, 0)
        counts[section] += 1
    for section, n in counts.items():
        total_lignes = sum(len(CACHE[(section, f)]['dates']) for f in [x for x, _ in CACHE if _ == section])
        print(f"   - {section} : {n} flux ({total_lignes} lignes cumulées)")
    print()

# -------------------------------------------------------------------------
# 5️⃣ Index annuel
# -------------------------------------------------------------------------
def _build_year_index():
    """
    Pour chaque (section, flux) dans CACHE :
      - regroupe les indices de lignes par année
      - détermine quels profils (colonnes) sont 'actifs' sur l'année (au moins une valeur != 0 / non None)
      - prépare la liste de headers nettoyés pour l'année
    Remplit YEAR_INDEX[(section, flux)].
    """
    YEAR_INDEX.clear()
    for (section, flux_name), B in CACHE.items():
        dates = B.get("dates", [])
        prev_vals = B.get("prev_vals", [])
        headers = B.get("prev_headers", [])

        # indices de lignes par année
        rows_by_year = {}
        for i, d in enumerate(dates):
            y = d.year
            rows_by_year.setdefault(y, []).append(i)

        years_map = {}
        for y, row_idx in rows_by_year.items():
            prof_idx = []
            for k, serie in enumerate(prev_vals):
                # activité = au moins une valeur non None et != 0 sur ces lignes
                active = False
                for i in row_idx:
                    if i < len(serie):
                        v = serie[i]
                        if v is None:
                            continue
                        try:
                            if float(v) != 0.0:
                                active = True
                                break
                        except Exception:
                            active = True
                            break
                if active:
                    prof_idx.append(k)

            headers_year = [_clean_profil_label(headers[k] if k < len(headers) else None, k) for k in prof_idx]

            years_map[y] = {
                "row_idx": row_idx,       # indices de lignes pour cette année
                "prof_idx": prof_idx,     # indices de colonnes actives pour cette année
                "headers": headers_year,  # labels propres des profils actifs
            }

        YEAR_INDEX[(section, flux_name)] = {"years": years_map}

# -------------------------------------------------------------------------
# 6️⃣ Initialisation complète
# -------------------------------------------------------------------------
def _init_full_load():
    print("🚀 Initialisation complète du cache de données...")
    files = _lister_fichiers_mensuels()
    if not files:
        print("ℹ️ Aucun fichier trouvé — initialisation du cache ignorée.\n")
        return

    ref_path = files[-1][2]  # dernier mois = référence
    _lire_structure_reference(ref_path)
    _accumuler_valeurs_tous_mois(files)
    _build_year_index()  # ✅ construire l’index après remplissage du CACHE
    print("✅ Chargement complet terminé.\n")


# Lance le chargement maintenant que tout est défini
_init_full_load()

# -------------------------------------------------------------------------
# 7️⃣ Variables dérivées compatibles avec l'existant
# -------------------------------------------------------------------------
previsions_triees = sorted(PREV_UNION)
nb_prev = len(previsions_triees)
taille_bloc = 2 + 2 * nb_prev + 1
print(f"📈 Nombre total de prévisions : {nb_prev}")
print(f"📦 Taille d’un bloc de flux : {taille_bloc}\n")


# -------------------------------------------------------------------------
# 8️⃣ Fonctions publiques
# -------------------------------------------------------------------------
def charger_donnees(feuille, taille_bloc_param):
    """Retourne le nom de la feuille (section) et la liste des flux détectés, en excluant certains flux inutiles."""
    print(f"➡️ charger_donnees() appelé pour '{feuille}'")

    # Liste des flux à ignorer
    noms_a_exclure = [
        "Trésorerie de fin",
        "Cashpool",
        "Emprunts",
        "Tirages Lignes CT",
        "Variation de collatéral",
        "Créances CDP",
        "Placements",
        "CC financiers",
        "Emprunts / Prêts - Groupe",
        "Encours de financement",
        "Endettement Net"
    ]

    # Récupère tous les flux
    all_noms = list(TOKENS.get(feuille, []))

    # Filtrage : on garde uniquement ceux qui ne sont pas exclus
    noms = [(name, col) for name, col in all_noms if name not in noms_a_exclure]
    for j in noms :
        print(f"noms : {j}")
    print(len(noms))
    
    if not noms:
        print(f"⚠️ Aucun flux valide trouvé pour la feuille '{feuille}'")
    else:
        print(f"   ✅ {len(noms)} flux conservés (sur {len(all_noms)} totaux)")

    return feuille, noms

# Profils à exclure pour 2023 (d’après tes listes / cases à cocher)
EXCLUDED_DATES_2023 = {
    "02/01",
    "09/01",
    "16/01",
    "30/05",
    "05/06",
    "19/06",
    "26/06",
    "03/07",
    "10/07",
    "17/07",
}

def extraire_valeurs(ws, col_start, nb_prev_param, annee=None, annee_min=None, annee_max=None):
    """
    Retourne (dates, reel, previsions, noms_profils)

    - Si 'annee' est fourni : utilise YEAR_INDEX pour ne garder que
      *les lignes* de l'année et *les profils actifs* de l'année,
      puis on filtre éventuellement certains profils (2023) → le RÉEL reste inchangé.
    - Sinon : filtre sur plage ou aucun filtre (comportement historique).
    """

    DEBUG_EXTRAIRE = True
    def _dbg(*args):
        if DEBUG_EXTRAIRE:
            print("[extraire_valeurs]", *args)

    section = ws
    flux_name = _flux_name_from_token(section, col_start)
    _dbg(f"CALL: section={section!r} col_start={col_start} nb_prev_param={nb_prev_param} "
         f"annee={annee} annee_min={annee_min} annee_max={annee_max}")

    if not flux_name:
        _dbg("ABORT: flux_name introuvable -> retour listes vides/pad prév.")
        return [], [], [[] for _ in range(nb_prev_param)], []

    B = CACHE.get((section, flux_name))
    if not B or not B.get("dates"):
        _dbg(f"ABORT: pas de cache ou dates vides pour {(section, flux_name)}")
        return [], [], [[] for _ in range(nb_prev_param)], []

    dates_all = B["dates"]
    reel_all  = B["reel"]
    prev_all  = B.get("prev_vals", [])
    headers   = B.get("prev_headers", [])   # ex: "Prévision 16/01 (K€)"

    _dbg(f"CACHE: len(dates_all)={len(dates_all)}, len(reel_all)={len(reel_all)}, "
         f"nb_profils_total={len(prev_all)}, nb_headers={len(headers)}")

    # ==============================
    # CAS 1 : année précise
    # ==============================
    if annee is not None:
        years = YEAR_INDEX.get((section, flux_name), {}).get("years", {})
        info = years.get(annee)
        if not info:
            _dbg(f"ANNEE={annee}: YEAR_INDEX absent/vide -> retour listes vides.")
            return [], [], [], []

        # Ces indices NE CHANGENT PAS pour les lignes → le RÉEL reste intact
        row_idx  = info["row_idx"]      # indices des dates pour cette année
        prof_idx = info["prof_idx"]     # indices des colonnes de prévision actives

        _dbg(f"ANNEE={annee}: rows={len(row_idx)} profils_actifs={len(prof_idx)}")

        # ---- Filtre spécial 2023 : on enlève certains profils, mais on garde le réel ----
        if annee == 2023:
            def _is_excluded(k: int) -> bool:
                if k < 0 or k >= len(headers):
                    return False
                lab = str(headers[k])  # "Prévision 16/01 (K€)"
                for d in EXCLUDED_DATES_2023:
                    if d in lab:
                        return True
                return False

            prof_idx_before = list(prof_idx)
            prof_idx = [k for k in prof_idx if not _is_excluded(k)]
            _dbg(f"ANNEE=2023: prof_idx avant filtre = {prof_idx_before}")
            _dbg(f"ANNEE=2023: prof_idx après  filtre = {prof_idx}")

        # ---- Extraction des lignes (réel inchangé) ----
        dates = [dates_all[i] for i in row_idx]
        reel  = [reel_all[i]  for i in row_idx]

        # ---- Extraction des prévisions uniquement pour les profils retenus ----
        previsions = []
        for k in prof_idx:
            serie = prev_all[k] if k < len(prev_all) else []
            previsions.append([serie[i] if i < len(serie) else None for i in row_idx])

        # Noms de profils alignés sur prof_idx
        noms_profils = [
            _clean_profil_label(headers[k] if k < len(headers) else None, k)
            for k in prof_idx
        ]

        _dbg(f"RETURN(annee): len(dates)={len(dates)}, len(reel)={len(reel)}, "
             f"nb_profils={len(previsions)}, noms_profils={noms_profils}")
        return dates, reel, previsions, noms_profils

    # ==============================
    # CAS 2 : plage / historique
    # ==============================
    def _keep(d):
        y = d.year
        if annee_min is not None and annee_max is not None:
            return annee_min <= y <= annee_max
        if annee_min is not None:
            return y >= annee_min
        if annee_max is not None:
            return y <= annee_max
        return True

    idxs = [i for i, d in enumerate(dates_all) if _keep(d)]
    dates = [dates_all[i] for i in idxs]
    reel  = [reel_all[i]  for i in idxs]
    previsions = [[serie[i] if i < len(serie) else None for i in idxs] for serie in prev_all]

    noms_profils = [_clean_profil_label(h, i) for i, h in enumerate(headers)]

    _dbg(f"PLAGE/HISTO: indices_retenus={len(idxs)} "
         f"(min_year={min((d.year for d in dates), default=None)}, "
         f"max_year={max((d.year for d in dates), default=None)})")
    _dbg(f"RETURN(plage): len(dates)={len(dates)}, len(reel)={len(reel)}, "
         f"nb_profils={len(previsions)}, len(noms_profils)={len(noms_profils)}; "
         f"exemple_date={dates[0] if dates else None}")

    return dates, reel, previsions, noms_profils


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Visualisation Réel vs Prévisions")
        self.configure(bg='#001f3f')
        self.attributes("-fullscreen", True)
        self.geometry("1300x900")
        self.bind("<Escape>", lambda e: self.attributes("-fullscreen", False))

        self.style = ttk.Style()
        self.style.theme_use('clam')

        self.style.configure("TLabel", background="#001f3f", foreground="white", font=('Arial', 16))
        self.style.configure("TButton", font=('Arial', 14), padding=10)
        self.style.configure("Treeview", background="#001f3f", foreground="white",
                             fieldbackground="#001f3f", font=('Arial', 11))
        self.style.configure("Treeview.Heading", background="#004080",
                             foreground="white", font=('Arial', 12, 'bold'))

        self.creer_accueil()
        self.canvas = None

#===================Page Accueil + fonctions de navigation===================
    def creer_accueil(self):
        # -- RESET & imports locaux
        import customtkinter as ctk
        from PIL import Image
        import tkinter as tk
        from customtkinter import CTkImage
        from datetime import datetime
        import os

        # === THEME / BASE ===
        try:
            ctk.set_appearance_mode("dark")
            ctk.set_default_color_theme("blue")
            self.configure(fg_color="#0b1220")
        except Exception:
            pass

        # Détruit proprement les enfants
        self.vider_fenetre()

        # =========================
        #  Helpers data pour KPI
        # =========================
        def _fmt_dt(ts):
            try:
                return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")
            except Exception:
                return "—"

        # Fichiers mensuels
        try:
            fichiers = _lister_fichiers_mensuels()  # supposé renvoyer (annee, mois, path)
            nb_fichiers = len(fichiers)

            if fichiers:
                last_mtime = max(os.path.getmtime(p) for _, _, p in fichiers if os.path.exists(p))
                derniere_maj = _fmt_dt(last_mtime)
                # années disponibles (1er champ du tuple)
                annees_dispos = sorted({int(a) for a, _, _ in fichiers if a is not None})
            else:
                derniere_maj = "—"
                annees_dispos = []
        except Exception:
            nb_fichiers, derniere_maj = 0, "—"
            annees_dispos = []

        # Filiales
        try:
            if SECTIONS_CONFIG:
                filiales_set = {s.get("dest") for s in SECTIONS_CONFIG if s.get("dest")}
            else:
                filiales_set = set(_dest_names) if '_dest_names' in globals() else set()
            filiales_list = sorted(f for f in filiales_set if f)
            nb_filiales = len(filiales_list)
        except Exception:
            filiales_list = []
            nb_filiales = 0

        # =========================
        #  GRID RACINE (window)
        # =========================
        self.grid_rowconfigure(0, weight=0)
        self.grid_rowconfigure(1, weight=1)
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)

        # =========================
        #  HEADER
        # =========================
        header = ctk.CTkFrame(self, fg_color="#0f1b31", corner_radius=0, height=72)
        header.grid(row=0, column=0, columnspan=2, sticky="nsew")
        header.grid_columnconfigure(0, weight=0)
        header.grid_columnconfigure(1, weight=1)
        header.grid_columnconfigure(2, weight=0)

        # Barre blanche sous le header
        barre_header = ctk.CTkFrame(self, fg_color="white", height=2)
        barre_header.grid(row=1, column=0, columnspan=2, sticky="ew")

        # Logo à gauche
        logo_wrap = ctk.CTkFrame(header, fg_color="transparent")
        logo_wrap.grid(row=0, column=0, sticky="w", padx=24, pady=12)

        try:
            _img = Image.open(image_path)
            ratio = _img.width / max(_img.height, 1)
            new_h, new_w = 44, int(44 * ratio)
            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS
            _img = _img.resize((new_w, new_h), resample_mode)
            cimg = CTkImage(light_image=_img, dark_image=_img, size=(_img.width, _img.height))
            logo = ctk.CTkLabel(logo_wrap, image=cimg, text="")
            logo.image = cimg
            logo.pack(side="left")
        except Exception:
            ctk.CTkLabel(logo_wrap, text="PULSE", font=("Segoe UI Semibold", 20),
                        text_color="white").pack(side="left")

        titre = ctk.CTkLabel(
            header,
            text="PROJET PULSE — ANALYSE DE L’EXISTANT",
            font=("Segoe UI Semibold", 18, "bold"),
            text_color="white",
        )
        titre.grid(row=0, column=1, sticky="w", padx=8)

        # Actions header à droite
        header_actions = ctk.CTkFrame(header, fg_color="transparent")
        header_actions.grid(row=0, column=2, sticky="e", padx=20)

        ctk.CTkButton(
            header_actions, text="❌ Quitter",
            fg_color="#C21515", hover_color="#A00000",
            command=self.demander_confirmation_quit
        ).grid(row=0, column=1)

        # =========================
        #  SIDEBAR
        # =========================
        sidebar = ctk.CTkFrame(self, fg_color="#0e162a", corner_radius=0, width=260)
        sidebar.grid(row=1, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self.grid_rowconfigure(1, weight=1)
        sidebar.grid_rowconfigure(11, weight=1)

        inner = ctk.CTkFrame(sidebar, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=14, pady=14)

        def _nav_btn(parent, text, cmd, emoji=""):
            btn = ctk.CTkButton(
                parent, text=f"{emoji}  {text}", anchor="w",
                width=220, height=42, corner_radius=12,
                fg_color="#15305b", hover_color="#1d3e77",
                font=("Segoe UI", 13, "bold"),
                command=cmd
            )

            def _enter(_):
                btn.configure(fg_color="#1d3e77")

            def _leave(_):
                btn.configure(fg_color="#15305b")

            btn.bind("<Enter>", _enter)
            btn.bind("<Leave>", _leave)
            return btn

        ctk.CTkLabel(inner, text="Navigation", text_color="#9fb7dd",
                    font=("Segoe UI Semibold", 14)).pack(anchor="w", pady=(4, 12))

        _nav_btn(inner, "IA — Prédiction", self.creer_page_ia_prediction, "🤖").pack(fill="x", pady=6)

        separator = ctk.CTkFrame(sidebar, fg_color="#1a2745", height=2)
        separator.pack(fill="x", padx=14, pady=(4, 2))

        foot = ctk.CTkLabel(
            sidebar,
            text=f"v1.0 • Dernière MAJ : {derniere_maj}",
            text_color="#6e86ad",
            font=("Segoe UI", 11)
        )
        foot.pack(anchor="w", padx=20, pady=(8, 10))

        # =========================
        #  MAIN CONTENT
        # =========================
        main = ctk.CTkScrollableFrame(self, fg_color="#0b1220", corner_radius=0)
        main.grid(row=1, column=1, sticky="nsew", padx=0, pady=0)
        self.grid_columnconfigure(1, weight=1)
        main.grid_columnconfigure(0, weight=1)

        # ====== HERO ======
        hero = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        hero.grid(row=0, column=0, sticky="ew", padx=24, pady=(20, 14))
        hero.grid_columnconfigure(0, weight=1)
        hero.grid_columnconfigure(1, weight=0)

        ctk.CTkLabel(
            hero,
            text="Bienvenue 👋",
            font=("Segoe UI Semibold", 24, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(16, 0))

        subtitle_text = (
            "Console d'analyse des flux de trésorerie : réels, prévisions, écarts et détection d'anomalies."
        )
        subtitle_label = ctk.CTkLabel(
            hero,
            text="",
            font=("Segoe UI", 13),
            text_color="#c9defa",
            wraplength=520,
            justify="left"
        )
        subtitle_label.grid(row=1, column=0, sticky="w", padx=18, pady=(6, 10))

        def _typewriter(i=0):
            if i <= len(subtitle_text):
                subtitle_label.configure(text=subtitle_text[:i])
                self.after(12, _typewriter, i + 1)

        _typewriter()

        # Pills
        pills = ctk.CTkFrame(hero, fg_color="transparent")
        pills.grid(row=2, column=0, sticky="w", padx=16, pady=(0, 16))

        def _pill(parent, text, emoji=""):
            f = ctk.CTkFrame(parent, fg_color="#152544", corner_radius=20)
            f.pack(side="left", padx=6)
            ctk.CTkLabel(
                f,
                text=f"{emoji}  {text}",
                font=("Segoe UI", 11),
                text_color="#c9defa"
            ).pack(padx=12, pady=5)

        _pill(pills, "Réel vs prévisions multi-horizon", "📈")
        _pill(pills, "Analyse d'écarts & anomalies", "🧠")
        _pill(pills, "Vue filiales / flux / profils", "📊")

        # Illustration graphique (3 mini graphes)
        viz = ctk.CTkFrame(hero, fg_color="#111b33", corner_radius=18)
        viz.grid(row=0, column=1, rowspan=3, sticky="e", padx=18, pady=12)
        viz.grid_columnconfigure((0, 1, 2), weight=1)
        viz.grid_rowconfigure(0, weight=1)

        # --- Graph 1 : barres animées ---
        graph1 = ctk.CTkFrame(viz, fg_color="transparent")
        graph1.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        bars_h = [52, 34, 78, 60]
        bar_widgets = []
        for h_val in bars_h:
            col = ctk.CTkFrame(graph1, fg_color="transparent")
            col.pack(side="left", padx=4, fill="y", expand=True)

            base = ctk.CTkFrame(col, fg_color="#1b2945", corner_radius=8, width=18, height=85)
            base.pack(side="bottom", pady=(0, 4))

            bar = ctk.CTkFrame(col, fg_color="#2563eb", corner_radius=8, width=18, height=1)
            bar.place(relx=0.5, rely=1.0, anchor="s")
            bar_widgets.append((bar, h_val))

        def _animate_bars(step=0, steps=18):
            factor = min(1.0, step / float(steps))
            for bar, target in bar_widgets:
                bar.configure(height=int(target * factor))
            if step < steps:
                self.after(25, _animate_bars, step + 1, steps)

        _animate_bars()

        # --- Graph 2 : courbe ---
        graph2 = ctk.CTkFrame(viz, fg_color="transparent")
        graph2.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

        pts = [10, 40, 22, 70, 50, 90, 60, 80, 55, 100]
        w, h = 120, 90

        canvas2 = ctk.CTkCanvas(graph2, width=w, height=h, bg="#111b33", bd=0, highlightthickness=0)
        canvas2.pack(fill="both", expand=True)

        max_v = max(pts)
        coords = [(i * (w // (len(pts) - 1)), h - int((v / max_v) * (h - 10))) for i, v in enumerate(pts)]
        for i in range(len(coords) - 1):
            canvas2.create_line(
                coords[i][0], coords[i][1],
                coords[i + 1][0], coords[i + 1][1],
                fill="#2563eb", width=3, smooth=True
            )
        for x, y in coords:
            canvas2.create_oval(x - 4, y - 4, x + 4, y + 4, fill="#3c82ff", outline="")

        # --- Graph 3 : donut ---
        graph3 = ctk.CTkFrame(viz, fg_color="transparent")
        graph3.grid(row=0, column=2, sticky="nsew", padx=10, pady=10)

        canvas3 = tk.Canvas(graph3, width=110, height=110,
                            bg="#111b33", bd=0, highlightthickness=0)
        canvas3.pack(fill="both", expand=True, padx=6, pady=6)

        values = [30, 15, 22, 18, 15]
        total = sum(values)
        start_angle = 0
        colors = ["#1d4ed8", "#2563eb", "#3b82f6", "#60a5fa", "#93c5fd"]
        bbox_outer = (8, 8, 102, 102)

        for i, v in enumerate(values):
            extent = 360 * v / total
            canvas3.create_arc(
                *bbox_outer,
                start=start_angle,
                extent=extent,
                fill=colors[i % len(colors)],
                outline="#111b33",
                style="pieslice"
            )
            start_angle += extent

        bbox_inner = (28, 28, 82, 82)
        canvas3.create_oval(*bbox_inner, fill="#111b33", outline="#111b33")

        # ====== KPI CARDS ======
        kpi = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        kpi.grid(row=1, column=0, sticky="ew", padx=24, pady=10)
        for i in range(3):
            kpi.grid_columnconfigure(i, weight=1)

        # ----- Zone "Détails" dynamique juste sous les KPI -----
        details = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        details.grid(row=2, column=0, sticky="ew", padx=24, pady=(0, 10))
        details.grid_columnconfigure(0, weight=1)

        details_title = ctk.CTkLabel(
            details,
            text="Détails",
            font=("Segoe UI Semibold", 15, "bold"),
            text_color="white"
        )
        details_title.grid(row=0, column=0, sticky="w", padx=16, pady=(14, 4))

        details_box = ctk.CTkTextbox(
            details,
            height=110,
            wrap="word",
            fg_color="#142544",
            text_color="white",
            font=("Segoe UI", 12),
            border_width=1,
            border_color="#223658",
            corner_radius=12
        )
        details_box.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 14))
        details_box.insert("1.0", "Clique sur un KPI pour afficher les détails (années ou filiales)…")
        details_box.configure(state="disabled")

        def _set_details(title: str, lines: list[str]):
            details_title.configure(text=title)
            details_box.configure(state="normal")
            details_box.delete("1.0", "end")
            if not lines:
                details_box.insert("1.0", "Aucune donnée disponible.")
            else:
                details_box.insert("1.0", "\n".join(lines))
            details_box.configure(state="disabled")

        def _show_years():
            if annees_dispos:
                lignes = ["Années disponibles :"]
                lignes += [f"• {a}" for a in annees_dispos]
            else:
                lignes = ["Aucune année détectée (pas de fichiers mensuels trouvés)."]
            _set_details("Détails — Fichiers mensuels détectés", lignes)

        def _show_filiales():
            if filiales_list:
                lignes = ["Filiales configurées :"]
                lignes += [f"• {f}" for f in filiales_list]
            else:
                lignes = ["Aucune filiale détectée dans la configuration."]
            _set_details("Détails — Filiales", lignes)

        def _kpi_card(parent, title, value, subtitle, col, detail_cb=None):
            card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16,
                                border_width=1, border_color="#223658")
            card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
            ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
            ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 28, "bold"),
                        text_color="white") \
                .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 4))
            ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11),
                        text_color="#7ea2d8") \
                .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 6))

            if detail_cb is not None:
                btn = ctk.CTkButton(
                    card,
                    text="Voir détails",
                    height=28,
                    corner_radius=10,
                    fg_color="#2563eb",
                    hover_color="#1d4ed8",
                    command=detail_cb
                )
                btn.grid(row=3, column=0, sticky="w", padx=14, pady=(0, 10))

        _kpi_card(
            kpi,
            "Fichiers mensuels détectés",
            f"{nb_fichiers:,}".replace(",", " "),
            "Historique *_YYYY_MM.xlsx*",
            0,
            detail_cb=_show_years
        )
        _kpi_card(
            kpi,
            "Nombre de filiales",
            f"{nb_filiales:,}".replace(",", " "),
            "Distinctes dans la config",
            1,
            detail_cb=_show_filiales
        )
        _kpi_card(
            kpi,
            "Dernière actualisation",
            f"{derniere_maj}",
            "Basée sur les mtime",
            2,
            detail_cb=None
        )



        # ====== ACTIVITÉ RÉCENTE ======
        recent = ctk.CTkFrame(main, fg_color="#0f1b31", corner_radius=18)
        recent.grid(row=4, column=0, sticky="ew", padx=24, pady=(10, 20))
        recent.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            recent, text="Activité récente",
            font=("Segoe UI Semibold", 16, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=16, pady=(14, 8))

        table = ctk.CTkFrame(recent, fg_color="#142544", corner_radius=12,
                            border_width=1, border_color="#223658")
        table.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 16))
        for i in range(3):
            table.grid_columnconfigure(i, weight=[3, 1, 1][i])

        def _header(r, text):
            ctk.CTkLabel(table, text=text, text_color="#9fb7dd",
                        font=("Segoe UI Semibold", 12)) \
                .grid(row=r, column=["Fichier", "Taille", "Modifié le"].index(text),
                    sticky="w", padx=12, pady=(10, 6))

        _header(0, "Fichier")
        _header(0, "Taille")
        _header(0, "Modifié le")

        recent_files = []
        try:
            if fichiers:
                rec = sorted(
                    [(p, os.path.getmtime(p)) for _, _, p in fichiers if os.path.exists(p)],
                    key=lambda x: x[1], reverse=True
                )[:5]
                for p, ts in rec:
                    size = f"{(os.path.getsize(p) / 1024 / 1024):.1f} Mo"
                    recent_files.append((os.path.basename(p), size, _fmt_dt(ts)))
        except Exception:
            pass

        if not recent_files:
            recent_files = [("—", "—", "—")]

        for i, (name, size, dt_str) in enumerate(recent_files, start=1):
            ctk.CTkLabel(table, text=name, text_color="white", font=("Segoe UI", 12)) \
                .grid(row=i, column=0, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=size, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=1, sticky="w", padx=12, pady=6)
            ctk.CTkLabel(table, text=dt_str, text_color="#c9defa", font=("Segoe UI", 12)) \
                .grid(row=i, column=2, sticky="w", padx=12, pady=6)

    def vider_fenetre(self):
        for widget in self.winfo_children():
            widget.destroy()

    def retour_menu(self):
        self.vider_fenetre()
        self.creer_accueil()

    def demander_confirmation_quit(self):
        if messagebox.askokcancel("Quitter", "Voulez-vous vraiment quitter l'application ?"):
            self.destroy()

#===================Page Prédiction IA===================
    def _rebuild_profils_ui2(self, section, flux_name, annee):
        """Reconstruit la select-box de profils selon le flux et l'année sélectionnée."""
        import tkinter as tk

        # Nettoyage du frame
        for w in self.profils_frame.winfo_children():
            w.destroy()

        # nouveau : on n'utilise plus self.vars_prev ici
        self.vars_prev = []
        self.profils_names_order = []   # ordre des profils pour mapping
        self.profils_listbox = None     # widget de sélection multiple

        if annee is None:
            return

        noms_profils, _ = self._profils_for_year(section, flux_name, annee)
        self.profils_names_order = noms_profils

        if not noms_profils:
            info = tk.Label(
                self.profils_frame,
                text="Aucun profil actif pour l'année sélectionnée.",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
            )
            info.pack(anchor="w")
            return

        # Label d'info
        lbl = tk.Label(
            self.profils_frame,
            text="Profils (Ctrl+clic pour multi-sélection) :",
            bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
        )
        lbl.pack(anchor="w", padx=4, pady=(0, 4))

        # Frame contenant la listbox + scrollbar
        container = tk.Frame(self.profils_frame, bg="#00122e")
        container.pack(fill="x", padx=4, pady=(0, 6))

        scrollbar = tk.Scrollbar(container, orient="vertical")
        self.profils_listbox = tk.Listbox(
            container,
            selectmode="multiple",
            exportselection=False,
            bg="#00122e",
            fg="white",
            font=('Segoe UI', 10),
            height=min(8, len(noms_profils)),
            highlightthickness=0,
            activestyle="dotbox"
        )
        self.profils_listbox.pack(side="left", fill="x", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.profils_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.profils_listbox.yview)

        for name in noms_profils:
            self.profils_listbox.insert("end", name)

        # petite sélection par défaut : tout sélectionner
        self.profils_listbox.select_set(0, "end")

    def creer_page_ia_prediction(self):
        """
        Page IA de prévision (N -> N+1) pilotée par un flux sélectionnable.
        Cette méthode :
        1) Vérifie les dépendances
        2) Réinitialise la fenêtre
        3) Construit toute la page IA (UI + callbacks + entraînement)
        -> le dataset sera construit dynamiquement selon le flux choisi
        """
        from tkinter import messagebox

        # sécurité sklearn
        if RandomForestRegressor is None:
            messagebox.showerror(
                "Dépendance manquante",
                "Le module 'scikit-learn' n'est pas installé.\n"
                "Installe-le avec : pip install scikit-learn"
            )
            return

        # reset page
        self.vider_fenetre()

        # Build page IA (dataset construit selon flux dans la page)
        self._ia_build_prediction_page(df=None)
        
    def _ia_build_dataset_for_flux(self, flux_cible: str):
        """
        Construit le DataFrame global df sur le flux choisi (ex: 'Trafic Voyageurs')
        pour toutes les filiales, avec colonnes :
        section, date, y, year, month, dayofyear, section_id, roll_mean_7, roll_mean_30
        Retourne df ou None si aucune donnée.
        """
        import pandas as pd
        import numpy as np
        from tkinter import messagebox

        lignes = []

        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        for section_name, feuille in sections.items():
            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees pour {feuille} : {e}")
                continue

            cible = [t for t in noms_flux if t[0] == flux_cible]
            if not cible:
                continue

            _, col_start = cible[0]
            try:
                dates, reel, _previsions, _noms_profils = extraire_valeurs(
                    ws, col_start, nb_prev, annee=None
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs pour {section_name}/{flux_cible} : {e}")
                continue

            for d, r in zip(dates, reel):
                y_val = _to_number(r)
                if y_val is None:
                    continue
                try:
                    d_ts = pd.to_datetime(d)
                except Exception:
                    continue
                lignes.append({
                    "section": section_name,
                    "date": d_ts,
                    "y": float(y_val)
                })

        if not lignes:
            messagebox.showinfo(
                "Information",
                f"Aucune donnée trouvée pour le flux '{flux_cible}' sur les filiales."
            )
            self.retour_menu()
            return None

        df = pd.DataFrame(lignes).sort_values(["section", "date"]).reset_index(drop=True)
        df["year"] = df["date"].dt.year
        df["month"] = df["date"].dt.month
        df["dayofyear"] = df["date"].dt.dayofyear

        cat = df["section"].astype("category")
        df["section_id"] = cat.cat.codes

        df["roll_mean_7"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(7, min_periods=1).mean()
        )
        df["roll_mean_30"] = df.groupby("section")["y"].transform(
            lambda s: s.rolling(30, min_periods=1).mean()
        )

        return df

    def _ia_build_prediction_page(self, df):
        """
        Construit toute l'interface IA + callbacks d'entraînement
        à partir du DataFrame global df déjà préparé.

        Version + SELECT BOX FLUX (rebuild df_current selon flux choisi).
        """
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt
        from PIL import Image
        from customtkinter import CTkImage
        import customtkinter as ctk
        import tkinter as tk
        from tkinter import ttk, messagebox, filedialog
        import pandas as pd
        import numpy as np
        import traceback

        # ---------- helpers numériques locaux ----------
        def _to_number(x):
            if x is None:
                return None
            if isinstance(x, str):
                s = x.strip().replace("\xa0", " ").replace(" ", "")
                if s in {"", "-", "—", "NA", "N/A"}:
                    return None
                s = s.replace(",", ".")
                try:
                    return float(s)
                except Exception:
                    return None
            try:
                return float(x)
            except Exception:
                return None

        def _to_float_or_nan(x):
            val = _to_number(x)
            return float(val) if val is not None else np.nan

        default_n_estimators = 200

        # ---------- variables profils & graph 2 ----------
        ia_profils_vars = []        # liste de tk.BooleanVar
        ia_profils_names = []       # noms de profils actifs N+1
        ia_profils_dates = []       # dates N+1
        ia_profils_series = []      # séries (une par profil actif)

        current_pred_df = None          # df_future_all pour N+1
        current_real_target_df = None   # réel sur N+1 (si dispo)
        current_target_year = None
        current_filiale_name = None

        ia_graph2_widget = None
        graph2_container_packed = False

        exported_pred_df = None         # pour export Excel prédictions
        analysis_table_frame = None
        analysis_tree = None
        analysis_export_button = None
        export_button = None            # bouton d’export des prédictions
        export_graph2_button = None

        # NEW: dataset courant selon flux
        df_current = None

        model_artifacts = {
            "X_train": None,
            "y_train": None,
            "features": None,
            "cls_model": None,
            "reg_model": None,
        }

        # ============================================================
        # NEW: récupérer la liste des flux disponibles (toutes filiales)
        # ============================================================
        def _get_all_flux_names():
            flux_set = set()
            for _section_name, feuille in sections.items():
                try:
                    _ws, noms_flux = charger_donnees(feuille, taille_bloc)
                    for nom, _col in noms_flux:
                        flux_set.add(nom)
                except Exception:
                    pass
            out = sorted(flux_set)
            return out

        # ============ HEADER ============
        header_frame = ctk.CTkFrame(self, fg_color="#001f3f", corner_radius=0)
        header_frame.pack(side="top", fill="x", pady=(20, 5), padx=30)

        titre_font = ("Segoe UI Semibold", 26, "bold")
        ctk.CTkLabel(
            header_frame,
            text="PROJET PULSE - IA PRÉDICTION  (N → N+1)",
            font=titre_font,
            text_color="white"
        ).pack(side="left", anchor="w")

        # Logo
        try:
            logo_img = Image.open(image_path)
            test = tk.Label(self, text="Test", font=titre_font)
            test.update_idletasks()
            text_h = test.winfo_reqheight()
            test.destroy()
            ratio = logo_img.width / max(logo_img.height, 1)
            new_w, new_h = int(text_h * ratio), text_h

            try:
                resample_mode = Image.Resampling.LANCZOS
            except AttributeError:
                resample_mode = Image.ANTIALIAS

            resized_logo = logo_img.resize((new_w, new_h), resample_mode)
            ctk_logo = CTkImage(light_image=resized_logo, dark_image=resized_logo, size=(new_w, new_h))
            logo_label = ctk.CTkLabel(header_frame, image=ctk_logo, text="", fg_color="#001f3f")
            logo_label.image = ctk_logo
            logo_label.pack(side="right", anchor="e", padx=(10, 0))
        except Exception as e:
            print(f"[IA] Erreur chargement logo IA : {e}")

        ctk.CTkFrame(self, height=2, fg_color="white").pack(side="top", fill="x")

        # ============ CONTAINER SCROLLABLE ============
        container = ctk.CTkFrame(self, fg_color="#00122e", corner_radius=15)
        container.pack(side="top", fill="both", expand=True, padx=30, pady=30)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        canvas_frame = tk.Frame(container, bg="#00122e")
        canvas_frame.grid(row=0, column=0, sticky="nsew")
        main_canvas = tk.Canvas(canvas_frame, bg="#00122e", highlightthickness=0)
        main_canvas.grid(row=0, column=0, sticky="nsew")
        canvas_frame.grid_rowconfigure(0, weight=1)
        canvas_frame.grid_columnconfigure(0, weight=1)

        v_scrollbar = tk.Scrollbar(container, orient="vertical", command=main_canvas.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = tk.Scrollbar(container, orient="horizontal", command=main_canvas.xview)
        h_scrollbar.grid(row=1, column=0, columnspan=2, sticky="ew")
        main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        scrollable_frame = ctk.CTkFrame(main_canvas, fg_color="#00122e", corner_radius=0)
        window_id = main_canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")

        def _on_frame_configure(_):
            main_canvas.configure(scrollregion=main_canvas.bbox("all"))
        scrollable_frame.bind("<Configure>", _on_frame_configure)

        def _on_canvas_configure(event):
            main_canvas.itemconfig(window_id, width=event.width)
        main_canvas.bind("<Configure>", _on_canvas_configure)

        # ============ TITRE / TEXTE ============
        title_label = ctk.CTkLabel(
            scrollable_frame,
            text="IA - Prédiction (prévision année N+1)",
            font=("Segoe UI", 18, "bold"),
            text_color="white"
        )
        title_label.pack(pady=15)

        desc_label = ctk.CTkLabel(
            scrollable_frame,
            text=(
                "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                "Il apprend comment la valeur évolue d'une année à l'autre (N → N+1) pour chaque jour.\n"
                "Il prédit ensuite l’ensemble de l’année N+1.\n"
                "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
            ),
            font=("Segoe UI", 12),
            text_color="#c9defa",
            justify="left"
        )
        desc_label.pack(pady=(0, 10))

        # ============ ZONE PARAMÈTRES (CARD) ============
        params_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        params_frame.pack(fill="x", padx=10, pady=(5, 10))

        # ------------------- Flux -------------------
        ctk.CTkLabel(
            params_frame,
            text="Flux :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 2))

        flux_list = _get_all_flux_names()
        if not flux_list:
            flux_list = ["Trafic Voyageurs"]

        default_flux = "Trafic Voyageurs" if "Trafic Voyageurs" in flux_list else flux_list[0]
        selected_flux = tk.StringVar(value=default_flux)

        flux_box = ttk.Combobox(
            params_frame,
            textvariable=selected_flux,
            values=flux_list,
            state="readonly",
            width=28
        )
        flux_box.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Filiale -------------------
        ctk.CTkLabel(
            params_frame,
            text="Filiale :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=1, sticky="w", padx=12, pady=(10, 2))

        selected_filiale = tk.StringVar(value="")
        filiale_box = ttk.Combobox(
            params_frame,
            textvariable=selected_filiale,
            values=[],
            state="readonly",
            width=28
        )
        filiale_box.grid(row=1, column=1, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Année -------------------
        ctk.CTkLabel(
            params_frame,
            text="Année N (historique → N+1) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=2, sticky="w", padx=12, pady=(10, 2))

        annees_var = tk.StringVar(value="")
        annees_box = ttk.Combobox(
            params_frame,
            textvariable=annees_var,
            values=[],
            state="readonly",
            width=18
        )
        annees_box.grid(row=1, column=2, sticky="w", padx=12, pady=(0, 10))

        # ------------------- Hyperparams -------------------
        ctk.CTkLabel(
            params_frame,
            text="Hyperparamètres Random Forest (point de départ) :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).grid(row=0, column=3, columnspan=2, sticky="w", padx=12, pady=(10, 2))

        # n_estimators
        ctk.CTkLabel(
            params_frame,
            text="n_estimators",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=3, sticky="w", padx=12, pady=(0, 2))

        n_estimators_var = tk.IntVar(value=default_n_estimators)
        slider_n = ctk.CTkSlider(
            params_frame,
            from_=50,
            to=600,
            number_of_steps=11,
            command=lambda v: n_estimators_var.set(int(v))
        )
        slider_n.set(default_n_estimators)
        slider_n.grid(row=2, column=3, sticky="we", padx=12, pady=(0, 8))

        lbl_n = ctk.CTkLabel(
            params_frame,
            text=f"{default_n_estimators}",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_n.grid(row=3, column=3, sticky="w", padx=12, pady=(0, 6))

        def _update_lbl_n(_=None):
            lbl_n.configure(text=str(n_estimators_var.get()))
        slider_n.configure(command=lambda v: (n_estimators_var.set(int(v)), _update_lbl_n()))

        # max_depth
        ctk.CTkLabel(
            params_frame,
            text="max_depth",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        ).grid(row=1, column=4, sticky="w", padx=12, pady=(0, 2))

        max_depth_var = tk.IntVar(value=15)
        slider_d = ctk.CTkSlider(
            params_frame,
            from_=3,
            to=25,
            number_of_steps=11,
            command=lambda v: max_depth_var.set(int(v))
        )
        slider_d.set(15)
        slider_d.grid(row=2, column=4, sticky="we", padx=12, pady=(0, 8))

        lbl_d = ctk.CTkLabel(
            params_frame,
            text="15",
            font=("Segoe UI", 11),
            text_color="#c9defa"
        )
        lbl_d.grid(row=3, column=4, sticky="w", padx=12, pady=(0, 2))

        def _update_lbl_d(_=None):
            lbl_d.configure(text=str(max_depth_var.get()))
        slider_d.configure(command=lambda v: (max_depth_var.set(int(v)), _update_lbl_d()))

        use_depth_var = tk.BooleanVar(value=True)
        chk_depth = ctk.CTkCheckBox(
            params_frame,
            text="Limiter la profondeur",
            variable=use_depth_var,
            text_color="#c9defa",
            font=("Segoe UI", 11)
        )
        chk_depth.grid(row=4, column=4, sticky="w", padx=12, pady=(0, 10))

        # Bouton entraîner
        bouton_train = ctk.CTkButton(
            params_frame,
            text="🚀 Entraîner / ré-entraîner le modèle",
            width=240,
            height=40,
            corner_radius=16,
            fg_color="#008C4B",
            hover_color="#006C39",
            text_color="white",
            font=("Segoe UI", 13, "bold")
        )
        bouton_train.grid(row=1, column=5, rowspan=3, padx=14, pady=(0, 10), sticky="e")

        for c in range(6):
            params_frame.grid_columnconfigure(c, weight=1 if c in (3, 4) else 0)

        # ============================================================
        # NEW: rebuild dataset selon flux + refresh combobox filiale/année
        # ============================================================
        def _refresh_filiale_year_boxes_from_df():
            nonlocal df_current
            if df_current is None or df_current.empty:
                return

            filiales = sorted(df_current["section"].unique().tolist())
            filiale_box.configure(values=filiales)
            if filiales:
                if selected_filiale.get() not in filiales:
                    selected_filiale.set(filiales[0])

            years_sorted = sorted(df_current["year"].unique().tolist())
            annees_box.configure(values=[str(y) for y in years_sorted])
            if years_sorted:
                if annees_var.get() not in [str(y) for y in years_sorted]:
                    annees_var.set(str(years_sorted[-1]))

        def _rebuild_df_for_current_flux():
            nonlocal df_current
            flux = selected_flux.get()
            df_new = self._ia_build_dataset_for_flux(flux)
            if df_new is None:
                return False
            df_current = df_new
            _refresh_filiale_year_boxes_from_df()

            # titre / texte dynamiques
            try:
                title_label.configure(text=f"IA - Prédiction sur le flux '{flux}' (prévision année N+1)")
                desc_label.configure(text=(
                    "Le modèle utilise tout l’historique disponible jusqu'à l'année N choisie pour la filiale.\n"
                    f"Il apprend comment la valeur de '{flux}' évolue d'une année à l'autre (N → N+1) pour chaque jour.\n"
                    "Il prédit ensuite l’ensemble de l’année N+1 pour cette filiale.\n"
                    "Les profils de prévision N+1 peuvent être affichés sur le graphe détaillé via des cases à cocher."
                ))
            except Exception:
                pass

            return True

        # init dataset
        if not _rebuild_df_for_current_flux():
            return

        # ============ CARD PROFILS N+1 (CASES À COCHER) ============
        profils_card = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
        profils_card.pack(fill="x", padx=10, pady=(0, 15))

        ctk.CTkLabel(
            profils_card,
            text="Profils de prévision disponibles sur l'année N+1 :",
            font=("Segoe UI", 12, "bold"),
            text_color="white"
        ).pack(anchor="w", padx=12, pady=(10, 4))

        ia_profils_frame = tk.Frame(profils_card, bg="#00122e")
        ia_profils_frame.pack(fill="x", padx=12, pady=(0, 10))

        def _export_graph2_to_excel():
            """
            Export Excel des valeurs utilisées pour le graphe 2 :
            - date
            - réel (si dispo)
            - prédiction IA (si dispo)
            - profils cochés (si dispo)
            """
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal ia_profils_names, ia_profils_vars, ia_profils_dates, ia_profils_series

            if (current_pred_df is None or current_pred_df.empty) and (current_real_target_df is None or current_real_target_df.empty):
                messagebox.showinfo("Export", "Aucune donnée à exporter (ni réel, ni prévision).")
                return

            try:
                all_dates = set()

                if current_real_target_df is not None and not current_real_target_df.empty:
                    all_dates |= set(pd.to_datetime(current_real_target_df["date"]).dt.normalize())

                if current_pred_df is not None and not current_pred_df.empty:
                    all_dates |= set(pd.to_datetime(current_pred_df["date"]).dt.normalize())

                if ia_profils_dates:
                    all_dates |= set(pd.to_datetime(ia_profils_dates).normalize())

                if not all_dates:
                    messagebox.showinfo("Export", "Aucune date exploitable à exporter.")
                    return

                dates_sorted = sorted(all_dates)
                df_out = pd.DataFrame({"date": pd.to_datetime(dates_sorted)})

                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = (
                        current_real_target_df.assign(date_norm=pd.to_datetime(current_real_target_df["date"]).dt.normalize())
                        .set_index("date_norm")["y"]
                        .to_dict()
                    )
                    df_out["reel"] = df_out["date"].dt.normalize().map(real_map)
                else:
                    df_out["reel"] = np.nan

                if current_pred_df is not None and not current_pred_df.empty:
                    pred_map = (
                        current_pred_df.assign(date_norm=pd.to_datetime(current_pred_df["date"]).dt.normalize())
                        .set_index("date_norm")["pred_value"]
                        .to_dict()
                    )
                    df_out["prev_ia"] = df_out["date"].dt.normalize().map(pred_map)
                else:
                    df_out["prev_ia"] = np.nan

                if ia_profils_names and ia_profils_series and ia_profils_dates and ia_profils_vars:
                    prof_dates = pd.to_datetime(ia_profils_dates).normalize()

                    for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                        if not var.get():
                            continue

                        vals = [_to_float_or_nan(v) for v in serie]
                        prof_map = pd.Series(vals, index=prof_dates, dtype="float64").to_dict()

                        col_name = f"profil_{name}".replace("\n", " ").strip()
                        df_out[col_name] = df_out["date"].dt.normalize().map(prof_map)

                df_out.insert(0, "filiale", current_filiale_name or selected_filiale.get())
                df_out.insert(1, "annee", current_target_year or (int(annees_var.get()) + 1))
                df_out.insert(2, "flux", selected_flux.get())

                file_path = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Fichiers Excel", "*.xlsx")],
                    title="Exporter les données du graphe détaillé (Graph2) en Excel"
                )
                if not file_path:
                    return

                df_out.to_excel(file_path, index=False)
                messagebox.showinfo("Export", f"Export Graph2 OK :\n{file_path}")

            except Exception as e:
                messagebox.showerror("Erreur export Graph2", str(e))

        def _show_pyvista_3d():
            import numpy as np
            import pandas as pd

            try:
                import pyvista as pv
            except Exception as e:
                messagebox.showerror(
                    "PyVista",
                    "PyVista n'est pas dispo dans cet environnement.\n"
                    "Installe-le dans le même Python/venv que ton script.\n\n"
                    f"Détail: {e}"
                )
                return

            feats = model_artifacts.get("features")
            cls_model = model_artifacts.get("cls_model")
            reg_model = model_artifacts.get("reg_model")
            X_train = model_artifacts.get("X_train")

            if feats is None or cls_model is None or reg_model is None or X_train is None:
                messagebox.showinfo("3D", "Lance d'abord l'entraînement du modèle.")
                return

            features = list(feats)

            # -----------------------
            # Helpers
            # -----------------------
            def _norm01(x):
                x = np.asarray(x, dtype=float)
                if x.size == 0:
                    return x.astype(np.float32)
                mn = np.nanmin(x)
                mx = np.nanmax(x)
                if not np.isfinite(mn) or not np.isfinite(mx) or abs(mx - mn) < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return ((x - mn) / (mx - mn)).astype(np.float32)

            def _norm_signed(x):
                """Normalize to [-1,1] using max abs."""
                x = np.asarray(x, dtype=float)
                m = np.nanmax(np.abs(x))
                if not np.isfinite(m) or m < 1e-12:
                    return np.zeros_like(x, dtype=np.float32)
                return (x / m).astype(np.float32)

            def build_lines(src_pts, dst_pts, w):
                e = len(src_pts)
                if e == 0:
                    m = pv.PolyData(np.zeros((0, 3), dtype=np.float32))
                    m.lines = np.array([], dtype=np.int64)
                    m.point_data["w"] = np.array([], dtype=np.float32)
                    return m

                pts = np.empty((2 * e, 3), dtype=np.float32)
                pts[0::2] = src_pts
                pts[1::2] = dst_pts

                lines = np.empty((3 * e,), dtype=np.int64)
                lines[0::3] = 2
                lines[1::3] = np.arange(0, 2 * e, 2, dtype=np.int64)
                lines[2::3] = np.arange(1, 2 * e, 2, dtype=np.int64)

                m = pv.PolyData(pts)
                m.lines = lines
                m.point_data["w"] = np.repeat(np.asarray(w, dtype=np.float32), 2)
                return m

            def get_gain_importance(model, features):
                imp = np.zeros(len(features), dtype=np.float32)
                try:
                    booster = getattr(model, "booster_", None)
                    if booster is not None:
                        names = booster.feature_name()
                        gains = booster.feature_importance(importance_type="gain")
                        d = dict(zip(names, gains))
                        for i, f in enumerate(features):
                            imp[i] = float(d.get(f, 0.0))
                    else:
                        gains = getattr(model, "feature_importances_", None)
                        if gains is not None:
                            imp[:] = np.asarray(gains[: len(features)], dtype=np.float32)
                except Exception:
                    pass
                return imp

            gain_cls = get_gain_importance(cls_model, features)
            gain_reg = get_gain_importance(reg_model, features)

            # -----------------------
            # SHAP (best effort)
            # -----------------------
            shap_ok = False
            shap_cls = np.zeros(len(features), dtype=np.float32)
            shap_reg = np.zeros(len(features), dtype=np.float32)

            def try_compute_shap():
                nonlocal shap_ok, shap_cls, shap_reg
                try:
                    import shap
                except Exception:
                    shap_ok = False
                    return

                try:
                    Xs = X_train[features].copy()
                except Exception:
                    Xs = pd.DataFrame(X_train, columns=features).copy()

                n_sample = min(800, len(Xs))
                if n_sample < 80:
                    shap_ok = False
                    return
                Xs = Xs.sample(n=n_sample, random_state=42)

                ok_any = False
                # CLS
                try:
                    expl = shap.TreeExplainer(cls_model)
                    sv = expl.shap_values(Xs)
                    if isinstance(sv, list) and len(sv) >= 2:
                        sv = sv[1]
                    shap_cls[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_cls)) > 0)
                except Exception:
                    shap_cls[:] = 0.0

                # REG
                try:
                    expl = shap.TreeExplainer(reg_model)
                    sv = expl.shap_values(Xs)
                    shap_reg[:] = np.mean(np.abs(np.asarray(sv)), axis=0).astype(np.float32)
                    ok_any = ok_any or (float(np.max(shap_reg)) > 0)
                except Exception:
                    shap_reg[:] = 0.0

                shap_ok = bool(ok_any)

            try_compute_shap()

            # -----------------------
            # Scene config
            # -----------------------
            TOPK = 18          # nb features affichées par arc (CLS + REG)
            LABEL_TOP = 10     # nb labels par arc
            NODE_R = 0.20

            pv.set_plot_theme("dark")
            plotter = pv.Plotter(window_size=(1600, 900))
            plotter.set_background("#0b1220")

            # pipeline positions
            P_OUT   = np.array([-9.0, 0.0, -2.6], dtype=np.float32)
            P_RULES = np.array([-6.0, 0.0, -1.0], dtype=np.float32)
            P_COMB  = np.array([-3.0, 0.0,  0.4], dtype=np.float32)
            P_CLS   = np.array([ 0.0, -1.2,  1.8], dtype=np.float32)
            P_REG   = np.array([ 0.0, +1.2,  1.8], dtype=np.float32)

            def add_block(center, size=(2.6, 1.0, 0.9), label=""):
                box = pv.Cube(center=center, x_length=size[0], y_length=size[1], z_length=size[2])
                box.point_data["v"] = np.full(box.n_points, 1.0, dtype=np.float32)
                plotter.add_mesh(box, scalars="v", cmap="turbo", opacity=0.90, show_edges=True)
                plotter.add_point_labels([center], [label], font_size=14, text_color="white",
                                        point_color="white", shape=False)

            def add_arrow(p0, p1):
                arr = pv.Arrow(start=p0, direction=(p1 - p0),
                            tip_length=0.25, tip_radius=0.07, shaft_radius=0.03, scale="auto")
                plotter.add_mesh(arr, color="white", opacity=0.9)

            add_block(P_CLS, label="LGBM CLS\nP(y>0)")
            add_block(P_REG, label="LGBM REG\nE[y|y>0)")
            add_block(P_COMB, label="Combine\np*amt")
            add_block(P_RULES, label="Rules\nWE/Férié=0\ncarry")
            add_block(P_OUT, label="Output\npred/p10/p90")

            add_arrow(P_CLS, P_COMB)
            add_arrow(P_REG, P_COMB)
            add_arrow(P_COMB, P_RULES)
            add_arrow(P_RULES, P_OUT)

            # -----------------------
            # Arcs placement (2 arcs)
            # -----------------------
            def arc_points(n, z_shift=0.0, y_shift=0.0):
                ang = np.linspace(-1.05, 1.05, n)
                r = 6.2
                x = 6.6 + 0.28 * np.cos(ang)
                y = (r * np.cos(ang) * 0.18) + y_shift
                z = (r * np.sin(ang) * 0.55) + 2.2 + z_shift
                return np.stack([x, y, z], axis=1).astype(np.float32)

            # -----------------------
            # State
            # -----------------------
            state = {
                "edge_mode": 3,          # 1=CLS 2=REG 3=BOTH
                "score_mode": ("shap" if shap_ok else "gain"),  # gain/shap
                "diff_mode": False,      # False: turbo on v ; True: diverging on diff
            }

            a_cls_nodes = None
            a_reg_nodes = None
            a_cls_edges = None
            a_reg_edges = None
            label_actors = []

            def clear_labels():
                nonlocal label_actors
                for act in label_actors:
                    try:
                        plotter.remove_actor(act)
                    except Exception:
                        pass
                label_actors = []

            def scores():
                if state["score_mode"] == "shap" and shap_ok:
                    s_cls = shap_cls.copy()
                    s_reg = shap_reg.copy()
                else:
                    s_cls = gain_cls.copy()
                    s_reg = gain_reg.copy()
                return s_cls, s_reg

            def hud_text():
                s = f"Edges 1/2/3=CLS/REG/BOTH | Score G/S=GAIN/SHAP | D=DIFF"
                s += f"  ||  score={state['score_mode'].upper()}  diff={'ON' if state['diff_mode'] else 'OFF'}"
                if state["score_mode"] == "shap" and not shap_ok:
                    s += " (SHAP indispo -> GAIN)"
                return s

            plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            # -----------------------
            # Rebuild
            # -----------------------
            def rebuild():
                nonlocal a_cls_nodes, a_reg_nodes, a_cls_edges, a_reg_edges

                s_cls, s_reg = scores()

                # TOPK CLS et TOPK REG (union)
                idx_cls = np.argsort(-s_cls)[: min(TOPK, len(features))]
                idx_reg = np.argsort(-s_reg)[: min(TOPK, len(features))]

                f_cls = [features[i] for i in idx_cls]
                f_reg = [features[i] for i in idx_reg]

                v_cls = s_cls[idx_cls]
                v_reg = s_reg[idx_reg]

                # node colors:
                # - normal : importance normalisée
                # - diff : (reg - cls) normalisé signé
                if state["diff_mode"]:
                    # pour diff, on calcule diff sur l'ensemble des features de chaque arc
                    d_cls = (s_reg[idx_cls] - s_cls[idx_cls])  # sur les features TOP CLS
                    d_reg = (s_reg[idx_reg] - s_cls[idx_reg])  # sur les features TOP REG
                    c_cls = _norm_signed(d_cls)  # [-1,1]
                    c_reg = _norm_signed(d_reg)  # [-1,1]
                else:
                    c_cls = _norm01(v_cls)       # [0,1]
                    c_reg = _norm01(v_reg)       # [0,1]

                # edges weights (normalisés séparément)
                w_cls = _norm01(v_cls)
                w_reg = _norm01(v_reg)

                pts_cls = arc_points(len(f_cls), z_shift=-0.9, y_shift=-0.25)
                pts_reg = arc_points(len(f_reg), z_shift=+0.9, y_shift=+0.25)

                # nodes glyphs
                sphere = pv.Sphere(radius=NODE_R, theta_resolution=16, phi_resolution=16)

                cloud_cls = pv.PolyData(pts_cls)
                cloud_reg = pv.PolyData(pts_reg)

                # scalar name
                scalar_name = "d" if state["diff_mode"] else "v"
                cloud_cls[scalar_name] = c_cls.astype(np.float32)
                cloud_reg[scalar_name] = c_reg.astype(np.float32)

                glyph_cls = cloud_cls.glyph(scale=False, geom=sphere, orient=False)
                glyph_reg = cloud_reg.glyph(scale=False, geom=sphere, orient=False)

                # edges meshes (feature -> model)
                PCLS = np.repeat(P_CLS[None, :], len(f_cls), axis=0)
                PREG = np.repeat(P_REG[None, :], len(f_reg), axis=0)

                e_cls = build_lines(pts_cls, PCLS, w_cls)
                e_reg = build_lines(pts_reg, PREG, w_reg)

                # create/update actors
                if a_cls_nodes is None:
                    if state["diff_mode"]:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_cls_nodes = plotter.add_mesh(glyph_cls, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_cls_nodes.mapper.dataset.shallow_copy(glyph_cls)

                if a_reg_nodes is None:
                    if state["diff_mode"]:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="coolwarm",
                                                    clim=[-1, 1], opacity=0.98, smooth_shading=True)
                    else:
                        a_reg_nodes = plotter.add_mesh(glyph_reg, scalars=scalar_name, cmap="turbo",
                                                    opacity=0.98, smooth_shading=True)
                else:
                    a_reg_nodes.mapper.dataset.shallow_copy(glyph_reg)

                if a_cls_edges is None:
                    a_cls_edges = plotter.add_mesh(e_cls, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_cls_edges.mapper.dataset.shallow_copy(e_cls)

                if a_reg_edges is None:
                    a_reg_edges = plotter.add_mesh(e_reg, scalars="w", cmap="turbo",
                                                opacity=0.10, line_width=2, lighting=False)
                else:
                    a_reg_edges.mapper.dataset.shallow_copy(e_reg)

                # edge visibility
                if state["edge_mode"] == 1:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(False)
                elif state["edge_mode"] == 2:
                    a_cls_edges.SetVisibility(False)
                    a_reg_edges.SetVisibility(True)
                else:
                    a_cls_edges.SetVisibility(True)
                    a_reg_edges.SetVisibility(True)

                # make edge thickness depend on mean importance (simple)
                # (PyVista can't vary per-line width easily, but you can tune global)
                a_cls_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_cls)))
                a_reg_edges.GetProperty().SetLineWidth(float(1.0 + 4.0 * np.mean(w_reg)))

                # labels
                clear_labels()
                k1 = min(LABEL_TOP, len(f_cls))
                k2 = min(LABEL_TOP, len(f_reg))
                if k1 > 0:
                    act = plotter.add_point_labels(pts_cls[:k1], [f"CLS: {x}" for x in f_cls[:k1]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)
                if k2 > 0:
                    act = plotter.add_point_labels(pts_reg[:k2], [f"REG: {x}" for x in f_reg[:k2]],
                                                font_size=10, text_color="white",
                                                point_color="white", shape=False)
                    if act is not None:
                        label_actors.append(act)

                plotter.add_text(hud_text(), color="white", font_size=11, position="upper_left", name="hud")

            rebuild()

            # -----------------------
            # Key events
            # -----------------------
            def set_edge_mode(m):
                state["edge_mode"] = m
                rebuild()

            def set_score_mode(mode):
                if mode == "shap" and not shap_ok:
                    state["score_mode"] = "gain"
                else:
                    state["score_mode"] = mode
                rebuild()

            def toggle_diff():
                state["diff_mode"] = not state["diff_mode"]
                rebuild()

            plotter.add_key_event("1", lambda: set_edge_mode(1))
            plotter.add_key_event("2", lambda: set_edge_mode(2))
            plotter.add_key_event("3", lambda: set_edge_mode(3))
            plotter.add_key_event("g", lambda: set_score_mode("gain"))
            plotter.add_key_event("G", lambda: set_score_mode("gain"))
            plotter.add_key_event("s", lambda: set_score_mode("shap"))
            plotter.add_key_event("S", lambda: set_score_mode("shap"))
            plotter.add_key_event("d", toggle_diff)
            plotter.add_key_event("D", toggle_diff)

            plotter.add_axes()
            plotter.camera_position = "iso"
            plotter.camera.zoom(1.35)
            plotter.show()

        # ============ ZONE RÉSULTATS ============
        graph_widgets = []  # kpi, graphe 1, tableaux

        graph2_container = tk.Frame(scrollable_frame, bg="#00122e")      # déjà ok chez toi
        monthly_container = tk.Frame(scrollable_frame, bg="#00122e")     # déjà ok chez toi

        valid_container = tk.Frame(scrollable_frame, bg="#00122e")       # VALID
        valid_graph_widget = None
        valid_container_packed = False

        
        mc_container = tk.Frame(scrollable_frame, bg="#00122e")          # NEW: Monte Carlo paths
        mc_graph_widget = None
        mc_container_packed = False
        ia_valid_widget = None
        monthly_graph_widget = None
        monthly_container_packed = False
        # --- Containers nouveaux ---
        mc_fan_container = tk.Frame(scrollable_frame, bg="#00122e")
        mc_fan_widget = None
        mc_fan_container_packed = False

        resid_container = tk.Frame(scrollable_frame, bg="#00122e")
        resid_widget = None
        resid_container_packed = False

        mc3d_container_packed = False
        mc3d_widget = None
        mc3d_container = tk.Frame(scrollable_frame, bg="#00122e")

        cm_container = tk.Frame(scrollable_frame, bg="#00122e")
        cm_widget = None
        cm_container_packed = False

        cm_pred_container = tk.Frame(scrollable_frame, bg="#00122e")
        cm_pred_widget = None
        cm_pred_container_packed = False

        # globals/etat
        current_valid_df = None  # si tu en as besoin ailleurs (optionnel)

        export_tools_frame = ctk.CTkFrame(scrollable_frame, fg_color="#00122e", corner_radius=0)
        export_tools_frame.pack(fill="x", padx=10, pady=(0, 10))

        export_graph2_button = ctk.CTkButton(
            export_tools_frame,
            text="📤 Export Graph2 (Réel / IA / Profils) - Excel",
            width=320, height=34,
            corner_radius=10,
            fg_color="#2563eb", hover_color="#1d4ed8",
            text_color="white",
            state="disabled",
            command=_export_graph2_to_excel
        )
        export_graph2_button.pack(anchor="w")

        btn_3d = ctk.CTkButton(
            export_tools_frame,
            text="🧊 Voir 3D (PyVista)",
            width=200, height=34,
            corner_radius=10,
            fg_color="#7c3aed", hover_color="#6d28d9",
            text_color="white",
            state="disabled",
            command=_show_pyvista_3d
        )
        btn_3d.pack(anchor="w", pady=(6, 0))

        # ---------- helpers graphiques ----------
        GRAPH_EXPLANATIONS = {
            "graph2": (
                "Vue détaillée – réel, prévision IA et profils",
                "Ce graphique compare l’évolution journalière de l’année cible entre le réel observé, la prévision IA et les profils activés. "
                "Il permet de juger la cohérence globale de la trajectoire prévue sur l’ensemble de l’année. "
                "Pour le lire, compare d’abord la forme de la courbe IA au réel, puis observe l’écart avec les profils pour situer la prévision par rapport aux scénarios de référence."
            ),
            "monthly": (
                "Vue mensuelle – cumuls par mois",
                "Ce graphique regroupe les valeurs par mois pour comparer les cumuls mensuels entre le réel, la prévision IA et les profils éventuels. "
                "Il sert à vérifier si la répartition annuelle reste cohérente à un niveau plus agrégé que le journalier. "
                "Pour le lire, regarde la hauteur des barres mois par mois afin d’identifier rapidement les périodes sur-prévues, sous-prévues ou atypiques."
            ),
            "valid": (
                "Validation interne – réel vs prédiction",
                "Ce graphique montre la performance du modèle sur la période de validation interne en comparant les valeurs réelles et les valeurs prédites. "
                "Il sert à visualiser la qualité d’ajustement avant utilisation sur l’année cible. "
                "Pour le lire, regarde si la courbe prédite suit correctement le niveau, les variations et les points de rupture observés sur la série réelle."
            ),
            "mc_fan": (
                "Fan chart Monte Carlo – incertitude de prévision",
                "Ce graphique représente la prévision centrale ainsi qu’une bande d’incertitude issue des simulations Monte Carlo. "
                "Il permet d’évaluer non seulement la valeur attendue mais aussi l’étendue plausible des résultats autour de cette trajectoire. "
                "Pour le lire, suis la courbe centrale puis observe la largeur de la bande : plus elle est large, plus l’incertitude estimée est importante."
            ),
            "residuals": (
                "Résidus – erreur journalière de la prévision",
                "Ce graphique montre l’écart entre le réel et la prévision centrale jour par jour, ainsi qu’une moyenne glissante pour en résumer la tendance. "
                "Il permet de repérer les périodes où le modèle surestime ou sous-estime durablement la série. "
                "Pour le lire, regarde la position par rapport à zéro : au-dessus le réel dépasse la prévision, en dessous la prévision dépasse le réel."
            ),
            "mc_paths": (
                "Trajectoires Monte Carlo – scénarios simulés",
                "Ce graphique affiche un ensemble de trajectoires simulées autour de la prévision finale afin d’illustrer la dispersion possible des résultats. "
                "Il sert à visualiser la variabilité potentielle de la trajectoire au fil du temps, au-delà de la seule courbe centrale. "
                "Pour le lire, observe la densité et l’ouverture du faisceau de courbes : plus il s’élargit, plus l’incertitude augmente sur la période."
            ),
            "mc_3d": (
                "Surface 3D Monte Carlo – distribution des quantiles",
                "Ce graphique représente la distribution simulée de la prévision dans le temps selon différents quantiles, sous forme de surface 3D. "
                "Il permet d’analyser simultanément la date, le niveau prévu et la structure de dispersion des simulations. "
                "Pour le lire, suis la surface selon l’axe des quantiles : plus elle est épaisse ou étirée, plus l’éventail des valeurs possibles est large."
            ),
            "cm": (
                "Matrice de confusion – qualité de la classification",
                "Ce graphique évalue la capacité du classifieur à distinguer les jours prédits comme nuls ou non nuls sur l’échantillon de validation. "
                "Il sert à comprendre si le modèle détecte correctement la présence ou l’absence de trafic attendu. "
                "Pour le lire, regarde la diagonale principale : plus elle concentre les valeurs, plus la classification est correcte ; les cases hors diagonale représentent les erreurs."
            ),
        }
        
        def _add_section_separator():
            """Ajoute une barre blanche pleine largeur pour structurer la page."""
            sep = tk.Frame(scrollable_frame, bg="white", height=2)
            sep.pack(fill="x", padx=10, pady=(12, 10))
            graph_widgets.append(sep)
            return sep

        def _add_graph_explanation(title: str, text: str):
            """Ajoute un bloc texte explicatif au-dessus d'un graphe."""
            box = ctk.CTkFrame(
                scrollable_frame,
                fg_color="#0b1730",
                corner_radius=12,
                border_width=1,
                border_color="#223658"
            )
            box.pack(fill="x", padx=10, pady=(6, 8))
            graph_widgets.append(box)

            ctk.CTkLabel(
                box,
                text=title,
                font=("Segoe UI Semibold", 14, "bold"),
                text_color="white",
                anchor="w",
                justify="left"
            ).pack(fill="x", padx=14, pady=(10, 4))

            ctk.CTkLabel(
                box,
                text=text,
                font=("Segoe UI", 12),
                text_color="#d7e3f4",
                anchor="w",
                justify="left",
                wraplength=1200
            ).pack(fill="x", padx=14, pady=(0, 10))

            return box
        
        def _clear_graph_widgets():
            """Supprime les widgets graphiques (sauf le conteneur du graphe 2)."""
            nonlocal graph_widgets
            for w in graph_widgets:
                try:
                    w.destroy()
                except Exception:
                    pass
            graph_widgets = []

        def _format_time_axis_like_graph2(ax, dates):
            """Applique le même format d'axe X que _redraw_graph2()."""
            import matplotlib.dates as mdates
            import pandas as pd

            d = pd.to_datetime(dates)
            if len(d) < 2:
                return

            # Même logique que graph2: rotation 30 + couleur blanche
            ax.tick_params(axis='x', colors="white", rotation=30)

            # Force les limites exactes sur les données
            ax.set_xlim(d.min(), d.max())

            # Locator "propre" et stable (évite un x-axis bizarre)
            # AutoDateLocator choisit un nombre de ticks raisonnable
            locator = mdates.AutoDateLocator(minticks=6, maxticks=12)
            formatter = mdates.ConciseDateFormatter(locator)

            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(formatter)

            # Petite marge visuelle (optionnel)
            ax.margins(x=0.01)

        def _redraw_graph2():
            """Redessine le graphe détaillé N+1 avec réel / IA / profils cochés."""
            nonlocal ia_graph2_widget, graph2_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            if current_pred_df is None and current_real_target_df is None and not ia_profils_names:
                return

            if not graph2_container_packed:
                graph2_container.pack(pady=10, fill="both", expand=True)
                graph2_container_packed = True

            for child in graph2_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            ia_graph2_widget = None

            fig2, ax2 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax2.set_facecolor("#00122e")

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            flux = selected_flux.get()

            if current_real_target_df is not None and not current_real_target_df.empty:
                ax2.plot(
                    current_real_target_df["date"], current_real_target_df["y"],
                    label=f"Réel {target_year}", linewidth=2, color="#5DADE2"
                )

            if current_pred_df is not None and not current_pred_df.empty:
                ax2.plot(
                    current_pred_df["date"], current_pred_df["pred_value"],
                    label=f"Prévision IA {target_year}",
                    linewidth=2, linestyle="--", color="#F4D03F"
                )

            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = [pd.to_datetime(d) for d in ia_profils_dates]
                palette = plt.cm.tab10.colors
                prof_idx = 0
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue
                    y_prof = [_to_float_or_nan(v) for v in serie]
                    ax2.plot(
                        dates_prof, y_prof,
                        label=f"Profil '{name}' {target_year}",
                        linewidth=1.8,
                        linestyle="-.",
                        color=palette[prof_idx % len(palette)]
                    )
                    prof_idx += 1

            filiale_for_title = current_filiale_name or selected_filiale.get()
            ax2.set_title(f"{flux} – année {target_year} – {filiale_for_title}", color="white", fontsize=14)
            ax2.set_xlabel("Date", color="white", fontsize=12)
            ax2.set_ylabel("Valeur", color="white", fontsize=12)
            ax2.tick_params(axis='x', colors="white", rotation=30)
            # après avoir choisi les séries tracées, récupère les dates "d" de référence
            if current_pred_df is not None and not current_pred_df.empty:
                dref = current_pred_df["date"]
            elif current_real_target_df is not None and not current_real_target_df.empty:
                dref = current_real_target_df["date"]
            else:
                dref = dates_prof

            _format_time_axis_like_graph2(ax2, dref)
            ax2.tick_params(axis='y', colors="white")
            
            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["graph2"])
            
            canvas_fig2 = FigureCanvasTkAgg(fig2, master=graph2_container)
            canvas_fig2.draw()
            ia_graph2_widget = canvas_fig2.get_tk_widget()
            ia_graph2_widget.pack(fill="both", expand=True)
            plt.close(fig2)

        def _redraw_monthly_graph():
            """Ton code existant mensuel (inchangé) — je garde le tien en l'état."""
            nonlocal monthly_graph_widget, monthly_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            if current_pred_df is None or current_pred_df.empty:
                return

            if not monthly_container_packed:
                monthly_container.pack(pady=10, fill="both", expand=True)
                monthly_container_packed = True

            for child in monthly_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            monthly_graph_widget = None

            import numpy as np
            from matplotlib.ticker import FuncFormatter

            target_year = current_target_year
            try:
                if target_year is None and annees_var.get():
                    target_year = int(annees_var.get()) + 1
            except Exception:
                pass

            color_real = "#1f77b4"
            color_pred = "#F4D03F"
            profile_colors = [
                "#e74c3c", "#9b59b6", "#2ecc71", "#1abc9c", "#f1c40f",
                "#d35400", "#8e44ad", "#27ae60", "#16a085", "#c0392b",
                "#7f8c8d", "#95a5a6", "#34495e", "#bdc3c7", "#f39c12",
                "#c27ba0", "#76d7c4", "#7dcea0", "#af7ac5", "#5dade2",
            ]

            df_pred = current_pred_df.copy()
            df_pred["month"] = df_pred["date"].dt.month
            pred_monthly = (df_pred.groupby("month", as_index=False)["pred_value"].sum()
                            ).rename(columns={"pred_value": "pred_value"})

            if current_real_target_df is not None and not current_real_target_df.empty:
                df_real = current_real_target_df.copy()
                df_real["month"] = df_real["date"].dt.month
                real_monthly = (df_real.groupby("month", as_index=False)["y"].sum()
                                ).rename(columns={"y": "real_value"})

                monthly_cmp = pd.merge(real_monthly, pred_monthly, on="month", how="outer").fillna(0.0)
            else:
                monthly_cmp = pred_monthly.copy()
                monthly_cmp["real_value"] = 0.0

            monthly_cmp = monthly_cmp.sort_values("month")

            active_profiles = []
            if ia_profils_names and ia_profils_series and ia_profils_dates:
                dates_prof = pd.to_datetime(ia_profils_dates)
                df_day = pd.DataFrame({"date": dates_prof, "month": dates_prof.month})

                if current_real_target_df is not None and not current_real_target_df.empty:
                    real_map = dict(zip(current_real_target_df["date"], current_real_target_df["y"]))
                    df_day["real"] = df_day["date"].map(real_map).fillna(0.0)
                else:
                    df_day["real"] = 0.0

                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    vals = [_to_float_or_nan(v) for v in serie]
                    df_day["prev"] = vals
                    df_day["comb"] = df_day["prev"]
                    df_day.loc[df_day["comb"].isna(), "comb"] = df_day["real"]

                    mois_avec_prev = set(df_day.loc[df_day["prev"].notna(), "month"].unique())
                    mois_tous = set(monthly_cmp["month"].unique())
                    mois_masques = sorted(mois_tous - mois_avec_prev)

                    df_day["comb_masked"] = df_day["comb"]
                    df_day.loc[df_day["month"].isin(mois_masques), "comb_masked"] = np.nan

                    prof_month = df_day.groupby("month")["comb_masked"].sum(min_count=1)
                    yvals = [prof_month.get(m, np.nan) for m in monthly_cmp["month"]]
                    active_profiles.append((name, yvals))

            x = np.arange(len(monthly_cmp))
            nb_series = 2 + len(active_profiles)
            width = 0.8 / max(nb_series, 1)
            offsets = (np.arange(nb_series) - (nb_series - 1) / 2.0) * width

            fig_m, ax_m = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax_m.set_facecolor("#00122e")

            ax_m.bar(x + offsets[0], monthly_cmp["real_value"], width, label=f"Réel {target_year}", color=color_real)
            ax_m.bar(x + offsets[1], monthly_cmp["pred_value"], width, label=f"Prévision IA {target_year}", color=color_pred)

            if active_profiles:
                for i, (name, y_vals) in enumerate(active_profiles):
                    serie_idx = 2 + i
                    profile_color = profile_colors[i % len(profile_colors)]
                    ax_m.bar(x + offsets[serie_idx], y_vals, width, label=f"Profil '{name}' {target_year}", color=profile_color)

            month_nums = monthly_cmp["month"].astype(int)
            month_dates = pd.to_datetime({"year": [target_year] * len(month_nums), "month": month_nums, "day": 1})
            mois_labels = month_dates.dt.strftime("%Y-%m")

            ax_m.set_xticks(x)
            ax_m.set_xticklabels(mois_labels, rotation=45, ha="right", fontsize=9, color="white")

            ax_m.set_xlabel("Mois", fontsize=11, color="white")
            ax_m.set_ylabel("Valeur cumulée", fontsize=11, color="white")
            ax_m.tick_params(axis="y", colors="white")

            def _fmt_milliers(val, pos):
                try:
                    return f"{int(val):,}".replace(",", " ")
                except Exception:
                    return ""
            ax_m.yaxis.set_major_formatter(FuncFormatter(_fmt_milliers))

            filiale_for_title = current_filiale_name or selected_filiale.get()
            flux = selected_flux.get()
            ax_m.set_title(f"{filiale_for_title} - {flux} - {target_year}", fontsize=13, fontweight="bold", color="white", pad=12)

            fig_m.patch.set_edgecolor("#00122e")
            fig_m.patch.set_linewidth(0)
            for spine in ax_m.spines.values():
                spine.set_visible(False)

            leg_m = ax_m.legend(frameon=False, facecolor="#00122e")
            for text in leg_m.get_texts():
                text.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["monthly"])

            canvas_m = FigureCanvasTkAgg(fig_m, master=monthly_container)
            canvas_m.draw()
            monthly_graph_widget = canvas_m.get_tk_widget()
            monthly_graph_widget.pack(fill="both", expand=True)
            plt.close(fig_m)

        def _redraw_valid_graph():
            """Redessine le graphe VALID (réel vs préd) en prenant l'année depuis valid_vis."""
            nonlocal ia_valid_widget, valid_container_packed
            nonlocal current_filiale_name

            df_valid_vis = model_artifacts.get("valid_vis", None)
            if df_valid_vis is None or df_valid_vis.empty:
                return

            if not valid_container_packed:
                valid_container.pack(pady=10, fill="both", expand=True)
                valid_container_packed = True

            for child in valid_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            valid_graph_widget = None

            # -------- année VALID : robuste --------
            valid_year = None
            if "year_target" in df_valid_vis.columns and df_valid_vis["year_target"].notna().any():
                try:
                    valid_year = int(df_valid_vis["year_target"].dropna().astype(int).mode().iloc[0])
                except Exception:
                    valid_year = None

            # dates
            if "date_tgt" in df_valid_vis.columns:
                d = pd.to_datetime(df_valid_vis["date_tgt"])
            elif "date" in df_valid_vis.columns:
                d = pd.to_datetime(df_valid_vis["date"])
            else:
                return

            if valid_year is None:
                try:
                    valid_year = int(d.dt.year.mode().iloc[0])
                except Exception:
                    valid_year = ""

            if "y_true" not in df_valid_vis.columns or "y_pred" not in df_valid_vis.columns:
                return

            y_true = df_valid_vis["y_true"].astype(float).values
            y_pred = df_valid_vis["y_pred"].astype(float).values

            # -------- plot --------
            figv, axv = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            axv.set_facecolor("#00122e")

            axv.plot(d, y_true, label="VALID réel", linewidth=2, color="#5DADE2")
            axv.plot(d, y_pred, label="VALID prédiction", linewidth=2, linestyle="--", color="#F4D03F")

            # Bande P10–P90 (si calib dispo)
            qs = model_artifacts.get("calib_qs", None)
            q_low_by_bin = model_artifacts.get("calib_q_low_by_bin", None)
            q_high_by_bin = model_artifacts.get("calib_q_high_by_bin", None)

            if qs is not None and q_low_by_bin is not None and q_high_by_bin is not None:
                try:
                    import numpy as np
                    qs = np.asarray(qs, dtype=float)
                    q_low_by_bin = np.asarray(q_low_by_bin, dtype=float)
                    q_high_by_bin = np.asarray(q_high_by_bin, dtype=float)

                    n_bins = len(q_low_by_bin)
                    bins = np.digitize(y_pred, qs[1:-1], right=True)
                    bins = np.clip(bins, 0, n_bins - 1)

                    p10 = np.clip(y_pred + q_low_by_bin[bins], 0.0, None)
                    p90 = np.clip(y_pred + q_high_by_bin[bins], 0.0, None)

                    axv.fill_between(d, p10, p90, alpha=0.20, label="Bande VALID P10–P90")
                except Exception:
                    pass

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            axv.set_title(f"VALID – {filiale} – année {valid_year}", color="white", fontsize=14)
            axv.set_xlabel("Date", color="white", fontsize=12)
            axv.set_ylabel("Valeur", color="white", fontsize=12)
            axv.set_xlabel("Date", color="white", fontsize=12)
            axv.set_ylabel("Valeur", color="white", fontsize=12)

            _format_time_axis_like_graph2(axv, d)
            axv.tick_params(axis="y", colors="white")

            axv.legend(facecolor="#00122e", edgecolor="white", labelcolor="white")

            # IMPORTANT : évite “mauvaises limites” d’axe
            try:
                axv.set_xlim(d.min(), d.max())
            except Exception:
                pass
            
            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["valid"])

            canvas_v = FigureCanvasTkAgg(figv, master=valid_container)
            canvas_v.draw()
            valid_graph_widget = canvas_v.get_tk_widget()
            valid_graph_widget.pack(fill="both", expand=True)
            plt.close(figv)

        def _redraw_mc_fan_graph():
            """Fan chart MC: P10/P50/P90 + réel année cible si dispo, axe x limité à l'année cible."""
            nonlocal mc_fan_widget, mc_fan_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            # source forecast
            if current_pred_df is None or current_pred_df.empty:
                return

            dfp = current_pred_df.copy()
            if "date" not in dfp.columns:
                return
            if not all(c in dfp.columns for c in ["pred_p10", "pred_value", "pred_p90"]):
                return

            d_all = pd.to_datetime(dfp["date"])
            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(d_all.dt.year).mode().iloc[0])

            # filtre strict année cible
            mask = (d_all.dt.year == ty)
            if mask.sum() < 2:
                mask = np.ones(len(d_all), dtype=bool)

            dfp = dfp.loc[mask].copy()
            d = pd.to_datetime(dfp["date"])

            if not mc_fan_container_packed:
                mc_fan_container.pack(pady=10, fill="both", expand=True)
                mc_fan_container_packed = True

            for child in mc_fan_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc_fan_widget = None

            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            # réel si dispo
            if current_real_target_df is not None and not current_real_target_df.empty:
                d_real = pd.to_datetime(current_real_target_df["date"])
                mask_r = (d_real.dt.year == ty)
                rr = current_real_target_df.loc[mask_r].copy()
                if not rr.empty:
                    ax.plot(rr["date"], rr["y"], linewidth=2.0, color="#5DADE2", label=f"Réel {ty}")

            ax.plot(d, dfp["pred_value"].astype(float).values, linewidth=2.2, linestyle="--",
                    color="#F4D03F", label=f"P50 {ty}")
            ax.fill_between(d,
                            dfp["pred_p10"].astype(float).values,
                            dfp["pred_p90"].astype(float).values,
                            alpha=0.22, label="Bande P10–P90")

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Monte Carlo – fan chart – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Valeur", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            # même “fenêtre année” que graph2
            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                ax.set_xlim(d.min(), d.max())

            # si tu as déjà ton helper :
            _format_time_axis_like_graph2(ax, d)

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_fan"])

            canvas = FigureCanvasTkAgg(fig, master=mc_fan_container)
            canvas.draw()
            mc_fan_widget = canvas.get_tk_widget()
            mc_fan_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_residuals_graph():
            """Résidus sur l'année cible: resid = réel - P50 + moyenne glissante 7j.
            + Marqueurs rouges sur les jours 'alerte' (même logique que _compute_metrics: sMAPE > rel_seuil).
            """
            nonlocal resid_widget, resid_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            if current_pred_df is None or current_pred_df.empty:
                return
            if current_real_target_df is None or current_real_target_df.empty:
                return
            if "pred_value" not in current_pred_df.columns:
                return

            # --- paramètres alignés sur _compute_metrics ---
            rel_seuil = 40.0       # seuil alerte (%)
            min_abs_ref = 10.0     # ignore si |ref| < 10
            eps = 1e-9
            mark_only_peaks = True  # mets False si tu veux un point rouge sur TOUS les jours en alerte

            dfp = current_pred_df[["date", "pred_value"]].copy()
            dfr = current_real_target_df[["date", "y"]].copy()
            dfp["date"] = pd.to_datetime(dfp["date"])
            dfr["date"] = pd.to_datetime(dfr["date"])

            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(dfp["date"].dt.year).mode().iloc[0])

            df = pd.merge(dfr, dfp, on="date", how="inner").sort_values("date")
            df = df[df["date"].dt.year == ty].copy()
            if df.shape[0] < 5:
                return

            # --- résidus ---
            y = df["y"].astype(float).values
            p = df["pred_value"].astype(float).values
            df["resid"] = y - p
            df["roll7"] = df["resid"].rolling(7, min_periods=1).mean()

            # --- logique alerte (sMAPE) comme _compute_metrics ---
            abs_err = np.abs(p - y)
            denom = (np.abs(y) + np.abs(p))
            denom = np.clip(denom, eps, None)
            smape = (2.0 * abs_err / denom) * 100.0
            df["smape"] = smape

            # filtre min_abs_ref (même philosophie que _compute_metrics)
            valid_ref = np.abs(y) >= float(min_abs_ref)
            df["is_alert"] = (df["smape"] > float(rel_seuil)) & valid_ref

            # --- option: ne marquer que les pics locaux parmi les alertes ---
            if mark_only_peaks:
                rabs = np.abs(df["resid"].values)
                # pic local si > voisin gauche et >= voisin droite (simple et robuste)
                is_peak = np.zeros(len(df), dtype=bool)
                if len(df) >= 3:
                    is_peak[1:-1] = (rabs[1:-1] > rabs[:-2]) & (rabs[1:-1] >= rabs[2:])
                df["is_peak"] = is_peak
                df["mark_red"] = df["is_alert"] & df["is_peak"]
            else:
                df["mark_red"] = df["is_alert"]

            # --- UI pack ---
            if not resid_container_packed:
                resid_container.pack(pady=10, fill="both", expand=True)
                resid_container_packed = True

            for child in resid_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            resid_widget = None

            # --- plot ---
            fig, ax = plt.subplots(figsize=(11, 4.0), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            ax.axhline(0.0, linewidth=1.2, alpha=0.5)
            ax.plot(df["date"], df["resid"], linewidth=1.6, color="#5DADE2", label="Résidu (réel - P50)")
            ax.plot(df["date"], df["roll7"], linewidth=2.2, linestyle="--", color="#F4D03F", label="Moyenne 7j")

            # --- points rouges (alertes) ---
            df_red = df[df["mark_red"]].copy()
            if not df_red.empty:
                ax.scatter(
                    df_red["date"], df_red["resid"],
                    s=42, marker="o",
                    color="#ff3b30", edgecolors="white", linewidths=0.6,
                    label=f"Alerte (sMAPE>{rel_seuil:.0f}%)"
                )

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Résidus – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Erreur", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                pass

            _format_time_axis_like_graph2(ax, df["date"])

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["residuals"])

            canvas = FigureCanvasTkAgg(fig, master=resid_container)
            canvas.draw()
            resid_widget = canvas.get_tk_widget()
            resid_widget.pack(fill="both", expand=True)
            plt.close(fig)

        def _redraw_mc_graph(show_all: bool = False, max_paths: int = 300):
            """Graphe séparé : MC trajectories + trajectoire finale, X limité à l'année cible."""
            nonlocal mc_graph_widget, mc_container_packed
            nonlocal current_target_year, current_filiale_name

            sims = model_artifacts.get("mc_sims", None)      # (B,N) ou None
            dates = model_artifacts.get("mc_dates", None)    # (N,)
            det = model_artifacts.get("mc_det", None)        # (N,)
            p50 = model_artifacts.get("mc_p50", None)        # (N,)

            if dates is None:
                return

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import matplotlib.dates as mdates

            # IMPORTANT: DatetimeIndex -> pas de .dt
            d_all = pd.to_datetime(dates)
            if len(d_all) < 2:
                return

            # année cible
            ty = model_artifacts.get("target_year", None)
            if ty is None:
                ty = current_target_year

            try:
                ty = int(ty)
            except Exception:
                try:
                    ty = int(pd.Series(d_all.year).mode().iloc[0])
                except Exception:
                    ty = int(d_all[0].year)

            # filtre strict sur l'année cible (évite débordement 2026)
            mask_year = (d_all.year == ty)
            if mask_year.sum() < 2:
                mask_year = np.ones(len(d_all), dtype=bool)

            d = d_all[mask_year]

            # sims filtrées
            sims_ok = None
            if sims is not None:
                sims = np.asarray(sims, dtype=float)
                if sims.ndim == 2 and sims.shape[1] == len(d_all):
                    sims_ok = sims[:, mask_year]

            # vecteurs filtrés (trajectoire finale)
            def _vec(v):
                if v is None:
                    return None
                v = np.asarray(v, dtype=float)
                if len(v) != len(d_all):
                    return None
                return v[mask_year]

            det_f = _vec(det)
            p50_f = _vec(p50)

            if not mc_container_packed:
                mc_container.pack(pady=10, fill="both", expand=True)
                mc_container_packed = True

            # clear container
            for child in mc_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc_graph_widget = None

            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            # spaghetti
            if sims_ok is not None:
                B = sims_ok.shape[0]
                if show_all:
                    idx = range(B)
                    alpha = 0.08
                    lw = 0.6
                else:
                    k = int(min(max_paths, B))
                    idx = np.linspace(0, B - 1, num=k, dtype=int)
                    alpha = 0.12
                    lw = 0.7

                for i in idx:
                    ax.plot(d, sims_ok[i, :], linewidth=lw, alpha=alpha)

            # trajectoire finale (P50 prioritaire)
            if p50_f is not None:
                ax.plot(d, p50_f, linewidth=2.6, linestyle="--", color="#F4D03F", label="Trajectoire finale (P50)")
            elif det_f is not None:
                ax.plot(d, det_f, linewidth=2.6, linestyle="-", color="#F4D03F", label="Trajectoire finale (déterministe)")

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"Monte Carlo – trajectoires – {filiale} – {ty}", color="white", fontsize=14)
            ax.set_xlabel("Date", color="white", fontsize=12)
            ax.set_ylabel("Valeur simulée", color="white", fontsize=12)

            ax.tick_params(axis="x", colors="white", rotation=30)
            ax.tick_params(axis="y", colors="white")

            # limites strictes année cible (évite 2026-01)
            try:
                ax.set_xlim(pd.Timestamp(ty, 1, 1), pd.Timestamp(ty, 12, 31))
            except Exception:
                ax.set_xlim(d.min(), d.max())

            # ticks propres
            locator = mdates.AutoDateLocator(minticks=6, maxticks=12)
            ax.xaxis.set_major_locator(locator)
            ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(locator))

            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_paths"])

            canvas = FigureCanvasTkAgg(fig, master=mc_container)
            canvas.draw()
            mc_graph_widget = canvas.get_tk_widget()
            mc_graph_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_mc_3d_fan_surface(
            q_grid=None,
            show_surface=True,
            show_quantile_curves=True,
            alert_rel_seuil=40.0,
            min_abs_ref=10.0,
        ):
            """
            3D fan chart : X=Date, Y=Quantile, Z=Valeur simulée (MC)
            + points rouges sur la courbe P50 si alerte (sMAPE>seuil) vs réel.
            """
            nonlocal mc3d_widget, mc3d_container_packed
            nonlocal current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from mpl_toolkits.mplot3d import Axes3D  # noqa: F401
            import matplotlib.dates as mdates

            sims = model_artifacts.get("mc_sims", None)      # (B,N)
            dates = model_artifacts.get("mc_dates", None)    # (N,)
            if sims is None or dates is None:
                return

            sims = np.asarray(sims, dtype=float)
            if sims.ndim != 2 or sims.shape[1] != len(dates):
                return

            d_all = pd.to_datetime(dates)
            if len(d_all) < 5:
                return

            # année cible
            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(d_all.year).mode().iloc[0])

            mask_year = (d_all.year == ty)
            if mask_year.sum() < 5:
                mask_year = np.ones(len(d_all), dtype=bool)

            d = d_all[mask_year]
            sims_y = sims[:, mask_year]

            # grille de quantiles
            if q_grid is None:
                q_grid = np.array([0.05, 0.10, 0.20, 0.35, 0.50, 0.65, 0.80, 0.90, 0.95], dtype=float)
            q_grid = np.asarray(q_grid, dtype=float)

            # compute quantiles (len(q), N)
            Q = np.quantile(sims_y, q_grid, axis=0)

            # container UI
            if not mc3d_container_packed:
                mc3d_container.pack(pady=10, fill="both", expand=True)
                mc3d_container_packed = True

            for child in mc3d_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            mc3d_widget = None

            # X axis numeric dates
            x_num = mdates.date2num(d.to_pydatetime())
            X, Y = np.meshgrid(x_num, q_grid)  # (len(q), N) after broadcast with Q

            fig = plt.figure(figsize=(11, 5.2), facecolor="#00122e", constrained_layout=True)
            ax = fig.add_subplot(111, projection="3d")
            ax.set_facecolor("#00122e")

            # surface fan
            if show_surface:
                # IMPORTANT: ne pas mettre de couleurs custom si tu veux rester neutre
                ax.plot_surface(X, Y, Q, rstride=1, cstride=1, linewidth=0, antialiased=True, alpha=0.35)

            # courbes quantiles (lisible)
            if show_quantile_curves:
                for i, qq in enumerate(q_grid):
                    lw = 2.2 if np.isclose(qq, 0.50) else 1.0
                    ls = "--" if np.isclose(qq, 0.50) else "-"
                    ax.plot(x_num, np.full_like(x_num, qq, dtype=float), Q[i, :], linewidth=lw, linestyle=ls)

            # points rouges d'alerte sur la courbe P50 (si réel dispo)
            if current_real_target_df is not None and not current_real_target_df.empty:
                dfr = current_real_target_df[["date", "y"]].copy()
                dfr["date"] = pd.to_datetime(dfr["date"])
                dfr = dfr[dfr["date"].dt.year == ty].copy()
                if not dfr.empty:
                    # merge sur dates filtrées
                    dfm = pd.DataFrame({"date": d})
                    dfm = dfm.merge(dfr, on="date", how="left")
                    p50 = Q[np.argmin(np.abs(q_grid - 0.50)), :]
                    dfm["p50"] = p50
                    dfm = dfm.dropna(subset=["y"])

                    y_true = dfm["y"].astype(float).values
                    y_pred = dfm["p50"].astype(float).values
                    abs_err = np.abs(y_pred - y_true)
                    denom = np.clip(np.abs(y_true) + np.abs(y_pred), 1e-9, None)
                    smape = (2.0 * abs_err / denom) * 100.0

                    mask_valid = (np.abs(y_true) >= float(min_abs_ref))
                    is_alert = (smape > float(alert_rel_seuil)) & mask_valid

                    if np.any(is_alert):
                        x_alert = mdates.date2num(dfm.loc[is_alert, "date"].dt.to_pydatetime())
                        z_alert = dfm.loc[is_alert, "p50"].values.astype(float)
                        ax.scatter(
                            x_alert,
                            np.full_like(x_alert, 0.50, dtype=float),
                            z_alert,
                            s=28,
                            color="#ff3b30",
                            depthshade=False
                        )

            # axes labels
            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax.set_title(f"3D Fan chart (MC) – {filiale} – {ty}", color="white", fontsize=13, pad=12)

            ax.set_xlabel("Date", color="white", labelpad=10)
            ax.set_ylabel("Quantile", color="white", labelpad=10)
            ax.set_zlabel("Valeur", color="white", labelpad=8)

            # ticks colors
            ax.tick_params(axis="x", colors="white")
            ax.tick_params(axis="y", colors="white")
            ax.tick_params(axis="z", colors="white")

            # format date ticks
            ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=6, maxticks=10))
            ax.xaxis.set_major_formatter(mdates.ConciseDateFormatter(ax.xaxis.get_major_locator()))

            # limits
            ax.set_xlim(mdates.date2num(pd.Timestamp(ty, 1, 1)), mdates.date2num(pd.Timestamp(ty, 12, 31)))
            ax.set_ylim(float(q_grid.min()), float(q_grid.max()))

            _add_section_separator()
            _add_graph_explanation(*GRAPH_EXPLANATIONS["mc_3d"])

            # canvas
            canvas = FigureCanvasTkAgg(fig, master=mc3d_container)
            canvas.draw()
            mc3d_widget = canvas.get_tk_widget()
            mc3d_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_confusion_matrix_graph():
            """Redessine la matrice de confusion + courbe ROC du classifieur sur VALID."""
            nonlocal cm_widget, cm_container_packed

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from sklearn.metrics import confusion_matrix, roc_curve, auc

            X_valid = model_artifacts.get("X_valid", None)
            y_cls_valid = model_artifacts.get("y_cls_valid", None)
            cls_models = model_artifacts.get("cls_models", None)
            cls_weights = model_artifacts.get("cls_model_weights", None)
            cls_model_names = model_artifacts.get("cls_model_names", None)

            filiale = model_artifacts.get("filiale", "") or ""
            base_year = model_artifacts.get("base_year", None)

            if X_valid is None or y_cls_valid is None or cls_models is None or cls_weights is None:
                return
            if len(X_valid) == 0 or len(y_cls_valid) == 0 or len(cls_models) == 0:
                return

            X_arr = X_valid.values if hasattr(X_valid, "values") else np.asarray(X_valid)
            y_true = np.asarray(y_cls_valid).astype(int)

            try:
                # Ensemble pondéré
                probas = [m.predict_proba(X_arr)[:, 1] for m in cls_models]
                y_prob = np.average(np.vstack(probas), axis=0, weights=np.asarray(cls_weights, dtype=float))
            except Exception:
                return

            threshold = 0.5
            y_pred = (y_prob >= threshold).astype(int)

            # Matrice de confusion
            cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
            if cm.shape != (2, 2):
                return

            tn, fp, fn, tp = cm.ravel()
            total = cm.sum()

            acc = (tp + tn) / total if total > 0 else 0.0
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

            # ROC
            has_both_classes = len(np.unique(y_true)) > 1
            if has_both_classes:
                fpr, tpr, _ = roc_curve(y_true, y_prob)
                roc_auc = auc(fpr, tpr)
            else:
                fpr, tpr, roc_auc = None, None, np.nan

            if not cm_container_packed:
                cm_container.pack(pady=10, fill="both", expand=True)
                cm_container_packed = True

            for child in cm_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            cm_widget = None

            # Figure 2 panneaux
            fig, (ax_cm, ax_roc) = plt.subplots(
                1, 2,
                figsize=(11.8, 5.2),
                facecolor="#00122e",
                constrained_layout=True
            )
            ax_cm.set_facecolor("#00122e")
            ax_roc.set_facecolor("#00122e")

            # =========================
            # Panneau 1 : Confusion Matrix
            # =========================
            im = ax_cm.imshow(cm, interpolation="nearest", cmap="Blues")

            cbar = fig.colorbar(im, ax=ax_cm, fraction=0.046, pad=0.04)
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="white")

            classes = ["Prédit : nul", "Prédit : non nul"]
            rows = ["Réel : nul", "Réel : non nul"]

            ax_cm.set_xticks(np.arange(2))
            ax_cm.set_yticks(np.arange(2))
            ax_cm.set_xticklabels(classes, color="white", fontsize=11)
            ax_cm.set_yticklabels(rows, color="white", fontsize=11)

            ax_cm.set_xlabel("Classe prédite", color="white", fontsize=12)
            ax_cm.set_ylabel("Classe réelle", color="white", fontsize=12)

            year_txt = f" – VALID {base_year}" if base_year is not None else ""
            ax_cm.set_title(f"Matrice de confusion – {filiale}{year_txt}", color="white", fontsize=14)

            vmax = cm.max() if cm.size > 0 else 1
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    val = cm[i, j]
                    color_txt = "white" if val > vmax / 2 else "#00122e"
                    ax_cm.text(
                        j, i, f"{val}",
                        ha="center", va="center",
                        color=color_txt, fontsize=14, fontweight="bold"
                    )

            # =========================
            # Panneau 2 : ROC Curve
            # =========================
            if has_both_classes:
                # Nom du meilleur modèle pour le titre si dispo
                best_name = "Ensemble"
                try:
                    if cls_model_names is not None and len(cls_model_names) == len(cls_weights):
                        best_idx = int(np.argmax(np.asarray(cls_weights, dtype=float)))
                        best_name = f"Ensemble (dominant: {cls_model_names[best_idx]})"
                except Exception:
                    pass

                ax_roc.plot(fpr, tpr, linewidth=2.2, color="#5DADE2", label=f"{best_name} AUC = {roc_auc:.3f}")
                ax_roc.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="white", alpha=0.7, label="Aléatoire")

                ax_roc.set_xlim(0.0, 1.0)
                ax_roc.set_ylim(0.0, 1.05)
                ax_roc.set_xlabel("False Positive Rate", color="white", fontsize=12)
                ax_roc.set_ylabel("True Positive Rate", color="white", fontsize=12)
                ax_roc.set_title("Courbe ROC – classification VALID", color="white", fontsize=14)

                ax_roc.tick_params(axis="x", colors="white")
                ax_roc.tick_params(axis="y", colors="white")

                leg = ax_roc.legend(facecolor="#00122e", edgecolor="white", fontsize=9)
                for t in leg.get_texts():
                    t.set_color("white")
            else:
                ax_roc.text(
                    0.5, 0.5,
                    "ROC indisponible\n(une seule classe dans VALID)",
                    ha="center", va="center",
                    color="white", fontsize=12
                )
                ax_roc.set_title("Courbe ROC – classification VALID", color="white", fontsize=14)
                ax_roc.set_xticks([])
                ax_roc.set_yticks([])

            # Style commun
            for ax in (ax_cm, ax_roc):
                for spine in ax.spines.values():
                    spine.set_color("white")

            # KPI en bas
            kpi_txt = (
                f"Accuracy: {acc:.1%}   |   "
                f"Precision: {precision:.1%}   |   "
                f"Recall: {recall:.1%}   |   "
                f"Specificity: {specificity:.1%}"
            )
            if has_both_classes:
                kpi_txt += f"   |   AUC: {roc_auc:.3f}"

            fig.text(0.5, 0.02, kpi_txt, ha="center", color="white", fontsize=10)

            canvas = FigureCanvasTkAgg(fig, master=cm_container)
            canvas.draw()
            cm_widget = canvas.get_tk_widget()
            cm_widget.pack(fill="both", expand=True)
            plt.close(fig)
        
        def _redraw_confusion_matrix_pred_graph():
            """Redessine la matrice de confusion + courbe ROC sur les prédictions de l'année cible."""
            nonlocal cm_pred_widget, cm_pred_container_packed
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            from sklearn.metrics import confusion_matrix, roc_curve, auc

            if current_pred_df is None or current_pred_df.empty:
                return
            if current_real_target_df is None or current_real_target_df.empty:
                return

            if "date" not in current_pred_df.columns or "date" not in current_real_target_df.columns:
                return
            if "y" not in current_real_target_df.columns:
                return
            if "pred_cls_proba" not in current_pred_df.columns:
                return

            df_pred = current_pred_df.copy()
            df_real = current_real_target_df.copy()

            df_pred["date"] = pd.to_datetime(df_pred["date"])
            df_real["date"] = pd.to_datetime(df_real["date"])

            ty = model_artifacts.get("target_year", None) or current_target_year
            try:
                ty = int(ty)
            except Exception:
                ty = int(pd.Series(df_pred["date"].dt.year).mode().iloc[0])

            # Merge réel / prédit
            cols_pred = ["date", "pred_cls_proba"]
            if "pred_cls" in df_pred.columns:
                cols_pred.append("pred_cls")

            df_eval = pd.merge(
                df_real[["date", "y"]].copy(),
                df_pred[cols_pred].copy(),
                on="date",
                how="inner"
            )

            df_eval = df_eval[df_eval["date"].dt.year == ty].copy()
            if df_eval.empty:
                return

            # Vérité terrain : 0 si nul, 1 si non nul
            y_true = (df_eval["y"].astype(float).values > 0).astype(int)

            # Probabilités / classes prédites
            y_prob = df_eval["pred_cls_proba"].astype(float).values
            y_prob = np.clip(y_prob, 0.0, 1.0)

            if "pred_cls" in df_eval.columns:
                y_pred = df_eval["pred_cls"].astype(int).values
            else:
                threshold = 0.5
                y_pred = (y_prob >= threshold).astype(int)

            cm = confusion_matrix(y_true, y_pred, labels=[0, 1])
            if cm.shape != (2, 2):
                return

            tn, fp, fn, tp = cm.ravel()
            total = cm.sum()

            acc = (tp + tn) / total if total > 0 else 0.0
            precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
            recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
            specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0

            has_both_classes = len(np.unique(y_true)) > 1
            if has_both_classes:
                fpr, tpr, _ = roc_curve(y_true, y_prob)
                roc_auc = auc(fpr, tpr)
            else:
                fpr, tpr, roc_auc = None, None, np.nan

            if not cm_pred_container_packed:
                cm_pred_container.pack(pady=10, fill="both", expand=True)
                cm_pred_container_packed = True

            for child in cm_pred_container.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass
            cm_pred_widget = None

            fig, (ax_cm, ax_roc) = plt.subplots(
                1, 2,
                figsize=(11.8, 5.2),
                facecolor="#00122e",
                constrained_layout=True
            )
            ax_cm.set_facecolor("#00122e")
            ax_roc.set_facecolor("#00122e")

            # =========================
            # Panneau 1 : Matrice de confusion
            # =========================
            im = ax_cm.imshow(cm, interpolation="nearest", cmap="Blues")

            cbar = fig.colorbar(im, ax=ax_cm, fraction=0.046, pad=0.04)
            cbar.ax.yaxis.set_tick_params(color="white")
            plt.setp(plt.getp(cbar.ax.axes, "yticklabels"), color="white")

            classes = ["Prédit : nul", "Prédit : non nul"]
            rows = ["Réel : nul", "Réel : non nul"]

            ax_cm.set_xticks(np.arange(2))
            ax_cm.set_yticks(np.arange(2))
            ax_cm.set_xticklabels(classes, color="white", fontsize=11)
            ax_cm.set_yticklabels(rows, color="white", fontsize=11)

            ax_cm.set_xlabel("Classe prédite", color="white", fontsize=12)
            ax_cm.set_ylabel("Classe réelle", color="white", fontsize=12)

            filiale = model_artifacts.get("filiale", None) or current_filiale_name or ""
            ax_cm.set_title(f"Matrice de confusion – Prévision {ty} – {filiale}", color="white", fontsize=14)

            vmax = cm.max() if cm.size > 0 else 1
            for i in range(cm.shape[0]):
                for j in range(cm.shape[1]):
                    val = cm[i, j]
                    color_txt = "white" if val > vmax / 2 else "#00122e"
                    ax_cm.text(
                        j, i, f"{val}",
                        ha="center", va="center",
                        color=color_txt, fontsize=14, fontweight="bold"
                    )

            # =========================
            # Panneau 2 : ROC curve
            # =========================
            if has_both_classes:
                ax_roc.plot(fpr, tpr, linewidth=2.2, color="#5DADE2", label=f"Prévision {ty} AUC = {roc_auc:.3f}")
                ax_roc.plot([0, 1], [0, 1], linestyle="--", linewidth=1.0, color="white", alpha=0.7, label="Aléatoire")

                ax_roc.set_xlim(0.0, 1.0)
                ax_roc.set_ylim(0.0, 1.05)
                ax_roc.set_xlabel("False Positive Rate", color="white", fontsize=12)
                ax_roc.set_ylabel("True Positive Rate", color="white", fontsize=12)
                ax_roc.set_title(f"Courbe ROC – Prévision {ty}", color="white", fontsize=14)

                ax_roc.tick_params(axis="x", colors="white")
                ax_roc.tick_params(axis="y", colors="white")

                leg = ax_roc.legend(facecolor="#00122e", edgecolor="white", fontsize=9)
                for t in leg.get_texts():
                    t.set_color("white")
            else:
                ax_roc.text(
                    0.5, 0.5,
                    "ROC indisponible\n(une seule classe réelle)",
                    ha="center", va="center",
                    color="white", fontsize=12
                )
                ax_roc.set_title(f"Courbe ROC – Prévision {ty}", color="white", fontsize=14)
                ax_roc.set_xticks([])
                ax_roc.set_yticks([])

            for ax in (ax_cm, ax_roc):
                for spine in ax.spines.values():
                    spine.set_color("white")

            kpi_txt = (
                f"Accuracy: {acc:.1%}   |   "
                f"Precision: {precision:.1%}   |   "
                f"Recall: {recall:.1%}   |   "
                f"Specificity: {specificity:.1%}"
            )
            if has_both_classes:
                kpi_txt += f"   |   AUC: {roc_auc:.3f}"

            fig.text(0.5, 0.02, kpi_txt, ha="center", color="white", fontsize=10)

            canvas = FigureCanvasTkAgg(fig, master=cm_pred_container)
            canvas.draw()
            cm_pred_widget = canvas.get_tk_widget()
            cm_pred_widget.pack(fill="both", expand=True)
            plt.close(fig)

        # ---------- analyse des écarts ----------
        def _compute_metrics(
            label,
            ref_dates, ref_values,
            cmp_dates, cmp_values,
            min_abs_ref=10.0,
            rel_seuil=40.0,
            rel_mode="smape",
            eps=1e-9,
            exclude_above_k=None,
            restrict_dates=None,
        ):
            if ref_dates is None or cmp_dates is None:
                return None

            ref_vals = [_to_float_or_nan(v) for v in ref_values]
            cmp_vals = [_to_float_or_nan(v) for v in cmp_values]

            def _norm_dates(x):
                d = pd.to_datetime(x, errors="coerce")
                if isinstance(d, (pd.DatetimeIndex, pd.Index)):
                    return d.normalize()
                return d.dt.normalize()

            ref_idx = _norm_dates(ref_dates)
            cmp_idx = _norm_dates(cmp_dates)

            ref = pd.Series(ref_vals, index=ref_idx, dtype="float64").dropna()
            cmp_ = pd.Series(cmp_vals, index=cmp_idx, dtype="float64").dropna()

            common_idx = ref.index.intersection(cmp_.index)
            if len(common_idx) == 0:
                return None

            df_m = pd.DataFrame({
                "ref": ref.loc[common_idx],
                "cmp": cmp_.loc[common_idx],
            }).dropna()

            if restrict_dates is not None and not df_m.empty:
                try:
                    ridx = pd.to_datetime(restrict_dates, errors="coerce")
                    ridx = pd.DatetimeIndex(ridx).normalize()
                    df_m = df_m.loc[df_m.index.intersection(ridx)]
                except Exception:
                    pass

            if df_m.empty:
                return None

            if min_abs_ref is not None and min_abs_ref > 0:
                df_m = df_m[df_m["ref"].abs() >= float(min_abs_ref)]

            if exclude_above_k is not None and not df_m.empty:
                df_m = df_m[df_m["ref"].abs() <= float(exclude_above_k)]

            if df_m.empty:
                return None

            err = df_m["cmp"] - df_m["ref"]
            abs_err = err.abs()

            mae = float(abs_err.mean())
            rmse = float(np.sqrt(np.mean(err.values ** 2)))
            bias = float(err.mean())

            if rel_mode == "mape_ref":
                denom = df_m["ref"].abs().clip(lower=eps)
                rel = (abs_err / denom) * 100.0
            else:
                denom = (df_m["ref"].abs() + df_m["cmp"].abs()).clip(lower=eps)
                rel = (2.0 * abs_err / denom) * 100.0

            rel = rel.replace([np.inf, -np.inf], np.nan)
            df_m = df_m.assign(rel=rel).dropna(subset=["rel"])

            if df_m.empty:
                return None

            mask = df_m["rel"] > float(rel_seuil)
            nb_points = int(len(df_m))
            nb_alert = int(mask.sum())
            taux_alert = 0.0 if nb_points == 0 else 100.0 * nb_alert / nb_points
            valo_alert = float(abs_err[mask].sum())

            sum_ref = float(df_m["ref"].abs().sum())
            wmape = float(abs_err.sum() / (sum_ref + eps) * 100.0)
            smape = float(df_m["rel"].mean())

            # 🔹 Détermination du type
            if label.startswith("Prévision IA vs Réel (période"):
                type_label = "Prévision IA (période Profil)"
            elif label.startswith("Prévision"):
                type_label = "Prévision IA"
            else:
                type_label = "Profil"

            return {
                "Type": type_label,
                "Comparaison": label,
                "Nb points (utilisés)": nb_points,
                "MAE (K€)": round(mae, 1),
                "RMSE (K€)": round(rmse, 1),
                "Biais (K€)": round(bias, 1),
                "sMAPE (%)": round(smape, 1),
                "WMAPE (%)": round(wmape, 1),
                f"Nb alertes (> {rel_seuil}%)": nb_alert,
                f"Taux alertes (> {rel_seuil}%)": round(taux_alert, 1),
                "Valeur totale écarts alertes (K€)": round(valo_alert, 1),
            }

        def _rebuild_analysis_table():
            nonlocal analysis_table_frame
            nonlocal current_pred_df, current_real_target_df, current_target_year

            for child in analysis_table_frame.winfo_children():
                try:
                    child.destroy()
                except Exception:
                    pass

            if current_real_target_df is None or current_real_target_df.empty:
                return

            rows = []

            rel_seuil = 40.0
            min_abs_ref = 10.0
            rel_mode = "smape"

            # 🔹 IA vs Réel (normal)
            if current_pred_df is not None and not current_pred_df.empty:
                m = _compute_metrics(
                    label="Prévision IA vs Réel",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    min_abs_ref=min_abs_ref,
                    rel_seuil=rel_seuil,
                    rel_mode=rel_mode,
                )
                if m:
                    rows.append(m)

                # 🔹 IA vs Réel (sans jours > 100 000)
                m_np = _compute_metrics(
                    label="Prévision Réel vs IA (sans jours > 100 000)",
                    ref_dates=current_real_target_df["date"],
                    ref_values=current_real_target_df["y"],
                    cmp_dates=current_pred_df["date"],
                    cmp_values=current_pred_df["pred_value"],
                    min_abs_ref=min_abs_ref,
                    rel_seuil=rel_seuil,
                    rel_mode=rel_mode,
                    exclude_above_k=100000.0,  # ✅ 100 M€
                )
                if m_np:
                    rows.append(m_np)

            # 🔹 Profils cochés
            if ia_profils_names and ia_profils_series:
                for name, var, serie in zip(ia_profils_names, ia_profils_vars, ia_profils_series):
                    if not var.get():
                        continue

                    # Profil vs Réel
                    m_prof = _compute_metrics(
                        label=f"Profil '{name}' vs Réel",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=ia_profils_dates,
                        cmp_values=serie,
                        min_abs_ref=min_abs_ref,
                        rel_seuil=rel_seuil,
                        rel_mode=rel_mode,
                    )
                    if m_prof:
                        rows.append(m_prof)

                    # IA vs Réel sur même période que le profil
                    prof_idx = pd.to_datetime(ia_profils_dates, errors="coerce").normalize()
                    prof_vals = pd.Series(
                        [_to_float_or_nan(v) for v in serie],
                        index=prof_idx
                    ).dropna()

                    m_ia_same = _compute_metrics(
                        label=f"Prévision IA vs Réel (période Profil '{name}')",
                        ref_dates=current_real_target_df["date"],
                        ref_values=current_real_target_df["y"],
                        cmp_dates=current_pred_df["date"],
                        cmp_values=current_pred_df["pred_value"],
                        min_abs_ref=min_abs_ref,
                        rel_seuil=rel_seuil,
                        rel_mode=rel_mode,
                        restrict_dates=prof_vals.index,
                    )
                    if m_ia_same:
                        rows.append(m_ia_same)

            if not rows:
                return

            cols = list(rows[0].keys())

            tree = ttk.Treeview(
                analysis_table_frame,
                columns=cols,
                show="headings",
                height=len(rows)
            )

            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, anchor="center", width=140)

            for r in rows:
                tree.insert("", "end", values=[r.get(c, "") for c in cols])

            tree.pack(fill="x", expand=True)

            # ===================== ✅ Bouton d'extraction (même format CTk) =====================
            export_tools_frame = ctk.CTkFrame(analysis_table_frame, fg_color="#00122e", corner_radius=0)
            export_tools_frame.pack(fill="x", padx=10, pady=(6, 10))

            def _export_treeview_to_xlsx(tree: ttk.Treeview):
                cols_ = list(tree["columns"])
                rows_ = [tree.item(i, "values") for i in tree.get_children()]

                if not cols_ or not rows_:
                    messagebox.showwarning("Export", "Aucune donnée à exporter.")
                    return

                fpath = filedialog.asksaveasfilename(
                    title="Exporter l'analyse des écarts",
                    defaultextension=".xlsx",
                    filetypes=[("Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
                )
                if not fpath:
                    return

                try:
                    df = pd.DataFrame(rows_, columns=cols_)
                    df.to_excel(fpath, index=False)
                    messagebox.showinfo("Export", f"Export Excel OK :\n{fpath}")
                except Exception as e:
                    messagebox.showerror("Export", f"Erreur export Excel :\n{e}")

            export_analysis_button = ctk.CTkButton(
                export_tools_frame,
                text="📤 Export Analyse Écarts - Excel",
                width=320, height=34,
                corner_radius=10,
                fg_color="#2563eb", hover_color="#1d4ed8",
                text_color="white",
                state="normal",   # ou "disabled" si tu veux l'activer plus tard
                command=lambda t=tree: _export_treeview_to_xlsx(t)
            )
            export_analysis_button.pack(pady=6)



        # ---------- PROFILS UI ----------
        def _build_ia_profils_ui(filiale, base_year):
            """
            Construit les cases à cocher de profils pour l'année N+1 de la filiale sélectionnée.
            """
            nonlocal ia_profils_vars, ia_profils_names, ia_profils_dates, ia_profils_series

            for w in ia_profils_frame.winfo_children():
                try:
                    w.destroy()
                except Exception:
                    pass

            ia_profils_vars = []
            ia_profils_names = []
            ia_profils_dates = []
            ia_profils_series = []

            if not filiale:
                tk.Label(
                    ia_profils_frame,
                    text="Aucune filiale sélectionnée.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            try:
                base_year_int = int(base_year)
            except Exception:
                tk.Label(
                    ia_profils_frame,
                    text="Année N invalide.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            target_year = base_year_int + 1
            feuille = sections.get(filiale)
            if not feuille:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune feuille trouvée pour {filiale}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            try:
                ws, noms_flux = charger_donnees(feuille, taille_bloc)
            except Exception as e:
                print(f"[IA] Erreur charger_donnees (profils N+1) pour {feuille} : {e}")
                tk.Label(
                    ia_profils_frame,
                    text="Erreur de chargement des données N+1.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            flux_cible_local = selected_flux.get()
            cible = [t for t in noms_flux if t[0] == flux_cible_local]
            if not cible:
                tk.Label(
                    ia_profils_frame,
                    text=f"Flux '{flux_cible_local}' introuvable pour cette filiale.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            _, col_start = cible[0]

            try:
                dates_p, reel_p, previsions_p, noms_profils_p = extraire_valeurs(
                    ws, col_start, nb_prev, annee=target_year
                )
            except Exception as e:
                print(f"[IA] Erreur extraire_valeurs (profils N+1) {filiale}/{flux_cible_local}/{target_year} : {e}")
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée de profils pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            if not dates_p:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucune donnée pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            actifs = []
            for serie in previsions_p:
                exist = any(v not in (None, 0, 0.0, "") for v in serie)
                actifs.append(exist)

            noms_actifs = [n for n, ok in zip(noms_profils_p, actifs) if ok]
            series_actives = [s for s, ok in zip(previsions_p, actifs) if ok]

            if not noms_actifs:
                tk.Label(
                    ia_profils_frame,
                    text=f"Aucun profil actif pour {target_year}.",
                    bg="#00122e", fg="white", font=('Segoe UI', 10, 'italic')
                ).pack(anchor="w")
                return

            ia_profils_dates = dates_p
            ia_profils_names = noms_actifs
            ia_profils_series = series_actives

            lbl = tk.Label(
                ia_profils_frame,
                text=f"Profils N+1 ({target_year}) : coche pour afficher dans le graphe détaillé et l'analyse",
                bg="#00122e", fg="white", font=('Segoe UI', 10, 'bold')
            )
            lbl.grid(row=0, column=0, columnspan=4, sticky="w", padx=4, pady=(0, 6))

            nb_lignes = 4
            for i, name in enumerate(noms_actifs):
                var = tk.BooleanVar(value=False)

                def _on_toggle(v=var):
                    _redraw_graph2()
                    _redraw_monthly_graph()
                    _rebuild_analysis_table()

                cb = tk.Checkbutton(
                    ia_profils_frame, text=name, variable=var,
                    bg="#00122e", fg="white", font=('Segoe UI', 10),
                    selectcolor="#00aced", activebackground="#003366",
                    activeforeground="white",
                    command=_on_toggle
                )

                row = 1 + (i % nb_lignes)
                col = i // nb_lignes

                cb.grid(row=row, column=col, sticky="w", padx=12, pady=4)
                cb.bind("<Enter>", lambda e, c=cb: c.config(bg="#003366"))
                cb.bind("<Leave>", lambda e, c=cb: c.config(bg="#00122e"))

                ia_profils_vars.append(var)

        def _on_filiale_or_year_change(_event=None):
            filiale = selected_filiale.get()
            base_year = annees_var.get()
            if not filiale or not base_year:
                return
            _build_ia_profils_ui(filiale, base_year)

        filiale_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)
        annees_box.bind("<<ComboboxSelected>>", _on_filiale_or_year_change)

        # init profils
        _on_filiale_or_year_change()

        # ============================================================
        # NEW: callback flux change
        # ============================================================
        def _on_flux_change(_event=None):
            ok = _rebuild_df_for_current_flux()
            if not ok:
                return
            _on_filiale_or_year_change()

            # si un modèle a déjà été entrainé, on redessine ce qui dépend des profils
            try:
                _redraw_monthly_graph()
                _redraw_graph2()
                _rebuild_analysis_table()
            except Exception:
                pass

        flux_box.bind("<<ComboboxSelected>>", _on_flux_change)


        # ---------- ENTRAÎNEMENT DU MODÈLE & GRAPHIQUES ----------
        import re

        def _norm(s: str) -> str:
            s = (s or "").strip().lower()
            s = s.replace("œ", "oe")
            s = re.sub(r"\s+", " ", s)
            return s

        def _train_model():
            """
            Routeur: appelle le bon trainer selon flux + filiale.
            df_current est supposé déjà filtré sur le flux sélectionné.
            """
            flux_raw = selected_flux.get()
            filiale_raw = selected_filiale.get()

            flux = _norm(flux_raw)
            filiale = (filiale_raw or "").strip().upper()

            print(f"[ROUTER] filiale={filiale} | flux='{flux_raw}' -> norm='{flux}'")

            # --- Cas 1 : Trafic Voyageurs (ton V3.3)
            if flux in {"trafic voyageurs", "traficvoyageurs"}:
                return _train_trafic_voyageurs_v33()   # ton trainer existant

            # --- Cas 2 : RESEAU + ACE & Investissements (négatif <= 0)
            # robuste aux variantes d'orthographe
            if filiale == "RESEAU" and ("ace" in flux) and ("invest" in flux):
                return _train_ace_invest_reseau_negative_v33()

            # --- Fallback
            messagebox.showinfo(
                "Info",
                f"Aucun modèle spécialisé pour :\nFiliale={filiale_raw}\nFlux={flux_raw}\n\n→ modèle générique."
            )
            return _train_generic_baseline()

        def _train_generic_baseline():
            messagebox.showinfo(
                "Non supporté",
                f"Aucun modèle spécialisé pour:\nFiliale={selected_filiale.get()}\nFlux={selected_flux.get()}"
            )

        def _train_trafic_voyageurs_v33():
            """
            SA_VOYAGEURS – V4.0 ENSEMBLE ULTIME (XGBoost + LightGBM + CatBoost + ExtraTrees)
            Version figée / reproductible au maximum.
            """

            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame, export_button

            _clear_graph_widgets()

            # ======================================================================
            # 0) Reproductibilité MAX
            # ======================================================================
            import os
            import random

            SEED = 42

            # Idéalement à mettre tout en haut du script principal, avant imports globaux.
            os.environ["PYTHONHASHSEED"] = str(SEED)
            os.environ["OMP_NUM_THREADS"] = "1"
            os.environ["MKL_NUM_THREADS"] = "1"
            os.environ["OPENBLAS_NUM_THREADS"] = "1"
            os.environ["VECLIB_MAXIMUM_THREADS"] = "1"
            os.environ["NUMEXPR_NUM_THREADS"] = "1"

            random.seed(SEED)

            import numpy as np
            np.random.seed(SEED)

            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            import calendar
            import traceback
            import time

            from sklearn.metrics import (
                mean_absolute_error,
                mean_squared_error,
                r2_score,
                roc_auc_score
            )
            from sklearn.model_selection import RandomizedSearchCV, TimeSeriesSplit
            from scipy.stats import randint, uniform

            # --- Tentative d'import de XGBoost ---
            try:
                from xgboost import XGBClassifier, XGBRegressor
                HAS_XGB = True
            except ImportError:
                HAS_XGB = False
                print("[WARN] XGBoost n'est pas installé. Le modèle tournera sans, mais installe-le avec 'pip install xgboost' pour des perfs maximales.")

            t0_global = time.perf_counter()

            def _log(msg):
                dt = time.perf_counter() - t0_global
                print(f"[{dt:8.1f}s] {msg}", flush=True)

            _log("======================== TRAIN V4.0 ENSEMBLE ULTIME (SA_VOYAGEURS) ========================")
            _log(f"[REPRO] SEED={SEED}")

            # ---------------- Filiale fixée ----------------
            filiale = "SA_VOYAGEURS"
            current_filiale_name = filiale
            source_df = df_current if (df_current is not None and not df_current.empty) else df
            df_filiale = source_df[source_df["section"] == filiale].copy()

            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée trouvée pour la filiale {filiale}.")
                return

            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["year"].max())

            _log(f"Base year for training {filiale} : {base_year}")

            try:
                # ======================================================================
                # 1) Checks
                # ======================================================================
                if df_filiale["year"].nunique() < 2:
                    messagebox.showinfo("Information", f"Pas assez d'historique pour {filiale} (au moins 2 années).")
                    return

                df_filiale = df_filiale.sort_values(["date"]).reset_index(drop=True).copy()

                # ======================================================================
                # 2) Holidays
                # ======================================================================
                _log("[STEP] Début calcul jours fériés")

                holiday_sets = {}
                if "is_holiday" not in df_filiale.columns:
                    df_filiale["is_holiday"] = 0
                    try:
                        import holidays
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            fr_h = holidays.country_holidays("FR", years=[int(yy)])
                            holiday_sets[int(yy)] = set(fr_h.keys())
                        df_filiale["is_holiday"] = df_filiale["date"].dt.date.apply(
                            lambda d: 1 if d in holiday_sets.get(int(d.year), set()) else 0
                        )
                        _log(f"[HOLIDAYS] Calcul jours fériés FR OK (years={years})")
                    except Exception:
                        _log("[HOLIDAYS] fallback is_holiday=0")
                        holiday_sets = {}
                else:
                    try:
                        years = sorted(df_filiale["year"].unique().tolist())
                        for yy in years:
                            sub = df_filiale[df_filiale["year"] == int(yy)]
                            holiday_sets[int(yy)] = set(sub.loc[sub["is_holiday"] == 1, "date"].dt.date.tolist())
                        _log(f"[HOLIDAYS] Reprise colonne existante OK (years={years})")
                    except Exception:
                        holiday_sets = {}
                        _log("[HOLIDAYS] Reprise colonne existante impossible -> fallback vide")

                def _is_open_day(ts: pd.Timestamp, holiday_set: set) -> bool:
                    return (ts.weekday() < 5) and (ts.date() not in holiday_set)

                def _effective_22_date(year: int, month: int, holiday_set: set) -> pd.Timestamp:
                    d = pd.Timestamp(year=year, month=month, day=22)
                    while not _is_open_day(d, holiday_set):
                        d += pd.Timedelta(days=1)
                    return d.normalize()

                def _last_business_day(year: int, month: int, holiday_set: set) -> pd.Timestamp:
                    d = pd.Timestamp(year=year, month=month, day=1) + pd.offsets.MonthEnd(0)
                    d = d.normalize()
                    while not _is_open_day(d, holiday_set):
                        d -= pd.Timedelta(days=1)
                    return d.normalize()

                def _apply_closed_day_report(df_in: pd.DataFrame, holiday_sets_: dict, y_col="y") -> pd.DataFrame:
                    dd = df_in.sort_values("date").copy()
                    dd["dow"] = dd["date"].dt.weekday

                    if "is_holiday" not in dd.columns:
                        dd["is_holiday"] = dd["date"].dt.date.apply(
                            lambda d: 1 if d in holiday_sets_.get(int(pd.Timestamp(d).year), set()) else 0
                        )

                    y = dd[y_col].astype(float).values
                    out = np.zeros_like(y, dtype=float)
                    carry = 0.0

                    dow = dd["dow"].values
                    ish = dd["is_holiday"].values

                    for i in range(len(dd)):
                        closed = (dow[i] >= 5) or (ish[i] == 1)
                        if closed:
                            carry += y[i]
                            out[i] = 0.0
                        else:
                            out[i] = y[i] + carry
                            carry = 0.0

                    if carry > 0:
                        idx_open = np.where((dow < 5) & (ish == 0))[0]
                        if len(idx_open) > 0:
                            out[idx_open[-1]] += carry

                    dd[y_col] = out
                    return dd

                def _safe_date_from_year_doy(year: int, doy: int) -> pd.Timestamp:
                    days_in_year = 366 if calendar.isleap(int(year)) else 365
                    doy = int(max(1, min(int(doy), days_in_year)))
                    return (pd.Timestamp(year=int(year), month=1, day=1) + pd.Timedelta(days=doy - 1)).normalize()

                _log("[STEP] Fin calcul jours fériés")

                # ======================================================================
                # 3) Série métier + features
                # ======================================================================
                _log("[STEP] Début feature engineering (incl. Volatilité)")

                df_filiale["y_raw"] = df_filiale["y"].astype(float)
                df_filiale = _apply_closed_day_report(df_filiale, holiday_sets, y_col="y")

                df_filiale["dow"] = df_filiale["date"].dt.weekday
                df_filiale["month"] = df_filiale["date"].dt.month
                df_filiale["dom"] = df_filiale["date"].dt.day
                df_filiale["year"] = df_filiale["date"].dt.year
                df_filiale["dayofyear"] = df_filiale["date"].dt.dayofyear
                df_filiale["is_eom"] = df_filiale["date"].dt.is_month_end.astype(int)

                # lags/rolls série métier
                df_filiale["lag_1"] = df_filiale["y"].shift(1)
                df_filiale["lag_7"] = df_filiale["y"].shift(7)
                df_filiale["roll_mean_7"] = df_filiale["y"].rolling(7, min_periods=1).mean()
                df_filiale["roll_mean_30"] = df_filiale["y"].rolling(30, min_periods=1).mean()

                # Volatilité
                df_filiale["roll_std_7"] = df_filiale["y"].rolling(7, min_periods=1).std().fillna(0)
                df_filiale["roll_std_30"] = df_filiale["y"].rolling(30, min_periods=1).std().fillna(0)

                # lags/rolls série brute
                df_filiale["lag_1_raw"] = df_filiale["y_raw"].shift(1)
                df_filiale["lag_7_raw"] = df_filiale["y_raw"].shift(7)
                df_filiale["roll_mean_7_raw"] = df_filiale["y_raw"].rolling(7, min_periods=1).mean()
                df_filiale["roll_mean_30_raw"] = df_filiale["y_raw"].rolling(30, min_periods=1).mean()

                years_present = sorted(df_filiale["year"].unique().tolist())
                map_rows = []
                for yy in years_present:
                    hset = holiday_sets.get(int(yy), set())
                    for mm in range(1, 13):
                        try:
                            map_rows.append({
                                "year": int(yy),
                                "month": int(mm),
                                "eff22_date": _effective_22_date(int(yy), int(mm), hset),
                                "lbd_date": _last_business_day(int(yy), int(mm), hset),
                            })
                        except Exception:
                            pass
                calmap = pd.DataFrame(map_rows)

                df_filiale = df_filiale.merge(calmap, on=["year", "month"], how="left")
                df_filiale["date_norm"] = df_filiale["date"].dt.normalize()

                def _business_day_features_for_month(year: int, month: int, holiday_set: set):
                    start = pd.Timestamp(year=year, month=month, day=1)
                    end = (start + pd.offsets.MonthEnd(0)).normalize()
                    all_days = pd.date_range(start, end, freq="D")

                    def is_open(ts):
                        return (ts.weekday() < 5) and (ts.date() not in holiday_set)

                    open_days = [d.normalize() for d in all_days if is_open(d)]
                    bd_rank = {d: i + 1 for i, d in enumerate(open_days)}
                    K = len(open_days)
                    bd_to_eom = {d: (K - bd_rank[d]) for d in bd_rank}
                    return bd_rank, bd_to_eom

                bd_rows = []
                for yy in years_present:
                    hset = holiday_sets.get(int(yy), set())
                    for mm in range(1, 13):
                        bd_rank, bd_to_eom = _business_day_features_for_month(int(yy), int(mm), hset)
                        for d, rnk in bd_rank.items():
                            bd_rows.append({
                                "date_norm": d,
                                "business_day_of_month": int(rnk),
                                "business_days_to_eom": int(bd_to_eom.get(d, 0))
                            })

                bd_map = pd.DataFrame(bd_rows)
                df_filiale = df_filiale.merge(bd_map, on="date_norm", how="left")
                df_filiale["business_day_of_month"] = df_filiale["business_day_of_month"].fillna(0).astype(int)
                df_filiale["business_days_to_eom"] = df_filiale["business_days_to_eom"].fillna(0).astype(int)

                # spikes
                df_filiale["is_dom_22"] = (df_filiale["dom"] == 22).astype(int)
                df_filiale["days_to_22"] = (df_filiale["dom"] - 22).astype(int)

                df_filiale["is_22_effective"] = (df_filiale["date_norm"] == df_filiale["eff22_date"]).astype(int)
                df_filiale["days_to_22_effective"] = (df_filiale["date_norm"] - df_filiale["eff22_date"]).dt.days.fillna(0).astype(int)
                df_filiale["abs_days_to_22_effective"] = df_filiale["days_to_22_effective"].abs()
                df_filiale["is_near_22_effective"] = (df_filiale["abs_days_to_22_effective"] <= 2).astype(int)

                df_filiale["is_mid_month_window"] = df_filiale["dom"].between(20, 22).astype(int)
                df_filiale["abs_days_to_20"] = (df_filiale["dom"] - 20).abs().astype(int)
                df_filiale["is_near_20"] = (df_filiale["abs_days_to_20"] <= 1).astype(int)
                df_filiale["is_monday"] = (df_filiale["dow"] == 0).astype(int)
                df_filiale["is_monday_after_20"] = ((df_filiale["dow"] == 0) & (df_filiale["dom"].between(20, 22))).astype(int)

                df_filiale["is_lbd"] = (df_filiale["date_norm"] == df_filiale["lbd_date"]).astype(int)
                df_filiale["days_to_lbd"] = (df_filiale["date_norm"] - df_filiale["lbd_date"]).dt.days.fillna(0).astype(int)
                df_filiale["abs_days_to_lbd"] = df_filiale["days_to_lbd"].abs()
                df_filiale["is_near_lbd"] = (df_filiale["abs_days_to_lbd"] <= 1).astype(int)

                df_filiale["spike_risk"] = (
                    (df_filiale["is_near_22_effective"] == 1) |
                    (df_filiale["is_mid_month_window"] == 1) |
                    (df_filiale["is_monday_after_20"] == 1) |
                    (df_filiale["is_near_lbd"] == 1)
                ).astype(int)

                _log(f"[FEATURES] nb_rows={len(df_filiale)} | years={sorted(df_filiale['year'].unique().tolist())}")
                _log("[STEP] Fin feature engineering")

                # ======================================================================
                # 4) Paires YoY
                # ======================================================================
                _log("[STEP] Début merge YoY")

                s = df_filiale.copy()
                s_prev = s.copy()
                s_prev["year_target"] = s_prev["year"] + 1

                s_prev = s_prev.rename(columns={
                    "y": "y_prev_year",
                    "y_raw": "y_raw_prev_year",
                    "roll_mean_7": "roll_prev_7",
                    "roll_mean_30": "roll_prev_30",
                    "roll_std_7": "roll_std_prev_7",
                    "roll_std_30": "roll_std_prev_30",
                    "roll_mean_7_raw": "roll_prev_7_raw",
                    "roll_mean_30_raw": "roll_prev_30_raw",
                    "dow": "dow_prev",
                    "lag_1": "lag_1_prev",
                    "lag_7": "lag_7_prev",
                    "lag_1_raw": "lag_1_prev_raw",
                    "lag_7_raw": "lag_7_prev_raw",
                    "month": "month_prev",
                    "dom": "dom_prev",
                    "is_eom": "is_eom_prev",
                    "is_holiday": "is_holiday_prev",
                    "business_day_of_month": "bdm_prev",
                    "business_days_to_eom": "bdeom_prev",
                    "is_dom_22": "is_dom_22_prev",
                    "days_to_22": "days_to_22_prev",
                    "is_22_effective": "is_22_effective_prev",
                    "days_to_22_effective": "days_to_22_effective_prev",
                    "abs_days_to_22_effective": "abs_days_to_22_effective_prev",
                    "is_near_22_effective": "is_near_22_effective_prev",
                    "is_mid_month_window": "is_mid_month_window_prev",
                    "abs_days_to_20": "abs_days_to_20_prev",
                    "is_near_20": "is_near_20_prev",
                    "is_monday": "is_monday_prev",
                    "is_monday_after_20": "is_monday_after_20_prev",
                    "is_lbd": "is_lbd_prev",
                    "days_to_lbd": "days_to_lbd_prev",
                    "abs_days_to_lbd": "abs_days_to_lbd_prev",
                    "is_near_lbd": "is_near_lbd_prev",
                    "spike_risk": "spike_risk_prev",
                })

                tgt_cols = [
                    "section", "year", "dayofyear", "y", "y_raw",
                    "dow", "month", "dom", "is_eom", "is_holiday",
                    "roll_mean_7", "roll_mean_30", "roll_std_7", "roll_std_30",
                    "roll_mean_7_raw", "roll_mean_30_raw",
                    "lag_1", "lag_7", "lag_1_raw", "lag_7_raw",
                    "business_day_of_month", "business_days_to_eom",
                    "is_dom_22", "days_to_22",
                    "is_22_effective", "days_to_22_effective",
                    "abs_days_to_22_effective", "is_near_22_effective",
                    "is_mid_month_window", "abs_days_to_20", "is_near_20",
                    "is_monday", "is_monday_after_20",
                    "is_lbd", "days_to_lbd", "abs_days_to_lbd", "is_near_lbd",
                    "spike_risk",
                ]

                merged = pd.merge(
                    s_prev,
                    s[tgt_cols],
                    left_on=["section", "year_target", "dayofyear"],
                    right_on=["section", "year", "dayofyear"],
                    how="inner",
                )

                if merged.empty:
                    messagebox.showinfo("Information", f"Aucune paire année→année pour {filiale}.")
                    return

                merged = merged.rename(columns={"y": "y_target"})
                merged = merged.rename(columns={
                    "dow": "dow_tgt",
                    "month": "month_tgt",
                    "dom": "dom_tgt",
                    "is_eom": "is_eom_tgt",
                    "is_holiday": "is_holiday_tgt",
                    "y_raw": "y_raw_tgt",
                    "roll_mean_7": "roll_7_tgt",
                    "roll_mean_30": "roll_30_tgt",
                    "roll_std_7": "roll_std_7_tgt",
                    "roll_std_30": "roll_std_30_tgt",
                    "roll_mean_7_raw": "roll_7_raw_tgt",
                    "roll_mean_30_raw": "roll_30_raw_tgt",
                    "lag_1": "lag_1_tgt",
                    "lag_7": "lag_7_tgt",
                    "lag_1_raw": "lag_1_raw_tgt",
                    "lag_7_raw": "lag_7_raw_tgt",
                    "business_day_of_month": "bdm_tgt",
                    "business_days_to_eom": "bdeom_tgt",
                    "is_dom_22": "is_dom_22_tgt",
                    "days_to_22": "days_to_22_tgt",
                    "is_22_effective": "is_22_effective_tgt",
                    "days_to_22_effective": "days_to_22_effective_tgt",
                    "abs_days_to_22_effective": "abs_days_to_22_effective_tgt",
                    "is_near_22_effective": "is_near_22_effective_tgt",
                    "is_mid_month_window": "is_mid_month_window_tgt",
                    "abs_days_to_20": "abs_days_to_20_tgt",
                    "is_near_20": "is_near_20_tgt",
                    "is_monday": "is_monday_tgt",
                    "is_monday_after_20": "is_monday_after_20_tgt",
                    "is_lbd": "is_lbd_tgt",
                    "days_to_lbd": "days_to_lbd_tgt",
                    "abs_days_to_lbd": "abs_days_to_lbd_tgt",
                    "is_near_lbd": "is_near_lbd_tgt",
                    "spike_risk": "spike_risk_tgt",
                })

                df_pairs = merged.copy()
                df_pairs["section_id"] = 0

                _log(f"[YOY] nb_pairs={len(df_pairs)}")
                _log("[STEP] Fin merge YoY")

                # ======================================================================
                # 5) Features modèle
                # ======================================================================
                base_features = [
                    "y_prev_year", "y_raw_prev_year",
                    "roll_prev_7", "roll_prev_30",
                    "roll_std_prev_7", "roll_std_prev_30",
                    "roll_prev_7_raw", "roll_prev_30_raw",
                    "dayofyear", "section_id",
                    "dow_prev",
                    "lag_1_prev", "lag_7_prev",
                    "lag_1_prev_raw", "lag_7_prev_raw",
                    "month_prev", "dom_prev", "is_eom_prev", "is_holiday_prev",
                    "bdm_prev", "bdeom_prev",
                    "is_dom_22_prev", "days_to_22_prev",
                    "is_22_effective_prev", "days_to_22_effective_prev",
                    "abs_days_to_22_effective_prev", "is_near_22_effective_prev",
                    "is_mid_month_window_prev", "abs_days_to_20_prev", "is_near_20_prev",
                    "is_monday_prev", "is_monday_after_20_prev",
                    "is_lbd_prev", "days_to_lbd_prev", "abs_days_to_lbd_prev", "is_near_lbd_prev",
                    "dow_tgt", "month_tgt", "dom_tgt", "is_eom_tgt", "is_holiday_tgt",
                    "bdm_tgt", "bdeom_tgt",
                    "is_dom_22_tgt", "days_to_22_tgt",
                    "is_22_effective_tgt", "days_to_22_effective_tgt",
                    "abs_days_to_22_effective_tgt", "is_near_22_effective_tgt",
                    "is_mid_month_window_tgt", "abs_days_to_20_tgt", "is_near_20_tgt",
                    "is_monday_tgt", "is_monday_after_20_tgt",
                    "is_lbd_tgt", "days_to_lbd_tgt", "abs_days_to_lbd_tgt", "is_near_lbd_tgt",
                ]

                corr_features = base_features + ["spike_risk_prev", "spike_risk_tgt"]

                df_pairs = df_pairs.dropna(subset=list(set(corr_features)) + ["y_target"]).reset_index(drop=True)
                _log(f"[CLEAN] nb_pairs_after_dropna={len(df_pairs)}")

                # ======================================================================
                # 6) Split
                # ======================================================================
                df_train_pairs = df_pairs[df_pairs["year_target"] <= (base_year - 1)].copy()
                df_valid_pairs = df_pairs[df_pairs["year_target"] == base_year].copy()

                if df_train_pairs.shape[0] >= 90 and df_valid_pairs.shape[0] >= 60:
                    X_train = df_train_pairs[base_features].copy().reset_index(drop=True)
                    y_train = df_train_pairs["y_target"].astype(float).reset_index(drop=True)
                    X_valid = df_valid_pairs[base_features].copy().reset_index(drop=True)
                    y_valid = df_valid_pairs["y_target"].astype(float).reset_index(drop=True)

                    X_train_corr = df_train_pairs[corr_features].copy().reset_index(drop=True)
                    X_valid_corr = df_valid_pairs[corr_features].copy().reset_index(drop=True)

                    spike_train = df_train_pairs["spike_risk_tgt"].astype(int).values
                    spike_valid = df_valid_pairs["spike_risk_tgt"].astype(int).values

                    _log(f"[SPLIT] YEAR | train={len(X_train)} valid={len(X_valid)} val_year_target={base_year}")
                else:
                    _log("[WARN] fallback split chrono 80/20")

                    X_all = df_pairs[base_features].copy()
                    y_all = df_pairs["y_target"].astype(float).copy()
                    X_all_corr = df_pairs[corr_features].copy()
                    spike_all = df_pairs["spike_risk_tgt"].astype(int).values

                    n_total = len(X_all)
                    cut = int(n_total * 0.8)
                    cut = max(60, min(cut, n_total - 20))

                    X_train = X_all.iloc[:cut].reset_index(drop=True)
                    X_valid = X_all.iloc[cut:].reset_index(drop=True)
                    y_train = y_all.iloc[:cut].reset_index(drop=True)
                    y_valid = y_all.iloc[cut:].reset_index(drop=True)

                    X_train_corr = X_all_corr.iloc[:cut].reset_index(drop=True)
                    X_valid_corr = X_all_corr.iloc[cut:].reset_index(drop=True)

                    spike_train = spike_all[:cut]
                    spike_valid = spike_all[cut:]

                    _log(f"[SPLIT] CHRONO | total={n_total} train={len(X_train)} valid={len(X_valid)}")

                # ======================================================================
                # 7) Models
                # ======================================================================
                try:
                    from catboost import CatBoostClassifier, CatBoostRegressor
                    from lightgbm import LGBMClassifier, LGBMRegressor
                    from sklearn.ensemble import ExtraTreesClassifier, ExtraTreesRegressor
                except Exception as e:
                    messagebox.showerror("Erreur", f"Dépendance ML manquante : {e}")
                    return

                _log("[STEP] Début training modèles")
                tscv = TimeSeriesSplit(n_splits=3)

                def _tune(model, param_dist, X_t, y_t, scoring, n_iter=15, sample_weight=None):
                    _log(f"[TUNE START] {model.__class__.__name__} | n={len(X_t)} | n_iter={n_iter}")
                    rs = RandomizedSearchCV(
                        model,
                        param_distributions=param_dist,
                        n_iter=n_iter,
                        cv=tscv,
                        scoring=scoring,
                        random_state=SEED,
                        n_jobs=1,
                        verbose=0
                    )
                    if sample_weight is not None and model.__class__.__name__ != 'ExtraTreesClassifier':
                        try:
                            rs.fit(X_t, y_t, sample_weight=sample_weight)
                        except Exception:
                            rs.fit(X_t, y_t)
                    else:
                        rs.fit(X_t, y_t)
                    _log(f"[TUNE DONE] {model.__class__.__name__}")
                    return rs.best_estimator_, rs.best_params_, rs.best_score_

                def _normalize_weights(scores, gamma=2.0, min_w=0.05):
                    s = np.asarray(scores, dtype=float)
                    s = np.where(np.isfinite(s), s, 0.0)
                    s = np.clip(s, a_min=0.0, a_max=None)
                    if float(s.sum()) <= 0:
                        return np.ones_like(s) / len(s)
                    s_exp = np.exp(gamma * (s - np.max(s)))
                    w = s_exp / s_exp.sum()
                    w = np.maximum(w, min_w)
                    return w / w.sum()

                # ----------------------------------------------------------------------
                # 7A) Classification ensemble
                # ----------------------------------------------------------------------
                _log("[STEP] Début classifieur ensemble")

                y_cls_train = (y_train > 0).astype(int).values
                y_cls_valid = (y_valid > 0).astype(int).values

                w_cls = np.ones(len(X_train), dtype=float)
                w_cls *= np.where(spike_train == 1, 3.0, 1.0)
                w_cls *= np.where(y_cls_train == 1, 1.5, 1.0)

                cls_models = []
                cls_model_names = []

                cls_cat_base = CatBoostClassifier(
                    random_seed=SEED,
                    verbose=0,
                    loss_function="Logloss",
                    eval_metric="AUC",
                    auto_class_weights="Balanced",
                    thread_count=1
                )
                cls_cat_params = {
                    "depth": randint(4, 9),
                    "learning_rate": uniform(0.01, 0.08),
                    "iterations": randint(350, 1000),
                    "l2_leaf_reg": uniform(1.0, 8.0)
                }
                cls_cat, _, _ = _tune(cls_cat_base, cls_cat_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                cls_models.append(cls_cat)
                cls_model_names.append("CatBoost")

                cls_lgb_base = LGBMClassifier(
                    random_state=SEED,
                    objective="binary",
                    verbosity=-1,
                    n_jobs=1,
                    deterministic=True,
                    force_col_wise=True
                )
                cls_lgb_params = {
                    "n_estimators": randint(250, 800),
                    "num_leaves": randint(15, 63),
                    "max_depth": randint(3, 8),
                    "learning_rate": uniform(0.01, 0.1)
                }
                cls_lgb, _, _ = _tune(cls_lgb_base, cls_lgb_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                cls_models.append(cls_lgb)
                cls_model_names.append("LightGBM")

                if HAS_XGB:
                    cls_xgb_base = XGBClassifier(
                        random_state=SEED,
                        eval_metric="auc",
                        verbosity=0,
                        n_jobs=1,
                        tree_method="hist",
                        subsample=1.0,
                        colsample_bytree=1.0
                    )
                    cls_xgb_params = {
                        "n_estimators": randint(200, 700),
                        "max_depth": randint(3, 8),
                        "learning_rate": uniform(0.01, 0.1),
                        "subsample": uniform(0.6, 0.4)
                    }
                    cls_xgb, _, _ = _tune(cls_xgb_base, cls_xgb_params, X_train, y_cls_train, scoring="roc_auc", n_iter=12, sample_weight=w_cls)
                    cls_models.append(cls_xgb)
                    cls_model_names.append("XGBoost")

                cls_et = ExtraTreesClassifier(
                    n_estimators=800,
                    max_depth=15,
                    min_samples_leaf=2,
                    class_weight="balanced",
                    random_state=SEED,
                    n_jobs=1
                )
                cls_et.fit(X_train, y_cls_train, sample_weight=w_cls)
                cls_models.append(cls_et)
                cls_model_names.append("ExtraTrees")

                valid_cls_probas = []
                cls_scores = []
                for name, mdl in zip(cls_model_names, cls_models):
                    p = mdl.predict_proba(X_valid)[:, 1]
                    valid_cls_probas.append(p)
                    auc_i = roc_auc_score(y_cls_valid, p) if len(np.unique(y_cls_valid)) > 1 else 0.5
                    _log(f"[VALID CLS] {name} AUC={auc_i:.4f}")
                    cls_scores.append(max(auc_i - 0.50, 1e-4))

                cls_weights = _normalize_weights(cls_scores, gamma=3.0, min_w=0.05)
                _log("[CLS WEIGHTS] " + " | ".join([f"{n}={w:.3f}" for n, w in zip(cls_model_names, cls_weights)]))

                def _predict_cls_mean(X_):
                    preds = [m.predict_proba(X_)[:, 1] for m in cls_models]
                    return np.average(np.vstack(preds), axis=0, weights=cls_weights)

                _log("[STEP] Fin classifieur ensemble")

                # ----------------------------------------------------------------------
                # 7B) Régression ensemble
                # ----------------------------------------------------------------------
                _log("[STEP] Début régression ensemble")

                mask_pos = (y_train > 0).values
                X_train_pos = X_train.loc[mask_pos].reset_index(drop=True)
                y_train_pos = y_train.loc[mask_pos].reset_index(drop=True)

                if len(X_train_pos) < 40:
                    _log("[WARN] Peu de jours >0 : reg entraînée sur tout le train.")
                    X_train_pos = X_train.copy()
                    y_train_pos = y_train.copy()
                    spike_train_pos = spike_train.copy()
                else:
                    spike_train_pos = spike_train[mask_pos]

                y_train_pos_log = np.log1p(y_train_pos.values)
                q90_pos = float(np.quantile(y_train_pos.values, 0.90)) if len(y_train_pos) > 10 else float(np.max(y_train_pos.values))

                w_reg = np.ones(len(X_train_pos), dtype=float)
                w_reg *= np.where(spike_train_pos == 1, 3.0, 1.0)
                w_reg *= np.where(y_train_pos.values >= q90_pos, 2.5, 1.0)

                reg_models = []
                reg_model_names = []

                reg_cat_base = CatBoostRegressor(
                    random_seed=SEED,
                    verbose=0,
                    loss_function="RMSE",
                    thread_count=1
                )
                reg_cat_params = {
                    "depth": randint(4, 9),
                    "learning_rate": uniform(0.01, 0.08),
                    "iterations": randint(400, 1200)
                }
                reg_cat, _, _ = _tune(reg_cat_base, reg_cat_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                reg_models.append(reg_cat)
                reg_model_names.append("CatBoost")

                reg_lgb_base = LGBMRegressor(
                    random_state=SEED,
                    objective="regression",
                    verbosity=-1,
                    n_jobs=1,
                    deterministic=True,
                    force_col_wise=True
                )
                reg_lgb_params = {
                    "n_estimators": randint(300, 1000),
                    "num_leaves": randint(15, 63),
                    "max_depth": randint(3, 8),
                    "learning_rate": uniform(0.01, 0.08)
                }
                reg_lgb, _, _ = _tune(reg_lgb_base, reg_lgb_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                reg_models.append(reg_lgb)
                reg_model_names.append("LightGBM")

                if HAS_XGB:
                    reg_xgb_base = XGBRegressor(
                        random_state=SEED,
                        objective="reg:squarederror",
                        verbosity=0,
                        n_jobs=1,
                        tree_method="hist",
                        subsample=1.0,
                        colsample_bytree=1.0
                    )
                    reg_xgb_params = {
                        "n_estimators": randint(250, 800),
                        "max_depth": randint(3, 8),
                        "learning_rate": uniform(0.01, 0.08),
                        "subsample": uniform(0.6, 0.4)
                    }
                    reg_xgb, _, _ = _tune(reg_xgb_base, reg_xgb_params, X_train_pos, y_train_pos_log, scoring="neg_root_mean_squared_error", n_iter=15, sample_weight=w_reg)
                    reg_models.append(reg_xgb)
                    reg_model_names.append("XGBoost")

                reg_et = ExtraTreesRegressor(
                    n_estimators=1000,
                    max_depth=15,
                    min_samples_leaf=2,
                    random_state=SEED,
                    n_jobs=1
                )
                reg_et.fit(X_train_pos, y_train_pos_log, sample_weight=w_reg)
                reg_models.append(reg_et)
                reg_model_names.append("ExtraTrees")

                valid_log_preds = []
                reg_scores = []
                for name, mdl in zip(reg_model_names, reg_models):
                    pred_log_i = mdl.predict(X_valid)
                    valid_log_preds.append(pred_log_i)
                    pred_i = np.expm1(np.clip(pred_log_i, a_min=0.0, a_max=None))
                    rmse_i = mean_squared_error(y_valid, pred_i) ** 0.5
                    _log(f"[VALID REG] {name} RMSE(level)={rmse_i:.2f}")
                    reg_scores.append(1.0 / max(rmse_i ** 2, 1e-6))

                reg_weights = _normalize_weights(reg_scores, gamma=1.5, min_w=0.05)
                _log("[REG WEIGHTS] " + " | ".join([f"{n}={w:.3f}" for n, w in zip(reg_model_names, reg_weights)]))

                def _predict_reg_mean_log(X_):
                    preds = [m.predict(X_) for m in reg_models]
                    return np.average(np.vstack(preds), axis=0, weights=reg_weights)

                def _predict_base(X_):
                    p = _predict_cls_mean(X_)
                    pred_log = _predict_reg_mean_log(X_)
                    amt = np.expm1(np.clip(pred_log, a_min=0.0, a_max=None))
                    return p * amt

                pred_train_base = _predict_base(X_train)
                pred_valid_base = _predict_base(X_valid)

                spike_gate = float(np.quantile(pred_train_base, 0.75))
                spike_q85 = float(np.quantile(pred_train_base, 0.85))
                _log(f"[SPIKE] gate(q75)={spike_gate:.1f} | q85={spike_q85:.1f}")

                def _soft_gate(pred_base_arr, gate, sharp=0.0025):
                    z = (pred_base_arr - gate) * sharp
                    z = np.clip(z, -25, 25)
                    return 1.0 / (1.0 + np.exp(-z))

                w_train = _soft_gate(pred_train_base, spike_gate)
                w_valid = _soft_gate(pred_valid_base, spike_gate)

                _log("[STEP] Fin régression ensemble")

                # ======================================================================
                # 8) Correcteurs spike
                # ======================================================================
                _log("[STEP] Début correcteurs spike")

                resid_train = (y_train.values - pred_train_base)

                X_train_corr = X_train_corr.copy()
                X_valid_corr = X_valid_corr.copy()
                X_train_corr["pred_base"] = pred_train_base
                X_valid_corr["pred_base"] = pred_valid_base
                corr_features_plus = corr_features + ["pred_base"]

                idx_corr = np.where((spike_train == 1) | (pred_train_base >= spike_q85))[0]
                if len(idx_corr) < 80:
                    idx_corr = np.where(spike_train == 1)[0]

                def _avg_corr_predict(models, weights, X_):
                    if not models:
                        return np.zeros(len(X_), dtype=float)
                    pp = [np.clip(m.predict(X_), 0.0, None) for m in models]
                    if weights is None or len(weights) != len(models):
                        return np.mean(np.vstack(pp), axis=0)
                    return np.average(np.vstack(pp), axis=0, weights=weights)

                if len(idx_corr) < 60:
                    _log("[SPIKE CORR] Pas assez de jours corr -> OFF.")
                    corr_pos_models, corr_neg_models = [], []
                    corr_pos_weights, corr_neg_weights = [], []
                    pred_train_corr = np.zeros_like(pred_train_base)
                    pred_valid_corr = np.zeros_like(pred_valid_base)
                else:
                    X_corr_train = X_train_corr.iloc[idx_corr].reset_index(drop=True)
                    resid_corr = resid_train[idx_corr]
                    y_corr_pos = np.clip(resid_corr, 0.0, None)
                    y_corr_neg = np.clip(-resid_corr, 0.0, None)

                    def _fit_corr(y_corr, label):
                        nz = int(np.sum(y_corr > 0))
                        _log(f"[SPIKE CORR {label}] start | n={len(y_corr)} | nonzero={nz}")

                        if nz < 30:
                            _log(f"[SPIKE CORR {label}] Trop peu de non-nuls -> OFF")
                            return [], []

                        w_corr = np.ones(len(y_corr), dtype=float)
                        w_corr *= np.where(y_corr > 0, 2.0, 1.0)
                        w_corr *= np.where(X_corr_train["spike_risk_tgt"].values == 1, 3.0, 1.0)

                        corr_models_local = []
                        corr_names_local = []

                        mdl_cat = CatBoostRegressor(
                            iterations=600,
                            learning_rate=0.04,
                            depth=6,
                            random_seed=SEED,
                            verbose=0,
                            thread_count=1
                        )
                        mdl_cat.fit(X_corr_train, y_corr, sample_weight=w_corr)
                        corr_models_local.append(mdl_cat)
                        corr_names_local.append("CatBoost")

                        mdl_lgb = LGBMRegressor(
                            n_estimators=400,
                            learning_rate=0.04,
                            num_leaves=31,
                            random_state=SEED,
                            verbosity=-1,
                            n_jobs=1,
                            deterministic=True,
                            force_col_wise=True
                        )
                        mdl_lgb.fit(X_corr_train, y_corr, sample_weight=w_corr)
                        corr_models_local.append(mdl_lgb)
                        corr_names_local.append("LightGBM")

                        if HAS_XGB:
                            mdl_xgb = XGBRegressor(
                                n_estimators=300,
                                learning_rate=0.04,
                                max_depth=5,
                                random_state=SEED,
                                verbosity=0,
                                n_jobs=1,
                                tree_method="hist",
                                subsample=1.0,
                                colsample_bytree=1.0
                            )
                            mdl_xgb.fit(X_corr_train, y_corr, sample_weight=w_corr)
                            corr_models_local.append(mdl_xgb)
                            corr_names_local.append("XGBoost")

                        scores_local = []
                        for nm, mdl in zip(corr_names_local, corr_models_local):
                            p = np.clip(mdl.predict(X_corr_train), 0.0, None)
                            rmse_local = mean_squared_error(y_corr, p) ** 0.5
                            scores_local.append(1.0 / max(rmse_local ** 2, 1e-4))

                        weights_local = _normalize_weights(scores_local, min_w=0.10)
                        return corr_models_local, weights_local

                    corr_pos_models, corr_pos_weights = _fit_corr(y_corr_pos, "POS")
                    corr_neg_models, corr_neg_weights = _fit_corr(y_corr_neg, "NEG")

                    pred_train_corr = np.zeros_like(pred_train_base, dtype=float)
                    pred_valid_corr = np.zeros_like(pred_valid_base, dtype=float)

                    if corr_pos_models:
                        pred_train_corr[idx_corr] += _avg_corr_predict(corr_pos_models, corr_pos_weights, X_train_corr.iloc[idx_corr])
                    if corr_neg_models:
                        pred_train_corr[idx_corr] -= _avg_corr_predict(corr_neg_models, corr_neg_weights, X_train_corr.iloc[idx_corr])

                    idx_spike_val = np.where(spike_valid == 1)[0]
                    if len(idx_spike_val) > 0:
                        if corr_pos_models:
                            pred_valid_corr[idx_spike_val] += _avg_corr_predict(corr_pos_models, corr_pos_weights, X_valid_corr.iloc[idx_spike_val])
                        if corr_neg_models:
                            pred_valid_corr[idx_spike_val] -= _avg_corr_predict(corr_neg_models, corr_neg_weights, X_valid_corr.iloc[idx_spike_val])

                pred_train = pred_train_base + pred_train_corr * w_train
                pred_valid = pred_valid_base + pred_valid_corr * w_valid

                _log("[STEP] Fin correcteurs spike")
                _log("[STEP] Fin training modèles")

                # ======================================================================
                # 9) Scale + Offset
                # ======================================================================
                denom = float(np.sum(pred_valid))
                if denom > 1e-9:
                    scale = float(np.sum(y_valid.values) / denom)
                    scale = max(0.85, min(scale, 1.15))
                else:
                    scale = 1.0

                pred_train *= scale
                pred_valid *= scale
                _log(f"[SCALE] scale={scale:.4f}")

                offset = float(np.median(y_valid.values - pred_valid))
                offset = max(-5000.0, min(offset, 5000.0))
                pred_train = np.clip(pred_train + offset, 0.0, None)
                pred_valid = np.clip(pred_valid + offset, 0.0, None)
                _log(f"[OFFSET] median residual offset={offset:.1f}")

                # ======================================================================
                # 10) VALID viz + résidus
                # ======================================================================
                df_valid_vis = df_valid_pairs.iloc[:len(X_valid)].copy().reset_index(drop=True)
                yv = y_valid.reset_index(drop=True).values if hasattr(y_valid, "reset_index") else y_valid.values

                df_valid_vis["y_true"] = yv
                df_valid_vis["y_pred"] = pred_valid
                df_valid_vis["resid"] = df_valid_vis["y_true"] - df_valid_vis["y_pred"]
                df_valid_vis["spike_risk_tgt"] = spike_valid

                df_valid_vis["date_tgt"] = [
                    _safe_date_from_year_doy(y, d)
                    for y, d in zip(df_valid_vis["year_target"].astype(int).values,
                                    df_valid_vis["dayofyear"].astype(int).values)
                ]
                df_valid_vis = df_valid_vis.sort_values("date_tgt").reset_index(drop=True)

                model_artifacts["valid_vis"] = df_valid_vis
                model_artifacts["valid_resid_series"] = df_valid_vis["resid"].astype(float).values
                model_artifacts["valid_dates"] = df_valid_vis["date_tgt"].values
                model_artifacts["valid_spike"] = df_valid_vis["spike_risk_tgt"].values.astype(int)
                model_artifacts["base_year"] = base_year
                model_artifacts["filiale"] = filiale

                # ======================================================================
                # 11) KPI P50
                # ======================================================================
                mae = mean_absolute_error(y_valid, pred_valid)
                rmse = mean_squared_error(y_valid, pred_valid) ** 0.5
                r2 = r2_score(y_valid, pred_valid)

                _log(f"[MODEL V4.0 ENSEMBLE ULTIME] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")

                # ======================================================================
                # 12) Calibration 2D
                # ======================================================================
                _log("[STEP] Début calibration/conformal")

                res = (y_valid.values - pred_valid)

                n_bins = 5 if len(pred_valid) < 180 else 7
                _log(f"[CALIB] n_valid={len(pred_valid)} | n_bins={n_bins}")

                qs = np.quantile(pred_valid, np.linspace(0, 1, n_bins + 1))
                bins = np.digitize(pred_valid, qs[1:-1], right=True)
                spk = spike_valid

                q_low_2d = np.zeros((n_bins, 2), dtype=float)
                q_high_2d = np.zeros((n_bins, 2), dtype=float)

                glob_low = float(np.quantile(res, 0.05))
                glob_high = float(np.quantile(res, 0.95))

                for b in range(n_bins):
                    for s_ in (0, 1):
                        rr = res[(bins == b) & (spk == s_)]
                        if len(rr) < 25:
                            q_low_2d[b, s_] = glob_low
                            q_high_2d[b, s_] = glob_high
                        else:
                            q_low_2d[b, s_] = float(np.quantile(rr, 0.05))
                            q_high_2d[b, s_] = float(np.quantile(rr, 0.95))

                adj_low = np.array([q_low_2d[b, s_] for b, s_ in zip(bins, spk)])
                adj_high = np.array([q_high_2d[b, s_] for b, s_ in zip(bins, spk)])

                pred_p05_valid = np.clip(pred_valid + adj_low, 0.0, None)
                pred_p95_valid = np.clip(pred_valid + adj_high, 0.0, None)

                coverage_95 = float(np.mean((y_valid.values >= pred_p05_valid) & (y_valid.values <= pred_p95_valid)))
                width_95 = float(np.mean(pred_p95_valid - pred_p05_valid))
                alert_95 = 1.0 - coverage_95

                model_artifacts["calib_qs"] = qs
                model_artifacts["calib_q_low_2d_05"] = q_low_2d
                model_artifacts["calib_q_high_2d_95"] = q_high_2d

                _log(f"[CALIB 2D 95% bins={n_bins}] Coverage={coverage_95:.3f} | Alert={alert_95:.3f} | Width={width_95:.1f}")

                # ======================================================================
                # 13) Conformal conditionnel
                # ======================================================================
                alpha = 0.05

                s_conf = np.maximum.reduce([
                    pred_p05_valid - y_valid.values,
                    y_valid.values - pred_p95_valid,
                    np.zeros_like(y_valid.values)
                ])

                q_conf_by_spk = {}
                for g in (0, 1):
                    ss = s_conf[spk == g]
                    if len(ss) < 40:
                        q_conf_by_spk[g] = float(np.quantile(s_conf, 1 - alpha))
                    else:
                        q_conf_by_spk[g] = float(np.quantile(ss, 1 - alpha))

                model_artifacts["conformal_alpha"] = alpha
                model_artifacts["conformal_q_by_spk"] = q_conf_by_spk

                q_adj = np.array([q_conf_by_spk[int(s)] for s in spk])
                pred_p05_valid_c = np.clip(pred_p05_valid - q_adj, 0.0, None)
                pred_p95_valid_c = pred_p95_valid + q_adj

                coverage_c = float(np.mean((y_valid.values >= pred_p05_valid_c) & (y_valid.values <= pred_p95_valid_c)))
                width_c = float(np.mean(pred_p95_valid_c - pred_p05_valid_c))
                alert_c = 1.0 - coverage_c

                _log(f"[CONFORMAL 95%] q(spk0)={q_conf_by_spk[0]:.1f} | q(spk1)={q_conf_by_spk[1]:.1f}")
                _log(f"[CONFORMAL CHECK 95%] Coverage={coverage_c:.3f} | Alert={alert_c:.3f} | Width={width_c:.1f}")
                _log("[STEP] Fin calibration/conformal")

                # ======================================================================
                # 14) Forecast N+1 + MC
                # ======================================================================
                _log("[STEP] Début forecast N+1")

                target_year = base_year + 1
                current_target_year = target_year
                model_artifacts["target_year"] = target_year

                df_real_target = df_filiale[df_filiale["year"] == target_year].copy().sort_values("date")
                current_real_target_df = df_real_target

                df_prev_year = df_filiale[df_filiale["year"] == base_year].copy().sort_values("date")
                target_holidays = holiday_sets.get(int(target_year), set())

                calmap_tgt_rows = []
                for mm in range(1, 13):
                    try:
                        calmap_tgt_rows.append({
                            "month": int(mm),
                            "eff22_date": _effective_22_date(int(target_year), int(mm), target_holidays),
                            "lbd_date": _last_business_day(int(target_year), int(mm), target_holidays),
                        })
                    except Exception:
                        pass
                calmap_tgt = pd.DataFrame(calmap_tgt_rows).set_index("month")

                def _predict_base_row(feat_base_df):
                    p = float(_predict_cls_mean(feat_base_df)[0])
                    pred_log = float(_predict_reg_mean_log(feat_base_df)[0])
                    amt = float(np.expm1(max(0.0, pred_log)))
                    return p * amt

                def _block_bootstrap_residual_paths_weekday_conditional(
                    resid_series: np.ndarray,
                    valid_dates: np.ndarray,
                    valid_spike: np.ndarray,
                    horizon_dates: np.ndarray,
                    horizon_spike: np.ndarray,
                    B: int = 2000,
                    block_len: int = 7,
                    seed: int = SEED
                ):
                    rng = np.random.default_rng(seed)

                    r = np.asarray(resid_series, dtype=float)
                    vd = pd.to_datetime(valid_dates)
                    vs = np.asarray(valid_spike, dtype=int)

                    hd = pd.to_datetime(horizon_dates)
                    hs = np.asarray(horizon_spike, dtype=int)

                    mask = np.isfinite(r)
                    r, vd, vs = r[mask], vd[mask], vs[mask]

                    T = len(r)
                    N = len(hd)
                    if T < max(30, block_len + 5):
                        idx = rng.integers(0, T, size=(B, N))
                        return r[idx]

                    vd_w = vd.dayofweek.values
                    hd_w = hd.dayofweek.values

                    max_start = T - block_len
                    all_starts = list(range(max_start)) if max_start > 0 else [0]

                    starts = {(w, s): [] for w in range(7) for s in (0, 1)}
                    for s0 in range(max_start):
                        starts[(int(vd_w[s0]), int(vs[s0]))].append(s0)

                    for w in range(7):
                        for s_ in (0, 1):
                            if len(starts[(w, s_)]) == 0:
                                starts[(w, s_)] = all_starts

                    paths = np.zeros((B, N), dtype=float)
                    for b in range(B):
                        pos = 0
                        while pos < N:
                            w0 = int(hd_w[pos])
                            sp0 = int(hs[pos])
                            s0 = int(rng.choice(starts[(w0, sp0)]))
                            block = r[s0:s0 + block_len]
                            take = min(block_len, N - pos)
                            paths[b, pos:pos + take] = block[:take]
                            pos += take
                    return paths

                def _mc_dynamic_band_weekday_conditional(
                    pred_values: np.ndarray,
                    horizon_dates: np.ndarray,
                    horizon_spike: np.ndarray,
                    resid_series: np.ndarray,
                    valid_dates: np.ndarray,
                    valid_spike: np.ndarray,
                    B: int = 2000,
                    block_len: int = 7,
                    seed: int = SEED,
                    return_sims: bool = False
                ):
                    pv = np.asarray(pred_values, dtype=float)
                    resid_paths = _block_bootstrap_residual_paths_weekday_conditional(
                        resid_series=resid_series,
                        valid_dates=valid_dates,
                        valid_spike=valid_spike,
                        horizon_dates=horizon_dates,
                        horizon_spike=horizon_spike,
                        B=B,
                        block_len=block_len,
                        seed=seed
                    )
                    sims = np.clip(pv[None, :] + resid_paths, 0.0, None)
                    p10 = np.quantile(sims, 0.10, axis=0)
                    p50 = np.quantile(sims, 0.50, axis=0)
                    p90 = np.quantile(sims, 0.90, axis=0)
                    if return_sims:
                        return p10.astype(float), p50.astype(float), p90.astype(float), sims.astype(float)
                    return p10.astype(float), p50.astype(float), p90.astype(float)

                future_rows = []
                carryover = 0.0

                for i_row, (_, row) in enumerate(df_prev_year.iterrows(), start=1):
                    if i_row % 50 == 0:
                        _log(f"[FORECAST LOOP] {i_row}/{len(df_prev_year)}")

                    d_prev = row["date"]
                    d_next = d_prev + pd.DateOffset(years=1)
                    if int(d_next.year) != int(target_year):
                        continue

                    dow_next = int(d_next.weekday())
                    is_h_next = 1 if (d_next.date() in target_holidays) else 0

                    month_tgt = int(d_next.month)
                    dom_tgt = int(d_next.day)

                    eff22 = calmap_tgt.loc[month_tgt, "eff22_date"] if month_tgt in calmap_tgt.index else None
                    is_22_eff_tgt = 1 if (eff22 is not None and d_next.normalize() == eff22) else 0
                    days_to_22_eff_tgt = int((d_next.normalize() - eff22).days) if eff22 is not None else 0
                    abs_days_to_22_eff_tgt = abs(days_to_22_eff_tgt)
                    is_near_22_eff_tgt = 1 if abs_days_to_22_eff_tgt <= 2 else 0

                    is_dom_22_tgt = 1 if dom_tgt == 22 else 0
                    days_to_22_tgt = int(dom_tgt - 22)

                    is_mid_window_tgt = 1 if (20 <= dom_tgt <= 22) else 0
                    abs_days_to_20_tgt = abs(dom_tgt - 20)
                    is_near_20_tgt = 1 if abs_days_to_20_tgt <= 1 else 0
                    is_monday_tgt = 1 if dow_next == 0 else 0
                    is_monday_after_20_tgt = 1 if (dow_next == 0 and 20 <= dom_tgt <= 22) else 0

                    lbd = calmap_tgt.loc[month_tgt, "lbd_date"] if month_tgt in calmap_tgt.index else None
                    is_lbd_tgt = 1 if (lbd is not None and d_next.normalize() == lbd) else 0
                    days_to_lbd_tgt = int((d_next.normalize() - lbd).days) if lbd is not None else 0
                    abs_days_to_lbd_tgt = abs(days_to_lbd_tgt)
                    is_near_lbd_tgt = 1 if abs_days_to_lbd_tgt <= 1 else 0

                    spike_risk_tgt = 1 if (is_near_22_eff_tgt or is_mid_window_tgt or is_monday_after_20_tgt or is_near_lbd_tgt) else 0

                    try:
                        bdm_tgt = int(df_filiale.loc[df_filiale["date_norm"] == d_next.normalize(), "business_day_of_month"].iloc[0])
                        bdeom_tgt = int(df_filiale.loc[df_filiale["date_norm"] == d_next.normalize(), "business_days_to_eom"].iloc[0])
                    except Exception:
                        bdm_tgt = 0
                        bdeom_tgt = 0

                    feat = {c: 0 for c in base_features}

                    feat["y_prev_year"] = float(row["y"])
                    feat["y_raw_prev_year"] = float(row.get("y_raw", row["y"]))
                    feat["roll_prev_7"] = float(row["roll_mean_7"])
                    feat["roll_prev_30"] = float(row["roll_mean_30"])
                    feat["roll_std_prev_7"] = float(row.get("roll_std_7", 0))
                    feat["roll_std_prev_30"] = float(row.get("roll_std_30", 0))
                    feat["roll_prev_7_raw"] = float(row.get("roll_mean_7_raw", row["roll_mean_7"]))
                    feat["roll_prev_30_raw"] = float(row.get("roll_mean_30_raw", row["roll_mean_30"]))
                    feat["dayofyear"] = int(row["dayofyear"])
                    feat["section_id"] = 0
                    feat["dow_prev"] = int(row["dow"])
                    feat["lag_1_prev"] = float(row["lag_1"]) if pd.notna(row["lag_1"]) else 0.0
                    feat["lag_7_prev"] = float(row["lag_7"]) if pd.notna(row["lag_7"]) else 0.0
                    feat["lag_1_prev_raw"] = float(row.get("lag_1_raw", row["lag_1"])) if pd.notna(row.get("lag_1_raw", np.nan)) else 0.0
                    feat["lag_7_prev_raw"] = float(row.get("lag_7_raw", row["lag_7"])) if pd.notna(row.get("lag_7_raw", np.nan)) else 0.0
                    feat["month_prev"] = int(row["month"])
                    feat["dom_prev"] = int(row["dom"])
                    feat["is_eom_prev"] = int(row["is_eom"])
                    feat["is_holiday_prev"] = int(row["is_holiday"])
                    feat["bdm_prev"] = int(row.get("business_day_of_month", 0))
                    feat["bdeom_prev"] = int(row.get("business_days_to_eom", 0))
                    feat["is_dom_22_prev"] = int(row.get("is_dom_22", 0))
                    feat["days_to_22_prev"] = int(row.get("days_to_22", int(row["dom"]) - 22))
                    feat["is_22_effective_prev"] = int(row.get("is_22_effective", 0))
                    feat["days_to_22_effective_prev"] = int(row.get("days_to_22_effective", 0))
                    feat["abs_days_to_22_effective_prev"] = int(row.get("abs_days_to_22_effective", abs(int(row.get("days_to_22_effective", 0)))))
                    feat["is_near_22_effective_prev"] = int(row.get("is_near_22_effective", 0))
                    feat["is_mid_month_window_prev"] = int(row.get("is_mid_month_window", 1 if 20 <= int(row["dom"]) <= 22 else 0))
                    feat["abs_days_to_20_prev"] = int(row.get("abs_days_to_20", abs(int(row["dom"]) - 20)))
                    feat["is_near_20_prev"] = int(row.get("is_near_20", 1 if abs(int(row["dom"]) - 20) <= 1 else 0))
                    feat["is_monday_prev"] = int(row.get("is_monday", 1 if int(row["dow"]) == 0 else 0))
                    feat["is_monday_after_20_prev"] = int(row.get("is_monday_after_20", 1 if (int(row["dow"]) == 0 and 20 <= int(row["dom"]) <= 22) else 0))
                    feat["is_lbd_prev"] = int(row.get("is_lbd", 0))
                    feat["days_to_lbd_prev"] = int(row.get("days_to_lbd", 0))
                    feat["abs_days_to_lbd_prev"] = int(row.get("abs_days_to_lbd", abs(int(row.get("days_to_lbd", 0)))))
                    feat["is_near_lbd_prev"] = int(row.get("is_near_lbd", 0))

                    feat["dow_tgt"] = dow_next
                    feat["month_tgt"] = month_tgt
                    feat["dom_tgt"] = dom_tgt
                    feat["is_eom_tgt"] = int(d_next.is_month_end)
                    feat["is_holiday_tgt"] = is_h_next
                    feat["bdm_tgt"] = bdm_tgt
                    feat["bdeom_tgt"] = bdeom_tgt
                    feat["is_dom_22_tgt"] = is_dom_22_tgt
                    feat["days_to_22_tgt"] = days_to_22_tgt
                    feat["is_22_effective_tgt"] = is_22_eff_tgt
                    feat["days_to_22_effective_tgt"] = days_to_22_eff_tgt
                    feat["abs_days_to_22_effective_tgt"] = abs_days_to_22_eff_tgt
                    feat["is_near_22_effective_tgt"] = is_near_22_eff_tgt
                    feat["is_mid_month_window_tgt"] = is_mid_window_tgt
                    feat["abs_days_to_20_tgt"] = abs_days_to_20_tgt
                    feat["is_near_20_tgt"] = is_near_20_tgt
                    feat["is_monday_tgt"] = is_monday_tgt
                    feat["is_monday_after_20_tgt"] = is_monday_after_20_tgt
                    feat["is_lbd_tgt"] = is_lbd_tgt
                    feat["days_to_lbd_tgt"] = days_to_lbd_tgt
                    feat["abs_days_to_lbd_tgt"] = abs_days_to_lbd_tgt
                    feat["is_near_lbd_tgt"] = is_near_lbd_tgt

                    feat_base = pd.DataFrame([feat], columns=base_features)
                    pred_base = _predict_base_row(feat_base)

                    pred_corr = 0.0
                    if spike_risk_tgt == 1:
                        w = float(_soft_gate(np.array([pred_base]), spike_gate)[0])
                        feat_corr = pd.DataFrame([feat_base.iloc[0].tolist() + [
                            int(row.get("spike_risk", 0)),
                            int(spike_risk_tgt),
                            float(pred_base)
                        ]], columns=corr_features_plus)

                        if corr_pos_models:
                            pred_corr += float(_avg_corr_predict(corr_pos_models, corr_pos_weights, feat_corr)[0])
                        if corr_neg_models:
                            pred_corr -= float(_avg_corr_predict(corr_neg_models, corr_neg_weights, feat_corr)[0])

                        pred_corr *= w

                    pred_raw = pred_base + pred_corr

                    is_closed = (dow_next >= 5) or (is_h_next == 1)
                    if is_closed:
                        carryover += pred_raw
                        pred_det = 0.0
                    else:
                        pred_det = pred_raw + carryover
                        carryover = 0.0

                    pred_det = pred_det * scale + offset
                    if pred_det < 0:
                        pred_det = 0.0

                    future_rows.append({
                        "section": filiale,
                        "date": d_next,
                        "year": int(d_next.year),
                        "month": int(d_next.month),
                        "dayofyear": int(d_next.dayofyear),
                        "spike_risk_tgt": int(spike_risk_tgt),
                        "pred_det": float(pred_det),
                    })

                df_future_all = pd.DataFrame(future_rows).sort_values("date").reset_index(drop=True)
                _log(f"[FORECAST] nb_future_rows={len(df_future_all)}")
                _log("[STEP] Fin forecast N+1")

                # --- MC ---
                B_MC = 5000
                BLOCK_LEN = 7
                resid_series = model_artifacts.get("valid_resid_series", None)
                valid_dates = model_artifacts.get("valid_dates", None)
                valid_spike = model_artifacts.get("valid_spike", None)

                _log("[STEP] Début MC")
                _log(f"[MC CONFIG] B={B_MC} | block_len={BLOCK_LEN} | horizon={len(df_future_all)}")

                if resid_series is None or valid_dates is None or valid_spike is None or len(resid_series) < 50 or df_future_all.empty:
                    _log("[MC] Pas assez de résidus VALID -> fallback deterministe.")
                    df_future_all["pred_p10"] = df_future_all["pred_det"]
                    df_future_all["pred_value"] = df_future_all["pred_det"]
                    df_future_all["pred_p90"] = df_future_all["pred_det"]
                    model_artifacts["mc_sims"] = None
                else:
                    p10_mc, p50_mc, p90_mc, sims = _mc_dynamic_band_weekday_conditional(
                        pred_values=df_future_all["pred_det"].values,
                        horizon_dates=df_future_all["date"].values,
                        horizon_spike=df_future_all["spike_risk_tgt"].fillna(0).astype(int).values,
                        resid_series=resid_series,
                        valid_dates=valid_dates,
                        valid_spike=valid_spike,
                        B=B_MC,
                        block_len=BLOCK_LEN,
                        seed=SEED,
                        return_sims=True
                    )
                    df_future_all["pred_value"] = p50_mc
                    model_artifacts["mc_sims"] = sims
                    _log(f"[MC] Conditional weekday-block OK (B={B_MC}, block_len={BLOCK_LEN}).")

                _log("[STEP] Fin MC")

                # ======================================================================
                # 15) Bandes forecast
                # ======================================================================
                qs = model_artifacts.get("calib_qs", None)
                q_low_2d = model_artifacts.get("calib_q_low_2d_05", None)
                q_high_2d = model_artifacts.get("calib_q_high_2d_95", None)
                q_conf_by_spk = model_artifacts.get("conformal_q_by_spk", {0: 0.0, 1: 0.0})

                pv = df_future_all["pred_value"].values.astype(float)
                spk_f = df_future_all["spike_risk_tgt"].fillna(0).astype(int).values

                if qs is not None and q_low_2d is not None and q_high_2d is not None:
                    bins_f = np.digitize(pv, qs[1:-1], right=True)
                    adj_low_f = np.array([q_low_2d[b, s_] for b, s_ in zip(bins_f, spk_f)])
                    adj_high_f = np.array([q_high_2d[b, s_] for b, s_ in zip(bins_f, spk_f)])
                    p05 = np.clip(pv + adj_low_f, 0.0, None)
                    p95 = np.clip(pv + adj_high_f, 0.0, None)
                else:
                    p05 = pv.copy()
                    p95 = pv.copy()

                q_adj_f = np.array([q_conf_by_spk[int(s)] for s in spk_f])
                df_future_all["pred_p05"] = np.clip(p05 - q_adj_f, 0.0, None)
                df_future_all["pred_p95"] = p95 + q_adj_f

                df_future_all["pred_p10"] = df_future_all["pred_p05"]
                df_future_all["pred_p90"] = df_future_all["pred_p95"]

                model_artifacts["mc_dates"] = df_future_all["date"].values
                model_artifacts["mc_det"] = df_future_all["pred_det"].values
                model_artifacts["mc_p50"] = df_future_all["pred_value"].values

                if current_real_target_df is not None and not current_real_target_df.empty:
                    df_future_all = pd.merge(
                        current_real_target_df[["date"]].copy(),
                        df_future_all,
                        on="date",
                        how="left"
                    )
                    for c in ["pred_det", "pred_p05", "pred_value", "pred_p95", "pred_p10", "pred_p90", "spike_risk_tgt"]:
                        if c in df_future_all.columns:
                            df_future_all[c] = df_future_all[c].fillna(0.0)
                    df_future_all["section"] = filiale
                    df_future_all["year"] = df_future_all["date"].dt.year
                    df_future_all["month"] = df_future_all["date"].dt.month
                    df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

                # ======================================================================
                # 16) Graphe N+1
                # ======================================================================
                fig1, ax1 = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
                ax1.set_facecolor("#00122e")

                df_hist_plot = df_filiale[df_filiale["year"] <= base_year]
                ax1.plot(df_hist_plot["date"], df_hist_plot["y"], label=f"Réel (≤ {base_year})", linewidth=2)

                ax1.plot(df_future_all["date"], df_future_all["pred_value"],
                        label=f"Prévision {target_year} (P50)", linewidth=2, linestyle="--")
                ax1.fill_between(df_future_all["date"], df_future_all["pred_p05"], df_future_all["pred_p95"],
                                alpha=0.25, label=f"Bande 95% (Calib 2D + conformal, alpha={alpha})")

                ax1.set_title("SA_VOYAGEURS – Trafic voyageur - Prévision N+1 (ENSEMBLE V4.0)", color="white")
                ax1.tick_params(axis='x', colors="white", rotation=30)
                ax1.tick_params(axis='y', colors="white")
                for spine in ax1.spines.values():
                    spine.set_color("#00122e")

                leg1 = ax1.legend(facecolor="#00122e", edgecolor="white")
                for text in leg1.get_texts():
                    text.set_color("white")

                canvas_fig1 = FigureCanvasTkAgg(fig1, master=scrollable_frame)
                canvas_fig1.draw()
                w_fig1 = canvas_fig1.get_tk_widget()
                w_fig1.pack(pady=10, fill="both", expand=True)
                graph_widgets.append(w_fig1)
                plt.close(fig1)

                # ======================================================================
                # 17) Update globals + artifacts + UI
                # ======================================================================
                current_pred_df = df_future_all
                exported_pred_df = df_future_all.copy()

                model_artifacts["X_train"] = X_train.copy()
                model_artifacts["y_train"] = y_train.copy()
                model_artifacts["X_valid"] = X_valid.copy()
                model_artifacts["y_valid"] = y_valid.copy()
                model_artifacts["y_cls_valid"] = (y_valid > 0).astype(int)

                model_artifacts["features"] = base_features[:]
                model_artifacts["corr_features"] = corr_features_plus[:]

                model_artifacts["cls_models"] = cls_models
                model_artifacts["cls_model_names"] = cls_model_names
                model_artifacts["cls_model_weights"] = cls_weights

                model_artifacts["reg_models"] = reg_models
                model_artifacts["reg_model_names"] = reg_model_names
                model_artifacts["reg_model_weights"] = reg_weights

                model_artifacts["corr_pos_models"] = corr_pos_models
                model_artifacts["corr_neg_models"] = corr_neg_models
                model_artifacts["corr_pos_weights"] = corr_pos_weights if 'corr_pos_weights' in locals() else []
                model_artifacts["corr_neg_weights"] = corr_neg_weights if 'corr_neg_weights' in locals() else []

                model_artifacts["scale"] = scale
                model_artifacts["offset"] = offset
                model_artifacts["spike_gate"] = spike_gate
                model_artifacts["spike_q85"] = spike_q85
                model_artifacts["soft_gate_sharp"] = 0.0025
                model_artifacts["mc_B"] = B_MC
                model_artifacts["mc_block_len"] = BLOCK_LEN
                model_artifacts["seed"] = SEED

                try:
                    btn_3d.configure(state="normal")
                except Exception:
                    pass

                try:
                    export_graph2_button.configure(state="normal")
                except Exception:
                    pass

                _redraw_monthly_graph()
                _redraw_graph2()
                _redraw_mc_fan_graph()
                _redraw_residuals_graph()
                _redraw_confusion_matrix_graph()
                _redraw_confusion_matrix_pred_graph()

                try:
                    _redraw_cls_diagnostics_graph()
                except Exception:
                    pass

                if analysis_table_frame is not None:
                    analysis_table_frame.destroy()

                analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
                analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
                graph_widgets.append(analysis_table_frame)

                _rebuild_analysis_table()

                _log("[DONE] Entraînement terminé")
                _log(f"[TIME] total = {(time.perf_counter() - t0_global)/60.0:.2f} min")

            except Exception:
                messagebox.showerror("Erreur", traceback.format_exc())
                print("[IA] ERREUR:\n", traceback.format_exc())
        
        def _train_ace_invest_reseau_negative_v33():
            """
            RESEAU — ACE & Investissements — NEG “V5.0 ULTRA PRÉCIS” (jour-par-jour)

            OBJECTIF (basé sur tes constats 2024–2025) :
            - Flux très concentrés J1–J3 (19/26) => modèle fortement "event-driven" (calendrier / BDM / lundi)
            - Montants DISCRETS très récurrents : -30/-35/-45/-55 (intermédiaires) et gros autour de -120 (-110 à -140)
            - Quelques spikes rares (>= -150 / jusqu’à -381)
            - Lundi très lourd => features + règles de calibration
            - Jours fermés (WE + fériés) => y=0 (règle métier)
            - y_pred <= 0

            DESIGN (anti-lissage, précis jour-par-jour) :
            1) Modèle “RÉGIME” multiclass (0=0, 1=petit récurrent, 2=gros récurrent (~120), 3=spike/other)
            2) Modèle “ANCHOR” pour régime 1 (choisit parmi les montants discrets fréquents)
            3) Modèle “ANCHOR” pour régime 2 (choisit parmi les gros montants fréquents)
            4) Modèle “SPIKE REG” (régression log1p magnitude) pour régime 3
            5) Post-calibration métier : début de mois + lundi => pousse vers activité si le modèle hésite
            6) CAP : très haut (quantiles) uniquement pour éviter les explosions artificielles, mais ne coupe pas les vrais spikes

            IMPORTANT :
            - Ici on privilégie la précision "point" => on réduit l'influence des lags/rolling (qui lissent).
            Le modèle est majoritairement calendrier + BDM + interactions.
            """

            nonlocal graph_widgets
            nonlocal current_pred_df, current_real_target_df, current_target_year, current_filiale_name
            nonlocal exported_pred_df, analysis_table_frame

            _clear_graph_widgets()
            print("======================== TRAIN NEG V5.0 ULTRA (RESEAU / ACE & Investissements) ========================")

            import numpy as np
            import pandas as pd
            import matplotlib.pyplot as plt
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

            from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
            from sklearn.model_selection import RandomizedSearchCV
            from scipy.stats import randint, uniform

            try:
                from lightgbm import LGBMClassifier, LGBMRegressor
            except Exception:
                messagebox.showerror("Erreur", "LightGBM n'est pas dispo. Installe lightgbm.")
                return

            # ---------------- Filiale + data ----------------
            filiale = "RESEAU"
            current_filiale_name = filiale

            df_filiale = df_current[df_current["section"] == filiale].copy()
            if df_filiale.empty:
                messagebox.showinfo("Information", f"Aucune donnée pour {filiale} (flux={selected_flux.get()}).")
                return
            df_filiale = df_filiale.sort_values("date").copy()

            # ---------------- base year ----------------
            try:
                base_year = int(annees_var.get())
            except Exception:
                base_year = int(df_filiale["date"].dt.year.max())
            target_year = base_year + 1
            current_target_year = target_year
            print(f"[NEG] base_year={base_year} -> target_year={target_year}")

            # ---------------- Holidays (best effort) ----------------
            holiday_sets = {}
            if "is_holiday" not in df_filiale.columns:
                df_filiale["is_holiday"] = 0
                try:
                    import holidays
                    years = sorted(set(df_filiale["date"].dt.year.unique().tolist() + [target_year]))
                    for yy in years:
                        fr_h = holidays.country_holidays("FR", years=[int(yy)])
                        holiday_sets[int(yy)] = set(fr_h.keys())
                    df_filiale["is_holiday"] = df_filiale["date"].dt.date.apply(
                        lambda d: 1 if d in holiday_sets.get(int(pd.Timestamp(d).year), set()) else 0
                    )
                    print(f"[HOLIDAYS] Calcul FR OK years={sorted(holiday_sets.keys())}")
                except Exception:
                    print("[HOLIDAYS] fallback is_holiday=0")
                    holiday_sets = {}
            else:
                try:
                    years = sorted(set(df_filiale["date"].dt.year.unique().tolist() + [target_year]))
                    for yy in years:
                        sub = df_filiale[df_filiale["date"].dt.year == int(yy)]
                        holiday_sets[int(yy)] = set(sub.loc[sub["is_holiday"] == 1, "date"].dt.date.tolist())
                except Exception:
                    holiday_sets = {}

            # ---------------- Closed days ----------------
            df_filiale["dow"] = df_filiale["date"].dt.weekday
            df_filiale["year"] = df_filiale["date"].dt.year
            df_filiale["is_closed"] = ((df_filiale["dow"] >= 5) | (df_filiale["is_holiday"] == 1)).astype(int)

            # règle métier: WE/ferié = 0
            df_filiale.loc[df_filiale["is_closed"] == 1, "y"] = 0.0

            # ---------------- Feature engineering (calendrier riche, anti-lissage) ----------------
            def _add_calendar_features(dd: pd.DataFrame, holiday_sets_: dict) -> pd.DataFrame:
                d = dd.copy()
                d = d.sort_values("date").copy()
                d["date_norm"] = d["date"].dt.normalize()
                d["month"] = d["date"].dt.month
                d["dom"] = d["date"].dt.day
                d["dayofyear"] = d["date"].dt.dayofyear
                d["weekofyear"] = d["date"].dt.isocalendar().week.astype(int)

                d["is_month_start"] = d["date"].dt.is_month_start.astype(int)
                d["is_month_end"] = d["date"].dt.is_month_end.astype(int)
                d["is_quarter_start"] = d["date"].dt.is_quarter_start.astype(int)
                d["is_quarter_end"] = d["date"].dt.is_quarter_end.astype(int)

                d["is_monday"] = (d["dow"] == 0).astype(int)

                # BOM (très important)
                d["dom_1"] = (d["dom"] == 1).astype(int)
                d["dom_2"] = (d["dom"] == 2).astype(int)
                d["dom_3"] = (d["dom"] == 3).astype(int)
                d["is_bom_1_3"] = (d["dom"] <= 3).astype(int)
                d["is_bom_1_5"] = (d["dom"] <= 5).astype(int)
                d["is_bom_1_7"] = (d["dom"] <= 7).astype(int)

                # days to end of month
                eom = (d["date"] + pd.offsets.MonthEnd(0)).dt.normalize()
                d["days_to_eom"] = (eom - d["date_norm"]).dt.days.astype(int)
                d["is_near_eom3"] = (d["days_to_eom"] <= 3).astype(int)
                d["is_near_eom7"] = (d["days_to_eom"] <= 7).astype(int)

                # business day of month (WE + fériés)
                bdm, bdl = [], []
                for (yy, mm), grp in d.groupby(["year", "month"], sort=False):
                    hset = holiday_sets_.get(int(yy), set())
                    all_days = pd.date_range(
                        start=pd.Timestamp(int(yy), int(mm), 1),
                        end=pd.Timestamp(int(yy), int(mm), 1) + pd.offsets.MonthEnd(0),
                        freq="D"
                    )
                    open_days = []
                    for dt in all_days:
                        closed = (dt.weekday() >= 5) or (dt.date() in hset)
                        if not closed:
                            open_days.append(dt.normalize())

                    pos_map = {od: i for i, od in enumerate(open_days, start=1)}
                    total_open = len(open_days)
                    open_set = set(open_days)

                    for dn in grp["date_norm"]:
                        if dn in open_set:
                            idx = pos_map.get(dn, 0)
                            bdm.append(idx)
                            bdl.append(max(0, total_open - idx))
                        else:
                            bdm.append(0)
                            bdl.append(total_open)

                d["bdm"] = np.array(bdm, dtype=int)
                d["bdl"] = np.array(bdl, dtype=int)

                d["bdm_1"] = (d["bdm"] == 1).astype(int)
                d["bdm_2"] = (d["bdm"] == 2).astype(int)
                d["bdm_3"] = (d["bdm"] == 3).astype(int)
                d["is_bdm_1_3"] = ((d["bdm"] >= 1) & (d["bdm"] <= 3)).astype(int)

                d["is_first_bizday"] = (d["bdm"] == 1).astype(int)
                d["is_last_bizday"] = ((d["bdl"] == 0) & (d["is_closed"] == 0)).astype(int)

                # after closed (WE/ferié)
                d["prev_is_closed"] = d["is_closed"].shift(1).fillna(0).astype(int)
                d["is_after_closed"] = ((d["prev_is_closed"] == 1) & (d["is_closed"] == 0)).astype(int)

                # interactions (lundi + BOM)
                d["mon_bom_1_3"] = ((d["is_monday"] == 1) & (d["is_bom_1_3"] == 1)).astype(int)
                d["mon_bdm_1_3"] = ((d["is_monday"] == 1) & (d["is_bdm_1_3"] == 1)).astype(int)
                d["after_bom_1_3"] = ((d["is_after_closed"] == 1) & (d["is_bom_1_3"] == 1)).astype(int)
                d["after_bdm_1_3"] = ((d["is_after_closed"] == 1) & (d["is_bdm_1_3"] == 1)).astype(int)

                # spike-risk (jours où l'activité est probable)
                d["spike_risk"] = (
                    (d["is_bom_1_3"] == 1) |
                    (d["is_bdm_1_3"] == 1) |
                    (d["is_quarter_end"] == 1) |
                    (d["is_after_closed"] == 1) |
                    (d["is_monday"] == 1) |
                    (d["mon_bom_1_3"] == 1) |
                    (d["mon_bdm_1_3"] == 1) |
                    (d["is_near_eom3"] == 1)
                ).astype(int)

                # “mois” cyclique (aide sur pattern annuel sans lisser)
                d["month_sin"] = np.sin(2 * np.pi * d["month"] / 12.0)
                d["month_cos"] = np.cos(2 * np.pi * d["month"] / 12.0)

                # DOW cyclique
                d["dow_sin"] = np.sin(2 * np.pi * d["dow"] / 7.0)
                d["dow_cos"] = np.cos(2 * np.pi * d["dow"] / 7.0)

                return d

            df_filiale = _add_calendar_features(df_filiale, holiday_sets)

            # ---------------- Split ----------------
            df_train = df_filiale[df_filiale["year"] < base_year].copy()
            df_valid = df_filiale[df_filiale["year"] == base_year].copy()

            if df_train.shape[0] < 120 or df_valid.shape[0] < 80:
                print("[WARN] Peu de données pour split année -> fallback chrono 80/20")
                df_all = df_filiale.sort_values("date").copy()
                cut = int(len(df_all) * 0.8)
                cut = max(120, min(cut, len(df_all) - 60))
                df_train = df_all.iloc[:cut].copy()
                df_valid = df_all.iloc[cut:].copy()

            # ---------------- Features ----------------
            features = [
                "dow", "dow_sin", "dow_cos",
                "weekofyear", "month", "month_sin", "month_cos",
                "dom", "dayofyear",
                "is_month_start", "is_month_end",
                "is_quarter_start", "is_quarter_end",
                "days_to_eom", "is_near_eom3", "is_near_eom7",
                "dom_1", "dom_2", "dom_3", "is_bom_1_3", "is_bom_1_5", "is_bom_1_7",
                "bdm", "bdl", "bdm_1", "bdm_2", "bdm_3", "is_bdm_1_3",
                "is_first_bizday", "is_last_bizday",
                "is_holiday", "is_closed", "is_after_closed",
                "is_monday", "mon_bom_1_3", "mon_bdm_1_3",
                "after_bom_1_3", "after_bdm_1_3",
                "spike_risk",
            ]

            df_train_m = df_train.dropna(subset=features + ["y"]).copy()
            df_valid_m = df_valid.dropna(subset=features + ["y"]).copy()

            if df_train_m.shape[0] < 80 or df_valid_m.shape[0] < 40:
                messagebox.showinfo("Information", "Pas assez de lignes pour entraîner le modèle.")
                return

            # sécurise closed=0
            df_train_m.loc[df_train_m["is_closed"] == 1, "y"] = 0.0
            df_valid_m.loc[df_valid_m["is_closed"] == 1, "y"] = 0.0

            X_train = df_train_m[features].copy()
            y_train = df_train_m["y"].astype(float).copy()
            X_valid = df_valid_m[features].copy()
            y_valid = df_valid_m["y"].astype(float).copy()

            is_open_train = (df_train_m["is_closed"].values == 0)
            is_open_valid = (df_valid_m["is_closed"].values == 0)

            m_train = np.clip(-y_train.values, 0.0, None)
            m_valid = np.clip(-y_valid.values, 0.0, None)

            # ======================================================================
            # 1) Détection d'échelle + quantification des montants (pour anchors)
            # ======================================================================
            # On quantifie en "pas" adapté à l'échelle, pour détecter -30/-35/-45/-55/-120 etc sans dépendre de l'unité.
            nonzero = m_train[(m_train > 0) & np.isfinite(m_train)]
            if len(nonzero) == 0:
                messagebox.showinfo("Information", "Aucun flux négatif sur train.")
                return

            med_mag = float(np.median(nonzero))
            # heuristique de pas
            # - si magnitude ~1e8 (100M) => step = 1e6
            # - si magnitude ~1e5 (100k) => step = 1e3
            # - sinon step = 1
            if med_mag >= 5e7:
                step = 1e6
            elif med_mag >= 5e4:
                step = 1e3
            else:
                step = 1.0

            def qround(x):
                return float(np.round(x / step) * step)

            m_train_q = np.array([qround(v) for v in nonzero], dtype=float)
            vc = pd.Series(m_train_q).value_counts()

            # anchors "petits récurrents": on prend top fréquents sous ~80% de la médiane (pour éviter les -120)
            small_cut = 0.80 * med_mag
            small_candidates = vc[vc.index <= small_cut]
            small_anchors = small_candidates.head(6).index.values.astype(float).tolist()

            # anchors "gros récurrents": autour de la médiane +/- 20% (capte 110-140 autour de 120)
            big_lo, big_hi = 0.80 * med_mag, 1.20 * med_mag
            big_candidates = vc[(vc.index >= big_lo) & (vc.index <= big_hi)]
            big_anchors = big_candidates.head(7).index.values.astype(float).tolist()

            # si detection insuffisante, fallback simple
            if len(small_anchors) < 2:
                # fallback: on prend des valeurs basses fréquentes
                small_anchors = vc.head(6).index.values.astype(float).tolist()
            if len(big_anchors) < 2:
                big_anchors = big_candidates.index.values.astype(float).tolist()
                if len(big_anchors) < 2:
                    big_anchors = [med_mag]

            small_anchors = sorted(list(set(small_anchors)))
            big_anchors = sorted(list(set(big_anchors)))

            print(f"[SCALE] med_mag={med_mag:,.0f} step={step:,.0f}")
            print(f"[ANCHORS] small={small_anchors}")
            print(f"[ANCHORS] big={big_anchors}")

            # ======================================================================
            # 2) Labelisation "régime" (anti-lissage)
            # ======================================================================
            # regime:
            # 0: m=0
            # 1: petit récurrent (proche d'un small anchor)
            # 2: gros récurrent (proche d'un big anchor)
            # 3: spike/other (reste)
            def nearest_anchor_class(m, anchors, tol):
                if m <= 0:
                    return None
                a = np.array(anchors, dtype=float)
                idx = int(np.argmin(np.abs(a - m)))
                if abs(a[idx] - m) <= tol:
                    return idx
                return None

            # tolérance = 2*step pour absorber les petites variations / arrondis
            tol = 2.0 * step

            regime_train = np.zeros(len(df_train_m), dtype=int)
            # jours fermés => regime 0 (0)
            for i in range(len(df_train_m)):
                if int(df_train_m.iloc[i]["is_closed"]) == 1:
                    regime_train[i] = 0
                    continue
                m = float(m_train[i])
                if m <= 0:
                    regime_train[i] = 0
                    continue
                if nearest_anchor_class(m, small_anchors, tol) is not None:
                    regime_train[i] = 1
                elif nearest_anchor_class(m, big_anchors, tol) is not None:
                    regime_train[i] = 2
                else:
                    # spike si très gros
                    regime_train[i] = 3

            # ======================================================================
            # 3) Modèles : Régime + Anchors + SpikeReg
            # ======================================================================
            def _tune(model, param_dist, X_t, y_t, scoring, n_iter=18):
                rs = RandomizedSearchCV(
                    model,
                    param_distributions=param_dist,
                    n_iter=n_iter,
                    cv=3,
                    scoring=scoring,
                    random_state=42,
                    n_jobs=-1,
                    verbose=0
                )
                rs.fit(X_t, y_t)
                return rs.best_estimator_, rs.best_params_, rs.best_score_

            # --- Régime multiclass (précis jour-par-jour) ---
            # On entraîne sur jours ouverts + y (regime), mais closed gardés pour apprendre "0"
            cls_regime_base = LGBMClassifier(
                random_state=42,
                objective="multiclass",
                num_class=4,
                n_estimators=1800,
                learning_rate=0.03,
                subsample=0.9,
                colsample_bytree=0.9,
                num_leaves=63,
                min_child_samples=25,
                class_weight="balanced",
            )
            regime_params = {
                "n_estimators": randint(600, 2600),
                "num_leaves": randint(15, 127),
                "max_depth": randint(2, 10),
                "min_child_samples": randint(10, 140),
                "learning_rate": uniform(0.01, 0.08),
                "subsample": uniform(0.70, 0.30),
                "colsample_bytree": uniform(0.70, 0.30),
                "reg_lambda": uniform(0.0, 10.0),
            }

            # scoring: accuracy macro (LightGBM via sklearn -> "accuracy")
            cls_regime, bp, bs = _tune(cls_regime_base, regime_params, X_train, regime_train, scoring="accuracy", n_iter=20)
            print(f"[TUNING] REGIME best params: {bp} | CV acc={bs:.4f}")

            # --- Anchor classifier petit ---
            # y_small_class = index of anchor
            mask_small = (regime_train == 1)
            cls_small = None
            if np.sum(mask_small) >= 60 and len(small_anchors) >= 2:
                y_small = []
                for idx, rowi in enumerate(np.where(mask_small)[0]):
                    m = float(m_train[rowi])
                    c = nearest_anchor_class(m, small_anchors, tol)
                    y_small.append(int(c) if c is not None else 0)
                y_small = np.array(y_small, dtype=int)
                X_small = X_train.iloc[np.where(mask_small)[0]].copy()

                cls_small = LGBMClassifier(
                    random_state=42,
                    objective="multiclass",
                    num_class=len(small_anchors),
                    n_estimators=1200,
                    learning_rate=0.04,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                )
                cls_small.fit(X_small, y_small)
                print(f"[ANCHOR SMALL] ON classes={len(small_anchors)}")
            else:
                print("[ANCHOR SMALL] OFF (pas assez de points ou anchors insuffisants)")

            # --- Anchor classifier gros ---
            mask_big = (regime_train == 2)
            cls_big = None
            if np.sum(mask_big) >= 60 and len(big_anchors) >= 2:
                y_big = []
                for idx, rowi in enumerate(np.where(mask_big)[0]):
                    m = float(m_train[rowi])
                    c = nearest_anchor_class(m, big_anchors, tol)
                    y_big.append(int(c) if c is not None else 0)
                y_big = np.array(y_big, dtype=int)
                X_big = X_train.iloc[np.where(mask_big)[0]].copy()

                cls_big = LGBMClassifier(
                    random_state=42,
                    objective="multiclass",
                    num_class=len(big_anchors),
                    n_estimators=1400,
                    learning_rate=0.035,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                )
                cls_big.fit(X_big, y_big)
                print(f"[ANCHOR BIG] ON classes={len(big_anchors)}")
            else:
                print("[ANCHOR BIG] OFF (pas assez de points ou anchors insuffisants)")

            # --- Spike reg (régime 3) ---
            mask_spike = (regime_train == 3) & is_open_train
            # si peu de spikes, on inclut aussi gros non-anchors
            if np.sum(mask_spike) < 40:
                mask_spike = ((m_train > 0) & is_open_train & (regime_train == 3))

            X_spike = X_train.iloc[np.where(mask_spike)[0]].copy()
            m_spike = m_train[np.where(mask_spike)[0]].astype(float)

            # log1p magnitude
            y_spike_log = np.log1p(np.clip(m_spike, 0.0, None))

            # poids: surpondère gros spikes
            if len(m_spike) > 0:
                q90s = float(np.quantile(m_spike, 0.90))
            else:
                q90s = 1.0
            w_spike = 0.5 + 6.0 * np.sqrt(np.clip(m_spike / max(q90s, 1e-9), 0.0, 16.0))

            reg_spike = None
            reg_spike_q10 = None
            reg_spike_q90 = None

            if len(X_spike) >= 40:
                reg_spike = LGBMRegressor(
                    random_state=42,
                    objective="regression",
                    n_estimators=2000,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike.fit(X_spike, y_spike_log, sample_weight=w_spike)

                reg_spike_q10 = LGBMRegressor(
                    random_state=42,
                    objective="quantile",
                    alpha=0.10,
                    n_estimators=1400,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike_q90 = LGBMRegressor(
                    random_state=42,
                    objective="quantile",
                    alpha=0.90,
                    n_estimators=1400,
                    learning_rate=0.03,
                    subsample=0.9,
                    colsample_bytree=0.9,
                    num_leaves=63,
                    min_child_samples=20,
                    reg_lambda=2.0,
                )
                reg_spike_q10.fit(X_spike, y_spike_log, sample_weight=w_spike)
                reg_spike_q90.fit(X_spike, y_spike_log, sample_weight=w_spike)

                print(f"[SPIKE REG] ON points={len(X_spike)}")
            else:
                print("[SPIKE REG] OFF (pas assez de spikes)")

            # ======================================================================
            # 4) CAP (évite explosion artificielle, mais laisse passer les vrais pics)
            # ======================================================================
            # cap très haut : on se base sur quantiles open days
            open_mags = m_train[(m_train > 0) & is_open_train]
            q995 = float(np.quantile(open_mags, 0.995)) if len(open_mags) else float(med_mag)
            q999 = float(np.quantile(open_mags, 0.999)) if len(open_mags) else float(med_mag * 3)
            hard_cap = max(q999, q995, med_mag * 3)  # laisse passer très haut si vraiment présent
            print(f"[CAP] q995={q995:,.0f} q999={q999:,.0f} hard_cap={hard_cap:,.0f}")

            # ======================================================================
            # 5) Prédiction (régime -> montant) + calibration métier (BOM+Lundi)
            # ======================================================================
            def _predict_one(df_row, X_row):
                # closed => 0
                if int(df_row["is_closed"]) == 1:
                    return 0.0, 0.0, 0.0, 0  # y, p10, p90, regime

                proba = cls_regime.predict_proba(X_row)[0]  # size 4
                # calibration métier : BOM+Lundi => si le modèle hésite à activer, on pousse un peu
                # (évite le "raté" sur des jours critiques)
                if int(df_row["is_bom_1_3"]) == 1:
                    proba[1] *= 1.20  # petit
                    proba[2] *= 1.35  # gros (plus probable)
                    if int(df_row["is_monday"]) == 1:
                        proba[2] *= 1.25
                        proba[3] *= 1.10
                if int(df_row["mon_bom_1_3"]) == 1:
                    proba[2] *= 1.40
                    proba[3] *= 1.15

                # renormalize
                s = float(np.sum(proba))
                if s > 0:
                    proba = proba / s

                regime = int(np.argmax(proba))

                # base output
                mag = 0.0
                mag10, mag90 = 0.0, 0.0

                if regime == 0:
                    mag = 0.0
                    mag10, mag90 = 0.0, 0.0

                elif regime == 1:
                    # petit récurrent -> anchor
                    if cls_small is not None:
                        c = int(cls_small.predict(X_row)[0])
                        c = int(np.clip(c, 0, len(small_anchors) - 1))
                        mag = float(small_anchors[c])
                        mag10, mag90 = mag, mag
                    else:
                        # fallback: prend l'ancre la plus fréquente "small"
                        mag = float(small_anchors[0]) if len(small_anchors) else float(min(med_mag, q995))
                        mag10, mag90 = mag, mag

                elif regime == 2:
                    # gros récurrent -> anchor "big"
                    if cls_big is not None:
                        c = int(cls_big.predict(X_row)[0])
                        c = int(np.clip(c, 0, len(big_anchors) - 1))
                        mag = float(big_anchors[c])
                        mag10, mag90 = mag, mag
                    else:
                        # fallback: médiane
                        mag = float(med_mag)
                        mag10, mag90 = mag, mag

                else:
                    # regime 3: spike/other
                    if reg_spike is not None:
                        p = float(reg_spike.predict(X_row)[0])
                        mag = float(np.expm1(max(0.0, p)))
                        p10 = float(reg_spike_q10.predict(X_row)[0]) if reg_spike_q10 is not None else p
                        p90 = float(reg_spike_q90.predict(X_row)[0]) if reg_spike_q90 is not None else p
                        mag10 = float(np.expm1(max(0.0, p10)))
                        mag90 = float(np.expm1(max(0.0, p90)))
                    else:
                        # fallback: très haut quantile (mais pas absurde)
                        mag = float(q995)
                        mag10, mag90 = float(q995 * 0.7), float(min(hard_cap, q999))

                # cap (évite explosions artificielles)
                mag = min(max(mag, 0.0), hard_cap)
                mag10 = min(max(mag10, 0.0), hard_cap)
                mag90 = min(max(mag90, 0.0), hard_cap)

                # convertir en y<=0
                y = min(-mag, 0.0)
                y_p10 = min(-mag90, 0.0)  # P10 plus négatif (borne basse)
                y_p90 = min(-mag10, 0.0)  # P90 moins négatif (borne haute)

                return float(y), float(y_p10), float(y_p90), regime

            # ======================================================================
            # 6) VALID metrics
            # ======================================================================
            y_pred_valid = np.zeros(len(df_valid_m), dtype=float)
            y_p10_valid = np.zeros(len(df_valid_m), dtype=float)
            y_p90_valid = np.zeros(len(df_valid_m), dtype=float)
            r_valid = np.zeros(len(df_valid_m), dtype=int)

            for i in range(len(df_valid_m)):
                Xr = pd.DataFrame([X_valid.iloc[i].values], columns=features)
                yhat, p10, p90, rr = _predict_one(df_valid_m.iloc[i], Xr)
                y_pred_valid[i] = yhat
                y_p10_valid[i] = p10
                y_p90_valid[i] = p90
                r_valid[i] = rr

            mae = mean_absolute_error(y_valid.values, y_pred_valid)
            rmse = mean_squared_error(y_valid.values, y_pred_valid) ** 0.5
            r2 = r2_score(y_valid.values, y_pred_valid)
            bias = float(np.mean(y_pred_valid - y_valid.values))

            # couverture bande si dispo
            coverage = float(np.mean((y_valid.values >= y_p10_valid) & (y_valid.values <= y_p90_valid)))
            width = float(np.mean(y_p90_valid - y_p10_valid))

            print(f"[V5.0] MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}, Bias={bias:.2f}")
            print(f"[BAND] Coverage={coverage:.3f} Width={width:.1f}")

            # ===================== KPI UI =====================
            kpi_frame = ctk.CTkFrame(scrollable_frame, fg_color="#0f1b31", corner_radius=18)
            kpi_frame.pack(fill="x", padx=10, pady=(0, 10))
            graph_widgets.append(kpi_frame)
            for i in range(10):
                kpi_frame.grid_columnconfigure(i, weight=1)

            def _kpi(parent, title, value, subtitle, col):
                card = ctk.CTkFrame(parent, fg_color="#142544", corner_radius=16, border_width=1, border_color="#223658")
                card.grid(row=0, column=col, sticky="nsew", padx=10, pady=10)
                ctk.CTkLabel(card, text=title, font=("Segoe UI", 12), text_color="#9fb7dd") \
                    .grid(row=0, column=0, sticky="w", padx=14, pady=(12, 0))
                ctk.CTkLabel(card, text=value, font=("Segoe UI Semibold", 22, "bold"), text_color="white") \
                    .grid(row=1, column=0, sticky="w", padx=14, pady=(2, 6))
                ctk.CTkLabel(card, text=subtitle, font=("Segoe UI", 11), text_color="#7ea2d8") \
                    .grid(row=2, column=0, sticky="w", padx=14, pady=(0, 12))

            _kpi(kpi_frame, "MAE", f"{mae:.1f}", "Erreur moyenne", 0)
            _kpi(kpi_frame, "RMSE", f"{rmse:.1f}", "Risque spikes", 1)
            _kpi(kpi_frame, "R²", f"{r2:.3f}", "Pouvoir explicatif", 2)
            _kpi(kpi_frame, "Bias", f"{bias:.0f}", "Moy(y_pred - y)", 3)
            _kpi(kpi_frame, "Coverage", f"{coverage:.1%}", "Réel ∈ [P10,P90]", 4)
            _kpi(kpi_frame, "Width", f"{width:.0f}", "Largeur bande", 5)
            _kpi(kpi_frame, "Step", f"{step:.0f}", "Quantif anchors", 6)
            _kpi(kpi_frame, "SmallAnch", f"{len(small_anchors)}", "Petits récurrents", 7)
            _kpi(kpi_frame, "BigAnch", f"{len(big_anchors)}", "Gros ~120", 8)
            _kpi(kpi_frame, "SpikeReg", "ON" if reg_spike is not None else "OFF", "Modèle spikes", 9)

            # ======================================================================
            # Forecast N+1 (non lissé, jour-par-jour)
            # ======================================================================
            df_real_target = df_filiale[df_filiale["year"] == target_year].copy().sort_values("date")
            current_real_target_df = df_real_target

            start = pd.Timestamp(year=target_year, month=1, day=1)
            end = pd.Timestamp(year=target_year, month=12, day=31)
            all_dates_tgt = pd.date_range(start=start, end=end, freq="D")

            target_holidays = holiday_sets.get(int(target_year), set())

            # Build df_future skeleton
            future_rows = []
            for dt in all_dates_tgt:
                dt = pd.Timestamp(dt).normalize()
                dow = int(dt.weekday())
                is_h = int(dt.date() in target_holidays)
                is_closed = int((dow >= 5) or (is_h == 1))

                row = {
                    "section": filiale,
                    "date": dt,
                    "year": int(dt.year),
                    "dow": dow,
                    "is_holiday": is_h,
                    "is_closed": is_closed,
                    "y": 0.0,  # placeholder
                }
                future_rows.append(row)

            df_future = pd.DataFrame(future_rows).sort_values("date").reset_index(drop=True)

            # recompute full calendar features for target year (avec holiday_sets)
            df_future = _add_calendar_features(df_future, holiday_sets)

            # align features
            for c in features:
                if c not in df_future.columns:
                    df_future[c] = 0

            # predict day-by-day
            pred_vals = []
            pred_p10s = []
            pred_p90s = []
            pred_regimes = []

            for i in range(len(df_future)):
                Xr = pd.DataFrame([df_future.loc[i, features].values], columns=features)
                yhat, p10, p90, rr = _predict_one(df_future.iloc[i], Xr)
                pred_vals.append(yhat)
                pred_p10s.append(p10)
                pred_p90s.append(p90)
                pred_regimes.append(rr)

            df_future["pred_value"] = np.array(pred_vals, dtype=float)
            df_future["pred_p10"] = np.array(pred_p10s, dtype=float)
            df_future["pred_p90"] = np.array(pred_p90s, dtype=float)
            df_future["pred_regime"] = np.array(pred_regimes, dtype=int)

            # merge real target if present
            df_future_all = df_future[["section", "date", "year", "month", "dayofyear", "pred_p10", "pred_value", "pred_p90", "pred_regime"]].copy()
            df_future_all["month"] = df_future_all["date"].dt.month
            df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

            if df_real_target is not None and not df_real_target.empty:
                df_future_all = pd.merge(df_real_target[["date"]].copy(), df_future_all, on="date", how="left")
                for c in ["pred_p10", "pred_value", "pred_p90"]:
                    df_future_all[c] = df_future_all[c].fillna(0.0)
                df_future_all["section"] = filiale
                df_future_all["year"] = df_future_all["date"].dt.year
                df_future_all["month"] = df_future_all["date"].dt.month
                df_future_all["dayofyear"] = df_future_all["date"].dt.dayofyear

            current_pred_df = df_future_all
            exported_pred_df = df_future_all.copy()

            # store artifacts
            model_artifacts["features"] = features[:]
            model_artifacts["cls_regime"] = cls_regime
            model_artifacts["cls_small"] = cls_small
            model_artifacts["cls_big"] = cls_big
            model_artifacts["reg_spike"] = reg_spike
            model_artifacts["reg_spike_q10"] = reg_spike_q10
            model_artifacts["reg_spike_q90"] = reg_spike_q90
            model_artifacts["small_anchors"] = small_anchors[:]
            model_artifacts["big_anchors"] = big_anchors[:]
            model_artifacts["step"] = float(step)
            model_artifacts["hard_cap"] = float(hard_cap)

            # ---------------- Plot ----------------
            fig, ax = plt.subplots(figsize=(11, 4.5), facecolor="#00122e", constrained_layout=True)
            ax.set_facecolor("#00122e")

            df_hist_plot = df_filiale[df_filiale["year"] <= base_year].copy().sort_values("date")
            ax.plot(df_hist_plot["date"], df_hist_plot["y"], label=f"Réel (≤ {base_year})", linewidth=2)

            ax.plot(df_future_all["date"], df_future_all["pred_value"], label=f"Prévision IA {target_year} (ULTRA, ≤0)", linewidth=2, linestyle="--")
            ax.fill_between(df_future_all["date"], df_future_all["pred_p10"], df_future_all["pred_p90"], alpha=0.25,
                            label="Bande P10–P90 (spikes)")

            if df_real_target is not None and not df_real_target.empty:
                ax.plot(df_real_target["date"], df_real_target["y"], label=f"Réel {target_year}", linewidth=2)

            ax.set_title(f"ACE & Investissements – année {target_year} – {filiale} (NEG V5.0 ULTRA)", color="white")
            ax.tick_params(axis='x', colors="white", rotation=30)
            ax.tick_params(axis='y', colors="white")
            leg = ax.legend(facecolor="#00122e", edgecolor="white")
            for t in leg.get_texts():
                t.set_color("white")

            canvas = FigureCanvasTkAgg(fig, master=scrollable_frame)
            canvas.draw()
            w = canvas.get_tk_widget()
            w.pack(pady=10, fill="both", expand=True)
            graph_widgets.append(w)
            plt.close(fig)

            _redraw_monthly_graph()
            _redraw_graph2()

            if analysis_table_frame is not None:
                analysis_table_frame.destroy()

            analysis_table_frame = ctk.CTkFrame(scrollable_frame, fg_color="#001838", corner_radius=12)
            analysis_table_frame.pack(fill="x", padx=10, pady=(10, 20))
            graph_widgets.append(analysis_table_frame)

            _rebuild_analysis_table()    

        bouton_train.configure(command=_train_model)



        # ============ SCROLL MOLETTE ============
        def _on_mousewheel(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.yview_scroll(step, "units")
            return "break"

        def _on_linux_scroll_up(event):
            main_canvas.yview_scroll(-1, "units")
            return "break"

        def _on_linux_scroll_down(event):
            main_canvas.yview_scroll(1, "units")
            return "break"

        def _on_mousewheel_shift(event):
            if event.delta == 0:
                return "break"
            step = -1 if event.delta > 0 else 1
            main_canvas.xview_scroll(step, "units")
            return "break"

        def _bind_mousewheel(_event=None):
            main_canvas.bind_all("<MouseWheel>", _on_mousewheel, add="+")
            main_canvas.bind_all("<Shift-MouseWheel>", _on_mousewheel_shift, add="+")
            main_canvas.bind_all("<Button-4>", _on_linux_scroll_up, add="+")
            main_canvas.bind_all("<Button-5>", _on_linux_scroll_down, add="+")

        def _unbind_mousewheel(_event=None):
            main_canvas.unbind_all("<MouseWheel>")
            main_canvas.unbind_all("<Shift-MouseWheel>")
            main_canvas.unbind_all("<Button-4>")
            main_canvas.unbind_all("<Button-5>")

        main_canvas.bind("<Enter>", _bind_mousewheel, add="+")
        main_canvas.bind("<Leave>", _unbind_mousewheel, add="+")

        # ============ BOUTON RETOUR ============
        ctk.CTkButton(
            scrollable_frame,
            text="⬅️ Retour au menu",
            command=self.retour_menu,
            width=200, height=40, corner_radius=15,
            fg_color="#444", hover_color="#666", text_color="white",
            font=("Segoe UI", 13, "bold")
        ).pack(pady=15)

if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()

    app = Application()
    app.mainloop()